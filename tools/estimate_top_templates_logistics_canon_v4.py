#!/usr/bin/env python3
# === ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4 ===
from __future__ import annotations

import io
import json
import os
import re
import sqlite3
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

BASE = Path("/root/.areal-neva-core")
sys.path.insert(0, str(BASE))

AI_ORCHESTRA_FOLDER_ID = "13No7_E7Mwj1n1awNQ-lzbohWGOiEM2PB"
MEMORY_DB = BASE / "data" / "memory.db"
REGISTRY_PATH = BASE / "config" / "estimate_template_registry.json"
CANON_PATH = BASE / "docs" / "CANON_FINAL" / "ESTIMATE_TEMPLATE_M80_M110_CANON.md"
REPORT_PATH = BASE / "docs" / "REPORTS" / "ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4_REPORT.md"
FORMULA_INDEX_PATH = BASE / "data" / "templates" / "estimate_logic" / "estimate_template_formula_index.json"

TEMPLATES = [
    {"key": "M80", "aliases": ["М-80.xlsx", "M-80.xlsx"], "role": "full_house_estimate_template", "description": "Эталон полной сметы М-80"},
    {"key": "M110", "aliases": ["М-110.xlsx", "M-110.xlsx"], "role": "full_house_estimate_template", "description": "Эталон полной сметы М-110"},
    {"key": "ROOF_FLOORS", "aliases": ["крыша и перекр.xlsx"], "role": "roof_and_floor_estimate_template", "description": "Эталон расчёта кровли и перекрытий"},
    {"key": "FOUNDATION_WAREHOUSE", "aliases": ["фундамент_Склад2.xlsx"], "role": "foundation_estimate_template", "description": "Эталон расчёта фундамента"},
    {"key": "AREAL_NEVA", "aliases": ["Ареал Нева.xlsx"], "role": "general_company_estimate_template", "description": "Общий эталон сметной структуры Ареал-Нева"},
]

SECTION_ORDER = [
    "Фундамент",
    "Каркас",
    "Стены",
    "Перекрытия",
    "Кровля",
    "Окна, двери",
    "Внешняя отделка",
    "Внутренняя отделка",
    "Инженерные коммуникации",
    "Логистика",
    "Накладные расходы",
]

UNIVERSAL_MATERIAL_GROUPS = {
    "стены": ["кирпич", "газобетон", "керамоблок", "арболит", "монолит", "каркас", "брус"],
    "фундамент": ["монолитная плита", "лента", "сваи", "ростверк", "утеплённая плита", "складской фундамент"],
    "кровля": ["металлочерепица", "профнастил", "гибкая черепица", "фальц", "мембрана", "стропильная система"],
    "перекрытия": ["деревянные балки", "монолит", "плиты", "металлические балки"],
    "утепление": ["минвата", "роквул", "пеноплэкс", "pir", "эковата"],
    "отделка": ["имитация бруса", "штукатурка", "плитка", "гкл", "цсп", "фасадная доска"],
    "инженерия": ["электрика", "водоснабжение", "канализация", "отопление", "вентиляция"],
    "логистика": ["доставка", "разгрузка", "манипулятор", "кран", "проживание", "транспорт бригады", "удалённость"],
}

FORMULA_POLICY = [
    "Топовые сметы являются эталонами логики расчёта, а не прайс-листами",
    "Новые сметы считаются по такой же структуре: разделы, строки, колонки, формулы, итоги, примечания, исключения",
    "Материал может быть любым: кирпич, газобетон, каркас, монолит, кровля, перекрытия, отделка, инженерия",
    "При замене материала сохраняется расчётная логика: количество × цена = сумма; работа + материалы = всего; разделы = итоги; финальный итог = сумма разделов",
    "Каркасный сценарий, газобетон/монолитная плита, кровля/перекрытия и фундамент считаются как разные сценарии и не смешиваются",
    "Если объёмов не хватает — оркестр спрашивает только недостающие объёмы",
    "Если пользователь прислал файл как образец — сначала принять как образец, а не запускать поиск цен",
]

PRICE_CONFIRMATION_FLOW = [
    "Интернет-цены материалов и техники не подставляются молча",
    "Для финальной сметы оркестр ищет актуальные цены по материалам, технике, доставке и разгрузке",
    "По каждой позиции показывает: источник, цена, единица, дата/регион, ссылка",
    "Оркестр предлагает среднюю/медианную цену без явных выбросов",
    "Пользователь выбирает: средняя / минимальная / максимальная / конкретная ссылка / ручная цена",
    "Пользователь может добавить наценку, запас, скидку, поправку по позиции, разделу или всей смете",
    "До подтверждения цен финальный XLSX/PDF не выпускается",
    "После подтверждения цены пересчитываются по формулам шаблона",
]

LOGISTICS_POLICY = [
    "Перед финальной сметой оркестр обязан запросить локацию объекта или расстояние от города",
    "Стоимость объекта рядом с городом и объекта за 200 км не может быть одинаковой",
    "Оркестр обязан учитывать доставку материалов, транспорт бригады, разгрузку, манипулятор/кран, проживание, удалённость, дорожные условия",
    "Если логистика неизвестна — оркестр задаёт один короткий вопрос: город/населённый пункт или расстояние от города, подъезд для грузовой техники, нужна ли разгрузка/манипулятор",
    "Логистика считается отдельным блоком сметы или отдельным коэффициентом, но не смешивается молча с ценами материалов",
    "Перед финальным результатом оркестр показывает логистические допущения и спрашивает подтверждение",
]

def now() -> str:
    return datetime.now(timezone.utc).isoformat()

def s(v: Any) -> str:
    return "" if v is None else str(v)

def clean(v: Any) -> str:
    return re.sub(r"\s+", " ", s(v)).strip()

def get_drive_service():
    from dotenv import load_dotenv
    load_dotenv("/root/.areal-neva-core/.env", override=True)

    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    cid = os.getenv("GDRIVE_CLIENT_ID", "").strip()
    sec = os.getenv("GDRIVE_CLIENT_SECRET", "").strip()
    ref = os.getenv("GDRIVE_REFRESH_TOKEN", "").strip()

    if cid and sec and ref:
        creds = Credentials(
            None,
            refresh_token=ref,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=cid,
            client_secret=sec,
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        creds.refresh(Request())
        return build("drive", "v3", credentials=creds)

    import google_io
    return google_io.get_drive_service()

def find_file(service, aliases: List[str]) -> Dict[str, Any]:
    for name in aliases:
        safe_name = name.replace("'", "\\'")
        for q in [
            f"name='{safe_name}' and '{AI_ORCHESTRA_FOLDER_ID}' in parents and trashed=false",
            f"name='{safe_name}' and trashed=false",
        ]:
            res = service.files().list(
                q=q,
                fields="files(id,name,mimeType,modifiedTime,size,webViewLink,parents)",
                pageSize=20,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute()
            files = res.get("files", [])
            if files:
                files.sort(key=lambda x: x.get("modifiedTime") or "", reverse=True)
                return files[0]
    raise RuntimeError("DRIVE_TEMPLATE_NOT_FOUND_OR_NOT_ACCESSIBLE: " + " / ".join(aliases))

def download_xlsx(service, meta: Dict[str, Any]) -> bytes:
    from googleapiclient.http import MediaIoBaseDownload

    mime = meta.get("mimeType") or ""
    file_id = meta["id"]

    if mime == "application/vnd.google-apps.spreadsheet":
        req = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=file_id, supportsAllDrives=True)

    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()

def row_text(row: List[Any]) -> str:
    return " ".join(clean(x) for x in row if clean(x))

def detect_scenario(text: str, title: str) -> str:
    title_low = (title or "").lower()
    low = (title + " " + text).lower()

    # ВАЖНО: сначала название файла/листа, потому что полные сметы М-80/М-110
    # содержат внутри кровлю и перекрытия, но листы называются "Каркас" и "Газобетон"
    if any(x in title_low for x in ["каркас", "frame"]):
        return "frame_house"

    if any(x in title_low for x in ["газобетон", "газо", "кладка", "masonry"]):
        return "gasbeton_or_masonry_with_monolithic_foundation"

    if any(x in title_low for x in ["фундамент", "склад", "foundation"]):
        return "foundation"

    if any(x in title_low for x in ["крыш", "кров", "перекр", "roof", "floor"]):
        return "roof_and_floors"

    # Потом fallback по содержимому
    if any(x in low for x in ["газобетон", "кладка стен", "арматурного каркаса", "бетон в20", "бетон в22"]):
        return "gasbeton_or_masonry_with_monolithic_foundation"

    if any(x in low for x in ["каркас", "свая винтовая", "свайный фундамент", "обвязка свай", "доска с/к"]):
        return "frame_house"

    if any(x in low for x in ["фундамент", "монолитная плита", "ростверк", "свая", "склад"]):
        if not any(y in low for y in ["кровля", "кровель", "стропил", "перекрыт"]):
            return "foundation"

    if any(x in low for x in ["кров", "стропил", "перекр", "балк"]):
        return "roof_and_floors"

    return "general_estimate"

def extract_formula_cells(ws) -> List[Dict[str, str]]:
    out = []
    for row in ws.iter_rows():
        for c in row:
            val = c.value
            if isinstance(val, str) and val.startswith("="):
                out.append({"sheet": ws.title, "cell": c.coordinate, "formula": val[:500]})
    return out

def extract_structure(ws_values, file_title: str) -> Dict[str, Any]:
    rows = [list(r) for r in ws_values.iter_rows(values_only=True)]
    sections = []
    header_rows = []
    total_rows = []
    sample_rows = []
    material_rows = 0
    work_rows = 0
    logistics_rows = 0

    for i, r in enumerate(rows, start=1):
        txt = row_text(r)
        low = txt.lower()
        if not txt:
            continue

        for sec in SECTION_ORDER:
            if low.strip(" :") == sec.lower() and sec not in sections:
                sections.append(sec)

        if "№ п/п" in txt and ("Наименование" in txt or "Наименование работ" in txt):
            header_rows.append(i)

        if low.startswith("итого") or "итого сметная стоимость" in low or "всего" == low.strip():
            total_rows.append({"row": i, "text": txt[:300]})

        if any(x in low for x in ["логист", "достав", "транспорт", "разгруз", "манипулятор", "кран", "проживан", "удален", "удалён", "км"]):
            logistics_rows += 1

        name = clean(r[1] if len(r) > 1 else "")
        unit = clean(r[2] if len(r) > 2 else "")
        qty = clean(r[3] if len(r) > 3 else "")
        work_price = clean(r[4] if len(r) > 4 else "")
        material_price = clean(r[6] if len(r) > 6 else "")

        if name and (unit or qty):
            if work_price and work_price not in ("0", "0.0", "0,0", "-"):
                work_rows += 1
            if material_price and material_price not in ("0", "0.0", "0,0", "-"):
                material_rows += 1
            if len(sample_rows) < 35:
                sample_rows.append({
                    "row": i,
                    "name": name[:180],
                    "unit": unit,
                    "qty": qty,
                    "work_price": work_price,
                    "material_price": material_price,
                })

    hay = "\n".join(row_text(r) for r in rows[:250])
    return {
        "scenario": detect_scenario(hay, file_title),
        "sections": sections,
        "header_rows": header_rows,
        "total_rows": total_rows[:50],
        "material_rows": material_rows,
        "work_rows": work_rows,
        "logistics_rows": logistics_rows,
        "sample_rows": sample_rows,
        "row_count": len(rows),
    }

def analyze_template(service, template: Dict[str, Any], meta: Dict[str, Any]) -> Dict[str, Any]:
    import openpyxl

    raw = download_xlsx(service, meta)
    wb_formula = openpyxl.load_workbook(io.BytesIO(raw), data_only=False, read_only=False)
    wb_values = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)

    sheets = []
    formula_total = 0
    formula_samples = []

    for ws_f, ws_v in zip(wb_formula.worksheets, wb_values.worksheets):
        formulas = extract_formula_cells(ws_f)
        struct = extract_structure(ws_v, f"{meta.get('name') or ''} {ws_f.title}")
        formula_total += len(formulas)
        formula_samples.extend(formulas[:50])
        sheets.append({
            "sheet_name": ws_f.title,
            "scenario": struct["scenario"],
            "sections": struct["sections"],
            "header_rows": struct["header_rows"],
            "total_rows": struct["total_rows"],
            "material_rows": struct["material_rows"],
            "work_rows": struct["work_rows"],
            "logistics_rows": struct["logistics_rows"],
            "sample_rows": struct["sample_rows"],
            "formula_count": len(formulas),
            "formula_samples": formulas[:30],
            "row_count": struct["row_count"],
        })

    return {
        "key": template["key"],
        "title": meta["name"],
        "template_role": template["role"],
        "description": template["description"],
        "file_id": meta["id"],
        "drive_url": meta.get("webViewLink") or f"https://drive.google.com/file/d/{meta['id']}/view",
        "mimeType": meta.get("mimeType"),
        "modifiedTime": meta.get("modifiedTime"),
        "parents": meta.get("parents") or [],
        "formula_total": formula_total,
        "formula_samples": formula_samples[:120],
        "sheets": sheets,
    }

def build_policy(source_files: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "version": "ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4",
        "status": "ACTIVE_CANON",
        "updated_at": now(),
        "purpose": "Use top estimate files as scalable estimate calculation logic templates with mandatory logistics and web price confirmation",
        "source_files": source_files,
        "canonical_columns": [
            "№ п/п",
            "Наименование",
            "Ед. изм.",
            "Кол-во",
            "Работа Цена",
            "Работа Стоимость",
            "Материалы Цена",
            "Материалы Стоимость",
            "Всего",
            "Примечание",
        ],
        "canonical_sections": SECTION_ORDER,
        "universal_material_groups": UNIVERSAL_MATERIAL_GROUPS,
        "formula_policy": FORMULA_POLICY,
        "price_confirmation_flow": PRICE_CONFIRMATION_FLOW,
        "logistics_policy": LOGISTICS_POLICY,
        "runtime_rule": "ai_router injects this context through core.estimate_template_policy.build_estimate_template_context",
    }

def write_canon(policy: Dict[str, Any]) -> None:
    CANON_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = []
    lines.append("# ESTIMATE_TEMPLATE_TOP_CANON")
    lines.append("")
    lines.append("status: ACTIVE_CANON")
    lines.append("version: ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4")
    lines.append("updated_at: " + policy["updated_at"])
    lines.append("")
    lines.append("## ГЛАВНОЕ")
    lines.append("")
    lines.append("М-80.xlsx, М-110.xlsx, крыша и перекр.xlsx, фундамент_Склад2.xlsx, Ареал Нева.xlsx — топовые эталонные сметы")
    lines.append("Они являются образцами логики построения смет, формул, разделов, колонок, итогов, примечаний и исключений")
    lines.append("Они не являются фиксированным прайсом")
    lines.append("Оркестр обязан переносить их расчётную логику на любые новые задачи и любые материалы")
    lines.append("")
    lines.append("## ЧТО СОХРАНЯТЬ")
    for r in policy["formula_policy"]:
        lines.append("- " + r)
    lines.append("")
    lines.append("## ЦЕНЫ ИЗ ИНТЕРНЕТА")
    for r in policy["price_confirmation_flow"]:
        lines.append("- " + r)
    lines.append("")
    lines.append("## ЛОГИСТИКА И НАКЛАДНЫЕ")
    for r in policy["logistics_policy"]:
        lines.append("- " + r)
    lines.append("")
    lines.append("## КОЛОНКИ")
    lines.append(" | ".join(policy["canonical_columns"]))
    lines.append("")
    lines.append("## РАЗДЕЛЫ")
    for i, sec in enumerate(policy["canonical_sections"], 1):
        lines.append(f"{i}. {sec}")
    lines.append("")
    lines.append("## МАТЕРИАЛЫ")
    for group, values in policy["universal_material_groups"].items():
        lines.append(f"- {group}: " + ", ".join(values))
    lines.append("")
    lines.append("## ПРОЧИТАННЫЕ ШАБЛОНЫ")
    for src in policy["source_files"]:
        lines.append("")
        lines.append(f"### {src['title']}")
        lines.append(f"- role: `{src['template_role']}`")
        lines.append(f"- file_id: `{src['file_id']}`")
        lines.append(f"- drive_url: {src['drive_url']}")
        lines.append(f"- formula_total: {src['formula_total']}")
        for sh in src["sheets"]:
            lines.append(f"  - sheet: {sh['sheet_name']} | scenario={sh['scenario']} | formulas={sh['formula_count']} | material_rows={sh['material_rows']} | work_rows={sh['work_rows']} | logistics_rows={sh['logistics_rows']}")
    lines.append("")
    lines.append("## ОБЯЗАТЕЛЬНОЕ ПОВЕДЕНИЕ")
    lines.append("")
    lines.append("При новой смете оркестр обязан брать структуру и формулы из топовых эталонов")
    lines.append("Оркестр обязан подставлять конкретные объёмы и материалы задачи")
    lines.append("Оркестр обязан запросить локацию/удалённость/доступ/разгрузку до финального расчёта")
    lines.append("Оркестр обязан обновлять цены материалов и логистики через интернет только с подтверждением пользователя")
    lines.append("Оркестр обязан показывать найденные цены, источники, ссылки и среднюю/медианную цену")
    lines.append("Пользователь выбирает цену или задаёт ручную, может добавить наценку/скидку/запас")
    lines.append("Финальный XLSX/PDF запрещён до подтверждения цен и логистики")
    lines.append("")
    CANON_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")

def write_registry(policy: Dict[str, Any]) -> None:
    REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
    old = {}
    if REGISTRY_PATH.exists():
        try:
            old = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
        except Exception:
            old = {}
    old["estimate_top_templates_logistics_canon_v4"] = policy
    old["active_estimate_template_policy"] = "ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4"
    old["estimate_formula_logic_preserve_required"] = True
    old["estimate_material_price_web_refresh_required"] = True
    old["estimate_price_confirmation_required"] = True
    old["estimate_logistics_required"] = True
    old["estimate_final_xlsx_forbidden_before_price_and_logistics_confirmation"] = True
    REGISTRY_PATH.write_text(json.dumps(old, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def write_formula_index(policy: Dict[str, Any]) -> None:
    FORMULA_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    FORMULA_INDEX_PATH.write_text(json.dumps(policy, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def save_memory_sqlite(policy: Dict[str, Any]) -> None:
    if not MEMORY_DB.exists():
        raise RuntimeError(f"MEMORY_DB_MISSING: {MEMORY_DB}")

    value = json.dumps(policy, ensure_ascii=False, indent=2)
    ts = now()
    keys = [
        "estimate_top_templates_logistics_canon_v4",
        "topic_0_estimate_top_templates_logistics_canon_v4",
        "topic_2_estimate_top_templates_logistics_canon_v4",
        "topic_210_estimate_top_templates_logistics_canon_v4",
        "estimate_universal_material_calculation_policy_v4",
        "estimate_price_confirmation_required_v4",
        "estimate_logistics_required_v4",
    ]

    conn = sqlite3.connect(str(MEMORY_DB))
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(memory)").fetchall()]
        for key in keys:
            data = {
                "id": str(uuid.uuid4()),
                "chat_id": "-1003725299009",
                "key": key,
                "value": value,
                "timestamp": ts,
                "topic_id": 2,
                "scope": "topic",
            }
            use_cols = [c for c in ["id", "chat_id", "key", "value", "timestamp", "topic_id", "scope"] if c in cols]
            sql = f"INSERT INTO memory({','.join(use_cols)}) VALUES ({','.join(['?'] * len(use_cols))})"
            conn.execute(sql, [data[c] for c in use_cols])
        conn.commit()
    finally:
        conn.close()

def write_report(policy: Dict[str, Any]) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4_REPORT",
        "",
        "status: OK",
        "updated_at: " + policy["updated_at"],
        "canon: docs/CANON_FINAL/ESTIMATE_TEMPLATE_M80_M110_CANON.md",
        "registry: config/estimate_template_registry.json",
        "formula_index: data/templates/estimate_logic/estimate_template_formula_index.json",
        "",
        "## CLOSED",
        "- top estimate templates resolved from Drive",
        "- XLSX formulas extracted",
        "- universal material logic registered",
        "- web price confirmation registered",
        "- logistics and overhead clarification registered",
        "- direct sqlite memory write completed",
        "- ai_router context hook enabled",
        "",
        "## RAW_POLICY",
        "```json",
        json.dumps(policy, ensure_ascii=False, indent=2),
        "```",
    ]
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")

def main() -> int:
    service = get_drive_service()
    about = service.about().get(fields="user").execute()
    print("DRIVE_ACCOUNT", about.get("user", {}).get("emailAddress"))

    source_files = []
    for template in TEMPLATES:
        meta = find_file(service, template["aliases"])
        print("TEMPLATE_FOUND", template["key"], meta.get("name"), meta.get("id"), meta.get("parents"))
        source_files.append(analyze_template(service, template, meta))

    if not source_files:
        raise RuntimeError("NO_TEMPLATES_ANALYZED")

    policy = build_policy(source_files)
    write_canon(policy)
    write_registry(policy)
    write_formula_index(policy)
    save_memory_sqlite(policy)
    write_report(policy)

    print("ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4_OK")
    for src in source_files:
        print("SOURCE", src["title"], src["file_id"], "role", src["template_role"], "formulas", src["formula_total"])
        for sh in src["sheets"]:
            print("SHEET", sh["sheet_name"], sh["scenario"], "formulas", sh["formula_count"], "materials", sh["material_rows"], "works", sh["work_rows"], "logistics", sh["logistics_rows"])
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
# === END_ESTIMATE_TOP_TEMPLATES_LOGISTICS_CANON_V4 ===
