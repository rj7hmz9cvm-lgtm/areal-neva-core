#!/usr/bin/env python3
# TOPIC2_DRAINAGE_PRICE_ENRICHMENT_CANON_FIX_V1
# Canonical price flow only:
#   _openrouter_price_search → _price_prompt → user choice → XLSX/PDF
# No custom Sonar prompts. No regex price parsing. No fallback prices.
# No XLSX/PDF before TOPIC2_PRICE_CHOICE_CONFIRMED.
from __future__ import annotations
import asyncio, glob, json, os, re, sqlite3, subprocess, sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import requests
from dotenv import load_dotenv

BASE = Path("/root/.areal-neva-core")
sys.path.insert(0, str(BASE))
DB   = BASE / "data" / "core.db"
OUT_DIR = BASE / "runtime" / "stroyka_estimates" / "drainage_repair"
TASK_ID  = "043e5c9f-e8bc-434c-9dad-a66c7e50f917"
CACHE_FILE = OUT_DIR / f"price_cache_{TASK_ID[:8]}.json"
OUT_DIR.mkdir(parents=True, exist_ok=True)
load_dotenv(BASE / ".env", override=False)
VAT_RATE = 0.22

# ---------------------------------------------------------------------------
# Canonical imports — must not be replaced with custom equivalents
# ---------------------------------------------------------------------------
from core.price_enrichment import (
    _openrouter_price_search,
    _detect_price_choice,
    _price_prompt,
    _select_price,
    _apply_selected_prices,
)

# ---------------------------------------------------------------------------
# Source classification helpers
# ---------------------------------------------------------------------------
DRAINAGE_STRONG = ["нвд","наружные водостоки","наружные водостоки и дренажи",
    "схема дренажной и ливневой канализации","дренажная насосная станция",
    "пескоуловитель","линейный водоотвод","d=160","i=0,005","дк","лк"]
GEO_STRONG = ["инженерно-геологические","бурение геотехнических скважин","скважин",
    "игэ","грунтовых вод","нормативная глубина промерзания","супесь","насыпные грунты"]

def low(t): return str(t or "").lower().replace("ё","е")

def hist(conn, tid, action):
    conn.execute(
        "INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
        (tid, action[:900]),
    )

def pdf_text(path):
    try:
        r = subprocess.run(["pdftotext","-layout","-q",str(path),"-"],
            stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True, timeout=25)
        return r.stdout or ""
    except Exception as e:
        return f"PDFTOTEXT_ERR={e}"

def is_artifact(path, text):
    p = str(path)
    if "/runtime/stroyka_estimates/" in p: return True
    if path.name.lower().startswith("drainage_estimate_"): return True
    h = low(text[:600])
    if "смета:" in h and "дренаж" in h: return True
    return False

def classify(path, text):
    t = low(text); name = low(path.name)
    geo = sum(1 for m in GEO_STRONG if m in t)
    drn = sum(1 for m in DRAINAGE_STRONG if m in t)
    if "отчет" in name or "отчёт" in name or "мистолово" in name: geo += 3
    if "дренаж" in name or "схема" in name: drn += 3
    if geo >= 3 and drn < 5: return "geology_report"
    if drn >= 2: return "drainage_scheme"
    if geo >= 3: return "geology_report"
    return "other_pdf"

def friendly(kind):
    return {"drainage_scheme":"Схема глубинного дренажа.pdf",
            "geology_report":"Отчет_Мистолово_03.26.pdf"}.get(kind, "source.pdf")

def find_user_pdfs():
    now = datetime.now().timestamp()
    candidates = []
    for raw in glob.glob("/var/lib/telegram-bot-api/*/documents/*.pdf"):
        p = Path(raw)
        try:
            if p.is_file() and now - p.stat().st_mtime <= 12*3600:
                candidates.append(p)
        except: pass
    out = []; seen = set()
    for p in sorted(set(candidates), key=lambda x: x.stat().st_mtime, reverse=True):
        txt = pdf_text(p)
        if is_artifact(p, txt): continue
        kind = classify(p, txt)
        if kind in ("drainage_scheme","geology_report") and kind not in seen:
            out.append({"path":p,"kind":kind,"name":friendly(kind),"text":txt,"chars":len(txt)})
            seen.add(kind)
    return out

# ---------------------------------------------------------------------------
# VAT helpers
# ---------------------------------------------------------------------------
def infer_vat(conn, tid, raw_in, result):
    texts = [raw_in or "", result or ""]
    for row in conn.execute(
        "SELECT action FROM task_history WHERE task_id=? ORDER BY rowid ASC", (tid,)
    ).fetchall():
        texts.append(str(row[0] or ""))
    t = low("\n".join(texts))
    if "topic2_vat_mode_confirmed:with_vat_22" in t: return "WITH_VAT_22"
    if "topic2_vat_mode_confirmed:without_vat" in t: return "WITHOUT_VAT"
    wo = ["без ндс","ндс не нужен","без налога","без учета ндс"]
    wv = ["с ндс","с учетом ндс","добавь ндс","посчитай с ндс"]
    if any(p in t for p in wo): return "WITHOUT_VAT"
    if any(p in t for p in wv): return "WITH_VAT_22"
    return None

# ---------------------------------------------------------------------------
# Messaging
# ---------------------------------------------------------------------------
def send_msg(chat_id, topic_id, text):
    from core.reply_sender import send_reply_ex
    if len(text) > 3900: text = text[:3800]+"\n\n[сокращено]"
    res = send_reply_ex(chat_id=str(chat_id), text=text,
                        message_thread_id=int(topic_id) if int(topic_id) else None)
    if not res.get("ok"): raise RuntimeError(f"SEND_FAILED:{res}")
    return int(res.get("bot_message_id") or 0)

def ask_vat(conn, task):
    tid=str(task["id"]); chat_id=str(task["chat_id"]); topic_id=int(task["topic_id"] or 2)
    msg = "Считать с НДС 22% или без НДС?"
    bot_msg = send_msg(chat_id, topic_id, msg)
    conn.execute("UPDATE tasks SET state='WAITING_CLARIFICATION',result=?,bot_message_id=?,"
                 "error_message='TOPIC2_VAT_MODE_REQUIRED',updated_at=datetime('now') WHERE id=?",
                 (msg, bot_msg, tid))
    hist(conn, tid, "TOPIC2_VAT_GATE_CHECKED")
    hist(conn, tid, "TOPIC2_VAT_MODE_REQUIRED")
    hist(conn, tid, f"TOPIC2_VAT_QUESTION_SENT:{bot_msg}")
    conn.commit()
    print(f"VAT_MODE_REQUIRED\nBOT_MESSAGE_ID={bot_msg}")

# ---------------------------------------------------------------------------
# Length extraction (PDF + user reply in recent tasks)
# ---------------------------------------------------------------------------
def num(x): return float(x.replace(",","."))

def extract_lengths_from_pdf(text):
    LEGEND_SKIP = ("уклон","длина","диаметр")
    vals = []
    for line in text.splitlines():
        ll = low(line)
        if all(k in ll for k in LEGEND_SKIP): continue
        if not any(k in ll for k in ["i=","d=","дрен","водоотвод","труб","ливнев"]): continue
        for pat in [r"(?i)\bl\s*=\s*(\d+(?:[,.]\d+)?)\s*м\b",
                    r"(?i)длина\s*[-:=]?\s*(\d+(?:[,.]\d+)?)\s*м\b"]:
            for m in re.finditer(pat, line):
                try:
                    v = num(m.group(1))
                    if 0.5 <= v <= 500 and v not in vals: vals.append(round(v,2))
                except: pass
    return vals

def extract_depths(text):
    vals = []
    for pat in [r"(?i)на глубине\s*(\d+(?:[,.]\d+)?)\s*м",
                r"(?i)глубин[а-я]*\s*(?:до|от)?\s*(\d+(?:[,.]\d+)?)\s*м"]:
        for m in re.finditer(pat, text):
            try:
                v = num(m.group(1))
                if 0.2 <= v <= 12 and v not in vals: vals.append(round(v,2))
            except: pass
    return vals

def count_unique(text, prefix):
    return len(set(re.findall(rf"{re.escape(prefix)}\s*[-–]?\s*(\d+)", text, flags=re.I)))

def has(text, marker): return low(marker) in low(text)

def read_user_provided_length(conn, tid):
    """Check task_history and recent topic_2 tasks for user-provided length."""
    # Check history markers first
    for row in conn.execute(
        "SELECT action FROM task_history WHERE task_id=? ORDER BY rowid DESC LIMIT 50", (tid,)
    ).fetchall():
        a = str(row[0] or "")
        m = re.match(r"USER_PROVIDED_LENGTH:(\d+(?:\.\d+)?)", a)
        if m:
            return float(m.group(1))
    # Check recent topic_2 text tasks (user's reply after WC was sent)
    rows = conn.execute(
        "SELECT raw_input FROM tasks WHERE topic_id=2 AND id!=? "
        "AND input_type='text' ORDER BY rowid DESC LIMIT 15",
        (tid,),
    ).fetchall()
    for row in rows:
        text = str(row[0] or "").lower()
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:м(?:\.|\.п\.|\s|$)|метр)", text)
        if m:
            try:
                v = float(m.group(1).replace(",","."))
                if 5 <= v <= 2000:
                    return v
            except: pass
    return 0.0

# ---------------------------------------------------------------------------
# History state readers
# ---------------------------------------------------------------------------
def read_history_markers(conn, tid):
    rows = conn.execute(
        "SELECT action FROM task_history WHERE task_id=? ORDER BY rowid ASC", (tid,)
    ).fetchall()
    return [str(r[0] or "") for r in rows]

def find_marker(markers, prefix):
    for m in reversed(markers):
        if m.startswith(prefix): return m
    return ""

# ---------------------------------------------------------------------------
# Read recent user reply (for price choice)
# ---------------------------------------------------------------------------
def read_recent_user_reply(conn, tid):
    """Return most recent user text input in topic_2 (not the parent task)."""
    rows = conn.execute(
        "SELECT raw_input FROM tasks WHERE topic_id=2 AND id!=? "
        "AND input_type='text' ORDER BY rowid DESC LIMIT 10",
        (tid,),
    ).fetchall()
    for row in rows:
        text = str(row[0] or "").strip()
        if text: return text
    return ""

# ---------------------------------------------------------------------------
# Cache file (stores item definitions + offers between script runs)
# ---------------------------------------------------------------------------
def save_cache(cache: dict):
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2))

def load_cache() -> Optional[dict]:
    if CACHE_FILE.exists():
        try: return json.loads(CACHE_FILE.read_text())
        except: pass
    return None

# ---------------------------------------------------------------------------
# Drainage item definitions
# (name_xlsx, search_query, unit, work_price, раздел)
# qty_fn(L, ex, wd, wl) computed in build_cache
# ---------------------------------------------------------------------------
def build_drainage_cache(L, ex, wd, wl, has_dns, has_pu, sources):
    """Build the canonical cache dict for price enrichment."""
    items = []

    # Material items that need online price search
    mat_defs = [
        ("Геотекстиль в траншее",
         "Геотекстиль нетканый 150-200 г/м² Terram Typar ТехноНИКОЛЬ",
         "м²", round(L*1.8,2), 180, "Геотекстиль / щебень / песок"),
        ("Песчаная подготовка",
         "Песок строительный намывной",
         "м³", round(L*0.12,2), 1300, "Геотекстиль / щебень / песок"),
        ("Щебёночный фильтр 20-40мм",
         "Щебень гранитный фракция 20-40мм",
         "м³", round(L*0.35,2), 1600, "Геотекстиль / щебень / песок"),
        ("Труба дренажная/водоотводящая d=160",
         "Труба дренажная гофрированная двустенная d=160мм КОРСИС SN8",
         "п.м.", L, 850, "Дренажные трубы и обратный фильтр"),
    ]
    if wd:
        mat_defs.append((
            "Дренажный ревизионный колодец Дк ∅500",
            "Колодец дренажный ревизионный диаметр 500мм полимерный Wavin Политрон",
            "шт", float(wd), 6500, "Колодцы и дождеприёмники",
        ))
    if wl:
        mat_defs.append((
            "Ливневый ревизионный колодец Лк ∅500",
            "Колодец ливневый ревизионный диаметр 500мм полимерный",
            "шт", float(wl), 6500, "Колодцы и дождеприёмники",
        ))
    if has_dns:
        mat_defs.append((
            "Дренажная насосная станция ДНС-1",
            "Дренажная насосная станция 0.55кВт Grundfos Unilift Wilo Джилекс",
            "шт", 1.0, 28000, "ДНС / насосное оборудование",
        ))
    if has_pu:
        mat_defs.append((
            "Пескоуловитель ПУ-1",
            "Пескоуловитель дорожный пластиковый ПУ-1 Ecoteck Gidrostroy",
            "шт", 1.0, 6500, "Пескоуловители / линейный водоотвод",
        ))
    mat_defs.append((
        "Линейный водоотвод / лотки DN100",
        "Лоток водоотводный пластиковый DN100 с решёткой Hauraton Gidrostroy",
        "п.м.", max(round(L*0.2,2), 1.0), 1100, "Ливневая канализация",
    ))

    for (name, search, unit, qty, work_price, раздел) in mat_defs:
        items.append({
            "name": name, "search": search, "unit": unit,
            "qty": qty, "work_price": float(work_price),
            "раздел": раздел, "offers": [],
        })

    # Pure work items (no material price search)
    work_defs = [
        ("Разметка трасс дренажа/ливнёвки",        "м.п.",   L,               450.0,     0.0, "Подготовительные и земляные работы"),
        ("Разработка траншей",                       "м³",     ex,             1900.0,     0.0, "Подготовительные и земляные работы"),
        ("Вывоз/перемещение лишнего грунта",         "м³",     round(ex*0.35,2),1400.0,    0.0, "Подготовительные и земляные работы"),
        ("Укладка трубы с уклоном i=0,005",         "м.п.",   L,               950.0,     0.0, "Дренажные трубы и обратный фильтр"),
        ("Сборка узлов, подключение колодцев",      "компл",  1.0,           45000.0,     0.0, "Монтажные работы"),
        ("Доставка материалов и инструмента",        "рейс",   2.0,               0.0, 18000.0, "Логистика"),
    ]
    for (name, unit, qty, work_price, mat_price, раздел) in work_defs:
        items.append({
            "name": name, "search": None, "unit": unit,
            "qty": qty, "work_price": work_price,
            "раздел": раздел, "offers": [],
            "_fixed_mat_price": mat_price,  # for delivery etc.
        })

    return {
        "length": L, "ex": ex, "wd": wd, "wl": wl,
        "has_dns": has_dns, "has_pu": has_pu,
        "sources": sources,
        "items": items,
    }

# ---------------------------------------------------------------------------
# Price enrichment: call _openrouter_price_search for each material item
# ---------------------------------------------------------------------------
async def enrich_cache(conn, tid, cache):
    hist(conn, tid, "TOPIC2_PRICE_ENRICHMENT_STARTED")
    conn.commit()
    for item in cache["items"]:
        if not item.get("search"):
            continue
        name = item["name"]; unit = item["unit"]
        print(f"  SEARCHING: {name} ({unit})")
        try:
            offers = await asyncio.wait_for(
                _openrouter_price_search(item["search"], unit, "Санкт-Петербург"),
                timeout=30.0,
            )
        except Exception as e:
            print(f"  SEARCH_ERR {name}: {e}")
            offers = []
        item["offers"] = offers
        if offers:
            sup = offers[0].get("supplier","")
            hist(conn, tid, f"TOPIC2_PRICE_SOURCE_FOUND:{name}:{sup}")
            print(f"    → {len(offers)} offers, best: {offers[0].get('price')} {unit} @ {sup}")
        else:
            hist(conn, tid, f"TOPIC2_PRICE_SOURCE_MISSING:{name}")
            print(f"    → no offers found")
    hist(conn, tid, "TOPIC2_PRICE_ENRICHMENT_DONE")
    conn.commit()
    return cache

# ---------------------------------------------------------------------------
# Send price choice menu to user
# ---------------------------------------------------------------------------
def send_price_menu(conn, tid, chat_id, topic_id, cache):
    menu_text = _price_prompt(cache)
    bot_msg = send_msg(chat_id, topic_id, menu_text)
    conn.execute(
        "UPDATE tasks SET state='WAITING_CLARIFICATION', result=?, bot_message_id=?, "
        "error_message='TOPIC2_DRAINAGE_PRICE_CHOICE_REQUIRED', updated_at=datetime('now') WHERE id=?",
        (menu_text, bot_msg, tid),
    )
    hist(conn, tid, f"TOPIC2_PRICE_CHOICE_MENU_SENT:{bot_msg}")
    conn.commit()
    print(f"PRICE_MENU_SENT BOT_MESSAGE_ID={bot_msg}")
    return bot_msg

# ---------------------------------------------------------------------------
# XLSX + PDF generation (only after TOPIC2_PRICE_CHOICE_CONFIRMED)
# ---------------------------------------------------------------------------
def build_xlsx_rows(cache, mode, vat_mode):
    """Apply selected prices and build full XLSX row dicts."""
    rows = []
    for i, item in enumerate(cache["items"], 1):
        offers = item.get("offers") or []
        fixed_mat = item.get("_fixed_mat_price", 0.0)

        if offers:
            mat_price = _select_price(offers, mode)
            best = offers[0]
            src = best.get("status", "UNVERIFIED")
            supplier = best.get("supplier", "—")
            url = best.get("url", "—")
            checked = best.get("checked_at", datetime.now().strftime("%Y-%m-%d"))
        else:
            mat_price = fixed_mat
            src = "MANUAL" if fixed_mat > 0 else "WORK_ONLY"
            supplier = "—"; url = "—"
            checked = datetime.now().strftime("%Y-%m-%d")

        qty = float(item["qty"])
        work = float(item["work_price"])
        rows.append({
            "№": i,
            "Раздел": item.get("раздел",""),
            "Наименование": item["name"],
            "Ед изм": item["unit"],
            "Кол-во": qty,
            "Цена работ": work,
            "Стоимость работ": round(qty*work, 2),
            "Цена материалов": mat_price,
            "Стоимость материалов": round(qty*mat_price, 2),
            "Всего": round(qty*(work+mat_price), 2),
            "Источник цены": src,
            "Поставщик": supplier,
            "URL": url,
            "checked_at": checked,
            "Примечание": f"mode={mode}",
        })
    return rows

def calc_totals(rows, vat_mode):
    works = sum(r["Стоимость работ"] for r in rows)
    mats  = sum(r["Стоимость материалов"] for r in rows)
    no_vat = works + mats
    vat = no_vat * VAT_RATE if vat_mode == "WITH_VAT_22" else 0.0
    return {"works":works,"mats":mats,"no_vat":no_vat,"vat":vat,"grand":no_vat+vat}

def create_xlsx(path, rows, meta, vat_mode, mode):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    H = ["№","Раздел","Наименование","Ед изм","Кол-во","Цена работ","Стоимость работ",
         "Цена материалов","Стоимость материалов","Всего","Источник цены","Поставщик","URL","checked_at","Примечание"]
    wb = Workbook(); ws = wb.active; ws.title = "DRAINAGE_CALC"
    ws["A1"] = "Смета: дренаж / ливневая канализация / наружные сети"
    ws["A2"] = f"Исходные файлы: {', '.join(meta['file_names'])}"
    ws["A3"] = f"Длина: {meta['total_len']} м; глубина: {meta['avg_depth']} м"
    ws["A4"] = f"Цены: онлайн-поиск OpenRouter/Sonar, выбор пользователя: {mode}"
    ws["A5"] = "НДС: 22%" if vat_mode=="WITH_VAT_22" else "НДС: не применяется"
    for r in range(1,6): ws.cell(r,1).font = Font(bold=True)
    start = 7
    for c,h in enumerate(H,1):
        cell = ws.cell(start,c,h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid",fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    for r,row in enumerate(rows, start+1):
        for c,h in enumerate(H,1): ws.cell(r,c,row[h])
        ws.cell(r,7,f"=E{r}*F{r}"); ws.cell(r,9,f"=E{r}*H{r}"); ws.cell(r,10,f"=G{r}+I{r}")
    last = start+len(rows); tr=last+2
    ws.cell(tr,2,"ИТОГО без НДС" if vat_mode=="WITH_VAT_22" else "ИТОГО")
    ws.cell(tr,7,f"=SUM(G{start+1}:G{last})")
    ws.cell(tr,9,f"=SUM(I{start+1}:I{last})")
    ws.cell(tr,10,f"=SUM(J{start+1}:J{last})")
    if vat_mode=="WITH_VAT_22":
        vr=tr+1; gr=vr+1
        ws.cell(vr,2,"НДС 22%"); ws.cell(vr,10,f"=J{tr}*0.22")
        ws.cell(gr,2,"ИТОГО с НДС"); ws.cell(gr,10,f"=J{tr}+J{vr}")
    else:
        vr=tr+1; gr=vr
        ws.cell(vr,2,"НДС не применяется"); ws.cell(vr,10,0)
    for r in range(tr,gr+1):
        for c in range(1,16): ws.cell(r,c).font = Font(bold=True)
    for i,w in enumerate([6,26,46,10,12,14,16,16,18,16,22,28,16,14,20],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    thin = Side(style="thin",color="999999")
    for row in ws.iter_rows(min_row=start,max_row=gr,min_col=1,max_col=15):
        for cell in row:
            cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)
            cell.alignment = Alignment(vertical="top",wrap_text=True)
    wb.save(path)

def create_pdf(path, rows, meta, totals, vat_mode, mode):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    fp = next((p for p in ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
               "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"] if Path(p).exists()), None)
    if fp: pdfmetrics.registerFont(TTFont("RU",fp)); font="RU"
    else: font="Helvetica"
    vat_label = "НДС: 22%" if vat_mode=="WITH_VAT_22" else "НДС: не применяется"
    doc = SimpleDocTemplate(str(path),pagesize=landscape(A4),leftMargin=18,rightMargin=18,topMargin=18,bottomMargin=18)
    sty = getSampleStyleSheet()
    N = ParagraphStyle("n",parent=sty["Normal"],fontName=font,fontSize=8,leading=10)
    T = ParagraphStyle("t",parent=sty["Title"],fontName=font,fontSize=14,leading=16)
    story = [
        Paragraph("Смета: дренаж / ливневая канализация / наружные сети",T),Spacer(1,8),
        Paragraph(f"Исходные файлы: {', '.join(meta['file_names'])}",N),
        Paragraph(f"Длина: {meta['total_len']} м; глубина: {meta['avg_depth']} м",N),
        Paragraph(f"Цены: онлайн-поиск OpenRouter/Sonar, режим: {mode}; {vat_label}",N),
        Spacer(1,8),
    ]
    data=[["Раздел","Наименование","Ед","Кол-во","Работы","Материалы","Всего"]]
    for r in rows:
        data.append([
            Paragraph(r["Раздел"],N), Paragraph(r["Наименование"],N), r["Ед изм"],
            f"{r['Кол-во']:.1f}",
            f"{r['Стоимость работ']:,.0f}".replace(",","  "),
            f"{r['Стоимость материалов']:,.0f}".replace(",","  "),
            f"{r['Всего']:,.0f}".replace(",","  "),
        ])
    table=Table(data,colWidths=[105,230,42,55,75,85,75])
    table.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),font),("FONTSIZE",(0,0),(-1,-1),7),
        ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
        ("GRID",(0,0),(-1,-1),0.25,colors.grey),("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    story += [table, Spacer(1,8),
              Paragraph(f"Материалы: {totals['mats']:,.0f} руб".replace(",","  "),N),
              Paragraph(f"Работы: {totals['works']:,.0f} руб".replace(",","  "),N)]
    if vat_mode=="WITH_VAT_22":
        story += [
            Paragraph(f"Без НДС: {totals['no_vat']:,.0f} руб".replace(",","  "),N),
            Paragraph(f"НДС 22%: {totals['vat']:,.0f} руб".replace(",","  "),N),
            Paragraph(f"С НДС: {totals['grand']:,.0f} руб".replace(",","  "),N),
        ]
    else:
        story += [
            Paragraph(f"Итого без НДС: {totals['grand']:,.0f} руб".replace(",","  "),N),
            Paragraph("НДС: не применяется",N),
        ]
    doc.build(story)

def send_doc(chat_id, topic_id, path, caption):
    token = os.getenv("TELEGRAM_BOT_TOKEN","").strip()
    if not token: raise RuntimeError("TELEGRAM_BOT_TOKEN_MISSING")
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(path,"rb") as fh:
        data={"chat_id":str(chat_id),"caption":caption[:900]}
        if int(topic_id)!=0: data["message_thread_id"]=str(int(topic_id))
        r=requests.post(url,data=data,files={"document":(path.name,fh)},timeout=120)
    js=r.json()
    if r.status_code!=200 or not js.get("ok"):
        raise RuntimeError(f"SEND_DOC_FAILED:{r.status_code}:{r.text[:200]}")
    return int(js["result"]["message_id"])

async def maybe_upload(path, chat_id, topic_id):
    import inspect
    try: from core.topic_drive_oauth import upload_file_to_topic
    except: return ""
    for fn in [
        lambda: upload_file_to_topic(file_path=str(path),file_name=path.name,chat_id=str(chat_id),topic_id=int(topic_id)),
        lambda: upload_file_to_topic(str(path),path.name,str(chat_id),int(topic_id)),
    ]:
        try:
            res = fn()
            if inspect.isawaitable(res): res = await res
            if isinstance(res,dict):
                for k in ("webViewLink","link","url","drive_link","view_link"):
                    if res.get(k): return str(res[k])
                if res.get("file_id"): return f"https://drive.google.com/file/d/{res['file_id']}/view"
            if isinstance(res,str) and res.startswith("http"): return res
        except: continue
    return ""

# ---------------------------------------------------------------------------
# Main state machine
# ---------------------------------------------------------------------------
async def main():
    conn = sqlite3.connect(str(DB)); conn.row_factory = sqlite3.Row
    task = conn.execute("SELECT * FROM tasks WHERE id=? LIMIT 1",(TASK_ID,)).fetchone()
    if not task: raise SystemExit(f"TASK_NOT_FOUND:{TASK_ID}")
    tid = str(task["id"]); chat_id = str(task["chat_id"]); topic_id = int(task["topic_id"] or 2)
    raw_in = str(task["raw_input"] or ""); result = str(task["result"] or "")

    # VAT gate
    vat_mode = infer_vat(conn, tid, raw_in, result)
    if vat_mode is None:
        ask_vat(conn, task); conn.close(); return

    markers = read_history_markers(conn, tid)

    # ── STATE: TOPIC2_PRICE_CHOICE_CONFIRMED exists → generate estimate ──
    confirmed_marker = find_marker(markers, "TOPIC2_PRICE_CHOICE_CONFIRMED:")
    if confirmed_marker:
        mode = confirmed_marker.split(":", 1)[1].strip()
        print(f"PRICE_CHOICE_CONFIRMED:{mode} — generating estimate")
        cache = load_cache()
        if cache is None:
            raise SystemExit("PRICE_CACHE_FILE_MISSING — re-run from price search state")
        await _generate_and_send(conn, tid, chat_id, topic_id, vat_mode, mode, cache, markers)
        conn.close(); return

    # ── STATE: price menu already sent → check for user reply ──
    menu_marker = find_marker(markers, "TOPIC2_PRICE_CHOICE_MENU_SENT:")
    if menu_marker:
        user_reply = read_recent_user_reply(conn, tid)
        if user_reply:
            choice = _detect_price_choice(user_reply)
            if choice:
                print(f"USER_CHOICE_DETECTED:{choice} from '{user_reply[:40]}'")
                hist(conn, tid, f"TOPIC2_PRICE_CHOICE_CONFIRMED:{choice}")
                conn.commit()
                cache = load_cache()
                if cache is None:
                    raise SystemExit("PRICE_CACHE_FILE_MISSING")
                await _generate_and_send(conn, tid, chat_id, topic_id, vat_mode, choice, cache, markers)
                conn.close(); return
            # Check if it's a length (user replied to wrong WC)
            m = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:м\b|метр)", user_reply.lower())
            if m:
                try:
                    L_reply = float(m.group(1).replace(",","."))
                    if 5 <= L_reply <= 2000:
                        print(f"LENGTH_FROM_USER_REPLY:{L_reply} — rebuilding cache")
                        hist(conn, tid, f"USER_PROVIDED_LENGTH:{L_reply}")
                        conn.commit()
                        await _do_price_search(conn, tid, chat_id, topic_id, L_reply, vat_mode)
                        conn.close(); return
                except: pass
        print("WAITING_FOR_PRICE_CHOICE — no actionable reply yet")
        conn.close(); return

    # ── STATE: find length ──
    sources = find_user_pdfs()
    if not sources:
        raise SystemExit("NO_USER_SOURCE_PDFS")
    drainage = [x for x in sources if x["kind"]=="drainage_scheme"]
    if not drainage:
        raise SystemExit("DRAINAGE_SOURCE_NOT_FOUND")

    scheme_text = "\n".join(x["text"] for x in drainage)
    geo_text    = "\n".join(x["text"] for x in sources if x["kind"]=="geology_report")

    pdf_lengths = extract_lengths_from_pdf(scheme_text)
    total_len   = round(sum(pdf_lengths), 2)
    print(f"PDF_LENGTHS={pdf_lengths} TOTAL_LEN={total_len}")

    if total_len <= 0:
        user_len = read_user_provided_length(conn, tid)
        if user_len > 0:
            print(f"USER_PROVIDED_LENGTH:{user_len}")
            hist(conn, tid, f"USER_PROVIDED_LENGTH:{user_len}")
            conn.commit()
            total_len = user_len
        else:
            # Ask user for length
            wc_msg = (
                "Длина трасс дренажа и ливнёвки в PDF не читается — схема графическая.\n\n"
                "Пришли, пожалуйста:\n"
                "• Общую длину дренажных труб (в метрах)\n"
                "• Или длины по участкам: Дк-1→Дк-2, Дк-2→ДНС, и т.д.\n\n"
                "После этого запрошу актуальные цены и покажу смету."
            )
            bot_msg = send_msg(chat_id, topic_id, wc_msg)
            conn.execute(
                "UPDATE tasks SET state='WAITING_CLARIFICATION', result=?, bot_message_id=?, "
                "error_message='TOPIC2_DRAINAGE_LENGTH_NOT_PROVEN', updated_at=datetime('now') WHERE id=?",
                (wc_msg, bot_msg, tid),
            )
            for a in ["TOPIC2_DRAINAGE_LENGTH_PROOF_GATE_V1",
                      f"TOPIC2_DRAINAGE_LENGTH_NOT_PROVEN:lines={len(pdf_lengths)}:total={total_len}",
                      "TOPIC2_DRAINAGE_FINAL_ARTIFACTS_BLOCKED",
                      f"TOPIC2_DRAINAGE_WC_SENT:{bot_msg}"]:
                hist(conn, tid, a)
            conn.commit(); conn.close()
            print(f"DRAINAGE_LENGTH_GATE_WC_SENT BOT_MESSAGE_ID={bot_msg}")
            return

    await _do_price_search(conn, tid, chat_id, topic_id, total_len, vat_mode)
    conn.close()


async def _do_price_search(conn, tid, chat_id, topic_id, L, vat_mode):
    """Search prices via canonical _openrouter_price_search + send menu."""
    task = conn.execute("SELECT * FROM tasks WHERE id=? LIMIT 1",(tid,)).fetchone()
    sources = find_user_pdfs()
    drainage = [x for x in sources if x["kind"]=="drainage_scheme"]
    geo      = [x for x in sources if x["kind"]=="geology_report"]
    scheme_text = "\n".join(x["text"] for x in drainage)
    geo_text    = "\n".join(x["text"] for x in geo)
    depths = extract_depths(geo_text)
    avg_depth = round(max(1.2, min(depths)), 2) if depths else 1.2
    ex = round(L * avg_depth * 0.6, 2)
    wd = count_unique(scheme_text, "Дк")
    wl = count_unique(scheme_text, "Лк")
    has_dns = has(scheme_text, "ДНС")
    has_pu  = has(scheme_text, "пескоуловитель") or has(scheme_text, "ПУ-1")

    print(f"LENGTH={L} DEPTH={avg_depth} WELLS_DK={wd} WELLS_LK={wl} DNS={has_dns} PU={has_pu}")

    cache = build_drainage_cache(L, ex, wd, wl, has_dns, has_pu,
                                  [x["name"] for x in sources])
    cache = await enrich_cache(conn, tid, cache)
    save_cache(cache)
    print(f"CACHE_SAVED:{CACHE_FILE}")

    hist(conn, tid, f"TOPIC2_DRAINAGE_LENGTHS_STATUS:PROVEN:total_len={L}")
    hist(conn, tid, f"TOPIC2_DRAINAGE_VAT_MODE:{vat_mode}")
    conn.commit()

    send_price_menu(conn, tid, chat_id, topic_id, cache)


async def _generate_and_send(conn, tid, chat_id, topic_id, vat_mode, mode, cache, markers):
    """Generate XLSX/PDF after confirmed price choice and send to Telegram."""
    hist(conn, tid, "TOPIC2_DRAINAGE_GENERATE_STARTED")
    conn.commit()

    rows    = build_xlsx_rows(cache, mode, vat_mode)
    totals  = calc_totals(rows, vat_mode)
    sources = cache.get("sources", [])
    L       = cache["length"]
    depth   = cache.get("ex",0) / (cache["length"] * 0.6) if cache["length"] else 1.2
    meta    = {"file_names": sources, "total_len": L, "avg_depth": round(depth,2)}

    stamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx    = OUT_DIR / f"drainage_estimate_clean_{tid[:8]}_{stamp}.xlsx"
    pdf_out = OUT_DIR / f"drainage_estimate_clean_{tid[:8]}_{stamp}.pdf"

    create_xlsx(xlsx, rows, meta, vat_mode, mode)
    create_pdf(pdf_out, rows, meta, totals, vat_mode, mode)

    xlsx_link = await maybe_upload(xlsx, chat_id, topic_id)
    pdf_link  = await maybe_upload(pdf_out, chat_id, topic_id)
    if not xlsx_link:
        mid = send_doc(chat_id, topic_id, xlsx, "Excel: смета дренажа"); print(f"XLSX_DOC_SENT:{mid}")
    if not pdf_link:
        mid = send_doc(chat_id, topic_id, pdf_out, "PDF: смета дренажа"); print(f"PDF_DOC_SENT:{mid}")

    if vat_mode=="WITH_VAT_22":
        totals_block=(f"Без НДС: {totals['no_vat']:,.0f} руб\n"
                      f" НДС 22%: {totals['vat']:,.0f} руб\n"
                      f" С НДС: {totals['grand']:,.0f} руб").replace(",","  ")
    else:
        totals_block=f"Итого без НДС: {totals['grand']:,.0f} руб\n НДС: не применяется".replace(",","  ")

    excel_line = f"Excel: {xlsx_link}" if xlsx_link else "Excel: отправлен файлом"
    pdf_line   = f"PDF: {pdf_link}"   if pdf_link  else "PDF: отправлен файлом"

    public = (
        f"✅ Смета дренажа готова\n\n"
        f"Объект: наружные сети / дренаж / ливневая канализация\n"
        f"Файлы учтены: {', '.join(sources)}\n"
        f"Цены: онлайн-поиск OpenRouter/Sonar, режим: {mode}\n"
        f"Длина: {L} м\n\n"
        f"Итого:\n Материалы: {totals['mats']:,.0f} руб\n"
        f" Работы: {totals['works']:,.0f} руб\n {totals_block}\n\n"
        f"{excel_line}\n{pdf_line}\n\nПодтверди или пришли правки"
    ).replace(",","  ")

    dirty = [x for x in ["/root/","runtime","drainage_estimate_"] if x in public]
    if dirty: raise SystemExit(f"PUBLIC_OUTPUT_DIRTY:{dirty}")

    bot_msg = send_msg(chat_id, topic_id, public)
    conn.execute(
        "UPDATE tasks SET state='AWAITING_CONFIRMATION', result=?, bot_message_id=?, "
        "error_message=NULL, updated_at=datetime('now') WHERE id=?",
        (public, bot_msg, tid),
    )
    for action in [
        f"TOPIC2_DRAINAGE_SOURCE_FILTER_OK:user_pdfs={len(sources)}",
        "TOPIC2_DRAINAGE_NO_GENERATED_ARTIFACT_INPUT",
        f"TOPIC2_DRAINAGE_PRICES_SOURCE:OpenRouter/Sonar:mode={mode}",
        f"TOPIC2_DRAINAGE_XLSX_CREATED:{xlsx.name}",
        f"TOPIC2_DRAINAGE_PDF_CREATED:{pdf_out.name}",
        f"TOPIC2_DRAINAGE_DRIVE_XLSX_OK:{xlsx_link}" if xlsx_link else "TOPIC2_DRAINAGE_TELEGRAM_XLSX_FALLBACK_SENT",
        f"TOPIC2_DRAINAGE_DRIVE_PDF_OK:{pdf_link}"   if pdf_link  else "TOPIC2_DRAINAGE_TELEGRAM_PDF_FALLBACK_SENT",
        f"TOPIC2_DRAINAGE_TELEGRAM_SENT:{bot_msg}",
        "TOPIC2_VAT_PUBLIC_OUTPUT_OK",
        "TOPIC2_DRAINAGE_AWAITING_CONFIRMATION_CLEAN_V1",
    ]:
        hist(conn, tid, action)
    conn.commit()
    print(f"DRAINAGE_ESTIMATE_OK BOT_MESSAGE_ID={bot_msg} GRAND={totals['grand']}")


if __name__ == "__main__":
    asyncio.run(main())
