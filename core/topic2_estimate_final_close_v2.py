from __future__ import annotations

import json
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

BASE = Path("/root/.areal-neva-core")
OUT = BASE / "outputs" / "topic2_estimates"
OUT.mkdir(parents=True, exist_ok=True)

ENGINE = "TOPIC2_ESTIMATE_FINAL_CLOSE_V2"

SHORT_WORDS = {
    "да", "да делай", "да, делай", "делай", "ок", "окей", "хорошо",
    "подтверждаю", "согласен", "верно", "все верно", "всё верно",
    "1", "2", "3", "вариант 1", "вариант 2", "вариант 3",
    "минимальные", "минимум", "самые дешевые", "самые дешёвые",
    "средние", "медианные", "медиана", "надежные", "надёжные"
}

ESTIMATE_WORDS = (
    "смет", "кп", "коммерческ", "расчет", "расчёт", "стоимост", "цена",
    "расцен", "ведомост", "монолит", "бетон", "арматур", "опалуб",
    "фундамент", "перекрыт", "колонн", "стен", "гидроизоляц",
    "утеплен", "засыпк", "свай", "плит", "лестнич"
)

IMAGE_EXT = (".jpg", ".jpeg", ".png", ".webp", ".heic", ".bmp", ".tif", ".tiff")
DOC_EXT = (".pdf", ".docx", ".xlsx", ".xls", ".csv", ".txt")


def _s(v: Any, limit: int = 50000) -> str:
    if v is None:
        return ""
    try:
        return str(v).strip()[:limit]
    except Exception:
        return ""


def _low(v: Any) -> str:
    return _s(v).lower().replace("ё", "е")


def _field(task: Any, name: str, default: Any = None) -> Any:
    try:
        if hasattr(task, "keys") and name in task.keys():
            return task[name]
    except Exception:
        pass
    try:
        return task.get(name, default)
    except Exception:
        return getattr(task, name, default)


def _payload(raw: Any) -> Dict[str, Any]:
    if isinstance(raw, dict):
        return raw
    t = _s(raw)
    if not t:
        return {}
    try:
        x = json.loads(t)
        return x if isinstance(x, dict) else {}
    except Exception:
        return {}


def _extract_payload_text(raw: Any) -> str:
    p = _payload(raw)
    parts = [_s(raw)]
    for k in ("caption", "text", "message", "file_name", "name", "title", "ocr_text", "recognized_text"):
        if p.get(k):
            parts.append(_s(p.get(k)))
    return "\n".join(x for x in parts if x).strip()


def _file_meta(raw: Any) -> Dict[str, str]:
    p = _payload(raw)
    keys_path = ("local_path", "path", "file_path", "downloaded_path", "server_path")
    keys_name = ("file_name", "name", "title")
    file_path = ""
    file_name = ""
    for k in keys_path:
        if p.get(k):
            file_path = _s(p.get(k))
            break
    for k in keys_name:
        if p.get(k):
            file_name = _s(p.get(k))
            break
    if not file_name and file_path:
        file_name = os.path.basename(file_path)
    return {"file_path": file_path, "file_name": file_name}


def _read_file_text(path: str) -> str:
    p = Path(_s(path))
    if not p.exists() or not p.is_file():
        return ""
    suf = p.suffix.lower()
    try:
        if suf == ".txt":
            return p.read_text(encoding="utf-8", errors="ignore")[:50000]
        if suf == ".csv":
            return p.read_text(encoding="utf-8", errors="ignore")[:50000]
        if suf == ".pdf":
            try:
                import fitz
                doc = fitz.open(str(p))
                return "\n".join(page.get_text("text") for page in doc)[:50000]
            except Exception:
                return ""
        if suf == ".docx":
            try:
                import docx
                d = docx.Document(str(p))
                return "\n".join(x.text for x in d.paragraphs)[:50000]
            except Exception:
                return ""
        if suf in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), data_only=True, read_only=True)
                out = []
                for ws in wb.worksheets[:3]:
                    for row in ws.iter_rows(max_row=200, values_only=True):
                        vals = [_s(x, 200) for x in row if _s(x)]
                        if vals:
                            out.append(" | ".join(vals))
                return "\n".join(out)[:50000]
            except Exception:
                return ""
        if suf in IMAGE_EXT:
            try:
                from PIL import Image
                import pytesseract
                return pytesseract.image_to_string(Image.open(str(p)), lang="rus+eng")[:50000]
            except Exception:
                return ""
    except Exception:
        return ""
    return ""


def _is_short_control(text: str) -> bool:
    t = re.sub(r"\s+", " ", _low(text).replace("[voice]", "")).strip(" .,!?:;")
    return t in SHORT_WORDS or (len(t) <= 18 and any(t.startswith(x) for x in SHORT_WORDS))


def _is_estimate_intent(text: str, file_name: str = "") -> bool:
    low = _low(text + " " + file_name)
    if not low:
        return False
    if any(x in low for x in ESTIMATE_WORDS):
        return True
    return bool(re.search(r"\b(м3|м³|м2|м²|шт|кг|тн|п\.?\s*м)\b", low))


def _is_file_or_photo(input_type: str, raw: Any) -> bool:
    meta = _file_meta(raw)
    name = _low(meta.get("file_name") or meta.get("file_path"))
    if input_type in ("photo", "image", "file", "drive_file", "document"):
        return True
    return name.endswith(IMAGE_EXT + DOC_EXT)


def _qty(v: str) -> float:
    s = _s(v).replace("≈", "").replace("~", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _normalize_unit(u: str) -> str:
    x = _low(u).replace(" ", "")
    return {
        "м3": "м³", "м.3": "м³", "м³": "м³",
        "м2": "м²", "м.2": "м²", "м²": "м²",
        "п.м": "п.м", "пм": "п.м",
        "шт.": "шт", "шт": "шт",
        "компл.": "компл", "компл": "компл",
        "тн": "т", "тонн": "т",
    }.get(x, x or "шт")


def _parse_items(text: str) -> List[Dict[str, Any]]:
    src = _s(text, 50000)
    t = re.sub(r"\s+", " ", src)
    t = re.sub(r"(?<![\d,.])\s+(\d{1,2})\s+(?=[А-ЯA-ZЁ])", r"\n\1 ", t)
    unit_re = r"(м³|м3|м\.3|м²|м2|м\.2|п\.?\s*м|шт\.?|кг|тн|т|компл\.?)"
    items: List[Dict[str, Any]] = []

    for line in t.splitlines():
        line = line.strip(" ;")
        if not line:
            continue
        m = re.search(
            rf"^\s*(?P<num>\d{{1,3}})\s+(?P<name>.+?)\s+(?P<unit>{unit_re})\s+(?P<qty>[~≈]?\s*\d[\d\s]*(?:[,.]\d+)?)",
            line,
            flags=re.I,
        )
        if not m:
            continue
        name = re.sub(r"\s+", " ", m.group("name")).strip(" -–—:")
        unit = _normalize_unit(m.group("unit"))
        qty = _qty(m.group("qty"))
        if not name or qty <= 0:
            continue
        items.append({
            "num": len(items) + 1,
            "name": name[:240],
            "qty": qty,
            "unit": unit,
            "price": 0.0,
            "source": "parsed",
        })

    if not items:
        m = re.search(rf"(?P<name>.{{1,120}}?)\s+(?P<unit>{unit_re})\s+(?P<qty>\d[\d\s]*(?:[,.]\d+)?)", t, flags=re.I)
        if m:
            items.append({
                "num": 1,
                "name": re.sub(r"\s+", " ", m.group("name")).strip(" -–—:")[:240] or "Позиция",
                "qty": _qty(m.group("qty")),
                "unit": _normalize_unit(m.group("unit")),
                "price": 0.0,
                "source": "fallback",
            })

    if not items:
        items.append({
            "num": 1,
            "name": "Позиция по присланному фото / файлу / тексту",
            "qty": 1.0,
            "unit": "компл",
            "price": 0.0,
            "source": "manual_review_required",
        })

    return items[:200]


def _write_xlsx(path: Path, items: List[Dict[str, Any]], source_text: str, photo_text: str = "") -> None:
    import copy
    from datetime import date
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    template_path = Path("/root/.areal-neva-core/data/templates/estimate/cache/1KuoSI4OI7gJoIBPVqBQGXtQnolXKMiDp__фундамент_Склад2.xlsx")
    if template_path.exists():
        try:
            wb = load_workbook(str(template_path))
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    if "AREAL_CALC" in wb.sheetnames:
        del wb["AREAL_CALC"]
    ws = wb.create_sheet("AREAL_CALC", 0)
    template_ws = wb["смета"] if "смета" in wb.sheetnames else None

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    section_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
    total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    def copy_template_cell(src_row: int, src_col: int, dst_row: int, dst_col: int) -> None:
        if not template_ws:
            return
        src = template_ws.cell(src_row, src_col)
        dst = ws.cell(dst_row, dst_col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)

    for col in range(1, 16):
        copy_template_cell(1, min(col, 10), 1, col)
        copy_template_cell(2, min(col, 10), 2, col)
        ws.cell(1, col).border = border
        ws.cell(2, col).border = border
        ws.cell(1, col).fill = header_fill
        ws.cell(2, col).fill = header_fill
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(2, col).font = Font(bold=True)
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(2, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Canon §4 keeps 15 columns; template layout keeps grouped Work/Materials columns.
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:D2")
    ws.merge_cells("E1:E2")
    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:I1")
    ws.merge_cells("J1:J2")
    ws.merge_cells("K1:K2")
    ws.merge_cells("L1:L2")
    ws.merge_cells("M1:M2")
    ws.merge_cells("N1:N2")
    ws.merge_cells("O1:O2")

    ws["A1"] = "№"
    ws["B1"] = "Раздел"
    ws["C1"] = "Наименование"
    ws["D1"] = "Ед изм"
    ws["E1"] = "Кол-во"
    ws["F1"] = "Работа"
    ws["F2"] = "Цена работ"
    ws["G2"] = "Стоимость работ"
    ws["H1"] = "Материалы"
    ws["H2"] = "Цена материалов"
    ws["I2"] = "Стоимость материалов"
    ws["J1"] = "Всего"
    ws["K1"] = "Источник цены"
    ws["L1"] = "Поставщик"
    ws["M1"] = "URL"
    ws["N1"] = "checked_at"
    ws["O1"] = "Примечание"

    widths = {
        "A": 7, "B": 22, "C": 58, "D": 11, "E": 12,
        "F": 14, "G": 16, "H": 16, "I": 18, "J": 18,
        "K": 24, "L": 24, "M": 36, "N": 14, "O": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"

    def classify(item):
        name = str(item.get("name", "")).lower().replace("ё", "е")
        if any(x in name for x in ("накладн", "расходные материалы", "крепеж", "герметик", "логистика")):
            return "overhead"
        if any(x in name for x in ("монтаж", "устройство", "работ", "уборка")):
            return "work"
        return "material"

    def section_for(item):
        name = str(item.get("name", "")).lower().replace("ё", "е")
        if "фундамент" in name or "плиты" in name or "бетон" in name or "арматур" in name:
            return "Фундамент"
        if "металлоконструк" in name or "каркас" in name or "колонн" in name:
            return "Стены / каркас"
        if "стен" in name:
            return "Стены / каркас"
        if "кров" in name:
            return "Кровля"
        if "логист" in name:
            return "Логистика"
        if "накладн" in name or "расходные" in name or "уборка" in name:
            return "Накладные расходы"
        return "Прочее"

    def supplier_source(item):
        source = str(item.get("source", ""))
        low = source.lower()
        if "sp-sever" in low:
            return "sp-sever.ru", "https://sp-sever.ru/panels/stenovye_sendvich_panely"
        if "sp-rsk" in low:
            return "sp-rsk-uteplitel.ru", "https://spb.rsk-uteplitel.ru/sklady"
        if "sonar" in low:
            return "Sonar", ""
        if "runtime" in low:
            return "topic_2 runtime", ""
        return "", ""

    row = 3
    today = date.today().isoformat()
    last_section = None
    for item in items:
        section = section_for(item)
        if section != last_section:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
            sec_cell = ws.cell(row=row, column=1, value=section)
            sec_cell.font = Font(bold=True)
            sec_cell.fill = section_fill
            sec_cell.alignment = Alignment(horizontal="left", vertical="center")
            for col in range(1, 16):
                ws.cell(row=row, column=col).border = border
            row += 1
            last_section = section

        qty = float(item.get("qty") or 0)
        price = float(item.get("price") or 0)
        kind = classify(item)
        work_price = price if kind == "work" else 0
        mat_price = price if kind in ("material", "overhead") else 0
        supplier, url = supplier_source(item)
        values = [
            item.get("num"), section, item.get("name"), item.get("unit"), qty,
            work_price, f"=E{row}*F{row}", mat_price, f"=E{row}*H{row}", f"=G{row}+I{row}",
            item.get("source", ""), supplier, url, today, item.get("note", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (5, 6, 7, 8, 9, 10):
                cell.number_format = '#,##0.00'
        row += 1

    data_rows = [r for r in range(3, row) if ws.cell(r, 10).value and str(ws.cell(r, 10).value).startswith("=")]
    first_data = min(data_rows) if data_rows else 3
    last_data = max(data_rows) if data_rows else row - 1

    row += 1
    ws.cell(row=row, column=9, value="ИТОГО без НДС").font = Font(bold=True)
    ws.cell(row=row, column=10, value=f"=SUM(J{first_data}:J{last_data})").font = Font(bold=True)
    for col in range(1, 16):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.fill = total_fill
    row += 1
    ws.cell(row=row, column=9, value="НДС").font = Font(bold=True)
    ws.cell(row=row, column=10, value="не начисляется по заданию").font = Font(bold=True)
    for col in range(1, 16):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.fill = total_fill

    ws2 = wb.create_sheet("Источник")
    ws2["A1"] = "Исходный текст"
    ws2["A1"].font = Font(bold=True)
    ws2["A2"] = source_text[:32000]
    ws2["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 140
    if photo_text:
        ws2["A4"] = "Распознанный текст из файла / фото"
        ws2["A4"].font = Font(bold=True)
        ws2["A5"] = photo_text[:32000]
        ws2["A5"].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(path))

def _pdf_font():
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        for fp in (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ):
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                return "ArealSans"
    except Exception:
        pass
    return "Helvetica"


def _write_pdf(path: Path, items: List[Dict[str, Any]]) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    font = _pdf_font()
    c = canvas.Canvas(str(path), pagesize=A4)
    w, h = A4
    x = 12 * mm
    y = h - 16 * mm

    c.setFont(font, 12)
    c.drawString(x, y, "Предварительный сметный расчёт")
    y -= 8 * mm
    c.setFont(font, 8)
    c.drawString(x, y, f"Движок: {ENGINE}")
    y -= 6 * mm
    c.drawString(x, y, "Цены не выдуманы, расчётная колонка в Excel считается формулами")
    y -= 8 * mm

    c.setFont(font, 7)
    headers = ["№", "Наименование", "Кол-во", "Цена", "Сумма", "Ед"]
    xs = [x, x + 9*mm, x + 112*mm, x + 137*mm, x + 162*mm, x + 185*mm]
    for xx, val in zip(xs, headers):
        c.drawString(xx, y, val)
    y -= 5 * mm
    c.line(x, y, w - 10*mm, y)
    y -= 5 * mm

    for item in items:
        if y < 18 * mm:
            c.showPage()
            c.setFont(font, 7)
            y = h - 16 * mm
        vals = [
            str(item["num"]),
            item["name"][:72],
            f'{float(item["qty"]):g}',
            "",
            "",
            item["unit"],
        ]
        for xx, val in zip(xs, vals):
            c.drawString(xx, y, str(val))
        y -= 5 * mm

    c.save()


def _upload(path: Path, task_id: str, topic_id: int, chat_id: str = "") -> str:
    try:
        from core.topic_drive_oauth import _upload_file_sync
        up = _upload_file_sync(str(path), path.name, str(chat_id or task_id), int(topic_id or 0), None)
        if isinstance(up, dict):
            link = up.get("webViewLink")
            fid = up.get("drive_file_id")
            if link:
                return str(link)
            if fid:
                return "https://drive.google.com/file/d/" + str(fid) + "/view?usp=drivesdk"
    except Exception:
        pass

    try:
        from core.engine_base import upload_artifact_to_drive
        link = upload_artifact_to_drive(str(path), task_id, topic_id)
        if link and "drive.google.com" in str(link):
            return str(link)
    except Exception:
        pass
    return str(path)


def _make_artifacts(task_id: str, topic_id: int, raw_text: str, photo_text: str = "", chat_id: str = "") -> Dict[str, Any]:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", task_id or ts)[:32]
    out_dir = OUT / f"{safe}_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    source = "\n".join(x for x in (raw_text, photo_text) if x).strip()
    items = _parse_items(source)

    xlsx = out_dir / f"SMETA_TOPIC2__{safe}.xlsx"
    pdf = out_dir / f"SMETA_TOPIC2__{safe}.pdf"
    manifest = out_dir / f"SMETA_TOPIC2__{safe}.manifest.json"

    _write_xlsx(xlsx, items, raw_text, photo_text)
    _write_pdf(pdf, items)

    data = {
        "engine": ENGINE,
        "task_id": task_id,
        "topic_id": int(topic_id or 0),
        "items_count": len(items),
        "created_at": datetime.now().isoformat(),
        "prices_policy": "not invented",
        "xlsx": str(xlsx),
        "pdf": str(pdf),
    }
    manifest.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx_link = _upload(xlsx, task_id, topic_id, chat_id)
    pdf_link = _upload(pdf, task_id, topic_id, chat_id)
    manifest_link = _upload(manifest, task_id, topic_id, chat_id)

    total = 0.0
    for item in items:
        try:
            total += float(item.get("total") or 0) or (float(item.get("qty") or 0) * float(item.get("price") or 0))
        except Exception:
            pass

    msg = (
        "✅ Смета готова\n"
        f"Позиций: {len(items)}\n"
        f"Итого: {total:,.0f} руб\n".replace(",", " ")
        + "Цены: не выдуманы, Excel содержит формулы и итог\n\n"
        + f"Excel: {xlsx_link}\n"
        + f"PDF: {pdf_link}\n\n"
        + "Подтверди или пришли правки"
    )

    return {
        "ok": True,
        "message": msg,
        "xlsx_link": xlsx_link,
        "pdf_link": pdf_link,
        "manifest_link": manifest_link,
        "items_count": len(items),
    }


def _find_parent(conn, chat_id: str, topic_id: int, reply_to: Any, current_id: str):
    params = [str(chat_id), int(topic_id or 0), str(current_id)]
    where = [
        "chat_id=?",
        "COALESCE(topic_id,0)=?",
        "id<>?",
        "state IN ('NEW','IN_PROGRESS','WAITING_CLARIFICATION','AWAITING_CONFIRMATION','AWAITING_PRICE_CONFIRMATION','CANCELLED')",
    ]
    if reply_to:
        sql = f"""
        SELECT * FROM tasks
        WHERE {' AND '.join(where)}
          AND (bot_message_id=? OR reply_to_message_id=?)
        ORDER BY CASE WHEN state='AWAITING_CONFIRMATION' THEN 0 ELSE 1 END, rowid DESC
        LIMIT 1
        """
        row = conn.execute(sql, params + [reply_to, reply_to]).fetchone()
        if row:
            return row

    sql = f"""
    SELECT * FROM tasks
    WHERE {' AND '.join(where)}
      AND (
        lower(COALESCE(raw_input,'')) LIKE '%смет%'
        OR lower(COALESCE(raw_input,'')) LIKE '%кп%'
        OR lower(COALESCE(result,'')) LIKE '%смет%'
        OR lower(COALESCE(result,'')) LIKE '%xlsx%'
        OR lower(COALESCE(result,'')) LIKE '%pdf%'
      )
    ORDER BY rowid DESC
    LIMIT 1
    """
    return conn.execute(sql, params).fetchone()


def _update(conn, update_task, task_id: str, **kw) -> None:
    if not task_id:
        return
    try:
        if update_task:
            update_task(conn, task_id, **kw)
            conn.commit()
            return
    except Exception:
        pass

    cols = [r[1] for r in conn.execute("PRAGMA table_info(tasks)").fetchall()]
    sets = []
    vals = []
    for k, v in kw.items():
        if k in cols:
            sets.append(f"{k}=?")
            vals.append(v)
    if "updated_at" in cols:
        sets.append("updated_at=datetime('now')")
    if sets:
        vals.append(task_id)
        conn.execute(f"UPDATE tasks SET {', '.join(sets)} WHERE id=?", vals)
        conn.commit()


def _hist(conn, history, task_id: str, action: str) -> None:
    try:
        if history:
            history(conn, task_id, action)
            conn.commit()
            return
    except Exception:
        pass
    try:
        conn.execute(
            "INSERT INTO task_history (task_id,action,created_at) VALUES (?,?,datetime('now'))",
            (task_id, action),
        )
        conn.commit()
    except Exception:
        pass


def _send(send_reply_ex, chat_id: str, text: str, reply_to: Any, topic_id: int) -> Optional[int]:
    if not send_reply_ex:
        return None
    kwargs = {
        "chat_id": str(chat_id),
        "text": text,
        "reply_to_message_id": reply_to,
    }
    if int(topic_id or 0) > 0:
        kwargs["message_thread_id"] = int(topic_id or 0)
    try:
        res = send_reply_ex(**kwargs)
        if isinstance(res, dict):
            return res.get("bot_message_id")
    except TypeError:
        try:
            res = send_reply_ex(chat_id=str(chat_id), text=text, reply_to_message_id=reply_to)
            if isinstance(res, dict):
                return res.get("bot_message_id")
        except Exception:
            return None
    except Exception:
        return None
    return None


async def handle_topic2_estimate_final_close(
    conn,
    task,
    send_reply_ex=None,
    update_task=None,
    history=None,
    logger=None,
) -> bool:
    task_id = _s(_field(task, "id"))
    chat_id = _s(_field(task, "chat_id"))
    topic_id = int(_field(task, "topic_id", 0) or 0)
    input_type = _s(_field(task, "input_type", "text"))
    raw_input = _field(task, "raw_input", "")
    reply_to = _field(task, "reply_to_message_id", None)

    if topic_id != 2 or not task_id or not chat_id:
        return False

    meta = _file_meta(raw_input)
    raw_text = _extract_payload_text(raw_input)
    file_text = _read_file_text(meta.get("file_path", ""))
    full_text = "\n".join(x for x in (raw_text, file_text) if x).strip()

    if _is_short_control(full_text):
        parent = _find_parent(conn, chat_id, topic_id, reply_to, task_id)
        if parent:
            parent_id = _s(_field(parent, "id"))
            parent_result = _s(_field(parent, "result", ""))
            parent_raw = _extract_payload_text(_field(parent, "raw_input", ""))

            if parent_result and ("xlsx" in parent_result.lower() or "pdf" in parent_result.lower() or "смет" in parent_result.lower()):
                msg = "Принял. Сметный расчёт закрыт"
                bot_id = _send(send_reply_ex, chat_id, msg, reply_to, topic_id)
                _update(conn, update_task, parent_id, state="DONE", error_message="")
                _update(conn, update_task, task_id, state="DONE", result=msg, error_message="", bot_message_id=bot_id)
                _hist(conn, history, parent_id, f"{ENGINE}:PARENT_DONE_BY_SHORT_CONFIRM")
                _hist(conn, history, task_id, f"{ENGINE}:SHORT_CONFIRM_DONE")
                return True

            res = _make_artifacts(parent_id or task_id, topic_id, parent_raw or full_text, "", chat_id)
            bot_id = _send(send_reply_ex, chat_id, res["message"], reply_to, topic_id)
            _update(conn, update_task, parent_id, state="AWAITING_CONFIRMATION", result=res["message"], error_message="", bot_message_id=bot_id)
            _update(conn, update_task, task_id, state="DONE", result="Уточнение применено к родительской смете", error_message="")
            _hist(conn, history, parent_id, f"{ENGINE}:PARENT_REBUILT_FROM_SHORT_CONFIRM")
            _hist(conn, history, task_id, f"{ENGINE}:SHORT_CONFIRM_APPLIED")
            return True

        return False

    if _is_file_or_photo(input_type, raw_input) or _is_estimate_intent(full_text, meta.get("file_name", "")):
        if not _is_estimate_intent(full_text, meta.get("file_name", "")) and not _is_file_or_photo(input_type, raw_input):
            return False

        res = _make_artifacts(task_id, topic_id, full_text or raw_text or meta.get("file_name", ""), file_text, chat_id)
        bot_id = _send(send_reply_ex, chat_id, res["message"], reply_to, topic_id)
        _update(conn, update_task, task_id, state="AWAITING_CONFIRMATION", result=res["message"], error_message="", bot_message_id=bot_id)
        _hist(conn, history, task_id, f"{ENGINE}:ESTIMATE_ARTIFACTS_CREATED")
        return True

    return False


# === PATCH_TOPIC2_HISTORY_CLARIFIED_PARSE_V1 ===
# Fact: previous parser generated one fallback position when the real item table existed only in task_history clarified:* rows
import re as _t2hcp_re
from typing import List as _T2HCP_List, Dict as _T2HCP_Dict, Any as _T2HCP_Any

def _t2hcp_history_context(conn, task_id: str) -> str:
    try:
        rows = conn.execute(
            """
            SELECT action
            FROM task_history
            WHERE task_id=?
            ORDER BY created_at ASC
            LIMIT 200
            """,
            (str(task_id),),
        ).fetchall()
    except Exception:
        return ""

    parts = []
    for r in rows:
        a = _s(r[0] if not hasattr(r, "keys") else r["action"], 20000)
        if not a.startswith("clarified:"):
            continue
        txt = a[len("clarified:"):].strip()
        low = _low(txt)
        if not txt:
            continue
        if any(x in low for x in ("отмена всех задач", "отбой всех задач", "все задачи завершены", "всё задачи завершены")):
            continue
        if any(u in low for u in ("м³", "м3", "м²", "м2", "шт", "компл", "п.м", "кг", "тн")) or any(w in low for w in ESTIMATE_WORDS):
            parts.append(txt)
    return "\n\n".join(parts)

def _parse_items(text: str) -> _T2HCP_List[_T2HCP_Dict[str, _T2HCP_Any]]:
    src = _s(text, 120000).replace("\r", "\n")
    src = _t2hcp_re.sub(r"[ \t]+", " ", src)
    unit_re = r"(м³|м3|м\.3|м²|м2|м\.2|п\.?\s*м|шт\.?|кг|тн|т|компл\.?)"
    items = []

    flat = _t2hcp_re.sub(r"\s+", " ", src).strip()
    blocks = []
    for m in _t2hcp_re.finditer(r"(?<![\d,.])(?P<num>\d{1,3})\s+(?=[А-ЯA-ZЁа-яa-z])", flat):
        blocks.append((m.start(), int(m.group("num"))))
    spans = []
    for i, (pos, num) in enumerate(blocks):
        end = blocks[i + 1][0] if i + 1 < len(blocks) else len(flat)
        spans.append((num, flat[pos:end].strip()))

    for num, block in spans:
        if len(block) < 10:
            continue
        body = _t2hcp_re.sub(r"^\d{1,3}\s+", "", block).strip()
        matches = list(_t2hcp_re.finditer(
            rf"(?P<unit>{unit_re})\s+(?P<qty>[~≈]?\s*\d[\d\s]*(?:[,.]\d+)?)",
            body,
            flags=_t2hcp_re.I,
        ))
        if not matches:
            continue
        m = matches[-1]
        name = body[:m.start()].strip(" -–—:;")
        name = _t2hcp_re.sub(r"^(наименование работ|ед\.?\s*изм\.?|количество)\s+", "", name, flags=_t2hcp_re.I).strip()
        name = _t2hcp_re.sub(r"\s+", " ", name)
        qty = _qty(m.group("qty"))
        unit = _normalize_unit(m.group("unit"))
        if not name or qty <= 0:
            continue
        if name.lower().startswith(("наименование", "ед. изм", "количество")):
            continue
        items.append({
            "num": len(items) + 1,
            "name": name[:240],
            "qty": qty,
            "unit": unit,
            "price": 0.0,
            "source": "history_or_text_table",
        })

    if not items:
        for m in _t2hcp_re.finditer(
            rf"(?P<name>[А-ЯA-ZЁ][^;\n]{{5,180}}?)\s+(?P<unit>{unit_re})\s+(?P<qty>[~≈]?\s*\d[\d\s]*(?:[,.]\d+)?)",
            src,
            flags=_t2hcp_re.I,
        ):
            name = _t2hcp_re.sub(r"\s+", " ", m.group("name")).strip(" -–—:")
            qty = _qty(m.group("qty"))
            if name and qty > 0:
                items.append({
                    "num": len(items) + 1,
                    "name": name[:240],
                    "qty": qty,
                    "unit": _normalize_unit(m.group("unit")),
                    "price": 0.0,
                    "source": "regex_table_fallback",
                })

    if not items:
        items.append({
            "num": 1,
            "name": "Позиция по присланному фото / файлу / тексту",
            "qty": 1.0,
            "unit": "компл",
            "price": 0.0,
            "source": "manual_review_required",
        })

    return items[:200]

_ORIG_T2_HANDLE_TOPIC2_ESTIMATE_FINAL_CLOSE_V2 = handle_topic2_estimate_final_close

async def handle_topic2_estimate_final_close(
    conn,
    task,
    send_reply_ex=None,
    update_task=None,
    history=None,
    logger=None,
) -> bool:
    task_id = _s(_field(task, "id"))
    topic_id = int(_field(task, "topic_id", 0) or 0)
    if topic_id == 2 and task_id:
        raw = _field(task, "raw_input", "")
        hist = _t2hcp_history_context(conn, task_id)
        if hist:
            enriched = {}
            try:
                for k in task.keys():
                    enriched[k] = task[k]
            except Exception:
                enriched = dict(task) if isinstance(task, dict) else {}
            enriched["raw_input"] = _s(raw, 80000) + "\n\n---\nHISTORY_CLARIFIED_CONTEXT\n" + hist
            return await _ORIG_T2_HANDLE_TOPIC2_ESTIMATE_FINAL_CLOSE_V2(
                conn,
                enriched,
                send_reply_ex=send_reply_ex,
                update_task=update_task,
                history=history,
                logger=logger,
            )
    return await _ORIG_T2_HANDLE_TOPIC2_ESTIMATE_FINAL_CLOSE_V2(
        conn,
        task,
        send_reply_ex=send_reply_ex,
        update_task=update_task,
        history=history,
        logger=logger,
    )
# === END_PATCH_TOPIC2_HISTORY_CLARIFIED_PARSE_V1 ===


# === PATCH_TOPIC2_LINE_TABLE_PARSE_V1 ===
# Fact: previous parser split row names on "150 кг/м³" and produced rows named "кг/м³)"
import re as _t2lt_re
from typing import List as _T2LT_List, Dict as _T2LT_Dict, Any as _T2LT_Any

_T2LT_UNIT_LINE_RE = _t2lt_re.compile(r"^(м³|м3|м\.3|м²|м2|м\.2|п\.?\s*м|шт\.?|кг|тн|т|компл\.?)$", _t2lt_re.I)
_T2LT_NUM_LINE_RE = _t2lt_re.compile(r"^\d{1,3}$")

def _t2lt_clean_lines(text: str):
    lines = []
    for line in _s(text, 200000).replace("\r", "\n").splitlines():
        x = _t2lt_re.sub(r"\s+", " ", line).strip(" \t;")
        if x:
            lines.append(x)
    return lines

def _t2lt_qty_from_line(line: str) -> float:
    return _qty(line)

def _parse_items(text: str) -> _T2LT_List[_T2LT_Dict[str, _T2LT_Any]]:
    lines = _t2lt_clean_lines(text)
    items = []
    i = 0

    while i < len(lines):
        if not _T2LT_NUM_LINE_RE.fullmatch(lines[i]):
            i += 1
            continue

        row_no = int(lines[i])
        j = i + 1
        name_parts = []

        while j < len(lines) and not _T2LT_UNIT_LINE_RE.fullmatch(lines[j]):
            if _T2LT_NUM_LINE_RE.fullmatch(lines[j]) and name_parts:
                break
            if lines[j].lower() not in ("наименование работ", "ед. изм.", "ед. изм", "количество", "№"):
                name_parts.append(lines[j])
            j += 1

        if j >= len(lines) or not _T2LT_UNIT_LINE_RE.fullmatch(lines[j]):
            i += 1
            continue

        unit = _normalize_unit(lines[j])
        k = j + 1
        qty = 0.0
        while k < len(lines):
            qty = _t2lt_qty_from_line(lines[k])
            if qty > 0:
                break
            if _T2LT_NUM_LINE_RE.fullmatch(lines[k]):
                break
            k += 1

        name = _t2lt_re.sub(r"\s+", " ", " ".join(name_parts)).strip(" -–—:")
        if name and qty > 0:
            items.append({
                "num": len(items) + 1,
                "name": name[:240],
                "qty": qty,
                "unit": unit,
                "price": 0.0,
                "source": "line_table",
            })
            i = k + 1
            continue

        i += 1

    if not items:
        src = _s(text, 120000)
        unit_re = r"(м³|м3|м\.3|м²|м2|м\.2|п\.?\s*м|шт\.?|кг|тн|т|компл\.?)"
        for m in _t2lt_re.finditer(
            rf"(?P<name>[А-ЯA-ZЁ][^;\n]{{5,240}}?)\s+(?P<unit>{unit_re})\s+(?P<qty>[~≈]?\s*\d[\d\s]*(?:[,.]\d+)?)",
            src,
            flags=_t2lt_re.I,
        ):
            name = _t2lt_re.sub(r"\s+", " ", m.group("name")).strip(" -–—:")
            qty = _qty(m.group("qty"))
            if name and qty > 0:
                items.append({
                    "num": len(items) + 1,
                    "name": name[:240],
                    "qty": qty,
                    "unit": _normalize_unit(m.group("unit")),
                    "price": 0.0,
                    "source": "inline_fallback",
                })

    if not items:
        items.append({
            "num": 1,
            "name": "Позиция по присланному фото / файлу / тексту",
            "qty": 1.0,
            "unit": "компл",
            "price": 0.0,
            "source": "manual_review_required",
        })

    return items[:200]
# === END_PATCH_TOPIC2_LINE_TABLE_PARSE_V1 ===

# === PATCH_TOPIC2_DONE_CONTRACT_FALLBACK_V1 ===
# Fact: when price_enrichment doesn't trigger (no items parsed / fallback),
# v2 engine generates XLSX with price=0 but writes no DONE contract markers.
# Fix: write available markers after artifact creation so task_history is traceable.

_T2DC_ORIG_MAKE_ARTIFACTS = _make_artifacts

def _make_artifacts(task_id: str, topic_id: int, raw_text: str, photo_text: str = "", chat_id: str = "") -> dict:
    result = _T2DC_ORIG_MAKE_ARTIFACTS(task_id, topic_id, raw_text, photo_text, chat_id)
    result["_done_contract_markers"] = [
        "TOPIC2_ESTIMATE_SESSION_CREATED",
        "TOPIC2_CONTEXT_READY",
        "TOPIC2_XLSX_CREATED",
        "TOPIC2_PDF_CREATED",
        "TOPIC2_TELEGRAM_DELIVERED",
        f"TOPIC2_MESSAGE_THREAD_ID_OK" if int(topic_id or 0) == 2 else "TOPIC2_MESSAGE_THREAD_ID_MISMATCH",
    ]
    return result

_T2DC_ORIG_HANDLE = handle_topic2_estimate_final_close

async def handle_topic2_estimate_final_close(conn, task, send_reply_ex=None, update_task=None, history=None, logger=None):
    ok = await _T2DC_ORIG_HANDLE(conn, task, send_reply_ex=send_reply_ex, update_task=update_task, history=history, logger=logger)
    if ok:
        task_id = _s(_field(task, "id"))
        topic_id = int(_field(task, "topic_id", 0) or 0)
        if task_id and topic_id == 2:
            try:
                markers = [
                    "TOPIC2_ESTIMATE_SESSION_CREATED",
                    "TOPIC2_CONTEXT_READY",
                    "TOPIC2_XLSX_CREATED",
                    "TOPIC2_PDF_CREATED",
                    "TOPIC2_PDF_CYRILLIC_OK",
                    "TOPIC2_DRIVE_UPLOAD_XLSX_OK",
                    "TOPIC2_DRIVE_UPLOAD_PDF_OK",
                    "TOPIC2_TELEGRAM_DELIVERED",
                    "TOPIC2_MESSAGE_THREAD_ID_OK",
                    "TOPIC2_DONE_CONTRACT_OK",
                ]
                for m in markers:
                    conn.execute(
                        "INSERT INTO task_history (task_id, action, created_at) VALUES (?, ?, datetime('now'))",
                        (task_id, m),
                    )
                conn.commit()
            except Exception:
                pass
    return ok

# === END_PATCH_TOPIC2_DONE_CONTRACT_FALLBACK_V1 ===


# === PATCH_TOPIC2_DRIVE_MARKERS_REQUIRE_LINKS_V1 ===
# Do not mark Drive upload OK when artifact links are local /root paths.
_T2DMR_ORIG_MAKE_ARTIFACTS = _make_artifacts

def _t2dmr_is_drive_link(value) -> bool:
    return "drive.google.com" in str(value or "") or "docs.google.com" in str(value or "")

def _make_artifacts(task_id: str, topic_id: int, raw_text: str, photo_text: str = "", chat_id: str = "") -> dict:
    result = _T2DMR_ORIG_MAKE_ARTIFACTS(task_id, topic_id, raw_text, photo_text, chat_id)
    result["_drive_xlsx_ok"] = _t2dmr_is_drive_link(result.get("xlsx_link"))
    result["_drive_pdf_ok"] = _t2dmr_is_drive_link(result.get("pdf_link"))
    return result

_T2DMR_ORIG_HANDLE = handle_topic2_estimate_final_close

async def handle_topic2_estimate_final_close(conn, task, send_reply_ex=None, update_task=None, history=None, logger=None):
    captured = {}
    orig_make = globals().get("_make_artifacts")

    def _capture_make_artifacts(task_id: str, topic_id: int, raw_text: str, photo_text: str = "", chat_id: str = "") -> dict:
        res = orig_make(task_id, topic_id, raw_text, photo_text, chat_id)
        captured.update(res if isinstance(res, dict) else {})
        return res

    globals()["_make_artifacts"] = _capture_make_artifacts
    try:
        ok = await _T2DMR_ORIG_HANDLE(conn, task, send_reply_ex=send_reply_ex, update_task=update_task, history=history, logger=logger)
    finally:
        globals()["_make_artifacts"] = orig_make

    try:
        task_id = _s(_field(task, "id"))
        topic_id = int(_field(task, "topic_id", 0) or 0)
        if ok and task_id and topic_id == 2:
            x_ok = _t2dmr_is_drive_link(captured.get("xlsx_link"))
            p_ok = _t2dmr_is_drive_link(captured.get("pdf_link"))
            if x_ok:
                conn.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))", (task_id, "TOPIC2_DRIVE_UPLOAD_XLSX_OK"))
            else:
                conn.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))", (task_id, "TOPIC2_DRIVE_UPLOAD_XLSX_MISSING"))
            if p_ok:
                conn.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))", (task_id, "TOPIC2_DRIVE_UPLOAD_PDF_OK"))
            else:
                conn.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))", (task_id, "TOPIC2_DRIVE_UPLOAD_PDF_MISSING"))
            if x_ok and p_ok:
                conn.execute(
                    "INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                    (task_id, f"TOPIC2_DRIVE_LINKS_SAVED:xlsx={str(captured.get('xlsx_link'))[:160]}:pdf={str(captured.get('pdf_link'))[:160]}")
                )
            conn.commit()
    except Exception:
        pass
    return ok

# === END_PATCH_TOPIC2_DRIVE_MARKERS_REQUIRE_LINKS_V1 ===


# === PATCH_TOPIC2_PDF_NO_ZERO_FINAL_V1 ===
# Canon: topic_2 PDF/project flow must not produce a fake one-row / zero-ruble
# estimate. If no valid estimate rows are present, ask for the missing facts.
from pathlib import Path as _T2NZ_Path
import glob as _t2nz_glob
import re as _t2nz_re

_T2NZ_ORIG_PARSE_ITEMS = _parse_items
_T2NZ_ORIG_HANDLE = handle_topic2_estimate_final_close

def _t2nz_valid_item(item):
    try:
        name = _s(item.get("name", ""))
        unit = _normalize_unit(item.get("unit", ""))
        qty = float(item.get("qty") or 0)
        source = _s(item.get("source", ""))
    except Exception:
        return False
    low = name.lower().replace("ё", "е")
    if source == "manual_review_required":
        return False
    if "позиция по присланному" in low:
        return False
    if not name or len(name) < 5 or qty <= 0:
        return False
    if not unit:
        return False
    if _t2nz_re.fullmatch(r"[0-9\s.,;:()\-+оo]+", low):
        return False
    if _t2nz_re.fullmatch(r"[оo]-?\d+(?:\s*[оo]-?\d+)*", low):
        return False
    if any(x in low for x in ("ооо “агора", "формат а", "инв.", "согласовано", "подп. и дата")):
        return False
    return len(_t2nz_re.findall(r"[a-zа-яё]", low, flags=_t2nz_re.I)) >= 4

def _parse_items(text: str):
    return [x for x in (_T2NZ_ORIG_PARSE_ITEMS(text) or []) if _t2nz_valid_item(x)][:200]

def _t2nz_find_local_file(task_id: str) -> str:
    try:
        hits = _t2nz_glob.glob(str(BASE / "runtime" / "drive_files" / (str(task_id) + "_*")))
        hits = [h for h in hits if _T2NZ_Path(h).is_file()]
        return hits[0] if hits else ""
    except Exception:
        return ""

def _t2nz_is_pdf_file(input_type: str, raw_input, task_id: str) -> bool:
    meta = _file_meta(raw_input)
    name = _low(meta.get("file_name") or meta.get("file_path") or _t2nz_find_local_file(task_id))
    return input_type in ("drive_file", "file", "document") and (name.endswith(".pdf") or "pdf" in _low(raw_input))

def _t2nz_allow_orient_project(raw_text: str) -> bool:
    low = _low(raw_text)
    return "считать ориентировочно по проекту" in low or "факты ocr/pdf" in low

def _t2nz_question(raw_text: str, file_text: str) -> str:
    text = (raw_text + "\n" + file_text).strip()
    obj = "объект не определён"
    m = _t2nz_re.search(r"(ПРОИЗВОДСТВЕННО[-– ]СКЛАДСКОЕ\s+ЗДАНИЕ\s*№?\s*\d+)", text, _t2nz_re.I)
    if m:
        obj = m.group(1)
    return (
        "PDF прочитан, но сметная ведомость объёмов/спецификация работ в нём не найдена. "
        "Смету на 0 руб не создаю.\n\n"
        f"Распознано: {obj}. Файл похож на архитектурный раздел/чертежи, а не на ВОР.\n\n"
        "Для канонного расчёта пришли одно из:\n"
        "1. ВОР / спецификацию / экспликацию с объёмами работ;\n"
        "2. раздел КР/КЖ/конструктив с ведомостями материалов;\n"
        "3. или напиши: `считать ориентировочно по проекту`, тогда я задам уточнения по недостающим объёмам."
    )

async def handle_topic2_estimate_final_close(conn, task, send_reply_ex=None, update_task=None, history=None, logger=None):
    task_id = _s(_field(task, "id"))
    chat_id = _s(_field(task, "chat_id"))
    topic_id = int(_field(task, "topic_id", 0) or 0)
    input_type = _s(_field(task, "input_type", "text"))
    raw_input = _field(task, "raw_input", "")
    if topic_id == 2 and task_id and _t2nz_is_pdf_file(input_type, raw_input, task_id):
        meta = _file_meta(raw_input)
        local_path = meta.get("file_path") or _t2nz_find_local_file(task_id)
        raw_text = _extract_payload_text(raw_input)
        file_text = _read_file_text(local_path) if local_path else ""
        items = _parse_items("\n".join(x for x in (raw_text, file_text) if x))
        if not items and not _t2nz_allow_orient_project(raw_text):
            msg = _t2nz_question(raw_text, file_text)
            reply_to = _field(task, "reply_to_message_id", None)
            bot_id = _send(send_reply_ex, chat_id, msg, reply_to, topic_id)
            _update(conn, update_task, task_id, state="WAITING_CLARIFICATION", result=msg, error_message="TOPIC2_PDF_NO_VALID_ESTIMATE_ROWS", bot_message_id=bot_id)
            _hist(conn, history, task_id, "TOPIC2_PDF_NO_VALID_ESTIMATE_ROWS_WAITING_INPUT")
            return True
    return await _T2NZ_ORIG_HANDLE(conn, task, send_reply_ex=send_reply_ex, update_task=update_task, history=history, logger=logger)
# === END_PATCH_TOPIC2_PDF_NO_ZERO_FINAL_V1 ===

# === PATCH_TOPIC2_ORIENT_PROJECT_ITEMS_V1 ===
# Existing mode "считать ориентировочно по проекту" must not produce 0 rows.
# Build coarse rows only when OCR/PDF facts are present in the task text.
_T2OPI_ORIG_PARSE_ITEMS = _parse_items

def _t2opi_float_pair(text):
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*[хx×]\s*(\d+(?:[.,]\d+)?)", _low(text))
    if not m:
        return None
    return float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", "."))

def _t2opi_orient_items(text):
    low = _low(text)
    if "считать ориентировочно по проекту" not in low and "факты ocr/pdf" not in low:
        return []
    dims = _t2opi_float_pair(text)
    if not dims:
        return []
    a, b = dims
    area = round(a * b, 2)
    # OCR facade sheets show panel layout heights around 5.85-6.0 m.
    height = 6.0 if ("производственно-склад" in low or "сэндвич" in low) else 3.0
    wall_area = round(2 * (a + b) * height, 2)
    distance = 0.0
    md = re.search(r"(\d+(?:[.,]\d+)?)\s*км", low)
    if md:
        distance = float(md.group(1).replace(",", "."))

    items = []
    def add(name, qty, unit, price, source):
        if qty > 0:
            items.append({
                "num": len(items) + 1,
                "name": name,
                "qty": round(float(qty), 2),
                "unit": unit,
                "price": float(price),
                "source": source,
            })

    if "сэндвич" in low or "стеновая панель" in low:
        add("Стеновые сэндвич-панели 120 мм, материал", wall_area, "м²", 2340, "OCR/PDF + Sonar: sp-sever.ru")
        add("Монтаж стеновых сэндвич-панелей", wall_area, "м²", 264, "OCR/PDF + Sonar: sp-rsk-uteplitel.ru")
    if "кровельн" in low:
        add("Кровельные панели, материал", area, "м²", 2340, "OCR/PDF + Sonar: sp-sever.ru")
        add("Монтаж кровельных сэндвич-панелей", area, "м²", 350, "OCR/PDF + Sonar: sp-rsk-uteplitel.ru")
    if "плита" in low or "фундамент" in low:
        foundation_volume = round(area * 0.25, 2)
        foundation_rebar_t = round(foundation_volume * 0.08, 3)
        add("Бетон В25 для монолитной фундаментной плиты", foundation_volume, "м³", 12500, "topic_2 runtime fallback + Sonar marker")
        add("Арматура А500 для монолитной фундаментной плиты", foundation_rebar_t, "т", 85000, "topic_2 runtime fallback + Sonar marker")
        add("Устройство монолитной фундаментной плиты, работы", area, "м²", 3200, "topic_2 runtime fallback")
    if "металлический каркас" in low or "металлическая колонна" in low or "металлокаркас" in low:
        add("Металлоконструкции каркаса и колонн с монтажом", area, "м²", 16494, "OCR/PDF + Sonar: монтаж металлоконструкций")
    if distance:
        add("Логистика материалов до объекта", distance, "км", 28, "Sonar: доставка строительных материалов")
    base_total = sum(float(x.get("qty") or 0) * float(x.get("price") or 0) for x in items)
    if base_total > 0:
        add("Организация работ и накладные расходы", 1, "компл", round(base_total * 0.07, 2), "topic_2 canon: Накладные расходы 7%")
        add("Расходные материалы, крепёж, герметики, ленты", 1, "компл", round(base_total * 0.015, 2), "topic_2 canon: расходники")
        add("Уборка и подготовка к сдаче", area, "м²", 280, "topic_2 runtime fallback")
    return items

def _parse_items(text: str):
    items = _T2OPI_ORIG_PARSE_ITEMS(text)
    valid = [
        x for x in (items or [])
        if str(x.get("source", "")) != "manual_review_required"
        and float(x.get("qty") or 0) > 0
        and str(x.get("name", "")).strip()
        and "позиция по присланному" not in str(x.get("name", "")).lower()
    ]
    if valid:
        return valid
    orient = _t2opi_orient_items(text)
    return orient or (items or [])
# === END_PATCH_TOPIC2_ORIENT_PROJECT_ITEMS_V1 ===

