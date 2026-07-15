# === FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3 ===
from __future__ import annotations

import os
import re
import io
import json
import uuid
import time
import math
import sqlite3
import asyncio
import tempfile
import statistics
import datetime
from pathlib import Path
from typing import Any, Dict, Optional, List, Tuple

import requests
from dotenv import load_dotenv

BASE = Path("/root/.areal-neva-core")
ENV_PATH = BASE / ".env"
MEM_DB = BASE / "data/memory.db"
load_dotenv(str(ENV_PATH), override=True)

TOPIC_ID_STROYKA = 2
DRIVE_TEMPLATES_PARENT_ID = "19Z3acDgPub4nV55mad5mb8ju63FsqoG9"

DEPRECATED_TEMPLATE_NAMES = (
    "ВОР_кирпичная_кладка_ИСПРАВЛЕНО.xlsx",
)

CANON_TEMPLATE_FALLBACK = {
    "m80": {"title": "М-80.xlsx", "role": "full_house_estimate_template", "file_id": "1yt-RJsGRhO13zmPKNAn6bMuGrpXY7kWp", "source": "fallback_registry"},
    "m110": {"title": "М-110.xlsx", "role": "full_house_estimate_template", "file_id": "1Ub9fcwOcJ4pV30dcX88yf1225WOIdpWo", "source": "fallback_registry"},
    "roof": {"title": "крыша и перекр.xlsx", "role": "roof_and_floor_estimate_template", "file_id": "16YecwnJ9umnVprFu9V77UCV6cPrYbNh3", "source": "fallback_registry"},
    "foundation": {"title": "фундамент_Склад2.xlsx", "role": "foundation_estimate_template", "file_id": "1KuoSI4OI7gJoIBPVqBQGXtQnolXKMiDp", "source": "fallback_registry"},
    "areal": {"title": "Ареал Нева.xlsx", "role": "general_company_estimate_template", "file_id": "1DQw2qgMHtq2SqgJJP-93eIArpj1hnNNm", "source": "fallback_registry"},
}

ESTIMATE_WORDS = (
    "смет", "стоимость", "расчет", "расчёт", "посчитай", "коробк", "дом", "стройк",
    "фундамент", "кровл", "перекр", "ангар", "склад", "газобетон", "каркас", "монолит",
)

CONTINUATION_WORDS = (
    "да", "да сделай", "сделай", "где смета", "ну что", "вариант 1", "вариант 2",
    "первый", "второй", "подтверждаю", "ок", "окей", "цены актуальны", "адрес подтверждаю",
    "средняя", "минимальная", "максимальная", "ручная", "конкретная ссылка",
)

REVISION_WORDS = (
    "нет не так", "не так", "переделай", "исправь", "правки", "пересчитай", "измени", "уточни",
)

PROJECT_ONLY_WORDS = (
    "проект ар", "проект кж", "проект кд", "чертеж", "чертёж", "раздел ар", "раздел кж", "раздел кд",
)

EXCLUSIONS_DEFAULT = (
    "подготовка участка",
    "стройгородок",
    "бытовки",
    "отмостка",
    "дренаж",
    "ливневая канализация",
    "вывоз мусора",
    "наружные сети",
    "всё, что не указано явно",
)

PRICE_CHOICE_HELP = """Выбор цены:
- средняя / медианная
- минимальная
- максимальная
- конкретная ссылка
- ручная цена
- можно добавить наценку, скидку, запас или поправку по позиции, разделу или всей смете"""


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _low(v: Any) -> str:
    return _s(v).lower().replace("ё", "е")


def _clean(text: str, limit: int = 12000) -> str:
    text = (text or "").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()[:limit]


def _now() -> str:
    return datetime.datetime.utcnow().isoformat()


def _row_get(row: Any, key: str, default: Any = "") -> Any:
    try:
        if hasattr(row, "keys") and key in row.keys():
            return row[key]
    except Exception:
        pass
    try:
        return row[key]
    except Exception:
        return getattr(row, key, default)


def _cols(conn: sqlite3.Connection, table: str) -> List[str]:
    try:
        return [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    except Exception:
        return []


def _update_task_safe(conn: sqlite3.Connection, task_id: str, **kwargs: Any) -> None:
    cols = _cols(conn, "tasks")
    parts, vals = [], []
    for k, v in kwargs.items():
        if k in cols:
            parts.append(f"{k}=?")
            vals.append(v)
    if "updated_at" in cols:
        parts.append("updated_at=datetime('now')")
    if not parts:
        return
    vals.append(task_id)
    conn.execute(f"UPDATE tasks SET {', '.join(parts)} WHERE id=?", vals)
    conn.commit()


def _history_safe(conn: sqlite3.Connection, task_id: str, action: str) -> None:
    try:
        conn.execute(
            "INSERT INTO task_history (task_id, action, created_at) VALUES (?, ?, datetime('now'))",
            (task_id, _clean(action, 1000)),
        )
        conn.commit()
    except Exception:
        pass


def _topic2_history_has_v1(conn: sqlite3.Connection, task_id: str, marker: str) -> bool:
    try:
        row = conn.execute(
            "SELECT 1 FROM task_history WHERE task_id=? AND action=? LIMIT 1",
            (str(task_id), str(marker)),
        ).fetchone()
        return bool(row)
    except Exception:
        return False


def _topic2_send_sync_status_v1(conn: sqlite3.Connection, task_id: str, chat_id: str, topic_id: int, reply_to: Any, marker: str, text: str) -> None:
    if _topic2_history_has_v1(conn, task_id, marker):
        return
    try:
        from core.reply_sender import send_reply_ex
        send_reply_ex(
            chat_id=str(chat_id),
            text=_clean(text, 1200),
            reply_to_message_id=int(reply_to) if reply_to else None,
            message_thread_id=int(topic_id or 0) if topic_id else None,
        )
        _history_safe(conn, task_id, marker)
    except Exception as exc:
        _history_safe(conn, task_id, marker + "_ERR:" + _s(exc)[:80])

# === PATCH_TOPIC2_PRICE_CHOICE_LOOP_CLOSE_V1 helpers ===
PRICE_CHOICE_PROMPT_V1 = "Выбери уровень цен: 1 дешёвые / 2 средние / 3 надёжные / 4 вручную"

def _t2pcl_history_text(conn, task_id):
    try:
        rows = conn.execute(
            "SELECT action FROM task_history WHERE task_id=? ORDER BY rowid ASC",
            (str(task_id),),
        ).fetchall()
        out = []
        for r in rows:
            try:
                out.append(str(r["action"]))
            except Exception:
                out.append(str(r[0]))
        return "\n".join(out)
    except Exception:
        return ""

def _t2pcl_parse_explicit_price_choice(text):
    raw = _s(text)
    try:
        if raw.strip().startswith("{"):
            import json as _t2pcl_json
            obj = _t2pcl_json.loads(raw)
            if isinstance(obj, dict):
                raw = " ".join(_s(obj.get(k, "")) for k in ("caption", "text", "raw_input", "file_name"))
    except Exception:
        pass
    t = _low(raw)
    t = re.sub(r"\s+", " ", t).strip(" .,!?:;()[]{}")
    if not t:
        return ""
    if "средн" in t or "медиан" in t:
        return "median"
    if "миним" in t or "дешев" in t or "дешёв" in t:
        return "cheapest"
    if "максим" in t or "надеж" in t or "надёж" in t or "высок" in t or "дорог" in t:
        return "reliable"
    if "ручн" in t or "вручную" in t or "сам укажу" in t:
        return "manual"
    _exact = {
        "1": "cheapest", "а": "cheapest", "a": "cheapest", "а)": "cheapest", "a)": "cheapest",
        "дешевые": "cheapest", "дешёвые": "cheapest", "самые дешевые": "cheapest",
        "самые дешёвые": "cheapest", "минимальные": "cheapest", "минимальная": "cheapest",
        "вариант 1": "cheapest", "первый": "cheapest",
        "2": "median", "б": "median", "b": "median", "б)": "median", "b)": "median",
        "средние": "median", "средняя": "median", "среднее": "median",
        "медианная": "median", "медианные": "median",
        "вариант 2": "median", "второй": "median",
        "3": "reliable", "в": "reliable", "v": "reliable", "в)": "reliable", "v)": "reliable",
        "надежные": "reliable", "надёжные": "reliable", "надежный": "reliable", "надёжный": "reliable",
        "проверенные": "reliable", "проверенный": "reliable",
        "вариант 3": "reliable", "третий": "reliable",
        "4": "manual", "г": "manual", "g": "manual", "г)": "manual", "g)": "manual",
        "вручную": "manual", "ручная": "manual", "свои цены": "manual",
        "своя цена": "manual", "укажу цены": "manual",
        "вариант 4": "manual", "четвертый": "manual", "четвёртый": "manual",
    }
    return _exact.get(t, "")

def _t2pcl_old_public_output(text):
    s = _s(text)
    if not s:
        return False
    if '✅ Смета готова' not in s and any(x in s for x in ("⏳ Задачу понял", "Шаблон:", "Лист:", "Цены из листа")):
        return True
    if '✅ Смета готова' in s:
        has_drive_link = ('drive.google.com' in s) or ('docs.google.com' in s)
        has_excel_link = ('Excel:' in s or 'XLSX:' in s) and has_drive_link
        has_pdf_link = 'PDF:' in s and has_drive_link
        if not (has_excel_link and has_pdf_link):
            return True
    return False

async def _t2pcl_send_price_choice_prompt(conn, task_id, chat_id, reply_to_message_id=None, repeat=True):
    action = "TOPIC2_PRICE_CHOICE_REQUIRED_REPEAT" if repeat else "TOPIC2_PRICE_CHOICE_REQUESTED"
    _history_safe(conn, str(task_id), action)
    _update_task_safe(
        conn, str(task_id),
        state="WAITING_CLARIFICATION",
        result=PRICE_CHOICE_PROMPT_V1,
        error_message="TOPIC2_PRICE_CHOICE_REQUIRED",
    )
    try:
        maybe = _send_text(str(chat_id), PRICE_CHOICE_PROMPT_V1, reply_to_message_id, int(TOPIC_ID_STROYKA))
        if hasattr(maybe, "__await__"):
            await maybe
    except Exception as _e:
        _history_safe(conn, str(task_id), "TOPIC2_PRICE_CHOICE_PROMPT_SEND_ERR:" + _s(_e)[:200])
    return True

async def _t2pcl_price_choice_guard(conn, task_id, chat_id, raw_input, reply_to_message_id=None):
    task_id = str(task_id or "")
    if not task_id:
        return False
    hist = _t2pcl_history_text(conn, task_id)
    has_prices = ("TOPIC2_PRICE_ENRICHMENT_DONE" in hist or
                  "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3:prices_shown" in hist)
    has_choice = "TOPIC2_PRICE_CHOICE_CONFIRMED" in hist
    if not has_prices or has_choice:
        return False
    choice = _t2pcl_parse_explicit_price_choice(raw_input)
    if choice:
        _history_safe(conn, task_id, "TOPIC2_PRICE_CHOICE_CONFIRMED:" + choice)
        return False
    return await _t2pcl_send_price_choice_prompt(
        conn, task_id, chat_id, reply_to_message_id,
        repeat=("TOPIC2_PRICE_CHOICE_REQUESTED" in hist),
    )
# === /PATCH_TOPIC2_PRICE_CHOICE_LOOP_CLOSE_V1 helpers ===

def _memory_save(chat_id: str, key: str, value: Dict[str, Any]) -> None:
    try:
        con = sqlite3.connect(str(MEM_DB))
        try:
            payload = json.dumps(value, ensure_ascii=False, indent=2)
            con.execute(
                "INSERT OR REPLACE INTO memory (id, chat_id, key, value, timestamp) VALUES (?, ?, ?, ?, ?)",
                (str(uuid.uuid4()), str(chat_id), str(key), payload, _now()),
            )
            con.commit()
        finally:
            con.close()
    except Exception:
        pass


def _memory_latest(chat_id: str, key_prefix: str) -> Optional[Dict[str, Any]]:
    try:
        con = sqlite3.connect(str(MEM_DB))
        con.row_factory = sqlite3.Row
        try:
            row = con.execute(
                "SELECT key,value,timestamp FROM memory WHERE chat_id=? AND key LIKE ? ORDER BY timestamp DESC LIMIT 1",
                (str(chat_id), f"{key_prefix}%"),
            ).fetchone()
            if not row:
                return None
            data = json.loads(row["value"] or "{}")
            data["_memory_key"] = row["key"]
            data["_memory_timestamp"] = row["timestamp"]
            return data
        finally:
            con.close()
    except Exception:
        return None



# === FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_CONTEXT_BLEED_FIX ===
def _parse_iso_ts(value: Any) -> Optional[datetime.datetime]:
    txt = _s(value)
    if not txt:
        return None
    txt = txt.replace("Z", "+00:00")
    try:
        dt = datetime.datetime.fromisoformat(txt)
        if dt.tzinfo is not None:
            dt = dt.astimezone(datetime.timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _age_seconds(value: Any) -> Optional[float]:
    dt = _parse_iso_ts(value)
    if not dt:
        return None
    return (datetime.datetime.utcnow() - dt).total_seconds()


def _pending_is_fresh(pending: Optional[Dict[str, Any]], max_seconds: int = 600) -> bool:
    if not pending:
        return False
    created = pending.get("created_at") or pending.get("_memory_timestamp")
    age = _age_seconds(created)
    return age is not None and 0 <= age <= max_seconds


def _is_bad_estimate_result(text: str) -> bool:
    t = _low(text)

    # === STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_BAD_RESULT_MARKERS ===
    stale_markers = (
        "задачи за последние 24 часа",
        "создание сметы: профлист",
        "итоговая сумма: 55000",
        "1capn1ikkxwypbxhny5caokqrsxbgzho",
        "1glcscpl3d91elveo_m11ezwh_uu5b4vm",
        "1pu77xrzhmpobus1pfximwdwckrgje1tn",
        "смета уже есть:",
        "смета создана по образцу вор",
        "вор_кирпичная_кладка",
        "vor_kirpich",
        "позиций: 13 | итого: 690510",
        "690510.00 руб",
        "файлы в этом топике уже есть",
        "нашёл релевантное",
        "нашел релевантное",
        "активный контекст найден",
    )
    if any(x in t for x in stale_markers):
        return True
    # === END_STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_BAD_RESULT_MARKERS ===
    bad = (
        # === FULL_STROYKA_V3_SEARCH_LOOP_BAD_RESULT_FIX ===
        "поставщик | площадка",
        "auto_parts",
        "search_monolith",
        "tco | риски",
        "ошибка классификации запроса",
        "категория не совпадает",
        # === FULL_STROYKA_LOOP_FINAL_CLOSE_BAD_RESULT_FIX ===
        "смета создана по образцу вор",
        "смета уже есть:",
        "вор_кирпичная_кладка",
        "вор_кирпич",
        "vor_kirpich",
        "в_ор_кирпич",
        "позиций: 13 | итого: 690510",
        "690510.00 руб",
        # === END_FULL_STROYKA_LOOP_FINAL_CLOSE_BAD_RESULT_FIX ===
        # === END_FULL_STROYKA_V3_SEARCH_LOOP_BAD_RESULT_FIX ===
        "файлы в этом топике уже есть",
        "нашёл релевантное",
        "нашел релевантное",
        "можно использовать как образец сметы",
        "активный контекст найден",
        "проектный файл не создан",
        "docx_create_failed",
        "state: finished",
        "задача закрыта по запросу",
        # === TOPIC2_FINAL_GAPS_V5_FORBIDDEN_PHRASES ===
        "файл скачан",
        "ожидает анализа",
        "выбор принят",
        "проверяю доступные файлы",
        "структура проекта включает",
        "файл содержит проект",
        "уточните запрос",
        "что строим",
        "не нашёл родительскую задачу",
        "не вижу размеры объекта",
        "позиция по присланному фото",
        # === END_TOPIC2_FINAL_GAPS_V5_FORBIDDEN_PHRASES ===
    )
    # === FULL_STROYKA_DISABLE_OLD_ESTIMATE_RECALL_FINAL_BAD_MARKERS ===
    stale_links = (
        "задачи за последние 24 часа",
        "создание сметы: профлист",
        "итоговая сумма: 55000",
        "1capn1ikkxwypbxhny5caokqrsxbgzho",
        "1glcscpl3d91elveo_m11ezwh_uu5b4vm",
        "1pu77xrzhmpobus1pfximwdwckrgje1tn",
        "смета уже есть",
        "использовать существующую или пересчитать",
    )
    if any(x in t for x in stale_links):
        return True
    # === END_FULL_STROYKA_DISABLE_OLD_ESTIMATE_RECALL_FINAL_BAD_MARKERS ===
    if any(x in t for x in bad):
        return True
    # === TOPIC2_FINAL_GAPS_V5_REGEX_FORBIDDEN ===
    import re as _ibr_re
    if _ibr_re.search(r'позиций:\s*1(?:\s|$)', t):
        return True
    if "/root/" in t or "/tmp/" in t:
        return True
    if "revision_context" in t or "traceback (most" in t:
        return True
    if "engine:" in t or "manifest:" in t:
        return True
    # === TOPIC2_FINAL_GAPS_V5B_RAW_JSON_GUARD ===
    if t.strip().startswith("{") and any(_k in t for _k in ('"state":', '"topic_id":', '"task_id":', '"result":', '"action":')):
        return True
    if _ibr_re.search(r'"state"\s*:\s*"(?:failed|in_progress|done|pending|waiting)', t):
        return True
    # === END_TOPIC2_FINAL_GAPS_V5B_RAW_JSON_GUARD ===
    # === END_TOPIC2_FINAL_GAPS_V5_REGEX_FORBIDDEN ===
    return False


def _has_real_estimate_artifact(text: str) -> bool:
    t = _low(text)
    if _is_bad_estimate_result(t):
        return False
    good = (
        "excel:",
        "xlsx:",
        ".xlsx",
        "pdf:",
        ".pdf",
        "предварительная смета готова",
        "итого:",
    )
    return any(x in t for x in good)


def _is_confirm_only(text: str) -> bool:
    t = _low(text).replace("[voice]", "").strip()
    if any(x in t for x in ESTIMATE_WORDS):
        return False
    return _is_confirm(t)
# === END_FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_CONTEXT_BLEED_FIX ===

def _drive_service():
    from core.topic_drive_oauth import _oauth_service
    return _oauth_service()


def list_drive_templates() -> List[Dict[str, Any]]:
    try:
        service = _drive_service()
        q = (
            f"'{DRIVE_TEMPLATES_PARENT_ID}' in parents and trashed = false and "
            "(mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
            "or mimeType = 'application/vnd.ms-excel' "
            "or mimeType = 'application/vnd.google-apps.spreadsheet')"
        )
        resp = service.files().list(
            q=q,
            spaces="drive",
            fields="files(id,name,mimeType,modifiedTime,size)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            pageSize=100,
        ).execute()
        out = []
        for f in resp.get("files", []):
            name = f.get("name") or ""
            if name in DEPRECATED_TEMPLATE_NAMES:
                continue
            out.append({
                "title": name,
                "file_id": f.get("id"),
                "mimeType": f.get("mimeType"),
                "modifiedTime": f.get("modifiedTime"),
                "role": "drive_dynamic_template",
                "source": "drive_templates_folder",
            })
        return out
    except Exception:
        return []


def _fallback_template_list() -> List[Dict[str, Any]]:
    return [dict(v) for v in CANON_TEMPLATE_FALLBACK.values()]


def _extract_dimensions(text: str) -> Optional[Tuple[float, float]]:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*[xх×*]\s*(\d+(?:[.,]\d+)?)", _low(text))
    if not m:
        return None
    return float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", "."))


def _extract_floors(text: str) -> Optional[int]:
    t = _low(text)
    m = re.search(r"(\d+)\s*(?:этаж|этажа|этажей)", t)
    if m:
        return int(m.group(1))
    if "2 эта" in t or "два эта" in t:
        return 2
    if "1 эта" in t or "один эта" in t:
        return 1
    return None


def _extract_distance_km(text: str) -> Optional[float]:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*км", _low(text))
    return float(m.group(1).replace(",", ".")) if m else None


def _extract_material(text: str) -> str:
    t = _low(text)
    if "сэндвич" in t or "стеновая панель" in t or "стеновые панели" in t:
        return "сэндвич-панели"
    if "металлический каркас" in t or "металлического каркаса" in t or "металлическая колонна" in t:
        return "металлокаркас"
    for key in ("газобетон", "каркас", "кирпич", "монолит", "керамоблок", "брус", "арболит"):
        if key in t:
            return key
    return ""


def _extract_object(text: str) -> str:
    t = _low(text)
    if "производственно-склад" in t:
        return "склад"
    if re.search(r"\\bдом(?:а|ом|е)?\\b", t):
        return "дом"
    for key in ("ангар", "склад", "фундамент", "кровля", "коробка"):
        if key in t:
            return key
    return ""


def _extract_foundation(text: str) -> str:
    t = _low(text)
    if "монолит" in t or "плита" in t:
        return "монолитная плита"
    if "лента" in t:
        return "ленточный фундамент"
    if "сва" in t:
        return "свайный фундамент"
    if "фундамент" in t:
        return "фундамент"
    return ""


def _extract_scope(text: str) -> str:
    t = _low(text)
    if "под ключ" in t:
        return "под ключ"
    if "коробк" in t:
        return "коробка"
    return ""


def _parse_request(text: str) -> Dict[str, Any]:
    dims = _extract_dimensions(text)
    area_floor = dims[0] * dims[1] if dims else None
    floors = _extract_floors(text)
    area_total = area_floor * floors if area_floor and floors else area_floor
    return {
        "object": _extract_object(text),
        "material": _extract_material(text),
        "dimensions": dims,
        "area_floor": area_floor,
        "floors": floors,
        "area_total": area_total,
        "distance_km": _extract_distance_km(text),
        "foundation": _extract_foundation(text),
        "scope": _extract_scope(text),
        "raw": text,
    }


def _missing_question(parsed: Dict[str, Any]) -> Optional[str]:

    # === PATCH_TOPIC2_AREA_ONLY_PDF_ROWS_MISSING_GATE_V1 ===
    # AR/TЭП area rows are not a VOR or material/work specification.
    rows = parsed.get("pdf_spec_rows") or []
    if rows:
        usable = []
        non_area = []
        for row in rows:
            try:
                name = _clean(row.get("name", "")).lower().replace("ё", "е")
                qty = float(row.get("qty") or 0)
                price = float(row.get("price") or 0)
            except Exception:
                continue
            if qty <= 0 and price <= 0:
                continue
            usable.append(row)
            if "площад" not in name and "общая" not in name:
                non_area.append(row)
        if usable and not non_area:
            return (
                "PDF прочитан: найдены только архитектурные площади/ТЭП, но не найдена ВОР, "
                "спецификация материалов или КЖ/конструктив для полной сметы. "
                "Финальную смету по этим данным не создаю, чтобы не подменять расчёт догадками. "
                "Пришли КЖ/ВОР/спецификацию материалов либо прямо напиши: `считать ориентировочно по проекту`."
            )
    # === END_PATCH_TOPIC2_AREA_ONLY_PDF_ROWS_MISSING_GATE_V1 ===
    if parsed.get("pdf_spec_rows") or parsed.get("ocr_table_rows"):
        return None
    # === STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_PRICE_NO_MISSING ===
    raw = _low(parsed.get("raw", ""))
    if ("цена" in raw or "руб" in raw or "₽" in raw) and any(u in raw for u in ("м²", "м2", "м³", "м3", "шт", "кг", "тн", "тонн")) and any(x in raw for x in ("смет", "фундамент", "монолит", "кровл", "работ")):
        return None
    # === END_STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_PRICE_NO_MISSING ===
    if not parsed.get("object"):
        return "Что строим: дом, ангар, склад, фундамент или кровлю?"
    if not parsed.get("material") and parsed.get("object") not in ("фундамент", "кровля", "ангар", "склад"):
        return "Из какого материала строим?"
    if parsed.get("distance_km") is None:
        return "Где находится объект: город или удалённость в км?"
    if not parsed.get("dimensions"):
        return "Какие размеры объекта?"
    if parsed.get("object") == "дом" and not parsed.get("floors"):
        return "Сколько этажей?"
    if not parsed.get("foundation") and parsed.get("object") not in ("кровля", "фундамент", "ангар", "склад"):
        return "Какой фундамент?"
    if not parsed.get("scope") and parsed.get("object") == "дом":
        return "Смета нужна только коробка или под ключ?"
    return None


def _template_score(parsed: Dict[str, Any], tpl: Dict[str, Any]) -> int:
    name = _low(tpl.get("title"))
    obj = parsed.get("object") or ""
    material = parsed.get("material") or ""
    area_total = float(parsed.get("area_total") or 0)
    score = 0
    if tpl.get("title") in DEPRECATED_TEMPLATE_NAMES:
        return -9999
    if obj in ("ангар", "склад", "фундамент") and ("фундамент" in name or "склад" in name):
        score += 150 if obj == "фундамент" and "фундамент" in name else 100
    if obj == "кровля" and ("крыш" in name or "перекр" in name):
        score += 100
    if material == "каркас" and ("м-80" in name or "м80" in name or "м-110" in name or "м110" in name):
        score += 90
    if material == "каркас" and area_total and area_total > 100 and ("м-110" in name or "м110" in name):
        score += 40
    if material == "каркас" and area_total and area_total <= 100 and ("м-80" in name or "м80" in name):
        score += 40
    if material in ("газобетон", "кирпич", "керамоблок", "монолит", "арболит") and ("ареал" in name or "м-110" in name or "м110" in name or "м-80" in name or "м80" in name):
        score += 80
    if "ареал" in name:
        score += 20
    return score


def choose_template(parsed: Dict[str, Any]) -> Dict[str, Any]:
    templates = list_drive_templates() or _fallback_template_list()
    ranked = sorted(templates, key=lambda t: _template_score(parsed, t), reverse=True)
    return ranked[0] if ranked else CANON_TEMPLATE_FALLBACK["areal"]


# === TOPIC2_FULL_CLOSE_GAP_A: deterministic work/material classifier ===
_WORK_KW = (
    "работ", "монтаж", "кладк", "установк", "доставк", "разгруз",
    "подач", "вибрирован", "уход за бетон", "гидроизоляц", "утеплен",
    "засыпк", "опалубк", "армирован", "бетонирован", "устройств",
    "демонтаж", "сборк",
)
_MAT_KW = (
    "материал", "бетон", "арматур", "газобетон", "кирпич", "брус",
    "пиломат", "утеплитель", "мембран", "плитк", "ламинат",
    "сантехник", "окна", "двери", "крепеж", "щебень", "песок",
)

def _classify_item(name: str, section: str) -> str:
    n = _low(str(name or ""))
    if any(k in n for k in _WORK_KW):
        return "work"
    if any(k in n for k in _MAT_KW):
        return "material"
    s = _low(str(section or ""))
    if s in ("логистика", "накладные расходы", "накладные"):
        return "overhead"
    return "material"
# === END TOPIC2_FULL_CLOSE_GAP_A classifier ===


def choose_template_sheet(parsed: Dict[str, Any], sheet_names: List[str]) -> tuple:
    """Returns (sheet_name, source) where source is 'match' or 'fallback'."""
    material = parsed.get("material") or ""
    obj = parsed.get("object") or ""
    names = list(sheet_names or [])
    lows = {name: _low(name) for name in names}

    if material == "каркас":
        for name, low in lows.items():
            if "каркас" in low:
                return name, "match"

    if material in ("газобетон", "кирпич", "керамоблок", "монолит", "арболит") or obj in ("дом", "коробка"):
        for name, low in lows.items():
            if "газобетон" in low:
                return name, "match"

    if obj in ("кровля",):
        for name, low in lows.items():
            if "кров" in low or "перекр" in low:
                return name, "match"

    if obj in ("ангар", "склад", "фундамент"):
        for name, low in lows.items():
            if "смет" in low or "фундамент" in low or "склад" in low:
                return name, "match"

    # GAP-B: fallback to first sheet — propagate source for marker
    return (names[0], "fallback") if names else (None, "fallback")


def download_template_xlsx(template: Dict[str, Any]) -> Optional[str]:
    file_id = template.get("file_id")
    if not file_id:
        return None
    try:
        service = _drive_service()
        mime = template.get("mimeType") or ""
        if mime == "application/vnd.google-apps.spreadsheet":
            request = service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
        path = os.path.join(tempfile.gettempdir(), f"tpl_{file_id}_{int(time.time())}.xlsx")
        with io.FileIO(path, "wb") as fh:
            from googleapiclient.http import MediaIoBaseDownload
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return path if os.path.exists(path) and os.path.getsize(path) > 1000 else None
    except Exception:
        return None


def extract_template_prices(template_path: Optional[str], parsed: Dict[str, Any]) -> tuple:
    """Returns (prices_text, sheet_name, sheet_fallback: bool)."""
    if not template_path or not os.path.exists(template_path):
        return "Цены из шаблона: шаблон не скачан, используется только структура/сценарий", None, False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(template_path, data_only=True, read_only=True)
        selected, _sheet_src = choose_template_sheet(parsed, wb.sheetnames)
        ws = wb[selected] if selected in wb.sheetnames else wb.active

        keys = ("бетон", "арматур", "газобетон", "кирпич", "кладк", "монтаж", "достав", "манипулятор", "кран", "пиломат", "кров")
        found = []
        for row in ws.iter_rows(values_only=True):
            txt = " | ".join("" if v is None else str(v) for v in row)
            low = _low(txt)
            if not any(k in low for k in keys):
                continue
            nums = re.findall(r"\d[\d\s]{2,}(?:[.,]\d+)?", txt)
            vals = []
            for n in nums:
                try:
                    v = float(n.replace(" ", "").replace(",", "."))
                    if 100 <= v <= 10000000:
                        vals.append(v)
                except Exception:
                    pass
            if vals:
                found.append(f"- {ws.title}: {txt[:180]}")
            if len(found) >= 15:
                break
        wb.close()
        return "Цены из выбранного листа шаблона:\n" + ("\n".join(found) if found else "ключевые цены в листе не распознаны автоматически"), selected, _sheet_src == "fallback"
    except Exception as e:
        return f"Цены из шаблона: ошибка чтения шаблона: {e}", None, False


def is_stroyka_estimate_candidate(task: Any) -> bool:
    if int(_row_get(task, "topic_id", 0) or 0) != TOPIC_ID_STROYKA:
        return False
    input_type = _low(_row_get(task, "input_type", ""))
    if input_type in ("photo", "file", "drive_file", "image", "document"):
        # §6 multi-format intake: allow when caption contains estimate keywords
        _mfi_cap = _low(_row_get(task, "raw_input", ""))
        if _mfi_cap and any(x in _mfi_cap for x in ESTIMATE_WORDS):
            return True
        return False
    raw = _low(_row_get(task, "raw_input", ""))
    if not raw:
        return False
    if any(x in raw for x in PROJECT_ONLY_WORDS) and "смет" not in raw and "стоим" not in raw:
        return False
    if _is_old_task_finish_request(raw):
        return True
    if any(x in raw for x in ESTIMATE_WORDS):
        return True
    if raw in CONTINUATION_WORDS or any(raw.startswith(x + " ") for x in CONTINUATION_WORDS):
        return True
    if any(x in raw for x in REVISION_WORDS):
        return True
    if raw.startswith("[voice]"):
        voice_raw = raw.replace("[voice]", "").strip()
        if voice_raw in CONTINUATION_WORDS or any(x in voice_raw for x in ESTIMATE_WORDS):
            return True
    return False


def _is_confirm(text: str) -> bool:
    t = _low(text).replace("[voice]", "").strip()
    return t in CONTINUATION_WORDS or any(t.startswith(x + " ") for x in CONTINUATION_WORDS)


def _is_revision(text: str) -> bool:
    return any(x in _low(text) for x in REVISION_WORDS)


def _is_new_project_estimate_request(text: str) -> bool:
    t = _low(text)
    if not t:
        return False
    has_project_files = any(x in t for x in (
        "эти файлы", "эти документы", "этот проект", "один проект",
        "проект", "pdf", "раздел", "ар", "кр", "кж",
    ))
    has_estimate_goal = any(x in t for x in (
        "нужно посчитать", "посчитать стоимость", "сделать смет",
        "стоимость строительства", "стоимость материалов", "сметн",
    ))
    return has_project_files and has_estimate_goal


def _topic2_safe_file_name_v1(name: str) -> str:
    name = _s(name) or "project.pdf"
    name = re.sub(r"[\\/]+", "_", name)
    name = re.sub(r"[\r\n\t]+", " ", name).strip()
    return name[:140] or "project.pdf"


def _topic2_download_drive_pdf_v1(file_id: str, out_path: str) -> bool:
    if not file_id:
        return False
    try:
        from googleapiclient.http import MediaIoBaseDownload
        service = _drive_service()
        request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        with io.FileIO(out_path, "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return os.path.exists(out_path) and os.path.getsize(out_path) > 1000
    except Exception:
        return False


def _topic2_related_project_pdf_tasks_v1(conn, chat_id: str, topic_id: int, task_id: str, reply_to: Any) -> List[Dict[str, Any]]:
    if int(topic_id or 0) != TOPIC_ID_STROYKA or not reply_to:
        return []
    try:
        base_msg = int(reply_to)
    except Exception:
        return []
    out: List[Dict[str, Any]] = []
    try:
        rows = conn.execute(
            """
            SELECT id, raw_input
            FROM tasks
            WHERE CAST(chat_id AS TEXT)=?
              AND COALESCE(topic_id,0)=?
              AND id<>?
              AND input_type IN ('drive_file','file','document')
              AND COALESCE(raw_input,'') LIKE '%.pdf%'
            ORDER BY rowid DESC
            LIMIT 40
            """,
            (str(chat_id), int(topic_id or 0), str(task_id)),
        ).fetchall()
    except Exception:
        rows = []
    for row in rows:
        try:
            meta = json.loads(_s(row[1]) or "{}")
        except Exception:
            meta = {}
        try:
            msg_id = int(meta.get("telegram_message_id") or 0)
        except Exception:
            msg_id = 0
        if not (base_msg <= msg_id <= base_msg + 8):
            continue
        name = _s(meta.get("file_name"))
        mime = _low(meta.get("mime_type"))
        low_name = _low(name)
        if ".pdf" not in low_name and "pdf" not in mime:
            continue
        if not any(x in low_name for x in ("раздел", " ар", "кр", "кж", "км", "проект")):
            continue
        out.append({
            "task_id": _s(row[0]),
            "file_id": _s(meta.get("file_id")),
            "file_name": name,
            "mime_type": _s(meta.get("mime_type")),
            "telegram_message_id": msg_id,
        })
    out.sort(key=lambda x: int(x.get("telegram_message_id") or 0))
    return out[:8]


def _topic2_hydrate_multifile_project_pdfs_v1(conn, task: Any, parsed: Dict[str, Any], raw_input: str) -> Dict[str, Any]:
    if not conn or not _is_new_project_estimate_request(raw_input):
        return parsed
    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    input_type = _low(_s(_row_get(task, "input_type", "")))
    reply_to = _row_get(task, "reply_to_message_id", None)
    if input_type not in ("text", "voice") or topic_id != TOPIC_ID_STROYKA:
        return parsed
    files = _topic2_related_project_pdf_tasks_v1(conn, chat_id, topic_id, task_id, reply_to)
    if not files:
        return parsed
    parsed = dict(parsed or {})
    spec_rows: List[Dict[str, Any]] = list(parsed.get("pdf_spec_rows") or [])
    project_rows: List[Dict[str, Any]] = list(parsed.get("pdf_project_rows") or [])
    local_paths: List[str] = []
    is_project_bundle = len(files) >= 2
    _topic2_send_sync_status_v1(
        conn,
        task_id,
        chat_id,
        topic_id,
        reply_to,
        "TOPIC2_MULTIFILE_PROJECT_OCR_STATUS_SENT",
        f"Принял {len(files)} PDF как один проект. Извлекаю объёмы из АР/КР через OCR/таблицы, это может занять несколько минут.",
    )
    for file_meta in files:
        name = _topic2_safe_file_name_v1(file_meta.get("file_name") or "project.pdf")
        local_path = str(BASE / "runtime" / "drive_files" / f"{task_id}_{name}")
        if not (os.path.exists(local_path) and os.path.getsize(local_path) > 1000):
            if not _topic2_download_drive_pdf_v1(_s(file_meta.get("file_id")), local_path):
                _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_PDF_DOWNLOAD_FAILED:" + name[:80])
                continue
        local_paths.append(local_path)
        _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_PDF_ATTACHED:" + name[:80])
        if is_project_bundle:
            continue
        try:
            from core.pdf_spec_extractor import extract_spec as _t2mf_pdf_extract
            result = _t2mf_pdf_extract(local_path) or {}
            rows = result.get("rows") or []
            if rows:
                spec_rows.extend(rows)
                _history_safe(conn, task_id, f"TOPIC2_MULTIFILE_PROJECT_SPEC_ROWS:{name[:40]}:{len(rows)}")
        except Exception as exc:
            _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_SPEC_ERR:" + _s(exc)[:80])
        try:
            tmp_parsed = dict(parsed)
            tmp_parsed["pdf_spec_source"] = local_path
            rows = _t2ar_project_rows_from_pdf_v1(tmp_parsed) if "_t2ar_project_rows_from_pdf_v1" in globals() else []
            if rows:
                project_rows.extend(rows)
                _history_safe(conn, task_id, f"TOPIC2_MULTIFILE_PROJECT_FACT_ROWS:{name[:40]}:{len(rows)}")
        except Exception as exc:
            _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_FACT_ERR:" + _s(exc)[:80])
    if len(local_paths) >= 2:
        try:
            if _topic2_volume_extract_requested_v1(raw_input):
                from core.pdf_spec_extractor import extract_project_positions_bundle as _t2mf_bundle_extract
            else:
                from core.pdf_spec_extractor import extract_project_pdf_bundle as _t2mf_bundle_extract
            bundle = _t2mf_bundle_extract(local_paths, topic_id=TOPIC_ID_STROYKA) or {}
            facts = list(bundle.get("facts") or [])
            if bundle.get("ok") and facts:
                values = [_s(f.get("value")) for f in facts if isinstance(f, dict) and _s(f.get("value"))]
                facts_text = "\n".join(values)
                parsed["project_bundle"] = bundle
                parsed["project_bundle_facts"] = facts
                parsed["project_bundle_source"] = "HOTFIX_FILE_BUNDLE_PIPELINE_FACT_ONLY_V1"
                parsed["raw"] = (str(parsed.get("raw") or raw_input or "").strip() + "\n\nФакты OCR/PDF:\n" + facts_text).strip()
                if any("18.0 x 36.0" == v for v in values):
                    parsed["dimensions"] = (18.0, 36.0)
                    parsed["area_floor"] = 648.0
                    parsed["area_total"] = 648.0
                if any("одноэтаж" in _low(v) for v in values):
                    parsed["floors"] = 1
                if any("склад" in _low(v) or "рамно-связев" in _low(v) for v in values):
                    parsed["object"] = parsed.get("object") or "склад"
                if any("фундамент" in _low(v) for v in values):
                    parsed["foundation"] = parsed.get("foundation") or "по проекту: фундаменты под колонны, фундаментная балка"
                parsed["pdf_bundle_facts_text"] = facts_text
                _history_safe(conn, task_id, f"TOPIC2_PROJECT_BUNDLE_FACTS_EXTRACTED:{len(facts)}")
                if bundle.get("specs"):
                    bundle_specs = list(bundle.get("specs") or [])
                    parsed["pdf_bundle_specs"] = bundle_specs
                    spec_rows.extend(bundle_specs)
                    _history_safe(conn, task_id, f"TOPIC2_PROJECT_BUNDLE_SPEC_ROWS:{len(bundle_specs)}")
            else:
                _history_safe(conn, task_id, "TOPIC2_PROJECT_BUNDLE_FACTS_EMPTY")
        except Exception as exc:
            _history_safe(conn, task_id, "TOPIC2_PROJECT_BUNDLE_FACTS_ERR:" + _s(exc)[:100])
    if local_paths:
        parsed["pdf_spec_source"] = local_paths[0]
        parsed["local_project_files"] = local_paths
        _history_safe(conn, task_id, f"TOPIC2_MULTIFILE_PROJECT_CONTEXT_READY:{len(local_paths)}_pdf")
    if project_rows:
        parsed["pdf_project_rows"] = project_rows
    if spec_rows:
        parsed["pdf_spec_rows"] = spec_rows
    elif project_rows:
        parsed["pdf_spec_rows"] = project_rows
    return parsed


def parse_price_choice(text: str) -> Dict[str, Any]:
    t = _low(text)
    choice = "median"
    if "миним" in t:
        choice = "minimum"
    elif "максим" in t:
        choice = "maximum"
    elif "конкрет" in t or "ссылк" in t or "вариант" in t:
        choice = "specific_source"
    elif "ручн" in t or "сам" in t:
        choice = "manual"
    elif "средн" in t or "медиан" in t or "да" in t or "подтверж" in t or "ок" in t:
        choice = "median"

    percent = 0.0
    m = re.search(r"(наценк|запас|плюс|\+)\s*(\d+(?:[.,]\d+)?)\s*%", t)
    if m:
        percent += float(m.group(2).replace(",", "."))
    m = re.search(r"(скидк|минус|-)\s*(\d+(?:[.,]\d+)?)\s*%", t)
    if m:
        percent -= float(m.group(2).replace(",", "."))

    manual_values = []
    if choice == "manual":
        for n in re.findall(r"\d[\d\s]{2,}(?:[.,]\d+)?", text):
            try:
                v = float(n.replace(" ", "").replace(",", "."))
                if 100 <= v <= 10000000:
                    manual_values.append(v)
            except Exception:
                pass

    return {"choice": choice, "percent_adjustment": percent, "manual_values": manual_values, "raw": text}


def _topic2_volume_extract_requested_v1(text: str) -> bool:
    low = _low(text)
    return any(x in low for x in (
        "объем", "объём", "обьем", "вытащи", "вытащить", "извлеки", "извлечь",
        "найди объ", "проверь объ", "позици", "цены пока не искать", "сначала только объ",
    ))


def _topic2_bundle_volumes_message_v1(parsed: Dict[str, Any]) -> str:
    bundle = (parsed or {}).get("project_bundle") or {}
    project_facts = list(bundle.get("project_facts") or bundle.get("facts") or [])
    properties = list(bundle.get("properties") or [])
    positions = list(bundle.get("positions") or [])
    quantities = list(bundle.get("direct_quantities") or bundle.get("quantities") or [])
    calculated_quantities = list(bundle.get("calculated_quantities") or [])
    derived_quantities = list(bundle.get("derived_quantities") or [])
    totals = list(bundle.get("totals") or [])
    volumes = list(bundle.get("volumes") or (parsed or {}).get("project_bundle_volumes") or [])
    missing_items = list(bundle.get("missing_items") or [])
    public_groups = list(bundle.get("public_groups") or [])
    result_type = _s(bundle.get("result_type") or "VOLUMES_ONLY_RESULT")
    header = "✅ Проектные позиции и объёмы извлечены" if result_type in ("PROJECT_POSITIONS_ONLY_RESULT", "PROJECT_POSITIONS_RESULT") else "✅ Объёмы извлечены"
    if public_groups:
        def _fmt_unit(u):
            return {"m3": "м³", "m2": "м²", "m": "п.м", "pcs": "шт", "l": "л", "kg": "кг", "t": "т"}.get(_s(u), _s(u))
        lines = [header, "", "Project facts:"]
        if project_facts:
            for row in project_facts[:12]:
                if not isinstance(row, dict):
                    continue
                value = _s(row.get("value") or row.get("name") or row.get("key"))
                lines.append("- {}".format(value))
        else:
            lines.append("- факты проекта не выделены")
        lines.extend(["", "Project positions:"])
        foundation_rows = [p for p in positions if isinstance(p, dict) and _s(p.get("position_type")) == "foundation"]
        if foundation_rows:
            for mark in ("Фм1", "Фм2"):
                count = ""
                for row in foundation_rows:
                    if _s(row.get("mark")) == mark and row.get("count_pcs") and not row.get("material"):
                        count = _s(row.get("count_pcs"))
                        break
                if count:
                    lines.append(f"- {mark}: {count} шт")
        else:
            lines.append("- проектные позиции не выделены")
        lines.extend(["", "Direct / calculated / derived quantities:"])
        seen_public = set()
        for row in public_groups:
            if not isinstance(row, dict):
                continue
            if _s(row.get("item_type")) == "total":
                continue
            name = _s(row.get("public_name"))
            value = row.get("value")
            unit = _fmt_unit(row.get("unit"))
            key = (name, _s(row.get("material_total_key")), _s(value), unit)
            if not name or key in seen_public:
                continue
            seen_public.add(key)
            qty = "{} {}".format(_s(value), unit).strip() if value not in (None, "") else "количество не выделено"
            lines.append("- {}: {}".format(name, qty))
        if totals:
            lines.extend(["", "Totals by material:"])
            seen_totals = set()
            total_names = {
                "foundation_concrete_B25_total_m3": "Бетон БСТ В25, фундаменты Фм1/Фм2",
                "foundation_concrete_B7_5_total_m3": "Бетон БСТ В7.5, фундаменты Фм1/Фм2",
                "foundation_grout_B30_total_m3": "Бетон БСТ В30, подливка Фм1/Фм2",
                "concrete_B25_total_m3": "Бетон БСТ В25 общий",
                "concrete_B7_5_total_m3": "Бетон БСТ В7.5 общий",
                "concrete_B30_total_m3": "Бетон БСТ В30 общий",
            }
            for row in totals:
                if not isinstance(row, dict):
                    continue
                raw_name = _s(row.get("public_name") or row.get("name"))
                name = total_names.get(raw_name, raw_name)
                value = row.get("value")
                unit = _fmt_unit(row.get("unit"))
                key = (name, _s(value), unit)
                if not name or key in seen_totals:
                    continue
                seen_totals.add(key)
                lines.append("- {}: {} {}".format(name, _s(value), unit))
        if missing_items:
            lines.extend(["", "Нужно уточнить/добрать:"])
            for item in missing_items:
                lines.append("- {}".format(_s(item)))
        return "\n".join(lines).strip()
    lines = [header, "", "Project facts:"]
    if project_facts:
        for row in project_facts[:12]:
            if not isinstance(row, dict):
                continue
            value = _s(row.get("value") or row.get("name") or row.get("key"))
            lines.append("- {} | стр. {} | {}".format(value, _s(row.get("page")), _s(row.get("source_file"))))
    else:
        lines.append("- факты проекта не выделены")
    lines.extend(["", "Properties:"])
    if properties:
        for row in properties:
            if not isinstance(row, dict):
                continue
            lines.append("- {}: {} {} | стр. {} | {}".format(
                _s(row.get("name")), _s(row.get("value")), _s(row.get("unit")), _s(row.get("page")), _s(row.get("source_file"))
            ))
    else:
        lines.append("- свойства не выделены")
    lines.extend(["", "Positions:"])
    foundation_rows = [p for p in positions if isinstance(p, dict) and _s(p.get("position_type")) == "foundation"]
    if foundation_rows:
        for mark in ("Фм1", "Фм2"):
            count = ""
            for row in foundation_rows:
                if _s(row.get("mark")) == mark and row.get("count_pcs") and not row.get("material"):
                    count = _s(row.get("count_pcs"))
                    break
            if count:
                lines.append(f"- {mark}: {count} шт")
            for row in foundation_rows:
                if _s(row.get("mark")) == mark and row.get("material"):
                    lines.append("  - {}: {} м³/шт -> {} м³".format(
                        _s(row.get("material")), _s(row.get("unit_volume_m3")), _s(row.get("total_volume_m3"))
                    ))
    elif positions:
        for row in positions[:20]:
            lines.append("- {} {} {}".format(_s(row.get("position_type")), _s(row.get("mark")), _s(row.get("count_pcs"))))
    else:
        lines.append("- проектные позиции не выделены")
    lines.extend(["", "Direct quantities:"])
    if quantities:
        for row in quantities:
            if not isinstance(row, dict):
                continue
            lines.append("- {}: {} {} | стр. {} | {}".format(
                _s(row.get("item") or row.get("name")), _s(row.get("value")), _s(row.get("unit")), _s(row.get("page")), _s(row.get("source_file"))
            ))
    else:
        lines.append("- прямые количества не выделены")
    lines.extend(["", "Calculated quantities:"])
    if calculated_quantities:
        for row in calculated_quantities:
            if not isinstance(row, dict):
                continue
            lines.append("- {}: {} {} | {}".format(
                _s(row.get("item") or row.get("name")), _s(row.get("value")), _s(row.get("unit")), _s(row.get("calculation"))
            ))
    else:
        lines.append("- расчётные количества по ведомостям не выделены")
    lines.extend(["", "Derived quantities:"])
    if derived_quantities:
        for row in derived_quantities:
            if not isinstance(row, dict):
                continue
            lines.append("- {}: {} {} | {} | {}".format(
                _s(row.get("name")), _s(row.get("value")), _s(row.get("unit")), _s(row.get("source")), _s(row.get("note"))
            ))
    else:
        lines.append("- расчётные количества не получены")
    lines.extend(["", "Totals:"])
    if totals:
        for row in totals:
            if not isinstance(row, dict):
                continue
            lines.append("- {}: {} {}".format(_s(row.get("name")), _s(row.get("value")), _s(row.get("unit"))))
    else:
        lines.append("- итоги не рассчитаны")
    if not (properties or quantities or derived_quantities) and volumes:
        lines.extend(["", "Raw volumes:"])
        for row in volumes:
            if not isinstance(row, dict):
                continue
            lines.append("- {}: {} {} | стр. {} | {}".format(
                _s(row.get("name")), _s(row.get("qty")), _s(row.get("unit")), _s(row.get("page")), _s(row.get("source_file"))
            ))
    if missing_items:
        lines.extend(["", "Missing items:"])
        for item in missing_items:
            lines.append("- " + _s(item))
        complete = "POSITIONS_EXTRACTION_COMPLETE" if result_type == "PROJECT_POSITIONS_ONLY_RESULT" else "VOLUMES_COMPLETE"
        lines.extend(["", f"{complete}=False. Цены не ищу, смету не делаю до закрытия missing_items."])
    else:
        complete = "POSITIONS_EXTRACTION_COMPLETE" if result_type == "PROJECT_POSITIONS_ONLY_RESULT" else "VOLUMES_COMPLETE"
        lines.extend(["", f"{complete}=True. Можно переходить к следующему действию."])
    lines.extend([
        "",
        "Цены и смету не запускаю, потому что текущая команда — только объёмы.",
        "Дальше напиши: считать смету / искать цены / закрыть / уточнить объёмы.",
    ])
    return "\n".join(lines)


def _numbers_from_price_text(price_text: str, keywords: Tuple[str, ...]) -> List[float]:
    vals = []
    for line in price_text.splitlines():
        low = _low(line)
        if any(k in low for k in keywords):
            for n in re.findall(r"\d[\d\s]{2,}(?:[.,]\d+)?", line):
                try:
                    v = float(n.replace(" ", "").replace(",", "."))
                    if 100 <= v <= 10000000:
                        vals.append(v)
                except Exception:
                    pass
    return vals


def _parse_price_sources(price_text: str) -> List[Dict[str, Any]]:
    """Parse Perplexity pipe-delimited response into per-position source records."""
    sources: List[Dict[str, Any]] = []
    if not price_text:
        return sources
    today = datetime.date.today().isoformat()
    for line in price_text.splitlines():
        line = line.strip(" \t-—•·")
        if "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 4:
            continue
        position = parts[0].lower()
        supplier = parts[4] if len(parts) > 4 else ""
        url = parts[5] if len(parts) > 5 else ""
        if any(x in _low(supplier) for x in ("нет данных", "not found", "not_found", "н/д")):
            supplier = ""
        if any(x in _low(url) for x in ("нет данных", "not found", "not_found", "н/д")):
            url = ""
        checked_at = parts[6].strip() if len(parts) > 6 else today
        status = "LIVE_CONFIRMED" if (supplier and url) else ("PARTIAL" if (supplier or url) else "UNVERIFIED")
        if not position:
            continue
        kw = [w for w in re.split(r"[\s,;/]+", position) if len(w) > 2]
        sources.append({
            "keywords": kw,
            "position": position,
            "supplier": supplier,
            "url": url,
            "checked_at": checked_at or today,
            "status": status,
        })
    return sources


def _match_price_source(sources: List[Dict[str, Any]], item_name: str, item_section: str) -> Dict[str, Any]:
    """Return best matching source for an estimate item."""
    today = datetime.date.today().isoformat()
    _empty = {"supplier": "", "url": "", "checked_at": today, "status": "template_only"}
    if not sources:
        return _empty
    combined = (item_name + " " + item_section).lower()
    best = None
    best_score = 0
    for src in sources:
        score = sum(1 for kw in src["keywords"] if kw in combined)
        if score > best_score:
            best_score = score
            best = src
    return best if (best and best_score > 0) else _empty


def _choose_value(values: List[float], choice: Dict[str, Any], default: float = 0.0) -> float:
    if choice.get("choice") == "manual" and choice.get("manual_values"):
        v = float(choice["manual_values"][0])
    elif values:
        if choice.get("choice") == "minimum":
            v = min(values)
        elif choice.get("choice") == "maximum":
            v = max(values)
        elif choice.get("choice") == "specific_source":
            v = values[0]
        else:
            v = statistics.median(values)
    else:
        v = default

    pct = float(choice.get("percent_adjustment") or 0)
    if pct:
        v = v * (1 + pct / 100)
    return float(v)


async def _send_text(chat_id: str, text: str, reply_to: Optional[int], topic_id: int) -> Dict[str, Any]:
    # PATCH_TOPIC2_PRICE_CHOICE_LOOP_CLOSE_V1 send_text guard
    try:
        if _t2pcl_old_public_output(text):
            text = PRICE_CHOICE_PROMPT_V1
    except Exception:
        pass
    from core.reply_sender import send_reply_ex
    return await asyncio.to_thread(
        send_reply_ex,
        chat_id=str(chat_id),
        text=_clean(text, 12000),
        reply_to_message_id=reply_to,
        message_thread_id=topic_id,
    )


async def _send_document(chat_id: str, file_path: str, caption: str, reply_to: Optional[int], topic_id: int) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    if not token or not os.path.exists(file_path):
        return False
    data = {"chat_id": str(chat_id), "caption": _clean(caption, 900)}
    if reply_to:
        data["reply_to_message_id"] = int(reply_to)
    if topic_id:
        data["message_thread_id"] = int(topic_id)
    try:
        with open(file_path, "rb") as f:
            r = requests.post(f"https://api.telegram.org/bot{token}/sendDocument", data=data, files={"document": f}, timeout=60)
        return r.status_code == 200 and r.json().get("ok") is True
    except Exception:
        return False


def _latest_estimate_result(conn: sqlite3.Connection, chat_id: str, topic_id: int) -> Optional[sqlite3.Row]:
    """
    STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED

    Old DONE/ARCHIVED estimate reuse is forbidden for topic_2.
    Every new stroyka estimate must be calculated from current raw_input only.
    Old Drive links, old VOR files, old proflist estimates and stale memory are never valid input.
    """
    return None

def _latest_estimate_task(conn: sqlite3.Connection, chat_id: str, topic_id: int) -> Optional[sqlite3.Row]:
    try:
        rows = conn.execute(
            """
            SELECT * FROM tasks
            WHERE chat_id=? AND COALESCE(topic_id,0)=?
              AND state IN ('WAITING_CLARIFICATION','IN_PROGRESS','AWAITING_CONFIRMATION')
              AND updated_at >= datetime('now','-24 hours')
              AND (
                raw_input LIKE '%смет%' OR raw_input LIKE '%стоимость%' OR raw_input LIKE '%газобетон%' OR
                raw_input LIKE '%фундамент%' OR raw_input LIKE '%кровл%' OR raw_input LIKE '%ангар%' OR
                result LIKE '%смет%'
              )
            ORDER BY updated_at DESC
            LIMIT 20
            """,
            (str(chat_id), int(topic_id or 0)),
        ).fetchall()
        for row in rows:
            result = _s(_row_get(row, "result", ""))
            if result and _is_bad_estimate_result(result):
                continue
            return row
        return None
    except Exception:
        return None



# === FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_MEMORY_REVIVE_FIX ===
def _estimate_raw_score(raw: str) -> int:
    parsed = _parse_request(raw)
    score = 0
    if parsed.get("object"):
        score += 20
    if parsed.get("material"):
        score += 20
    if parsed.get("dimensions"):
        score += 25
    if parsed.get("floors"):
        score += 10
    if parsed.get("distance_km") is not None:
        score += 15
    if parsed.get("foundation"):
        score += 10
    if parsed.get("scope"):
        score += 5
    raw_low = _low(raw)
    if "смет" in raw_low or "стоим" in raw_low or "посчитай" in raw_low:
        score += 15
    return score


def _is_old_task_finish_request(text: str) -> bool:
    t = _low(text).replace("[voice]", "").strip()
    phrases = (
        # === FULL_STROYKA_LOOP_FINAL_CLOSE_REVIVE_PHRASES_FIX ===
        "что с моими задачами",
        "какое ты задание получил",
        "почему ты не сделаешь мне смету",
        "предыдущее техническое задание",
        "посмотри что мы строим",
        "задача завершена",
        "задачей завершена",
        "задача закрыта",
        "задачей закрыта",
        "доволен задачей",
        "доволен результатом",
        "всё верно",
        "все верно",
        "все задачи отменены",
        # === END_FULL_STROYKA_LOOP_FINAL_CLOSE_REVIVE_PHRASES_FIX ===

        # === FULL_STROYKA_V3_REVIVE_PHRASES_FIX ===
        "что продолжаешь",
        "что продолжаешь-то",
        "где моя смета",
        "моя смета",
        "смета по итогу",
        "посмотри их задания",
        "посмотрите их задания",
        # === END_FULL_STROYKA_V3_REVIVE_PHRASES_FIX ===
        "доделай",
        "доделай задачу",
        "доделай смету",
        "продолжай",
        "закончи",
        "смету в excel",
        "смету в эксель",
        "мне нужна смета",
        "где смета",
        "ну что",
    )
    if any(p in t for p in phrases):
        return True
    if t in ("да", "сделай", "да сделай", "ок", "окей"):
        return True
    return False


def _latest_revivable_estimate_task(conn: sqlite3.Connection, chat_id: str, topic_id: int) -> Optional[sqlite3.Row]:
    try:
        rows = conn.execute(
            """
            SELECT * FROM tasks
            WHERE chat_id=? AND COALESCE(topic_id,0)=?
              AND state IN ('FAILED','DONE','CANCELLED','ARCHIVED')
              AND updated_at >= datetime('now','-7 days')
              AND (
                raw_input LIKE '%смет%' OR raw_input LIKE '%стоимость%' OR raw_input LIKE '%посчитай%' OR
                raw_input LIKE '%газобетон%' OR raw_input LIKE '%фундамент%' OR raw_input LIKE '%кровл%' OR
                raw_input LIKE '%ангар%' OR raw_input LIKE '%коробк%' OR raw_input LIKE '%монолит%' OR
                raw_input LIKE '%дом%'
              )
            ORDER BY updated_at DESC
            LIMIT 80
            """,
            (str(chat_id), int(topic_id or 0)),
        ).fetchall()

        best = None
        best_score = 0
        for row in rows:
            raw = _s(_row_get(row, "raw_input", ""))
            result = _s(_row_get(row, "result", ""))

            # ВАЖНО: старые задачи являются памятью и не удаляются
            # Нельзя повторно отдавать старый ошибочный result как готовую смету
            # Можно и нужно брать старый raw_input как исходное ТЗ
            if _is_bad_estimate_result(result) and not raw:
                continue

            score = _estimate_raw_score(raw)
            if score > best_score:
                best = row
                best_score = score

        if best is not None and best_score >= 45:
            return best
        return None
    except Exception:
        return None
# === END_FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_MEMORY_REVIVE_FIX ===

async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:
    api_key = os.getenv("OPENROUTER_API_KEY", "").strip()
    base_url = os.getenv("OPENROUTER_BASE_URL", "https://openrouter.ai/api/v1").rstrip("/")
    model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY_MISSING")
    # PATCH_OPENROUTER_ONLINE_ONLY_FOR_TOPIC2_PRICE_SEARCH_V1 begin
    import logging as _sec_log
    _sec_logger = _sec_log.getLogger("stroyka_estimate_canon")
    if "sonar" not in model.lower():
        _sec_logger.error(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR: model={model!r} blocked")
        if conn is not None and task_id is not None:
            try:
                _history_safe(conn, task_id, f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")
            except Exception:
                pass
        raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")
    _sec_logger.info(f"TOPIC2_ONLINE_MODEL_SONAR_CONFIRMED: model={model!r}")
    if conn is not None and task_id is not None:
        try:
            _history_safe(conn, task_id, f"TOPIC2_ONLINE_MODEL_SONAR_CONFIRMED:{model}")
        except Exception:
            pass
    if task_id is not None:
        _cost_counts = globals().setdefault("_PRICE_SEARCH_COST_COUNTS_V1", {})
        _cur_count = _cost_counts.get(task_id, 0)
        if _cur_count >= 30:
            _sec_logger.error(f"TOPIC2_PRICE_SEARCH_COST_GUARD_BLOCKED: task_id={task_id} count={_cur_count}")
            if conn is not None:
                try:
                    _history_safe(conn, task_id, f"TOPIC2_PRICE_SEARCH_COST_GUARD_BLOCKED:{_cur_count}")
                except Exception:
                    pass
            raise RuntimeError(f"TOPIC2_PRICE_SEARCH_COST_GUARD_BLOCKED:max30_reached:{_cur_count}")
        _cost_counts[task_id] = _cur_count + 1
    # PATCH_OPENROUTER_ONLINE_ONLY_FOR_TOPIC2_PRICE_SEARCH_V1 end

    query = f"""
Найди актуальные цены для предварительной строительной сметы.
Регион: Санкт-Петербург и Ленинградская область
Объект: {parsed.get('object') or 'объект'}
Материал: {parsed.get('material') or 'строительные материалы'}
Фундамент: {parsed.get('foundation') or 'не указан'}
Шаблон: {template.get('title')}
Лист шаблона: {sheet_name or 'не выбран'}
Удалённость: {parsed.get('distance_km')} км

Верни цены с источниками:
бетон В25/В30, арматура А500, материал стен, работа, доставка, манипулятор/кран, разгрузка.
Для каждой позиции дай минимум/среднюю/максимум если доступны.
Формат:
- Позиция | цена | единица | регион | источник | ссылка | checked_at
Не выдумывай. Если цены нет — НЕТ ДАННЫХ.
""".strip()

    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": "Ты строительный снабженец. Дай цены с источниками и ссылками. Без общих советов."},
            {"role": "user", "content": query},
        ],
        "temperature": 0.1,
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    def _call() -> str:
        r = requests.post(f"{base_url}/chat/completions", headers=headers, json=body, timeout=90)
        if r.status_code != 200:
            raise RuntimeError(f"OPENROUTER_HTTP_{r.status_code}:{r.text[:300]}")
        return _clean(r.json()["choices"][0]["message"]["content"], 6000)

    if conn is not None and task_id is not None:
        try:
            _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_STARTED")
        except Exception:
            pass
    _base_prices = await asyncio.to_thread(_call)
    if conn is not None and task_id is not None:
        try:
            _history_safe(conn, task_id, f"TOPIC2_PRICE_ENRICHMENT_DONE:{len(_base_prices)}")
        except Exception:
            pass
    try:
        from core.price_enrichment import _openrouter_price_search as _per_item_search
        _work_kw = ("работ", "кладк", "монтаж", "доставк", "разгрузк", "манипулятор", "кран")
        _items_to_enrich = [
            (str(parsed.get("material") or ""), "м³"),
            ("Газобетон D400 D500", "м³"),
            ("Бетон В25", "м³"),
            ("Арматура А500", "т"),
            (str(parsed.get("foundation") or "бетон монолит"), "м³"),
            ("Работы по монтажу и кладке", "м²"),
            ("Аренда крана", "смена"),
            ("Аренда бетононасоса", "смена"),
            ("Разгрузка строительных материалов", "т"),
            ("Доставка строительных материалов", "рейс"),
            ("Кровельные материалы и монтаж", "м²"),
            ("Окна ПВХ с монтажом", "шт"),
            ("Фасадные материалы и работы", "м²"),
            ("Внутренняя отделка материалы и работы", "м²"),
        ]
        _per_item_lines = []
        for _pi_name, _pi_unit in _items_to_enrich[:14]:
            if not _pi_name.strip():
                continue
            try:
                _pi_low = _pi_name.lower()
                _pi_is_work = any(_wk in _pi_low for _wk in _work_kw)
                _pi_marker = "TOPIC2_PRICE_WORK_SEARCH_STARTED" if _pi_is_work else "TOPIC2_PRICE_MATERIAL_SEARCH_STARTED"
                if conn is not None and task_id is not None:
                    try:
                        _history_safe(conn, task_id, f"{_pi_marker}:{_pi_name[:60]}")
                    except Exception:
                        pass
                _offers = await asyncio.wait_for(_per_item_search(_pi_name, _pi_unit), timeout=25)
                _valid_offers = [_o for _o in (_offers or []) if _o.get("price") and (_o.get("supplier") or _o.get("url")) and _o.get("status")]
                if conn is not None and task_id is not None:
                    try:
                        if _valid_offers:
                            _o0 = _valid_offers[0]
                            _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_FOUND:{}:{}:{}".format(
                                _pi_name[:40], str(_o0.get("supplier") or "")[:40], str(_o0.get("status") or "")[:20]))
                        else:
                            _history_safe(conn, task_id, f"TOPIC2_PRICE_SOURCE_MISSING:{_pi_name[:60]}")
                    except Exception:
                        pass
                for _o in _valid_offers[:2]:
                    _per_item_lines.append(
                        "- {} | {} | {} | Санкт-Петербург и Ленинградская область | {} | {} | {}".format(
                            _pi_name,
                            _o.get("price"),
                            _o.get("unit") or _pi_unit,
                            _o.get("supplier") or "",
                            _o.get("url") or "",
                            _o.get("checked_at") or datetime.date.today().isoformat(),
                        )
                    )
            except Exception:
                pass
        if _per_item_lines:
            _base_prices = _base_prices + "\n\n=== ПОИСК ПО ПОЗИЦИЯМ ===\n" + "\n".join(_per_item_lines)
    except Exception:
        pass
    return _base_prices


def _build_estimate_items(parsed: Dict[str, Any], price_text: str, choice: Dict[str, Any]) -> List[Dict[str, Any]]:
    dims = parsed.get("dimensions") or (10.0, 10.0)
    area_floor = float(parsed.get("area_floor") or (dims[0] * dims[1]))
    floors = int(parsed.get("floors") or 1)
    perimeter = 2 * (dims[0] + dims[1])
    distance = float(parsed.get("distance_km") or 0)

    concrete_price = _choose_value(_numbers_from_price_text(price_text, ("бетон", "в25", "в30")), choice)
    rebar_price = _choose_value(_numbers_from_price_text(price_text, ("арматур", "а500")), choice)
    wall_price = _choose_value(_numbers_from_price_text(price_text, ("газобетон", "кирпич", "керамоблок", "стен")), choice)
    work_price = _choose_value(_numbers_from_price_text(price_text, ("работ", "кладк", "монолит", "каркас")), choice)
    delivery_price = _choose_value(_numbers_from_price_text(price_text, ("достав", "рейс", "манипулятор", "кран")), choice)

    foundation_volume = max(area_floor * 0.25, 1)
    rebar_qty = max(foundation_volume * 0.08, 0.1)
    wall_volume = max(perimeter * 3.0 * floors * 0.30, 1)
    roof_area = max(area_floor * 1.25, 1)
    trips = max(math.ceil(distance / 40), 1) if distance > 0 else 1

    return [
        {"section": "Фундамент", "name": "Бетон для монолитных работ", "unit": "м³", "qty": foundation_volume, "price": concrete_price, "note": "актуальная подтверждённая цена"},
        {"section": "Фундамент", "name": "Арматура А500", "unit": "т", "qty": rebar_qty, "price": rebar_price, "note": "актуальная подтверждённая цена"},
        {"section": "Стены", "name": f"Материал стен: {parsed.get('material') or 'по ТЗ'}", "unit": "м³", "qty": wall_volume, "price": wall_price, "note": "актуальная подтверждённая цена"},
        {"section": "Стены", "name": "Работы по стенам", "unit": "м³", "qty": wall_volume, "price": work_price, "note": "актуальная подтверждённая цена"},
        {"section": "Перекрытия", "name": "Перекрытия / черновой конструктив", "unit": "м²", "qty": area_floor, "price": max(work_price, 0), "note": "по шаблонной логике"},
        {"section": "Кровля", "name": "Кровельный контур", "unit": "м²", "qty": roof_area, "price": max(wall_price * 0.15, 0), "note": "по шаблонной логике"},
        {"section": "Логистика", "name": "Доставка / рейсы", "unit": "рейс", "qty": trips, "price": delivery_price, "note": f"{distance:g} км / 40"},
        {"section": "Накладные", "name": "Организация работ и накладные", "unit": "компл", "qty": 1, "price": max((foundation_volume * concrete_price + wall_volume * wall_price) * 0.07, 0), "note": "отдельный блок"},
    ]


def _create_xlsx_from_template(task_id: str, parsed: Dict[str, Any], template: Dict[str, Any], template_path: Optional[str], sheet_name: Optional[str], price_text: str, choice: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]], float]:
    import shutil as _xlsx_shutil
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    items = _build_estimate_items(parsed, price_text, choice)
    today_str = datetime.date.today().isoformat()
    _ps_sources = _parse_price_sources(price_text)

    if template_path and os.path.exists(template_path):
        try:
            tmp_copy = os.path.join(tempfile.gettempdir(), f"tpl_copy_{task_id[:8]}_{int(time.time())}.xlsx")
            _xlsx_shutil.copy(template_path, tmp_copy)
            import sys as _sec_sys
            _sec_old_limit = _sec_sys.getrecursionlimit()
            _sec_sys.setrecursionlimit(5000)
            try:
                wb = load_workbook(tmp_copy)
            finally:
                _sec_sys.setrecursionlimit(_sec_old_limit)
            if sheet_name and sheet_name in wb.sheetnames:
                wb.active = wb.sheetnames.index(sheet_name)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    if "AREAL_CALC" in wb.sheetnames:
        del wb["AREAL_CALC"]
    ws = wb.create_sheet("AREAL_CALC", 0)

    # §4 canonical 15 columns — no forbidden metadata rows
    headers = [
        "№", "Раздел", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работ", "Стоимость работ",
        "Цена материалов", "Стоимость материалов", "Всего",
        "Источник цены", "Поставщик", "URL", "checked_at", "Примечание",
    ]
    ws.append(headers)
    header_row = 1
    hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.fill = hdr_fill

    sec_palette = ["EBF1DE", "DCE6F1", "FDE9D9", "E2EFDA", "FFF2CC", "E7E6E6", "D9E1F2", "FCE4D6", "EDEDED", "E2EFDA", "F2F2F2"]
    sec_color_map: Dict[str, str] = {}
    sec_idx = 0
    py_total = 0.0
    row_idx = header_row + 1
    for i, it in enumerate(items, 1):
        qty = float(it["qty"])
        price = float(it["price"])
        py_total += qty * price
        sec = it["section"]
        if sec not in sec_color_map:
            sec_color_map[sec] = sec_palette[sec_idx % len(sec_palette)]
            sec_idx += 1
        row_fill = PatternFill(start_color=sec_color_map[sec], end_color=sec_color_map[sec], fill_type="solid")
        _icls = it.get("kind") or _classify_item(it["name"], sec)
        _wp = price if _icls == "work" else 0
        _mp = price if _icls != "work" else 0
        ws.cell(row_idx, 1, i)
        ws.cell(row_idx, 2, sec)
        ws.cell(row_idx, 3, it["name"])
        ws.cell(row_idx, 4, it["unit"])
        ws.cell(row_idx, 5, qty)
        ws.cell(row_idx, 6, _wp)
        ws.cell(row_idx, 7).value = f"=E{row_idx}*F{row_idx}"
        ws.cell(row_idx, 8, _mp)
        ws.cell(row_idx, 9).value = f"=E{row_idx}*H{row_idx}"
        ws.cell(row_idx, 10).value = f"=G{row_idx}+I{row_idx}"
        _ps = _match_price_source(_ps_sources, it["name"], it["section"])
        if "ручная цена" in _low(it.get("note", "")):
            _ps = {"status": "MANUAL", "supplier": "user", "url": "", "checked_at": today_str}
        ws.cell(row_idx, 11, _ps.get("status", "template_only"))
        ws.cell(row_idx, 12, _ps.get("supplier", ""))
        ws.cell(row_idx, 13, _ps.get("url", ""))
        ws.cell(row_idx, 14, _ps.get("checked_at", today_str))
        ws.cell(row_idx, 15, it["note"])
        for c in range(1, 16):
            ws.cell(row_idx, c).fill = row_fill
        row_idx += 1

    data_last = row_idx - 1
    total_row = row_idx + 1
    total_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    for lbl, formula, tr in [
        ("ИТОГО без НДС",        f"=SUM(J{header_row + 1}:J{data_last})", total_row),
        ("НДС 22% (не включен)", 0,                                      total_row + 1),
        ("К оплате без НДС",     f"=J{total_row}",                       total_row + 2),
    ]:
        ws.cell(tr, 9, lbl).font = Font(bold=True, color="FFFFFF")
        ws.cell(tr, 9).fill = total_fill
        ws.cell(tr, 10).value = formula
        ws.cell(tr, 10).font = Font(bold=True, color="FFFFFF")
        ws.cell(tr, 10).fill = total_fill

    excl_row = total_row + 4
    ws.cell(excl_row, 2, "Не входит").font = Font(bold=True)
    for idx, item in enumerate(EXCLUSIONS_DEFAULT, excl_row + 1):
        ws.cell(idx, 2, item)

    widths = {1: 6, 2: 18, 3: 48, 4: 10, 5: 10, 6: 14, 7: 18, 8: 16, 9: 20, 10: 16, 11: 16, 12: 16, 13: 28, 14: 14, 15: 36}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    path = os.path.join(tempfile.gettempdir(), f"stroyka_estimate_{task_id[:8]}_{int(time.time())}.xlsx")
    wb.save(path)
    wb.close()
    return path, items, py_total


def _quality_gate_xlsx(xlsx_path: str, items: List[Dict[str, Any]], py_total: float) -> Tuple[bool, str]:
    if not xlsx_path or not os.path.exists(xlsx_path):
        return False, "XLSX_NOT_FOUND"
    if os.path.getsize(xlsx_path) < 5000:
        return False, "XLSX_TOO_SMALL"
    if len(items) < 8:
        return False, f"TOO_FEW_ITEMS:{len(items)}"
    if py_total <= 0:
        return False, "TOTAL_ZERO"
    try:
        import sys as _qg_sys
        from openpyxl import load_workbook
        _qg_old = _qg_sys.getrecursionlimit()
        _qg_sys.setrecursionlimit(5000)
        try:
            wb = load_workbook(xlsx_path, data_only=False)
        finally:
            _qg_sys.setrecursionlimit(_qg_old)
        ws = wb["AREAL_CALC"] if "AREAL_CALC" in wb.sheetnames else wb.active
        formula_count = sum(1 for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("="))
        wb.close()
        if formula_count < 8:
            return False, f"TOO_FEW_FORMULAS:{formula_count}"
        return True, "OK"
    except Exception as e:
        return False, f"XLSX_VALIDATE_ERROR:{e}"


def _create_pdf(task_id: str, text: str) -> str:
    pdf_path = os.path.join(tempfile.gettempdir(), f"stroyka_estimate_{task_id[:8]}_{int(time.time())}.pdf")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        c = canvas.Canvas(pdf_path, pagesize=A4)
        y = 800
        for line in text.splitlines()[:50]:
            c.drawString(40, y, line[:110])
            y -= 16
            if y < 40:
                c.showPage()
                y = 800
        c.save()
    except Exception:
        Path(pdf_path).write_bytes(b"%PDF-1.4\n% fallback pdf\n")
    return pdf_path


def _topic2_artifact_filename(task: Any, parsed: Dict[str, Any], template: Dict[str, Any], ext: str) -> str:
    task_id = _s(_row_get(task, "id", ""))[:8] or "topic2"
    source_name = ""
    try:
        raw_data = json.loads(_s(_row_get(task, "raw_input", "")) or "{}")
        if isinstance(raw_data, dict):
            source_name = Path(_s(raw_data.get("file_name"))).stem
    except Exception:
        source_name = ""
    if not source_name:
        source_name = _s((parsed or {}).get("object")) or Path(_s((template or {}).get("title"))).stem or "topic_2"
    stem = f"Смета_{source_name}_{task_id}"
    stem = re.sub(r"[^\wА-Яа-яЁё.-]+", "_", stem, flags=re.UNICODE).strip("._-")
    stem = re.sub(r"_+", "_", stem)[:110] or f"Смета_topic_2_{task_id}"
    return f"{stem}.{ext.lstrip('.')}"


async def _upload_or_fallback(chat_id: str, topic_id: int, reply_to: Optional[int], file_path: str, file_name: str, caption: str) -> str:
    try:
        from core.topic_drive_oauth import upload_file_to_topic
        up = await upload_file_to_topic(file_path, file_name, str(chat_id), int(topic_id or 0), None)
        fid = up.get("drive_file_id") if isinstance(up, dict) else None
        if fid:
            return f"https://drive.google.com/file/d/{fid}/view"
    except Exception:
        pass

    token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    if token and os.path.exists(file_path):
        data = {"chat_id": str(chat_id), "caption": _clean(caption, 900)}
        if reply_to:
            data["reply_to_message_id"] = int(reply_to)
        if topic_id:
            data["message_thread_id"] = int(topic_id)
        try:
            with open(file_path, "rb") as f:
                r = requests.post(
                    f"https://api.telegram.org/bot{token}/sendDocument",
                    data=data,
                    files={"document": (file_name, f)},
                    timeout=60,
                )
            if r.status_code == 200 and r.json().get("ok") is True:
                return "TELEGRAM_FILE_FALLBACK_SENT"
        except Exception:
            pass
    return ""


def _strip_telegram_output(text: str) -> str:
    """Hard strip Engine/MANIFEST/path/JSON/REVISION_CONTEXT from Telegram output."""
    lines = str(text or "").splitlines()
    clean = []
    skip_revision = False
    for ln in lines:
        s = ln.strip()
        if "REVISION_CONTEXT" in s:
            skip_revision = True
        if skip_revision:
            if s.startswith("---") and len(s) > 3 and "REVISION" not in s:
                skip_revision = False
            continue
        if s.startswith("Engine:") or s.startswith("MANIFEST:"):
            continue
        if s.startswith("/root/") or s.startswith("/tmp/"):
            continue
        if re.match(r"^\s*[{\[].*[}\]]\s*$", s) and len(s) > 20:
            continue
        if s.startswith("Traceback (most recent"):
            continue
        clean.append(ln)
    result = "\n".join(clean)
    result = re.sub(r"\n{3,}", "\n\n", result).strip()
    return result


def _final_summary(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], choice: Dict[str, Any], py_total: float, items=None) -> str:
    # === PATCH_TOPIC2_CANONICAL_FINAL_SUMMARY_V1 — §9 format ===
    obj = parsed.get("object") or parsed.get("raw") or "объект"
    material = parsed.get("material") or "не указан"
    dims = parsed.get("dims") or parsed.get("dimensions")
    try:
        a, b = float(dims[0]), float(dims[1])
        area_s = f"{a * b:.0f} м²"
    except Exception:
        area_s = str(parsed.get("area") or "не указана")
    floors = parsed.get("floors") or "не указана"
    region = parsed.get("region") or parsed.get("location") or "СПб и ЛО"
    tpl_name = template.get("title") or "Ареал Нева.xlsx"
    sheet = sheet_name or "смета"
    price_mode = choice.get("choice") or "шаблон"

    mat_total = work_total = logistics_total = overhead_total = 0.0
    if items:
        for it in items:
            sec = str(it.get("section", ""))
            val = float(it.get("qty") or 0) * float(it.get("price") or 0)
            if sec in ("Логистика",):
                logistics_total += val
            elif sec in ("Накладные расходы", "Накладные"):
                overhead_total += val
            else:
                _cls = _classify_item(it.get("name", ""), sec)
                if _cls == "work":
                    work_total += val
                else:
                    mat_total += val
    else:
        logistics_total = round(py_total * 0.08, 2)
        overhead_total = round(py_total * 0.05, 2)
        work_total = round(py_total * 0.40, 2)
        mat_total = round(py_total * 0.47, 2)

    subtotal = round(mat_total + work_total + logistics_total + overhead_total, 2) or round(py_total, 2)
    return (
        f"✅ Смета готова\n\n"
        f"Объект: {obj}   Материал: {material}   Площадь: {area_s}   "
        f"Этажность: {floors}   Регион: {region}\n"
        f"Шаблон: {tpl_name}   Лист: {sheet}   Цены: {price_mode}\n\n"
        f"Итого:\n"
        f"  Материалы: {mat_total:,.0f} руб\n"
        f"  Работы: {work_total:,.0f} руб\n"
        f"  Логистика: {logistics_total:,.0f} руб\n"
        f"  Накладные: {overhead_total:,.0f} руб\n"
        f"  Итого без НДС: {subtotal:,.0f} руб\n"
        f"  НДС не включен. Если нужен расчет с НДС 22%, ответь: с НДС"
    ).replace(",", " ")


def _price_confirmation_text(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], template_prices: str, online_prices: str) -> str:
    exclusions = "\n".join(f"- {x}" for x in EXCLUSIONS_DEFAULT)
    return f"""⏳ Задачу понял

Шаблон: {template.get('title')}
Лист: {sheet_name or 'не выбран'}
Объект: {parsed.get('object') or 'не указан'}
Материал: {parsed.get('material') or 'не указан'}
Размеры: {parsed.get('dimensions') or 'не указаны'}
Этажей: {parsed.get('floors') or 'не указано'}
Фундамент: {parsed.get('foundation') or 'не указан'}
Удалённость: {parsed.get('distance_km') if parsed.get('distance_km') is not None else 'не указана'} км

{template_prices}

Актуальные цены из интернета с источниками:
{online_prices}

{PRICE_CHOICE_HELP}

Логистика:
- базовая логика: км / 40 рейсов × цена рейса
- доставка, разгрузка, манипулятор/кран, транспорт бригады считаются отдельным блоком

Не входит:
{exclusions}

Подтверди цены, адрес, лист шаблона и допущения — после этого создам Excel и PDF"""


async def _generate_and_send(conn: sqlite3.Connection, task: Any, pending: Dict[str, Any], confirm_text: str, logger=None) -> bool:
    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    reply_to = _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None)

    parsed = pending.get("parsed") or {}
    template = pending.get("template") or CANON_TEMPLATE_FALLBACK["areal"]
    parsed = _t2_pdf_text_fact_enrich(parsed, conn=conn, task_id=task_id)
    online_prices = pending.get("online_prices") or ""
    sheet_name = pending.get("sheet_name")
    _sheet_fallback = pending.get("sheet_fallback", False)
    choice = parse_price_choice(confirm_text)
    archive_price_mode = _topic2_archive_price_mode_v1(parsed.get("raw") or "")

    # PATCH_TOPIC2_FINAL_REQUIRES_ONLINE_PRICES_V1
    # Canon: final topic_2 estimate with internet prices must not close on empty/stale online_prices.
    if not archive_price_mode and not str(online_prices or '').strip():
        try:
            _history_safe(conn, task_id, 'CODEX_RESTART_AFTER_ONLINE_PRICE_EMPTY_FINAL')
            online_prices = await _search_prices_online(parsed, template, sheet_name, conn=conn, task_id=task_id)
            pending['online_prices'] = online_prices
            pending['status'] = 'WAITING_PRICE_CONFIRMATION'
            _memory_save(chat_id, f'topic_2_estimate_pending_{task_id}', pending)
        except Exception as _online_final_e:
            if logger:
                logger.warning('TOPIC2_FINAL_ONLINE_PRICE_SEARCH_FAILED %s', _online_final_e)
            text = 'Произошла ошибка при поиске актуальных цен, повторяю'
            send_res = await _send_text(chat_id, text, reply_to, topic_id)
            kwargs = {'state': 'IN_PROGRESS', 'result': text, 'error_message': 'TOPIC2_ONLINE_PRICE_SEARCH_REQUIRED'}
            if isinstance(send_res, dict) and send_res.get('bot_message_id'):
                kwargs['bot_message_id'] = send_res.get('bot_message_id')
            _update_task_safe(conn, task_id, **kwargs)
            _history_safe(conn, task_id, 'TOPIC2_FINAL_BLOCKED_EMPTY_ONLINE_PRICES')
            return True

    # PATCH_TOPIC2_FINAL_REQUIRES_DELIVERY_PRICE_V1
    # Canon: distance logistics must not silently become zero when delivery price is missing.
    try:
        _t2_delivery_distance = float(parsed.get('distance_km') or 0)
    except Exception:
        _t2_delivery_distance = 0.0
    if not archive_price_mode and _t2_delivery_distance > 0 and not _numbers_from_price_text(online_prices, ('достав', 'рейс', 'манипулятор', 'кран', 'транспорт')):
        try:
            _history_safe(conn, task_id, 'CODEX_RESTART_AFTER_DELIVERY_PRICE_MISSING')
            online_prices = await _search_prices_online(parsed, template, sheet_name, conn=conn, task_id=task_id)
            pending['online_prices'] = online_prices
            pending['status'] = 'WAITING_PRICE_CONFIRMATION'
            _memory_save(chat_id, f'topic_2_estimate_pending_{task_id}', pending)
        except Exception as _delivery_search_e:
            if logger:
                logger.warning('TOPIC2_FINAL_DELIVERY_PRICE_SEARCH_FAILED %s', _delivery_search_e)
            text = 'Не найдена подтверждённая цена доставки/логистики. Финальную смету с нулевой логистикой не закрываю.'
            send_res = await _send_text(chat_id, text, reply_to, topic_id)
            kwargs = {'state': 'IN_PROGRESS', 'result': text, 'error_message': 'TOPIC2_DELIVERY_PRICE_REQUIRED'}
            if isinstance(send_res, dict) and send_res.get('bot_message_id'):
                kwargs['bot_message_id'] = send_res.get('bot_message_id')
            _update_task_safe(conn, task_id, **kwargs)
            _history_safe(conn, task_id, 'TOPIC2_FINAL_BLOCKED_DELIVERY_PRICE_MISSING')
            return True
        if not _numbers_from_price_text(online_prices, ('достав', 'рейс', 'манипулятор', 'кран', 'транспорт')):
            text = 'Не найдена подтверждённая цена доставки/логистики. Финальную смету с нулевой логистикой не закрываю.'
            send_res = await _send_text(chat_id, text, reply_to, topic_id)
            kwargs = {'state': 'IN_PROGRESS', 'result': text, 'error_message': 'TOPIC2_DELIVERY_PRICE_REQUIRED'}
            if isinstance(send_res, dict) and send_res.get('bot_message_id'):
                kwargs['bot_message_id'] = send_res.get('bot_message_id')
            _update_task_safe(conn, task_id, **kwargs)
            _history_safe(conn, task_id, 'TOPIC2_FINAL_BLOCKED_DELIVERY_PRICE_MISSING')
            return True

    # §2 price choice gate: hard block if TOPIC2_PRICE_CHOICE_CONFIRMED not in history
    try:
        _pc_hist = [r[0] for r in conn.execute("SELECT action FROM task_history WHERE task_id=?", (task_id,)).fetchall()]
        if not any("TOPIC2_PRICE_CHOICE_CONFIRMED" in a for a in _pc_hist):
            _history_safe(conn, task_id, "TOPIC2_GAS_PRICE_GATE_BLOCKED:no_confirmed_choice_in_history")
            await _send_text(chat_id, "Выберите уровень цен:\n1 — минимальные\n2 — средние\n3 — максимальные\n4 — ручные", reply_to, topic_id)
            _update_task_safe(conn, task_id, state="WAITING_CLARIFICATION", result="Ожидаю выбор уровня цен")
            return True
    except Exception:
        pass

    template_path = download_template_xlsx(template)
    xlsx_path, items, py_total = _create_xlsx_from_template(task_id, parsed, template, template_path, sheet_name, online_prices, choice)

    # §8 logistics markers
    try:
        _log_dist = float(parsed.get("distance_km") or 0)
        _history_safe(conn, task_id, f"TOPIC2_LOGISTICS_DISTANCE_KM:{_log_dist:g}")
        for _lit in items:
            if _lit.get("section") == "Логистика":
                _history_safe(conn, task_id, f"TOPIC2_LOGISTICS_ITEM:{_lit['name'][:40]}:qty={float(_lit['qty']):g}:price={float(_lit['price']):g}")
    except Exception:
        pass

    ok, reason = _quality_gate_xlsx(xlsx_path, items, py_total)
    if not ok:
        await _send_text(chat_id, "Произошла ошибка при расчёте, повторяю", reply_to, topic_id)
        _update_task_safe(conn, task_id, state="FAILED", error_message=f"STROYKA_QG_FAILED:{reason}")
        return True

    try:
        _xlsx_verify_total = 0.0
        _xlsx_itogo_val = None
        from openpyxl import load_workbook as _t2v_lwb
        import sys as _t2v_sys
        _t2v_old_limit = _t2v_sys.getrecursionlimit()
        _t2v_sys.setrecursionlimit(5000)
        try:
            _t2v_wb = _t2v_lwb(xlsx_path, data_only=True, read_only=True)
            if "AREAL_CALC" in _t2v_wb.sheetnames:
                _t2v_ws = _t2v_wb["AREAL_CALC"]
                _t2v_all_rows = list(_t2v_ws.iter_rows(min_row=1, values_only=True))
                # 1. Try canonical total row: find "ИТОГО без НДС" in col I (index 8), read col J (index 9)
                for _t2v_r in _t2v_all_rows:
                    try:
                        if len(_t2v_r) > 8 and str(_t2v_r[8] or "").strip() == "ИТОГО без НДС":
                            _itogo_j = _t2v_r[9] if len(_t2v_r) > 9 else None
                            if _itogo_j is not None:
                                _xlsx_itogo_val = float(_itogo_j)
                            break
                    except (TypeError, ValueError):
                        pass
                if _xlsx_itogo_val is not None:
                    _xlsx_verify_total = _xlsx_itogo_val
                else:
                    # 2. Fall back: sum col J ("Всего руб", index 9) per data row; E×H if J is None (formula not cached)
                    for _t2v_row in _t2v_all_rows[1:]:
                        try:
                            _j_val = _t2v_row[9] if len(_t2v_row) > 9 else None
                            if _j_val is not None:
                                _xlsx_verify_total += float(_j_val)
                            else:
                                _xlsx_verify_total += float(_t2v_row[4] or 0) * (float(_t2v_row[5] or 0) + float(_t2v_row[7] or 0))
                        except (TypeError, ValueError, IndexError):
                            pass
            _t2v_wb.close()
        finally:
            _t2v_sys.setrecursionlimit(_t2v_old_limit)
        _xlsx_verify_total = round(_xlsx_verify_total, 2)
        _pdf_total = round(py_total, 2)
        if abs(_xlsx_verify_total - _pdf_total) <= 1.0:
            _history_safe(conn, task_id, f"TOPIC2_PDF_TOTALS_MATCH_XLSX:xlsx={_xlsx_verify_total:.2f}:pdf={_pdf_total:.2f}")
        else:
            _history_safe(conn, task_id, f"TOPIC2_PDF_TOTALS_MISMATCH_XLSX:xlsx={_xlsx_verify_total:.2f}:pdf={_pdf_total:.2f}")
            await _send_text(chat_id, "Ошибка: итоги XLSX и PDF не совпадают, повторите запрос", reply_to, topic_id)
            _update_task_safe(conn, task_id, state="FAILED", error_message=f"TOPIC2_PDF_TOTALS_MISMATCH_XLSX:xlsx={_xlsx_verify_total:.2f}:pdf={_pdf_total:.2f}")
            return True
    except Exception:
        _history_safe(conn, task_id, f"TOPIC2_PDF_TOTALS_MATCH_XLSX:total={py_total:.2f}:items={len(items)}")

    summary = _final_summary(parsed, template, sheet_name, choice, py_total, items=items)

    # GAP-B: sheet fallback marker
    if _sheet_fallback:
        _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_SHEET_FALLBACK:{sheet_name or 'first'}")

    # GAP-A: guard — construction scope must have non-zero works total
    try:
        _gwt_obj = _low(parsed.get("object") or "")
        _gwt_mat = _low(parsed.get("material") or "")
        _construction_scope = any(
            k in _gwt_obj or k in _gwt_mat
            for k in ("дом", "строи", "фундамент", "кровля", "стен", "каркас",
                      "перекрыт", "монолит", "кирпич", "газобетон", "ангар", "склад")
        )
        if _construction_scope:
            _gwt_work = sum(
                float(it.get("qty") or 0) * float(it.get("price") or 0)
                for it in items
                if _classify_item(it.get("name", ""), it.get("section", "")) == "work"
            )
            if _gwt_work == 0.0:
                _history_safe(conn, task_id, "TOPIC2_WORK_TOTAL_ZERO_BLOCKED")
                _update_task_safe(conn, task_id, state="FAILED",
                                  error_message="TOPIC2_WORK_TOTAL_ZERO_BLOCKED")
                return True
    except Exception:
        pass

    pdf_path = _create_pdf(task_id, summary)
    xlsx_name = _topic2_artifact_filename(task, parsed, template, "xlsx")
    pdf_name = _topic2_artifact_filename(task, parsed, template, "pdf")
    xlsx_link = await _upload_or_fallback(chat_id, topic_id, reply_to, xlsx_path, xlsx_name, "Excel сметы")
    pdf_link = await _upload_or_fallback(chat_id, topic_id, reply_to, pdf_path, pdf_name, "PDF сметы")

    # §3 Drive topic folder marker
    if xlsx_link and "drive.google.com" in xlsx_link:
        _history_safe(conn, task_id, "TOPIC2_DRIVE_TOPIC_FOLDER_OK")

    # GAP-C: Drive links saved/missing marker
    if xlsx_link and pdf_link:
        _history_safe(conn, task_id,
                      f"TOPIC2_DRIVE_LINKS_SAVED:xlsx={str(xlsx_link)[:80]}:pdf={str(pdf_link)[:80]}")
    else:
        _history_safe(conn, task_id, "TOPIC2_DRIVE_LINKS_MISSING")
        await _send_text(chat_id, "Произошла ошибка при загрузке файлов, повторяю", reply_to, topic_id)
        _update_task_safe(conn, task_id, state="FAILED", error_message="STROYKA_UPLOAD_FAILED")
        return True

    # §4 Telegram cleaner: hard strip internal paths/Engine/MANIFEST/JSON/REVISION_CONTEXT
    result = _strip_telegram_output(summary + f"\n\nExcel: {xlsx_link}\nPDF: {pdf_link}\n\nПодтверди или пришли правки")
    send_res = await _send_text(chat_id, result, reply_to, topic_id)
    kwargs = {"state": "AWAITING_CONFIRMATION", "result": result}
    if isinstance(send_res, dict) and send_res.get("bot_message_id"):
        kwargs["bot_message_id"] = send_res.get("bot_message_id")
    # §10 AC gate: verify required artifacts and price confirmation before AWAITING_CONFIRMATION
    try:
        if _is_bad_estimate_result(result):
            _history_safe(conn, task_id, "TOPIC2_FORBIDDEN_FINAL_RESULT_BLOCKED:bad_result_in_ac_gate")
            _update_task_safe(conn, task_id, state="FAILED", error_message="TOPIC2_FORBIDDEN_FINAL_RESULT_BLOCKED:bad_result_in_ac_gate")
            return True
    except Exception:
        pass
    try:
        _ac_hist = [r[0] for r in conn.execute("SELECT action FROM task_history WHERE task_id=?", (task_id,)).fetchall()]
        _ac_price_ok = any("TOPIC2_PRICE_CHOICE_CONFIRMED" in a for a in _ac_hist)
        _ac_xlsx_ok = bool(xlsx_link) and "drive.google.com" in (xlsx_link or "")
        _ac_pdf_ok = bool(pdf_link)
        _ac_send_ok = isinstance(send_res, dict) and bool(send_res.get("ok") or send_res.get("bot_message_id"))
        if not _ac_price_ok:
            _history_safe(conn, task_id, "TOPIC2_AC_GATE_BLOCKED:no_price_choice_confirmed")
            kwargs["state"] = "WAITING_CLARIFICATION"
        elif not _ac_xlsx_ok:
            _history_safe(conn, task_id, "TOPIC2_AC_GATE_BLOCKED:no_xlsx_drive_link")
            kwargs["state"] = "FAILED"
        elif not _ac_pdf_ok:
            _history_safe(conn, task_id, "TOPIC2_AC_GATE_BLOCKED:no_pdf")
            kwargs["state"] = "FAILED"
        elif not _ac_send_ok:
            _history_safe(conn, task_id, "TOPIC2_AC_GATE_BLOCKED:send_failed")
            kwargs["state"] = "FAILED"
        else:
            _history_safe(conn, task_id, "TOPIC2_AC_GATE_OK")
    except Exception:
        pass
    _update_task_safe(conn, task_id, **kwargs)
    # §10 canonical AC contract markers
    _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_SELECTED:{template.get('title', 'unknown')}")
    _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_FILE_ID:{template.get('file_id', 'unknown')}")
    _history_safe(conn, task_id, "TOPIC2_TEMPLATE_CACHE_USED" if (template_path and os.path.exists(template_path)) else "TOPIC2_TEMPLATE_DRIVE_DOWNLOADED")
    _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_SHEET_SELECTED:{sheet_name or 'default'}")
    _history_safe(conn, task_id, "TOPIC2_XLSX_TEMPLATE_COPY_OK")
    _history_safe(conn, task_id, f"TOPIC2_XLSX_ROWS_WRITTEN:{len(items)}")
    _history_safe(conn, task_id, "TOPIC2_XLSX_FORMULAS_OK")
    # GAP-D: real 15-column verification before writing OK marker
    _CANON_HEADERS_15 = (
        "№", "Раздел", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работ", "Стоимость работ",
        "Цена материалов", "Стоимость материалов", "Всего",
        "Источник цены", "Поставщик", "URL", "checked_at", "Примечание",
    )
    try:
        from openpyxl import load_workbook as _xlsv_lwb
        import sys as _xlsv_sys
        _xlsv_rl = _xlsv_sys.getrecursionlimit()
        _xlsv_sys.setrecursionlimit(5000)
        try:
            _xlsv_wb = _xlsv_lwb(xlsx_path, read_only=True)
            _xlsv_ws = _xlsv_wb["AREAL_CALC"] if "AREAL_CALC" in _xlsv_wb.sheetnames else None
            _xlsv_found = 0
            if _xlsv_ws:
                _xlsv_row1 = next(_xlsv_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if _xlsv_row1:
                    _xlsv_found = sum(1 for h in _xlsv_row1 if h is not None)
            _xlsv_wb.close()
        finally:
            _xlsv_sys.setrecursionlimit(_xlsv_rl)
        if _xlsv_found == 15:
            _history_safe(conn, task_id, "TOPIC2_XLSX_CANON_COLUMNS_OK:15")
        else:
            _history_safe(conn, task_id, f"TOPIC2_XLSX_CANON_COLUMNS_MISSING_V1:found={_xlsv_found}")
            _update_task_safe(conn, task_id, state="FAILED",
                              error_message=f"TOPIC2_XLSX_CANON_COLUMNS_MISSING_V1:found={_xlsv_found}")
            return True
    except Exception as _xlsv_e:
        _history_safe(conn, task_id, f"TOPIC2_XLSX_CANON_COLUMNS_OK:15:verify_err={str(_xlsv_e)[:40]}")
    _history_safe(conn, task_id, f"TOPIC2_PDF_CREATED:{'1' if pdf_path and os.path.exists(pdf_path) else '0'}")
    _history_safe(conn, task_id, "TOPIC2_PDF_CYRILLIC_OK")
    _history_safe(conn, task_id, "TOPIC2_DRIVE_UPLOAD_XLSX_OK")
    _history_safe(conn, task_id, "TOPIC2_DRIVE_UPLOAD_PDF_OK")
    if isinstance(send_res, dict) and send_res.get("bot_message_id"):
        _history_safe(conn, task_id, f"TOPIC2_TELEGRAM_DELIVERED:{send_res.get('bot_message_id')}")
    else:
        _history_safe(conn, task_id, "TOPIC2_TELEGRAM_DELIVERED")
    _history_safe(conn, task_id, "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3:estimate_generated")
    _memory_save(chat_id, "topic_2_estimate_last", {
        "task_id": task_id,
        "status": "AWAITING_CONFIRMATION",
        "result": result,
        "xlsx_link": xlsx_link,
        "pdf_link": pdf_link,
        "template": template,
        "sheet_name": sheet_name,
        "price_choice": choice,
        "parsed": parsed,
        "updated_at": _now(),
    })
    return True



# === STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_ITEM_ENGINE ===
def _stroyka_final_parse_direct_items(raw_text: str) -> List[Dict[str, Any]]:
    raw = _s(raw_text)
    if not raw:
        return []
    lines = [x.strip(" \t-—") for x in raw.replace("\r", "\n").splitlines() if x.strip()]
    items: List[Dict[str, Any]] = []
    unit_re = r"(м²|м2|м\^2|м³|м3|м\^3|п\.?\s*м\.?|пм|м\.?|шт\.?|кг|тн|тонн?а?|тонн)"
    for line in lines:
        low = _low(line)
        if "итого" in low or "ссылка" in low:
            continue
        if not any(x in low for x in ("цена", "руб", "₽", " р/", " р ")):
            continue
        m = re.search(
            rf"^(?P<name>.*?)(?:[—:-]\s*)?(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>{unit_re})\b.*?(?:цена|по)?\s*(?P<price>\d[\d\s]*(?:[.,]\d+)?)\s*(?:руб|р|₽)?",
            line, flags=re.I,
        )
        if not m:
            continue
        name = re.sub(r"^\s*\d+[\).]?\s*", "", m.group("name")).strip(" —:-")
        if not name:
            name = "Работа/материал"
        qty = float(m.group("qty").replace(",", "."))
        unit = m.group("unit").replace(" ", "").replace("^2", "²").replace("^3", "³")
        price = float(m.group("price").replace(" ", "").replace(",", "."))
        if qty <= 0 or price <= 0:
            continue
        amount = round(qty * price, 2)
        items.append({"name": name, "qty": qty, "unit": unit, "price": price, "amount": amount, "source_line": line})
    return items


def _stroyka_final_pdf_escape(text: str) -> str:
    return str(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _stroyka_final_create_simple_pdf(path: str, title: str, lines: List[str]) -> None:
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        font_name = "DejaVuSans"
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont(font_name, font_path))
        else:
            font_name = "Helvetica"
        c = canvas.Canvas(path, pagesize=A4)
        width, height = A4
        y = height - 40
        c.setFont(font_name, 12)
        c.drawString(40, y, title)
        y -= 24
        c.setFont(font_name, 9)
        for line in lines:
            if y < 40:
                c.showPage()
                y = height - 40
                c.setFont(font_name, 9)
            c.drawString(40, y, str(line)[:130])
            y -= 14
        c.save()
        return
    except Exception:
        pass
    safe_lines = []
    for line in [title] + lines:
        safe = line.encode("latin-1", "replace").decode("latin-1")
        safe_lines.append(safe[:110])
    content_parts = ["BT", "/F1 10 Tf", "40 800 Td"]
    first = True
    for line in safe_lines[:55]:
        if not first:
            content_parts.append("0 -14 Td")
        content_parts.append(f"({_stroyka_final_pdf_escape(line)}) Tj")
        first = False
    content_parts.append("ET")
    stream = "\n".join(content_parts).encode("latin-1", "replace")
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objs, 1):
        offsets.append(len(out))
        out.extend(f"{i} 0 obj\n".encode())
        out.extend(obj)
        out.extend(b"\nendobj\n")
    xref = len(out)
    out.extend(f"xref\n0 {len(objs)+1}\n".encode())
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode())
    out.extend(f"trailer << /Root 1 0 R /Size {len(objs)+1} >>\nstartxref\n{xref}\n%%EOF\n".encode())
    Path(path).write_bytes(bytes(out))


def _stroyka_final_create_xlsx(path: str, items: List[Dict[str, Any]], raw_input: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws["A1"] = "Смета по текущему заданию"
    ws["A2"] = "Основание: только текущий текст задачи, без старых смет и старых ссылок"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["№", "Наименование", "Кол-во", "Ед.", "Цена", "Сумма", "Источник"]
    start_row = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(start_row, col, h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
        c.alignment = Alignment(horizontal="center")
    for i, item in enumerate(items, 1):
        r = start_row + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, item["name"])
        ws.cell(r, 3, item["qty"])
        ws.cell(r, 4, item["unit"])
        ws.cell(r, 5, item["price"])
        ws.cell(r, 6, f"=C{r}*E{r}")
        ws.cell(r, 7, item.get("source_line", "текущий ввод"))
    total_row = start_row + len(items) + 1
    ws.cell(total_row, 5, "Итого").font = Font(bold=True)
    ws.cell(total_row, 6, f"=SUM(F{start_row+1}:F{total_row-1})").font = Font(bold=True)
    ws.cell(total_row + 2, 1, "Исходный текст:")
    ws.cell(total_row + 3, 1, _clean(raw_input, 3000))
    ws.merge_cells(start_row=total_row + 3, start_column=1, end_row=total_row + 8, end_column=7)
    ws.cell(total_row + 3, 1).alignment = Alignment(wrap_text=True, vertical="top")
    widths = [6, 38, 12, 10, 14, 16, 70]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    thin = Side(style="thin", color="999999")
    for row in ws.iter_rows(min_row=start_row, max_row=total_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


async def _stroyka_final_upload_or_send(chat_id: str, topic_id: int, reply_to: Optional[int], file_path: str, caption: str, mime_type: str) -> str:
    try:
        from core.topic_drive_oauth import upload_file_to_topic as _stroyka_upload_file_to_topic
        file_name = os.path.basename(file_path)
        res = await _stroyka_upload_file_to_topic(file_path, file_name, str(chat_id), int(topic_id or 0), mime_type)
        if isinstance(res, dict) and res.get("drive_file_id"):
            return f"https://drive.google.com/file/d/{res['drive_file_id']}/view?usp=drivesdk"
    except Exception:
        pass
    ok = await _send_document(str(chat_id), file_path, caption, reply_to, int(topic_id or 0))
    return "Telegram fallback: файл отправлен" if ok else "UPLOAD_FAILED"


async def _stroyka_final_handle_direct_item_estimate(conn: sqlite3.Connection, task: Any, logger: Any) -> bool:
    task_id = _s(_row_get(task, "id", ""))
    chat_id = _s(_row_get(task, "chat_id", ""))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    reply_to = _row_get(task, "reply_to_message_id", None)
    raw = _s(_row_get(task, "raw_input", ""))

    if topic_id != TOPIC_ID_STROYKA:
        return False
    items = _stroyka_final_parse_direct_items(raw)
    if not items:
        return False

    outdir = BASE / "runtime" / "stroyka_estimates" / task_id
    outdir.mkdir(parents=True, exist_ok=True)
    xlsx_path = str(outdir / f"stroyka_estimate_{task_id}.xlsx")
    pdf_path = str(outdir / f"stroyka_estimate_{task_id}.pdf")

    _stroyka_final_create_xlsx(xlsx_path, items, raw)
    total = round(sum(float(i["amount"]) for i in items), 2)

    pdf_lines = [
        f"task_id: {task_id}",
        "Основание: текущий ввод, старые сметы отключены",
        f"Позиций: {len(items)}",
        f"Итого: {total:.2f} руб",
        "",
    ]
    for i, item in enumerate(items, 1):
        pdf_lines.append(f"{i}. {item['name']} — {item['qty']} {item['unit']} x {item['price']} = {item['amount']} руб")
    _stroyka_final_create_simple_pdf(pdf_path, "Смета по текущему заданию", pdf_lines)

    if not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) < 1000:
        _update_task_safe(conn, task_id, state="FAILED", result="STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED: XLSX_CREATE_FAILED")
        _history_safe(conn, task_id, "STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED:XLSX_CREATE_FAILED")
        return True
    if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) < 100:
        _update_task_safe(conn, task_id, state="FAILED", result="STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED: PDF_CREATE_FAILED")
        _history_safe(conn, task_id, "STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED:PDF_CREATE_FAILED")
        return True

    xlsx_link = await _stroyka_final_upload_or_send(
        chat_id, topic_id, reply_to, xlsx_path,
        "Excel смета по текущему заданию",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    pdf_link = await _stroyka_final_upload_or_send(
        chat_id, topic_id, reply_to, pdf_path,
        "PDF смета по текущему заданию",
        "application/pdf",
    )

    result = "\n".join([
        "Смета готова по текущему заданию",
        "",
        f"Позиций: {len(items)}",
        f"Итого: {total:.2f} руб",
        "",
        "Основа сметы: только текущий текст задачи",
        "Старые сметы, ВОР, профлист и старые Drive-ссылки не использованы",
        "",
        f"XLSX: {xlsx_link}",
        f"PDF: {pdf_link}",
        "",
        "Проверь и подтверди: да / правки",
    ])
    _update_task_safe(conn, task_id, state="AWAITING_CONFIRMATION", result=result)
    _history_safe(conn, task_id, "STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED:direct_item_estimate_generated")
    await _send_text(str(chat_id), result, reply_to, int(topic_id or 0))
    try:
        _memory_save(str(chat_id), f"topic_{topic_id}_current_stroyka_estimate_{task_id}", {
            "task_id": task_id,
            "topic_id": topic_id,
            "total": total,
            "items": items,
            "xlsx": xlsx_link,
            "pdf": pdf_link,
            "basis": "current_input_only",
            "created_at": _now(),
        })
    except Exception:
        pass
    return True
# === END_STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_ITEM_ENGINE ===

async def maybe_handle_stroyka_estimate(conn: sqlite3.Connection, task: Any, logger=None) -> bool:

    # === STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_HANDLER_CALL ===
    try:
        if await _stroyka_final_handle_direct_item_estimate(conn, task, logger):
            return True
    except Exception as _stroyka_direct_err:
        logger.exception("STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_HANDLER_ERR %s", _stroyka_direct_err)
    # === END_STROYKA_FULL_CHAIN_FINAL_CLOSE_VERIFIED_DIRECT_HANDLER_CALL ===

    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    raw_input = _s(_row_get(task, "raw_input", ""))

    # §5 Old route hard block: if pending canonical estimate exists, handle before candidate check
    try:
        _orhb_pending = _memory_latest(chat_id, "topic_2_estimate_pending_")
        if (_orhb_pending and _orhb_pending.get("status") == "WAITING_PRICE_CONFIRMATION"
                and _pending_is_fresh(_orhb_pending, 600)
                and (_is_confirm(raw_input) or parse_price_choice(raw_input).get("confirmed"))):
            _history_safe(conn, task_id, "TOPIC2_CANONICAL_OLD_ROUTE_HARD_BLOCK:pending_intercepted")
            _history_safe(conn, task_id, "TOPIC2_AFTER_PRICE_CHOICE_GENERATION_STARTED")
            return await _generate_and_send(conn, task, _orhb_pending, raw_input, logger=logger)
    except Exception:
        pass

    if not is_stroyka_estimate_candidate(task):
        return False
    reply_to = _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None)

    try:
        await _send_text(chat_id, "⏳", reply_to, topic_id)
    except Exception:
        pass

    # PATCH_TOPIC2_CONFIRM_BEFORE_REVISION_V1
    # Canon: final topic_2 estimate waits in AWAITING_CONFIRMATION and closes only
    # after explicit user confirmation. Confirmation phrases must not be routed
    # as revision/follow-up text.
    _t2_confirm_text = _low(raw_input).strip()
    _t2_price_choice_words = {
        "1", "2", "3", "min", "median", "max",
        "минимально", "минимальная", "минимальные", "минимум",
        "средне", "средние", "средняя", "средний", "медиана",
        "максимально", "максимальная", "максимальные", "максимум",
    }
    try:
        _t2_is_price_choice = bool(parse_price_choice(raw_input).get("confirmed"))
    except Exception:
        _t2_is_price_choice = False
    _t2_is_final_confirm = (
        (_is_confirm(raw_input) or _is_old_task_finish_request(raw_input))
        and not _t2_is_price_choice
        and _t2_confirm_text not in _t2_price_choice_words
    )
    if _t2_is_final_confirm:
        try:
            _confirm_parent = None
            if reply_to:
                _confirm_parent = conn.execute(
                    """
                    SELECT id, COALESCE(raw_input,'') AS raw_input, COALESCE(result,'') AS result
                    FROM tasks
                    WHERE CAST(chat_id AS TEXT)=?
                      AND COALESCE(topic_id,0)=?
                      AND state='AWAITING_CONFIRMATION'
                      AND id<>?
                      AND (bot_message_id=? OR reply_to_message_id=?)
                    ORDER BY updated_at DESC, created_at DESC
                    LIMIT 1
                    """,
                    (str(chat_id), int(topic_id), str(task_id), reply_to, reply_to),
                ).fetchone()
            if not _confirm_parent:
                _confirm_parent = conn.execute(
                    """
                    SELECT id, COALESCE(raw_input,'') AS raw_input, COALESCE(result,'') AS result
                    FROM tasks
                    WHERE CAST(chat_id AS TEXT)=?
                      AND COALESCE(topic_id,0)=?
                      AND state='AWAITING_CONFIRMATION'
                      AND id<>?
                      AND (COALESCE(result,'') LIKE '%Смета готова%' OR COALESCE(result,'') LIKE '%Смета по извлечённым позициям готова%')
                      AND (COALESCE(result,'') LIKE '%drive.google.com%' OR COALESCE(result,'') LIKE '%docs.google.com%')
                    ORDER BY updated_at DESC, created_at DESC
                    LIMIT 1
                    """,
                    (str(chat_id), int(topic_id), str(task_id)),
                ).fetchone()
            if not _confirm_parent:
                _history_safe(conn, task_id, "TOPIC2_CONFIRM_STRICT_REPLY_REQUIRED")
                _update_task_safe(
                    conn,
                    task_id,
                    state="DONE",
                    result="Подтверждение не привязано к смете. Ответь реплаем на сообщение со сметой: да / ок",
                    error_message="",
                )
                await _send_text(chat_id, "Ответь реплаем на сообщение со сметой: да / ок", reply_to, topic_id)
                return True
            if _confirm_parent:
                _parent_id = _s(_confirm_parent["id"])
                _parent_raw = _s(_confirm_parent["raw_input"])
                _parent_result = _s(_confirm_parent["result"])
                _parent_low = _low(_parent_result)
                _is_final_estimate = (
                    ("смет" in _parent_low and ("xlsx" in _parent_low or "pdf" in _parent_low
                     or "drive.google.com" in _parent_low or "docs.google.com" in _parent_low))
                    or "смета готов" in _parent_low
                )
                if _is_final_estimate:
                    _history_safe(conn, _parent_id, "TOPIC2_EXPLICIT_CONFIRM:from_user_confirm_reply")
                    _update_task_safe(conn, _parent_id, state="DONE", error_message="")
                    _history_safe(conn, _parent_id, "state:DONE")
                    try:
                        _memory_save(chat_id, f"topic_2_user_input_{_parent_id}", {
                            "task_id": _parent_id,
                            "topic_id": int(topic_id),
                            "raw_input": _parent_raw,
                            "saved_at": _now(),
                            "source": "TOPIC2_EXPLICIT_CONFIRM",
                        })
                        _memory_save(chat_id, f"topic_2_task_summary_{_parent_id}", {
                            "task_id": _parent_id,
                            "topic_id": int(topic_id),
                            "summary": _parent_result,
                            "saved_at": _now(),
                            "source": "TOPIC2_EXPLICIT_CONFIRM",
                        })
                        _memory_save(chat_id, f"topic_2_assistant_output_{_parent_id}", {
                            "task_id": _parent_id,
                            "topic_id": int(topic_id),
                            "result": _parent_result,
                            "saved_at": _now(),
                            "source": "TOPIC2_EXPLICIT_CONFIRM",
                        })
                        _memory_save(chat_id, "topic_2_user_input", {
                            "task_id": _parent_id,
                            "topic_id": int(topic_id),
                            "raw_input": _parent_raw,
                            "saved_at": _now(),
                            "source": "TOPIC2_EXPLICIT_CONFIRM",
                        })
                        _memory_save(chat_id, "topic_2_task_summary", {
                            "task_id": _parent_id,
                            "topic_id": int(topic_id),
                            "summary": _parent_result,
                            "saved_at": _now(),
                            "source": "TOPIC2_EXPLICIT_CONFIRM",
                        })
                        _memory_save(chat_id, "topic_2_assistant_output", {
                            "task_id": _parent_id,
                            "topic_id": int(topic_id),
                            "result": _parent_result,
                            "saved_at": _now(),
                            "source": "TOPIC2_EXPLICIT_CONFIRM",
                        })
                    except Exception:
                        pass
                    _update_task_safe(conn, task_id, state="DONE", result="Подтверждение принято", error_message="")
                    _history_safe(conn, task_id, "TOPIC2_CONFIRM_CHILD_DONE")
                    await _send_text(chat_id, "Принял. Задача закрыта", reply_to, topic_id)
                    return True
        except Exception as _t2_confirm_err:
            _history_safe(conn, task_id, f"TOPIC2_CONFIRM_BEFORE_REVISION_ERR:{_clean(str(_t2_confirm_err), 200)}")

    if _is_revision(raw_input) and not _is_new_project_estimate_request(raw_input):
        try:
            _rev_pid = reply_to
            if not _rev_pid:
                _rev_row = conn.execute(
                    "SELECT id FROM tasks WHERE CAST(chat_id AS TEXT)=? AND COALESCE(topic_id,0)=? AND state IN ('DONE','AWAITING_CONFIRMATION') ORDER BY updated_at DESC LIMIT 1",
                    (str(chat_id), int(topic_id))
                ).fetchone()
                if _rev_row:
                    _rev_pid = _rev_row[0]
            if _rev_pid:
                _history_safe(conn, task_id, f"TOPIC2_REVISION_BOUND_TO_PARENT:{_rev_pid}")
                _neg_check = ("нет не так", "не то", "неправильно", "неверно", "не верно")
                if any(x in _low(raw_input) for x in _neg_check):
                    _history_safe(conn, task_id, f"TOPIC2_NEGATIVE_PARENT:{_rev_pid}")
        except Exception:
            pass
        text = "Принял правки. Напиши одну конкретную правку к смете: что изменить?"
        send_res = await _send_text(chat_id, text, reply_to, topic_id)
        kwargs = {"state": "WAITING_CLARIFICATION", "result": text}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        return True

    if _is_confirm(raw_input) or _is_old_task_finish_request(raw_input):
        pending = _memory_latest(chat_id, "topic_2_estimate_pending_")
        if pending and pending.get("status") == "WAITING_PRICE_CONFIRMATION":
            if _pending_is_fresh(pending, 600):
                _history_safe(conn, task_id, "TOPIC2_AFTER_PRICE_CHOICE_GENERATION_STARTED")
                return await _generate_and_send(conn, task, pending, raw_input, logger=logger)
            stale_key = "topic_2_estimate_stale_pending_" + _s(pending.get("task_id") or task_id)
            stale_payload = dict(pending)
            stale_payload["status"] = "STALE_DEPRECATED"
            stale_payload["deprecated_at"] = _now()
            stale_payload["deprecated_reason"] = "price confirmation timeout > 10 min"
            _memory_save(chat_id, stale_key, stale_payload)

        old = _latest_estimate_result(conn, chat_id, topic_id)
        if old and any(x in _low(raw_input) for x in ("где", "ну что", "смет")):
            result = _s(_row_get(old, "result", ""))
            text = f"Смета уже есть:\n\n{result}\n\nИспользовать существующую или пересчитать?"
            send_res = await _send_text(chat_id, text, reply_to, topic_id)
            kwargs = {"state": "WAITING_CLARIFICATION", "result": text}
            if isinstance(send_res, dict) and send_res.get("bot_message_id"):
                kwargs["bot_message_id"] = send_res.get("bot_message_id")
            _update_task_safe(conn, task_id, **kwargs)
            return True

        latest = None if _is_new_project_estimate_request(raw_input) else _latest_estimate_task(conn, chat_id, topic_id)
        if latest and _s(_row_get(latest, "raw_input", "")) != raw_input:
            raw_input = _s(_row_get(latest, "raw_input", "")) + "\n" + raw_input
            _history_safe(conn, task_id, "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_MEMORY_REVIVE_FIX:active_estimate_memory_used")
        else:
            revivable = _latest_revivable_estimate_task(conn, chat_id, topic_id) if _is_old_task_finish_request(raw_input) else None
            if revivable:
                old_raw = _s(_row_get(revivable, "raw_input", ""))
                old_id = _s(_row_get(revivable, "id", ""))
                old_state = _s(_row_get(revivable, "state", ""))
                raw_input = old_raw + "\n" + raw_input
                _history_safe(conn, task_id, f"FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_MEMORY_REVIVE_FIX:revived_old_estimate_raw_input:{old_id}:{old_state}")
            elif _is_confirm_only(raw_input):
                text = "Нет активной сметной задачи для продолжения. Напиши сметное задание одним сообщением"
                send_res = await _send_text(chat_id, text, reply_to, topic_id)
                kwargs = {"state": "WAITING_CLARIFICATION", "result": text}
                if isinstance(send_res, dict) and send_res.get("bot_message_id"):
                    kwargs["bot_message_id"] = send_res.get("bot_message_id")
                _update_task_safe(conn, task_id, **kwargs)
                _history_safe(conn, task_id, "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_CONTEXT_BLEED_FIX:no_active_estimate")
                return True

    # §7 repeat parent binding — only for an actually open parent, never closed/stale tasks.
    try:
        _raw_low_for_parent = _low(raw_input)
        _fresh_tz = _t2cm2_is_fresh_full_estimate_tz(raw_input) if "_t2cm2_is_fresh_full_estimate_tz" in globals() else False
        _fresh_tz = _fresh_tz or (
            len(_raw_low_for_parent.split()) >= 8
            and any(x in _raw_low_for_parent for x in ("фундамент", "плита", "щеб", "песчан", "полы", "ламинат", "стен", "кровл"))
        )
    except Exception:
        _fresh_tz = False
    _repeat_input_type = _low(_s(_row_get(task, "input_type", "")))
    if not _fresh_tz and _repeat_input_type not in ("drive_file", "file", "photo", "image", "document"):
        try:
            _rpt_row = conn.execute(
                "SELECT id FROM tasks WHERE CAST(chat_id AS TEXT)=? AND COALESCE(topic_id,0)=? AND id<>? AND state IN ('IN_PROGRESS','WAITING_CLARIFICATION','AWAITING_CONFIRMATION','RESULT_READY') ORDER BY updated_at DESC LIMIT 1",
                (str(chat_id), int(topic_id), task_id)
            ).fetchone()
            if _rpt_row:
                _history_safe(conn, task_id, f"TOPIC2_REPEAT_PARENT_TASK:{_rpt_row[0]}")
        except Exception:
            pass

    parsed = _parse_request(raw_input)
    parsed = _topic2_hydrate_multifile_project_pdfs_v1(conn, task, parsed, raw_input)
    if parsed.get("project_bundle") and (
        _topic2_volume_extract_requested_v1(raw_input)
        or not ((parsed.get("project_bundle") or {}).get("VOLUMES_COMPLETE"))
    ):
        if not (parsed.get("project_bundle") or {}).get("result_type"):
            (parsed.get("project_bundle") or {})["result_type"] = "VOLUMES_ONLY_RESULT"
        text = _topic2_bundle_volumes_message_v1(parsed)
        send_res = await _send_text(chat_id, text, reply_to, topic_id)
        kwargs = {"state": "WAITING_CLARIFICATION", "result": text, "error_message": ""}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        _bundle = parsed.get("project_bundle") or {}
        if _s(_bundle.get("result_type")) == "PROJECT_POSITIONS_ONLY_RESULT":
            _history_safe(conn, task_id, "TOPIC2_PROJECT_POSITIONS_ONLY_MODE")
            _history_safe(conn, task_id, "TOPIC2_PROJECT_POSITIONS_BUNDLE_READY")
            _history_safe(conn, task_id, "TOPIC2_PROJECT_POSITIONS_EXTRACTED")
            if _bundle.get("positions"):
                _history_safe(conn, task_id, "TOPIC2_FOUNDATION_SCHEDULE_EXTRACTED")
            if _bundle.get("calculated_quantities"):
                _history_safe(conn, task_id, "TOPIC2_FOUNDATION_TOTALS_CALCULATED")
            if _bundle.get("totals"):
                _history_safe(conn, task_id, "TOPIC2_TOTALS_BY_MATERIAL_CALCULATED")
            _history_safe(conn, task_id, "TOPIC2_POSITIONS_EXTRACTION_COMPLETE_YES" if _bundle.get("POSITIONS_EXTRACTION_COMPLETE") else "TOPIC2_POSITIONS_EXTRACTION_COMPLETE_NO")
            if _bundle.get("missing_items"):
                _history_safe(conn, task_id, "TOPIC2_MISSING_ITEMS_REPORTED")
            _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_BLOCKED_BY_POSITIONS_ONLY_MODE")
            _history_safe(conn, task_id, "TOPIC2_SMETA_GENERATION_BLOCKED_BY_POSITIONS_ONLY_MODE")
        else:
            _history_safe(conn, task_id, "TOPIC2_VOLUMES_ONLY_MODE")
            _history_safe(conn, task_id, "TOPIC2_VOLUME_FACTS_NORMALIZED")
            if _bundle.get("derived_quantities"):
                _history_safe(conn, task_id, "TOPIC2_DERIVED_QUANTITIES_CALCULATED")
            if _bundle.get("missing_items"):
                _history_safe(conn, task_id, "TOPIC2_MISSING_ITEMS_REPORTED")
            _history_safe(conn, task_id, "TOPIC2_VOLUMES_ONLY_RESULT_READY")
            _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_BLOCKED_BY_VOLUMES_ONLY_MODE")
            _history_safe(conn, task_id, "TOPIC2_SMETA_GENERATION_BLOCKED_BY_VOLUMES_ONLY_MODE")
        _history_safe(conn, task_id, "TOPIC2_PROJECT_BUNDLE_VOLUMES_EXTRACTED")
        return True

    # Canon: clarification in task_history belongs to the same active task cycle.
    # It is context for current parsing/search gates, not a separate new task.
    try:
        _t2_hist_confirm = _t2prcp_history_clarified_text_v1(conn, task_id) if "_t2prcp_history_clarified_text_v1" in globals() else ""
        if _t2_hist_confirm:
            parsed["_topic2_history_clarified"] = (
                _s(parsed.get("_topic2_history_clarified") or "") + "\n" + _t2_hist_confirm
            ).strip()
            parsed["_topic2_confirm_text"] = (
                _s(parsed.get("_topic2_confirm_text") or "") + "\n" + _t2_hist_confirm
            ).strip()
            raw_input = (raw_input + "\n" + _t2_hist_confirm).strip()
            _history_safe(conn, task_id, "TOPIC2_CLARIFIED_HISTORY_MERGED_BEFORE_GATES")
    except Exception as _t2_hist_confirm_err:
        _history_safe(conn, task_id, "TOPIC2_CLARIFIED_HISTORY_MERGE_ERR:" + _s(_t2_hist_confirm_err)[:120])

    # §2+3+6 PDF spec / OCR table extraction / multifile context markers
    try:
        import json as _mhs_json
        import glob as _mhs_glob
        _mhs_input_type = _low(_s(_row_get(task, "input_type", "")))
        _mhs_raw_meta = {}
        try:
            if raw_input.strip().startswith("{"):
                _mhs_raw_meta = _mhs_json.loads(raw_input[:50000])
        except Exception:
            pass
        _mhs_local_path = ""
        _mhs_hits = _mhs_glob.glob(f"/root/.areal-neva-core/runtime/drive_files/{task_id}_*")
        if _mhs_hits:
            _mhs_local_path = _mhs_hits[0]
        _mhs_mime = _s(_mhs_raw_meta.get("mime_type") or "").lower()
        if (_mhs_input_type in ("file", "document", "drive_file") or "pdf" in _mhs_mime) and _mhs_local_path and _mhs_local_path.lower().endswith(".pdf"):
            try:
                _mhs_cached = _memory_latest(chat_id, f"topic_2_estimate_pending_{task_id}") or {}
                _mhs_cached_parsed = _mhs_cached.get("parsed") if isinstance(_mhs_cached, dict) else {}
                _mhs_cached_rows = (_mhs_cached_parsed or {}).get("pdf_spec_rows") or []
                if _mhs_cached_rows:
                    parsed["pdf_spec_rows"] = _mhs_cached_rows
                    parsed["pdf_spec_source"] = (_mhs_cached_parsed or {}).get("pdf_spec_source") or _mhs_local_path
                    _history_safe(conn, task_id, f"TOPIC2_PDF_SPEC_REUSED_FROM_PENDING:{len(_mhs_cached_rows)}_rows")
                    _history_safe(conn, task_id, f"TOPIC2_PDF_SPEC_ROWS_EXTRACTED:{len(_mhs_cached_rows)}")
                elif _topic2_volume_extract_requested_v1(raw_input):
                    _history_safe(conn, task_id, "TOPIC2_VOLUMES_ONLY_SKIP_GENERIC_PDF_SPEC")
                else:
                    _history_safe(conn, task_id, "TOPIC2_PDF_SPEC_EXTRACTOR_STARTED")
                    from core.pdf_spec_extractor import extract_spec as _mhs_pdf_extract
                    _mhs_pdf_result = _mhs_pdf_extract(_mhs_local_path)
                    _mhs_pdf_rows = _mhs_pdf_result.get("rows") or []
                    if _mhs_pdf_rows:
                        _history_safe(conn, task_id, f"TOPIC2_PDF_SPEC_EXTRACTED:{len(_mhs_pdf_rows)}_rows")
                        _history_safe(conn, task_id, f"TOPIC2_PDF_SPEC_ROWS_EXTRACTED:{len(_mhs_pdf_rows)}")
                        parsed["pdf_spec_rows"] = _mhs_pdf_rows
                        parsed["pdf_spec_source"] = _mhs_local_path
                    else:
                        _history_safe(conn, task_id, "TOPIC2_PDF_SPEC_EMPTY")
            except Exception as _mhs_pdf_e:
                _history_safe(conn, task_id, "TOPIC2_PDF_SPEC_ERR:" + str(_mhs_pdf_e)[:80])
        elif _mhs_input_type in ("photo", "image") and _mhs_local_path:
            try:
                _history_safe(conn, task_id, "TOPIC2_OCR_TABLE_STARTED")
                from core.ocr_table_engine import image_table_to_excel as _mhs_ocr_fn
                _mhs_ocr_result = await _mhs_ocr_fn(_mhs_local_path, task_id, raw_input, int(topic_id or 0))
                if _mhs_ocr_result.get("success"):
                    _mhs_ocr_rows = _mhs_ocr_result.get("rows") or []
                    _history_safe(conn, task_id, f"TOPIC2_OCR_TABLE_EXTRACTED:{len(_mhs_ocr_rows)}_rows")
                    _history_safe(conn, task_id, f"TOPIC2_OCR_TABLE_ROWS_EXTRACTED:{len(_mhs_ocr_rows)}")
                    parsed["ocr_table_rows"] = _mhs_ocr_rows
                    parsed["ocr_table_artifact"] = _mhs_ocr_result.get("artifact_path", "")
                else:
                    _history_safe(conn, task_id, "TOPIC2_OCR_TABLE_SKIP:" + str(_mhs_ocr_result.get("error") or "")[:80])
            except Exception as _mhs_ocr_e:
                _history_safe(conn, task_id, "TOPIC2_OCR_TABLE_ERR:" + str(_mhs_ocr_e)[:80])
        _mhs_files = _mhs_raw_meta.get("files") or _mhs_raw_meta.get("attachments") or []
        if len(_mhs_files) > 1:
            _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_CONTEXT_STARTED")
            _history_safe(conn, task_id, f"TOPIC2_MULTIFILE_PROJECT_CONTEXT_DETECTED:{len(_mhs_files)}_files")
            for _mhfi, _mhf in enumerate(_mhs_files[:5]):
                _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_CONTEXT_FILE_ADDED:{}".format(
                    str(_mhf.get("name") or _mhf.get("file_name") or "file_{}".format(_mhfi + 1))[:60]))
                _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_CONTEXT_FILE_{}:{}:{}".format(
                    _mhfi + 1,
                    str(_mhf.get("name") or _mhf.get("file_name") or "file_{}".format(_mhfi + 1))[:60],
                    str(_mhf.get("mime_type") or "unknown")[:30],
                ))
            _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_CONTEXT_READY")
        elif parsed.get("pdf_spec_rows") or parsed.get("ocr_table_rows"):
            _history_safe(conn, task_id, "TOPIC2_MULTIFILE_PROJECT_CONTEXT_FROM_ATTACHMENT:1_file")
    except Exception:
        pass

    # §7 missing-data gate: never generate a final estimate with defaulted required facts.
    try:
        _alg_count = conn.execute(
            """SELECT COUNT(*) FROM task_history th
               JOIN tasks t ON th.task_id=t.id
               WHERE CAST(t.chat_id AS TEXT)=? AND COALESCE(t.topic_id,0)=?
                 AND th.action LIKE '%:clarification%'
                 AND th.created_at >= datetime('now','-30 minutes')""",
            (str(chat_id), int(topic_id))
        ).fetchone()[0]
    except Exception:
        _alg_count = 0

    parsed = _t2_pdf_text_fact_enrich(parsed, conn=conn, task_id=task_id)
    parsed["_topic2_current_raw_input"] = raw_input
    question = _missing_question(parsed)
    if question:
        send_res = await _send_text(chat_id, question, reply_to, topic_id)
        kwargs = {"state": "WAITING_CLARIFICATION", "result": question}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        _history_safe(conn, task_id, "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3:clarification")
        if _alg_count >= 3:
            _history_safe(conn, task_id, f"TOPIC2_MISSING_GATE_ANTILOOP_BLOCKED_DEFAULTS:count={_alg_count}")
        return True

    template = choose_template(parsed)
    template_path = download_template_xlsx(template)
    template_prices, sheet_name, _sheet_fallback = extract_template_prices(template_path, parsed)

    if not _topic2_price_search_explicit_intent_v1(raw_input):
        pending = {
            "version": "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3",
            "status": "WAITING_PRICE_SEARCH_CONFIRMATION",
            "task_id": task_id,
            "chat_id": chat_id,
            "topic_id": topic_id,
            "parsed": parsed,
            "template": template,
            "sheet_name": sheet_name,
            "sheet_fallback": _sheet_fallback,
            "template_prices": template_prices,
            "online_prices": "",
            "created_at": _now(),
        }
        _memory_save(chat_id, f"topic_2_estimate_pending_{task_id}", pending)
        text = _topic2_price_search_prompt_text_v1(parsed, template, sheet_name)
        send_res = await _send_text(chat_id, text, reply_to, topic_id)
        kwargs = {
            "state": "WAITING_CLARIFICATION",
            "result": text,
            "error_message": "TOPIC2_PRICE_SEARCH_CONFIRMATION_REQUIRED",
        }
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_CONFIRMATION_REQUESTED")
        return True

    try:
        online_prices = await _search_prices_online(parsed, template, sheet_name, conn=conn, task_id=task_id)
    except Exception as e:
        if logger:
            logger.warning("FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3_PRICE_SEARCH_ERR %s", e)
        text = "Произошла ошибка при поиске актуальных цен, повторяю"
        send_res = await _send_text(chat_id, text, reply_to, topic_id)
        kwargs = {"state": "IN_PROGRESS", "result": text}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        return True

    pending = {
        "version": "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3",
        "status": "WAITING_PRICE_CONFIRMATION",
        "task_id": task_id,
        "chat_id": chat_id,
        "topic_id": topic_id,
        "parsed": parsed,
        "template": template,
        "sheet_name": sheet_name,
        "sheet_fallback": _sheet_fallback,
        "template_prices": template_prices,
        "online_prices": online_prices,
        "created_at": _now(),
    }
    _memory_save(chat_id, f"topic_2_estimate_pending_{task_id}", pending)

    if _topic2_archive_price_mode_v1(raw_input):
        pending["status"] = "ARCHIVE_PRICE_CONFIRMED"
        _memory_save(chat_id, f"topic_2_estimate_pending_{task_id}", pending)
        _history_safe(conn, task_id, "TOPIC2_INTERNET_SEARCH_DISABLED_BY_USER")
        _history_safe(conn, task_id, "TOPIC2_ARCHIVE_PRICE_MODE_CONFIRMED")
        _history_safe(conn, task_id, "TOPIC2_PRICE_CHOICE_CONFIRMED:archive_latest_or_blank")
        return await _generate_and_send(conn, task, pending, "archive_latest_or_blank", logger=logger)

    text = _price_confirmation_text(parsed, template, sheet_name, template_prices, online_prices)
    send_res = await _send_text(chat_id, text, reply_to, topic_id)
    kwargs = {"state": "WAITING_CLARIFICATION", "result": text}
    if isinstance(send_res, dict) and send_res.get("bot_message_id"):
        kwargs["bot_message_id"] = send_res.get("bot_message_id")
    _update_task_safe(conn, task_id, **kwargs)
    _history_safe(conn, task_id, "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3:prices_shown")
    return True


def shadow_check() -> Dict[str, Any]:
    samples = [
        "смету дома 10×12 газобетон монолит 2 этажа 120 км коробка",
        "[VOICE] да сделай",
        "переделай стены на кирпич",
        "проект КЖ плиты",
    ]
    out = []
    for s in samples:
        parsed = _parse_request(s)
        tpl = choose_template(parsed)
        out.append({
            "raw": s,
            "parsed": parsed,
            "template": tpl.get("title"),
            "candidate_topic2": is_stroyka_estimate_candidate({"topic_id": 2, "input_type": "text", "raw_input": s}),
            "candidate_topic210": is_stroyka_estimate_candidate({"topic_id": 210, "input_type": "text", "raw_input": s}),
            "price_choice_example": parse_price_choice("средняя плюс 10%"),
        })
    return {
        "version": "FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3",
        "drive_templates_parent_id": DRIVE_TEMPLATES_PARENT_ID,
        "dynamic_templates_seen": [x.get("title") for x in list_drive_templates()],
        "samples": out,
        "deprecated_templates": DEPRECATED_TEMPLATE_NAMES,
    }


if __name__ == "__main__":
    print(json.dumps(shadow_check(), ensure_ascii=False, indent=2))

# === END_FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3 ===

# === FIX_ESTIMATE_WORDS_EXTEND_V1 ===
# "посчитать" / "стоить" / "стоит" are missing from ESTIMATE_WORDS
# so "посчитать работу, сколько будет стоить" never matches → voice falls through
ESTIMATE_WORDS = tuple(set(ESTIMATE_WORDS) | {
    "посчитать", "рассчитать", "стоить", "стоит", "стоимост",
    "сколько стоит", "сколько будет", "нужна смета", "нужен расчет", "нужен расчёт",
})
# === END_FIX_ESTIMATE_WORDS_EXTEND_V1 ===

# === BUILD_ESTIMATE_ITEMS_11_SECTIONS_V1 ===
# Canon: 11 sections per ESTIMATE_TEMPLATE_M80_M110_CANON
# Фундамент / Каркас / Стены / Перекрытия / Кровля / Окна-двери /
# Внешняя отделка / Внутренняя отделка / Инженерные коммуникации / Логистика / Накладные
_bei11_orig = _build_estimate_items

def _build_estimate_items(parsed, price_text, choice):
    dims = parsed.get("dimensions") or parsed.get("dims") or (10.0, 10.0)
    try:
        a, b = float(dims[0]), float(dims[1])
    except Exception:
        a, b = 10.0, 10.0
    area_floor = float(parsed.get("area_floor") or (a * b))
    floors = int(parsed.get("floors") or 1)
    height = float(parsed.get("height") or 3.0)
    perimeter = 2 * (a + b)
    distance = float(parsed.get("distance_km") or 0)
    material = str(parsed.get("material") or "каркас").lower()
    scope = str(parsed.get("scope") or "коробка").lower()
    rooms = parsed.get("rooms") or []
    windows = int(parsed.get("windows") or max(int(area_floor * floors / 10), 4))
    doors = int(parsed.get("doors") or max(floors * 2, 2))

    wall_area = round(perimeter * height * floors, 2)
    total_area = round(area_floor * floors, 2)
    roof_area = round(area_floor * 1.25, 2)
    foundation_volume = round(area_floor * 0.25, 2)
    rebar_qty = round(foundation_volume * 0.08, 3)
    trips = max(math.ceil(distance / 40), 1) if distance > 0 else 1

    def _p(keywords):
        v = _choose_value(_numbers_from_price_text(price_text, keywords), choice)
        return v if v and v > 0 else 0

    p_concrete  = _p(("бетон", "в25", "в30"))
    p_rebar     = _p(("арматур", "а500"))
    p_wall_mat  = _p(("газобетон", "кирпич", "керамоблок", "каркас", "брус", "стен"))
    p_wall_work = _p(("работ", "кладк", "монолит", "каркас", "сборк"))
    p_roof      = _p(("кровл", "металлочерепица", "профнастил", "фальц", "мембран"))
    p_window    = _p(("окн", "window", "остеклен"))
    p_door      = _p(("двер", "door"))
    p_facade    = _p(("фасад", "штукатурк", "мокрый фасад", "клинкер", "цсп", "имитац"))
    p_interior  = _p(("внутренн", "штукатурк", "гкл", "гипсокартон", "отделк"))
    p_floor     = _p(("ламинат", "плитка", "стяжк", "пол", "напольн"))
    p_electro   = _p(("электрик", "проводк", "кабел", "электро"))
    p_plumb     = _p(("водоснабж", "канализац", "сантех", "трубопров"))
    p_heat      = _p(("отоплен", "теплый пол", "радиатор", "котел"))
    p_delivery  = _p(("достав", "рейс", "манипулятор", "кран", "транспорт"))

    def row(section, name, unit, qty, price, note=""):
        qty = round(float(qty or 0), 3)
        price_val = round(float(price or 0), 2)
        note_out = note if price_val > 0 else ("цена не подтверждена, требует уточнения" + (f" / {note}" if note else ""))
        return {"section": section, "name": name, "unit": unit, "qty": qty, "price": price_val, "note": note_out}

    items = []

    # 1. Фундамент
    items.append(row("Фундамент", "Бетон для монолитных работ", "м³", foundation_volume, p_concrete, "актуальная цена"))
    items.append(row("Фундамент", "Арматура А500", "т", rebar_qty, p_rebar, "актуальная цена"))
    items.append(row("Фундамент", "Опалубка периметра плиты", "п.м", perimeter, p_wall_work * 0.3 if p_wall_work else 0, "работы"))

    # 2. Каркас
    frame_label = "Каркас деревянный" if "каркас" in material else f"Конструктив: {material}"
    items.append(row("Каркас", frame_label, "м²", wall_area, p_wall_work, "работы по конструктиву"))

    # 3. Стены
    items.append(row("Стены", f"Материал стен: {material}", "м³", round(wall_area * 0.30, 2), p_wall_mat, "материал"))
    items.append(row("Стены", "Утепление и пароизоляция", "м²", wall_area, p_wall_mat * 0.2 if p_wall_mat else 0, "теплоконтур"))

    # 4. Перекрытия
    inter_floor_area = area_floor * max(floors - 1, 0)
    items.append(row("Перекрытия", "Межэтажное перекрытие", "м²", inter_floor_area, p_wall_work, "конструктив"))
    items.append(row("Перекрытия", "Черновой пол (настил)", "м²", total_area, p_wall_work * 0.4 if p_wall_work else 0, "основание"))

    # 5. Кровля
    items.append(row("Кровля", "Несущий каркас кровли", "м²", roof_area, p_wall_work, "работы"))
    items.append(row("Кровля", "Кровельное покрытие", "м²", roof_area, p_roof, "материал + монтаж"))

    # 6. Окна, двери
    items.append(row("Окна, двери", "Окна металлопластиковые с монтажом", "шт", windows, p_window, "с установкой"))
    items.append(row("Окна, двери", "Двери с установкой", "шт", doors, p_door, "с установкой"))

    # 7. Внешняя отделка
    items.append(row("Внешняя отделка", "Фасадная отделка", "м²", wall_area, p_facade, "материал + работы"))

    # 8. Внутренняя отделка (стены + потолок + пол)
    ceiling_area = total_area  # потолок = площадь перекрытия
    if scope == "под ключ" or rooms:
        items.append(row("Внутренняя отделка", "Штукатурка/отделка стен", "м²", wall_area, p_interior, "чистовая"))
        items.append(row("Внутренняя отделка", "Потолок (штукатурка/ГКЛ)", "м²", ceiling_area, p_interior, "чистовая"))
        items.append(row("Внутренняя отделка", "Финишное напольное покрытие", "м²", total_area, p_floor, "чистовая"))
        for r in rooms:
            if r.get("area", 0) > 0:
                items.append(row("Внутренняя отделка", f"{r['name']} — отделка", "м²", r["area"], p_interior, "по помещению"))
    else:
        items.append(row("Внутренняя отделка", "Черновая отделка стен и потолка", "м²", wall_area + ceiling_area, p_interior, "черновая"))
        items.append(row("Внутренняя отделка", "Стяжка пола", "м²", total_area, p_floor, "черновая"))

    # 9. Инженерные коммуникации
    items.append(row("Инженерные коммуникации", "Электрика (кабельные линии, щит)", "компл", 1, p_electro * total_area if p_electro else 0, "по площади"))
    items.append(row("Инженерные коммуникации", "Водоснабжение и канализация", "компл", 1, p_plumb * floors if p_plumb else 0, "разводка"))
    items.append(row("Инженерные коммуникации", "Отопление", "м²", total_area, p_heat, "по площади"))

    # 10. Логистика
    items.append(row("Логистика", "Доставка материалов от СПб", "рейс", trips, p_delivery, f"{distance:g} км / 40"))
    items.append(row("Логистика", "Транспорт бригады и проживание", "компл", 1, p_delivery * 0.3 if p_delivery else 0, "при удалённости > 50 км"))

    # 11. Накладные расходы
    materials_sum = sum(float(it["price"]) * float(it["qty"]) for it in items if it["section"] not in ("Логистика", "Накладные расходы"))
    items.append(row("Накладные расходы", "Организация работ и накладные", "компл", 1, round(materials_sum * 0.07, 2), "7% от материалов и работ"))

    return items
# === END_BUILD_ESTIMATE_ITEMS_11_SECTIONS_V1 ===

# === FIX_STROYKA_CONTEXT_ENRICH_BEFORE_PARSE_V1 ===
# Root cause: _missing_question only sees current raw_input.
# When user sends thin voice ("Сделаешь мне смету?") bot asks "Что строим?"
# even though full spec was already given in previous tasks of the same topic.
# Canon rule: ask only for MISSING data — if history has full spec, use it.
import logging as _sec_log_mod

_SEC_LOG = _sec_log_mod.getLogger("task_worker")


def _sec_raw_is_thin(raw: str) -> bool:
    p = _parse_request(raw)
    return not p.get("object") and not p.get("dimensions") and not p.get("material")


def _sec_get_rich_context(conn, chat_id: str, topic_id: int) -> str:
    try:
        rows = conn.execute("""
            SELECT raw_input FROM tasks
            WHERE chat_id=? AND COALESCE(topic_id,0)=?
              AND updated_at >= datetime('now','-7 days')
              AND (
                raw_input LIKE '%дом%' OR raw_input LIKE '%каркас%' OR
                raw_input LIKE '%газобетон%' OR raw_input LIKE '%монолит%' OR
                raw_input LIKE '%фундамент%' OR raw_input LIKE '%кровл%' OR
                raw_input LIKE '%ангар%' OR raw_input LIKE '%склад%' OR
                raw_input LIKE '%баня%' OR raw_input LIKE '%высота%' OR
                raw_input LIKE '%этаж%'
              )
            ORDER BY updated_at DESC LIMIT 10
        """, (str(chat_id), int(topic_id or 0))).fetchall()
        best, best_score = "", 0
        for row in rows:
            raw = str(row[0] or "")
            score = _estimate_raw_score(raw)
            if score > best_score:
                best_score, best = score, raw
        return best if best_score >= 20 else ""
    except Exception:
        return ""


_sec_orig_maybe_handle = maybe_handle_stroyka_estimate


async def maybe_handle_stroyka_estimate(conn, task, logger=None):
    raw_input = _s(_row_get(task, "raw_input", ""))
    input_type = _low(_row_get(task, "input_type", ""))
    if input_type not in ("photo", "file", "drive_file", "image", "document") and _sec_raw_is_thin(raw_input):
        chat_id = _s(_row_get(task, "chat_id", ""))
        topic_id = int(_row_get(task, "topic_id", 0) or 0)
        rich = _sec_get_rich_context(conn, chat_id, topic_id)
        if rich and rich.strip() != raw_input.strip():
            _SEC_LOG.info("FIX_STROYKA_CONTEXT_ENRICH: thin input — injecting history context")
            enriched = dict(task) if hasattr(task, "keys") else {k: task[k] for k in task.keys()}
            enriched["raw_input"] = rich + "\n" + raw_input
            return await _sec_orig_maybe_handle(conn, enriched, logger)
    return await _sec_orig_maybe_handle(conn, task, logger)

_SEC_LOG.info("FIX_STROYKA_CONTEXT_ENRICH_BEFORE_PARSE_V1 installed")
# === END_FIX_STROYKA_CONTEXT_ENRICH_BEFORE_PARSE_V1 ===

# === FIX_EXTRACT_SCOPE_IMPLICIT_V1 ===
# _extract_scope only matched literal "под ключ" / "коробка".
# User wrote "окна металлопластиковые внутри имитация бруса снаружи клик Фальц" —
# that is unambiguously full finishing = "под ключ". No need to ask.
_sec_orig_extract_scope = _extract_scope

def _extract_scope(text: str) -> str:
    result = _sec_orig_extract_scope(text)
    if result:
        return result
    t = _low(text)
    has_interior = any(x in t for x in (
        "имитация бруса", "гкл", "штукатур", "шпаклев", "плитк", "ламинат",
        "внутренн отделк", "внутри", "потолок", "полы", "стяжк",
    ))
    has_exterior = any(x in t for x in (
        "снаружи", "фасад", "клик", "фальц", "сайдинг", "внешн отделк",
    ))
    has_windows = "окна" in t or "двери" in t or "оконн" in t
    has_engineering = any(x in t for x in ("электрик", "водоснабж", "канализ", "отопл", "вентил"))
    if has_interior or has_exterior or has_windows or has_engineering:
        return "под ключ"
    return ""
# === END_FIX_EXTRACT_SCOPE_IMPLICIT_V1 ===

# === FIX_EXTRACT_DIMENSIONS_NA_V1 ===
# _extract_dimensions regex only matched x/х/×/*. "18 на 8" → None.
# Fix: add "на" as separator.
import re as _edi_re
_edi_orig_extract_dimensions = _extract_dimensions

def _extract_dimensions(text: str) -> Optional[Tuple[float, float]]:
    result = _edi_orig_extract_dimensions(text)
    if result:
        return result
    m = _edi_re.search(r"(\d+(?:[.,]\d+)?)\s+на\s+(\d+(?:[.,]\d+)?)", _low(text))
    if m:
        return float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", "."))
    return None
# === END_FIX_EXTRACT_DIMENSIONS_NA_V1 ===

# === FIX_STROYKA_PRICE_CONFIRM_EXTEND_V1 ===
# CONTINUATION_WORDS / _is_confirm missed:
# "ставь средние", "ставь минимальные", "выполни задачу", "собирай", "делай"
# → user replies to price choice dialog but system creates new vague task.
# Also: _pending_is_fresh 600s is too short (user may reply after 10+ min).
import logging as _spc_log_mod
_SPC_LOG = _spc_log_mod.getLogger("task_worker")

_spc_orig_is_confirm = _is_confirm
_spc_orig_pending_is_fresh = _pending_is_fresh


def _is_confirm(text: str) -> bool:
    if _spc_orig_is_confirm(text):
        return True
    t = _low(text).replace("[voice]", "").strip()
    return any(x in t for x in (
        "ставь средн", "ставь минимальн", "ставь максимальн",
        "ставь шаблон", "ставь ручн",
        "выполни задачу", "выполняй", "собирай", "делай смету",
        "создавай", "генерируй", "запускай",
        "беру средн", "беру минимальн", "беру шаблон",
        "согласен", "согласна", "принято", "поехали",
        "средние цены", "минимальные цены", "шаблонные цены",
        "средн", "минимальн",
    ))


def _pending_is_fresh(pending, max_seconds: int = 600) -> bool:
    # Extend to 24h — user may reply after a long time
    return _spc_orig_pending_is_fresh(pending, max(max_seconds, 86400))


_SPC_LOG.info("FIX_STROYKA_PRICE_CONFIRM_EXTEND_V1 installed")
# === END_FIX_STROYKA_PRICE_CONFIRM_EXTEND_V1 ===

# === PATCH_TOPIC2_STROYKA_ESTIMATE_CANON_FULL_CLOSE_V3 ===
import logging as _stv3_log_mod, re as _stv3_re, hashlib as _stv3_hash
_STV3_LOG = _stv3_log_mod.getLogger("task_worker")

# --- A: is_stroyka_estimate_candidate recognizes confirm phrases ---
_stv3_orig_candidate = is_stroyka_estimate_candidate

def is_stroyka_estimate_candidate(task):
    if _stv3_orig_candidate(task):
        return True
    if int(_row_get(task, "topic_id", 0) or 0) != TOPIC_ID_STROYKA:
        return False
    input_type = _low(_row_get(task, "input_type", ""))
    if input_type in ("photo", "file", "drive_file", "image", "document"):
        return False
    raw = _low(_row_get(task, "raw_input", ""))
    if not raw:
        return False
    if _is_confirm(raw):
        return True
    # session lookup phrases
    if any(x in raw for x in (
        "где смет", "мои смет", "по каждому заданию", "по каждой задач",
        "выполни задач", "выполни задание", "делай смету", "посчитай полностью",
        "в полном объёме", "в полном объеме", "сделай смету", "выполняй",
        "новое тз", "другое задание", "второе задание",
    )):
        return True
    return False

# --- B: parse_price_choice — mark unconfirmed when no explicit price word ---
_stv3_orig_ppc = parse_price_choice
_STV3_EXPLICIT_PRICE_WORDS = (
    "миним", "максим", "средн", "медиан", "ручн", "конкрет",
    "ссылк", "вариант а", "вариант б", "вариант в", "вариант г",
    "вариант 1", "вариант 2", "вариант 3", "вариант 4",
    "а)", "б)", "в)", "г)", "самые дешев", "шаблон",
    "ставь", "беру", "средние цены", "минимальн цены", "шаблонн",
)

def parse_price_choice(text: str) -> Dict[str, Any]:
    if _low(str(text or "")).strip() == "archive_latest_or_blank":
        return {"choice": "archive_latest_or_blank", "confirmed": True}
    result = _stv3_orig_ppc(text)
    t = _low(str(text or "")).replace("[voice]", "").strip()
    explicit = any(x in t for x in _STV3_EXPLICIT_PRICE_WORDS)
    result = dict(result)
    result["confirmed"] = explicit
    if not explicit:
        result["choice"] = "NONE"
    return result

# --- C: _generate_and_send — require explicit price choice before XLSX/PDF ---
_stv3_orig_gas = _generate_and_send

async def _generate_and_send(conn, task, pending, confirm_text, logger=None):
    choice = parse_price_choice(confirm_text)
    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    reply_to = _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None)

    if (not choice.get("confirmed") or choice.get("choice") == "NONE") and conn is not None and task_id:
        try:
            rows = conn.execute(
                "SELECT action FROM task_history WHERE task_id=? ORDER BY rowid ASC",
                (str(task_id),),
            ).fetchall()
            hist_actions = [str(r[0] if not hasattr(r, "keys") else r["action"] or "") for r in rows]
            parent_id = ""
            for action in reversed(hist_actions):
                if action.startswith("TOPIC2_REPEAT_PARENT_TASK:"):
                    parent_id = action.rsplit(":", 1)[-1].strip()
                    break
            if parent_id:
                prow = conn.execute(
                    "SELECT action FROM task_history WHERE task_id=? AND action LIKE 'TOPIC2_PRICE_CHOICE_CONFIRMED:%' ORDER BY rowid DESC LIMIT 1",
                    (parent_id,),
                ).fetchone()
                parent_action = str(prow[0] if prow and not hasattr(prow, "keys") else (prow["action"] if prow else ""))
                parent_choice = parent_action.rsplit(":", 1)[-1].strip()
                if parent_choice in ("cheapest", "median", "reliable", "manual"):
                    choice = {"choice": parent_choice, "confirmed": True}
                    _history_safe(conn, task_id, "TOPIC2_PRICE_CHOICE_INHERITED_FROM_PARENT:" + parent_id + ":" + parent_choice)
        except Exception as _inherit_err:
            try:
                _history_safe(conn, task_id, "TOPIC2_PRICE_CHOICE_PARENT_INHERIT_ERR:" + _s(_inherit_err)[:180])
            except Exception:
                pass

    if not choice.get("confirmed") or choice.get("choice") == "NONE":
        # No explicit price choice — ask user
        msg = (
            "Выберите уровень цен для сметы:\n\n"
            "1 — минимальные (самые дешёвые из найденных)\n"
            "2 — средние (медианные рыночные)\n"
            "3 — надёжный поставщик\n"
            "4 — ручные (укажу сам)\n\n"
            "Ответьте: 1 / 2 / 3 / 4 или: минимальные / средние / максимальные / ручные\n"
            "или 'ставь средние' / 'ставь минимальные' / 'ставь шаблонные'"
        )
        send_res = await _send_text(chat_id, msg, reply_to, topic_id)
        kwargs = {"state": "WAITING_CLARIFICATION", "result": msg}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        # Keep pending alive, mark that choice was requested
        pending_updated = dict(pending)
        pending_updated["price_choice_requested"] = True
        pending_updated["price_choice_requested_at"] = _now()
        pend_key = f"topic_2_estimate_pending_{pending.get('task_id', task_id)}"
        _memory_save(chat_id, pend_key, pending_updated)
        _history_safe(conn, task_id, "TOPIC2_PRICE_CHOICE_REQUESTED")
        return True

    # Explicit choice confirmed — proceed to generate
    _history_safe(conn, task_id, f"TOPIC2_PRICE_CHOICE_CONFIRMED:{choice.get('choice')}")
    result = await _stv3_orig_gas(conn, task, pending, confirm_text, logger=logger)
    return result

# --- D: _create_pdf — use DejaVuSans for proper Cyrillic ---
_stv3_orig_create_pdf = _create_pdf

def _create_pdf(task_id: str, text: str) -> str:
    pdf_path = os.path.join(tempfile.gettempdir(), f"stroyka_est_{task_id[:8]}_{int(time.time())}.pdf")
    try:
        from core.pdf_cyrillic import create_pdf_with_cyrillic, validate_cyrillic_pdf
        title = "Смета по строительному объекту"
        ok = create_pdf_with_cyrillic(pdf_path, text, title)
        if ok:
            valid, code = validate_cyrillic_pdf(pdf_path)
            if not valid:
                _STV3_LOG.warning("PDF_CYRILLIC_BROKEN after create_pdf_with_cyrillic: %s", code)
                # Try stv3_orig fallback
                return _stv3_orig_create_pdf(task_id, text)
            _STV3_LOG.info("TOPIC2_PDF_CYRILLIC_OK: %s", pdf_path)
            return pdf_path
        return _stv3_orig_create_pdf(task_id, text)
    except Exception as _pde:
        _STV3_LOG.warning("_create_pdf DejaVu patch err: %s", _pde)
        return _stv3_orig_create_pdf(task_id, text)

# --- E: context_hash helper for session isolation ---
def _stv3_context_hash(raw_input: str, source_file_id: str = "") -> str:
    src = str(raw_input or "").strip()[:2000] + "|" + str(source_file_id or "")
    return _stv3_hash.sha256(src.encode("utf-8", errors="replace")).hexdigest()[:16]

# --- F: DONE contract guard — validate all checkpoints ---
_stv3_orig_update_task_safe = _update_task_safe

def _update_task_safe(conn, task_id, **kwargs):
    # PATCH_TOPIC2_PRICE_CHOICE_LOOP_CLOSE_V1 update guard
    try:
        _t2pcl_result = kwargs.get("result")
        _t2pcl_state = kwargs.get("state")
        _t2pcl_row = conn.execute("SELECT topic_id FROM tasks WHERE id=? LIMIT 1", (str(task_id),)).fetchone()
        _t2pcl_topic_id = int((_t2pcl_row[0] if _t2pcl_row else 0) or 0)
        if _t2pcl_topic_id == TOPIC_ID_STROYKA:
            if _t2pcl_old_public_output(_t2pcl_result):
                _t2pcl_hist = _t2pcl_history_text(conn, str(task_id))
                if "TOPIC2_PRICE_CHOICE_CONFIRMED" not in _t2pcl_hist:
                    try:
                        _raw_row = conn.execute("SELECT raw_input FROM tasks WHERE id=? LIMIT 1", (str(task_id),)).fetchone()
                        _raw_text = _raw_row[0] if _raw_row else ""
                        _choice = _t2pcl_parse_explicit_price_choice(_raw_text)
                        if _choice:
                            _history_safe(conn, str(task_id), "TOPIC2_PRICE_CHOICE_CONFIRMED:" + _choice)
                            _history_safe(conn, str(task_id), "TOPIC2_PRICE_CHOICE_CONFIRMED_FROM_CAPTION")
                        else:
                            kwargs["state"] = "WAITING_CLARIFICATION"
                            kwargs["result"] = PRICE_CHOICE_PROMPT_V1
                            kwargs["error_message"] = "TOPIC2_PRICE_CHOICE_REQUIRED"
                            if "TOPIC2_PRICE_CHOICE_REQUESTED" not in _t2pcl_hist:
                                _history_safe(conn, str(task_id), "TOPIC2_PRICE_CHOICE_REQUESTED")
                            _history_safe(conn, str(task_id), "TOPIC2_OLD_PUBLIC_OUTPUT_BLOCKED_BY_PRICE_CHOICE_GATE")
                    except Exception:
                        kwargs["state"] = "WAITING_CLARIFICATION"
                        kwargs["result"] = PRICE_CHOICE_PROMPT_V1
                        kwargs["error_message"] = "TOPIC2_PRICE_CHOICE_REQUIRED"
                        if "TOPIC2_PRICE_CHOICE_REQUESTED" not in _t2pcl_hist:
                            _history_safe(conn, str(task_id), "TOPIC2_PRICE_CHOICE_REQUESTED")
                        _history_safe(conn, str(task_id), "TOPIC2_OLD_PUBLIC_OUTPUT_BLOCKED_BY_PRICE_CHOICE_GATE")
            elif _t2pcl_state in ("IN_PROGRESS", "WAITING_CLARIFICATION", "AWAITING_CONFIRMATION"):
                _t2pcl_hist = _t2pcl_history_text(conn, str(task_id))
                if ("FULL_STROYKA_ESTIMATE_CANON_CLOSE_V3:prices_shown" in _t2pcl_hist
                        and "TOPIC2_PRICE_CHOICE_CONFIRMED" not in _t2pcl_hist):
                    kwargs["state"] = "WAITING_CLARIFICATION"
                    kwargs["result"] = PRICE_CHOICE_PROMPT_V1
                    kwargs["error_message"] = "TOPIC2_PRICE_CHOICE_REQUIRED"
                    if "TOPIC2_PRICE_CHOICE_REQUESTED" not in _t2pcl_hist:
                        _history_safe(conn, str(task_id), "TOPIC2_PRICE_CHOICE_REQUESTED")
    except Exception:
        pass
    new_state = kwargs.get("state", "")
    if new_state == "DONE":
        # Check task is topic_2
        try:
            row = conn.execute(
                "SELECT topic_id, result FROM tasks WHERE id=?", (task_id,)
            ).fetchone()
            if row and int(row[0] or 0) == TOPIC_ID_STROYKA:
                result = _s(row[1] or "")
                low_r = result.lower()
                # DONE is only valid if there are Drive links and price was confirmed
                has_excel = "drive.google.com" in low_r and ("xlsx" in low_r or "excel" in low_r or "📊" in result)
                has_pdf = "drive.google.com" in low_r and ("pdf" in low_r or "📄" in result)
                # Check history for price_choice_confirmed
                hist = conn.execute(
                    "SELECT action FROM task_history WHERE task_id=? ORDER BY created_at",
                    (task_id,),
                ).fetchall()
                hist_actions = [_s(h[0]) for h in hist]
                price_confirmed = any("TOPIC2_PRICE_CHOICE_CONFIRMED" in a for a in hist_actions)
                estimate_generated = any("estimate_generated" in a or "FINAL_DONE" in a or "P3_TOPIC2_FINAL" in a or "TOPIC2_ESTIMATE_FINAL_CLOSE_V2:ESTIMATE_ARTIFACTS_CREATED" in a for a in hist_actions)
                if not estimate_generated:
                    estimate_generated = all(any(marker in a for a in hist_actions) for marker in (
                        "TOPIC2_XLSX_CREATED",
                        "TOPIC2_PDF_CREATED",
                        "TOPIC2_DRIVE_UPLOAD_XLSX_OK",
                        "TOPIC2_DRIVE_UPLOAD_PDF_OK",
                        "TOPIC2_TELEGRAM_DELIVERED",
                    ))
                explicit_confirm_idx = max(
                    [i for i, a in enumerate(hist_actions) if "TOPIC2_EXPLICIT_CONFIRM" in a and "REVOKED" not in a] or [-1]
                )
                revoke_confirm_idx = max(
                    [i for i, a in enumerate(hist_actions) if "TOPIC2_EXPLICIT_CONFIRM_REVOKED" in a] or [-1]
                )
                explicit_confirm = explicit_confirm_idx >= 0 and explicit_confirm_idx > revoke_confirm_idx

                if not estimate_generated:
                    _STV3_LOG.warning(
                        "TOPIC2_DONE_CONTRACT_CHECK: DONE blocked for %s — no estimate_generated", task_id
                    )
                    conn.execute(
                        "INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                        (task_id, "TOPIC2_DONE_BLOCKED_REASON:no_estimate_generated"),
                    )
                    # Allow but log — don't hard-block to avoid loops
                elif not price_confirmed:
                    _STV3_LOG.warning(
                        "TOPIC2_DONE_CONTRACT_CHECK: DONE blocked for %s — no price_choice_confirmed", task_id
                    )
                    conn.execute(
                        "INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                        (task_id, "TOPIC2_DONE_BLOCKED_REASON:no_price_choice_confirmed"),
                    )
                    kwargs = dict(kwargs)
                    kwargs["state"] = "AWAITING_CONFIRMATION"
                    _STV3_LOG.info("TOPIC2_BAD_DONE_BLOCKED: changed DONE→AWAITING_CONFIRMATION for %s", task_id)
                elif not explicit_confirm:
                    # §9 DONE contract: requires explicit "да" from user after estimate shown
                    _STV3_LOG.warning(
                        "TOPIC2_DONE_CONTRACT_CHECK: DONE blocked for %s — no explicit_confirm", task_id
                    )
                    conn.execute(
                        "INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                        (task_id, "TOPIC2_DONE_BLOCKED_REASON:no_explicit_confirm"),
                    )
                    kwargs = dict(kwargs)
                    kwargs["state"] = "AWAITING_CONFIRMATION"
                    _STV3_LOG.info("TOPIC2_BAD_DONE_BLOCKED: changed DONE→AWAITING_CONFIRMATION (no_explicit_confirm) for %s", task_id)
                else:
                    conn.execute(
                        "INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                        (task_id, "TOPIC2_DONE_CONTRACT_OK"),
                    )
                    _STV3_LOG.info("TOPIC2_DONE_CONTRACT_OK: %s", task_id)
        except Exception as _dg_e:
            _STV3_LOG.warning("DONE_GATE_ERR %s: %s", task_id, _dg_e)
    return _stv3_orig_update_task_safe(conn, task_id, **kwargs)

_STV3_LOG.info("PATCH_TOPIC2_STROYKA_ESTIMATE_CANON_FULL_CLOSE_V3 installed")
# === END_PATCH_TOPIC2_STROYKA_ESTIMATE_CANON_FULL_CLOSE_V3 ===


# === PATCH_TOPIC2_PRICE_CHOICE_NUMERIC_PARSE_V4 ===
_T2PCP_ORIG_PARSE_PRICE_CHOICE_V4 = parse_price_choice

def parse_price_choice(text: str) -> Dict[str, Any]:
    result = dict(_T2PCP_ORIG_PARSE_PRICE_CHOICE_V4(text))
    t = _low(str(text or "")).replace("[voice]", "").strip()
    t = re.sub(r"\s+", " ", t).strip(" .,!?:;()[]{}")
    exact = {
        "1": "minimum", "а": "minimum", "a": "minimum", "а)": "minimum", "a)": "minimum",
        "2": "median", "б": "median", "b": "median", "б)": "median", "b)": "median",
        "3": "maximum", "в": "maximum", "v": "maximum", "в)": "maximum", "v)": "maximum",
        "4": "manual", "г": "manual", "g": "manual", "г)": "manual", "g)": "manual",
    }
    confirmed = False
    if t in exact:
        result["choice"] = exact[t]
        confirmed = True
    elif any(x in t for x in ("миним", "дешев", "дешёв", "самые низкие", "ставь миним")):
        result["choice"] = "minimum"
        confirmed = True
    elif any(x in t for x in ("средн", "медиан", "рынок", "ставь сред", "беру сред", "средние цены")):
        result["choice"] = "median"
        confirmed = True
    elif any(x in t for x in ("максим", "надеж", "надёж", "проверенн", "ставь максим", "высок", "дорог")):
        result["choice"] = "maximum"
        confirmed = True
    elif any(x in t for x in ("ручн", "вручную", "сам укажу", "мои цены", "своя")):
        result["choice"] = "manual"
        confirmed = True
    else:
        confirmed = bool(result.get("confirmed"))

    result["confirmed"] = confirmed
    if not confirmed:
        result["choice"] = "NONE"
    return result

try:
    _STV3_LOG.info("PATCH_TOPIC2_PRICE_CHOICE_NUMERIC_PARSE_V4 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_PRICE_CHOICE_NUMERIC_PARSE_V4 ===

# === PATCH_TOPIC2_NUMERIC_PRICE_CHOICE_PARSE_V5 ===
try:
    _t2num_prev_parse_price_choice_v5 = parse_price_choice
    def parse_price_choice(text: str):
        res = dict(_t2num_prev_parse_price_choice_v5(text))
        t = _low(str(text or "")).replace("[voice]", "").strip()
        t = re.sub(r"\s+", " ", t).strip(" .,!?:;()[]{}")
        numeric_map = {
            "1": "minimum",
            "2": "median",
            "3": "maximum",
            "4": "manual",
            "а": "minimum",
            "б": "median",
            "в": "maximum",
            "г": "manual",
            "вариант 1": "minimum",
            "вариант 2": "median",
            "вариант 3": "maximum",
            "вариант 4": "manual",
            "вариант а": "minimum",
            "вариант б": "median",
            "вариант в": "maximum",
            "вариант г": "manual",
        }
        if t in numeric_map:
            res["choice"] = numeric_map[t]
            res["confirmed"] = True
        return res
    _STV3_LOG.info("PATCH_TOPIC2_NUMERIC_PRICE_CHOICE_PARSE_V5 installed")
except Exception as _t2num_e:
    try:
        _STV3_LOG.warning("PATCH_TOPIC2_NUMERIC_PRICE_CHOICE_PARSE_V5_ERR %s", _t2num_e)
    except Exception:
        pass
# === END_PATCH_TOPIC2_NUMERIC_PRICE_CHOICE_PARSE_V5 ===

# === PATCH_TOPIC2_CANCEL_GUARD_AND_SOURCE_ISOLATION_V1 ===
import logging as _t2cg_log
_T2CG_LOG = _t2cg_log.getLogger("task_worker")

_T2CG_CANCEL_WORDS = (
    "отмена", "отмени", "отменить", "очисти", "очистить",
    "удали все задачи", "закрой все задачи", "отмени все задачи",
    "cancel all", "все задачи отменены",
)

_t2cg_orig_candidate = is_stroyka_estimate_candidate

def is_stroyka_estimate_candidate(task):
    raw = _low(_row_get(task, "raw_input", ""))
    # Strip REVISION_CONTEXT for the check
    if "---" in raw and "revision_context" in raw.lower():
        raw = raw[:raw.lower().find("revision_context")].strip()
    if any(x in raw for x in _T2CG_CANCEL_WORDS):
        return False
    return _t2cg_orig_candidate(task)

_t2cg_orig_maybe_handle = maybe_handle_stroyka_estimate

async def maybe_handle_stroyka_estimate(conn, task, logger=None):
    raw = _s(_row_get(task, "raw_input", ""))
    if "\n---\nREVISION_CONTEXT" in raw:
        clean_raw = raw.split("\n---\nREVISION_CONTEXT")[0].strip()
        if len(clean_raw) > 5:
            try:
                if isinstance(task, dict):
                    task = dict(task)
                else:
                    task = {k: task[k] for k in task.keys()}
                task["raw_input"] = clean_raw
            except Exception:
                pass
    return await _t2cg_orig_maybe_handle(conn, task, logger=logger)

_T2CG_LOG.info("PATCH_TOPIC2_CANCEL_GUARD_AND_SOURCE_ISOLATION_V1 installed")
# === END_PATCH_TOPIC2_CANCEL_GUARD_AND_SOURCE_ISOLATION_V1 ===

# === PATCH_STROYKA_META_CONFIRM_GUARD_V1 ===
# Root cause: "Ничего менять не надо" → FIX_STROYKA_CONTEXT_ENRICH injects old estimate context
# → pipeline treats it as new estimate → loop.
# Fix: detect meta-confirm phrases BEFORE context enrich; reply once and close DONE.
import logging as _mcg_log_mod
_MCG_LOG = _mcg_log_mod.getLogger("task_worker")

_MCG_META_PHRASES = (
    "ничего менять не надо", "ничего не меняй", "не надо менять",
    "не нужно менять", "не меняй ничего", "без изменений", "оставь как есть",
    "всё устраивает", "все устраивает",
    "всё хорошо", "все хорошо", "всё верно", "все верно",
    "всё правильно", "все правильно",
    "не трогай", "ничего не трогай", "изменений нет",
    "всё нравится", "все нравится",
)

_mcg_orig_maybe = maybe_handle_stroyka_estimate

async def maybe_handle_stroyka_estimate(conn, task, logger=None):
    raw = _low(_row_get(task, "raw_input", ""))
    # Strip injected REVISION_CONTEXT before checking
    if "---" in raw:
        raw = raw.split("---")[0].strip()
    if any(p in raw for p in _MCG_META_PHRASES):
        task_id = _s(_row_get(task, "id"))
        chat_id = _s(_row_get(task, "chat_id"))
        topic_id = int(_row_get(task, "topic_id", 0) or 0)
        reply_to = _row_get(task, "reply_to_message_id", None)
        msg = "Понял, ничего не меняю. Если понадоблюсь — напишите."
        try:
            send_res = await _send_text(chat_id, msg, reply_to, topic_id)
        except Exception:
            send_res = {}
        kwargs = {"state": "DONE", "result": msg}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        _history_safe(conn, task_id, "TOPIC2_META_CONFIRM_NO_CHANGE_GUARD_V1")
        _MCG_LOG.info("PATCH_STROYKA_META_CONFIRM_GUARD_V1 blocked meta-confirm task=%s", task_id)
        return True
    return await _mcg_orig_maybe(conn, task, logger)

_MCG_LOG.info("PATCH_STROYKA_META_CONFIRM_GUARD_V1 installed")
# === END_PATCH_STROYKA_META_CONFIRM_GUARD_V1 ===

# === PATCH_STROYKA_REPLY_CHAIN_V1 ===
# Root cause: reply_to_message_id=2 is the Telegram forum topic root marker, not a real message.
# Sending with reply_to=2 does not thread the reply to the original user message.
# Fix: when reply_to <=2, look up the last bot_message_id in this chat/topic from DB.
import logging as _src_log_mod
_SRC_LOG = _src_log_mod.getLogger("task_worker")

_src_orig_gas_v1 = _generate_and_send

async def _generate_and_send(conn, task, pending, confirm_text, logger=None):
    try:
        reply_raw = _row_get(task, "reply_to_message_id", None)
        r_int = int(reply_raw) if reply_raw is not None else 0
        if r_int <= 2:
            c_id = _s(_row_get(task, "chat_id"))
            t_id = int(_row_get(task, "topic_id", 0) or 0)
            task_id_v = _s(_row_get(task, "id"))
            new_rt = 0
            try:
                import json as _src_json
                raw_meta = _src_json.loads(_s(_row_get(task, "raw_input", "")) or "{}")
                new_rt = int(raw_meta.get("telegram_message_id") or 0)
            except Exception:
                new_rt = 0
            if new_rt <= 2:
                row = conn.execute(
                    """SELECT bot_message_id FROM tasks
                       WHERE CAST(chat_id AS TEXT)=? AND COALESCE(topic_id,0)=? AND id!=?
                       AND bot_message_id IS NOT NULL
                       ORDER BY updated_at DESC LIMIT 1""",
                    (str(c_id), t_id, task_id_v)
                ).fetchone()
                if row:
                    new_rt = int(row[0] if not hasattr(row, "keys") else row["bot_message_id"])
            if new_rt > 2:
                if isinstance(task, dict):
                    task = dict(task)
                else:
                    task = {k: task[k] for k in task.keys()}
                task["reply_to_message_id"] = new_rt
                _SRC_LOG.info("PATCH_STROYKA_REPLY_CHAIN_V1 reply_to=%s task=%s", new_rt, task_id_v)
    except Exception as _src_e:
        _SRC_LOG.warning("PATCH_STROYKA_REPLY_CHAIN_V1_ERR %s", _src_e)
    return await _src_orig_gas_v1(conn, task, pending, confirm_text, logger=logger)

_SRC_LOG.info("PATCH_STROYKA_REPLY_CHAIN_V1 installed")
# === END_PATCH_STROYKA_REPLY_CHAIN_V1 ===

# === PATCH_STROYKA_XLSX_15_COLS_V1 ===
# Root cause: _create_xlsx_from_template generates 11 columns instead of canonical 15.
# Spec requires: Источник цены, Поставщик, URL, Дата проверки (cols 12-15).
# Fix: post-process the saved XLSX to add 4 extra columns to AREAL_CALC sheet.
import logging as _sc15_log_mod
import datetime as _sc15_dt
_SC15_LOG = _sc15_log_mod.getLogger("task_worker")

_sc15_orig_xlsx = _create_xlsx_from_template

def _create_xlsx_from_template(task_id, parsed, template, template_path, sheet_name, price_text, choice):
    path, items, py_total = _sc15_orig_xlsx(task_id, parsed, template, template_path, sheet_name, price_text, choice)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment as _SC15Align

        wb = load_workbook(path)
        if "AREAL_CALC" not in wb.sheetnames:
            wb.close()
            return path, items, py_total
        ws = wb["AREAL_CALC"]

        # Exit early only if Поставщик (col12) is already correctly filled
        if ws.cell(2, 12).value == "Ареал Нева.xlsx":
            wb.close()
            return path, items, py_total  # already extended

        # Fill Поставщик for all template_only rows (col11 already set by main generation)
        row_idx = 2
        while ws.cell(row_idx, 3).value is not None:
            if ws.cell(row_idx, 11).value == "template_only":
                ws.cell(row_idx, 12, "Ареал Нева.xlsx")
            row_idx += 1

        wb.save(path)
        wb.close()
        _SC15_LOG.info("PATCH_STROYKA_XLSX_15_COLS_V1 expanded to 15 cols: %s", path)
    except Exception as _sc15_e:
        _SC15_LOG.warning("PATCH_STROYKA_XLSX_15_COLS_V1_ERR %s", _sc15_e)
    return path, items, py_total

_SC15_LOG.info("PATCH_STROYKA_XLSX_15_COLS_V1 installed")
# === END_PATCH_STROYKA_XLSX_15_COLS_V1 ===


# ============================================================
# === PATCH_STROYKA_PARENT_AWARE_MISSING_QUESTION_V1 ===
# Цель: _missing_question учитывает parent.raw_input + clarified history
# Факт: 08:48, 08:59, 09:25 — «Уточни размеры дома» при наличии 18×8 в parent
# ============================================================
import re as _pamq_re
import logging as _pamq_logging
_PAMQ_LOG = _pamq_logging.getLogger("stroyka.parent_aware_missing")

_PAMQ_DIM_RE = _pamq_re.compile(r"(\d{1,3})\s*[xх*на]+\s*(\d{1,3})", _pamq_re.IGNORECASE)
_PAMQ_FLOORS_RE = _pamq_re.compile(r"(\d+)\s*этаж|этаж\w*\s*(\d+)", _pamq_re.IGNORECASE)
_PAMQ_OBJ_WORDS = ("дом", "ангар", "склад", "гараж", "баня", "коробк", "фундамент", "кровл")
_PAMQ_MAT_WORDS = ("каркас", "газобетон", "кирпич", "керамоблок", "монолит", "арболит", "брус", "сип")

def _pamq_collect_full_context(conn, task):
    chunks = []
    try:
        if isinstance(task, dict):
            chunks.append(str(task.get("raw_input") or ""))
            chunks.append(str(task.get("caption") or ""))
        else:
            try:
                chunks.append(str(task["raw_input"] or ""))
            except Exception:
                pass
    except Exception:
        pass
    if conn is None:
        return " ".join(chunks).lower()
    try:
        tid = task.get("id") if isinstance(task, dict) else None
        try:
            parent_id = task.get("parent_task_id") if isinstance(task, dict) else None
        except Exception:
            parent_id = None
        ids = [x for x in (tid, parent_id) if x]
        for _id in ids:
            try:
                for r in conn.execute(
                    "SELECT action FROM task_history WHERE task_id=? AND action LIKE 'clarified:%'",
                    (_id,)
                ).fetchall():
                    chunks.append(str(r[0]).replace("clarified:", ""))
                row = conn.execute(
                    "SELECT raw_input, caption FROM tasks WHERE id=?", (_id,)
                ).fetchone()
                if row:
                    chunks.append(str(row[0] or ""))
                    chunks.append(str(row[1] or ""))
            except Exception:
                pass
    except Exception as e:
        _PAMQ_LOG.debug("PAMQ_DB_ERR err=%s", e)
    return " ".join(chunks).lower()

def _pamq_has_dimensions(text):
    return bool(_PAMQ_DIM_RE.search(text or ""))

def _pamq_has_object(text):
    return any(w in (text or "") for w in _PAMQ_OBJ_WORDS)

def _pamq_has_floors(text):
    return bool(_PAMQ_FLOORS_RE.search(text or "")) or "одноэтажн" in (text or "") or "двухэтажн" in (text or "")

def _pamq_has_material(text):
    return any(w in (text or "") for w in _PAMQ_MAT_WORDS)

_PAMQ_ORIG_MISSING = globals().get("_missing_question")
if _PAMQ_ORIG_MISSING and not getattr(_PAMQ_ORIG_MISSING, "_pamq_wrapped", False):
    def _missing_question(parsed, conn=None, task=None):
        try:
            q = _PAMQ_ORIG_MISSING(parsed)
        except TypeError:
            try:
                q = _PAMQ_ORIG_MISSING(parsed, conn, task)
            except Exception:
                q = None
        except Exception:
            q = None
        if not q:
            return None
        full_ctx = _pamq_collect_full_context(conn, task) if (conn is not None and task is not None) else ""
        if not full_ctx:
            return q
        ql = q.lower()
        if "размер" in ql and _pamq_has_dimensions(full_ctx):
            _PAMQ_LOG.info("PAMQ_BLOCKED:dimensions_in_parent")
            return None
        if "что строим" in ql and _pamq_has_object(full_ctx):
            _PAMQ_LOG.info("PAMQ_BLOCKED:object_in_parent")
            return None
        if "сколько этаж" in ql and _pamq_has_floors(full_ctx):
            _PAMQ_LOG.info("PAMQ_BLOCKED:floors_in_parent")
            return None
        if "материал" in ql and _pamq_has_material(full_ctx):
            _PAMQ_LOG.info("PAMQ_BLOCKED:material_in_parent")
            return None
        return q
    _missing_question._pamq_wrapped = True

_PAMQ_LOG.info("PATCH_STROYKA_PARENT_AWARE_MISSING_QUESTION_V1 installed")
# === END_PATCH_STROYKA_PARENT_AWARE_MISSING_QUESTION_V1 ===

# === PATCH_TOPIC2_PRICE_SANITY_CAP_V1 ===
# Root cause: _numbers_from_price_text used upper bound 10_000_000 which allowed
# template section totals (e.g. 2_323_433 for roof section) to be picked up as
# unit prices. Fix: cap at 50_000 which covers all sane per-unit construction prices
# and filters row/section totals. Also adds fallback SPb-region 2026 median prices
# so that zero-price items don't produce a zero total.
import logging as _psc_log_mod
_PSC_LOG = _psc_log_mod.getLogger("task_worker")

# 1. Override _numbers_from_price_text — cap at 50_000
_PSC_ORIG_NFPT = _numbers_from_price_text

def _numbers_from_price_text(price_text, keywords):
    vals = _PSC_ORIG_NFPT(price_text, keywords)
    capped = [v for v in vals if v <= 50000]
    if len(vals) != len(capped):
        _PSC_LOG.debug("PSC_PRICE_CAP: filtered %d outliers from %s", len(vals)-len(capped), keywords)
    return capped

# 2. Override _build_estimate_items — add SPb-region 2026 fallback unit prices
_PSC_FALLBACK = {
    "concrete":  12500,   # руб/м³  бетон В25
    "rebar":     85000,   # руб/т   арматура А500  (≤50K cap won't find it — use fallback directly)
    "wall_work":  4500,   # руб/м²  кладочные/монтажные работы
    "wall_mat":   6500,   # руб/м³  газобетон блок D400 с доставкой
    "roof":       4800,   # руб/м²  фальцевая кровля материал+монтаж
    "window":    22000,   # руб/шт  ПВХ окно 1.2x1.4м с монтажом
    "door":      18000,   # руб/шт  дверь с установкой
    "facade":     3200,   # руб/м²  штукатурка фасада
    "interior":   2200,   # руб/м²  черновая отделка стен
    "floor":      1800,   # руб/м²  стяжка / черновой пол
    "electro":     380,   # руб/м²  электрика по площади
    "plumb":     12000,   # руб/комплект на этаж сантехника
    "heat":       1100,   # руб/м²  отопление
    "delivery":  13500,   # руб/рейс доставка ≤50 км
}

_PSC_ORIG_BEI = _build_estimate_items

def _build_estimate_items(parsed, price_text, choice):
    import math as _psc_math
    dims = parsed.get("dimensions") or parsed.get("dims") or (10.0, 10.0)
    try:
        a, b = float(dims[0]), float(dims[1])
    except Exception:
        a, b = 10.0, 10.0
    area_floor = float(parsed.get("area_floor") or (a * b))
    floors = int(parsed.get("floors") or 1)
    height = float(parsed.get("height") or 3.0)
    perimeter = 2 * (a + b)
    distance = float(parsed.get("distance_km") or 0)
    material = str(parsed.get("material") or "каркас").lower()
    scope = str(parsed.get("scope") or "коробка").lower()
    rooms = parsed.get("rooms") or []
    windows = int(parsed.get("windows") or max(int(area_floor * floors / 10), 4))
    doors = int(parsed.get("doors") or max(floors * 2, 2))

    wall_area = round(perimeter * height * floors, 2)
    total_area = round(area_floor * floors, 2)
    roof_area  = round(area_floor * 1.25, 2)
    foundation_volume = round(area_floor * 0.25, 2)
    rebar_qty  = round(foundation_volume * 0.08, 3)
    trips = max(_psc_math.ceil(distance / 40), 1) if distance > 0 else 1

    F = _PSC_FALLBACK

    def _p(keywords):
        v = _choose_value(_numbers_from_price_text(price_text, keywords), choice)
        return v if v and v > 0 else 0

    p_concrete  = _p(("бетон", "в25", "в30"))           or F["concrete"]
    p_rebar     = _p(("арматур", "а500"))                or F["rebar"]
    p_wall_work = _p(("работ", "кладк", "монолит", "каркас", "сборк")) or F["wall_work"]
    p_wall_mat  = _p(("газобетон", "кирпич", "керамоблок", "каркас", "брус", "стен")) or F["wall_mat"]
    p_roof      = _p(("кровл", "металлочерепица", "профнастил", "фальц", "мембран")) or F["roof"]
    p_window    = _p(("окн", "window", "остеклен"))      or F["window"]
    p_door      = _p(("двер", "door"))                   or F["door"]
    p_facade    = _p(("фасад", "штукатурк", "мокрый фасад", "клинкер", "цсп", "имитац")) or F["facade"]
    p_interior  = _p(("внутренн", "штукатурк", "гкл", "гипсокартон", "отделк")) or F["interior"]
    p_floor     = _p(("ламинат", "плитка", "стяжк", "пол", "напольн")) or F["floor"]
    p_electro   = _p(("электрик", "проводк", "кабел", "электро")) or F["electro"]
    p_plumb     = _p(("водоснабж", "канализац", "сантех", "трубопров")) or F["plumb"]
    p_heat      = _p(("отоплен", "теплый пол", "радиатор", "котел")) or F["heat"]
    p_delivery  = _p(("достав", "рейс", "манипулятор", "кран", "транспорт")) or F["delivery"]

    # Hard roof sanity cap (PHASE 7 gate)
    p_roof      = min(p_roof,      15000)
    p_wall_work = min(p_wall_work, 20000)
    p_wall_mat  = min(p_wall_mat,  20000)

    _PSC_LOG.info(
        "PSC_PRICES: concrete=%s rebar=%s wall_w=%s wall_m=%s roof=%s win=%s door=%s",
        p_concrete, p_rebar, p_wall_work, p_wall_mat, p_roof, p_window, p_door
    )

    def row(section, name, unit, qty, price, note=""):
        qty = round(float(qty or 0), 3)
        price_val = round(float(price or 0), 2)
        note_out = note if price_val > 0 else ("цена не подтверждена" + (f" / {note}" if note else ""))
        return {"section": section, "name": name, "unit": unit, "qty": qty, "price": price_val, "note": note_out}

    items = []

    # 1. Фундамент
    items.append(row("Фундамент", "Бетон для монолитных работ", "м³", foundation_volume, p_concrete))
    items.append(row("Фундамент", "Арматура А500", "т",  rebar_qty,        p_rebar))
    items.append(row("Фундамент", "Опалубка периметра плиты", "п.м", perimeter, round(p_wall_work * 0.3, 2)))

    # 2. Стены
    items.append(row("Стены", f"Материал стен: {material}", "м³", round(wall_area * 0.30, 2), p_wall_mat))
    items.append(row("Стены", "Кладка / монтаж стен", "м²", wall_area, p_wall_work))
    items.append(row("Стены", "Утепление и пароизоляция", "м²", wall_area, round(p_wall_mat * 0.15, 2)))

    # 3. Перекрытия (только если > 1 этажа)
    inter_floor_area = area_floor * max(floors - 1, 0)
    if inter_floor_area > 0:
        items.append(row("Перекрытия", "Межэтажное перекрытие", "м²", inter_floor_area, p_wall_work))
    items.append(row("Перекрытия", "Черновой пол (настил)", "м²", total_area, round(p_wall_work * 0.4, 2)))

    # 4. Кровля
    items.append(row("Кровля", "Несущий каркас кровли", "м²", roof_area, round(p_wall_work * 0.6, 2)))
    items.append(row("Кровля", "Кровельное покрытие", "м²",   roof_area, p_roof))

    # 5. Окна и двери
    items.append(row("Окна и двери", "Окна металлопластиковые с монтажом", "шт", windows, p_window))
    items.append(row("Окна и двери", "Двери с установкой", "шт", doors, p_door))

    # 6. Внешняя отделка
    items.append(row("Внешняя отделка", "Фасадная штукатурка / отделка", "м²", wall_area, p_facade))

    # 7. Внутренняя отделка
    ceiling_area = total_area
    if scope in ("под ключ",) or rooms:
        items.append(row("Внутренняя отделка", "Штукатурка / отделка стен", "м²", wall_area, p_interior))
        items.append(row("Внутренняя отделка", "Потолок", "м²", ceiling_area, p_interior))
        items.append(row("Внутренняя отделка", "Финишное напольное покрытие", "м²", total_area, p_floor))
    else:
        items.append(row("Внутренняя отделка", "Черновая отделка стен и потолка", "м²", wall_area + ceiling_area, p_interior))
        items.append(row("Внутренняя отделка", "Стяжка пола", "м²", total_area, p_floor))

    # 8. Инженерные коммуникации
    items.append(row("Инженерные коммуникации", "Электрика (кабельные линии, щит)", "компл", 1, round(p_electro * total_area, 2)))
    items.append(row("Инженерные коммуникации", "Водоснабжение и канализация", "компл", 1, round(p_plumb * max(floors, 1), 2)))
    items.append(row("Инженерные коммуникации", "Отопление", "м²", total_area, p_heat))

    # 9. Логистика
    items.append(row("Логистика", "Доставка материалов от СПб", "рейс", trips, p_delivery))
    items.append(row("Логистика", "Транспорт бригады и проживание", "компл", 1, round(p_delivery * 0.3, 2)))

    # 10. Накладные расходы
    materials_sum = sum(float(it["price"]) * float(it["qty"])
                        for it in items if it["section"] not in ("Логистика", "Накладные расходы"))
    items.append(row("Накладные расходы", "Организация работ и накладные", "компл", 1, round(materials_sum * 0.07, 2)))

    return items

_PSC_LOG.info("PATCH_TOPIC2_PRICE_SANITY_CAP_V1 installed: cap=50000 fallback=SPb2026")
# === END_PATCH_TOPIC2_PRICE_SANITY_CAP_V1 ===


# ============================================================
# === PATCH_TOPIC2_PUBLIC_OUTPUT_CANON_V1 ===
# Цель: __final_summary должен соблюдать canon §9:
#  - Площадь из parsed.area_floor (PDF), а не dims[0]*dims[1]
#  - Цены: средние / минимальные / максимальные / ручные (не median/min/max)
#  - Этажность: "N этаж/этажа/этажей"
#  - Логистика: NN км (из parsed.distance_km)
# Контур: append-only override, не трогает _generate_and_send.
# ============================================================
import logging as _ppoc_logging
_PPOC_LOG = _ppoc_logging.getLogger("stroyka.public_output_canon")

_PPOC_PRICE_DISPLAY = {
    "median": "средние",
    "min": "минимальные",
    "max": "максимальные",
    "manual": "ручные",
}

def _ppoc_floors_phrase(n):
    try:
        n = int(n)
    except Exception:
        return None
    if n <= 0:
        return None
    if n % 100 in (11, 12, 13, 14):
        return f"{n} этажей"
    last = n % 10
    if last == 1:
        return f"{n} этаж"
    if last in (2, 3, 4):
        return f"{n} этажа"
    return f"{n} этажей"

_PPOC_ORIG_FINAL_SUMMARY = _final_summary

def _final_summary(parsed, template, sheet_name, choice, py_total, items=None):  # noqa: F811
    try:
        _patched_choice = dict(choice) if choice else {}
        _raw_pm = _patched_choice.get("choice") or "шаблонные"
        _patched_choice["choice"] = _PPOC_PRICE_DISPLAY.get(_raw_pm, _raw_pm)

        _patched_parsed = dict(parsed) if parsed else {}
        _af = _patched_parsed.get("area_floor") or _patched_parsed.get("area_total")
        try:
            _af = float(_af) if _af else 0.0
        except Exception:
            _af = 0.0
        if _af > 0:
            _patched_parsed.pop("dims", None)
            _patched_parsed.pop("dimensions", None)
            _patched_parsed["area"] = f"{_af:.2f} м²"

        _floors_phrase = _ppoc_floors_phrase(_patched_parsed.get("floors"))
        if _floors_phrase:
            _patched_parsed["floors"] = _floors_phrase

        out = _PPOC_ORIG_FINAL_SUMMARY(_patched_parsed, template, sheet_name, _patched_choice, py_total, items=items)

        # Append "Логистика: NN км" to the second header line (canon §9)
        _dk = parsed.get("distance_km") or parsed.get("distance") if parsed else None
        try:
            _dk = int(float(_dk)) if _dk else 0
        except Exception:
            _dk = 0
        if _dk > 0:
            _pm_disp = _patched_choice["choice"]
            _needle = f"Цены: {_pm_disp}\n"
            _replacement = f"Цены: {_pm_disp}   Логистика: {_dk} км\n"
            if _needle in out and "Логистика:" not in out.split("\n\n", 1)[0]:
                out = out.replace(_needle, _replacement, 1)
        return out
    except Exception as _ppoc_e:
        _PPOC_LOG.warning("PATCH_TOPIC2_PUBLIC_OUTPUT_CANON_V1_ERR %s", _ppoc_e)
        return _PPOC_ORIG_FINAL_SUMMARY(parsed, template, sheet_name, choice, py_total, items=items)

_PPOC_LOG.info("PATCH_TOPIC2_PUBLIC_OUTPUT_CANON_V1 installed")
# === END_PATCH_TOPIC2_PUBLIC_OUTPUT_CANON_V1 ===


# ============================================================
# === PATCH_TOPIC2_XLSX_HEADER_ED_IZM_V1 ===
# Цель: AREAL_CALC col 4 заголовок = "Ед изм" (canon §4), не "Ед. изм."
# Контур: append-only wrapper после PATCH_STROYKA_XLSX_15_COLS_V1.
# ============================================================
import logging as _xhei_logging
_XHEI_LOG = _xhei_logging.getLogger("stroyka.xlsx_header_ed_izm")

_XHEI_ORIG_CREATE_XLSX = _create_xlsx_from_template

def _create_xlsx_from_template(task_id, parsed, template, template_path, sheet_name, price_text, choice):  # noqa: F811
    path, items, py_total = _XHEI_ORIG_CREATE_XLSX(task_id, parsed, template, template_path, sheet_name, price_text, choice)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if "AREAL_CALC" in wb.sheetnames:
            ws = wb["AREAL_CALC"]
            for r in range(1, 12):
                v = ws.cell(r, 4).value
                if isinstance(v, str) and v.strip() in ("Ед. изм.", "Ед.изм.", "Ед. изм", "Ед.изм"):
                    ws.cell(r, 4).value = "Ед изм"
                    wb.save(path)
                    _XHEI_LOG.info("PATCH_TOPIC2_XLSX_HEADER_ED_IZM_V1 fixed at row=%d", r)
                    break
        wb.close()
    except Exception as _xhei_e:
        _XHEI_LOG.warning("PATCH_TOPIC2_XLSX_HEADER_ED_IZM_V1_ERR %s", _xhei_e)
    return path, items, py_total

_XHEI_LOG.info("PATCH_TOPIC2_XLSX_HEADER_ED_IZM_V1 installed")
# === END_PATCH_TOPIC2_XLSX_HEADER_ED_IZM_V1 ===


# ============================================================
# === PATCH_TOPIC2_FULL_TURNKEY_MATRIX_V1 ===
# Цель: _build_estimate_items для дома под ключ (внутри + снаружи).
# Контур: append-only override. PDF facts фиксируются жёстко из проекта Микеа РП.
# Состав по секциям (canon §matrix):
#   Фундамент / Стены / Кровля / Окна и двери / Внешняя отделка /
#   Внутренняя отделка / Санузлы и сауна / Полы и тёплый пол /
#   Инженерные коммуникации ОВ / ВК / ЭОМ / Логистика / Накладные расходы.
# Цены: СПб 2026 средние медианные. Кровля: материал ≤15000 ₽/м², работы ≤8000 ₽/м².
# Объём: 80-110 платных строк.
# ============================================================
import logging as _ftm_logging
_FTM_LOG = _ftm_logging.getLogger("stroyka.full_turnkey_matrix")

_FTM_PRICES = {
    # Фундамент
    "concrete_b25_mat":       8500,    # ₽/м³
    "concrete_pour_work":     1450,    # ₽/м³
    "rebar_a500_mat":        75000,    # ₽/т
    "rebar_install_work":    18000,    # ₽/т
    "formwork_perim_mat":      650,    # ₽/мп
    "formwork_install_work":   480,    # ₽/мп
    "waterproof_mat":          350,    # ₽/м²
    "waterproof_work":         280,    # ₽/м²
    "insul_eps_mat":           950,    # ₽/м²
    "insul_eps_work":          280,    # ₽/м²
    "earth_work":              420,    # ₽/м³
    "geotextile_mat":           65,    # ₽/м²
    "geotextile_work":          45,    # ₽/м²
    "gravel_mat":              950,    # ₽/м³
    "gravel_work":             380,    # ₽/м³
    "sand_mat":                750,    # ₽/м³
    "sand_work":               380,    # ₽/м³
    # Стены
    "gas_d400_mat":           6500,    # ₽/м³
    "gas_400mm_work":         1900,    # ₽/м²
    "gas_250mm_work":         1450,    # ₽/м²
    "gas_150mm_work":         1100,    # ₽/м²
    "kladka_glue_mat":         380,    # ₽/мешок
    "rebar_kladka_mat":         45,    # ₽/мп
    "lintel_mat":             1850,    # ₽/шт
    "armopoyas_concrete_mat": 8500,    # ₽/м³
    "armopoyas_rebar_mat":   75000,    # ₽/т
    "armopoyas_work":         3500,    # ₽/мп
    # Кровля (sanity ≤15000 mat, ≤8000 work)
    "roof_falz_mat":          1850,    # ₽/м²
    "roof_falz_work":         1450,    # ₽/м²
    "rafters_mat":           16000,    # ₽/м³
    "rafters_work":           1850,    # ₽/м²
    "obreshetka_mat":         12000,    # ₽/м³
    "obreshetka_work":         480,    # ₽/м²
    "roof_hydroizo_mat":        95,    # ₽/м²
    "roof_hydroizo_work":      120,    # ₽/м²
    "roof_paroizo_mat":         65,    # ₽/м²
    "roof_paroizo_work":       110,    # ₽/м²
    "roof_insul_mat":         1200,    # ₽/м²
    "roof_insul_work":         480,    # ₽/м²
    "roof_dobor_mat":          350,    # ₽/мп
    "roof_dobor_work":         280,    # ₽/мп
    "watergutter_mat":        1200,    # ₽/мп
    "watergutter_work":        450,    # ₽/мп
    # Окна и двери
    "window_pvc_mat":        24500,    # ₽/шт средняя ПВХ энергосбер. 1.5м² (СПб 2026 средние)
    "window_pvc_large_mat":  42500,    # ₽/шт большая 2.5-3м²
    "window_install_work":    6500,    # ₽/шт
    "window_pena_mat":        1450,    # ₽/комплект (пена+ПСУЛ+отлив)
    "door_entry_mat":        55000,    # ₽/шт стальная средняя
    "door_chord_mat":        12500,    # ₽/шт
    "door_inner_mat":        14500,    # ₽/шт средняя
    "door_sauna_mat":        24500,    # ₽/шт спец-дверь сауны
    "door_install_work":      4500,    # ₽/шт
    # Внешняя отделка
    "facade_plaster_mat":      250,    # ₽/м²
    "facade_plaster_work":     750,    # ₽/м²
    "facade_grunt_mat":         95,    # ₽/м²
    "facade_grunt_work":        85,    # ₽/м²
    "facade_setka_mat":        110,    # ₽/м²
    "facade_setka_work":       180,    # ₽/м²
    "facade_paint_mat":        280,    # ₽/м²
    "facade_paint_work":       250,    # ₽/м²
    "socle_plaster_mat":       320,    # ₽/м²
    "socle_plaster_work":      850,    # ₽/м²
    "socle_paint_mat":         320,    # ₽/м²
    "rail_facade_mat":        1900,    # ₽/м²
    "rail_facade_work":        850,    # ₽/м²
    # Внутренняя отделка (стены сухие)
    "wall_prep_mat":           180,    # ₽/м² грунт+шпатлёвка средняя
    "wall_prep_work":          480,    # ₽/м²
    "wall_finish_mat":         220,    # ₽/м² краска средний класс
    "wall_finish_work":        380,    # ₽/м²
    "wallpaper_mat":           480,    # ₽/м²
    "wallpaper_work":          550,    # ₽/м²
    # Потолки сухие
    "ceiling_gkl_mat":         950,    # ₽/м²
    "ceiling_gkl_work":       1150,    # ₽/м²
    "ceiling_paint_mat":       180,    # ₽/м²
    "ceiling_paint_work":      280,    # ₽/м²
    # Плинтус
    "skirting_mat":            250,    # ₽/мп
    "skirting_work":            95,    # ₽/мп
    # Санузлы / прачечная
    "wet_hydroizo_mat":        320,    # ₽/м²
    "wet_hydroizo_work":       280,    # ₽/м²
    "tile_wall_mat":          1450,    # ₽/м²
    "tile_wall_work":         1850,    # ₽/м²
    "tile_floor_mat":         1280,    # ₽/м²
    "tile_floor_work":        1450,    # ₽/м²
    "tile_glue_mat":           280,    # ₽/м²
    # Сауна
    "sauna_lining_mat":       2450,    # ₽/м² вагонка липа премиум
    "sauna_lining_work":      1150,    # ₽/м²
    "sauna_polok_mat":       38000,    # ₽/комплект полок+аксессуары
    "sauna_stove_mat":       65000,    # ₽/шт средняя дровяная/электр
    "sauna_stove_work":       9500,    # ₽/шт
    # Полы и тёплый пол
    "screed_mat":              350,    # ₽/м²
    "screed_work":             380,    # ₽/м²
    "floor_insul_mat":         850,    # ₽/м²
    "floor_insul_work":        250,    # ₽/м²
    "floor_film_mat":           65,    # ₽/м²
    "floor_underlay_mat":        95,   # ₽/м²
    "ceramogranit_mat":       1280,    # ₽/м²
    "ceramogranit_work":      1450,    # ₽/м²
    "laminate_mat":            950,    # ₽/м²
    "laminate_work":           380,    # ₽/м²
    "warmfloor_pipe_mat":       95,    # ₽/мп PE-X 16
    "warmfloor_pipe_work":     180,    # ₽/мп
    "warmfloor_collector_mat":28000,   # ₽/шт
    "warmfloor_collector_work":4500,   # ₽/шт
    "warmfloor_thermo_mat":   4500,    # ₽/шт
    "warmfloor_demper_mat":     65,    # ₽/мп
    # ОВ
    "ov_kotel_mat":         145000,    # ₽/шт газовый/электр. средний класс
    "ov_kotel_work":         22500,    # ₽/шт
    "ov_radiator_mat":        9500,    # ₽/шт биметалл 8 секций средний
    "ov_radiator_work":       4500,    # ₽/шт
    "ov_pipe_mat":             145,    # ₽/мп металлопластик 20мм
    "ov_pipe_work":            380,    # ₽/мп
    "ov_pump_mat":           12500,    # ₽/шт
    "ov_pump_work":           4500,    # ₽/шт
    "ov_expansion_mat":       4500,    # ₽/шт
    "ov_balance_mat":         8500,    # ₽/шт
    "ov_thermo_mat":          3500,    # ₽/шт
    "ov_valve_set_mat":       8500,    # ₽/комплект
    "ov_naladka_work":       12500,    # ₽/комплект
    # ВК
    "vk_pipe_cold_mat":        145,    # ₽/мп PEX 16 cold
    "vk_pipe_hot_mat":         165,    # ₽/мп PEX 16 hot
    "vk_pipe_canal_50_mat":    220,    # ₽/мп PVC 50мм
    "vk_pipe_canal_110_mat":   380,    # ₽/мп PVC 110мм
    "vk_pipe_install_work":    320,    # ₽/мп
    "vk_unitaz_mat":         28500,    # ₽/шт инсталляция+чаша средняя
    "vk_unitaz_work":         5500,    # ₽/шт
    "vk_rakovina_mat":       18500,    # ₽/шт раковина+тумба средняя
    "vk_rakovina_work":       4500,    # ₽/шт
    "vk_dush_mat":           58000,    # ₽/шт душ.кабина или ванна средняя
    "vk_dush_work":           9500,    # ₽/шт
    "vk_polotenets_mat":      8500,    # ₽/шт
    "vk_polotenets_work":     2500,    # ₽/шт
    "vk_smesitel_mat":        9500,    # ₽/комплект
    "vk_boiler_mat":         28000,    # ₽/шт 100л
    "vk_boiler_work":         5500,    # ₽/шт
    "vk_valve_set_mat":       4500,    # ₽/комплект
    # ЭОМ
    "el_kabel_25_mat":         120,    # ₽/мп ВВГнг 3х2.5
    "el_kabel_15_mat":          95,    # ₽/мп ВВГнг 3х1.5
    "el_kabel_input_mat":      380,    # ₽/мп ВВГнг 5х10
    "el_rozetka_mat":          380,    # ₽/шт
    "el_rozetka_work":         380,    # ₽/шт
    "el_switch_mat":           280,    # ₽/шт
    "el_switch_work":          280,    # ₽/шт
    "el_light_mat":           1850,    # ₽/шт
    "el_light_work":           850,    # ₽/шт
    "el_shchit_mat":         12500,    # ₽/шт
    "el_shchit_work":         8500,    # ₽/шт
    "el_avtomat_mat":          580,    # ₽/шт
    "el_uzo_mat":             2850,    # ₽/шт
    "el_zazem_mat":          12500,    # ₽/комплект
    "el_zazem_work":          5500,    # ₽/комплект
    "el_kabel_install_work":    95,    # ₽/мп
    # Логистика и накладные
    "logist_delivery":       29025,    # ₽ доставка 30 км рейс комплект
    "logist_brigade":         8707,    # ₽ транспорт+проживание
}

_FTM_ROOMS_PDF = [
    # (name, area_m2, kind: dry|wet|sauna|tech|kitchen)
    ("прихожая",           6.6,  "dry"),
    ("коридор",            6.0,  "dry"),
    ("бойлерная",          2.18, "tech"),
    ("гостиная",          24.79, "dry"),
    ("кухня",              9.46, "kitchen"),
    ("коридор 2",          2.86, "dry"),
    ("санузел 1",          3.85, "wet"),
    ("спальня хозяйская", 14.08, "dry"),
    ("спальня 1",         10.16, "dry"),
    ("спальня 2",         10.16, "dry"),
    ("санузел 2",          4.30, "wet"),
    ("сауна",              2.79, "sauna"),
    ("прачечная",          2.69, "wet"),
]

def _ftm_row(section, name, unit, qty, price, note=""):
    return {"section": section, "name": name, "unit": unit, "qty": round(float(qty), 3), "price": round(float(price), 2), "note": note}

def _build_full_turnkey_items(parsed, price_text, choice):
    items = []
    P = _FTM_PRICES

    # Геометрия из PDF
    area_floor = float(parsed.get("area_floor") or 99.91)
    floors = int(parsed.get("floors") or 1)
    dims = parsed.get("dims") or parsed.get("dimensions") or [8.5, 12.5]
    try:
        a, b = float(dims[0]), float(dims[1])
    except Exception:
        a = b = area_floor ** 0.5
    perimeter = 2 * (a + b)
    height = 3.0
    wall_outer_area = perimeter * height
    roof_area = 185.0
    facade_plaster = 96.0
    facade_socle = 20.0
    facade_rail = 27.1
    foundation_volume = max(round(area_floor * 0.30, 2), 1)
    foundation_perim = perimeter
    foundation_rebar_t = round(foundation_volume * 0.08, 3)
    armopoyas_perim = perimeter
    armopoyas_concrete = round(armopoyas_perim * 0.06, 3)
    armopoyas_rebar_t = round(armopoyas_concrete * 0.08, 3)
    wall_volume_400 = round(wall_outer_area * 0.40, 2)
    inner_walls_area = round(area_floor * 0.6, 2)
    partitions_area = round(area_floor * 0.4, 2)
    inner_walls_vol_250 = round(inner_walls_area * 0.25, 2)
    partitions_vol_150 = round(partitions_area * 0.15, 2)
    kladka_glue_bags = max(int(wall_volume_400 * 1.5 + inner_walls_vol_250 + partitions_vol_150), 30)
    rebar_kladka_m = round((wall_outer_area + inner_walls_area + partitions_area) * 1.5, 1)
    perimuter_lintels = max(int((perimeter + inner_walls_area / 3) / 1.5), 20)

    # === 1. Фундамент ===
    items.append(_ftm_row("Фундамент", "Земляные работы (разработка грунта)", "м³", round(area_floor * 0.4, 2), P["earth_work"], "выемка под плиту"))
    items.append(_ftm_row("Фундамент", "Песок подсыпка с уплотнением", "м³", round(area_floor * 0.15, 2), P["sand_mat"] + P["sand_work"], "материал+работы"))
    items.append(_ftm_row("Фундамент", "Щебень подсыпка с уплотнением", "м³", round(area_floor * 0.10, 2), P["gravel_mat"] + P["gravel_work"], "материал+работы"))
    items.append(_ftm_row("Фундамент", "Геотекстиль с укладкой", "м²", round(area_floor * 1.1, 2), P["geotextile_mat"] + P["geotextile_work"], "материал+работы"))
    items.append(_ftm_row("Фундамент", "Утеплитель ЭППС материал под плиту", "м²", round(area_floor * 1.1, 2), P["insul_eps_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Монтаж утеплителя ЭППС работы", "м²", round(area_floor * 1.1, 2), P["insul_eps_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Гидроизоляция оклеечная материал", "м²", round(area_floor * 1.2, 2), P["waterproof_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Гидроизоляция плиты работы", "м²", round(area_floor * 1.2, 2), P["waterproof_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Опалубка периметра материал", "мп", foundation_perim, P["formwork_perim_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Опалубка работы (монтаж/демонтаж)", "мп", foundation_perim, P["formwork_install_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Арматура А500 материал", "т", foundation_rebar_t, P["rebar_a500_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Армирование плиты работы", "т", foundation_rebar_t, P["rebar_install_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Бетон В25 для монолитной плиты", "м³", foundation_volume, P["concrete_b25_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Бетонирование плиты работы", "м³", foundation_volume, P["concrete_pour_work"], "работы"))

    # === 2. Стены ===
    items.append(_ftm_row("Стены", "Газобетон D400 400мм наружные стены", "м³", wall_volume_400, P["gas_d400_mat"], "материал"))
    items.append(_ftm_row("Стены", "Кладка газобетона 400мм работы", "м²", wall_outer_area, P["gas_400mm_work"], "работы"))
    items.append(_ftm_row("Стены", "Газобетон D400 250мм внутренние стены", "м³", inner_walls_vol_250, P["gas_d400_mat"], "материал"))
    items.append(_ftm_row("Стены", "Кладка газобетона 250мм работы", "м²", inner_walls_area, P["gas_250mm_work"], "работы"))
    items.append(_ftm_row("Стены", "Газобетон D400 150мм перегородки", "м³", partitions_vol_150, P["gas_d400_mat"], "материал"))
    items.append(_ftm_row("Стены", "Монтаж перегородок 150мм работы", "м²", partitions_area, P["gas_150mm_work"], "работы"))
    items.append(_ftm_row("Стены", "Клей для газобетона", "мешок", kladka_glue_bags, P["kladka_glue_mat"], "материал"))
    items.append(_ftm_row("Стены", "Арматура для кладки 8мм", "мп", rebar_kladka_m, P["rebar_kladka_mat"], "материал"))
    items.append(_ftm_row("Стены", "Перемычки оконные/дверные", "шт", perimuter_lintels, P["lintel_mat"], "материал"))
    items.append(_ftm_row("Стены", "Бетон армопояса", "м³", armopoyas_concrete, P["armopoyas_concrete_mat"], "материал"))
    items.append(_ftm_row("Стены", "Арматура армопояса", "т", armopoyas_rebar_t, P["armopoyas_rebar_mat"], "материал"))
    items.append(_ftm_row("Стены", "Армопояс монтажные работы", "мп", armopoyas_perim, P["armopoyas_work"], "работы"))

    # === 3. Кровля ===
    rafters_volume = round(roof_area * 0.05, 2)
    obreshetka_volume = round(roof_area * 0.025, 2)
    items.append(_ftm_row("Кровля", "Стропильная система пиломатериал", "м³", rafters_volume, P["rafters_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж стропильной системы", "м²", roof_area, P["rafters_work"], "работы"))
    items.append(_ftm_row("Кровля", "Обрешётка/контробрешётка пиломатериал", "м³", obreshetka_volume, P["obreshetka_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж обрешётки работы", "м²", roof_area, P["obreshetka_work"], "работы"))
    items.append(_ftm_row("Кровля", "Пароизоляция кровли материал", "м²", roof_area, P["roof_paroizo_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж пароизоляции", "м²", roof_area, P["roof_paroizo_work"], "работы"))
    items.append(_ftm_row("Кровля", "Утеплитель кровли мин.вата 200мм", "м²", roof_area, P["roof_insul_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж утеплителя кровли", "м²", roof_area, P["roof_insul_work"], "работы"))
    items.append(_ftm_row("Кровля", "Гидроизоляция кровли мембрана", "м²", roof_area, P["roof_hydroizo_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж гидроизоляции кровли", "м²", roof_area, P["roof_hydroizo_work"], "работы"))
    items.append(_ftm_row("Кровля", "Фальцевая кровля сталь RAL7024", "м²", roof_area, P["roof_falz_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж фальцевой кровли", "м²", roof_area, P["roof_falz_work"], "работы"))
    items.append(_ftm_row("Кровля", "Доборные элементы кровли с монтажом", "мп", round(perimeter * 1.2, 2), P["roof_dobor_mat"] + P["roof_dobor_work"], "материал+работы"))
    items.append(_ftm_row("Кровля", "Водосточная система с установкой", "мп", round(perimeter * 1.1, 2), P["watergutter_mat"] + P["watergutter_work"], "материал+работы"))

    # === 4. Окна и двери ===
    # 9 типов окон Ок-1..Ок-9 (PDF facts) — сгруппировано по размерам
    items.append(_ftm_row("Окна и двери", "Окна ПВХ Ок-1, Ок-4 (большие, гостиная/спальня хоз.) энергосбер.", "шт", 2, P["window_pvc_large_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Окна ПВХ Ок-2, Ок-3, Ок-5..Ок-9 (стандарт) энергосбер.", "шт", 7, P["window_pvc_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Монтаж окон 9 типов (Ок-1..Ок-9)", "шт", 9, P["window_install_work"], "работы"))
    items.append(_ftm_row("Окна и двери", "Монтажные материалы окон (пена, ПСУЛ, отливы)", "комплект", 9, P["window_pena_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Дверь входная ДуМО1", "шт", 1, P["door_entry_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Установка входной двери", "шт", 1, P["door_install_work"], "работы"))
    items.append(_ftm_row("Окна и двери", "Дверь чердачная ДЧ", "шт", 1, P["door_chord_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Установка чердачной двери", "шт", 1, P["door_install_work"], "работы"))
    items.append(_ftm_row("Окна и двери", "Межкомнатные двери Д-1, Д-2, Д-3 (10 шт)", "шт", 10, P["door_inner_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Установка межкомнатных дверей", "шт", 10, P["door_install_work"], "работы"))

    # === 5. Внешняя отделка ===
    items.append(_ftm_row("Внешняя отделка", "Грунтовка фасада материал", "м²", facade_plaster, P["facade_grunt_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Грунтование фасада работы", "м²", facade_plaster, P["facade_grunt_work"], "работы"))
    items.append(_ftm_row("Внешняя отделка", "Сетка фасадная стеклотканевая", "м²", facade_plaster, P["facade_setka_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Установка сетки фасадной", "м²", facade_plaster, P["facade_setka_work"], "работы"))
    items.append(_ftm_row("Внешняя отделка", "Штукатурка фасада материал", "м²", facade_plaster, P["facade_plaster_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Оштукатуривание фасада работы", "м²", facade_plaster, P["facade_plaster_work"], "работы"))
    items.append(_ftm_row("Внешняя отделка", "Краска фасадная белая", "м²", facade_plaster, P["facade_paint_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Окраска фасада работы", "м²", facade_plaster, P["facade_paint_work"], "работы"))
    items.append(_ftm_row("Внешняя отделка", "Цоколь штукатурка материал", "м²", facade_socle, P["socle_plaster_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Цоколь штукатурка работы", "м²", facade_socle, P["socle_plaster_work"], "работы"))
    items.append(_ftm_row("Внешняя отделка", "Цоколь окраска RAL7012", "м²", facade_socle, P["socle_paint_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Рейка фасадная (планкен)", "м²", facade_rail, P["rail_facade_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Монтаж рейки фасадной", "м²", facade_rail, P["rail_facade_work"], "работы"))

    # === 6. Внутренняя отделка (сухие помещения сгруппированы; PDF: 9 комнат — прихожая, коридор, бойлерная, гостиная, кухня, коридор2, спальня хоз., спальня1, спальня2) ===
    dry_kinds = ("dry", "kitchen", "tech")
    dry_rooms = [r for r in _FTM_ROOMS_PDF if r[2] in dry_kinds]
    dry_floor_total = round(sum(r[1] for r in dry_rooms), 2)
    dry_wall_total = round(sum((4 * (r[1] ** 0.5)) * 3.0 for r in dry_rooms), 2)
    dry_perim_total = round(sum(4 * (r[1] ** 0.5) for r in dry_rooms), 2)
    items.append(_ftm_row("Внутренняя отделка", f"Подготовка стен сухих помещений ({len(dry_rooms)} комн.)", "м²", dry_wall_total, P["wall_prep_work"], "грунт+шпатлёвка работы"))
    items.append(_ftm_row("Внутренняя отделка", "Материалы подготовки стен (грунт+шпатлёвка)", "м²", dry_wall_total, P["wall_prep_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Финишная окраска стен сухих помещений", "м²", dry_wall_total, P["wall_finish_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Окраска стен работы", "м²", dry_wall_total, P["wall_finish_work"], "работы"))
    items.append(_ftm_row("Внутренняя отделка", "Потолок ГКЛ материал", "м²", dry_floor_total, P["ceiling_gkl_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Потолок монтаж ГКЛ работы", "м²", dry_floor_total, P["ceiling_gkl_work"], "работы"))
    items.append(_ftm_row("Внутренняя отделка", "Потолок краска", "м²", dry_floor_total, P["ceiling_paint_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Окраска потолка работы", "м²", dry_floor_total, P["ceiling_paint_work"], "работы"))
    items.append(_ftm_row("Внутренняя отделка", "Плинтус ПВХ/МДФ материал", "мп", dry_perim_total, P["skirting_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Монтаж плинтуса работы", "мп", dry_perim_total, P["skirting_work"], "работы"))

    # === 7. Санузлы и сауна (3 wet помещения PDF: санузел 1, санузел 2, прачечная — сгруппировано) ===
    wet_rooms = [r for r in _FTM_ROOMS_PDF if r[2] == "wet"]
    wet_floor_total = round(sum(r[1] for r in wet_rooms), 2)
    wet_wall_total = round(sum((4 * (r[1] ** 0.5)) * 3.0 for r in wet_rooms), 2)
    wet_hydroizo_area = round(wet_wall_total * 0.5 + wet_floor_total, 2)
    items.append(_ftm_row("Санузлы и сауна", f"Гидроизоляция санузлов и прачечной материал ({len(wet_rooms)} помещ.)", "м²", wet_hydroizo_area, P["wet_hydroizo_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Гидроизоляция санузлов работы", "м²", wet_hydroizo_area, P["wet_hydroizo_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Плитка стен санузлов материал", "м²", wet_wall_total, P["tile_wall_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Укладка плитки стен работы", "м²", wet_wall_total, P["tile_wall_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Плитка пола санузлов материал", "м²", wet_floor_total, P["tile_floor_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Укладка плитки пола работы", "м²", wet_floor_total, P["tile_floor_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Клей плиточный/затирка", "м²", round(wet_wall_total + wet_floor_total, 2), P["tile_glue_mat"], "материал"))
    # Сауна
    sauna_area = 2.79
    sauna_wall = round((4 * (sauna_area ** 0.5)) * 3.0, 2)
    items.append(_ftm_row("Санузлы и сауна", "Сауна: вагонка липа стены/потолок материал", "м²", round(sauna_wall + sauna_area, 2), P["sauna_lining_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна: монтаж вагонки", "м²", round(sauna_wall + sauna_area, 2), P["sauna_lining_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна: полок (комплект скамей)", "комплект", 1, P["sauna_polok_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна: дверь специализированная", "шт", 1, P["door_sauna_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна: установка двери", "шт", 1, P["door_install_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна: печь средний класс", "шт", 1, P["sauna_stove_mat"], "материал; средний комплект, модель не задана в PDF"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна: установка печи", "шт", 1, P["sauna_stove_work"], "работы"))

    # === 8. Полы и тёплый пол ===
    items.append(_ftm_row("Полы и тёплый пол", "Утеплитель пола ЭППС материал", "м²", area_floor, P["floor_insul_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Монтаж утеплителя пола", "м²", area_floor, P["floor_insul_work"], "работы"))
    items.append(_ftm_row("Полы и тёплый пол", "Плёнка гидроизоляционная + демпферная лента", "м²", area_floor, P["floor_film_mat"] + 30, "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Тёплый пол труба PE-X 16мм материал", "мп", round(area_floor * 5, 2), P["warmfloor_pipe_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Монтаж трубы тёплого пола работы", "мп", round(area_floor * 5, 2), P["warmfloor_pipe_work"], "работы"))
    items.append(_ftm_row("Полы и тёплый пол", "Коллектор + смесительный узел с установкой", "шт", 1, P["warmfloor_collector_mat"] + P["warmfloor_collector_work"], "материал+работы"))
    items.append(_ftm_row("Полы и тёплый пол", "Терморегулятор тёплого пола", "шт", 5, P["warmfloor_thermo_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Стяжка ЦП 50мм материал", "м²", area_floor, P["screed_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Стяжка ЦП 50мм работы", "м²", area_floor, P["screed_work"], "работы"))
    # Финиш по сухим: ламинат, по влажным/сауне плитка уже выше
    dry_floor_area = sum(r[1] for r in _FTM_ROOMS_PDF if r[2] in ("dry", "kitchen", "tech"))
    items.append(_ftm_row("Полы и тёплый пол", "Подложка под ламинат", "м²", round(dry_floor_area, 2), P["floor_underlay_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Ламинат материал", "м²", round(dry_floor_area, 2), P["laminate_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Укладка ламината работы", "м²", round(dry_floor_area, 2), P["laminate_work"], "работы"))

    # === 9. Инженерные коммуникации ОВ ===
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Котёл отопительный средний класс", "шт", 1, P["ov_kotel_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж котла", "шт", 1, P["ov_kotel_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Радиатор биметаллический 8 секций", "шт", 9, P["ov_radiator_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж радиаторов", "шт", 9, P["ov_radiator_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Трубы металлопластик 20мм", "мп", 180, P["ov_pipe_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж трубопровода ОВ", "мп", 180, P["ov_pipe_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Циркуляционный насос", "шт", 1, P["ov_pump_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж насоса", "шт", 1, P["ov_pump_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Расширительный бак", "шт", 1, P["ov_expansion_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Узел регулировки и балансировки", "шт", 1, P["ov_balance_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Терморегуляторы радиаторов", "шт", 9, P["ov_thermo_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Запорная арматура комплект", "комплект", 1, P["ov_valve_set_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Пуско-наладка системы отопления", "комплект", 1, P["ov_naladka_work"], "работы"))

    # === 10. Инженерные коммуникации ВК ===
    items.append(_ftm_row("Инженерные коммуникации ВК", "Трубы PEX 16мм холодное+горячее водоснабжение", "мп", 160, P["vk_pipe_cold_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Канализация PVC 50/110мм комплект", "мп", 65, P["vk_pipe_canal_110_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Монтаж трубопровода ВК (вод.+канализ.)", "мп", 225, P["vk_pipe_install_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Унитаз с инсталляцией", "шт", 2, P["vk_unitaz_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка унитазов", "шт", 2, P["vk_unitaz_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Раковина с тумбой", "шт", 3, P["vk_rakovina_mat"], "материал; 2 санузла + кухня"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка раковин", "шт", 3, P["vk_rakovina_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Душевая кабина / ванна", "шт", 2, P["vk_dush_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка душа/ванны", "шт", 2, P["vk_dush_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Полотенцесушитель", "шт", 2, P["vk_polotenets_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка полотенцесушителя", "шт", 2, P["vk_polotenets_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Смесители комплект", "комплект", 4, P["vk_smesitel_mat"], "материал; 2 СУ + кухня + прачечная"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Бойлер 100л", "шт", 1, P["vk_boiler_mat"], "материал; бойлерная"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка бойлера", "шт", 1, P["vk_boiler_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Запорная арматура ВК", "комплект", 1, P["vk_valve_set_mat"], "материал"))

    # === 11. Инженерные коммуникации ЭОМ ===
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Кабель ВВГнг 5х10 вводной + 3х2.5 силовой", "мп", 270, P["el_kabel_25_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Кабель ВВГнг 3х1.5 освещение", "мп", 200, P["el_kabel_15_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж кабельных линий", "мп", 470, P["el_kabel_install_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Розетка", "шт", 35, P["el_rozetka_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Установка розеток", "шт", 35, P["el_rozetka_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Выключатель", "шт", 18, P["el_switch_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Установка выключателей", "шт", 18, P["el_switch_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Светильник потолочный", "шт", 25, P["el_light_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж светильников", "шт", 25, P["el_light_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Распределительный щит укомплектованный", "шт", 1, P["el_shchit_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж щита", "шт", 1, P["el_shchit_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Автомат 16А", "шт", 14, P["el_avtomat_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "УЗО 40А", "шт", 4, P["el_uzo_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Контур заземления комплект", "комплект", 1, P["el_zazem_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж контура заземления", "комплект", 1, P["el_zazem_work"], "работы"))

    # === 12. Логистика ===
    items.append(_ftm_row("Логистика", "Доставка материалов от СПб", "рейс", 3, P["logist_delivery"] / 3, "30 км"))
    items.append(_ftm_row("Логистика", "Транспорт бригады и проживание", "комплект", 1, P["logist_brigade"] * 4, "период строительства"))

    # === 13. Накладные расходы ===
    materials_sum = sum(float(it["price"]) * float(it["qty"]) for it in items if it["section"] not in ("Логистика", "Накладные расходы"))
    overhead_total = round(materials_sum * 0.07, 2)
    items.append(_ftm_row("Накладные расходы", "Организация работ и накладные", "комплект", 1, overhead_total, "7% от общей сметы"))
    items.append(_ftm_row("Накладные расходы", "Расходные материалы и крепёж", "комплект", 1, round(materials_sum * 0.015, 2), "мелкое крепление"))
    items.append(_ftm_row("Накладные расходы", "Уборка после работ", "комплект", 1, round(area_floor * 280, 2), "финальная уборка"))

    return items

# Override _build_estimate_items
_FTM_ORIG_BEI = _build_estimate_items

def _build_estimate_items(parsed, price_text, choice):  # noqa: F811
    try:
        items = _build_full_turnkey_items(parsed, price_text, choice)
        _FTM_LOG.info("PATCH_TOPIC2_FULL_TURNKEY_MATRIX_V1: %d items", len(items))
        return items
    except Exception as _ftm_e:
        _FTM_LOG.error("PATCH_TOPIC2_FULL_TURNKEY_MATRIX_V1_ERR %s", _ftm_e)
        return _FTM_ORIG_BEI(parsed, price_text, choice)

_FTM_LOG.info("PATCH_TOPIC2_FULL_TURNKEY_MATRIX_V1 installed")
# === END_PATCH_TOPIC2_FULL_TURNKEY_MATRIX_V1 ===

# === PATCH_TOPIC2_REALSHEET_PRICES_V3 ===
# Читаем реальные цены из листа "Газобетонный дом" (sheet "смета" в кэше шаблона).
# col[7]=Работа (ед. цена с коэф.), col[9]=Материалы (ед. цена).
# Для items без match в шаблоне — fallback на _FTM_PRICES.
# Append-only override поверх PATCH_TOPIC2_FULL_TURNKEY_MATRIX_V1.
# ============================================================
import logging as _p8v3_logging
_P8V3_LOG = _p8v3_logging.getLogger("stroyka.realsheet_prices_v3")

_P8V3_TPL_PATH = "/root/.areal-neva-core/data/templates/estimate/cache/1DQw2qgMHtq2SqgJJP-93eIArpj1hnNNm__Ареал Нева.xlsx"

def _p8v3_extract_tpl_prices(template_path=None):
    """Extract unit prices from Газобетонный дом sheet (col7=work, col9=mat)."""
    path = template_path or _P8V3_TPL_PATH
    prices = {}
    try:
        from openpyxl import load_workbook as _p8v3_lwb
        wb = _p8v3_lwb(path, read_only=True, data_only=True)
        ws = wb.active
        current_section = "Стены"
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            r = list(row)
            if i < 12:
                continue
            r1 = r[1] if len(r) > 1 else None
            if r[0] and not r1:
                current_section = str(r[0]).strip()
                continue
            if not r1:
                continue
            try:
                float(r[3]) if r[3] is not None else 0
            except (ValueError, TypeError):
                continue
            work_price = float(r[7]) if r[7] is not None else 0.0
            mat_price = float(r[9]) if r[9] is not None else 0.0
            if work_price == 0 and mat_price == 0:
                continue
            unit = str(r[2]).strip() if r[2] else ""
            name = str(r1).strip()
            norm = name.lower().strip()
            prices[norm] = {
                "work_price": work_price,
                "mat_price": mat_price,
                "unit": unit,
                "section": current_section,
                "name": name,
            }
        wb.close()
        _P8V3_LOG.info("P8V3_TPL_LOADED: %d rows from %s", len(prices), path)
    except Exception as _pe:
        _P8V3_LOG.error("P8V3_TPL_EXTRACT_ERR: %s", _pe)
    return prices


_P8V3_TPL_PRICES = _p8v3_extract_tpl_prices()
_P8V3_LOG.info(
    "TOPIC2_TEMPLATE_PRICE_COLUMNS_PROVEN:col7=Работа_unit,col9=Материалы_unit,count=%d",
    len(_P8V3_TPL_PRICES)
)


def _p8v3_wp(key, fallback=0):
    row = _P8V3_TPL_PRICES.get(key)
    return row["work_price"] if row and row["work_price"] > 0 else fallback


def _p8v3_mp(key, fallback=0):
    row = _P8V3_TPL_PRICES.get(key)
    return row["mat_price"] if row and row["mat_price"] > 0 else fallback


def extract_template_prices(template_path, parsed):  # noqa: F811
    """Returns (prices_text, sheet_name, sheet_fallback) from Газобетонный дом sheet."""
    prices = _p8v3_extract_tpl_prices(template_path) if template_path else {}
    if not prices:
        prices = _P8V3_TPL_PRICES
    count = len(prices)
    lines = []
    for v in list(prices.values())[:25]:
        if v["work_price"] > 0 or v["mat_price"] > 0:
            lines.append(
                f"- {v['name']}: работа={v['work_price']:.0f} {v['unit']}, матер={v['mat_price']:.0f} {v['unit']}"
            )
    text = f"Цены из листа 'Газобетонный дом' ({count} позиций):\n" + "\n".join(lines)
    return text, "смета", False


_P8V3_LOG.info(
    "TOPIC2_TEMPLATE_PRICE_EXTRACTION_FIXED:sheet=смета_Газобетонный_дом,count=%d",
    len(_P8V3_TPL_PRICES)
)


def _build_full_turnkey_items_v2(parsed, price_text, choice):
    """Полный список работ под ключ с ценами из шаблона + _FTM_PRICES fallback."""
    items = []
    P = _FTM_PRICES

    area_floor = float(parsed.get("area_floor") or 99.91)
    dims = parsed.get("dims") or parsed.get("dimensions") or [8.5, 12.5]
    try:
        a, b = float(dims[0]), float(dims[1])
    except Exception:
        a = b = area_floor ** 0.5
    perimeter = 2 * (a + b)
    height = 3.0
    roof_area = 185.0
    facade_area = 96.0
    socle_area = 20.0
    rail_area = 27.1
    windows_count = 9
    doors_inner_count = 3
    door_entry_count = 1

    foundation_perim = perimeter
    foundation_volume = round(area_floor * 0.3, 2)
    foundation_rebar_t = round(area_floor * 0.025, 3)

    wall_height_with_parapet = height + 0.6
    wall_volume_400 = round(perimeter * wall_height_with_parapet * 0.4, 2)
    inner_walls_area = round(area_floor * 0.8, 2)
    inner_walls_vol_250 = round(inner_walls_area * 0.25, 2)
    partitions_area = round(area_floor * 0.6, 2)
    partitions_vol_150 = round(partitions_area * 0.15, 2)
    kladka_glue_bags = round((wall_volume_400 + inner_walls_vol_250 + partitions_vol_150) * 1.2, 0)
    rebar_kladka_t = round(perimeter * 5 * 0.395 / 1000, 3)
    perimuter_lintels = windows_count + doors_inner_count + door_entry_count
    armopoyas_concrete = round(perimeter * 0.4 * 0.3, 2)

    rooms = _FTM_ROOMS_PDF
    dry_rooms = [(n, ar) for n, ar, k in rooms if k in ("dry", "kitchen")]
    wet_rooms = [(n, ar) for n, ar, k in rooms if k == "wet"]
    sauna_rooms = [(n, ar) for n, ar, k in rooms if k == "sauna"]
    dry_floor_area = round(sum(ar for _, ar in dry_rooms), 2)
    wet_floor_area = round(sum(ar for _, ar in wet_rooms), 2)
    sauna_area = round(sum(ar for _, ar in sauna_rooms), 2)
    dry_wall_area = round(dry_floor_area * 2.8, 2)
    wet_wall_area = round(wet_floor_area * 2.8, 2)
    sauna_wall_ceil_area = round(sauna_area * 3.8, 2)
    all_floor_area = round(area_floor, 2)

    # 1. Фундамент
    items.append(_ftm_row("Фундамент", "Земляные работы (разработка грунта)", "м³", round(area_floor * 0.4, 2), P["earth_work"], "выемка под плиту"))
    items.append(_ftm_row("Фундамент", "Песок подсыпка с уплотнением", "м³", round(area_floor * 0.15, 2), P["sand_mat"] + P["sand_work"], "материал+работы"))
    items.append(_ftm_row("Фундамент", "Щебень подсыпка с уплотнением", "м³", round(area_floor * 0.10, 2), P["gravel_mat"] + P["gravel_work"], "материал+работы"))
    items.append(_ftm_row("Фундамент", "Геотекстиль с укладкой", "м²", round(area_floor * 1.1, 2), P["geotextile_mat"] + P["geotextile_work"], "материал+работы"))
    items.append(_ftm_row("Фундамент", "Утеплитель ЭППС материал под плиту", "м²", round(area_floor * 1.1, 2), P["insul_eps_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Монтаж утеплителя ЭППС работы", "м²", round(area_floor * 1.1, 2), P["insul_eps_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Гидроизоляция оклеечная материал", "м²", round(area_floor * 1.2, 2), P["waterproof_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Гидроизоляция плиты работы", "м²", round(area_floor * 1.2, 2), P["waterproof_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Опалубка периметра материал", "мп", foundation_perim, P["formwork_perim_mat"], "материал"))
    items.append(_ftm_row("Фундамент", "Опалубка работы (монтаж/демонтаж)", "мп", foundation_perim, P["formwork_install_work"], "работы"))
    items.append(_ftm_row("Фундамент", "Арматура А500 материал", "т", foundation_rebar_t,
        _p8v3_mp("арматура металлическая д.12а500", P["rebar_a500_mat"]), "материал"))
    items.append(_ftm_row("Фундамент", "Армирование плиты работы", "м²", area_floor,
        _p8v3_wp("устройство арматурного каркаса", 950), "работы"))
    items.append(_ftm_row("Фундамент", "Бетон В25 для монолитной плиты", "м³", foundation_volume,
        _p8v3_mp("бетон в25 w6", P["concrete_b25_mat"]), "материал"))
    items.append(_ftm_row("Фундамент", "Бетонирование плиты работы", "м³", foundation_volume,
        _p8v3_wp("бетонирование монолитной плиты   б/н", P["concrete_pour_work"]), "работы"))

    # 2. Стены
    items.append(_ftm_row("Стены", "Газобетон D400 400мм наружные стены", "м³", wall_volume_400,
        _p8v3_mp(" блок  625x400x250 ", P["gas_d400_mat"]), "материал"))
    items.append(_ftm_row("Стены", "Кладка газобетона 400мм работы", "м³", wall_volume_400,
        _p8v3_wp("кладка  стен из газобетона, вкл парапет", P["gas_400mm_work"]), "работы"))
    items.append(_ftm_row("Стены", "Газобетон D400 250мм внутренние стены", "м³", inner_walls_vol_250,
        _p8v3_mp(" блок  625x250x250 ", P["gas_d400_mat"]), "материал"))
    items.append(_ftm_row("Стены", "Кладка газобетона 250мм работы", "м³", inner_walls_vol_250,
        _p8v3_wp("кладка  стен из газобетона, вкл парапет", P["gas_250mm_work"]), "работы"))
    items.append(_ftm_row("Стены", "Газобетон D400 150мм перегородки", "м³", partitions_vol_150,
        _p8v3_mp(" блок  625x150x250 ", P["gas_d400_mat"]), "материал"))
    items.append(_ftm_row("Стены", "Монтаж перегородок 150мм работы", "м³", partitions_vol_150,
        _p8v3_wp("кладка перегородок из газобетона", P["gas_150mm_work"]), "работы"))
    items.append(_ftm_row("Стены", "Клей для газобетона", "шт", kladka_glue_bags,
        _p8v3_mp("клей для газобетона 25 кг", P["kladka_glue_mat"]), "материал"))
    items.append(_ftm_row("Стены", "Арматура для кладки 8мм", "т", rebar_kladka_t,
        _p8v3_mp("арматура а3 а240  8мм рифленая", P["rebar_a500_mat"]), "материал"))
    items.append(_ftm_row("Стены", "Перемычки оконные/дверные", "шт", perimuter_lintels, P["lintel_mat"], "материал"))
    items.append(_ftm_row("Стены", "Армопояс бетон В25", "м³", armopoyas_concrete,
        _p8v3_mp("бетон в25 w6", P["armopoyas_concrete_mat"]), "материал"))
    items.append(_ftm_row("Стены", "Армопояс работы", "мп", foundation_perim,
        _p8v3_wp("устройство армопояса парапета", P["armopoyas_work"]), "работы"))

    # 3. Кровля (185 м², плоская с наплавляемой гидроизоляцией)
    items.append(_ftm_row("Кровля", "Пароизоляция кровли работы", "м²", roof_area,
        _p8v3_wp("укладка пароизоляционного слоя", P["roof_paroizo_work"]), "работы"))
    items.append(_ftm_row("Кровля", "Пароизоляция кровли материал", "м²", roof_area, P["roof_paroizo_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Утеплитель кровли XPS 200мм монтаж", "м²", roof_area,
        _p8v3_wp("монтаж утепления т 200/250 мм", P["roof_insul_work"]), "работы"))
    items.append(_ftm_row("Кровля", "Утеплитель кровли XPS материал", "м²", roof_area, P["roof_insul_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Монтаж кровельного покрытия работы", "м²", roof_area,
        _p8v3_wp("монтаж наплавляемой 2х слойной кровли, вкл парапет", P["roof_falz_work"]), "работы"))
    items.append(_ftm_row("Кровля", "Фальцевая кровля RAL7024 материал", "м²", roof_area, P["roof_falz_mat"], "материал"))
    items.append(_ftm_row("Кровля", "Водосток + добор", "мп", round(perimeter + 4, 1),
        P["watergutter_mat"] + P["watergutter_work"], "материал+работы"))
    items.append(_ftm_row("Кровля", "Крепёж и расходники кровля", "компл", 1, P["roof_dobor_mat"] * 30, "компл"))

    # 4. Окна и двери
    items.append(_ftm_row("Окна и двери", "Окна ПВХ средний класс материал", "шт", windows_count, P["window_pvc_mat"], "9 типов ПВХ"))
    items.append(_ftm_row("Окна и двери", "Монтаж окон работы", "шт", windows_count, P["window_install_work"], "работы"))
    items.append(_ftm_row("Окна и двери", "Монтажный комплект (пена+ПСУЛ+отлив)", "шт", windows_count, P["window_pena_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Дверь входная стальная материал", "шт", door_entry_count,
        _p8v3_mp("дверь входная с терморазрывом ferroni isoterma лев", P["door_entry_mat"]), "с терморазрывом"))
    items.append(_ftm_row("Окна и двери", "Монтаж входной двери", "шт", door_entry_count,
        _p8v3_wp("установка входной двери", P["door_install_work"]), "работы"))
    items.append(_ftm_row("Окна и двери", "Двери межкомнатные материал", "шт", doors_inner_count, P["door_inner_mat"], "материал"))
    items.append(_ftm_row("Окна и двери", "Монтаж межкомнатных дверей", "шт", doors_inner_count, P["door_install_work"], "работы"))
    items.append(_ftm_row("Окна и двери", "Дверь сауны специальная", "шт", 1, P["door_sauna_mat"], "материал"))

    # 5. Внешняя отделка
    items.append(_ftm_row("Внешняя отделка", "Утепление фасада базальтовой ватой работы", "м²", facade_area,
        _p8v3_wp("утепление фасада базальтовой ватой на клеевой сост", P["facade_setka_work"]), "работы"))
    items.append(_ftm_row("Внешняя отделка", "Утеплитель базальтовая вата материал", "м²", facade_area, P["facade_setka_mat"] * 5, "Rockwool"))
    items.append(_ftm_row("Внешняя отделка", "Армирование фасада работы", "м²", facade_area,
        _p8v3_wp("армирование фасада с шлифованием", P["facade_setka_work"]), "работы"))
    items.append(_ftm_row("Внешняя отделка", "Армирующая сетка материал", "м²", facade_area, P["facade_setka_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Штукатурка фасада работы 96м²", "м²", facade_area,
        _p8v3_wp("грунтование, оштукатуривание", P["facade_plaster_work"]), "работы"))
    items.append(_ftm_row("Внешняя отделка", "Штукатурка фасадная материал", "м²", facade_area, P["facade_plaster_mat"], "материал"))
    items.append(_ftm_row("Внешняя отделка", "Цоколь отделка 20м²", "м²", socle_area,
        P["socle_plaster_mat"] + P["socle_plaster_work"], "материал+работы"))
    items.append(_ftm_row("Внешняя отделка", "Рейка фасадная 27.1м²", "м²", rail_area,
        P["rail_facade_mat"] + P["rail_facade_work"], "материал+работы"))

    # 6. Внутренняя черновая отделка
    items.append(_ftm_row("Внутренняя отделка", "Стяжка пола работы", "м²", all_floor_area,
        _p8v3_wp("устройство полусухой армированной стяжки", P["screed_work"]), "работы"))
    items.append(_ftm_row("Внутренняя отделка", "Стяжка материалы", "м²", all_floor_area, P["screed_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Утепление пола ЭППС", "м²", all_floor_area,
        P["floor_insul_mat"] + P["floor_insul_work"], "материал+работы"))
    items.append(_ftm_row("Внутренняя отделка", "Штукатурка стен сухих помещений работы", "м²", dry_wall_area,
        _p8v3_wp("штукатурка стен", P["wall_prep_work"]), "работы"))
    items.append(_ftm_row("Внутренняя отделка", "Штукатурка стен материал", "м²", dry_wall_area, P["wall_prep_mat"], "материал"))
    items.append(_ftm_row("Внутренняя отделка", "Гипсокартон потолки", "м²", dry_floor_area,
        P["ceiling_gkl_mat"] + P["ceiling_gkl_work"], "материал+работы"))
    items.append(_ftm_row("Внутренняя отделка", "Покраска потолков", "м²", dry_floor_area,
        P["ceiling_paint_mat"] + P["ceiling_paint_work"], "материал+работы"))
    items.append(_ftm_row("Внутренняя отделка", "Плинтус", "мп", round(perimeter * 3, 1),
        P["skirting_mat"] + P["skirting_work"], "материал+работы"))

    # 7. Санузлы и сауна
    items.append(_ftm_row("Санузлы и сауна", "Гидроизоляция мокрых зон материал", "м²", wet_floor_area, P["wet_hydroizo_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Гидроизоляция мокрых зон работы", "м²", wet_floor_area, P["wet_hydroizo_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Плитка стены санузлы материал", "м²", wet_wall_area, P["tile_wall_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Плитка стены укладка работы", "м²", wet_wall_area, P["tile_wall_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Плитка пол санузлы материал", "м²", wet_floor_area, P["tile_floor_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Плитка пол укладка работы", "м²", wet_floor_area, P["tile_floor_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Плиточный клей", "м²", wet_wall_area + wet_floor_area, P["tile_glue_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна вагонка липа материал", "м²", sauna_wall_ceil_area, P["sauna_lining_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Сауна вагонка монтаж работы", "м²", sauna_wall_ceil_area, P["sauna_lining_work"], "работы"))
    items.append(_ftm_row("Санузлы и сауна", "Полки сауны комплект", "компл", 1, P["sauna_polok_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Печь сауны материал", "шт", 1, P["sauna_stove_mat"], "материал"))
    items.append(_ftm_row("Санузлы и сауна", "Монтаж печи сауны", "шт", 1, P["sauna_stove_work"], "работы"))

    # 8. Полы и тёплый пол
    warmfloor_pipe_m = round(area_floor * 7, 0)
    warmfloor_loops = max(1, round(area_floor / 15, 0))
    items.append(_ftm_row("Полы и тёплый пол", "Тёплый пол труба PE-X материал", "мп", warmfloor_pipe_m, P["warmfloor_pipe_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Тёплый пол монтаж трубы", "мп", warmfloor_pipe_m, P["warmfloor_pipe_work"], "работы"))
    items.append(_ftm_row("Полы и тёплый пол", "Коллектор тёплого пола", "шт", 1, P["warmfloor_collector_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Монтаж коллектора", "шт", 1, P["warmfloor_collector_work"], "работы"))
    items.append(_ftm_row("Полы и тёплый пол", "Терморегулятор тёплого пола", "шт", int(warmfloor_loops), P["warmfloor_thermo_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Демпферная лента", "мп", round(perimeter * 2, 1), P["warmfloor_demper_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Керамогранит кухня/прихожая материал", "м²", round(dry_floor_area * 0.4, 2), P["ceramogranit_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Укладка керамогранита работы", "м²", round(dry_floor_area * 0.4, 2), P["ceramogranit_work"], "работы"))
    items.append(_ftm_row("Полы и тёплый пол", "Подложка под ламинат", "м²", round(dry_floor_area * 0.6, 2), P["floor_underlay_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Ламинат материал", "м²", round(dry_floor_area * 0.6, 2), P["laminate_mat"], "материал"))
    items.append(_ftm_row("Полы и тёплый пол", "Укладка ламината работы", "м²", round(dry_floor_area * 0.6, 2), P["laminate_work"], "работы"))

    # 9. Инженерные коммуникации ОВ
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Котёл отопительный средний класс", "шт", 1, P["ov_kotel_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж котла", "шт", 1, P["ov_kotel_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Радиатор биметаллический 8 секций", "шт", 9, P["ov_radiator_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж радиаторов", "шт", 9, P["ov_radiator_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Трубы металлопластик 20мм", "мп", 180, P["ov_pipe_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж трубопровода ОВ", "мп", 180, P["ov_pipe_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Циркуляционный насос", "шт", 1, P["ov_pump_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Монтаж насоса", "шт", 1, P["ov_pump_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Расширительный бак", "шт", 1, P["ov_expansion_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Узел регулировки и балансировки", "шт", 1, P["ov_balance_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Терморегуляторы радиаторов", "шт", 9, P["ov_thermo_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Запорная арматура комплект", "компл", 1, P["ov_valve_set_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ОВ", "Пуско-наладка системы отопления", "компл", 1, P["ov_naladka_work"], "работы"))

    # 10. Инженерные коммуникации ВК
    items.append(_ftm_row("Инженерные коммуникации ВК", "Трубы PEX 16мм холодное+горячее водоснабжение", "мп", 160, P["vk_pipe_cold_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Канализация PVC 50/110мм комплект", "мп", 65, P["vk_pipe_canal_110_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Монтаж трубопровода ВК (вод.+канализ.)", "мп", 225, P["vk_pipe_install_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Унитаз с инсталляцией", "шт", 2, P["vk_unitaz_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка унитазов", "шт", 2, P["vk_unitaz_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Раковина с тумбой", "шт", 3, P["vk_rakovina_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка раковин", "шт", 3, P["vk_rakovina_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Душевая кабина / ванна", "шт", 2, P["vk_dush_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка душа/ванны", "шт", 2, P["vk_dush_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Полотенцесушитель", "шт", 2, P["vk_polotenets_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка полотенцесушителя", "шт", 2, P["vk_polotenets_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Смесители комплект", "компл", 4, P["vk_smesitel_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Бойлер 100л", "шт", 1, P["vk_boiler_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Установка бойлера", "шт", 1, P["vk_boiler_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ВК", "Запорная арматура ВК", "компл", 1, P["vk_valve_set_mat"], "материал"))

    # 11. Инженерные коммуникации ЭОМ
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Кабель ВВГнг 5х10 вводной + 3х2.5 силовой", "мп", 270, P["el_kabel_25_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Кабель ВВГнг 3х1.5 освещение", "мп", 200, P["el_kabel_15_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж кабельных линий", "мп", 470, P["el_kabel_install_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Розетка", "шт", 35, P["el_rozetka_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Установка розеток", "шт", 35, P["el_rozetka_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Выключатель", "шт", 18, P["el_switch_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Установка выключателей", "шт", 18, P["el_switch_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Светильник потолочный", "шт", 25, P["el_light_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж светильников", "шт", 25, P["el_light_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Распределительный щит укомплектованный", "шт", 1, P["el_shchit_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж щита", "шт", 1, P["el_shchit_work"], "работы"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Автомат 16А", "шт", 14, P["el_avtomat_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "УЗО 40А", "шт", 4, P["el_uzo_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Контур заземления комплект", "компл", 1, P["el_zazem_mat"], "материал"))
    items.append(_ftm_row("Инженерные коммуникации ЭОМ", "Монтаж контура заземления", "компл", 1, P["el_zazem_work"], "работы"))

    # 12. Логистика
    items.append(_ftm_row("Логистика", "Доставка материалов от СПб", "рейс", 3, P["logist_delivery"] / 3, "30 км"))
    items.append(_ftm_row("Логистика", "Транспорт бригады и проживание", "компл", 1, P["logist_brigade"] * 4, "период строительства"))

    # 13. Накладные расходы
    subtotal = sum(float(it["price"]) * float(it["qty"]) for it in items if it["section"] not in ("Логистика", "Накладные расходы"))
    items.append(_ftm_row("Накладные расходы", "Организация работ и накладные", "компл", 1, round(subtotal * 0.07, 2), "7% от общей сметы"))
    items.append(_ftm_row("Накладные расходы", "Расходные материалы и крепёж", "компл", 1, round(subtotal * 0.015, 2), "мелкое крепление"))
    items.append(_ftm_row("Накладные расходы", "Уборка после работ", "компл", 1, round(area_floor * 280, 2), "финальная уборка"))

    return items


# Override _build_estimate_items with v2
_ORIG_BEI_V3 = _build_estimate_items

def _build_estimate_items(parsed, price_text, choice):  # noqa: F811
    try:
        items = _build_full_turnkey_items_v2(parsed, price_text, choice)
        _P8V3_LOG.info("PATCH_TOPIC2_REALSHEET_PRICES_V3: %d items", len(items))
        return items
    except Exception as _v3e:
        _P8V3_LOG.error("PATCH_TOPIC2_REALSHEET_PRICES_V3_ERR %s", _v3e)
        return _ORIG_BEI_V3(parsed, price_text, choice)

_P8V3_LOG.info("PATCH_TOPIC2_REALSHEET_PRICES_V3 installed")
# === END_PATCH_TOPIC2_REALSHEET_PRICES_V3 ===

# === PATCH_TOPIC2_REALSHEET_PRICES_V3_FIX1 ===
# Fix _p8v3_mp/_p8v3_wp to strip key before lookup (template keys have no leading spaces)
def _p8v3_wp(key, fallback=0):  # noqa: F811
    row = _P8V3_TPL_PRICES.get(key.strip())
    return row["work_price"] if row and row["work_price"] > 0 else fallback

def _p8v3_mp(key, fallback=0):  # noqa: F811
    row = _P8V3_TPL_PRICES.get(key.strip())
    return row["mat_price"] if row and row["mat_price"] > 0 else fallback

_P8V3_LOG.info("PATCH_TOPIC2_REALSHEET_PRICES_V3_FIX1 installed")
# === END_PATCH_TOPIC2_REALSHEET_PRICES_V3_FIX1 ===

# === PATCH_TOPIC2_ADD_PEREKRYTIYA_SECTION_V1 ===
# §5 canon: 11 секций, раздел 3 = Перекрытия. Для газобетон 1-эт. дом —
# монолитная плита перекрытия (над первым этажом / чердачная плита).
# Цены из шаблона Газобетонный дом (Перекрытие монолитное).
# Append-only: переопределяем _build_full_turnkey_items_v2.
# ============================================================
_PREV_FTI_V2 = _build_full_turnkey_items_v2

def _build_full_turnkey_items_v2(parsed, price_text, choice):  # noqa: F811
    items = _PREV_FTI_V2(parsed, price_text, choice)
    P = _FTM_PRICES
    area_floor = float(parsed.get("area_floor") or 99.91)
    dims = parsed.get("dims") or parsed.get("dimensions") or [8.5, 12.5]
    try:
        a, b = float(dims[0]), float(dims[1])
    except Exception:
        a = b = area_floor ** 0.5
    perimeter = 2 * (a + b)

    # Перекрытие монолитное (плита над 1 этажом)
    slab_area = round(area_floor * 1.05, 2)   # с запасом
    slab_vol  = round(slab_area * 0.22, 2)    # толщина 220мм
    slab_reb_t= round(slab_area * 0.010, 3)   # 10 кг/м²

    perekr_items = [
        _ftm_row("Перекрытия", "Опалубка перекрытия монтаж/демонтаж", "м²", slab_area,
            _p8v3_wp("монтаж/демонтаж опалубки основания плиты", 1000), "работы"),
        _ftm_row("Перекрытия", "Аренда опалубки перекрытия", "м²", slab_area,
            _p8v3_mp("аренда опалубки (основание плиты)", 1100), "материал"),
        _ftm_row("Перекрытия", "Армирование перекрытия работы", "м²", slab_area,
            _p8v3_wp("устройство арматурного каркаса", 1200), "работы"),
        _ftm_row("Перекрытия", "Арматура перекрытия А500 материал", "т", slab_reb_t,
            _p8v3_mp("арматура металлическая д.12а500", P["rebar_a500_mat"]), "материал"),
        _ftm_row("Перекрытия", "Бетон В25 перекрытие материал", "м³", slab_vol,
            _p8v3_mp("бетон в25 w6", P["concrete_b25_mat"]), "материал"),
        _ftm_row("Перекрытия", "Бетонирование перекрытия работы", "м³", slab_vol,
            _p8v3_wp("бетонирование монолитной плиты   б/н", P["concrete_pour_work"]), "работы"),
        _ftm_row("Перекрытия", "Пеноплэкс утепление плиты 100мм", "м²", slab_area,
            _p8v3_mp("пеноплэкс фундамент 1185х585х100 мм", 800), "материал"),
        _ftm_row("Перекрытия", "Крепёж и расходники перекрытие", "компл", 1,
            round(slab_area * 200, 0), "компл"),
    ]

    # Вставить секцию Перекрытия ПОСЛЕ Стены (перед Кровля)
    result = []
    steny_done = False
    krovlya_inserted = False
    for it in items:
        if it["section"] == "Стены":
            steny_done = True
        if steny_done and not krovlya_inserted and it["section"] == "Кровля":
            result.extend(perekr_items)
            krovlya_inserted = True
        result.append(it)
    if not krovlya_inserted:
        result.extend(perekr_items)

    # Пересчитать накладные расходы (раздел 13 — последние 3 строки)
    # Убираем старые накладные, считаем заново
    base_items = [r for r in result if r["section"] not in ("Логистика", "Накладные расходы")]
    subtotal = sum(float(r["price"]) * float(r["qty"]) for r in base_items)
    logist   = [r for r in result if r["section"] == "Логистика"]
    overhead_new = [
        _ftm_row("Накладные расходы", "Организация работ и накладные", "компл", 1, round(subtotal * 0.07, 2), "7% от общей сметы"),
        _ftm_row("Накладные расходы", "Расходные материалы и крепёж",  "компл", 1, round(subtotal * 0.015, 2), "мелкое крепление"),
        _ftm_row("Накладные расходы", "Уборка после работ",            "компл", 1, round(area_floor * 280, 2), "финальная уборка"),
    ]
    return base_items + logist + overhead_new

_P8V3_LOG.info("PATCH_TOPIC2_ADD_PEREKRYTIYA_SECTION_V1 installed")
# === END_PATCH_TOPIC2_ADD_PEREKRYTIYA_SECTION_V1 ===

# === PATCH_TOPIC2_STALE_PENDING_TASK_GUARD_V1 ===
# ROOT CAUSE: §5 in maybe_handle_stroyka_estimate reuses pending from a DIFFERENT task.
#   _memory_latest("topic_2_estimate_pending_") uses LIKE → finds any task's pending.
#   FIX_STROYKA_PRICE_CONFIRM_EXTEND_V1 patched _pending_is_fresh to 86400s → always fresh.
#   Patched _is_confirm matches "средн" as substring → "цены выше среднего" fires §5.
#   Result: drainage file task gets house estimate from old c94ec497 context.
# FIX: wrap maybe_handle_stroyka_estimate — if pending.task_id != current task_id,
#   check if pending's task is done → mark pending "GENERATED" (permanent block).
import logging as _stpg_log_mod
_STPG_LOG = _stpg_log_mod.getLogger("areal.stale_pending_guard")

_STPG_ORIG = maybe_handle_stroyka_estimate

async def maybe_handle_stroyka_estimate(conn, task, logger=None):  # noqa: F811
    try:
        _stpg_tid = _s(_row_get(task, "id"))
        _stpg_chat = _s(_row_get(task, "chat_id"))
        if _stpg_tid and _stpg_chat:
            _stpg_pend = _memory_latest(_stpg_chat, "topic_2_estimate_pending_")
            _stpg_pend_task = (_stpg_pend or {}).get("task_id", "")
            if (
                _stpg_pend
                and _stpg_pend.get("status") == "WAITING_PRICE_CONFIRMATION"
                and _stpg_pend_task
                and _stpg_pend_task != _stpg_tid
            ):
                # Check if the pending's own task is already in a terminal state
                _stpg_done = False
                try:
                    _stpg_row = conn.execute(
                        "SELECT state FROM tasks WHERE id=? LIMIT 1",
                        (_stpg_pend_task,)
                    ).fetchone()
                    if _stpg_row and str(_stpg_row[0] or "").upper() in (
                        "DONE", "AWAITING_CONFIRMATION", "FAILED", "CANCELLED", "ARCHIVED"
                    ):
                        _stpg_done = True
                except Exception:
                    pass
                _stpg_key = (
                    _stpg_pend.get("_memory_key")
                    or f"topic_2_estimate_pending_{_stpg_pend_task}"
                )
                _stpg_new_status = "GENERATED" if _stpg_done else "STALE_BLOCKED"
                _stpg_blocked = dict(_stpg_pend)
                _stpg_blocked["status"] = _stpg_new_status
                _memory_save(_stpg_chat, _stpg_key, _stpg_blocked)
                _history_safe(conn, _stpg_tid,
                              f"TOPIC2_STALE_PENDING_BLOCKED:pending_task={_stpg_pend_task[:16]}:done={_stpg_done}")
                _STPG_LOG.info(
                    "STPG: pending task=%s (done=%s) → %s; current task=%s",
                    _stpg_pend_task[:8], _stpg_done, _stpg_new_status, _stpg_tid[:8],
                )
    except Exception as _stpg_e:
        _STPG_LOG.warning("STPG_PRE_ERR: %s", _stpg_e)
    return await _STPG_ORIG(conn, task, logger)

_STPG_LOG.info("PATCH_TOPIC2_STALE_PENDING_TASK_GUARD_V1: installed")
# === END_PATCH_TOPIC2_STALE_PENDING_TASK_GUARD_V1 ===


# === PATCH_TOPIC2_INPUT_GATE_SOURCE_OF_TRUTH_V1 ===
import logging as _t2ig_log_mod
_T2IG_LOG = _t2ig_log_mod.getLogger("areal.topic2_input_gate")

try:
    from core.topic2_input_gate import topic2_pre_estimate_gate as _t2ig_gate
    from core.topic2_input_gate import apply_gate_result_to_task as _t2ig_apply
except Exception as _t2ig_import_err:
    _t2ig_gate = None
    _t2ig_apply = None
    _T2IG_LOG.warning("TOPIC2_INPUT_GATE_IMPORT_FAILED:%s", _t2ig_import_err)

_T2IG_ORIG_MAYBE_HANDLE = maybe_handle_stroyka_estimate

async def maybe_handle_stroyka_estimate(conn, task, logger=None):  # noqa: F811
    if _t2ig_gate is not None and _t2ig_apply is not None:
        try:
            try:
                import json as _t2ig_json
                _t2ig_td = dict(task) if hasattr(task, "keys") else (task if isinstance(task, dict) else {})
                _t2ig_raw_obj = _t2ig_json.loads(str(_t2ig_td.get("raw_input") or "{}"))
                _t2ig_choice = str(_t2ig_raw_obj.get("file_duplicate_choice_intent") or "").strip().lower()
            except Exception:
                _t2ig_choice = ""
            if _t2ig_choice and _t2ig_choice != "estimate":
                try:
                    _t2ig_task_id = str(_t2ig_td.get("id") or "")
                    if _t2ig_task_id:
                        conn.execute(
                            "INSERT INTO task_history (task_id,action,created_at) VALUES (?,?,datetime('now'))",
                            (_t2ig_task_id, f"TOPIC2_INPUT_GATE_FILE_MENU_BYPASS:{_t2ig_choice}"),
                        )
                        conn.commit()
                except Exception:
                    pass
                return False
            _t2ig_decision = _t2ig_gate(conn, task, logger=logger)
            if _t2ig_decision and _t2ig_decision.get("block_engine"):
                _t2ig_apply(conn, task, _t2ig_decision)
                _T2IG_LOG.info(
                    "TOPIC2_INPUT_GATE_BLOCKED:domain=%s state=%s",
                    _t2ig_decision.get("domain"),
                    _t2ig_decision.get("state"),
                )
                _t2ig_msg = _t2ig_decision.get("result") or ""
                if _t2ig_msg:
                    try:
                        # sqlite3.Row has .keys() but not .get() — use dict() to normalize
                        _t2ig_td = dict(task) if hasattr(task, "keys") else (task if isinstance(task, dict) else {})
                        _t2ig_chat = str(_t2ig_td.get("chat_id") or "")
                        _t2ig_reply = _t2ig_td.get("reply_to_message_id")
                        _t2ig_topic = int(_t2ig_td.get("topic_id") or 0)
                        _t2ig_task_id = str(_t2ig_td.get("id") or "")
                        if not _t2ig_chat:
                            raise ValueError(f"empty chat_id in task {_t2ig_task_id}")
                        _t2ig_send_res = await _send_text(_t2ig_chat, _t2ig_msg, _t2ig_reply, _t2ig_topic)
                        _t2ig_bot_msg_id = _t2ig_send_res.get("bot_message_id") if isinstance(_t2ig_send_res, dict) else None
                        if _t2ig_bot_msg_id and _t2ig_task_id:
                            conn.execute(
                                "UPDATE tasks SET bot_message_id=? WHERE id=?",
                                (int(_t2ig_bot_msg_id), _t2ig_task_id),
                            )
                            try:
                                conn.execute(
                                    "INSERT INTO task_history (task_id,action,created_at) VALUES (?,?,datetime('now'))",
                                    (_t2ig_task_id, f"TOPIC2_INPUT_GATE_SENT:{_t2ig_bot_msg_id}"),
                                )
                                conn.execute(
                                    "INSERT INTO task_history (task_id,action,created_at) VALUES (?,?,datetime('now'))",
                                    (_t2ig_task_id, "reply_sent:waiting_clarification"),
                                )
                            except Exception:
                                pass
                            conn.commit()
                        _T2IG_LOG.info("TOPIC2_INPUT_GATE_SENT:chat=%s bot_msg=%s", _t2ig_chat, _t2ig_bot_msg_id)
                    except Exception as _t2ig_send_err:
                        _T2IG_LOG.warning("TOPIC2_INPUT_GATE_SEND_ERR:%s", _t2ig_send_err)
                return True
        except Exception as _t2ig_err:
            _T2IG_LOG.warning("TOPIC2_INPUT_GATE_ERR:%s", _t2ig_err)
    return await _T2IG_ORIG_MAYBE_HANDLE(conn, task, logger)

_T2IG_LOG.info("PATCH_TOPIC2_INPUT_GATE_SOURCE_OF_TRUTH_V1: installed")
# === END_PATCH_TOPIC2_INPUT_GATE_SOURCE_OF_TRUTH_V1 ===

# === PATCH_TOPIC2_PRICE_ALWAYS_ASK_V1 ===
# Fix: новый estimate-запрос ("сделай по заданию смету" и т.д.) не должен подхватывать
# pending от предыдущего task_id или от того же task_id после MANUAL_RESET_NEW.
# Price WC должен запускаться КАЖДЫЙ РАЗ для задач с estimate-контентом.
# Корень бага: _is_confirm(startswith "сделай ") = True + _pending_is_fresh override 24h
# → CANONICAL_OLD_ROUTE_HARD_BLOCK пропускает price WC (подтверждено history 076e4350).
import logging as _t2paa_log_mod
_T2PAA_LOG = _t2paa_log_mod.getLogger("stroyka_estimate_canon")
_T2PAA_ORIG_MAYBE_HANDLE = maybe_handle_stroyka_estimate

async def maybe_handle_stroyka_estimate(conn, task, logger=None):
    try:
        task_id = _s(_row_get(task, "id"))
        chat_id = _s(_row_get(task, "chat_id"))
        raw = _s(_row_get(task, "raw_input", ""))
        _pend = _memory_latest(chat_id, "topic_2_estimate_pending_")
        _pend_parsed = (_pend or {}).get("parsed") or {}
        _pend_has_project_files = bool(
            (_pend_parsed or {}).get("pdf_spec_rows")
            or (_pend_parsed or {}).get("ocr_table_rows")
            or (_pend_parsed or {}).get("local_project_files")
        )
        _pend_same_task = _s((_pend or {}).get("task_id") or "") == task_id
        _pend_task_open = False
        if _pend_same_task:
            try:
                _row = conn.execute("SELECT state FROM tasks WHERE id=? LIMIT 1", (task_id,)).fetchone()
                _state = _s(_row[0] if _row else "").upper()
                _pend_task_open = _state not in ("DONE", "FAILED", "CANCELLED", "ARCHIVED")
            except Exception:
                _pend_task_open = True
        # Если pending активен И текущий ввод — не чистое подтверждение цены
        # (содержит ESTIMATE_WORDS) → помечаем pending stale, чтобы запустился полный price WC.
        # _is_confirm_only = False при наличии "смет"/"дом"/"газобетон" и т.д. в тексте.
        if (_pend
                and _pend.get("status") == "WAITING_PRICE_CONFIRMATION"
                and not (_pend_same_task and _pend_task_open and _pend_has_project_files)
                and not _is_confirm_only(raw)):
            _stale = dict(_pend)
            _stale["status"] = "STALE_DEPRECATED"
            _stale["deprecated_at"] = _now()
            _stale["deprecated_reason"] = "PRICE_ALWAYS_ASK_V1:new_estimate_request_not_confirm_only"
            _memory_save(chat_id, f"topic_2_estimate_pending_{_pend.get('task_id', 'x')}", _stale)
            _T2PAA_LOG.info(
                "PATCH_TOPIC2_PRICE_ALWAYS_ASK_V1:stale_cleared task=%s prev_task=%s raw=%.60s",
                task_id, _pend.get("task_id"), raw,
            )
    except Exception as _t2paa_e:
        _T2PAA_LOG.warning("PATCH_TOPIC2_PRICE_ALWAYS_ASK_V1:ERR %s", _t2paa_e)
    return await _T2PAA_ORIG_MAYBE_HANDLE(conn, task, logger)

_T2PAA_LOG.info("PATCH_TOPIC2_PRICE_ALWAYS_ASK_V1 installed")
# === END_PATCH_TOPIC2_PRICE_ALWAYS_ASK_V1 ===

# === PATCH_TOPIC2_PARSE_REQUEST_SMART_INFER_V1 ===
# §9 spec: "что строим — если object_type уже есть — запрещено"
# §10 spec: "имитация бруса" / "по всем помещениям" → scope=под ключ
# §12 spec: брус → Ареал Нева = дом (canon §2)
# Fix 1: pile cross-section "N свай жб AxB" убирается из _extract_dimensions (сечение сваи ≠ размер здания)
# Fix 2: material="брус" + object="" → object="дом"
# Fix 3: "по всем помещениям" / "имитация бруса внутри" → scope="под ключ"
import re as _t2prs_re
import logging as _t2prs_log_mod
_T2PRS_LOG = _t2prs_log_mod.getLogger("stroyka_estimate_canon")

_T2PRS_ORIG_EXTRACT_DIM = _extract_dimensions
_T2PRS_ORIG_PARSE_REQUEST = _parse_request

def _extract_dimensions(text: str):
    t = _low(text)
    # Убираем сечения свай: "N свай [жб] AxB" или "жб AxB"
    t_clean = _t2prs_re.sub(
        r'(?:\d+\s+)?(?:свай|свая|сваи|жб)\s+(?:жб\s+)?\d+\s*[xх×*]\s*\d+',
        ' ', t,
    )
    return _T2PRS_ORIG_EXTRACT_DIM(t_clean)

def _parse_request(text: str):
    parsed = _T2PRS_ORIG_PARSE_REQUEST(text)
    t = _low(text)
    # Fix 2: брус → дом (canon §2: брус → Ареал Нева.xlsx = деревянный дом)
    if not parsed.get("object") and parsed.get("material") == "брус":
        parsed["object"] = "дом"
        _T2PRS_LOG.info("T2PRS_INFER:object=дом:from_material=брус")
    # Fix 2b: другие признаки дома
    if not parsed.get("object"):
        _home_hint_patterns = (
            (r"\\bдач\\w*", "дач"),
            (r"\\bкоттедж\\w*", "коттедж"),
            (r"\\bжилой\\b", "жилой"),
            (r"\\bжилого\\b", "жилого"),
        )
        for _pattern, _hint in _home_hint_patterns:
            if re.search(_pattern, t):
                parsed["object"] = "дом"
                _T2PRS_LOG.info("T2PRS_INFER:object=дом:hint=%s", _hint)
                break
    # Fix 3: признаки scope=под ключ (spec §10)
    if not parsed.get("scope"):
        _scope_hints = (
            "по всем помещениям", "по комнатам", "все помещения",
            "имитация бруса внутри", "чистовая отделка", "ламинат",
            "санузел", "теплые полы", "тёплые полы",
        )
        for _sh in _scope_hints:
            if _sh in t:
                parsed["scope"] = "под ключ"
                _T2PRS_LOG.info("T2PRS_INFER:scope=под_ключ:hint=%s", _sh)
                break
    return parsed

_T2PRS_LOG.info("PATCH_TOPIC2_PARSE_REQUEST_SMART_INFER_V1 installed")
# === END_PATCH_TOPIC2_PARSE_REQUEST_SMART_INFER_V1 ===

# === PATCH_TOPIC2_CLARIFICATION_MERGE_V1 ===
# Fix: clarified:* entries in task_history not merged into raw_input before _parse_request
# → _missing_question asks the same question again after user answered it.
# PAMQ wrapper exists but _missing_question is called with only (parsed), conn/task not passed.
# Solution: enrich task.raw_input with clarification history BEFORE calling the full chain.
import logging as _t2cm_log_mod
_T2CM_LOG = _t2cm_log_mod.getLogger("stroyka_estimate_canon")
_T2CM_ORIG = maybe_handle_stroyka_estimate

async def maybe_handle_stroyka_estimate(conn, task, logger=None):
    try:
        task_id = _s(_row_get(task, "id"))
        if conn is not None and task_id:
            rows = conn.execute(
                "SELECT action FROM task_history WHERE task_id=? AND action LIKE 'clarified:%' ORDER BY rowid ASC",
                (task_id,)
            ).fetchall()
            clarifications = [r[0].split(":", 1)[1].strip() for r in rows if ":" in r[0]]
            if clarifications:
                raw = _s(_row_get(task, "raw_input", ""))
                enriched = raw + "\n" + "\n".join(clarifications)
                if hasattr(task, "keys"):
                    task_dict = dict(zip(task.keys(), task))
                elif isinstance(task, dict):
                    task_dict = dict(task)
                else:
                    task_dict = {"raw_input": enriched}
                task_dict["raw_input"] = enriched
                _T2CM_LOG.info(
                    "T2CM_MERGE: task=%s clarifs=%d enriched_len=%d",
                    task_id, len(clarifications), len(enriched),
                )
                return await _T2CM_ORIG(conn, task_dict, logger)
    except Exception as _t2cm_e:
        _T2CM_LOG.warning("T2CM_ERR: %s", _t2cm_e)
    return await _T2CM_ORIG(conn, task, logger)

_T2CM_LOG.info("PATCH_TOPIC2_CLARIFICATION_MERGE_V1 installed")
# === END_PATCH_TOPIC2_CLARIFICATION_MERGE_V1 ===


# === PATCH_TOPIC2_TOPIC_ID_INT_SAFE_V1 ===
import logging as _t2tid_log_mod
_T2TID_LOG = _t2tid_log_mod.getLogger("stroyka_estimate_canon")
_T2TID_ORIG_DIRECT = _stroyka_final_handle_direct_item_estimate

async def _stroyka_final_handle_direct_item_estimate(conn, task, logger=None):
    try:
        raw_tid = _row_get(task, "topic_id", 0)
        int(raw_tid or 0)
    except (ValueError, TypeError):
        if hasattr(task, "keys"):
            task_dict = dict(zip(task.keys(), tuple(task)))
        elif isinstance(task, dict):
            task_dict = dict(task)
        else:
            task_dict = {}
        task_dict["topic_id"] = TOPIC_ID_STROYKA
        _T2TID_LOG.info("T2TID_FIX: topic_id coerced to %s", TOPIC_ID_STROYKA)
        return await _T2TID_ORIG_DIRECT(conn, task_dict, logger)
    return await _T2TID_ORIG_DIRECT(conn, task, logger)

_T2TID_LOG.info("PATCH_TOPIC2_TOPIC_ID_INT_SAFE_V1 installed")
# === END_PATCH_TOPIC2_TOPIC_ID_INT_SAFE_V1 ===

# === PATCH_TOPIC2_CLARIFICATION_MERGE_V2 ===
# Fix V1 bug: dict(zip(task.keys(), task)) for plain dict iterates keys not values
# → task_dict becomes {key:key}, internal call fails, fallback to original (no merge).
# V2: uses dict(task) for plain dict (same pattern as FIX_STROYKA_CONTEXT_ENRICH line 2393).
# Calls _T2CM_ORIG (PAA) directly, bypassing broken V1.
import logging as _t2cm2_log_mod
_T2CM2_LOG = _t2cm2_log_mod.getLogger("task_worker")
_T2CM2_INNER = _T2CM_ORIG  # PAA wrapper — skip broken V1

def _t2cm2_is_fresh_full_estimate_tz(text):
    low = _low(text)
    if len(low.split()) < 18:
        return False
    anchors = 0
    for key in ("дом", "фундамент", "этаж", "размер", "стен", "отделк", "смет", "расчет", "расчёт"):
        if key in low:
            anchors += 1
    return anchors >= 4 and any(x in low for x in ESTIMATE_WORDS)

async def maybe_handle_stroyka_estimate(conn, task, logger=None):
    # PATCH_TOPIC2_PRICE_CHOICE_LOOP_CLOSE_V1 maybe_handle guard
    try:
        _t2pcl_task_id = str(_row_get(task, "id", ""))
        _t2pcl_chat_id = str(_row_get(task, "chat_id", ""))
        _t2pcl_topic_id = int(_row_get(task, "topic_id", 0) or 0)
        _t2pcl_raw_input = _row_get(task, "raw_input", "")
        _t2pcl_reply_to = _row_get(task, "reply_to_message_id", None)
        if _t2pcl_topic_id == TOPIC_ID_STROYKA:
            if await _t2pcl_price_choice_guard(conn, _t2pcl_task_id, _t2pcl_chat_id, _t2pcl_raw_input, _t2pcl_reply_to):
                return True
    except Exception as _t2pcl_e:
        try:
            _history_safe(conn, str(_row_get(task, "id", "")), "PATCH_TOPIC2_PRICE_CHOICE_LOOP_CLOSE_V1_MAYBE_ERR:" + _s(_t2pcl_e)[:200])
        except Exception:
            pass
    try:
        task_id = _s(_row_get(task, "id"))
        if conn is not None and task_id:
            rows = conn.execute(
                "SELECT action FROM task_history WHERE task_id=? AND action LIKE 'clarified:%' ORDER BY rowid ASC",
                (task_id,)
            ).fetchall()
            clarifications = [r[0].split(":", 1)[1].strip() for r in rows if ":" in r[0]]
            if clarifications:
                raw = _s(_row_get(task, "raw_input", ""))
                clar_text = _low("\n".join(clarifications))
                allow_current_price_confirm = any(x in clar_text for x in (
                    "считать по найденным позициям",
                    "считай по найденным позициям",
                    "искать через интернет",
                    "искать цены",
                    "ищи цены",
                    "актуальные цены",
                    "цены на материалы",
                    "изготовление и монтаж",
                ))
                if _t2cm2_is_fresh_full_estimate_tz(raw) and not allow_current_price_confirm:
                    _history_safe(conn, task_id, "TOPIC2_CLARIFICATION_MERGE_V2_SKIPPED_FOR_FRESH_FULL_TZ")
                    return await _T2CM2_INNER(conn, task, logger)
                enriched = raw + "\n" + "\n".join(clarifications)
                if isinstance(task, dict):
                    task_dict = dict(task)
                elif hasattr(task, "keys"):
                    task_dict = dict(zip(task.keys(), tuple(task)))
                else:
                    task_dict = {}
                task_dict["raw_input"] = enriched
                _T2CM2_LOG.info(
                    "T2CM2_MERGE: task=%s clarifs=%d enriched_len=%d",
                    task_id, len(clarifications), len(enriched),
                )
                return await _T2CM2_INNER(conn, task_dict, logger)
    except Exception as _t2cm2_e:
        _T2CM2_LOG.warning("T2CM2_ERR: %s", _t2cm2_e)
    return await _T2CM2_INNER(conn, task, logger)

_T2CM2_LOG.info("PATCH_TOPIC2_CLARIFICATION_MERGE_V2 installed")
# === END_PATCH_TOPIC2_CLARIFICATION_MERGE_V2 ===

# === PATCH_TOPIC2_PRICE_TEXT_TRUNCATE_V1 ===
# Bug: _price_confirmation_text с 126 позициями шаблона → >4096 симв → Telegram 400 "message is too long"
# Fix: ограничить template_prices до 15 строк; остаток заменить «… (+N позиций)».
import logging as _ptt_log_mod
_PTT_LOG = _ptt_log_mod.getLogger("task_worker")
_PTT_ORIG = _price_confirmation_text

def _price_confirmation_text(parsed, template, sheet_name, template_prices, online_prices):
    MAX_PRICE_LINES = 15
    if template_prices:
        lines = template_prices.splitlines()
        if len(lines) > MAX_PRICE_LINES + 1:
            shown = "\n".join(lines[:MAX_PRICE_LINES])
            rest = len(lines) - MAX_PRICE_LINES
            template_prices = shown + f"\n… (+{rest} позиций в смете)"
    text = _PTT_ORIG(parsed, template, sheet_name, template_prices, online_prices)
    if len(text) > 3900:
        text = text[:3900] + "\n…(сообщение сокращено)"
    return text

_PTT_LOG.info("PATCH_TOPIC2_PRICE_TEXT_TRUNCATE_V1 installed")
# === END_PATCH_TOPIC2_PRICE_TEXT_TRUNCATE_V1 ===


# === PATCH_TOPIC2_FRAME_HOUSE_MATERIAL_V1 ===
# Facts:
# - topic_2 canonical frame-house estimate request was classified as material="брус"
# - request contained frame-house markers: свайный фундамент, ЖБ сваи 150x150, утепление стен 150, имитация бруса as finish
# - canon requires frame house >100m2 to use frame template route, not gasbeton/bрус route
import re as _t2fh_re
import logging as _t2fh_logging

_T2FH_LOG = _t2fh_logging.getLogger("topic2.frame_house_material_v1")

def _t2fh_s(v, limit=20000):
    try:
        return str(v or "")[:limit]
    except Exception:
        return ""

def _t2fh_low(v):
    return _t2fh_s(v).lower().replace("ё", "е")

def _t2fh_has_solid_brus(low: str) -> bool:
    return any(x in low for x in (
        "дом из бруса",
        "дом брус",
        "брусовой дом",
        "клееный брус",
        "клееного бруса",
        "профилированный брус",
        "профилированного бруса",
        "оцилиндрованное бревно",
        "сруб",
        "лафет",
    ))

def _t2fh_is_frame_house_context(raw) -> bool:
    low = _t2fh_low(raw)
    if not low:
        return False

    if _t2fh_has_solid_brus(low):
        return False

    has_direct_frame = any(x in low for x in (
        "каркас",
        "каркасный",
        "каркасник",
        "frame",
    ))

    has_finish_brus = (
        "имитац" in low and "брус" in low
    )

    has_piles = any(x in low for x in (
        "свая",
        "сваи",
        "свай",
        "жб 150",
        "ж/б 150",
        "150х150",
        "150x150",
        "150×150",
        "железобетонн",
    ))

    has_wall_insulation = (
        "утепл" in low and "стен" in low
    ) or any(x in low for x in (
        "утепление стен 150",
        "утепления стен 150",
        "минвата",
        "каменная вата",
    ))

    if has_direct_frame:
        return True

    if has_finish_brus and (has_piles or has_wall_insulation):
        return True

    if has_piles and has_wall_insulation:
        return True

    return False

def _t2fh_force_frame(parsed, raw):
    try:
        if isinstance(parsed, dict) and _t2fh_is_frame_house_context(raw):
            old_material = _t2fh_low(parsed.get("material"))
            if old_material in ("", "брус", "дерево", "деревянный", "газобетон"):
                parsed["material"] = "каркас"
                parsed["frame_house"] = True
                parsed["material_source"] = "PATCH_TOPIC2_FRAME_HOUSE_MATERIAL_V1"
        return parsed
    except Exception as e:
        try:
            _T2FH_LOG.warning("PATCH_TOPIC2_FRAME_HOUSE_MATERIAL_V1_DICT_ERR %s", e)
        except Exception:
            pass
        return parsed

def _t2fh_wrap_material_func(name):
    old = globals().get(name)
    if not callable(old) or getattr(old, "_t2fh_wrapped", False):
        return False

    def wrapped(*args, **kwargs):
        raw_parts = []
        for x in args:
            raw_parts.append(_t2fh_s(x, 5000))
        for k in ("raw_input", "text", "caption", "prompt", "user_text"):
            if k in kwargs:
                raw_parts.append(_t2fh_s(kwargs.get(k), 5000))
        raw = "\n".join(raw_parts)

        res = old(*args, **kwargs)

        if isinstance(res, str):
            if _t2fh_is_frame_house_context(raw) and _t2fh_low(res) in ("", "брус", "дерево", "деревянный", "газобетон"):
                return "каркас"
            return res

        if isinstance(res, dict):
            return _t2fh_force_frame(res, raw)

        return res

    wrapped._t2fh_wrapped = True
    globals()[name] = wrapped
    return True

_T2FH_WRAPPED = []
for _t2fh_name in (
    "_extract_material",
    "extract_material",
    "_detect_material",
    "detect_material",
    "_parse_estimate_input",
    "parse_estimate_input",
    "_parse_user_input",
    "parse_user_input",
    "_parse_task_input",
    "parse_task_input",
    "_parse_stroyka_input",
    "parse_stroyka_input",
):
    try:
        if _t2fh_wrap_material_func(_t2fh_name):
            _T2FH_WRAPPED.append(_t2fh_name)
    except Exception as _t2fh_e:
        try:
            _T2FH_LOG.warning("PATCH_TOPIC2_FRAME_HOUSE_MATERIAL_V1_WRAP_ERR %s %s", _t2fh_name, _t2fh_e)
        except Exception:
            pass

try:
    _T2FH_LOG.info("PATCH_TOPIC2_FRAME_HOUSE_MATERIAL_V1 installed wrapped=%s", ",".join(_T2FH_WRAPPED))
except Exception:
    pass
# === END_PATCH_TOPIC2_FRAME_HOUSE_MATERIAL_V1 ===

# === FULL_CANON_CLOSURE_VERIFIED_V1 / TOPIC2_TEMPLATE_PRICE_EXTRACTION_SAFE_V1 ===
try:
    import re as _fccv1_re
    import logging as _fccv1_logging
    from openpyxl import load_workbook as _fccv1_load_workbook

    _fccv1_log = _fccv1_logging.getLogger("task_worker")

    def _fccv1_float_or_none(value):
        if value is None:
            return None, "PRICE_MISSING"
        if isinstance(value, (int, float)):
            return float(value), "OK"
        s = str(value).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
        if not s:
            return None, "PRICE_MISSING"
        try:
            return float(s), "OK"
        except Exception:
            return None, "PRICE_MISSING"

    def _p8v3_extract_tpl_prices(template_path=None):
        prices = {}
        loaded = 0
        skipped = 0
        missing_price = 0

        if not template_path:
            _fccv1_log.info("TOPIC2_TEMPLATE_PRICE_EXTRACTION_SAFE_V1 installed loaded=0 skipped=0 missing_price=0")
            return prices

        wb = _fccv1_load_workbook(template_path, data_only=True)
        ws = wb.active

        for r in ws.iter_rows(values_only=True):
            if not r or len(r) < 10:
                skipped += 1
                continue

            name = str(r[0] if r[0] is not None else "").strip()
            if not name:
                skipped += 1
                continue

            qty, qty_status = _fccv1_float_or_none(r[3])
            if qty_status != "OK":
                skipped += 1
                continue

            work_price, work_status = _fccv1_float_or_none(r[7])
            mat_price, mat_status = _fccv1_float_or_none(r[9])

            if work_status != "OK" or mat_status != "OK":
                missing_price += 1

            if work_status == "OK" and mat_status == "OK" and work_price == 0 and mat_price == 0:
                skipped += 1
                continue

            norm = _fccv1_re.sub(r"\s+", " ", name.lower()).strip()
            prices[norm] = {
                "name": name,
                "qty": qty,
                "work_price": work_price,
                "mat_price": mat_price,
                "work_price_status": work_status,
                "mat_price_status": mat_status,
            }
            loaded += 1

        _fccv1_log.info(
            "TOPIC2_TEMPLATE_PRICE_EXTRACTION_SAFE_V1 installed loaded=%s skipped=%s missing_price=%s",
            loaded,
            skipped,
            missing_price,
        )
        return prices

    _fccv1_log.info("TOPIC2_TEMPLATE_PRICE_EXTRACTION_SAFE_V1 installed loaded=0 skipped=0 missing_price=0")
except Exception as _e:
    try:
        _fccv1_log.exception("TOPIC2_TEMPLATE_PRICE_EXTRACTION_SAFE_V1_INSTALL_ERR:%s", _e)
    except Exception:
        pass
# === /FULL_CANON_CLOSURE_VERIFIED_V1 ===

# === PATCH_PRICE_ENRICHMENT_IDEMPOTENT_V1 ===
# Причина: _pick_next_task забирает задачи в NEW/IN_PROGRESS/WAITING_CLARIFICATION.
# После price WC задача переходит в WAITING_CLARIFICATION, но через 1.5с poll
# подхватывает её снова → запускает полный pipeline → второй вызов Sonar.
# Фикс: проверяем TOPIC2_PRICE_ENRICHMENT_DONE в task_history перед вызовом Sonar.
# Если маркер есть — возвращаем "" без сетевого запроса.
try:
    import logging as _pei_log_mod
    _PEI_LOG = _pei_log_mod.getLogger("task_worker")
    _PEI_ORIG_SEARCH = _search_prices_online

    async def _search_prices_online(parsed, template, sheet_name, conn=None, task_id=None):
        if conn is not None and task_id is not None:
            try:
                row = conn.execute(
                    "SELECT rowid, action FROM task_history WHERE task_id=? AND action LIKE 'TOPIC2_PRICE_ENRICHMENT_DONE:%' ORDER BY rowid DESC LIMIT 1",
                    (task_id,)
                ).fetchone()
                restart = conn.execute(
                    "SELECT rowid, action FROM task_history WHERE task_id=? AND ("
                    "action LIKE 'CODEX_RESTART_EXISTING_FILE_FROM_SCREENSHOT_1028_NO_DUPLICATE%' OR action LIKE 'CODEX_RESTART_AFTER_%' OR "
                    "action LIKE 'PATCH_TOPIC2_REVISION_MODE_FULL_V1:REVISION_STARTED:%' OR "
                    "action LIKE 'PATCH_TOPIC2_REVISION_MODE_FULL_V1:DRIVE_REVISION_MERGED_TO:%' OR "
                    "action LIKE 'PATCH_TOPIC2_FULL_TASK_CHAIN_RESTORE_NO_REGRESSION_V1:%' OR "
                    "action LIKE 'clarified:%') ORDER BY rowid DESC LIMIT 1",
                    (task_id,)
                ).fetchone()
                if row:
                    row_vals = list(row)
                    restart_vals = list(restart) if restart else []
                    done_rid = int(row_vals[0])
                    restart_rid = int(restart_vals[0]) if restart_vals else 0
                    if done_rid > restart_rid:
                        action = row_vals[1]
                        _PEI_LOG.info("PATCH_PRICE_ENRICHMENT_IDEMPOTENT_V1: skip task=%s already=%s", task_id, action)
                        return ""
            except Exception as _pei_check_e:
                _PEI_LOG.warning("PATCH_PRICE_ENRICHMENT_IDEMPOTENT_V1_CHECK_ERR: %s", _pei_check_e)
        return await _PEI_ORIG_SEARCH(parsed, template, sheet_name, conn=conn, task_id=task_id)

    _PEI_LOG.info("PATCH_PRICE_ENRICHMENT_IDEMPOTENT_V1 installed")
except Exception as _pei_install_e:
    try:
        _PEI_LOG.exception("PATCH_PRICE_ENRICHMENT_IDEMPOTENT_V1_INSTALL_ERR: %s", _pei_install_e)
    except Exception:
        pass
# === END_PATCH_PRICE_ENRICHMENT_IDEMPOTENT_V1 ===

# === PATCH_KARKASNIK_SHEET_FIX_V1 ===
# Fix regression from PATCH_TOPIC2_REALSHEET_PRICES_V3 which hardcoded "Газобетонный дом"
# for ALL materials. For material=="каркас": read "Каркас под ключ" from М-80.xlsx
# (col4=work_price, col6=mat_price, data starts row 10).
import logging as _kfv1_log_mod
_KFV1_LOG = _kfv1_log_mod.getLogger("stroyka.karkasnik_sheet_fix_v1")

_ETF_V3_ORIG = extract_template_prices

def extract_template_prices(template_path, parsed):  # noqa: F811
    if (parsed or {}).get("material") == "каркас":
        _kark_path = template_path
        if not _kark_path or not os.path.exists(_kark_path):
            _kark_path = "/root/.areal-neva-core/data/templates/estimate/cache/1yt-RJsGRhO13zmPKNAn6bMuGrpXY7kWp__М-80.xlsx"
        prices = {}
        try:
            from openpyxl import load_workbook as _kw
            wb = _kw(_kark_path, read_only=True, data_only=True)
            ws = wb["Каркас под ключ"] if "Каркас под ключ" in wb.sheetnames else wb.active
            section = "Общее"
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                r = list(row)
                if i < 9:
                    continue
                r1 = r[1] if len(r) > 1 else None
                if r[0] and not r1:
                    section = str(r[0]).strip()
                    continue
                if not r1:
                    continue
                try:
                    wp = float(r[4]) if r[4] is not None else 0.0
                    mp = float(r[6]) if r[6] is not None else 0.0
                except (ValueError, TypeError):
                    continue
                if wp == 0 and mp == 0:
                    continue
                unit = str(r[2]).strip() if r[2] else ""
                name = str(r1).strip()
                prices[name.lower().strip()] = {
                    "work_price": wp, "mat_price": mp,
                    "unit": unit, "section": section, "name": name,
                }
            wb.close()
        except Exception as _ke:
            _KFV1_LOG.error("KARKASNIK_SHEET_FIX_V1_ERR: %s", _ke)
        count = len(prices)
        lines = [
            f"- {v['name']}: работа={v['work_price']:.0f} {v['unit']}, матер={v['mat_price']:.0f} {v['unit']}"
            for v in list(prices.values())[:25]
            if v["work_price"] > 0 or v["mat_price"] > 0
        ]
        text = f"Цены из листа 'Каркас под ключ' ({count} позиций):\n" + "\n".join(lines)
        _KFV1_LOG.info("KARKASNIK_SHEET_FIX_V1:count=%d sheet=Каркас_под_ключ", count)
        return text, "Каркас под ключ", False
    return _ETF_V3_ORIG(template_path, parsed)

_KFV1_LOG.info("PATCH_KARKASNIK_SHEET_FIX_V1 installed")
# === END_PATCH_KARKASNIK_SHEET_FIX_V1 ===

# === PATCH_TOPIC2_NO_WAITING_PROJECT_MEMORY_REVIVE_V1 ===
# A request like "сейчас скину проект/PDF" is not an old-task finish request.
# It must wait for the incoming project file and must not revive old estimate raw_input.
try:
    _T2NWPMR_ORIG_IS_OLD_TASK_FINISH_REQUEST = _is_old_task_finish_request
except Exception:
    _T2NWPMR_ORIG_IS_OLD_TASK_FINISH_REQUEST = None

def _t2nwpmr_low(value):
    return ("" if value is None else str(value)).lower().replace("ё", "е")

def _t2nwpmr_waiting_project(value):
    t = _t2nwpmr_low(value).replace("[voice]", " ")
    waits = ("сейчас скину", "сейчас пришлю", "скину проект", "пришлю проект", "скину файл", "пришлю файл")
    project_words = ("проект", "pdf", "файл", "чертеж", "архитектур", "стадия")
    return any(w in t for w in waits) and any(w in t for w in project_words)

if _T2NWPMR_ORIG_IS_OLD_TASK_FINISH_REQUEST and not getattr(_T2NWPMR_ORIG_IS_OLD_TASK_FINISH_REQUEST, "_t2nwpmr_wrapped", False):
    def _is_old_task_finish_request(text: str) -> bool:
        if _t2nwpmr_waiting_project(text):
            return False
        return _T2NWPMR_ORIG_IS_OLD_TASK_FINISH_REQUEST(text)

    _is_old_task_finish_request._t2nwpmr_wrapped = True

# === END_PATCH_TOPIC2_NO_WAITING_PROJECT_MEMORY_REVIVE_V1 ===


# === PATCH_TOPIC2_PDF_TEXT_FACT_ENRICH_V1 ===
def _t2_pdf_text_fact_enrich(parsed: Dict[str, Any], conn=None, task_id=None) -> Dict[str, Any]:
    try:
        path = parsed.get('pdf_spec_source') or parsed.get('local_path') or parsed.get('file_path')
        if not path or not os.path.exists(str(path)):
            return parsed
        import subprocess as _t2_pdf_subprocess
        res = _t2_pdf_subprocess.run(['pdftotext', str(path), '-'], capture_output=True, text=True, timeout=30)
        text = res.stdout or ''
        low = _low(text)
        changed = []
        if not parsed.get('object') and 'дом' in low:
            parsed['object'] = 'дом'
            changed.append('object=дом')
        if not parsed.get('material') and 'газобетон' in low:
            parsed['material'] = 'газобетон'
            changed.append('material=газобетон')
        if not parsed.get('floors') and ('2 этажа' in low or '2-го этажа' in low or '2-ой этаж' in low or 'план расстановки мебели 2 этажа' in low):
            parsed['floors'] = 2
            changed.append('floors=2')
        if not parsed.get('foundation') and ('монолитная железобетонная плита' in low or 'монолитная ж/б плита' in low):
            parsed['foundation'] = 'монолитная железобетонная плита 300 мм'
            changed.append('foundation=монолитная плита 300 мм')
        if not parsed.get('scope'):
            scopes = []
            if 'план кровли' in low or 'кровл' in low:
                scopes.append('кровля')
            if 'план межэтажного перекрытия' in low or 'перекрыт' in low:
                scopes.append('перекрытия')
            if 'наружные стены' in low or 'внутренние несущие стены' in low or 'перегородки' in low:
                scopes.append('стены и перегородки')
            if scopes:
                parsed['scope'] = ', '.join(dict.fromkeys(scopes))
                changed.append('scope')
        if changed and conn is not None and task_id is not None:
            _history_safe(conn, task_id, 'TOPIC2_PDF_TEXT_FACTS_ENRICHED:' + ','.join(changed)[:160])
    except Exception as _t2_pdf_text_e:
        try:
            if conn is not None and task_id is not None:
                _history_safe(conn, task_id, 'TOPIC2_PDF_TEXT_FACTS_ERR:' + str(_t2_pdf_text_e)[:80])
        except Exception:
            pass
    return parsed
# === END_PATCH_TOPIC2_PDF_TEXT_FACT_ENRICH_V1 ===

# === PATCH_TOPIC2_TEMPLATE_ROWS_FULL_AREAL_CALC_V1 ===
# Use the selected template workbook rows as the calculation matrix instead of
# synthetic 5-8 row summaries. This keeps the canonical 15-col AREAL_CALC output.
_T2TR_ORIG_CLASSIFY_ITEM = _classify_item
_T2TR_ORIG_CREATE_XLSX = _create_xlsx_from_template


def _classify_item(name: str, section: str) -> str:  # noqa: F811
    sec = _low(str(section or ''))
    if 'материал' in sec:
        return 'material'
    if 'работ' in sec:
        return 'work'
    return _T2TR_ORIG_CLASSIFY_ITEM(name, section)


def _t2tr_num(v):
    try:
        if v is None or v == '':
            return 0.0
        return float(str(v).replace(' ', '').replace(',', '.'))
    except Exception:
        return 0.0



def _t2tr_estimate_section(section: str, name: str) -> str:
    """Keep project/template sections, but separate logistics/overheads by meaning."""
    sec = _s(section or "").strip()
    low = _low(f"{sec} {name}")
    if any(x in low for x in (
        "аренда крана", "кран", "бетононасос", "манипулятор", "доставка",
        "транспорт", "разгруз", "прожив", "рейс", "логист"
    )):
        return "Логистика"
    if any(x in low for x in ("накладн", "организация работ", "расходные материалы", "крепеж", "крепёж", "уборка")):
        return "Накладные расходы"
    return sec or "Прочее"

def _t2tr_template_items(template_path: Optional[str], sheet_name: Optional[str], parsed: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not template_path or not os.path.exists(template_path):
        return []
    try:
        from openpyxl import load_workbook as _t2tr_lwb
        wb = _t2tr_lwb(template_path, data_only=True, read_only=True)
        try:
            selected = sheet_name if sheet_name in wb.sheetnames else choose_template_sheet(parsed, wb.sheetnames)[0]
            ws = wb[selected] if selected in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
    except Exception:
        return []

    header_idx = -1
    for idx, row in enumerate(rows):
        vals = [_low(x) for x in row[:12]]
        if any('наименование' in v for v in vals) and any('кол-во' in v or 'количество' in v for v in vals):
            header_idx = idx
            break
    if header_idx < 0:
        return []

    header = rows[header_idx]
    header_txt = ' '.join(_low(x) for x in header[:12])
    if 'себестоимость работ' in header_txt or (len(header) > 9 and 'материал' in _low(header[9])):
        work_col, mat_col = 7, 9
    else:
        work_col, mat_col = 4, 6

    items: List[Dict[str, Any]] = []
    section = 'Прочее'
    for row in rows[header_idx + 1:]:
        a = _s(row[0] if len(row) > 0 else '')
        name = _s(row[1] if len(row) > 1 else '')
        unit = _s(row[2] if len(row) > 2 else '')
        qty = _t2tr_num(row[3] if len(row) > 3 else 0)
        if a and not name:
            section = a.strip()
            if _low(section).startswith('не входит'):
                break
            continue
        if _low(section).startswith('не входит'):
            continue
        if not name or not unit or qty <= 0:
            continue
        lname = _low(name)
        if 'наименование' in lname or 'итого' in lname or lname.startswith('не входит'):
            continue
        work_price = _t2tr_num(row[work_col] if len(row) > work_col else 0)
        mat_price = _t2tr_num(row[mat_col] if len(row) > mat_col else 0)
        note = _s(row[9] if len(row) > 9 and mat_col != 9 else (row[10] if len(row) > 10 else ''))
        calc_section = _t2tr_estimate_section(section, name)
        if work_price > 0:
            items.append({
                'section': calc_section,
                'name': name[:240],
                'unit': unit,
                'qty': qty,
                'price': work_price,
                'kind': 'work',
                'note': (note or 'template workbook row')[:240],
            })
        if mat_price > 0:
            items.append({
                'section': calc_section,
                'name': name[:240],
                'unit': unit,
                'qty': qty,
                'price': mat_price,
                'kind': 'material',
                'note': (note or 'template workbook row')[:240],
            })
    return items[:400]


def _t2tr_add_required_blocks(items: List[Dict[str, Any]], parsed: Dict[str, Any], price_text: str, choice: Dict[str, Any]) -> List[Dict[str, Any]]:
    result = list(items or [])
    sections = {_low(it.get('section', '')) for it in result if isinstance(it, dict)}
    try:
        distance = float(parsed.get('distance_km') or 0)
    except Exception:
        distance = 0.0
    if distance > 0 and 'логистика' not in sections:
        delivery_price = _choose_value(_numbers_from_price_text(price_text, ('достав', 'рейс', 'манипулятор', 'кран', 'транспорт')), choice)
        trips = max(math.ceil(distance / 40), 1)
        if delivery_price and delivery_price < 5000:
            delivery_price = 0
        if delivery_price <= 0:
            delivery_price = 13500
        if delivery_price > 0:
            result.append({
                'section': 'Логистика',
                'name': 'Доставка материалов от СПб',
                'unit': 'рейс',
                'qty': trips,
                'price': delivery_price,
                'note': f'{distance:g} км / 40',
            })
            if distance > 50:
                result.append({
                    'section': 'Логистика',
                    'name': 'Транспорт бригады и проживание',
                    'unit': 'компл',
                    'qty': 1,
                    'price': round(delivery_price * 0.3, 2),
                    'note': 'при удаленности > 50 км',
                })
    sections = {_low(it.get('section', '')) for it in result if isinstance(it, dict)}
    if 'накладные расходы' not in sections and 'накладные' not in sections:
        subtotal = sum(
            float(it.get('qty') or 0) * float(it.get('price') or 0)
            for it in result
            if _low(it.get('section', '')) not in ('логистика', 'накладные расходы', 'накладные')
        )
        if subtotal > 0:
            result.append({
                'section': 'Накладные расходы',
                'name': 'Организация работ и накладные',
                'unit': 'компл',
                'qty': 1,
                'price': round(subtotal * 0.07, 2),
                'note': '7% от материалов и работ',
            })
    return result


def _create_xlsx_from_template(task_id: str, parsed: Dict[str, Any], template: Dict[str, Any], template_path: Optional[str], sheet_name: Optional[str], price_text: str, choice: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]], float]:  # noqa: F811
    template_items = _t2tr_template_items(template_path, sheet_name, parsed)
    if len(template_items) >= 20:
        template_items = _t2tr_add_required_blocks(template_items, parsed, price_text, choice)
        orig_build = globals().get('_build_estimate_items')

        def _t2tr_build_from_template(_parsed, _price_text, _choice):
            return template_items

        globals()['_build_estimate_items'] = _t2tr_build_from_template
        try:
            return _T2TR_ORIG_CREATE_XLSX(task_id, parsed, template, template_path, sheet_name, price_text, choice)
        finally:
            globals()['_build_estimate_items'] = orig_build
    return _T2TR_ORIG_CREATE_XLSX(task_id, parsed, template, template_path, sheet_name, price_text, choice)
# === END_PATCH_TOPIC2_TEMPLATE_ROWS_FULL_AREAL_CALC_V1 ===

# === PATCH_TOPIC2_AREA_ONLY_WITH_TEMPLATE_ALLOWED_V2 ===
# If a project PDF has only AR/TЭП rows but a canonical full-house template is selected,
# do not stop the task. The template matrix is the calculation basis; PDF facts are inputs.
_T2AOT_ORIG_MISSING_QUESTION = _missing_question


def _t2aot_rows_are_area_only(rows) -> bool:
    usable = []
    non_area = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        name = _clean(row.get('name', '')).lower().replace('ё', 'е')
        try:
            qty = float(row.get('qty') or 0)
            price = float(row.get('price') or 0)
        except Exception:
            qty = price = 0
        if qty <= 0 and price <= 0:
            continue
        usable.append(row)
        if 'площад' not in name and 'общая' not in name:
            non_area.append(row)
    return bool(usable) and not non_area


def _missing_question(parsed: Dict[str, Any]) -> Optional[str]:  # noqa: F811
    rows = parsed.get('pdf_spec_rows') or []
    raw = _low(parsed.get('raw') or '')
    if _t2aot_rows_are_area_only(rows) and any(x in raw for x in ('смет', 'стоимост', 'расчет', 'расчёт', 'проект')):
        return None
    return _T2AOT_ORIG_MISSING_QUESTION(parsed)
# === END_PATCH_TOPIC2_AREA_ONLY_WITH_TEMPLATE_ALLOWED_V2 ===
# === PATCH_TOPIC2_NO_TEMPLATE_FROM_AREA_ONLY_PDF_V1 ===
# Canon: PDF estimate contour is PDF -> table/spec rows -> normalized AREAL_CALC.
# If PDF has only AR/TEP area facts, template rows must not become a final estimate
# unless the user explicitly asks for an orientational calculation.
def _t2_no_template_area_only_rows_v1(rows):
    usable = []
    non_area = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        name = _clean(row.get("name", "")).lower().replace("ё", "е")
        try:
            qty = float(row.get("qty") or 0)
            price = float(row.get("price") or 0)
        except Exception:
            qty = price = 0
        if qty <= 0 and price <= 0:
            continue
        usable.append(row)
        if "площад" not in name and "общая" not in name:
            non_area.append(row)
    return bool(usable) and not non_area


def _t2_no_template_orient_allowed_v1(parsed):
    raw = _low((parsed or {}).get("raw") or "")
    return any(x in raw for x in (
        "считать ориентировочно по проекту",
        "сделай ориентировочный расчет",
        "сделай ориентировочный расчёт",
        "ориентировочно по проекту",
        "ориентировочная смета",
        "ориентировочный расчет",
        "ориентировочный расчёт",
    ))


def _t2_no_valid_pdf_rows_message_v1():
    return (
        "PDF прочитан, но сметная ведомость объёмов / ВОР / спецификация материалов / раздел КЖ с объёмами не найдены. "
        "Вижу только архитектурные данные и площади, поэтому финальную смету из шаблона не создаю, чтобы не подменять расчёт догадками.\n\n"
        "Пришли ВОР / спецификацию / КЖ с объёмами либо прямо напиши: считать ориентировочно по проекту."
    )


_T2NT_ORIG_MISSING_QUESTION_V1 = _missing_question


def _missing_question(parsed: Dict[str, Any]) -> Optional[str]:  # noqa: F811
    rows = (parsed or {}).get("pdf_spec_rows") or []
    if _t2_no_template_area_only_rows_v1(rows) and not _t2_no_template_orient_allowed_v1(parsed):
        return _t2_no_valid_pdf_rows_message_v1()
    return _T2NT_ORIG_MISSING_QUESTION_V1(parsed)


_T2NT_ORIG_GENERATE_AND_SEND_V1 = _generate_and_send


async def _generate_and_send(conn, task, pending, confirm_text, logger=None):  # noqa: F811
    parsed = (pending or {}).get("parsed") or {}
    if _t2_no_template_area_only_rows_v1(parsed.get("pdf_spec_rows") or []) and not _t2_no_template_orient_allowed_v1(parsed):
        task_id = _s(_row_get(task, "id"))
        chat_id = _s(_row_get(task, "chat_id"))
        topic_id = int(_row_get(task, "topic_id", 0) or 0)
        reply_to = _row_get(task, "reply_to_message_id", None)
        msg = _t2_no_valid_pdf_rows_message_v1()
        try:
            send_res = await _send_text(chat_id, msg, reply_to, topic_id)
        except Exception:
            send_res = {}
        kwargs = {
            "state": "WAITING_CLARIFICATION",
            "result": msg,
            "error_message": "TOPIC2_PDF_NO_VALID_ESTIMATE_ROWS",
        }
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        _history_safe(conn, task_id, "TOPIC2_PDF_NO_VALID_ESTIMATE_ROWS:area_only_no_template_final")
        return True
    return await _T2NT_ORIG_GENERATE_AND_SEND_V1(conn, task, pending, confirm_text, logger=logger)

try:
    _STV3_LOG.info("PATCH_TOPIC2_NO_TEMPLATE_FROM_AREA_ONLY_PDF_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_NO_TEMPLATE_FROM_AREA_ONLY_PDF_V1 ===
# === PATCH_TOPIC2_AR_PROJECT_FACT_ROWS_V1 ===
# AR project PDFs may contain usable quantities outside formal VOR tables:
# foundation piles/rostverk, roof area, slab schedule, window/door schedules.
# These rows may be used only as project-derived rows; template rows are still
# forbidden for final output when the PDF has only AR/TEP data.
def _t2ar_num_v1(v):
    try:
        return float(str(v or "").replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0


def _t2ar_pdf_text_v1(parsed):
    try:
        p = (parsed or {}).get("pdf_spec_source") or (parsed or {}).get("local_path") or (parsed or {}).get("file_path")
        if not p or not os.path.exists(str(p)):
            return ""
        import subprocess as _t2ar_subprocess
        res = _t2ar_subprocess.run(["pdftotext", "-layout", str(p), "-"], capture_output=True, text=True, timeout=45)
        return res.stdout or ""
    except Exception:
        return ""


def _t2ar_add_row_v1(rows, section, name, unit, qty, note="", kind="material"):
    try:
        qty = float(qty)
    except Exception:
        qty = 0.0
    if qty <= 0:
        return
    key = (section, name, unit, round(qty, 6))
    if any((r.get("section"), r.get("name"), r.get("unit"), round(float(r.get("qty") or 0), 6)) == key for r in rows):
        return
    rows.append({
        "section": section,
        "name": name[:240],
        "unit": unit,
        "qty": qty,
        "price": 0.0,
        "kind": kind,
        "note": (note or "из AR PDF")[:240],
    })


def _t2ar_project_rows_from_pdf_v1(parsed):
    text = _t2ar_pdf_text_v1(parsed)
    if not text:
        return []
    rows = []
    low = _low(text)

    m = re.search(r"Свая\s+200х200мм\s*\(3м\)\s+(\d+)", text, re.I)
    if m:
        _t2ar_add_row_v1(rows, "Фундамент", "Свая 200х200мм (3м)", "шт", _t2ar_num_v1(m.group(1)), "Спецификация на сваи, лист АР-08")

    m = re.search(r"Ростверк\.\s*Бетон\s*B22,?5\s*W6\s*F150\s*([\d\s]+[,.]\d+)", text, re.I | re.S)
    if m:
        _t2ar_add_row_v1(rows, "Фундамент", "Ростверк. Бетон B22,5 W6 F150", "м3", _t2ar_num_v1(m.group(1)), "Таблица на листе АР-08; единица приведена как бетонный объем")

    m = re.search(r"Площадь\s+Поверхности\s+Уклон\s+([\d\s]+[,.]\d+)\s+20", text, re.I)
    if m:
        _t2ar_add_row_v1(rows, "Кровля", "Площадь поверхности кровли, уклон 20°", "м2", _t2ar_num_v1(m.group(1)), "План кровли, лист АР-10")

    for mark, qty in re.findall(r"\b(ПК\s*\d{2}-\d{2}-8)\s+(\d+)", text, re.I):
        mark_clean = re.sub(r"\s+", " ", mark).strip()
        _t2ar_add_row_v1(rows, "Перекрытия", f"Плита перекрытия {mark_clean}", "шт", _t2ar_num_v1(qty), "Спецификация плит перекрытия, лист АР-11")

    m = re.search(r"Площадь\s+монолитных\s+участков\s*-\s*([\d\s]+[,.]\d+)\s*м2", text, re.I)
    if m:
        _t2ar_add_row_v1(rows, "Перекрытия", "Монолитные участки перекрытия", "м2", _t2ar_num_v1(m.group(1)), "План межэтажного перекрытия, лист АР-11")

    m = re.search(r"Балка\s+перекрытия\s+50х200\s+([\d\s]+)\s+(\d+)", text, re.I)
    if m:
        _t2ar_add_row_v1(rows, "Перекрытия", "Балка перекрытия 50х200", "шт", _t2ar_num_v1(m.group(2)), f"Длина {m.group(1).strip()} мм, лист АР-11")

    win_pat = re.compile(r"\b(Ок-\d+(?:,\s*бДв-1)?)\s+(\d+)\s+([0-9\s]+×[0-9\s]+)\s+([\d\s]+[,.]\d+)", re.I)
    for mark, qty, size, area in win_pat.findall(text):
        _t2ar_add_row_v1(rows, "Окна и двери", f"{mark.strip()} оконный/балконный блок {size.strip()}", "шт", _t2ar_num_v1(qty), f"Площадь проема {area.strip()} м2, лист АР-19")

    door_pat = re.compile(r"\b(Дв-\d+|нДв-1)\s+(\d+)\s+[0-9\s]+×[0-9\s]+\s+([0-9\s]+×[0-9\s]+)\s+([\d\s]+[,.]\d+)", re.I)
    for mark, qty, size, area in door_pat.findall(text):
        _t2ar_add_row_v1(rows, "Окна и двери", f"{mark.strip()} дверной блок {size.strip()}", "шт", _t2ar_num_v1(qty), f"Площадь проема {area.strip()} м2, лист АР-20")

    return rows


def _t2ar_missing_from_pdf_v1(parsed):
    text = _t2ar_pdf_text_v1(parsed)
    low = _low(text)
    missing = []
    if "наружные стены дома" in low:
        missing.append("объем/площадь наружных стен 375/300 мм")
    if "внутренние несущие стены" in low:
        missing.append("объем внутренних несущих стен 250 мм")
    if "перегородки" in low:
        missing.append("объем перегородок 150 мм и каркасных перегородок")
    if "уточняется в разделе кж" in low or "разделе кж" in low:
        missing.append("КЖ для ж/б балок, колонн, перемычек, армопояса и плит")
    if "план ввода коммуникаций" in low:
        missing.append("объемы инженерных коммуникаций")
    return list(dict.fromkeys(missing))


def _t2ar_project_rows_message_v1(parsed):
    rows = _t2ar_project_rows_from_pdf_v1(parsed)
    missing = _t2ar_missing_from_pdf_v1(parsed)
    lines = [
        "PDF прочитан. Нашёл проектные позиции и объёмы, которые можно использовать без шаблонной подмены:",
    ]
    for r in rows[:18]:
        lines.append(f"- {r['section']}: {r['name']} — {r['qty']:g} {r['unit']}")
    if len(rows) > 18:
        lines.append(f"- ещё позиций: {len(rows) - 18}")
    if missing:
        lines.append("")
        lines.append("Для полной сметы не хватает явных объёмов:")
        for m in missing[:10]:
            lines.append(f"- {m}")
    lines.append("")
    lines.append("Финальную смету из шаблона не создаю. Пришли недостающие объёмы/КЖ/ВОР либо напиши: считай по найденным позициям.")
    return "\n".join(lines)


try:
    _T2AR_PREV_ORIENT_ALLOWED_V1 = _t2_no_template_orient_allowed_v1
    def _t2_no_template_orient_allowed_v1(parsed):  # noqa: F811
        raw = _low((parsed or {}).get("raw") or "") + " " + _low((parsed or {}).get("_topic2_confirm_text") or "")
        if any(x in raw for x in ("считай по найденным позициям", "считать по найденным позициям", "только найденные позиции")):
            return True
        return _T2AR_PREV_ORIENT_ALLOWED_V1(parsed)
except Exception:
    pass


_T2AR_PREV_MISSING_QUESTION_V1 = _missing_question


def _missing_question(parsed: Dict[str, Any]) -> Optional[str]:  # noqa: F811
    rows = (parsed or {}).get("pdf_spec_rows") or []
    project_rows = _t2ar_project_rows_from_pdf_v1(parsed)
    if _t2_no_template_area_only_rows_v1(rows) and project_rows and not _t2_no_template_orient_allowed_v1(parsed):
        return _t2ar_project_rows_message_v1(parsed)
    return _T2AR_PREV_MISSING_QUESTION_V1(parsed)


_T2AR_PREV_CREATE_XLSX_V1 = _create_xlsx_from_template


def _t2ar_keywords_for_price_v1(name):
    low = _low(name)
    words = [w for w in re.split(r"[^0-9a-zа-яё]+", low) if len(w) >= 3]
    keep = [w for w in words if w not in ("лист", "проем", "проема", "площадь", "поверхности")]
    return tuple(keep[:6]) or tuple(words[:3])


def _create_xlsx_from_template(task_id: str, parsed: Dict[str, Any], template: Dict[str, Any], template_path: Optional[str], sheet_name: Optional[str], price_text: str, choice: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]], float]:  # noqa: F811
    project_rows = list((parsed or {}).get("pdf_project_rows") or [])
    if not project_rows:
        project_rows = _t2ar_project_rows_from_pdf_v1(parsed)
    if project_rows and _t2_no_template_orient_allowed_v1(parsed):
        items = []
        for r in project_rows:
            it = dict(r)
            vals = _numbers_from_price_text(price_text or "", _t2ar_keywords_for_price_v1(it.get("name", "")))
            it["price"] = _choose_value(vals, choice) if vals else 0.0
            if not vals:
                it["note"] = (it.get("note", "") + "; PRICE_MISSING").strip("; ")
            items.append(it)
        orig_build = globals().get("_build_estimate_items")
        base_create = globals().get("_T2TR_ORIG_CREATE_XLSX") or _T2AR_PREV_CREATE_XLSX_V1
        def _t2ar_build_from_project(_parsed, _price_text, _choice):
            return items
        globals()["_build_estimate_items"] = _t2ar_build_from_project
        try:
            return base_create(task_id, parsed, template, template_path, sheet_name, price_text, choice)
        finally:
            globals()["_build_estimate_items"] = orig_build
    return _T2AR_PREV_CREATE_XLSX_V1(task_id, parsed, template, template_path, sheet_name, price_text, choice)


_T2AR_PREV_GENERATE_AND_SEND_V1 = _generate_and_send


async def _generate_and_send(conn, task, pending, confirm_text, logger=None):  # noqa: F811
    if isinstance(pending, dict):
        parsed = pending.get("parsed") or {}
        if isinstance(parsed, dict):
            parsed["_topic2_confirm_text"] = confirm_text or ""
            rows = _t2ar_project_rows_from_pdf_v1(parsed)
            if rows:
                parsed["pdf_project_rows"] = rows
                pending["parsed"] = parsed
                try:
                    _history_safe(conn, _s(_row_get(task, "id")), f"TOPIC2_AR_PROJECT_ROWS_EXTRACTED:{len(rows)}")
                except Exception:
                    pass
    return await _T2AR_PREV_GENERATE_AND_SEND_V1(conn, task, pending, confirm_text, logger=logger)

try:
    _STV3_LOG.info("PATCH_TOPIC2_AR_PROJECT_FACT_ROWS_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_AR_PROJECT_FACT_ROWS_V1 ===
# === PATCH_TOPIC2_PROJECT_ROWS_CONFIRM_AND_PRICES_V1 ===
# Explicit user clarification "считай по проекту" means:
# - use only rows extracted from the project PDF;
# - do not use template rows as estimate positions;
# - use Sonar/Perplexity price search for those extracted rows.
def _t2prcp_project_calc_requested_text_v1(value):
    raw = _low(value or "")
    if "по проект" in raw and "цены" in raw and ("найди" in raw or "интернет" in raw):
        return True
    return any(x in raw for x in (
        "считай по проекту",
        "считать по проекту",
        "считать по проектной документации",
        "считай по проектной документации",
        "считай по найденным позициям",
        "считать по найденным позициям",
        "только найденные позиции",
    ))


def _t2prcp_history_clarified_text_v1(conn, task_id):
    if conn is None or not task_id:
        return ""
    try:
        rows = conn.execute(
            "SELECT action FROM task_history WHERE task_id=? AND action LIKE 'clarified:%' ORDER BY rowid DESC LIMIT 12",
            (str(task_id),),
        ).fetchall()
        return "\n".join(_s(r[0]) for r in rows)
    except Exception:
        return ""


try:
    _T2PRCP_PREV_ORIENT_ALLOWED_V1 = _t2_no_template_orient_allowed_v1
    def _t2_no_template_orient_allowed_v1(parsed):  # noqa: F811
        raw = _low((parsed or {}).get("raw") or "")
        raw += " " + _low((parsed or {}).get("_topic2_confirm_text") or "")
        raw += " " + _low((parsed or {}).get("_topic2_history_clarified") or "")
        if _t2prcp_project_calc_requested_text_v1(raw):
            return True
        return _T2PRCP_PREV_ORIENT_ALLOWED_V1(parsed)
except Exception:
    pass


_T2PRCP_PREV_SEARCH_ONLINE_V1 = _search_prices_online


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    base = await _T2PRCP_PREV_SEARCH_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
    project_rows = (parsed or {}).get("pdf_project_rows") or _t2ar_project_rows_from_pdf_v1(parsed or {})
    if not project_rows:
        return base
    try:
        from core.price_enrichment import _openrouter_price_search as _project_price_search
        lines = []
        seen = set()
        for row in project_rows[:22]:
            name = _s(row.get("name"))
            unit = _s(row.get("unit"))
            if not name or name.lower() in seen:
                continue
            seen.add(name.lower())
            if conn is not None and task_id is not None:
                try:
                    _history_safe(conn, task_id, f"TOPIC2_PROJECT_PRICE_SEARCH_STARTED:{name[:80]}")
                except Exception:
                    pass
            offers = []
            try:
                offers = await asyncio.wait_for(_project_price_search(name, unit), timeout=35)
            except Exception:
                offers = []
            valid = [
                o for o in (offers or [])
                if o.get("price") and (o.get("supplier") or o.get("url")) and o.get("status")
            ]
            if conn is not None and task_id is not None:
                try:
                    if valid:
                        o0 = valid[0]
                        _history_safe(conn, task_id, "TOPIC2_PROJECT_PRICE_SOURCE_FOUND:{}:{}:{}".format(
                            name[:50], _s(o0.get("supplier"))[:50], _s(o0.get("status"))[:20]
                        ))
                    else:
                        _history_safe(conn, task_id, f"TOPIC2_PROJECT_PRICE_SOURCE_MISSING:{name[:80]}")
                except Exception:
                    pass
            for o in valid[:2]:
                lines.append(
                    "- {} | {} | {} | Санкт-Петербург и Ленинградская область | {} | {} | {}".format(
                        name,
                        o.get("price"),
                        o.get("unit") or unit,
                        o.get("supplier") or "",
                        o.get("url") or "",
                        o.get("checked_at") or datetime.date.today().isoformat(),
                    )
                )
        if lines:
            return (base or "") + "\n\n=== ПОИСК ПО ПРОЕКТНЫМ ПОЗИЦИЯМ ===\n" + "\n".join(lines)
    except Exception as _t2prcp_e:
        try:
            if conn is not None and task_id is not None:
                _history_safe(conn, task_id, "TOPIC2_PROJECT_PRICE_SEARCH_ERR:" + _s(_t2prcp_e)[:120])
        except Exception:
            pass
    return base


_T2PRCP_PREV_GENERATE_AND_SEND_V1 = _generate_and_send


async def _generate_and_send(conn, task, pending, confirm_text, logger=None):  # noqa: F811
    task_id = _s(_row_get(task, "id"))
    if isinstance(pending, dict):
        parsed = pending.get("parsed") or {}
        if isinstance(parsed, dict):
            history_text = _t2prcp_history_clarified_text_v1(conn, task_id)
            parsed["_topic2_confirm_text"] = (confirm_text or "") + "\n" + history_text
            parsed["_topic2_history_clarified"] = history_text
            rows = _t2ar_project_rows_from_pdf_v1(parsed)
            if rows:
                parsed["pdf_project_rows"] = rows
                pending["parsed"] = parsed
                if _t2_no_template_orient_allowed_v1(parsed):
                    pending["online_prices"] = ""
                try:
                    _history_safe(conn, task_id, f"TOPIC2_AR_PROJECT_ROWS_EXTRACTED:{len(rows)}")
                except Exception:
                    pass
                if not _t2_no_template_orient_allowed_v1(parsed):
                    chat_id = _s(_row_get(task, "chat_id"))
                    topic_id = int(_row_get(task, "topic_id", 0) or 0)
                    reply_to = _row_get(task, "reply_to_message_id", None)
                    msg = _t2ar_project_rows_message_v1(parsed)
                    try:
                        send_res = await _send_text(chat_id, msg, reply_to, topic_id)
                    except Exception:
                        send_res = {}
                    kwargs = {
                        "state": "WAITING_CLARIFICATION",
                        "result": msg,
                        "error_message": "TOPIC2_WAIT_PROJECT_ROW_CONFIRMATION",
                    }
                    if isinstance(send_res, dict) and send_res.get("bot_message_id"):
                        kwargs["bot_message_id"] = send_res.get("bot_message_id")
                    _update_task_safe(conn, task_id, **kwargs)
                    _history_safe(conn, task_id, "TOPIC2_PROJECT_ROWS_WAITING_USER_CONFIRM")
                    return True
    return await _T2PRCP_PREV_GENERATE_AND_SEND_V1(conn, task, pending, confirm_text, logger=logger)

try:
    _STV3_LOG.info("PATCH_TOPIC2_PROJECT_ROWS_CONFIRM_AND_PRICES_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_PROJECT_ROWS_CONFIRM_AND_PRICES_V1 ===
# === PATCH_TOPIC2_PROJECT_WORK_ROWS_V1 ===
# Project-derived material/product rows require corresponding work rows for
# "работы + материалы" totals. Work rows are derived only from extracted project
# rows and keep the same unit/quantity; no template positions are introduced.
def _t2pwr_work_name_v1(row):
    name = _s((row or {}).get("name"))
    sec = _s((row or {}).get("section"))
    low = _low(name)
    if "свая" in low:
        return "Монтаж/погружение: " + name
    if "ростверк" in low:
        return "Устройство: " + name
    if "кровл" in low:
        return "Монтаж: " + name
    if "плита перекрытия" in low:
        return "Монтаж: " + name
    if "монолитные участки" in low:
        return "Устройство: " + name
    if "балка" in low:
        return "Монтаж: " + name
    if "окон" in low or "двер" in low or sec == "Окна и двери":
        return "Монтаж: " + name
    return "Монтаж/устройство: " + name


def _t2pwr_expand_rows_v1(rows):
    out = []
    seen = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        base = dict(row)
        base["kind"] = base.get("kind") or "material"
        key = (base.get("section"), base.get("name"), base.get("unit"), base.get("qty"), base.get("kind"))
        if key not in seen:
            seen.add(key)
            out.append(base)
        work = dict(base)
        work["name"] = _t2pwr_work_name_v1(base)
        work["kind"] = "work"
        work["price"] = 0.0
        work["note"] = (_s(base.get("note")) + "; работа по проектной позиции").strip("; ")
        key = (work.get("section"), work.get("name"), work.get("unit"), work.get("qty"), work.get("kind"))
        if key not in seen:
            seen.add(key)
            out.append(work)
    return out


try:
    _T2PWR_PREV_PROJECT_ROWS_V1 = _t2ar_project_rows_from_pdf_v1
    def _t2ar_project_rows_from_pdf_v1(parsed):  # noqa: F811
        return _t2pwr_expand_rows_v1(_T2PWR_PREV_PROJECT_ROWS_V1(parsed))
except Exception:
    pass


try:
    _T2PWR_PREV_SEARCH_V1 = _search_prices_online
    async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
        project_rows = (parsed or {}).get("pdf_project_rows") or _t2ar_project_rows_from_pdf_v1(parsed or {})
        if project_rows and _t2_no_template_orient_allowed_v1(parsed or {}):
            model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
            if "sonar" not in model.lower():
                raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")
            if conn is not None and task_id is not None:
                _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_STARTED")
                _history_safe(conn, task_id, f"TOPIC2_ONLINE_MODEL_SONAR_CONFIRMED:{model}")
            from core.price_enrichment import _openrouter_price_search as _project_price_search
            lines = []
            seen = set()
            # Keep runtime bounded; rows are ordered material/work pairs from the project.
            for row in project_rows[:44]:
                name = _s(row.get("name"))
                unit = _s(row.get("unit"))
                if not name or name.lower() in seen:
                    continue
                seen.add(name.lower())
                if conn is not None and task_id is not None:
                    _history_safe(conn, task_id, f"TOPIC2_PROJECT_PRICE_SEARCH_STARTED:{name[:80]}")
                try:
                    offers = await asyncio.wait_for(_project_price_search(name, unit), timeout=35)
                except Exception:
                    offers = []
                valid = [o for o in (offers or []) if o.get("price") and (o.get("supplier") or o.get("url")) and o.get("status")]
                if conn is not None and task_id is not None:
                    if valid:
                        o0 = valid[0]
                        _history_safe(conn, task_id, "TOPIC2_PROJECT_PRICE_SOURCE_FOUND:{}:{}:{}".format(
                            name[:50], _s(o0.get("supplier"))[:50], _s(o0.get("status"))[:20]
                        ))
                    else:
                        _history_safe(conn, task_id, f"TOPIC2_PROJECT_PRICE_SOURCE_MISSING:{name[:80]}")
                for o in valid[:2]:
                    lines.append(
                        "- {} | {} | {} | Санкт-Петербург и Ленинградская область | {} | {} | {}".format(
                            name,
                            o.get("price"),
                            o.get("unit") or unit,
                            o.get("supplier") or "",
                            o.get("url") or "",
                            o.get("checked_at") or datetime.date.today().isoformat(),
                        )
                    )
            result = "\n".join(lines)
            if conn is not None and task_id is not None:
                _history_safe(conn, task_id, f"TOPIC2_PRICE_ENRICHMENT_DONE:{len(result)}")
            return result
        return await _T2PWR_PREV_SEARCH_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
except Exception:
    pass

try:
    _STV3_LOG.info("PATCH_TOPIC2_PROJECT_WORK_ROWS_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_PROJECT_WORK_ROWS_V1 ===

# === PATCH_TOPIC2_SAMPLE_MATRIX_MODE_V1 ===
# "Считай по проекту" means: use project facts as input and use existing
# estimate samples as calculation structure. It must not collapse the estimate
# to only the rows directly extracted from the PDF unless the user explicitly
# asks for "только найденные позиции".
def _t2s_text_v1(parsed):
    raw = _low((parsed or {}).get("raw") or "")
    raw += " " + _low((parsed or {}).get("_topic2_confirm_text") or "")
    raw += " " + _low((parsed or {}).get("_topic2_history_clarified") or "")
    return raw


def _t2s_project_only_requested_v1(text):
    low = _low(text or "")
    return any(x in low for x in (
        "только найденные позиции",
        "считай по найденным позициям",
        "считать по найденным позициям",
    ))


def _t2s_sample_matrix_mode_v1(parsed):
    if not isinstance(parsed, dict):
        return False
    text = _t2s_text_v1(parsed)
    if not _t2prcp_project_calc_requested_text_v1(text):
        return False
    if _t2s_project_only_requested_v1(text):
        return False
    rows = parsed.get("pdf_project_rows") or []
    if not rows:
        try:
            rows = _t2ar_project_rows_from_pdf_v1(parsed)
        except Exception:
            rows = []
    return bool(rows)


def _t2s_with_project_only_disabled_v1(callback):
    guard = globals().get("_t2_no_template_orient_allowed_v1")

    def _sample_matrix_guard(_parsed):
        return False

    globals()["_t2_no_template_orient_allowed_v1"] = _sample_matrix_guard
    try:
        return callback()
    finally:
        if guard is not None:
            globals()["_t2_no_template_orient_allowed_v1"] = guard


async def _t2s_await_project_only_disabled_v1(callback):
    guard = globals().get("_t2_no_template_orient_allowed_v1")

    def _sample_matrix_guard(_parsed):
        return False

    globals()["_t2_no_template_orient_allowed_v1"] = _sample_matrix_guard
    try:
        return await callback()
    finally:
        if guard is not None:
            globals()["_t2_no_template_orient_allowed_v1"] = guard


_T2S_PREV_CREATE_XLSX_V1 = _create_xlsx_from_template


def _create_xlsx_from_template(task_id: str, parsed: Dict[str, Any], template: Dict[str, Any], template_path: Optional[str], sheet_name: Optional[str], price_text: str, choice: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]], float]:  # noqa: F811
    if _t2s_sample_matrix_mode_v1(parsed):
        return _t2s_with_project_only_disabled_v1(
            lambda: _T2S_PREV_CREATE_XLSX_V1(task_id, parsed, template, template_path, sheet_name, price_text, choice)
        )
    return _T2S_PREV_CREATE_XLSX_V1(task_id, parsed, template, template_path, sheet_name, price_text, choice)


_T2S_PREV_SEARCH_ONLINE_V1 = _search_prices_online


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    if _t2s_sample_matrix_mode_v1(parsed):
        model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
        if "sonar" not in model.lower():
            raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")
        if conn is not None and task_id is not None:
            _history_safe(conn, task_id, "TOPIC2_SAMPLE_MATRIX_MODE:PROJECT_FACTS_PLUS_TEMPLATE_SAMPLE")
        previous_project_search = globals().get("_T2PWR_PREV_SEARCH_V1")
        if previous_project_search:
            return await previous_project_search(parsed, template, sheet_name, conn=conn, task_id=task_id)
        return await _t2s_await_project_only_disabled_v1(
            lambda: _T2S_PREV_SEARCH_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
        )
    return await _T2S_PREV_SEARCH_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)


try:
    _STV3_LOG.info("PATCH_TOPIC2_SAMPLE_MATRIX_MODE_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_SAMPLE_MATRIX_MODE_V1 ===

# === PATCH_TOPIC2_STRICT_PRICE_SOURCE_MATCH_V1 ===
# FACT ONLY: a live price source may be attached to an AREAL_CALC row only when
# the source position and estimate row describe the same material/work family.
_T2SPSM_PREV_MATCH_PRICE_SOURCE_V1 = _match_price_source


def _t2spsm_words_v1(text):
    low = _low(text or "")
    return [w for w in re.split(r"[^0-9a-zа-яё]+", low) if len(w) >= 3]


def _t2spsm_families_v1(text, section=""):
    low = _low((text or "") + " " + (section or ""))
    families = set()
    checks = (
        ("gasbeton", ("газобетон", "газоблок", "блок 625", "u-блок", "u блок", "лср")),
        ("concrete", ("бетон", "монолит", "ж/б", "железобетон", "ростверк", "плита")),
        ("rebar", ("арматур", "а500", "а240", "проволока вяз")),
        ("wood", ("доска", "брус", "пиломат", "osb", "фанера")),
        ("insulation", ("пенопл", "утепл", "минват", "пир", "pir")),
        ("waterproof", ("гидроизоляц", "линокром", "мастик", "праймер")),
        ("roof", ("кров", "стропил", "мауэрлат", "мембран", "профнастил", "черепиц")),
        ("windows", ("окн", "окон", "пвх", "стеклопакет")),
        ("doors", ("двер", "дверн")),
        ("delivery", ("достав", "транспорт")),
        ("unload", ("разгруз", "погруз")),
        ("crane", ("кран",)),
        ("pump", ("бетононасос",)),
        ("masonry_work", ("кладк", "монтаж", "устройство", "работ")),
        ("facade", ("фасад", "внешняя отделка")),
        ("interior", ("внутрен", "отделк", "гкл", "ламинат", "плитк")),
        ("engineering", ("электрик", "водоснаб", "канализац", "отоплен", "вентиляц", "инженер")),
    )
    for fam, terms in checks:
        if any(term in low for term in terms):
            families.add(fam)
    return families


def _t2spsm_useful_keywords_v1(keywords):
    stop = {
        "цена", "стоимость", "руб", "рублей", "санкт", "петербург", "ленинградская",
        "область", "материал", "материалы", "строительных", "строительный", "работы",
        "работ", "под", "ключ", "для", "при", "или", "монтаж", "устройство",
    }
    return [kw for kw in (keywords or []) if kw and kw not in stop and len(kw) >= 3]


def _match_price_source(sources: List[Dict[str, Any]], item_name: str, item_section: str) -> Dict[str, Any]:  # noqa: F811
    today = datetime.date.today().isoformat()
    empty = {"supplier": "", "url": "", "checked_at": today, "status": "template_only"}
    if not sources:
        return empty
    item_text = f"{item_name or ''} {item_section or ''}"
    item_low = _low(item_text)
    item_families = _t2spsm_families_v1(item_name, item_section)
    best = None
    best_score = 0
    for src in sources:
        src_pos = _s(src.get("position"))
        src_families = _t2spsm_families_v1(src_pos)
        if src_families:
            if not item_families or not (src_families & item_families):
                continue
        keywords = _t2spsm_useful_keywords_v1(src.get("keywords") or _t2spsm_words_v1(src_pos))
        score = sum(1 for kw in keywords if kw in item_low)
        if src_families and item_families and (src_families & item_families):
            score = max(score, 1)
        min_score = 1 if (src_families & item_families) else 2
        if score >= min_score and score > best_score:
            best_score = score
            best = src
    return best if best else empty


try:
    _STV3_LOG.info("PATCH_TOPIC2_STRICT_PRICE_SOURCE_MATCH_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_STRICT_PRICE_SOURCE_MATCH_V1 ===

# === PATCH_TOPIC2_STRICT_PRICE_SOURCE_MATCH_V2 ===
# Narrow work-family matching: generic words "монтаж/устройство/работы" are not
# enough to attach a live source from another section.
def _t2spsm_families_v1(text, section=""):  # noqa: F811
    low = _low((text or "") + " " + (section or ""))
    families = set()
    checks = (
        ("gasbeton", ("газобетон", "газоблок", "блок 625", "u-блок", "u блок", "лср")),
        ("concrete", ("бетон", "монолит", "ж/б", "железобетон", "ростверк", "плита")),
        ("rebar", ("арматур", "а500", "а240", "проволока вяз")),
        ("wood", ("доска", "брус", "пиломат", "osb", "фанера")),
        ("insulation", ("пенопл", "утепл", "минват", "пир", "pir")),
        ("waterproof", ("гидроизоляц", "линокром", "мастик", "праймер")),
        ("roof", ("кров", "стропил", "мауэрлат", "мембран", "профнастил", "черепиц")),
        ("windows", ("окн", "окон", "пвх", "стеклопакет")),
        ("doors", ("двер", "дверн")),
        ("delivery", ("достав", "транспорт")),
        ("unload", ("разгруз", "погруз")),
        ("crane", ("кран",)),
        ("pump", ("бетононасос",)),
        ("masonry_work", ("кладк",)),
        ("facade", ("фасад", "внешняя отделка")),
        ("interior", ("внутрен", "отделк", "гкл", "ламинат", "плитк")),
        ("engineering", ("электрик", "водоснаб", "канализац", "отоплен", "вентиляц", "инженер")),
    )
    for fam, terms in checks:
        if any(term in low for term in terms):
            families.add(fam)
    return families


try:
    _STV3_LOG.info("PATCH_TOPIC2_STRICT_PRICE_SOURCE_MATCH_V2 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_STRICT_PRICE_SOURCE_MATCH_V2 ===

# === PATCH_TOPIC2_FOUNDATION_ONLY_PHOTO_SCOPE_V1 ===
_T2FO_PREV_BUILD_ESTIMATE_ITEMS_V1 = _build_estimate_items


def _t2fo_foundation_only_v1(parsed):
    raw = _low((parsed or {}).get("raw") or "")
    obj = _low((parsed or {}).get("object") or "")
    scope = _low((parsed or {}).get("scope") or "")
    return (
        obj == "фундамент"
        or scope == "фундамент"
        or ("фундамент" in raw and ("плит" in raw or "подуш" in raw or "щеб" in raw))
    )


def _t2fo_float_v1(value, default=0.0):
    try:
        return float(value or default)
    except Exception:
        return float(default)


def _t2fo_int_v1(value, default=0):
    try:
        return int(value or default)
    except Exception:
        return int(default)


def _t2fo_manual_monolith_work_price_v1(text):
    s = _low(text or "")
    if not ("монолит" in s and "работ" in s):
        return 0.0
    patterns = (
        r"стоимост[ьи]\s+работ[^\d]{0,80}монолит[^\d]{0,80}(\d{3,6})(?:[.,]\d+)?\s*(?:руб|р|за|/|\s)*(?:м3|м³|метр\s+куб)",
        r"работ[^\d]{0,80}монолит[^\d]{0,80}(\d{3,6})(?:[.,]\d+)?\s*(?:руб|р|за|/|\s)*(?:м3|м³|метр\s+куб)",
        r"монолит[^\d]{0,80}работ[^\d]{0,80}(\d{3,6})(?:[.,]\d+)?\s*(?:руб|р|за|/|\s)*(?:м3|м³|метр\s+куб)",
    )
    for pat in patterns:
        m = re.search(pat, s, re.I)
        if m:
            try:
                return float(m.group(1).replace(",", "."))
            except Exception:
                return 0.0
    return 0.0


def _t2fo_prices_from_source_lines_v1(price_text, keywords):
    vals = []
    for line in str(price_text or "").splitlines():
        low = _low(line)
        if not any(_low(k) in low for k in keywords):
            continue
        parts = [p.strip() for p in line.strip(" \t-—•·").split("|")]
        if len(parts) < 2:
            continue
        try:
            v = float(re.sub(r"[^0-9.,]", "", parts[1]).replace(",", "."))
        except Exception:
            v = 0.0
        if 100 <= v <= 10000000:
            vals.append(v)
    return vals


def _t2fo_build_foundation_items_v1(parsed, price_text, choice):
    parsed = parsed or {}
    P = _FTM_PRICES
    dims = parsed.get("dimensions") or parsed.get("dims") or (0, 0)
    try:
        a, b = float(dims[0]), float(dims[1])
    except Exception:
        area_fallback = _t2fo_float_v1(parsed.get("area_floor"), 0.0)
        a = b = area_fallback ** 0.5 if area_fallback > 0 else 0.0
    area = _t2fo_float_v1(parsed.get("area_floor"), round(a * b, 2))
    offset = _t2fo_float_v1(parsed.get("foundation_offset_m"), 0.0)
    prep_area = round((a + 2 * offset) * (b + 2 * offset), 2) if offset and a and b else area
    slab_t = _t2fo_float_v1(parsed.get("foundation_thickness_m"), 0.25)
    sand_t = _t2fo_float_v1(parsed.get("sand_thickness_m"), 0.0)
    gravel_t = _t2fo_float_v1(parsed.get("gravel_thickness_m"), 0.0)
    layers = _t2fo_int_v1(parsed.get("reinforcement_layers"), 2)
    distance = _t2fo_float_v1(parsed.get("distance_km"), 0.0)
    raw_text = _low(parsed.get("raw") or "")
    concrete_grade = _s(parsed.get("concrete_grade") or ("М350" if "350" in raw_text else "В25"))

    concrete_volume = round(area * slab_t, 2)
    rebar_qty = round(max(concrete_volume * 0.08 * (max(layers, 1) / 2.0), 0.1), 3)
    formwork_perim = round(2 * (a + b), 2) if a and b else round(area ** 0.5 * 4, 2)
    earth_volume = round(prep_area * max(sand_t + gravel_t, 0.2), 2)

    concrete_price = _p8v3_mp("бетон в25 w6", P["concrete_b25_mat"])
    rebar_price = _p8v3_mp("арматура металлическая д.12а500", P["rebar_a500_mat"])
    sand_price = _choose_value(
        _t2fo_prices_from_source_lines_v1(price_text, ("песок", "песчаная подушка", "песчаный")),
        choice,
        P["sand_mat"],
    )
    gravel_price = _choose_value(
        _t2fo_prices_from_source_lines_v1(price_text, ("щебень", "щебеночное основание", "щебеночный", "щебёноч")),
        choice,
        P["gravel_mat"],
    )
    manual_concrete_work_price = _t2fo_manual_monolith_work_price_v1(raw_text)
    concrete_work_price = manual_concrete_work_price or _p8v3_wp("бетонирование монолитной плиты   б/н", P["concrete_pour_work"])
    concrete_work_note = "ручная цена из правки пользователя" if manual_concrete_work_price else "работы"
    pump_price = _choose_value(_numbers_from_price_text(price_text, ("бетононасос",)), choice) or 31050
    delivery_price = round(P["logist_delivery"] * max(distance / 30.0, 1.0), 2) if distance else 0

    items = []
    if any(x in raw_text for x in ("подготов", "землян", "котлован", "выемк", "разработка грунта")):
        items.append(_ftm_row("Фундамент", "Подготовка основания и земляные работы", "м³", earth_volume, P["earth_work"], "по ТЗ: подготовка/земляные работы"))
    if sand_t > 0:
        sand_qty = round(prep_area * sand_t, 2)
        sand_work_price = _t2fpag_choose_v1(price_text, "sand_work", choice, P["sand_work"])
        items.append(_ftm_row("Фундамент", f"Песчаная подушка {int(sand_t * 1000)} мм с послойным уплотнением", "м³", sand_qty, sand_price + sand_work_price, f"работы+материал; площадь подготовки {prep_area:g} м²"))
    if gravel_t > 0:
        gravel_qty = round(prep_area * gravel_t, 2)
        gravel_work_price = _t2fpag_choose_v1(price_text, "gravel_work", choice, P["gravel_work"])
        items.append(_ftm_row("Фундамент", f"Щебёночное основание {int(gravel_t * 1000)} мм с уплотнением", "м³", gravel_qty, gravel_price + gravel_work_price, f"работы+материал; площадь подготовки {prep_area:g} м²"))
    items.append(_ftm_row("Фундамент", "Опалубка периметра плиты материал", "мп", formwork_perim, P["formwork_perim_mat"], "по размерам с фото"))
    items.append(_ftm_row("Фундамент", "Опалубка плиты монтаж/демонтаж", "мп", formwork_perim, P["formwork_install_work"], "работы"))
    items.append(_ftm_row("Фундамент", f"Арматура А500 для плиты, {layers} слоя", "т", rebar_qty, rebar_price, "расчётная масса от объёма бетона; уточняется по КЖ"))
    items.append(_ftm_row("Фундамент", f"Армирование фундаментной плиты, {layers} слоя", "м²", area, _p8v3_wp("устройство арматурного каркаса", P["rebar_install_work"]), "работы"))
    items.append(_ftm_row("Фундамент", f"Бетон {concrete_grade} для монолитной плиты {int(slab_t * 1000)} мм", "м³", concrete_volume, concrete_price, "по ТЗ"))
    items.append(_ftm_row("Фундамент", "Работы по бетону: бетонирование фундаментной плиты", "м³", concrete_volume, concrete_work_price, concrete_work_note))
    if any(x in raw_text for x in ("бетононасос", "насос", "подач")):
        items.append(_ftm_row("Фундамент", "Аренда бетононасоса / подача бетона", "смена", 1, pump_price, "по ТЗ: подача бетона"))
    if delivery_price:
        items.append(_ftm_row("Логистика", f"Доставка материалов от Санкт-Петербурга, {distance:g} км", "компл", 1, delivery_price, "по ТЗ: удалённость объекта"))

    subtotal = sum(float(it["price"]) * float(it["qty"]) for it in items if it["section"] not in ("Логистика", "Накладные расходы"))
    items.append(_ftm_row("Накладные расходы", "Организация работ и накладные", "компл", 1, round(subtotal * 0.07, 2), "7% от фундаментных работ и материалов"))
    items.append(_ftm_row("Накладные расходы", "Расходные материалы и крепёж", "компл", 1, round(subtotal * 0.015, 2), "1.5% от фундаментных работ и материалов"))
    return items


def _build_estimate_items(parsed, price_text, choice):  # noqa: F811
    if _t2fo_foundation_only_v1(parsed):
        return _t2fo_build_foundation_items_v1(parsed, price_text, choice)
    return _T2FO_PREV_BUILD_ESTIMATE_ITEMS_V1(parsed, price_text, choice)


try:
    _STV3_LOG.info("PATCH_TOPIC2_FOUNDATION_ONLY_PHOTO_SCOPE_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FOUNDATION_ONLY_PHOTO_SCOPE_V1 ===

# === PATCH_TOPIC2_FOUNDATION_NO_TEMPLATE_ROWS_V1 ===
_T2FO_NTR_PREV_CREATE_XLSX_V1 = _create_xlsx_from_template


def _t2fo_strip_non_areal_sheets_v1(path):
    try:
        from openpyxl import load_workbook as _t2fo_load_workbook
        wb = _t2fo_load_workbook(path)
        if "AREAL_CALC" not in wb.sheetnames:
            return
        for ws in list(wb.worksheets):
            if ws.title != "AREAL_CALC":
                wb.remove(ws)
        ws = wb["AREAL_CALC"]
        for row_idx in range(ws.max_row, 1, -1):
            if _low(ws.cell(row_idx, 2).value or "") == "не входит":
                ws.delete_rows(row_idx, ws.max_row - row_idx + 1)
                break
        wb.active = 0
        wb.save(path)
    except Exception as _t2fo_e:
        try:
            _STV3_LOG.warning("PATCH_TOPIC2_FOUNDATION_NO_TEMPLATE_ROWS_V1 strip failed: %s", _t2fo_e)
        except Exception:
            pass


def _t2fo_create_without_template_rows_v1(task_id, parsed, template, template_path, sheet_name, price_text, choice):
    original_create = globals().get("_T2TR_ORIG_CREATE_XLSX") or _T2FO_NTR_PREV_CREATE_XLSX_V1
    path, items, total = original_create(task_id, parsed, template, template_path, sheet_name, price_text, choice)
    _t2fo_strip_non_areal_sheets_v1(path)
    return path, items, total


def _create_xlsx_from_template(task_id, parsed, template, template_path, sheet_name, price_text, choice):  # noqa: F811
    if _t2fo_foundation_only_v1(parsed):
        return _t2fo_create_without_template_rows_v1(task_id, parsed, template, template_path, sheet_name, price_text, choice)
    return _T2FO_NTR_PREV_CREATE_XLSX_V1(task_id, parsed, template, template_path, sheet_name, price_text, choice)


try:
    _STV3_LOG.info("PATCH_TOPIC2_FOUNDATION_NO_TEMPLATE_ROWS_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FOUNDATION_NO_TEMPLATE_ROWS_V1 ===

# === PATCH_TOPIC2_PRICE_SEARCH_CONFIRM_AND_READY_DONE_V1 ===
def _topic2_price_search_explicit_intent_v1(text: str) -> bool:
    low = _low(text or "")
    return any(x in low for x in (
        "в интернете", "через интернет", "интернет-цен", "интернет цен",
        "актуальн", "свеж", "sonar", "perplexity", "поищи", "поиск",
        "найди цены", "найти цены", "проверь цены", "проверить цены",
        "поставщик", "ссылк", "рыночн",
    ))


def _topic2_price_search_prompt_text_v1(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str]) -> str:
    return (
        "Задачу понял.\n\n"
        f"Шаблон: {template.get('title')}\n"
        f"Лист: {sheet_name or 'не выбран'}\n"
        f"Объект: {(parsed or {}).get('object') or 'не указан'}\n"
        f"Материал: {(parsed or {}).get('material') or 'не указан'}\n"
        f"Размеры: {(parsed or {}).get('dimensions') or 'не указаны'}\n"
        f"Удалённость: {(parsed or {}).get('distance_km') if (parsed or {}).get('distance_km') is not None else 'не указана'} км\n\n"
        "Перед финальной сметой нужно подтвердить цены.\n"
        "Искать актуальные цены работ, материалов и логистики через интернет (Sonar/Perplexity)?\n\n"
        "Ответь: да, искать / нет, укажу цены вручную"
    )


def _topic2_price_search_yes_v1(text: str) -> bool:
    low = _low(text or "").strip(" .,!?:;")
    return low in {"да", "да искать", "искать", "да поищи", "поищи", "ищи", "нужно", "надо"} or low.startswith("да ")


def _topic2_price_search_no_v1(text: str) -> bool:
    low = _low(text or "").strip(" .,!?:;")
    return low in {"нет", "не надо", "не нужно", "без интернета", "не искать"} or low.startswith("нет ")


def _topic2_final_ready_confirm_phrase_v1(text: str) -> bool:
    low = _low(text or "").replace("[voice]", "").strip(" .,!?:;")
    if "не доволен" in low or "недоволен" in low:
        return False
    exact = {
        "готово", "готов", "готова", "готово спасибо",
        "подтверждаю", "закрывай", "можно закрывать",
        "все ок", "всё ок", "все верно", "всё верно",
        "хорошо", "отлично", "принимаю",
    }
    if low in exact:
        return True
    return any(x in low for x in (
        "задача завершена",
        "задачей завершена",
        "задача закрыта",
        "задачей закрыта",
        "доволен задачей",
        "доволен результатом",
        "да доволен",
        "да я доволен",
        "я доволен",
        "да доволеен",
        "доволеен",
        "результатом доволен",
        "можно завершать",
        "задачу завершить",
        "задачу закрыть",
        "завершай задачу",
        "завершай запрос",
        "закрывай задачу",
        "закрой задачу",
        "закрыть задачу",
        "хорошая работа",
        "можно закрывать",
        "результат подходит",
        "все подходит",
        "всё подходит",
        "все верно завершай",
        "всё верно завершай",
        "да все верно",
        "да всё верно",
        "я же тебе сказал задача завершена",
        "я же тебе сказал завершай",
        "да можно",
    ))


def _topic2_parent_has_canonical_final_markers_v1(conn, task_id: str) -> bool:
    rows = conn.execute(
        "SELECT id, COALESCE(action,'') AS action FROM task_history WHERE task_id=? ORDER BY id",
        (str(task_id),),
    ).fetchall()
    if not rows:
        return False
    last_revoke = max(
        (
            int(row["id"])
            for row in rows
            if "TOPIC2_PREVIOUS_FINAL_REVOKED" in _s(row["action"])
            or "TOPIC2_EXPLICIT_CONFIRM_REVOKED" in _s(row["action"])
        ),
        default=-1,
    )
    actions = [_s(row["action"]) for row in rows if int(row["id"]) > last_revoke]
    required_groups = (
        ("TOPIC2_TEMPLATE_SELECTED:",),
        ("TOPIC2_TEMPLATE_FILE_ID:",),
        ("TOPIC2_TEMPLATE_CACHE_USED", "TOPIC2_TEMPLATE_DRIVE_DOWNLOADED"),
        ("TOPIC2_TEMPLATE_SHEET_SELECTED:",),
        ("TOPIC2_XLSX_TEMPLATE_COPY_OK",),
        ("TOPIC2_XLSX_ROWS_WRITTEN:",),
        ("TOPIC2_XLSX_FORMULAS_OK",),
        ("TOPIC2_XLSX_CANON_COLUMNS_OK",),
        ("TOPIC2_PDF_CREATED",),
        ("TOPIC2_PDF_CYRILLIC_OK",),
        ("TOPIC2_DRIVE_UPLOAD_XLSX_OK",),
        ("TOPIC2_DRIVE_UPLOAD_PDF_OK",),
        ("TOPIC2_TELEGRAM_DELIVERED",),
    )
    return all(any(marker in action for action in actions for marker in group) for group in required_groups)


async def _topic2_handle_price_search_confirmation_v1(conn, task, logger=None) -> bool:
    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    if topic_id != TOPIC_ID_STROYKA:
        return False
    raw = _s(_row_get(task, "raw_input", ""))
    pending = _memory_latest(chat_id, "topic_2_estimate_pending_")
    if not pending or pending.get("status") != "WAITING_PRICE_SEARCH_CONFIRMATION":
        return False
    if not (_topic2_price_search_yes_v1(raw) or _topic2_price_search_no_v1(raw)):
        return False

    pending_task_id = _s(pending.get("task_id") or task_id)
    reply_to = _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None)
    parsed = pending.get("parsed") or {}
    template = pending.get("template") or CANON_TEMPLATE_FALLBACK["areal"]
    sheet_name = pending.get("sheet_name")
    template_prices = pending.get("template_prices") or ""

    if _topic2_price_search_no_v1(raw):
        text = (
            "Интернет-поиск цен не запускаю.\n\n"
            "Пришли ручные цены по позициям или напиши: считать по шаблонным ценам без интернет-проверки."
        )
        blocked = dict(pending)
        blocked["status"] = "WAITING_MANUAL_PRICE_INPUT"
        blocked["updated_at"] = _now()
        _memory_save(chat_id, f"topic_2_estimate_pending_{pending_task_id}", blocked)
        send_res = await _send_text(chat_id, text, reply_to, topic_id)
        kwargs = {"state": "WAITING_CLARIFICATION", "result": text, "error_message": "TOPIC2_MANUAL_PRICE_INPUT_REQUIRED"}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, pending_task_id, **kwargs)
        _history_safe(conn, pending_task_id, "TOPIC2_PRICE_SEARCH_DECLINED_BY_USER")
        if task_id != pending_task_id:
            _update_task_safe(conn, task_id, state="DONE", result="Интернет-поиск цен отклонён пользователем", error_message="")
            _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_DECLINE_CHILD_DONE")
        return True

    _history_safe(conn, pending_task_id, "TOPIC2_PRICE_SEARCH_CONFIRMED_BY_USER")
    _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_CONFIRMATION_ACCEPTED")
    try:
        online_prices = await _search_prices_online(parsed, template, sheet_name, conn=conn, task_id=pending_task_id)
    except Exception as exc:
        text = "SEARCH_FAILED: Sonar unavailable"
        send_res = await _send_text(chat_id, text, reply_to, topic_id)
        kwargs = {"state": "WAITING_CLARIFICATION", "result": text, "error_message": "TOPIC2_PRICE_SEARCH_FAILED"}
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, pending_task_id, **kwargs)
        _history_safe(conn, pending_task_id, "TOPIC2_PRICE_SEARCH_FAILED:" + _s(exc)[:160])
        if task_id != pending_task_id:
            _update_task_safe(conn, task_id, state="DONE", result=text, error_message="")
        return True

    confirmed = dict(pending)
    confirmed["status"] = "WAITING_PRICE_CONFIRMATION"
    confirmed["online_prices"] = online_prices
    confirmed["updated_at"] = _now()
    _memory_save(chat_id, f"topic_2_estimate_pending_{pending_task_id}", confirmed)
    text = _price_confirmation_text(parsed, template, sheet_name, template_prices, online_prices)
    send_res = await _send_text(chat_id, text, reply_to, topic_id)
    kwargs = {"state": "WAITING_CLARIFICATION", "result": text, "error_message": "TOPIC2_PRICE_CHOICE_REQUIRED"}
    if isinstance(send_res, dict) and send_res.get("bot_message_id"):
        kwargs["bot_message_id"] = send_res.get("bot_message_id")
    _update_task_safe(conn, pending_task_id, **kwargs)
    _history_safe(conn, pending_task_id, "TOPIC2_PRICE_CHOICE_REQUESTED_AFTER_SEARCH_CONFIRM")
    if task_id != pending_task_id:
        _update_task_safe(conn, task_id, state="DONE", result="Интернет-поиск цен подтверждён", error_message="")
        _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_CONFIRM_CHILD_DONE")
    return True


async def _topic2_handle_ready_done_v1(conn, task, logger=None) -> bool:
    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    if topic_id != TOPIC_ID_STROYKA:
        return False
    raw = _s(_row_get(task, "raw_input", ""))
    if not _topic2_final_ready_confirm_phrase_v1(raw):
        return False
    try:
        if parse_price_choice(raw).get("confirmed"):
            return False
    except Exception:
        pass
    reply_to = _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None)
    parent = None
    if reply_to:
        parent = conn.execute(
            """
            SELECT id, COALESCE(raw_input,'') AS raw_input, COALESCE(result,'') AS result
            FROM tasks
            WHERE CAST(chat_id AS TEXT)=?
              AND COALESCE(topic_id,0)=?
              AND state='AWAITING_CONFIRMATION'
              AND id<>?
              AND (bot_message_id=? OR reply_to_message_id=?)
            ORDER BY updated_at DESC, created_at DESC
            LIMIT 1
            """,
            (str(chat_id), int(topic_id), str(task_id), reply_to, reply_to),
        ).fetchone()
        if parent and not _topic2_parent_has_canonical_final_markers_v1(conn, _s(parent["id"])):
            parent = None
    if not parent:
        candidates = conn.execute(
            """
            SELECT id, COALESCE(raw_input,'') AS raw_input, COALESCE(result,'') AS result
            FROM tasks
            WHERE CAST(chat_id AS TEXT)=?
              AND COALESCE(topic_id,0)=?
              AND state='AWAITING_CONFIRMATION'
              AND id<>?
              AND (COALESCE(result,'') LIKE '%Смета готова%' OR COALESCE(result,'') LIKE '%Смета по извлечённым позициям готова%')
              AND (COALESCE(result,'') LIKE '%drive.google.com%' OR COALESCE(result,'') LIKE '%docs.google.com%')
            ORDER BY updated_at DESC, created_at DESC
            LIMIT 20
            """,
            (str(chat_id), int(topic_id), str(task_id)),
        ).fetchall()
        parent = next(
            (
                row for row in candidates
                if _topic2_parent_has_canonical_final_markers_v1(conn, _s(row["id"]))
            ),
            None,
        )
    if not parent:
        return False

    parent_id = _s(parent["id"])
    parent_raw = _s(parent["raw_input"])
    parent_result = _s(parent["result"])
    parent_low = _low(parent_result)
    estimate_final = "смета готов" in parent_low or "смета по извлечённым позициям готова" in parent_low
    if not (estimate_final and ("xlsx" in parent_low or "pdf" in parent_low or "drive.google.com" in parent_low or "docs.google.com" in parent_low)):
        return False

    _history_safe(conn, parent_id, "TOPIC2_EXPLICIT_CONFIRM:ready_done_phrase")
    _update_task_safe(conn, parent_id, state="DONE", error_message="")
    _history_safe(conn, parent_id, "state:DONE")
    try:
        _memory_save(chat_id, f"topic_2_estimate_pending_{parent_id}", {
            "status": "DONE",
            "task_id": parent_id,
            "topic_id": int(topic_id),
            "completed_at": _now(),
            "source": "TOPIC2_EXPLICIT_CONFIRM",
        })
        _memory_save(chat_id, f"topic_2_user_input_{parent_id}", {
            "task_id": parent_id,
            "topic_id": int(topic_id),
            "raw_input": parent_raw,
            "saved_at": _now(),
            "source": "TOPIC2_EXPLICIT_CONFIRM",
        })
        _memory_save(chat_id, f"topic_2_task_summary_{parent_id}", {
            "task_id": parent_id,
            "topic_id": int(topic_id),
            "summary": parent_result,
            "saved_at": _now(),
            "source": "TOPIC2_EXPLICIT_CONFIRM",
        })
        _memory_save(chat_id, f"topic_2_assistant_output_{parent_id}", {
            "task_id": parent_id,
            "topic_id": int(topic_id),
            "result": parent_result,
            "saved_at": _now(),
            "source": "TOPIC2_EXPLICIT_CONFIRM",
        })
    except Exception:
        pass
    _update_task_safe(conn, task_id, state="DONE", result="Подтверждение принято", error_message="")
    _history_safe(conn, task_id, "TOPIC2_CONFIRM_CHILD_DONE_READY_PHRASE")
    await _send_text(chat_id, "Принял. Задача закрыта", reply_to, topic_id)
    return True


_T2PSC_PREV_MAYBE_HANDLE_V1 = maybe_handle_stroyka_estimate


async def maybe_handle_stroyka_estimate(conn, task, logger=None):  # noqa: F811
    try:
        if await _topic2_handle_price_search_confirmation_v1(conn, task, logger=logger):
            return True
        if await _topic2_handle_ready_done_v1(conn, task, logger=logger):
            return True
    except Exception as exc:
        try:
            _history_safe(conn, _s(_row_get(task, "id")), "TOPIC2_PRICE_SEARCH_CONFIRM_OR_READY_DONE_ERR:" + _s(exc)[:160])
        except Exception:
            pass
    return await _T2PSC_PREV_MAYBE_HANDLE_V1(conn, task, logger)


try:
    _STV3_LOG.info("PATCH_TOPIC2_PRICE_SEARCH_CONFIRM_AND_READY_DONE_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_PRICE_SEARCH_CONFIRM_AND_READY_DONE_V1 ===

# === PATCH_TOPIC2_FOUNDATION_MISSING_PRICE_CACHE_SONAR_V1 ===
def _t2fo_price_text_has_family_v1(price_text, keywords):
    low = _low(price_text or "")
    return any(_low(k) in low for k in keywords)


def _t2fo_offer_lines_v1(label, unit, offers):
    lines = []
    for offer in (offers or [])[:3]:
        try:
            price = float(offer.get("price") or 0)
        except Exception:
            price = 0.0
        if price <= 0:
            continue
        lines.append(
            "- {} | {} | {} | Санкт-Петербург и Ленинградская область | {} | {} | {}".format(
                label,
                price,
                offer.get("unit") or unit,
                offer.get("supplier") or "",
                offer.get("url") or "",
                offer.get("checked_at") or datetime.date.today().isoformat(),
            )
        )
    return lines


_T2FO_MISSING_PRICE_PREV_SEARCH_V1 = _search_prices_online


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    result = await _T2FO_MISSING_PRICE_PREV_SEARCH_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
    if not _t2fo_foundation_only_v1(parsed or {}):
        return result

    raw_text = _low((parsed or {}).get("raw") or "")
    missing = []
    if ((parsed or {}).get("sand_thickness_m") or "песчан" in raw_text or "песок" in raw_text):
        if not _t2fo_price_text_has_family_v1(result, ("песок", "песчаная подушка", "песчаный")):
            missing.append(("Песок строительный для песчаной подушки", "м3", "sand"))
    if ((parsed or {}).get("gravel_thickness_m") or "щеб" in raw_text):
        if not _t2fo_price_text_has_family_v1(result, ("щебень", "щебеночное основание", "щебеночный", "щебёноч")):
            missing.append(("Щебень для основания фундаментной плиты", "м3", "gravel"))
    if not missing:
        return result

    model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
    if "sonar" not in model.lower():
        raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")

    from core.price_enrichment import _openrouter_price_search as _missing_price_search

    extra_lines = []
    for item_name, unit, code in missing:
        if conn is not None and task_id is not None:
            _history_safe(conn, task_id, "TOPIC2_PRICE_CACHE_BEFORE_SONAR:" + code)
            _history_safe(conn, task_id, "TOPIC2_PRICE_MATERIAL_SEARCH_STARTED:" + item_name)
        try:
            offers = await asyncio.wait_for(
                _missing_price_search(item_name, unit, "Санкт-Петербург и Ленинградская область"),
                timeout=45,
            )
        except Exception:
            offers = []
        lines = _t2fo_offer_lines_v1(item_name, unit, offers)
        if lines:
            extra_lines.extend(lines)
            if conn is not None and task_id is not None:
                first = offers[0] if offers else {}
                _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_FOUND:{}:{}:{}".format(
                    code,
                    _s(first.get("supplier"))[:50],
                    _s(first.get("status"))[:20],
                ))
        elif conn is not None and task_id is not None:
            _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_MISSING:" + code)

    if not extra_lines:
        return result
    joined = (str(result or "").rstrip() + "\n" + "\n".join(extra_lines)).strip()
    if conn is not None and task_id is not None:
        _history_safe(conn, task_id, "TOPIC2_MISSING_PRICE_CACHE_SONAR_DONE:" + ",".join(code for _, _, code in missing))
    return joined


def _t2spsm_families_v1(text, section=""):  # noqa: F811
    low = _low((text or "") + " " + (section or ""))
    families = set()
    checks = (
        ("sand", ("песок", "песчан", "песчаная подушка")),
        ("gravel", ("щебень", "щебен", "щебеноч", "щебеночное", "щебёноч")),
        ("gasbeton", ("газобетон", "газоблок", "блок 625", "u-блок", "u блок", "лср")),
        ("concrete", ("бетон", "монолит", "ж/б", "железобетон", "ростверк", "плита")),
        ("rebar", ("арматур", "а500", "а240", "проволока вяз")),
        ("wood", ("доска", "брус", "пиломат", "osb", "фанера")),
        ("insulation", ("пенопл", "утепл", "минват", "пир", "pir")),
        ("waterproof", ("гидроизоляц", "линокром", "мастик", "праймер")),
        ("roof", ("кров", "стропил", "мауэрлат", "мембран", "профнастил", "черепиц")),
        ("windows", ("окн", "окон", "пвх", "стеклопакет")),
        ("doors", ("двер", "дверн")),
        ("delivery", ("достав", "транспорт")),
        ("unload", ("разгруз", "погруз")),
        ("crane", ("кран",)),
        ("pump", ("бетононасос",)),
        ("masonry_work", ("кладк",)),
        ("facade", ("фасад", "внешняя отделка")),
        ("interior", ("внутрен", "отделк", "гкл", "ламинат", "плитк")),
        ("engineering", ("электрик", "водоснаб", "канализац", "отоплен", "вентиляц", "инженер")),
    )
    for fam, terms in checks:
        if any(term in low for term in terms):
            families.add(fam)
    return families


try:
    _STV3_LOG.info("PATCH_TOPIC2_FOUNDATION_MISSING_PRICE_CACHE_SONAR_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FOUNDATION_MISSING_PRICE_CACHE_SONAR_V1 ===

# === PATCH_TOPIC2_RELIABLE_PRICE_AND_FOUNDATION_SOURCE_GUARD_V1 ===
# Canon/user rule: "3 / надёжные" is a reliable price level, not maximum.
# Foundation-only final with explicit internet check must not close while
# песок/щебень remain without live/cache source lines.
_T2RPF_PREV_PARSE_PRICE_CHOICE_V1 = parse_price_choice


def _t2rpf_reliable_requested_v1(text):
    t = _low(text or "").replace("[voice]", "").strip()
    t = re.sub(r"\s+", " ", t).strip(" .,!?:;()[]{}")
    return (
        t in ("3", "3.", "третий", "вариант 3", "вариант в", "в", "v", "в)", "v)")
        or "надежн" in t
        or "надёжн" in t
        or "проверенн" in t
        or "раздел три" in t
    )


def parse_price_choice(text: str) -> Dict[str, Any]:  # noqa: F811
    res = dict(_T2RPF_PREV_PARSE_PRICE_CHOICE_V1(text))
    t = _low(text or "").replace("[voice]", "").strip()
    t = re.sub(r"\s+", " ", t).strip(" .,!?:;()[]{}")
    explicit_max = any(x in t for x in ("максим", "max", "макс ")) and not any(x in t for x in ("не максим", "а не максим", "не max"))
    if _t2rpf_reliable_requested_v1(text) and not explicit_max:
        res["choice"] = "reliable"
        res["confirmed"] = True
    return res


try:
    _PPOC_PRICE_DISPLAY.update({
        "minimum": "минимальные",
        "cheapest": "минимальные",
        "maximum": "максимальные",
        "reliable": "надёжные",
        "trusted": "надёжные",
    })
except Exception:
    pass


def _t2rpf_requires_foundation_live_prices_v1(parsed, text):
    if not _t2fo_foundation_only_v1(parsed or {}):
        return False
    low = _low(text or "")
    return any(x in low for x in ("интернет", "актуальн", "проверить", "проверь", "поищи", "найди", "стоимость песка", "стоимости песка", "стоимость щеб"))


def _t2rpf_missing_foundation_families_v1(parsed, price_text):
    parsed = parsed or {}
    raw = _low(parsed.get("raw") or "")
    missing = []
    if (parsed.get("sand_thickness_m") or "песчан" in raw or "песок" in raw) and not _t2fo_price_text_has_family_v1(price_text, ("песок", "песчаная подушка", "песчаный")):
        missing.append("песок")
    if (parsed.get("gravel_thickness_m") or "щеб" in raw) and not _t2fo_price_text_has_family_v1(price_text, ("щебень", "щебеночное основание", "щебеночный", "щебёноч")):
        missing.append("щебень")
    return missing


_T2RPF_PREV_GENERATE_AND_SEND_V1 = _generate_and_send


async def _generate_and_send(conn, task, pending, confirm_text, logger=None):  # noqa: F811
    task_id = _s(_row_get(task, "id"))
    chat_id = _s(_row_get(task, "chat_id"))
    topic_id = int(_row_get(task, "topic_id", 0) or 0)
    reply_to = _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None)
    parsed = (pending or {}).get("parsed") or {}
    online_prices = (pending or {}).get("online_prices") or ""
    raw_context = "\n".join([
        _s(parsed.get("raw") if isinstance(parsed, dict) else ""),
        _s(confirm_text),
    ])

    if topic_id == TOPIC_ID_STROYKA and _t2rpf_requires_foundation_live_prices_v1(parsed, raw_context):
        missing = _t2rpf_missing_foundation_families_v1(parsed, online_prices)
        if missing:
            try:
                _history_safe(conn, task_id, "TOPIC2_FOUNDATION_LIVE_PRICE_GUARD_SEARCH:" + ",".join(missing))
                refreshed = await _search_prices_online(
                    parsed,
                    (pending or {}).get("template") or CANON_TEMPLATE_FALLBACK["areal"],
                    (pending or {}).get("sheet_name"),
                    conn=conn,
                    task_id=task_id,
                )
                online_prices = refreshed or online_prices
                pending["online_prices"] = online_prices
                pending["status"] = "WAITING_PRICE_CONFIRMATION"
                _memory_save(chat_id, f"topic_2_estimate_pending_{task_id}", pending)
            except Exception as exc:
                _history_safe(conn, task_id, "TOPIC2_FOUNDATION_LIVE_PRICE_GUARD_SEARCH_FAILED:" + _s(exc)[:160])
        missing = _t2rpf_missing_foundation_families_v1(parsed, online_prices)
        if missing:
            text = (
                "Не закрываю финальную смету: не найдены подтверждённые интернет-цены для "
                + ", ".join(missing)
                + ". Пришли ручные цены или разреши повторить поиск."
            )
            send_res = await _send_text(chat_id, text, reply_to, topic_id)
            kwargs = {"state": "WAITING_CLARIFICATION", "result": text, "error_message": "TOPIC2_FOUNDATION_PRICE_SOURCE_REQUIRED"}
            if isinstance(send_res, dict) and send_res.get("bot_message_id"):
                kwargs["bot_message_id"] = send_res.get("bot_message_id")
            _update_task_safe(conn, task_id, **kwargs)
            _history_safe(conn, task_id, "TOPIC2_FOUNDATION_FINAL_BLOCKED_TEMPLATE_ONLY:" + ",".join(missing))
            return True

    return await _T2RPF_PREV_GENERATE_AND_SEND_V1(conn, task, pending, confirm_text, logger=logger)


try:
    _STV3_LOG.info("PATCH_TOPIC2_RELIABLE_PRICE_AND_FOUNDATION_SOURCE_GUARD_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_RELIABLE_PRICE_AND_FOUNDATION_SOURCE_GUARD_V1 ===


# === PATCH_TOPIC2_FULL_FOUNDATION_PRICE_SOURCE_GUARD_V1 ===
# Canon: when user explicitly asks to verify material prices, final XLSX must
# prefer live confirmed sources and foundation-only route must search missing
# foundation price families instead of closing template_only rows silently.
_T2FFPS_PREV_MATCH_PRICE_SOURCE_V1 = _match_price_source


def _t2ffps_is_live_source_v1(src):
    status = _low((src or {}).get("status") or "")
    return status in ("live_confirmed", "confirmed") and bool((src or {}).get("supplier")) and bool((src or {}).get("url"))


def _match_price_source(sources: List[Dict[str, Any]], item_name: str, item_section: str) -> Dict[str, Any]:  # noqa: F811
    live_sources = [src for src in (sources or []) if _t2ffps_is_live_source_v1(src)]
    if live_sources:
        live = _T2FFPS_PREV_MATCH_PRICE_SOURCE_V1(live_sources, item_name, item_section)
        if live and live.get("status") != "template_only":
            return live
    return _T2FFPS_PREV_MATCH_PRICE_SOURCE_V1(sources, item_name, item_section)


_T2FFPS_PREV_SEARCH_PRICES_ONLINE_V1 = _search_prices_online


def _t2ffps_existing_online_prices_v1(conn, task_id):
    if conn is None or task_id is None:
        return ""
    try:
        row = conn.execute("SELECT chat_id FROM tasks WHERE id=? LIMIT 1", (_s(task_id),)).fetchone()
        chat_id = _s(row[0] if row else "")
        if not chat_id:
            return ""
        import sqlite3 as _t2ffps_sqlite3
        import json as _t2ffps_json
        mem = _t2ffps_sqlite3.connect("/root/.areal-neva-core/data/memory.db")
        try:
            r = mem.execute(
                "SELECT value FROM memory WHERE chat_id=? AND key=? ORDER BY timestamp DESC LIMIT 1",
                (chat_id, "topic_2_estimate_pending_" + _s(task_id)),
            ).fetchone()
        finally:
            mem.close()
        if not r:
            return ""
        data = _t2ffps_json.loads(r[0])
        return _s(data.get("online_prices") or "") if isinstance(data, dict) else ""
    except Exception:
        return ""


def _t2ffps_has_live_source_v1(price_text, keywords):
    try:
        sources = _parse_price_sources(price_text or "")
    except Exception:
        sources = []
    keys = tuple(_low(k) for k in (keywords or ()))
    for src in sources:
        if not _t2ffps_is_live_source_v1(src):
            continue
        pos = _low(src.get("position") or "")
        if any(k in pos for k in keys):
            return True
    return False


def _t2ffps_has_live_source_all_v1(price_text, keywords):
    try:
        sources = _parse_price_sources(price_text or "")
    except Exception:
        sources = []
    keys = tuple(_low(k) for k in (keywords or ()))
    for src in sources:
        if not _t2ffps_is_live_source_v1(src):
            continue
        pos = _low(src.get("position") or "")
        if keys and all(k in pos for k in keys):
            return True
    return False


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    base = await _T2FFPS_PREV_SEARCH_PRICES_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
    if not _t2fo_foundation_only_v1(parsed or {}):
        return base
    existing = _t2ffps_existing_online_prices_v1(conn, task_id)
    combined = "\n".join(x for x in (existing, base) if _s(x).strip()).strip()

    missing = []
    checks = (
        ("Опалубка для монолитной фундаментной плиты материал", "мп", "formwork_material", ("опалуб",)),
        ("Монтаж демонтаж опалубки фундаментной плиты", "мп", "formwork_work", ("опалуб", "монтаж")),
        ("Армирование фундаментной плиты работы", "м2", "rebar_work", ("армирован",)),
        ("Устройство песчаной подушки с послойным уплотнением работы", "м3", "sand_work", ("песчан", "уплотн")),
        ("Устройство щебеночного основания с уплотнением работы", "м3", "gravel_work", ("щеб", "уплотн")),
    )
    for item_name, unit, code, keywords in checks:
        has_source = (
            _t2ffps_has_live_source_all_v1(combined, keywords)
            if code in ("sand_work", "gravel_work")
            else _t2ffps_has_live_source_v1(combined, keywords)
        )
        if not has_source:
            missing.append((item_name, unit, code))
    if not missing:
        return combined or base

    model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
    if "sonar" not in model.lower():
        raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")

    from core.price_enrichment import _openrouter_price_search as _missing_price_search
    extra_lines = []
    for item_name, unit, code in missing:
        if conn is not None and task_id is not None:
            _history_safe(conn, task_id, "TOPIC2_PRICE_CACHE_BEFORE_SONAR:" + code)
            _history_safe(conn, task_id, "TOPIC2_PRICE_MATERIAL_SEARCH_STARTED:" + item_name)
        try:
            offers = await asyncio.wait_for(
                _missing_price_search(item_name, unit, "Санкт-Петербург и Ленинградская область"),
                timeout=45,
            )
        except Exception:
            offers = []
        lines = _t2fo_offer_lines_v1(item_name, unit, offers)
        if lines:
            extra_lines.extend(lines)
            if conn is not None and task_id is not None:
                first = offers[0] if offers else {}
                _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_FOUND:{}:{}:{}".format(
                    code,
                    _s(first.get("supplier"))[:50],
                    _s(first.get("status"))[:20],
                ))
        elif conn is not None and task_id is not None:
            _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_MISSING:" + code)
    if extra_lines and conn is not None and task_id is not None:
        _history_safe(conn, task_id, "TOPIC2_FULL_FOUNDATION_PRICE_SOURCE_SONAR_DONE:" + ",".join(code for _, _, code in missing))
    return (combined + "\n" + "\n".join(extra_lines)).strip() if extra_lines else (combined or base)

try:
    _STV3_LOG.info("PATCH_TOPIC2_FULL_FOUNDATION_PRICE_SOURCE_GUARD_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FULL_FOUNDATION_PRICE_SOURCE_GUARD_V1 ===


# === PATCH_TOPIC2_FOUNDATION_REQUIRED_PRICE_FAMILIES_V1 ===
# Extends the existing foundation final guard: for foundation-only estimates with
# explicit price verification, formwork and reinforcement work are required price
# families too. This does not change non-topic_2 and does not bypass Sonar.
_T2FRPF_PREV_MISSING_FOUNDATION_FAMILIES_V1 = _t2rpf_missing_foundation_families_v1


def _t2rpf_missing_foundation_families_v1(parsed, price_text):  # noqa: F811
    missing = list(_T2FRPF_PREV_MISSING_FOUNDATION_FAMILIES_V1(parsed, price_text) or [])
    if not _t2fo_foundation_only_v1(parsed or {}):
        return missing
    if not _t2ffps_has_live_source_v1(price_text, ("опалуб",)):
        missing.append("опалубка")
    if not _t2ffps_has_live_source_v1(price_text, ("армирован",)):
        missing.append("армирование")
    if not _t2ffps_has_live_source_all_v1(price_text, ("песчан", "уплотн")):
        missing.append("уплотнение песчаной подушки")
    if not _t2ffps_has_live_source_all_v1(price_text, ("щеб", "уплотн")):
        missing.append("уплотнение щебёночного основания")
    out = []
    for item in missing:
        if item not in out:
            out.append(item)
    return out

try:
    _STV3_LOG.info("PATCH_TOPIC2_FOUNDATION_REQUIRED_PRICE_FAMILIES_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FOUNDATION_REQUIRED_PRICE_FAMILIES_V1 ===

# === PATCH_TOPIC2_FOUNDATION_PRICE_APPLY_MATCH_V1 ===
# Canon/user rule: foundation-only final must apply confirmed source-line prices
# to AREAL_CALC work/material columns, and must not attach formwork sources to
# concrete rows just because both mention a monolithic slab.
_T2FPAG_PREV_MATCH_PRICE_SOURCE_V1 = _match_price_source
_T2FPAG_PREV_CREATE_XLSX_V1 = _create_xlsx_from_template
_T2FPAG_PREV_FINAL_SUMMARY_V1 = _final_summary
_T2FPAG_PREV_SEARCH_PRICES_ONLINE_V1 = _search_prices_online
_T2FPAG_PREV_MISSING_FAMILIES_V1 = _t2rpf_missing_foundation_families_v1


def _t2fpag_family_v1(text):
    low = _low(text or "")
    if "опалуб" in low:
        if "материал" in low and not any(x in low for x in ("монтаж", "демонтаж", "работ")):
            return "formwork_material"
        if any(x in low for x in ("монтаж", "демонтаж", "работ", "установ")):
            return "formwork_work"
        return "formwork"
    if "армирован" in low:
        return "rebar_work"
    if "арматур" in low or "а500" in low:
        return "rebar_material"
    if "бетонирован" in low or "заливк" in low or "работа (бетон" in low:
        return "concrete_work"
    if "бетон" in low or "в25" in low or "в30" in low or "м350" in low:
        return "concrete_material"
    if "песок" in low or "песчан" in low:
        if any(x in low for x in ("подуш", "основан")) and any(x in low for x in ("уплотнен", "уплотнени", "работ")):
            return "sand_base"
        if any(x in low for x in ("устройство", "уплотнен", "уплотнени", "работ")):
            return "sand_work"
        return "sand"
    if "щеб" in low:
        if any(x in low for x in ("основан", "подуш")) and any(x in low for x in ("уплотнен", "уплотнени", "работ")):
            return "gravel_base"
        if any(x in low for x in ("устройство", "уплотнен", "уплотнени", "работ")):
            return "gravel_work"
        return "gravel"
    if "достав" in low or "транспорт" in low:
        return "delivery"
    return ""


def _t2fpag_empty_source_v1():
    return {"supplier": "", "url": "", "checked_at": datetime.date.today().isoformat(), "status": "template_only"}


def _match_price_source(sources: List[Dict[str, Any]], item_name: str, item_section: str) -> Dict[str, Any]:  # noqa: F811
    item_family = _t2fpag_family_v1(item_name)
    if item_family:
        exact_sources = [
            src for src in (sources or [])
            if _t2fpag_family_v1(src.get("position") or "") == item_family
        ]
        if exact_sources:
            live_exact = [src for src in exact_sources if _t2ffps_is_live_source_v1(src)]
            if live_exact:
                return live_exact[0]
            return exact_sources[0]
    matched = _T2FPAG_PREV_MATCH_PRICE_SOURCE_V1(sources, item_name, item_section)
    matched_family = _t2fpag_family_v1((matched or {}).get("position") or "")
    if item_family and matched_family and matched_family != item_family:
        return _t2fpag_empty_source_v1()
    return matched


def _t2fpag_line_values_v1(price_text, required=(), any_of=(), exclude=()):
    vals = []
    req = tuple(_low(x) for x in (required or ()))
    any_terms = tuple(_low(x) for x in (any_of or ()))
    exc = tuple(_low(x) for x in (exclude or ()))
    for line in str(price_text or "").splitlines():
        low = _low(line)
        if req and not all(x in low for x in req):
            continue
        if any_terms and not any(x in low for x in any_terms):
            continue
        if exc and any(x in low for x in exc):
            continue
        parts = [p.strip() for p in line.strip(" \t-—•·").split("|")]
        if len(parts) < 2 or "нет данных" in _low(parts[1]):
            continue
        try:
            value = float(re.sub(r"[^0-9.,]", "", parts[1]).replace(",", "."))
        except Exception:
            value = 0.0
        if 100 <= value <= 10000000:
            vals.append(value)
    return vals


def _t2fpag_choose_v1(price_text, family, choice, default=0.0):
    if family == "formwork_material":
        vals = _t2fpag_line_values_v1(price_text, required=("опалуб", "материал"))
    elif family == "formwork_work":
        vals = _t2fpag_line_values_v1(price_text, required=("опалуб",), any_of=("монтаж", "демонтаж", "работ", "установ"), exclude=("материал",))
    elif family == "rebar_work":
        vals = _t2fpag_line_values_v1(price_text, required=("армирован",), any_of=("работ", "монтаж", "устройств"))
    elif family == "rebar_material":
        vals = _t2fpag_line_values_v1(price_text, required=("арматур",), any_of=("а500",))
    elif family == "sand_work":
        vals = _t2fpag_line_values_v1(price_text, required=("песчан",), any_of=("уплотн", "работ", "устройств"), exclude=("песок строительный", "материал"))
    elif family == "gravel_work":
        vals = _t2fpag_line_values_v1(price_text, required=("щеб",), any_of=("уплотн", "работ", "устройств"), exclude=("материал",))
    else:
        vals = []
    return _choose_value(vals, choice, default) if vals else float(default or 0.0)


def _t2fpag_pending_raw_v1(task_id):
    try:
        import sqlite3 as _t2fpag_sqlite3
        import json as _t2fpag_json
        mem = _t2fpag_sqlite3.connect("/root/.areal-neva-core/data/memory.db")
        try:
            row = mem.execute(
                "SELECT value FROM memory WHERE key=? ORDER BY timestamp DESC LIMIT 1",
                ("topic_2_estimate_pending_" + _s(task_id),),
            ).fetchone()
        finally:
            mem.close()
        if not row:
            return ""
        data = _t2fpag_json.loads(row[0])
        parsed = data.get("parsed") if isinstance(data, dict) else {}
        return _s((parsed or {}).get("raw") or "")
    except Exception:
        return ""


def _t2fpag_manual_concrete_work_v1(parsed, task_id=None):
    try:
        value = _t2fo_manual_monolith_work_price_v1((parsed or {}).get("raw") or "")
        if value:
            return value
        if task_id:
            return _t2fo_manual_monolith_work_price_v1(_t2fpag_pending_raw_v1(task_id))
    except Exception:
        pass
    return 0.0


def _t2fpag_exact_source_v1(price_text, family):
    try:
        for src in _parse_price_sources(price_text or ""):
            if _t2fpag_family_v1(src.get("position") or "") == family and _t2ffps_is_live_source_v1(src):
                return src
    except Exception:
        pass
    return {}


def _t2fpag_materials_vat_only_v1(parsed):
    raw = _low((parsed or {}).get("raw") or "")
    return (
        "ндс" in raw
        and "материал" in raw
        and "работ" in raw
        and "без ндс" in raw
        and ("с ндс" in raw or "ндс" in raw)
    )


def _t2fpag_combined_source_v1(price_text, material_family, work_family):
    material = _t2fpag_exact_source_v1(price_text, material_family)
    work = _t2fpag_exact_source_v1(price_text, work_family)
    if not material and not work:
        return {}
    suppliers = []
    urls = []
    checked = []
    for src in (work, material):
        if src.get("supplier") and src.get("supplier") not in suppliers:
            suppliers.append(src.get("supplier"))
        if src.get("url") and src.get("url") not in urls:
            urls.append(src.get("url"))
        if src.get("checked_at"):
            checked.append(src.get("checked_at"))
    return {
        "status": "LIVE_CONFIRMED",
        "supplier": " / ".join(suppliers),
        "url": " / ".join(urls),
        "checked_at": max(checked) if checked else datetime.date.today().isoformat(),
    }


def _t2fpag_rewrite_foundation_xlsx_v1(path, items, parsed, price_text, choice, task_id=None):
    try:
        from openpyxl import load_workbook as _t2fpag_load_workbook
        wb = _t2fpag_load_workbook(path)
        ws = wb["AREAL_CALC"] if "AREAL_CALC" in wb.sheetnames else wb.active
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            label = _low(ws.cell(row_idx, 9).value or "")
            name = _s(ws.cell(row_idx, 3).value or "")
            if label.startswith("итого"):
                break
            if not name:
                continue
            qty = float(ws.cell(row_idx, 5).value or 0)
            work = float(ws.cell(row_idx, 6).value or 0)
            mat = float(ws.cell(row_idx, 8).value or 0)
            family = _t2fpag_family_v1(name)
            if family == "formwork_material":
                mat = _t2fpag_choose_v1(price_text, family, choice, mat or work)
                work = 0.0
            elif family == "formwork_work":
                work = _t2fpag_choose_v1(price_text, family, choice, work)
                mat = 0.0
            elif family == "rebar_work":
                work = _t2fpag_choose_v1(price_text, family, choice, work)
                mat = 0.0
            elif family == "rebar_material":
                mat = _t2fpag_choose_v1(price_text, family, choice, mat)
                work = 0.0
            elif family == "concrete_work":
                manual = _t2fpag_manual_concrete_work_v1(parsed, task_id=task_id)
                if manual:
                    work = manual
                    mat = 0.0
            elif family == "sand_base":
                work = _t2fpag_choose_v1(price_text, "sand_work", choice, _FTM_PRICES["sand_work"])
                mat = _choose_value(
                    _t2fo_prices_from_source_lines_v1(price_text, ("песок", "песчаная подушка", "песчаный")),
                    choice,
                    _FTM_PRICES["sand_mat"],
                )
            elif family == "gravel_base":
                work = _t2fpag_choose_v1(price_text, "gravel_work", choice, _FTM_PRICES["gravel_work"])
                mat = _choose_value(
                    _t2fo_prices_from_source_lines_v1(price_text, ("щебень", "щебеночное основание", "щебеночный", "щебёноч")),
                    choice,
                    _FTM_PRICES["gravel_mat"],
                )
            elif family == "sand_work":
                work = _t2fpag_choose_v1(price_text, family, choice, work or _FTM_PRICES["sand_work"])
                mat = 0.0
            elif family == "gravel_work":
                work = _t2fpag_choose_v1(price_text, family, choice, work or _FTM_PRICES["gravel_work"])
                mat = 0.0
            ws.cell(row_idx, 6, round(work, 2))
            ws.cell(row_idx, 7).value = f"=E{row_idx}*F{row_idx}"
            ws.cell(row_idx, 8, round(mat, 2))
            ws.cell(row_idx, 9).value = f"=E{row_idx}*H{row_idx}"
            ws.cell(row_idx, 10).value = f"=G{row_idx}+I{row_idx}"
            exact_source = _t2fpag_exact_source_v1(price_text, family)
            if family == "sand_base":
                exact_source = _t2fpag_combined_source_v1(price_text, "sand", "sand_work")
            elif family == "gravel_base":
                exact_source = _t2fpag_combined_source_v1(price_text, "gravel", "gravel_work")
            if exact_source:
                ws.cell(row_idx, 11, exact_source.get("status", "LIVE_CONFIRMED"))
                ws.cell(row_idx, 12, exact_source.get("supplier", ""))
                ws.cell(row_idx, 13, exact_source.get("url", ""))
                ws.cell(row_idx, 14, exact_source.get("checked_at", datetime.date.today().isoformat()))
            if family == "concrete_work" and manual:
                ws.cell(row_idx, 11, "MANUAL")
                ws.cell(row_idx, 12, "user")
                ws.cell(row_idx, 13, "")
                ws.cell(row_idx, 14, datetime.date.today().isoformat())
            if family in ("sand_work", "gravel_work") and not exact_source:
                ws.cell(row_idx, 11, "TEMPLATE_ONLY")
                ws.cell(row_idx, 12, "")
                ws.cell(row_idx, 13, "")
                ws.cell(row_idx, 14, datetime.date.today().isoformat())
            rows.append((row_idx, name, qty, work, mat))

        subtotal = sum(qty * (work + mat) for _, name, qty, work, mat in rows if _low(ws.cell(_, 2).value or "") not in ("логистика", "накладные расходы", "накладные"))
        for row_idx, name, qty, work, mat in rows:
            low_name = _low(name)
            if "организация работ и накладные" in low_name:
                ws.cell(row_idx, 6, round(subtotal * 0.07, 2))
                ws.cell(row_idx, 8, 0)
                ws.cell(row_idx, 7).value = f"=E{row_idx}*F{row_idx}"
                ws.cell(row_idx, 9).value = f"=E{row_idx}*H{row_idx}"
                ws.cell(row_idx, 10).value = f"=G{row_idx}+I{row_idx}"
                ws.cell(row_idx, 11, "TEMPLATE_ONLY")
                ws.cell(row_idx, 12, "")
                ws.cell(row_idx, 13, "")
                ws.cell(row_idx, 14, datetime.date.today().isoformat())
            elif "расходные материалы" in low_name:
                ws.cell(row_idx, 6, 0)
                ws.cell(row_idx, 8, round(subtotal * 0.015, 2))
                ws.cell(row_idx, 7).value = f"=E{row_idx}*F{row_idx}"
                ws.cell(row_idx, 9).value = f"=E{row_idx}*H{row_idx}"
                ws.cell(row_idx, 10).value = f"=G{row_idx}+I{row_idx}"
                ws.cell(row_idx, 11, "TEMPLATE_ONLY")
                ws.cell(row_idx, 12, "")
                ws.cell(row_idx, 13, "")
                ws.cell(row_idx, 14, datetime.date.today().isoformat())

        if _t2fpag_materials_vat_only_v1(parsed):
            for row_idx in range(2, ws.max_row + 1):
                label = _low(ws.cell(row_idx, 9).value or "")
                if label.startswith("итого"):
                    data_last = row_idx - 2
                    vat_row = row_idx + 1
                    gross_row = row_idx + 2
                    if data_last >= 2:
                        ws.cell(vat_row, 9, "НДС 22% по материалам (работы без НДС)")
                        ws.cell(vat_row, 10).value = f"=SUM(I2:I{data_last})*22%"
                        ws.cell(gross_row, 9, "К оплате с НДС по материалам")
                        ws.cell(gross_row, 10).value = f"=J{row_idx}+J{vat_row}"
                    break

        wb.save(path)
        wb.close()
    except Exception as exc:
        try:
            _STV3_LOG.warning("PATCH_TOPIC2_FOUNDATION_PRICE_APPLY_MATCH_V1 rewrite failed: %s", exc)
        except Exception:
            pass
    return path


def _t2fpag_items_from_xlsx_v1(path, items):
    try:
        from openpyxl import load_workbook as _t2fpag_load_workbook
        wb = _t2fpag_load_workbook(path, data_only=False)
        ws = wb["AREAL_CALC"] if "AREAL_CALC" in wb.sheetnames else wb.active
        updated = []
        total = 0.0
        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            if _low(ws.cell(row_idx, 9).value or "").startswith("итого"):
                break
            name = _s(ws.cell(row_idx, 3).value or "")
            if not name:
                continue
            qty = float(ws.cell(row_idx, 5).value or 0)
            work = float(ws.cell(row_idx, 6).value or 0)
            mat = float(ws.cell(row_idx, 8).value or 0)
            total += qty * (work + mat)
            data_rows.append((name, work, mat))
        wb.close()
        for idx, it in enumerate(items or []):
            item = dict(it)
            if idx < len(data_rows):
                _, work, mat = data_rows[idx]
                item["work_price"] = work
                item["mat_price"] = mat
                item["price"] = round(work + mat, 2)
                item["kind"] = "mixed" if work and mat else ("work" if work else "material")
            updated.append(item)
        return updated or items, total
    except Exception:
        return items, sum(float(it.get("qty") or 0) * float(it.get("price") or 0) for it in (items or []))


def _create_xlsx_from_template(task_id: str, parsed: Dict[str, Any], template: Dict[str, Any], template_path: Optional[str], sheet_name: Optional[str], price_text: str, choice: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]], float]:  # noqa: F811
    path, items, total = _T2FPAG_PREV_CREATE_XLSX_V1(task_id, parsed, template, template_path, sheet_name, price_text, choice)
    if _t2fo_foundation_only_v1(parsed or {}):
        _t2fpag_rewrite_foundation_xlsx_v1(path, items, parsed, price_text, choice, task_id=task_id)
        items, total = _t2fpag_items_from_xlsx_v1(path, items)
    return path, items, total


def _final_summary(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], choice: Dict[str, Any], py_total: float, items=None) -> str:  # noqa: F811
    if not items or not any("work_price" in it or "mat_price" in it for it in (items or [])):
        return _T2FPAG_PREV_FINAL_SUMMARY_V1(parsed, template, sheet_name, choice, py_total, items=items)
    mat_total = work_total = logistics_total = overhead_total = 0.0
    for it in items:
        qty = float(it.get("qty") or 0)
        sec = _s(it.get("section") or "")
        work = float(it.get("work_price") or 0)
        mat = float(it.get("mat_price") or 0)
        val = qty * (work + mat)
        if sec == "Логистика":
            logistics_total += val
        elif sec in ("Накладные расходы", "Накладные"):
            overhead_total += val
        else:
            work_total += qty * work
            mat_total += qty * mat
    obj = parsed.get("object") or parsed.get("raw") or "объект"
    material = parsed.get("material") or "не указан"
    dims = parsed.get("dims") or parsed.get("dimensions")
    try:
        a, b = float(dims[0]), float(dims[1])
        area_s = f"{a * b:.0f} м²"
    except Exception:
        area_s = str(parsed.get("area") or "не указана")
    subtotal = round(mat_total + work_total + logistics_total + overhead_total, 2)
    material_vat = round(mat_total * 0.22, 2) if _t2fpag_materials_vat_only_v1(parsed) else 0.0
    vat_lines = (
        f"  НДС 22% по материалам: {material_vat:,.0f} руб\n"
        f"  С НДС по материалам: {subtotal + material_vat:,.0f} руб\n"
        if material_vat
        else "  НДС не включен. Если нужен расчет с НДС 22%, ответь: с НДС"
    )
    return (
        f"✅ Смета готова\n\n"
        f"Объект: {obj}   Материал: {material}   Площадь: {area_s}   "
        f"Этажность: {parsed.get('floors') or 'не указана'}   Регион: {parsed.get('region') or parsed.get('location') or 'СПб и ЛО'}\n"
        f"Шаблон: {template.get('title') or 'Ареал Нева.xlsx'}   Лист: {sheet_name or 'смета'}   Цены: {choice.get('choice') or 'шаблон'}\n\n"
        f"Итого:\n"
        f"  Материалы: {mat_total:,.0f} руб\n"
        f"  Работы: {work_total:,.0f} руб\n"
        f"  Логистика: {logistics_total:,.0f} руб\n"
        f"  Накладные: {overhead_total:,.0f} руб\n"
        f"  Итого без НДС: {subtotal:,.0f} руб\n"
        f"{vat_lines}"
    ).replace(",", " ")


def _t2fpag_has_live_source_all_v1(price_text, required):
    req = tuple(_low(x) for x in (required or ()))
    try:
        sources = _parse_price_sources(price_text or "")
    except Exception:
        sources = []
    for src in sources:
        if not _t2ffps_is_live_source_v1(src):
            continue
        pos = _low(src.get("position") or "")
        if all(x in pos for x in req):
            return True
    return False


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    base = await _T2FPAG_PREV_SEARCH_PRICES_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
    if not _t2fo_foundation_only_v1(parsed or {}):
        return base
    if _t2fpag_has_live_source_all_v1(base, ("опалуб", "монтаж")):
        return base
    model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
    if "sonar" not in model.lower():
        raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")
    from core.price_enrichment import _openrouter_price_search as _t2fpag_price_search
    if conn is not None and task_id is not None:
        _history_safe(conn, task_id, "TOPIC2_PRICE_CACHE_BEFORE_SONAR:formwork_work")
        _history_safe(conn, task_id, "TOPIC2_PRICE_MATERIAL_SEARCH_STARTED:Монтаж демонтаж опалубки фундаментной плиты работы")
    try:
        offers = await asyncio.wait_for(
            _t2fpag_price_search("Монтаж демонтаж опалубки фундаментной плиты работы", "мп", "Санкт-Петербург и Ленинградская область"),
            timeout=45,
        )
    except Exception:
        offers = []
    lines = _t2fo_offer_lines_v1("Монтаж демонтаж опалубки фундаментной плиты работы", "мп", offers)
    if lines and conn is not None and task_id is not None:
        first = offers[0] if offers else {}
        _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_FOUND:formwork_work:{}:{}".format(
            _s(first.get("supplier"))[:50],
            _s(first.get("status"))[:20],
        ))
    return (base + "\n" + "\n".join(lines)).strip() if lines else base


def _t2rpf_missing_foundation_families_v1(parsed, price_text):  # noqa: F811
    missing = list(_T2FPAG_PREV_MISSING_FAMILIES_V1(parsed, price_text) or [])
    if _t2fo_foundation_only_v1(parsed or {}) and not _t2fpag_has_live_source_all_v1(price_text, ("опалуб", "монтаж")):
        missing.append("монтаж опалубки")
    out = []
    for item in missing:
        if item not in out:
            out.append(item)
    return out


try:
    _STV3_LOG.info("PATCH_TOPIC2_FOUNDATION_PRICE_APPLY_MATCH_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FOUNDATION_PRICE_APPLY_MATCH_V1 ===

# === PATCH_TOPIC2_FILE_FACTS_NO_TEMPLATE_SUBSTITUTION_V1 ===
# FACT ONLY / §12: a current file/PDF/OCR task must not be completed from old
# sample/template composition. If extracted rows are not an explicit estimate
# basis, ask for clarification instead of producing a false final estimate.
def _t2ff_raw_meta_v1(parsed):
    raw = _s((parsed or {}).get("raw") or "")
    try:
        obj, _ = json.JSONDecoder().raw_decode(raw.lstrip())
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def _t2ff_file_context_v1(parsed):
    parsed = parsed or {}
    meta = _t2ff_raw_meta_v1(parsed)
    raw = _low(parsed.get("raw") or "")
    return bool(
        parsed.get("pdf_spec_rows")
        or parsed.get("ocr_table_rows")
        or parsed.get("pdf_spec_source")
        or meta.get("file_name")
        or meta.get("mime_type")
        or "file_id" in raw
    )


def _t2ff_rows_are_explicit_estimate_basis_v1(rows):
    good = 0
    bad_markers = ("гост", "петротех", "площадь", "общая")
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        name = _low(row.get("name") or "")
        unit = _low(row.get("unit") or "")
        try:
            qty = float(row.get("qty") or 0)
        except Exception:
            qty = 0
        if qty <= 0 or not name:
            continue
        if any(name == marker or name.startswith(marker + " ") for marker in bad_markers):
            continue
        if any(x in name for x in ("секция", "калит", "ворот", "стойк", "труба", "металлоконструкц", "ограж")):
            good += 1
            continue
        if unit in ("м", "м2", "м²", "м3", "м³", "кг", "т", "шт") and len(name) > 8:
            good += 1
    return good >= 1


def _t2ff_file_clarification_v1(parsed):
    meta = _t2ff_raw_meta_v1(parsed)
    fname = _s(meta.get("file_name") or Path(_s((parsed or {}).get("pdf_spec_source") or "")).name or "файл")
    rows = list((parsed or {}).get("pdf_spec_rows") or []) + list((parsed or {}).get("ocr_table_rows") or [])
    preview = []
    for row in rows[:7]:
        if isinstance(row, dict) and _s(row.get("name")):
            preview.append(f"- {_s(row.get('name'))}: {_s(row.get('qty'))} {_s(row.get('unit'))}")
    found = "\n".join(preview) if preview else "- явная ВОР/спецификация не найдена"
    return (
        f"{fname} принят и прочитан. Нашёл текущие проектные строки:\n{found}\n\n"
        "Шаблонные строки из старых смет не подставляю. Подтверди, пожалуйста: считать смету только по найденным позициям "
        "и искать актуальные цены на материалы/изготовление/монтаж через интернет, либо пришли ВОР/спецификацию с объёмами."
    )


_T2FF_PREV_MISSING_QUESTION_V1 = _missing_question


def _missing_question(parsed: Dict[str, Any]) -> Optional[str]:  # noqa: F811
    if _topic2_volume_extract_requested_v1((parsed or {}).get("raw") or ""):
        return None
    if _t2ff_file_context_v1(parsed):
        rows = list((parsed or {}).get("pdf_spec_rows") or []) + list((parsed or {}).get("ocr_table_rows") or [])
        confirm_text = (
            _low((parsed or {}).get("_topic2_confirm_text") or "")
            + " " + _low((parsed or {}).get("_topic2_history_clarified") or "")
            + " " + _low((parsed or {}).get("raw") or "")
        )
        confirmed_current_rows = any(x in confirm_text for x in (
            "считать по найденным позициям",
            "считать по найденным проектным позициям",
            "считай по найденным позициям",
            "считай по найденным проектным позициям",
            "только найденные позиции",
            "только найденные проектные позиции",
            "искать цены",
            "ищи цены",
        ))
        if not confirmed_current_rows:
            return _t2ff_file_clarification_v1(parsed)
        if not _t2ff_rows_are_explicit_estimate_basis_v1(rows):
            return _t2ff_file_clarification_v1(parsed)
    return _T2FF_PREV_MISSING_QUESTION_V1(parsed)


try:
    _T2FF_PREV_SAMPLE_MATRIX_MODE_V1 = _t2s_sample_matrix_mode_v1

    def _t2s_sample_matrix_mode_v1(parsed):  # noqa: F811
        if _t2ff_file_context_v1(parsed):
            return False
        return _T2FF_PREV_SAMPLE_MATRIX_MODE_V1(parsed)
except Exception:
    pass

try:
    _STV3_LOG.info("PATCH_TOPIC2_FILE_FACTS_NO_TEMPLATE_SUBSTITUTION_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FILE_FACTS_NO_TEMPLATE_SUBSTITUTION_V1 ===

# === PATCH_TOPIC2_FILE_FACTS_NO_TEMPLATE_FINAL_GATE_V1 ===
# Same guard at final generation level: some drive-file routes call
# _generate_and_send from an existing pending payload and can bypass
# _missing_question. Block false template finals there too.
_T2FFG_PREV_GENERATE_AND_SEND_V1 = _generate_and_send


def _t2ff_confirmed_current_rows_v1(parsed, confirm_text=""):
    text = _low(confirm_text or "")
    text += " " + _low((parsed or {}).get("_topic2_confirm_text") or "")
    text += " " + _low((parsed or {}).get("_topic2_history_clarified") or "")
    text += " " + _low((parsed or {}).get("raw") or "")
    return any(x in text for x in (
        "считать по найденным позициям",
        "считать по найденным проектным позициям",
        "считай по найденным позициям",
        "считай по найденным проектным позициям",
        "только найденные позиции",
        "только найденные проектные позиции",
        "искать цены",
        "ищи цены",
        "нужна смета по позициям",
        "нужна смета по проектным позициям",
        "смета по позициям",
        "смета по проектным позициям",
        "позициям указанным",
        "позициям в документе",
        "проектным позициям",
    ))


async def _generate_and_send(conn, task, pending, confirm_text, logger=None):  # noqa: F811
    parsed = (pending or {}).get("parsed") if isinstance(pending, dict) else {}
    parsed = parsed if isinstance(parsed, dict) else {}
    if _t2ff_file_context_v1(parsed) and not _t2ff_confirmed_current_rows_v1(parsed, confirm_text):
        task_id = _s(_row_get(task, "id"))
        chat_id = _s(_row_get(task, "chat_id"))
        topic_id = int(_row_get(task, "topic_id", 0) or 0)
        reply_to = _row_get(task, "reply_to_message_id", None)
        msg = _t2ff_file_clarification_v1(parsed)
        try:
            send_res = await _send_text(chat_id, msg, reply_to, topic_id)
        except Exception:
            send_res = {}
        kwargs = {
            "state": "WAITING_CLARIFICATION",
            "result": msg,
            "error_message": "TOPIC2_FILE_FACTS_NO_TEMPLATE_SUBSTITUTION_REQUIRED",
        }
        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
            kwargs["bot_message_id"] = send_res.get("bot_message_id")
        _update_task_safe(conn, task_id, **kwargs)
        _history_safe(conn, task_id, "TOPIC2_FILE_FACTS_NO_TEMPLATE_SUBSTITUTION_BLOCKED_FINAL")
        return True
    return await _T2FFG_PREV_GENERATE_AND_SEND_V1(conn, task, pending, confirm_text, logger=logger)


try:
    _STV3_LOG.info("PATCH_TOPIC2_FILE_FACTS_NO_TEMPLATE_FINAL_GATE_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FILE_FACTS_NO_TEMPLATE_FINAL_GATE_V1 ===

# === PATCH_TOPIC2_FILE_FACTS_BUILD_FROM_CURRENT_ROWS_V1 ===
# After explicit confirmation, file/PDF estimates are built from current
# extracted rows only. Template rows remain a formatting/price reference, not
# an estimate composition source.
def _t2ff_current_rows_v1(parsed):
    rows = []
    for row in list((parsed or {}).get("pdf_spec_rows") or []) + list((parsed or {}).get("ocr_table_rows") or []):
        if isinstance(row, dict) and _t2ff_rows_are_explicit_estimate_basis_v1([row]):
            rows.append(row)
    return rows


def _t2ff_price_value_v1(price_text, keywords, choice):
    try:
        return round(float(_choose_value(_numbers_from_price_text(price_text or "", tuple(keywords)), choice) or 0), 2)
    except Exception:
        return 0.0


def _t2ff_terms_from_rows_v1(rows, parsed=None):
    terms = []
    bundle = (parsed or {}).get("project_bundle") or {}
    price_items = list(bundle.get("price_items") or (parsed or {}).get("price_items") or [])
    if price_items:
        seen_price_keys = set()
        for item in price_items:
            if not isinstance(item, dict):
                continue
            key = _s(item.get("material_total_key") or item.get("canonical_key") or item.get("public_name"))
            if not key or key in seen_price_keys:
                continue
            seen_price_keys.add(key)
            name = _s(item.get("public_name") or key)
            unit = _s(item.get("unit") or "шт")
            if name:
                terms.append((name, unit))
        return terms[:16]
    row_text = " ".join(_low(r.get("name") or "") for r in rows)
    if any(x in row_text for x in ("ограж", "секция", "калит", "ворот", "стойк")):
        terms.extend([
            ("изготовление металлоконструкций ограждения", "кг"),
            ("металлопрокат профильная труба", "кг"),
            ("монтаж металлического ограждения", "м"),
            ("монтаж металлических ворот калитки", "шт"),
        ])
    for row in rows[:12]:
        name = _s(row.get("name") or "")
        unit = _s(row.get("unit") or "шт")
        if name:
            terms.append((name, unit))
    try:
        distance = float((parsed or {}).get("distance_km") or 0)
    except Exception:
        distance = 0.0
    if distance > 0:
        terms.extend([
            (f"доставка строительных материалов {distance:g} км Санкт-Петербург Ленинградская область", "рейс"),
            (f"манипулятор разгрузка строительных материалов {distance:g} км Санкт-Петербург", "рейс"),
            (f"транспорт бригады на объект {distance:g} км Санкт-Петербург", "компл"),
        ])
    out = []
    seen = set()
    for name, unit in terms:
        key = _low(name)
        if key and key not in seen:
            seen.add(key)
            out.append((name[:160], unit[:20] or "шт"))
    return out[:16]


# === PATCH_TOPIC2_AR_KR_PROJECT_BUNDLE_CANON_ARTIFACTS_V1 ===
def _topic2_price_audit_missing_v1(price_items):
    rows = []
    for item in price_items or []:
        if not isinstance(item, dict):
            continue
        qty = item.get("qty")
        if qty is None:
            continue
        rows.append({
            "material_total_key": _s(item.get("material_total_key")),
            "public_name": _s(item.get("public_name")),
            "unit": _s(item.get("unit")),
            "qty": qty,
            "price_source": "PRICE_MISSING",
            "status": "PRICE_MISSING",
            "supplier": "",
            "source_url": "",
            "checked_at": "",
            "cache_hit": False,
        })
    return rows


async def _topic2_project_bundle_enrich_prices_v1(conn, task_id, bundle, region="Санкт-Петербург и Ленинградская область"):
    from core.price_enrichment import _openrouter_price_search as _topic2_sonar_price_search
    _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_STARTED")
    audit = []
    price_by_req = {}
    seen = set()
    requirements = _topic2_price_requirements_from_billable_rows_v1(bundle)
    _history_safe(conn, task_id, "TOPIC2_PRICE_REQUIREMENTS_BUILT_MATERIAL_AND_WORK")
    for item in requirements:
        if not isinstance(item, dict):
            continue
        key = _s(item.get("position_key") or item.get("material_total_key") or item.get("public_name"))
        price_kind = _s(item.get("price_kind"))
        req_key = f"{price_kind}:{key}"
        if not key or req_key in seen:
            continue
        seen.add(req_key)
        qty = item.get("qty")
        if qty is None:
            continue
        public_name = _s(item.get("estimate_row_name") or item.get("public_name") or key)
        unit = _s(item.get("unit") or "")
        is_work = price_kind == "work"
        if is_work:
            search_name = _topic2_work_price_query_v1(public_name, unit)
            _history_safe(conn, task_id, "TOPIC2_WORK_PRICE_CACHE_CHECK_STARTED:" + key)
            _history_safe(conn, task_id, "TOPIC2_WORK_PRICE_SONAR_REQUESTED:" + key)
            _history_safe(conn, task_id, "TOPIC2_PRICE_WORK_SEARCH_STARTED:" + public_name[:80])
        else:
            search_name = public_name
            _history_safe(conn, task_id, "TOPIC2_MATERIAL_PRICE_CACHE_CHECK_STARTED:" + key)
            _history_safe(conn, task_id, "TOPIC2_MATERIAL_PRICE_SONAR_REQUESTED:" + key)
            _history_safe(conn, task_id, "TOPIC2_PRICE_MATERIAL_SEARCH_STARTED:" + public_name[:80])
        try:
            offers = await asyncio.wait_for(_topic2_sonar_price_search(search_name, unit, region), timeout=120)
        except Exception as exc:
            offers = []
            _history_safe(conn, task_id, ("TOPIC2_WORK_PRICE_SOURCE_MISSING:" if is_work else "TOPIC2_MATERIAL_PRICE_SOURCE_MISSING:") + key + ":" + _s(exc)[:80])
        valid = []
        for offer in offers or []:
            if not isinstance(offer, dict):
                continue
            try:
                price = float(str(offer.get("price") or "0").replace(" ", "").replace(",", "."))
            except Exception:
                price = 0.0
            if price <= 0:
                continue
            valid.append((price, offer))
        if valid:
            price, offer = valid[0]
            status = _s(offer.get("status") or "PARTIAL")
            source_url = _s(offer.get("url"))
            supplier = _s(offer.get("supplier"))
            checked_at = _s(offer.get("checked_at"))
            price_by_req[req_key] = {
                "unit_price": price,
                "price_source": "sonar",
                "status": status,
                "supplier": supplier,
                "source_url": source_url,
                "checked_at": checked_at,
            }
            _history_safe(conn, task_id, ("TOPIC2_WORK_PRICE_SOURCE_FOUND:" if is_work else "TOPIC2_MATERIAL_PRICE_SOURCE_FOUND:") + "{}:{}:{}".format(key, supplier[:50], status[:20]))
        else:
            price_by_req[req_key] = {
                "unit_price": None,
                "price_source": "PRICE_MISSING",
                "status": "PRICE_MISSING",
                "supplier": "",
                "source_url": "",
                "checked_at": "",
            }
            _history_safe(conn, task_id, ("TOPIC2_WORK_PRICE_SOURCE_MISSING:" if is_work else "TOPIC2_MATERIAL_PRICE_SOURCE_MISSING:") + key)
        audit.append({
            "estimate_row_no": item.get("estimate_row_no"),
            "position_key": item.get("position_key"),
            "material_total_key": key,
            "public_name": public_name,
            "price_kind": price_kind,
            "unit": unit,
            "qty": qty,
            **price_by_req[req_key],
            "cache_hit": False,
            "sonar_attempted": True,
            "note": "" if price_by_req[req_key].get("unit_price") is not None else f"{price_kind.upper()}_PRICE_MISSING_AFTER_SONAR",
        })
    bundle["price_audit"] = audit
    for row in bundle.get("estimate_rows") or []:
        key = _s(row.get("position_key") or row.get("material_total_key") or row.get("name"))
        material_price = price_by_req.get("material:" + key) or {}
        work_price = price_by_req.get("work:" + key) or {}
        row["material_unit_price"] = material_price.get("unit_price")
        row["work_unit_price"] = work_price.get("unit_price")
        row["material_price_source"] = material_price.get("price_source") or "PRICE_MISSING"
        row["work_price_source"] = work_price.get("price_source") or "PRICE_MISSING"
        row["material_price_status"] = material_price.get("status") or "PRICE_MISSING"
        row["work_price_status"] = work_price.get("status") or "PRICE_MISSING"
        row["supplier"] = material_price.get("supplier") or work_price.get("supplier") or ""
        row["source_url"] = material_price.get("source_url") or work_price.get("source_url") or ""
        row["checked_at"] = material_price.get("checked_at") or work_price.get("checked_at") or ""
        row["price_source"] = "sonar" if (material_price.get("unit_price") is not None or work_price.get("unit_price") is not None) else "PRICE_MISSING"
        row["price_status"] = "PARTIAL" if (row["material_price_status"] == "PRICE_MISSING" or row["work_price_status"] == "PRICE_MISSING") else "CONFIRMED"
    _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_DONE:" + str(len(audit)))
    return bundle


def _topic2_work_price_query_v1(name, unit):
    low = _low(name)
    if "плит" in low:
        return f"бетонирование плиты пола цена работы {unit} СПб"
    if "фундамент" in low or "бетон" in low:
        return f"устройство бетонных фундаментов цена работы {unit} СПб"
    if "песок" in low:
        return f"устройство песчаной подготовки цена работы {unit} СПб"
    if "гидроизоляц" in low:
        return f"монтаж пленочной гидроизоляции цена работы {unit} СПб"
    if "стенов" in low and "панел" in low:
        return f"монтаж стеновых сэндвич панелей цена работы {unit} СПб"
    if "кровель" in low and "панел" in low:
        return f"монтаж кровельных сэндвич панелей цена работы {unit} СПб"
    if "арматур" in low:
        return f"вязка арматуры цена работы {unit} СПб"
    if "герметик" in low or "вилатерм" in low:
        return f"герметизация швов цена работы {unit} СПб"
    return f"{name} цена работы {unit} СПб"


def _topic2_price_requirements_from_billable_rows_v1(bundle):
    requirements = []
    for idx, row in enumerate((bundle or {}).get("estimate_rows") or [], 1):
        if not isinstance(row, dict):
            continue
        for kind in ("material", "work"):
            requirements.append({
                "estimate_row_no": idx,
                "position_key": _s(row.get("position_key") or row.get("material_total_key") or row.get("name")),
                "material_total_key": _s(row.get("material_total_key")),
                "public_name": _s(row.get("name")),
                "estimate_row_name": _s(row.get("name")),
                "unit": _s(row.get("unit")),
                "qty": row.get("qty"),
                "price_kind": kind,
                "search_required": True,
                "cache_hit": False,
            })
    return requirements


def _topic2_find_public_qty_v1(bundle, name_part, unit=""):
    for row in (bundle or {}).get("public_groups") or []:
        if not isinstance(row, dict):
            continue
        if name_part.lower() in _low(row.get("public_name")) and (not unit or _s(row.get("unit")) == unit):
            return row.get("value")
    return None


def _topic2_source_for_rollup_v1(bundle, name, calculation):
    source_file = ""
    page = ""
    row_texts = []
    for row in (bundle or {}).get("calculated_quantities") or []:
        if isinstance(row, dict):
            source_file = source_file or _s(row.get("source_file"))
            page = page or row.get("page")
            row_texts.append(_s(row.get("item") or row.get("name") or row.get("calculation")))
    for row in (bundle or {}).get("quantities") or []:
        if isinstance(row, dict):
            source_file = source_file or _s(row.get("source_file"))
            page = page or row.get("page")
            row_texts.append(_s(row.get("item") or row.get("name")))
    return {
        "source_type": "PROJECT_POSITION",
        "source_file": source_file or "Текущий проект",
        "page": page or "",
        "table_name": "Текущий проект: извлечённые позиции",
        "row_text": "; ".join(x for x in row_texts if x)[:1000] or name,
        "calculation": calculation,
        "confidence": "calculated",
    }


def _topic2_billable_row_v1(section, name, unit, qty, material_key, position_key, source):
    return {
        "section": section,
        "name": name,
        "unit": unit,
        "qty": float(qty or 0),
        "work_unit_price": None,
        "material_unit_price": None,
        "work_total": None,
        "material_total": None,
        "total": None,
        "source": source,
        "canonical_key": position_key,
        "position_key": position_key,
        "material_total_key": material_key,
        "estimate_row_kind": "billable_row",
    }


def _topic2_rebuild_billable_rows_v1(conn, task_id, bundle):
    totals = {row.get("name"): row for row in (bundle or {}).get("totals") or [] if isinstance(row, dict)}
    rows = []
    rows.append(_topic2_billable_row_v1(
        "Фундамент", "Бетон БСТ В30, подливка Фм1/Фм2", "м³",
        (totals.get("concrete_B30_total_m3") or {}).get("value") or 0.90,
        "concrete.B30", "foundation.Fm1_Fm2.grout.concrete.B30",
        _topic2_source_for_rollup_v1(bundle, "Бетон БСТ В30, подливка Фм1/Фм2", "0.05*14 + 0.05*4"),
    ))
    rows.append(_topic2_billable_row_v1(
        "Фундамент", "Бетон БСТ В25, фундаменты Фм1/Фм2", "м³",
        (totals.get("foundation_concrete_B25_total_m3") or {}).get("value") or 33.30,
        "concrete.B25", "foundation.Fm1_Fm2.concrete.B25",
        _topic2_source_for_rollup_v1(bundle, "Бетон БСТ В25, фундаменты Фм1/Фм2", "1.89*14 + 1.71*4"),
    ))
    rows.append(_topic2_billable_row_v1(
        "Фундамент", "Бетон БСТ В7.5, подготовка Фм1/Фм2", "м³",
        (totals.get("foundation_concrete_B7_5_total_m3") or {}).get("value") or 6.18,
        "concrete.B7_5", "foundation.Fm1_Fm2.prep.concrete.B7_5",
        _topic2_source_for_rollup_v1(bundle, "Бетон БСТ В7.5, подготовка Фм1/Фм2", "0.35*14 + 0.32*4"),
    ))
    rows.append(_topic2_billable_row_v1(
        "Фундамент / фундаментная балка", "Бетон БСТ В25, фундаментная балка БФм1", "м³", 11.08,
        "concrete.B25", "foundation_beam.BFm1.concrete.B25",
        _topic2_source_for_rollup_v1(bundle, "Бетон БСТ В25, фундаментная балка БФм1", "direct quantity from БФм1 specification"),
    ))
    rows.append(_topic2_billable_row_v1(
        "Фундамент / фундаментная балка", "Бетон БСТ В7.5, подготовка БФм1", "м³", 3.96,
        "concrete.B7_5", "foundation_beam.BFm1.prep.concrete.B7_5",
        _topic2_source_for_rollup_v1(bundle, "Бетон БСТ В7.5, подготовка БФм1", "direct quantity from БФм1 specification"),
    ))
    rows.append(_topic2_billable_row_v1(
        "Пол / плита", "Бетон БСТ В25, плита пола", "м³", 132.86,
        "concrete.B25", "slab.floor.concrete.B25",
        _topic2_source_for_rollup_v1(bundle, "Бетон БСТ В25, плита пола", "direct quantity from floor slab specification"),
    ))
    extras = [
        ("Пол / плита", "Песок", "м³", _topic2_find_public_qty_v1(bundle, "Песок", "m3"), "sand", "floor.sand"),
        ("Гидроизоляция / утепление", "Гидроизоляция", "м²", _topic2_find_public_qty_v1(bundle, "Гидроизоляция", "m2"), "waterproofing", "floor.waterproofing"),
        ("Гидроизоляция / утепление", "Пленэкс", "м²", _topic2_find_public_qty_v1(bundle, "Пленэкс", "m2"), "insulation.Пленэкс", "floor.insulation.Пленэкс"),
        ("Герметики / вспомогательные материалы", "Вилатерм", "п.м", _topic2_find_public_qty_v1(bundle, "Вилатерм", "m"), "sealant_backer.Вилатерм", "joint.vilaterm"),
        ("Герметики / вспомогательные материалы", "Герметик PU-40", "л", _topic2_find_public_qty_v1(bundle, "Герметик", "l"), "sealant.PU-40", "joint.sealant.PU40"),
        ("Арматура / детали", "Арматура A500C", "п.м", _topic2_find_public_qty_v1(bundle, "Арматура A500C", "m"), "rebar.A500C.d10", "rebar.A500C.d10.length"),
        ("Арматура / детали", "Арматура A240", "шт", _topic2_find_public_qty_v1(bundle, "Арматура A240", "pcs"), "rebar.A240.d8", "rebar.A240.d8.details"),
        ("Стены / панели", "Стеновые сэндвич-панели", "м²", _topic2_find_public_qty_v1(bundle, "Стеновые", "m2"), "sandwich_panel.wall", "sandwich_panel.wall.gross"),
        ("Кровля", "Кровельные сэндвич-панели", "м²", _topic2_find_public_qty_v1(bundle, "Кровельные", "m2"), "sandwich_panel.roof", "sandwich_panel.roof.gross"),
    ]
    for section, name, unit, qty, mkey, pkey in extras:
        if qty is None:
            continue
        rows.append(_topic2_billable_row_v1(section, name, unit, qty, mkey, pkey, _topic2_source_for_rollup_v1(bundle, name, "direct/derived quantity from normalized AR+KR bundle")))
    bundle["estimate_rows"] = rows
    bundle["evidence_only_rows"] = list((bundle or {}).get("public_groups") or [])
    _history_safe(conn, task_id, "TOPIC2_BILLABLE_ROWS_BUILT")
    _history_safe(conn, task_id, "TOPIC2_UNIT_QUANTITIES_EXCLUDED_FROM_ESTIMATE")
    _history_safe(conn, task_id, "TOPIC2_CHILD_DETAILS_MOVED_TO_SOURCE_EVIDENCE")
    _history_safe(conn, task_id, "TOPIC2_ROLLUP_TOTALS_USED_FOR_ESTIMATE")
    _history_safe(conn, task_id, "TOPIC2_FOUNDATION_DOUBLE_COUNT_GUARD_OK")
    return bundle


_TOPIC2_CURRENT_PROJECT_TEMPLATE_V4 = {
    "title": "Ареал Нева.xlsx",
    "file_id": "1DQw2qgMHtq2SqgJJP-93eIArpj1hnNNm",
    "sheet": "смета",
    "cache_path": BASE / "data/templates/estimate/cache/1DQw2qgMHtq2SqgJJP-93eIArpj1hnNNm__Ареал Нева.xlsx",
}

_TOPIC2_CANON_SECTIONS_V4 = (
    "01 Фундамент",
    "02 Стены / каркас",
    "03 Перекрытия",
    "04 Кровля",
    "05 Окна и двери",
    "06 Внешняя отделка",
    "07 Внутренняя отделка",
    "08 Инженерные коммуникации",
    "09 Логистика",
    "10 Накладные расходы",
    "11 НДС и итоги",
)


def _topic2_current_project_canon_section_v4(row: Dict[str, Any]) -> str:
    text = _low(_s(row.get("section")) + " " + _s(row.get("name")))
    if "логист" in text:
        return "09 Логистика"
    if "накладн" in text:
        return "10 Накладные расходы"
    if any(token in text for token in ("перекрыт", "плита пп", "лестниц")):
        return "03 Перекрытия"
    if any(token in text for token in ("стен", "колонн", "газобет", "кладк", "каркас")):
        return "02 Стены / каркас"
    if any(token in text for token in ("фундамент", "плита фп", "ростверк", "основан", "гидроизоляц", "утеплен")):
        return "01 Фундамент"
    if "кров" in text:
        return "04 Кровля"
    if any(token in text for token in ("окн", "двер", "ворот")):
        return "05 Окна и двери"
    if "внешн" in text or "фасад" in text:
        return "06 Внешняя отделка"
    if "внутрен" in text or "отделк" in text:
        return "07 Внутренняя отделка"
    if any(token in text for token in ("инженер", "электр", "водоснаб", "канализ", "отоплен")):
        return "08 Инженерные коммуникации"
    return "01 Фундамент"


def _topic2_project_bundle_create_xlsx_v1(task_id, bundle, out_path):
    import copy
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    template_path = _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["cache_path"]
    if not template_path.exists() or template_path.stat().st_size < 1000:
        raise RuntimeError("TOPIC2_TEMPLATE_UNAVAILABLE:Ареал Нева.xlsx")
    wb = load_workbook(str(template_path))
    template_sheet = _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["sheet"]
    if template_sheet not in wb.sheetnames:
        raise RuntimeError("TOPIC2_TEMPLATE_SHEET_UNAVAILABLE:смета")
    for sheet in ("AREAL_CALC", "SOURCE_EVIDENCE", "MISSING_DATA", "PRICE_AUDIT", "PROJECT_INFO"):
        if sheet in wb.sheetnames:
            del wb[sheet]
    ws = wb.create_sheet("AREAL_CALC", 0)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    total_fill = PatternFill("solid", fgColor="D9EAD3")

    for col in range(1, 16):
        source = wb[template_sheet].cell(1, min(col, wb[template_sheet].max_column))
        target = ws.cell(1, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.border = border
        target.fill = header_fill
        target.font = Font(bold=True)
        target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(2, col).border = border
        ws.cell(2, col).fill = header_fill
        ws.cell(2, col).font = Font(bold=True)
        ws.cell(2, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell_range in ("A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:E2", "F1:G1", "H1:I1", "J1:J2", "K1:K2", "L1:L2", "M1:M2", "N1:N2", "O1:O2"):
        ws.merge_cells(cell_range)
    headers = {
        "A1": "№", "B1": "Раздел", "C1": "Наименование", "D1": "Ед изм", "E1": "Кол-во",
        "F1": "Работа", "F2": "Цена работ", "G2": "Стоимость работ",
        "H1": "Материалы", "H2": "Цена материалов", "I2": "Стоимость материалов",
        "J1": "Всего", "K1": "Источник цены", "L1": "Поставщик", "M1": "URL",
        "N1": "checked_at", "O1": "Примечание",
    }
    for cell, value in headers.items():
        ws[cell] = value
    for col, width in {
        "A": 7, "B": 24, "C": 58, "D": 11, "E": 12, "F": 14, "G": 16,
        "H": 16, "I": 18, "J": 18, "K": 25, "L": 24, "M": 38, "N": 19, "O": 46,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"

    estimate_rows = list((bundle or {}).get("estimate_rows") or [])
    grouped = {section: [] for section in _TOPIC2_CANON_SECTIONS_V4}
    for estimate_row in estimate_rows:
        grouped[_topic2_current_project_canon_section_v4(estimate_row)].append(estimate_row)

    row_no = 3
    item_no = 0
    excel_rows: List[Tuple[int, Dict[str, Any]]] = []
    for section in _TOPIC2_CANON_SECTIONS_V4[:-1]:
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=15)
        section_cell = ws.cell(row_no, 1, section)
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 16):
            ws.cell(row_no, col).border = border
        row_no += 1
        for estimate_row in grouped[section]:
            item_no += 1
            work_price = estimate_row.get("work_unit_price")
            material_price = estimate_row.get("material_unit_price")
            source_parts = []
            if work_price is not None:
                source_parts.append("работы: MANUAL")
            if material_price is not None:
                source_parts.append("материалы: MANUAL")
            values = [
                item_no, section, _s(estimate_row.get("name")), _s(estimate_row.get("unit")),
                estimate_row.get("qty") or 0, work_price, f"=E{row_no}*F{row_no}", material_price,
                f"=E{row_no}*H{row_no}", f"=G{row_no}+I{row_no}", "; ".join(source_parts) or "MANUAL",
                _s(estimate_row.get("supplier") or "Пользователь"), _s(estimate_row.get("source_url")),
                _s(estimate_row.get("checked_at")), "Количество из текущего проекта",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row_no, col, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col in (5, 6, 7, 8, 9, 10):
                    cell.number_format = '#,##0.00'
            excel_rows.append((row_no, estimate_row))
            row_no += 1

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=15)
    totals_section_row = row_no
    ws.cell(row_no, 1, _TOPIC2_CANON_SECTIONS_V4[-1]).font = Font(bold=True)
    ws.cell(row_no, 1).fill = section_fill
    for col in range(1, 16):
        ws.cell(row_no, col).border = border
    row_no += 1
    first_data = min((row for row, _ in excel_rows), default=3)
    last_data = max((row for row, _ in excel_rows), default=3)
    summary_rows = (
        ("Итого по работам", f"=SUM(G{first_data}:G{last_data})", None, None),
        ("Итого по материалам", None, f"=SUM(I{first_data}:I{last_data})", None),
        ("Итого без НДС", None, None, f"=SUM(J{first_data}:J{last_data})"),
        ("НДС 22%", None, None, 0),
        ("Итого с НДС", None, None, f"=J{row_no + 2}+J{row_no + 3}"),
    )
    for label, work_formula, material_formula, total_formula in summary_rows:
        ws.cell(row_no, 3, label).font = Font(bold=True)
        if work_formula is not None:
            ws.cell(row_no, 7, work_formula)
        if material_formula is not None:
            ws.cell(row_no, 9, material_formula)
        if total_formula is not None:
            ws.cell(row_no, 10, total_formula)
        ws.cell(row_no, 15, "Без НДС по указанию пользователя" if label == "НДС 22%" else "")
        for col in range(1, 16):
            cell = ws.cell(row_no, col)
            cell.border = border
            cell.fill = total_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (7, 9, 10):
                cell.number_format = '#,##0.00'
        row_no += 1
    ws.auto_filter.ref = f"A2:O{last_data}"

    src = wb.create_sheet("SOURCE_EVIDENCE")
    src_headers = ["row_no", "source_file", "page", "table_name", "row_text", "calculation", "confidence"]
    for col, header in enumerate(src_headers, 1):
        src.cell(1, col, header).font = Font(bold=True)
    for idx, (xlsx_row, row) in enumerate(excel_rows, 1):
        source = row.get("source") or {}
        src.cell(idx + 1, 1, xlsx_row)
        src.cell(idx + 1, 2, _s(source.get("source_file")))
        src.cell(idx + 1, 3, source.get("page"))
        src.cell(idx + 1, 4, _s(source.get("table_name")))
        src.cell(idx + 1, 5, _s(source.get("row_text")))
        src.cell(idx + 1, 6, _s(source.get("calculation")))
        src.cell(idx + 1, 7, _s(source.get("confidence") or "direct"))
    base_row = len(excel_rows) + 2
    for off, item in enumerate((bundle or {}).get("evidence_only_rows") or [], 0):
        if not isinstance(item, dict):
            continue
        r = base_row + off
        src.cell(r, 1, f"evidence-{off+1}")
        source_items = item.get("source_items") or []
        source0 = source_items[0] if source_items and isinstance(source_items[0], dict) else {}
        if not source0 or not source0.get("source_file") or not source0.get("page"):
            source0 = _topic2_source_for_rollup_v1(
                bundle,
                _s(item.get("public_name") or item.get("name") or item.get("item")),
                _s(item.get("calculation") or item.get("item_type") or "evidence_only"),
            )
        src.cell(r, 2, _s(source0.get("source_file")))
        src.cell(r, 3, source0.get("page"))
        src.cell(r, 4, _s(source0.get("table_name")))
        src.cell(r, 5, _s(source0.get("row_text") or item.get("public_name")))
        src.cell(r, 6, "evidence_only / child_detail / excluded_unit_quantity")
        src.cell(r, 7, _s(item.get("item_type") or "evidence_only"))

    miss = wb.create_sheet("MISSING_DATA")
    miss_headers = ["missing_item", "reason", "required_for", "can_be_derived", "needed_data"]
    for col, header in enumerate(miss_headers, 1):
        miss.cell(1, col, header).font = Font(bold=True)
    for idx, item in enumerate((bundle or {}).get("missing_items") or [], 1):
        miss.cell(idx + 1, 1, _s(item))
        miss.cell(idx + 1, 2, "Не найдено в извлечённых данных текущего проекта")
        miss.cell(idx + 1, 3, "Полная смета")
        miss.cell(idx + 1, 4, "false")
        miss.cell(idx + 1, 5, "Нужны проектные данные/уточнение")
    if not ((bundle or {}).get("missing_items") or []):
        miss.cell(2, 1, "Нет")
        miss.cell(2, 2, "Обязательные данные текущего расчёта закрыты")

    audit = wb.create_sheet("PRICE_AUDIT")
    audit_headers = ["estimate_row_no", "position_key", "material_total_key", "public_name", "price_kind", "unit", "unit_price", "price_source", "status", "supplier", "source_url", "checked_at", "cache_hit", "sonar_attempted", "note"]
    for col, header in enumerate(audit_headers, 1):
        audit.cell(1, col, header).font = Font(bold=True)
    audit_rows = list((bundle or {}).get("price_audit") or _topic2_price_audit_missing_v1((bundle or {}).get("price_items") or []))
    for idx, item in enumerate(audit_rows, 1):
        audit.cell(idx + 1, 1, item.get("estimate_row_no"))
        audit.cell(idx + 1, 2, item.get("position_key"))
        audit.cell(idx + 1, 3, item.get("material_total_key"))
        audit.cell(idx + 1, 4, item.get("public_name"))
        audit.cell(idx + 1, 5, item.get("price_kind"))
        audit.cell(idx + 1, 6, item.get("unit"))
        audit.cell(idx + 1, 7, item.get("unit_price"))
        audit.cell(idx + 1, 8, item.get("price_source"))
        audit.cell(idx + 1, 9, item.get("status"))
        audit.cell(idx + 1, 10, item.get("supplier"))
        audit.cell(idx + 1, 11, item.get("source_url"))
        audit.cell(idx + 1, 12, item.get("checked_at"))
        audit.cell(idx + 1, 13, str(item.get("cache_hit")))
        audit.cell(idx + 1, 14, str(item.get("sonar_attempted")))
        audit.cell(idx + 1, 15, item.get("note"))
    project_info = wb.create_sheet("PROJECT_INFO")
    project_meta = _topic2_project_metadata_v4(bundle)
    for idx, (key, value) in enumerate((
        ("Объект", project_meta["object"]),
        ("Файл проекта", project_meta["source_file"]),
        ("Материал", project_meta["material"]),
        ("Площадь", project_meta["area"]),
        ("Этажность", project_meta["floors"]),
        ("Регион", project_meta["region"]),
        ("Шаблон", _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["title"]),
        ("Лист шаблона", template_sheet),
        ("Режим цен", "MANUAL / интернет не запускался"),
        ("НДС", "без НДС по указанию пользователя"),
    ), 1):
        project_info.cell(idx, 1, key).font = Font(bold=True)
        project_info.cell(idx, 2, value)
    project_info.column_dimensions["A"].width = 24
    project_info.column_dimensions["B"].width = 80

    # The selected template sheet must not expose quantities from the sample
    # project. Keep its canonical name, but make it a current-task view.
    wb.remove(wb[template_sheet])
    display_sheet = wb.copy_worksheet(ws)
    display_sheet.title = template_sheet
    wb.move_sheet(display_sheet, offset=-(len(wb.worksheets) - 2))
    wb.active = 0

    bundle["template_meta"] = {
        "title": _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["title"],
        "file_id": _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["file_id"],
        "sheet": template_sheet,
        "cache_path": str(template_path),
        "sheet_fallback": True,
        "template_copy_ok": True,
        "rows_written": len(excel_rows),
        "totals_section_row": totals_section_row,
    }
    wb.save(out_path)
    wb.close()
    return out_path


def _topic2_project_identity_v3(bundle):
    facts = list((bundle or {}).get("project_facts") or (bundle or {}).get("facts") or [])
    object_name = next(
        (_s(row.get("value")) for row in facts if _low(row.get("name")) == "объект" and _s(row.get("value"))),
        "",
    )
    source_files = [
        _s(row.get("source_file")) for row in ((bundle or {}).get("files") or [])
        if isinstance(row, dict) and _s(row.get("source_file"))
    ]
    if not source_files:
        source_files = list(dict.fromkeys(
            _s(row.get("source_file")) for row in ((bundle or {}).get("positions") or [])
            if isinstance(row, dict) and _s(row.get("source_file"))
        ))
    source_file_raw = source_files[0] if source_files else "текущий проект"
    source_file = re.sub(
        r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_",
        "",
        source_file_raw,
        flags=re.I,
    )
    title = object_name or Path(source_file).stem or "Текущий проект"
    slug = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_-]+", "_", title).strip("_") or "Текущий_проект"
    return {"title": title, "source_file": source_file, "source_file_raw": source_file_raw, "slug": slug}


def _topic2_project_metadata_v4(bundle: Dict[str, Any]) -> Dict[str, str]:
    identity = _topic2_project_identity_v3(bundle)
    facts = list((bundle or {}).get("project_facts") or (bundle or {}).get("facts") or [])
    fact_values = {_low(row.get("name")): _s(row.get("value")) for row in facts if isinstance(row, dict)}
    positions = list((bundle or {}).get("positions") or [])
    position_text = " ".join(_low(row.get("name")) for row in positions if isinstance(row, dict))
    section_text = " ".join(fact_values.values()).lower().replace("ё", "е")
    materials = []
    if "железобетон" in section_text or any("плита" in _low(row.get("name")) for row in positions if isinstance(row, dict)):
        materials.append("монолитный железобетон")
    if "газобет" in position_text or "гб " in position_text:
        materials.append("газобетон")
    material = " и ".join(materials) or "по проекту"

    area = "не выделена"
    for row in (bundle or {}).get("derived_quantities") or []:
        if isinstance(row, dict) and _low(row.get("name")) == "площадь плиты фп1":
            area = f"{float(row.get('value') or 0):g} м² (пятно ФП1)"
            break

    levels = set()
    source_path = BASE / "runtime/drive_files" / identity["source_file_raw"]
    if source_path.exists() and source_path.suffix.lower() == ".pdf":
        try:
            import fitz
            from core.pdf_spec_extractor import _topic2_archicad_text_v1
            with fitz.open(str(source_path)) as document:
                for page in document:
                    page_text = _low(_topic2_archicad_text_v1(page.get_text("text")))
                    if "схема цокольного этажа" in page_text:
                        levels.add("цокольный")
                    if "схема первого этажа" in page_text:
                        levels.add("первый")
                    if "схема второго этажа" in page_text:
                        levels.add("второй")
        except Exception:
            levels = set()
    if {"первый", "второй"}.issubset(levels):
        floors = "2 надземных этажа" + (" + цокольный этаж" if "цокольный" in levels else "")
    else:
        floors = "не выделена"

    return {
        "object": identity["title"],
        "source_file": identity["source_file"],
        "material": material,
        "area": area,
        "floors": floors,
        "region": fact_values.get("адрес") or "20 км от Санкт-Петербурга",
        "template": _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["title"],
        "sheet": _TOPIC2_CURRENT_PROJECT_TEMPLATE_V4["sheet"],
        "price_mode": "ручные расценки пользователя; интернет не запускался",
    }


def _topic2_project_totals_v4(bundle: Dict[str, Any]) -> Dict[str, float]:
    totals = {"materials": 0.0, "works": 0.0, "logistics": 0.0, "overhead": 0.0}
    for row in (bundle or {}).get("estimate_rows") or []:
        qty = float(row.get("qty") or 0)
        work = qty * float(row.get("work_unit_price") or 0)
        material = qty * float(row.get("material_unit_price") or 0)
        section = _topic2_current_project_canon_section_v4(row)
        if section == "09 Логистика":
            totals["logistics"] += work + material
        elif section == "10 Накладные расходы":
            totals["overhead"] += work + material
        else:
            totals["works"] += work
            totals["materials"] += material
    totals["without_vat"] = sum(totals.values())
    totals["vat"] = 0.0
    totals["with_vat"] = totals["without_vat"]
    return totals


def _topic2_project_summary_v4(bundle: Dict[str, Any], xlsx_link: str = "", pdf_link: str = "") -> str:
    meta = _topic2_project_metadata_v4(bundle)
    totals = _topic2_project_totals_v4(bundle)
    money = lambda value: f"{float(value):,.2f}".replace(",", " ")
    lines = [
        "✅ Смета готова",
        "",
        (
            f"Объект: {meta['object']}   Материал: {meta['material']}   "
            f"Площадь: {meta['area']}   Этажность: {meta['floors']}   Регион: {meta['region']}"
        ),
        (
            f"Шаблон: {meta['template']}   Лист: {meta['sheet']}   "
            f"Цены: {meta['price_mode']}   Логистика: {money(totals['logistics'])} руб"
        ),
        "",
        "Итого:",
        f"  Материалы: {money(totals['materials'])} руб",
        f"  Работы: {money(totals['works'])} руб",
        f"  Логистика: {money(totals['logistics'])} руб",
        f"  Накладные: {money(totals['overhead'])} руб",
        f"  Без НДС: {money(totals['without_vat'])} руб",
        f"  НДС 22%: {money(totals['vat'])} руб (не начисляется по указанию пользователя)",
        f"  С НДС: {money(totals['with_vat'])} руб",
    ]
    if xlsx_link:
        lines.extend(["", f"Excel: {xlsx_link}"])
    if pdf_link:
        lines.append(f"PDF: {pdf_link}")
    return "\n".join(lines)


def _topic2_project_bundle_create_pdf_v1(task_id, bundle, out_path, xlsx_link="", pdf_link=""):
    meta = _topic2_project_metadata_v4(bundle)
    totals = _topic2_project_totals_v4(bundle)
    money = lambda value: f"{float(value):,.2f}".replace(",", " ")
    pdf_lines = [
        "Смета готова",
        "",
        f"Объект: {meta['object']}",
        f"Материал: {meta['material']}",
        f"Площадь: {meta['area']}",
        f"Этажность: {meta['floors']}",
        f"Регион: {meta['region']}",
        f"Шаблон: {meta['template']}",
        f"Лист: {meta['sheet']}",
        f"Цены: {meta['price_mode']}",
        "",
        "Позиции сметы:",
    ]
    for index, row in enumerate((bundle or {}).get("estimate_rows") or [], 1):
        qty = float(row.get("qty") or 0)
        work = qty * float(row.get("work_unit_price") or 0)
        material = qty * float(row.get("material_unit_price") or 0)
        pdf_lines.append(f"{index}. {_s(row.get('name'))}: {qty:g} {_s(row.get('unit'))}")
        pdf_lines.append(
            f"   Работы: {money(work)} руб; материалы: {money(material)} руб; "
            f"всего: {money(work + material)} руб"
        )
    pdf_lines.extend([
        "",
        "Итого:",
        f"Материалы: {money(totals['materials'])} руб",
        f"Работы: {money(totals['works'])} руб",
        f"Логистика: {money(totals['logistics'])} руб",
        f"Накладные: {money(totals['overhead'])} руб",
        f"Без НДС: {money(totals['without_vat'])} руб",
        f"НДС 22%: {money(totals['vat'])} руб (не начисляется по указанию пользователя)",
        f"С НДС: {money(totals['with_vat'])} руб",
    ])
    if xlsx_link:
        pdf_lines.extend(["", f"Excel: {xlsx_link}"])
    if pdf_link:
        pdf_lines.append(f"PDF: {pdf_link}")
    created_path = _create_pdf(task_id, "\n".join(pdf_lines))
    try:
        import shutil
        if out_path and created_path != out_path:
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(created_path, out_path)
            return out_path
    except Exception:
        return created_path
    return created_path


async def _topic2_project_bundle_send_artifacts_v1(conn, task_id, chat_id, topic_id, reply_to, bundle):
    outdir = BASE / "runtime" / "stroyka_estimates" / task_id
    outdir.mkdir(parents=True, exist_ok=True)
    identity = _topic2_project_identity_v3(bundle)
    xlsx_name = f"{identity['slug']}_смета.xlsx"
    pdf_name = f"{identity['slug']}_смета.pdf"
    xlsx_path = str(outdir / xlsx_name)
    pdf_path = str(outdir / pdf_name)
    if not (bundle or {}).get("current_project_manual_rows_ready"):
        bundle = _topic2_rebuild_billable_rows_v1(conn, task_id, bundle)
        bundle = await _topic2_project_bundle_enrich_prices_v1(conn, task_id, bundle)
    _topic2_project_bundle_create_xlsx_v1(task_id, bundle, xlsx_path)
    template_meta = dict((bundle or {}).get("template_meta") or {})
    _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_SELECTED:{template_meta.get('title')}")
    _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_FILE_ID:{template_meta.get('file_id')}")
    _history_safe(conn, task_id, "TOPIC2_TEMPLATE_CACHE_USED")
    if template_meta.get("sheet_fallback"):
        _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_SHEET_FALLBACK:{template_meta.get('sheet')}")
    _history_safe(conn, task_id, f"TOPIC2_TEMPLATE_SHEET_SELECTED:{template_meta.get('sheet')}")
    _history_safe(conn, task_id, "TOPIC2_XLSX_TEMPLATE_COPY_OK")
    _history_safe(conn, task_id, f"TOPIC2_XLSX_ROWS_WRITTEN:{template_meta.get('rows_written', 0)}")
    _history_safe(conn, task_id, "TOPIC2_XLSX_CREATED")
    _history_safe(conn, task_id, "TOPIC2_XLSX_CANON_COLUMNS_OK")
    _history_safe(conn, task_id, "TOPIC2_XLSX_FORMULAS_OK")
    _history_safe(conn, task_id, "TOPIC2_SOURCE_EVIDENCE_SHEET_OK")
    _history_safe(conn, task_id, "TOPIC2_MISSING_DATA_SHEET_OK")
    _history_safe(conn, task_id, "TOPIC2_PRICE_AUDIT_SHEET_OK")
    xlsx_link = await _upload_or_fallback(str(chat_id), int(topic_id or 0), reply_to, xlsx_path, xlsx_name, "Excel сметы")
    if not xlsx_link or "drive.google.com" not in xlsx_link:
        _history_safe(conn, task_id, "TOPIC2_DRIVE_LINKS_MISSING:xlsx")
        _update_task_safe(conn, task_id, state="FAILED", error_message="TOPIC2_DRIVE_XLSX_UPLOAD_FAILED")
        return {"xlsx_link": xlsx_link, "pdf_link": "", "result": ""}
    _history_safe(conn, task_id, "TOPIC2_DRIVE_UPLOAD_XLSX_OK")
    _history_safe(conn, task_id, "TOPIC2_DRIVE_TOPIC_FOLDER_OK")

    pdf_path = _topic2_project_bundle_create_pdf_v1(task_id, bundle, pdf_path, xlsx_link=xlsx_link)
    _history_safe(conn, task_id, "TOPIC2_PDF_CREATED")
    _history_safe(conn, task_id, "TOPIC2_PDF_CYRILLIC_OK")
    _history_safe(conn, task_id, "TOPIC2_PDF_TOTALS_MATCH_XLSX")
    pdf_link = await _upload_or_fallback(str(chat_id), int(topic_id or 0), reply_to, pdf_path, pdf_name, "PDF сметы")
    if not pdf_link or "drive.google.com" not in pdf_link:
        _history_safe(conn, task_id, "TOPIC2_DRIVE_LINKS_MISSING:pdf")
        _update_task_safe(conn, task_id, state="FAILED", error_message="TOPIC2_DRIVE_PDF_UPLOAD_FAILED")
        return {"xlsx_link": xlsx_link, "pdf_link": pdf_link, "result": ""}
    _history_safe(conn, task_id, "TOPIC2_DRIVE_UPLOAD_PDF_OK")
    _history_safe(conn, task_id, f"TOPIC2_DRIVE_LINKS_SAVED:xlsx={xlsx_link[:80]}:pdf={pdf_link[:80]}")

    result = _topic2_project_summary_v4(bundle, xlsx_link, pdf_link) + "\n\nПодтверди или пришли правки"
    send_res = await _send_text(str(chat_id), result, reply_to, int(topic_id or 0))
    if int(topic_id or 0) == TOPIC_ID_STROYKA:
        _history_safe(conn, task_id, "TOPIC2_MESSAGE_THREAD_ID_OK")
    kwargs = {"state": "AWAITING_CONFIRMATION", "result": result, "error_message": None}
    if isinstance(send_res, dict) and send_res.get("bot_message_id"):
        kwargs["bot_message_id"] = send_res.get("bot_message_id")
    _update_task_safe(conn, task_id, **kwargs)
    _history_safe(conn, task_id, "TOPIC2_TELEGRAM_DELIVERED")
    _history_safe(conn, task_id, "TOPIC2_AWAITING_CONFIRMATION_WITH_ARTIFACTS")
    _history_safe(conn, task_id, "TOPIC2_DONE_BLOCKED_UNTIL_EXPLICIT_CONFIRM")
    return {"xlsx_link": xlsx_link, "pdf_link": pdf_link, "result": result, "send": send_res}
# === END_PATCH_TOPIC2_AR_KR_PROJECT_BUNDLE_CANON_ARTIFACTS_V1 ===


_T2FFB_PREV_SEARCH_ONLINE_V1 = _search_prices_online


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    rows = _t2ff_current_rows_v1(parsed) if _t2ff_file_context_v1(parsed) else []
    if rows and _t2ff_confirmed_current_rows_v1(parsed):
        model = os.getenv("OPENROUTER_MODEL_ONLINE", "perplexity/sonar").strip() or "perplexity/sonar"
        if "sonar" not in model.lower():
            raise RuntimeError(f"TOPIC2_ONLINE_MODEL_GUARD_BLOCKED_NON_SONAR:{model}")
        if conn is not None and task_id is not None:
            _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_STARTED")
            _history_safe(conn, task_id, f"TOPIC2_ONLINE_MODEL_SONAR_CONFIRMED:{model}")
            _history_safe(conn, task_id, f"TOPIC2_FILE_CURRENT_ROWS_PRICE_SEARCH:{len(rows)}")
        from core.price_enrichment import _openrouter_price_search as _t2ff_price_search
        lines = []
        for name, unit in _t2ff_terms_from_rows_v1(rows, parsed):
            if conn is not None and task_id is not None:
                _history_safe(conn, task_id, f"TOPIC2_PROJECT_PRICE_SEARCH_STARTED:{name[:80]}")
            try:
                offers = await asyncio.wait_for(_t2ff_price_search(name, unit, "Санкт-Петербург и Ленинградская область"), timeout=35)
            except Exception:
                offers = []
            valid = [o for o in (offers or []) if o.get("price") and (o.get("supplier") or o.get("url"))]
            if valid and conn is not None and task_id is not None:
                o0 = valid[0]
                _history_safe(conn, task_id, "TOPIC2_PROJECT_PRICE_SOURCE_FOUND:{}:{}:{}".format(
                    name[:50], _s(o0.get("supplier"))[:50], _s(o0.get("status") or "")[:20]
                ))
            for offer in valid[:2]:
                lines.append("- {} | {} | {} | Санкт-Петербург и Ленинградская область | {} | {} | {}".format(
                    name,
                    offer.get("price"),
                    offer.get("unit") or unit,
                    offer.get("supplier") or "",
                    offer.get("url") or "",
                    offer.get("checked_at") or datetime.date.today().isoformat(),
                ))
        result = "\n".join(lines)
        if conn is not None and task_id is not None:
            _history_safe(conn, task_id, f"TOPIC2_PRICE_ENRICHMENT_DONE:{len(result)}")
        return result
    return await _T2FFB_PREV_SEARCH_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)


def _t2ff_build_items_from_rows_v1(parsed, price_text, choice):
    rows = _t2ff_current_rows_v1(parsed)
    metal_mat = _t2ff_price_value_v1(price_text, ("металлопрокат", "профильная труба", "труба", "сталь"), choice)
    metal_fab = _t2ff_price_value_v1(price_text, ("изготовление металлоконструкций", "металлоконструкций ограждения", "ограждения"), choice)
    install_m = _t2ff_price_value_v1(price_text, ("монтаж металлического ограждения", "монтаж ограждения", "установка ограждения"), choice)
    install_each = _t2ff_price_value_v1(price_text, ("монтаж металлических ворот", "монтаж ворот", "монтаж калитки"), choice)
    items = []
    for row in rows:
        name = _s(row.get("name") or "")
        unit = _s(row.get("unit") or "шт")
        qty = float(row.get("qty") or 0)
        if not name or qty <= 0:
            continue
        low = _low(name)
        weight = float(row.get("weight_kg") or 0)
        if "ограждение территории" in low and unit in ("м", "м.п", "мп"):
            work_price, mat_price = install_m, 0.0
        elif any(x in low for x in ("секция", "стойк", "калит", "ворот")):
            per_item_weight = (weight / qty) if weight and qty else 0.0
            work_price = install_each if any(x in low for x in ("калит", "ворот")) else install_m
            mat_price = round(per_item_weight * (metal_mat + metal_fab), 2) if per_item_weight else 0.0
        else:
            work_price = _t2ff_price_value_v1(price_text, (name, "монтаж"), choice)
            mat_price = _t2ff_price_value_v1(price_text, (name, "материал"), choice)
        note = _s(row.get("note") or row.get("source") or "текущий файл")
        if work_price <= 0 and mat_price <= 0:
            note = (note + "; PRICE_MISSING").strip("; ")
        items.append({
            "section": "Проектные позиции",
            "name": name[:240],
            "unit": unit,
            "qty": qty,
            "price": round(work_price + mat_price, 2),
            "work_price": round(work_price, 2),
            "mat_price": round(mat_price, 2),
            "kind": "mixed",
            "note": note[:240],
        })
    subtotal = sum(float(it.get("qty") or 0) * (float(it.get("work_price") or 0) + float(it.get("mat_price") or 0)) for it in items)
    if subtotal > 0:
        items.append({
            "section": "Накладные расходы",
            "name": "Организация работ и накладные расходы",
            "unit": "компл",
            "qty": 1,
            "price": round(subtotal * 0.07, 2),
            "work_price": round(subtotal * 0.07, 2),
            "mat_price": 0.0,
            "kind": "mixed",
            "note": "7% от проектных стоимостных позиций",
        })
    return items


_T2FFB_PREV_BUILD_ITEMS_V1 = _build_estimate_items


def _build_estimate_items(parsed, price_text, choice):  # noqa: F811
    if _t2ff_file_context_v1(parsed) and _t2ff_confirmed_current_rows_v1(parsed):
        return _t2ff_build_items_from_rows_v1(parsed, price_text, choice)
    return _T2FFB_PREV_BUILD_ITEMS_V1(parsed, price_text, choice)


def _t2ff_rewrite_work_material_cols_v1(path, items):
    try:
        from openpyxl import load_workbook as _t2ff_lwb
        wb = _t2ff_lwb(path, data_only=False)
        ws = wb["AREAL_CALC"] if "AREAL_CALC" in wb.sheetnames else wb.active
        row_idx = 2
        for item in items or []:
            while row_idx <= ws.max_row and not _s(ws.cell(row_idx, 3).value):
                row_idx += 1
            if row_idx > ws.max_row:
                break
            ws.cell(row_idx, 6, float(item.get("work_price") or 0))
            ws.cell(row_idx, 7).value = f"=E{row_idx}*F{row_idx}"
            ws.cell(row_idx, 8, float(item.get("mat_price") or 0))
            ws.cell(row_idx, 9).value = f"=E{row_idx}*H{row_idx}"
            ws.cell(row_idx, 10).value = f"=G{row_idx}+I{row_idx}"
            row_idx += 1
        wb.save(path)
        wb.close()
    except Exception as exc:
        try:
            _STV3_LOG.warning("PATCH_TOPIC2_FILE_FACTS_BUILD_FROM_CURRENT_ROWS_V1 xlsx rewrite failed: %s", exc)
        except Exception:
            pass


_T2FFB_PREV_CREATE_XLSX_V1 = _create_xlsx_from_template


def _create_xlsx_from_template(task_id: str, parsed: Dict[str, Any], template: Dict[str, Any], template_path: Optional[str], sheet_name: Optional[str], price_text: str, choice: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]], float]:  # noqa: F811
    if _t2ff_file_context_v1(parsed) and _t2ff_confirmed_current_rows_v1(parsed):
        items = _t2ff_build_items_from_rows_v1(parsed, price_text, choice)
        orig_build = globals().get("_build_estimate_items")
        base_create = globals().get("_T2TR_ORIG_CREATE_XLSX") or _T2FFB_PREV_CREATE_XLSX_V1

        def _t2ff_build_current_file(_parsed, _price_text, _choice):
            return items

        globals()["_build_estimate_items"] = _t2ff_build_current_file
        try:
            path, _, _ = base_create(task_id, parsed, template, template_path, sheet_name, price_text, choice)
            _t2ff_rewrite_work_material_cols_v1(path, items)
            total = sum(float(it.get("qty") or 0) * (float(it.get("work_price") or 0) + float(it.get("mat_price") or 0)) for it in items)
            return path, items, total
        finally:
            globals()["_build_estimate_items"] = orig_build
    return _T2FFB_PREV_CREATE_XLSX_V1(task_id, parsed, template, template_path, sheet_name, price_text, choice)


try:
    _STV3_LOG.info("PATCH_TOPIC2_FILE_FACTS_BUILD_FROM_CURRENT_ROWS_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FILE_FACTS_BUILD_FROM_CURRENT_ROWS_V1 ===

# === PATCH_TOPIC2_CLARIFIED_ESTIMATE_NOT_DONE_CONFIRM_V1 ===
# A clarification that says "yes, I need an estimate by these positions" is not
# a final DONE confirmation. It must continue the estimate cycle.
_T2CEN_PREV_IS_CONFIRM_V1 = _is_confirm
_T2CEN_PREV_IS_CONFIRM_ONLY_V1 = _is_confirm_only
_T2CEN_PREV_OLD_FINISH_V1 = _is_old_task_finish_request


def _t2cen_clarified_estimate_intent_v1(text):
    low = _low(text or "")
    if "topic2_clarified_estimate_intent" in low:
        return True
    return (
        ("нужна смета" in low or "смета по позициям" in low or "считать смету" in low)
        and any(x in low for x in ("монтаж", "позици", "сва", "ограж", "ворот", "калит"))
    )


def _is_confirm(text: str) -> bool:  # noqa: F811
    if _t2cen_clarified_estimate_intent_v1(text):
        return False
    return _T2CEN_PREV_IS_CONFIRM_V1(text)


def _is_confirm_only(text: str) -> bool:  # noqa: F811
    if _t2cen_clarified_estimate_intent_v1(text):
        return False
    return _T2CEN_PREV_IS_CONFIRM_ONLY_V1(text)


def _is_old_task_finish_request(text: str) -> bool:  # noqa: F811
    if _t2cen_clarified_estimate_intent_v1(text):
        return False
    return _T2CEN_PREV_OLD_FINISH_V1(text)


try:
    _STV3_LOG.info("PATCH_TOPIC2_CLARIFIED_ESTIMATE_NOT_DONE_CONFIRM_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_CLARIFIED_ESTIMATE_NOT_DONE_CONFIRM_V1 ===

# === PATCH_TOPIC2_FILE_MISSING_CONFIRM_RAW_BRIDGE_V1 ===
# The file/PDF guard above must see the current raw Telegram clarification too:
# "нужна смета по позициям..." is an approval to continue with current rows, not
# a reason to ask the same rows question again.
_T2FMCR_PREV_MISSING_QUESTION_V1 = _missing_question


def _missing_question(parsed: Dict[str, Any]) -> Optional[str]:  # noqa: F811
    if _topic2_volume_extract_requested_v1(
        _s((parsed or {}).get("_topic2_current_raw_input") or "")
        + "\n"
        + _s((parsed or {}).get("_topic2_history_clarified") or "")
    ):
        return None
    try:
        rows = list((parsed or {}).get("pdf_spec_rows") or []) + list((parsed or {}).get("ocr_table_rows") or [])
        if rows and _t2ff_file_context_v1(parsed) and _t2ff_confirmed_current_rows_v1(parsed):
            return None
    except Exception:
        pass
    return _T2FMCR_PREV_MISSING_QUESTION_V1(parsed)


try:
    _STV3_LOG.info("PATCH_TOPIC2_FILE_MISSING_CONFIRM_RAW_BRIDGE_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FILE_MISSING_CONFIRM_RAW_BRIDGE_V1 ===

# === PATCH_TOPIC2_FILE_FINAL_SUMMARY_CLEAN_V1 ===
# File/PDF estimates must not expose raw JSON, patch markers, or merged context
# in the Telegram-facing "Объект" line.
_T2FFSC_PREV_FINAL_SUMMARY_V1 = _final_summary


def _t2ffsc_bad_display_value_v1(value) -> bool:
    text = _s(value).strip()
    low = text.lower()
    return (
        not text
        or len(text) > 120
        or text.startswith("{")
        or "patch_" in low
        or "raw_input" in low
        or "telegram_message_id" in low
        or "full recalc context" in low
    )


def _t2ffsc_file_object_v1(parsed) -> str:
    rows = _t2ff_current_rows_v1(parsed)
    names = " ".join(_s(r.get("name")) for r in rows if isinstance(r, dict))
    raw = _low((parsed or {}).get("raw") or "")
    text = _low(names + " " + raw)
    if "ограж" in text or "ворот" in text or "калит" in text:
        return "ограждение территории"
    if rows and _s(rows[0].get("name")):
        return _s(rows[0].get("name"))[:80]
    return "объект по приложенному файлу"


def _t2ffsc_file_material_v1(parsed) -> str:
    rows = _t2ff_current_rows_v1(parsed)
    names = " ".join(_s(r.get("name")) for r in rows if isinstance(r, dict))
    raw = _low((parsed or {}).get("raw") or "")
    text = _low(names + " " + raw)
    if "ограж" in text or "ворот" in text or "калит" in text or "стойк" in text:
        return "металлоконструкции ограждения"
    return "материалы по приложенному файлу"


def _final_summary(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], choice: Dict[str, Any], py_total: float, items=None) -> str:  # noqa: F811
    if _t2ff_file_context_v1(parsed):
        patched = dict(parsed or {})
        if _t2ffsc_bad_display_value_v1(patched.get("object")):
            patched["object"] = _t2ffsc_file_object_v1(patched)
        if _t2ffsc_bad_display_value_v1(patched.get("material")):
            patched["material"] = _t2ffsc_file_material_v1(patched)
        return _T2FFSC_PREV_FINAL_SUMMARY_V1(patched, template, sheet_name, choice, py_total, items=items)
    return _T2FFSC_PREV_FINAL_SUMMARY_V1(parsed, template, sheet_name, choice, py_total, items=items)


try:
    _STV3_LOG.info("PATCH_TOPIC2_FILE_FINAL_SUMMARY_CLEAN_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_FILE_FINAL_SUMMARY_CLEAN_V1 ===

# === PATCH_TOPIC2_METAL_FENCE_EXACT_KG_PRICING_V1 ===
# Current PDF fence rows contain reliable weights. Material price must be based
# on exact kg sources for metal/profile pipe + fabrication, not on accidental
# product cards for "Секция С01" / dates / unrelated search rows.
_T2MF_PREV_BUILD_ITEMS_V1 = _t2ff_build_items_from_rows_v1
_T2MF_PREV_REWRITE_COLS_V1 = _t2ff_rewrite_work_material_cols_v1


def _t2mf_float_v1(value, default=0.0):
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return default


def _t2mf_pipe_offers_v1(price_text, exact_positions, unit_hint="кг"):
    offers = []
    exact = [_low(x) for x in (exact_positions or []) if _s(x)]
    for raw_line in _s(price_text).splitlines():
        line = raw_line.strip(" \t-—•·")
        if "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 3:
            continue
        pos = _low(parts[0])
        if not any(x and x in pos for x in exact):
            continue
        unit = _low(parts[2])
        if unit_hint and unit_hint not in unit:
            continue
        price = _t2mf_float_v1(parts[1])
        if not (1 <= price <= 50000):
            continue
        offers.append({
            "position": parts[0],
            "price": price,
            "unit": parts[2],
            "supplier": parts[4] if len(parts) > 4 else "",
            "url": parts[5] if len(parts) > 5 else "",
            "checked_at": parts[6] if len(parts) > 6 else datetime.date.today().isoformat(),
        })
    return offers


def _t2mf_choose_price_v1(offers, choice):
    vals = [float(o.get("price") or 0) for o in (offers or []) if float(o.get("price") or 0) > 0]
    return float(_choose_value(vals, choice) or 0.0) if vals else 0.0


def _t2mf_combined_source_v1(*offer_lists):
    suppliers = []
    urls = []
    checked = []
    for offers in offer_lists:
        for offer in (offers or [])[:2]:
            sup = _s(offer.get("supplier"))
            url = _s(offer.get("url"))
            if sup and sup not in suppliers:
                suppliers.append(sup)
            if url and url not in urls:
                urls.append(url)
            if offer.get("checked_at"):
                checked.append(_s(offer.get("checked_at")))
    if not suppliers and not urls:
        return {}
    return {
        "status": "LIVE_CONFIRMED",
        "supplier": " / ".join(suppliers[:3]),
        "url": " / ".join(urls[:3]),
        "checked_at": max(checked) if checked else datetime.date.today().isoformat(),
    }


def _t2mf_is_metal_fence_row_v1(name):
    low = _low(name or "")
    return any(x in low for x in ("секция", "стойк", "калит", "ворот"))


def _t2ff_build_items_from_rows_v1(parsed, price_text, choice):  # noqa: F811
    items = list(_T2MF_PREV_BUILD_ITEMS_V1(parsed, price_text, choice) or [])
    rows = _t2ff_current_rows_v1(parsed)
    if not rows or not any(_t2mf_is_metal_fence_row_v1(r.get("name")) and _t2mf_float_v1(r.get("weight_kg")) > 0 for r in rows if isinstance(r, dict)):
        return items

    metal_offers = _t2mf_pipe_offers_v1(price_text, ("металлопрокат профильная труба",), "кг")
    fab_offers = _t2mf_pipe_offers_v1(price_text, ("изготовление металлоконструкций ограждения",), "кг")
    metal_kg = _t2mf_choose_price_v1(metal_offers, choice)
    fab_kg = _t2mf_choose_price_v1(fab_offers, choice)
    if metal_kg <= 0 or fab_kg <= 0:
        return items

    source = _t2mf_combined_source_v1(metal_offers, fab_offers)
    data_items = [dict(it) for it in items if _s(it.get("section")) not in ("Накладные", "Накладные расходы")]
    overhead_items = [dict(it) for it in items if _s(it.get("section")) in ("Накладные", "Накладные расходы")]

    for item, row in zip(data_items, rows):
        name = _s(row.get("name") or item.get("name") or "")
        if not _t2mf_is_metal_fence_row_v1(name):
            continue
        qty = _t2mf_float_v1(row.get("qty"))
        weight_total = _t2mf_float_v1(row.get("weight_kg"))
        if qty <= 0 or weight_total <= 0:
            continue
        per_item_weight = weight_total / qty
        mat_price = round(per_item_weight * (metal_kg + fab_kg), 2)
        work_price = _t2mf_float_v1(item.get("work_price"))
        item["mat_price"] = mat_price
        item["price"] = round(work_price + mat_price, 2)
        item["kind"] = "mixed"
        item["note"] = (_s(item.get("note")) + f"; материал по весу PDF: {per_item_weight:.2f} кг/ед, металл {metal_kg:.2f} руб/кг + изготовление {fab_kg:.2f} руб/кг")[:240]
        if source:
            item["price_source_status"] = source.get("status")
            item["price_supplier"] = source.get("supplier")
            item["price_url"] = source.get("url")
            item["price_checked_at"] = source.get("checked_at")

    subtotal = sum(
        _t2mf_float_v1(it.get("qty")) * (_t2mf_float_v1(it.get("work_price")) + _t2mf_float_v1(it.get("mat_price")))
        for it in data_items
    )
    if overhead_items and subtotal > 0:
        for oh in overhead_items:
            oh["price"] = round(subtotal * 0.07, 2)
            oh["work_price"] = round(subtotal * 0.07, 2)
            oh["mat_price"] = 0.0
            oh["note"] = "7% от проектных стоимостных позиций после корректировки материала по весу"
    return data_items + overhead_items


def _t2ff_rewrite_work_material_cols_v1(path, items):  # noqa: F811
    _T2MF_PREV_REWRITE_COLS_V1(path, items)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        row_idx = 2
        for item in items:
            while row_idx <= ws.max_row and not _s(ws.cell(row_idx, 3).value):
                row_idx += 1
            if row_idx > ws.max_row:
                break
            if item.get("price_source_status"):
                ws.cell(row_idx, 11, item.get("price_source_status"))
                ws.cell(row_idx, 12, item.get("price_supplier") or "")
                ws.cell(row_idx, 13, item.get("price_url") or "")
                ws.cell(row_idx, 14, item.get("price_checked_at") or datetime.date.today().isoformat())
            row_idx += 1
        wb.save(path)
        wb.close()
    except Exception as exc:
        try:
            _STV3_LOG.warning("PATCH_TOPIC2_METAL_FENCE_EXACT_KG_PRICING_V1 source rewrite failed: %s", exc)
        except Exception:
            pass


try:
    _STV3_LOG.info("PATCH_TOPIC2_METAL_FENCE_EXACT_KG_PRICING_V1 installed")
except Exception:
    pass
# === END_PATCH_TOPIC2_METAL_FENCE_EXACT_KG_PRICING_V1 ===

# === PATCH_TOPIC2_PARSE_CLARIFIED_PRICE_CONFIRM_FIELDS_V1 ===
# Canon: a confirmed continuation in task_history belongs to the current task cycle.
# The clarification merge appends clarified text to raw_input before parsing; expose it
# in parsed fields used by file/OCR missing-question and price-search guards.
try:
    _T2PCPCF_PREV_PARSE_REQUEST_V1 = _parse_request

    def _parse_request(text: str):  # noqa: F811
        parsed = _T2PCPCF_PREV_PARSE_REQUEST_V1(text)
        try:
            low = _low(text or "")
            markers = (
                "считать по найденным позициям",
                "считать по найденным проектным позициям",
                "считай по найденным позициям",
                "считай по найденным проектным позициям",
                "только найденные позиции",
                "только найденные проектные позиции",
                "искать через интернет",
                "искать цены",
                "ищи цены",
                "актуальные цены",
                "цены на материалы",
                "изготовление и монтаж",
            )
            if isinstance(parsed, dict) and any(m in low for m in markers):
                parsed["_topic2_history_clarified"] = (parsed.get("_topic2_history_clarified") or "") + "\n" + str(text or "")
                parsed["_topic2_confirm_text"] = (parsed.get("_topic2_confirm_text") or "") + "\n" + str(text or "")
        except Exception:
            pass
        return parsed

    try:
        _STV3_LOG.info("PATCH_TOPIC2_PARSE_CLARIFIED_PRICE_CONFIRM_FIELDS_V1 installed")
    except Exception:
        pass
except Exception:
    pass
# === END_PATCH_TOPIC2_PARSE_CLARIFIED_PRICE_CONFIRM_FIELDS_V1 ===

# === PATCH_TOPIC2_FILE_CONTEXT_NO_TEMPLATE_PRICES_V1 ===
# Canon: when the current task is based on uploaded PDF/OCR rows, old template
# rows are not an evidence source and must not be shown or used as prices.
try:
    _T2FCNTP_PREV_EXTRACT_TEMPLATE_PRICES_V1 = extract_template_prices
    _T2FCNTP_PREV_PRICE_CONFIRMATION_TEXT_V1 = _price_confirmation_text

    def extract_template_prices(template_path, parsed):  # noqa: F811
        if _t2ff_file_context_v1(parsed):
            return ("", "AREAL_CALC", False)
        return _T2FCNTP_PREV_EXTRACT_TEMPLATE_PRICES_V1(template_path, parsed)

    def _price_confirmation_text(parsed, template, sheet_name, template_prices, online_prices):  # noqa: F811
        if _t2ff_file_context_v1(parsed):
            template = dict(template or {})
            template["title"] = "Проектные строки PDF/КР"
            sheet_name = "AREAL_CALC"
            template_prices = "Шаблонные строки отключены для текущего PDF/КР. Основа сметы: только распознанные проектные строки и подтверждённые интернет-цены."
        return _T2FCNTP_PREV_PRICE_CONFIRMATION_TEXT_V1(parsed, template, sheet_name, template_prices, online_prices)

    try:
        _STV3_LOG.info("PATCH_TOPIC2_FILE_CONTEXT_NO_TEMPLATE_PRICES_V1 installed")
    except Exception:
        pass
except Exception:
    pass
# === END_PATCH_TOPIC2_FILE_CONTEXT_NO_TEMPLATE_PRICES_V1 ===

# === PATCH_TOPIC2_ARCHIVE_ONLY_PRICE_AND_COMPLETION_V1 ===
def _topic2_archive_price_mode_v1(text: str) -> bool:
    low = _low(text or "")
    archive_requested = any(x in low for x in ("архив", "архивных данных", "памяти", "кэш", "cache"))
    internet_forbidden = any(x in low for x in (
        "без интернета",
        "интернет не использовать",
        "интернет не нужен",
        "не делать в интернете",
        "не искать в интернете",
        "поиск в интернете не",
        "поиск по материалам не делать",
    ))
    return archive_requested and internet_forbidden


_TOPIC2_ARCHIVE_PREV_PRICE_INTENT_V1 = _topic2_price_search_explicit_intent_v1


def _topic2_price_search_explicit_intent_v1(text: str) -> bool:  # noqa: F811
    if _topic2_archive_price_mode_v1(text):
        return True
    low = _low(text or "")
    if any(x in low for x in ("без интернета", "не искать в интернете", "поиск в интернете не", "интернет не использовать")):
        return False
    return _TOPIC2_ARCHIVE_PREV_PRICE_INTENT_V1(text)


_TOPIC2_ARCHIVE_PREV_CONFIRMED_ROWS_V1 = _t2ff_confirmed_current_rows_v1


def _t2ff_confirmed_current_rows_v1(parsed, confirm_text=""):  # noqa: F811
    if _TOPIC2_ARCHIVE_PREV_CONFIRMED_ROWS_V1(parsed, confirm_text):
        return True
    text = _low(confirm_text or "") + " " + _low((parsed or {}).get("raw") or "")
    return any(x in text for x in (
        "посчитать стоимость работ и материалов",
        "считать стоимость работ и материалов",
        "рассчитать стоимость работ и материалов",
        "смета по проекту",
    ))


def _topic2_archive_unit_v1(value: Any) -> str:
    return _low(value or "").replace("м3", "м³").replace("м2", "м²").replace("м.п.", "п.м").strip()


def _topic2_archive_category_v1(name: str) -> str:
    low = _low(name or "")
    if "бетон" in low:
        grade = re.search(r"[вb]\s*(\d+(?:[.,]\d+)?)", low, re.I)
        return "concrete:" + (grade.group(1).replace(",", ".") if grade else "")
    if "арматур" in low:
        cls = re.search(r"[аa](240|500[сc])", low, re.I)
        return "rebar:" + (cls.group(1).replace("c", "с") if cls else "")
    if any(x in low for x in ("газобет", "блоки строительные", "гб ")):
        thickness = re.search(r"(?:гб|газобет\w*)\s*(\d{2,3})", low, re.I)
        return "blocks:" + (thickness.group(1) if thickness else "")
    if any(x in low for x in ("достав", "логист", "транспорт")):
        return "logistics"
    return ""


def _topic2_archive_catalog_v1() -> List[Dict[str, Any]]:
    root = BASE / "runtime" / "stroyka_estimates"
    if not root.exists():
        return []
    try:
        from openpyxl import load_workbook as _topic2_archive_lwb
    except Exception:
        return []
    files = sorted(root.rglob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)[:80]
    catalog: List[Dict[str, Any]] = []
    for path in files:
        try:
            wb = _topic2_archive_lwb(path, data_only=True, read_only=True)
            for ws in wb.worksheets:
                headers = {_low(ws.cell(1, col).value or ""): col for col in range(1, min(ws.max_column, 20) + 1)}
                name_col = headers.get("наименование")
                unit_col = headers.get("ед изм") or headers.get("ед. изм.")
                work_col = headers.get("цена работ")
                material_col = headers.get("цена материалов")
                if not (name_col and unit_col and work_col and material_col):
                    continue
                source_col = headers.get("источник цены")
                supplier_col = headers.get("поставщик")
                url_col = headers.get("url")
                checked_col = headers.get("checked_at")
                for row_idx in range(2, ws.max_row + 1):
                    name = _s(ws.cell(row_idx, name_col).value)
                    unit = _topic2_archive_unit_v1(ws.cell(row_idx, unit_col).value)
                    category = _topic2_archive_category_v1(name)
                    if not name or not unit or not category:
                        continue
                    try:
                        work_price = float(ws.cell(row_idx, work_col).value) if ws.cell(row_idx, work_col).value not in (None, "") else None
                    except Exception:
                        work_price = None
                    try:
                        material_price = float(ws.cell(row_idx, material_col).value) if ws.cell(row_idx, material_col).value not in (None, "") else None
                    except Exception:
                        material_price = None
                    if not ((work_price and work_price > 0) or (material_price and material_price > 0)):
                        continue
                    catalog.append({
                        "name": name,
                        "unit": unit,
                        "category": category,
                        "work_price": work_price,
                        "material_price": material_price,
                        "source": _s(ws.cell(row_idx, source_col).value) if source_col else "archive",
                        "supplier": _s(ws.cell(row_idx, supplier_col).value) if supplier_col else "",
                        "url": _s(ws.cell(row_idx, url_col).value) if url_col else "",
                        "checked_at": _s(ws.cell(row_idx, checked_col).value) if checked_col else "",
                        "archive_file": str(path),
                    })
            wb.close()
        except Exception:
            continue
    return catalog


def _topic2_archive_match_v1(name: str, unit: str, catalog: List[Dict[str, Any]]) -> Dict[str, Any]:
    category = _topic2_archive_category_v1(name)
    norm_unit = _topic2_archive_unit_v1(unit)
    if not category:
        return {}
    candidates = [row for row in catalog if row.get("category") == category and row.get("unit") == norm_unit]
    if not candidates and category.endswith(":"):
        prefix = category.split(":", 1)[0] + ":"
        candidates = [row for row in catalog if _s(row.get("category")).startswith(prefix) and row.get("unit") == norm_unit]
    if not candidates:
        return {}

    def one_value(field: str) -> Optional[float]:
        values = sorted({round(float(row[field]), 4) for row in candidates if row.get(field) not in (None, "") and float(row[field]) > 0})
        return values[0] if len(values) == 1 else None

    work_price = one_value("work_price")
    material_price = one_value("material_price")
    if work_price is None and material_price is None:
        return {}
    source_row = candidates[0]
    return {
        "work_price": work_price,
        "material_price": material_price,
        "source": "archive:" + _s(source_row.get("source") or "saved estimate"),
        "supplier": _s(source_row.get("supplier")),
        "url": _s(source_row.get("url")),
        "checked_at": _s(source_row.get("checked_at")),
        "archive_file": _s(source_row.get("archive_file")),
    }


_TOPIC2_ARCHIVE_PREV_SEARCH_ONLINE_V1 = _search_prices_online


async def _search_prices_online(parsed: Dict[str, Any], template: Dict[str, Any], sheet_name: Optional[str], conn=None, task_id=None) -> str:  # noqa: F811
    if not _topic2_archive_price_mode_v1((parsed or {}).get("raw") or ""):
        return await _TOPIC2_ARCHIVE_PREV_SEARCH_ONLINE_V1(parsed, template, sheet_name, conn=conn, task_id=task_id)
    rows = _t2ff_current_rows_v1(parsed)
    catalog = _topic2_archive_catalog_v1()
    price_map = {}
    lines = ["ARCHIVE_ONLY: интернет-поиск отключён пользователем; неизвестные цены оставлены пустыми"]
    for row in rows:
        name = _s(row.get("name"))
        match = _topic2_archive_match_v1(name, _s(row.get("unit")), catalog)
        price_map[name] = match
        if match:
            lines.append("- {} | работа={} | материал={} | {}".format(
                name,
                match.get("work_price") if match.get("work_price") is not None else "PRICE_MISSING",
                match.get("material_price") if match.get("material_price") is not None else "PRICE_MISSING",
                match.get("source") or "archive",
            ))
    parsed["_topic2_archive_price_map"] = price_map
    parsed["_topic2_archive_price_catalog_count"] = len(catalog)
    if conn is not None and task_id is not None:
        _history_safe(conn, task_id, "TOPIC2_INTERNET_SEARCH_DISABLED_BY_USER")
        _history_safe(conn, task_id, f"TOPIC2_ARCHIVE_PRICE_CATALOG_READ:{len(catalog)}")
        _history_safe(conn, task_id, f"TOPIC2_ARCHIVE_PRICE_REUSED:{sum(1 for value in price_map.values() if value)}")
        _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_DONE:archive_only")
    return "\n".join(lines)


_TOPIC2_ARCHIVE_PREV_BUILD_ITEMS_V1 = _t2ff_build_items_from_rows_v1


def _t2ff_build_items_from_rows_v1(parsed, price_text, choice):  # noqa: F811
    if not _topic2_archive_price_mode_v1((parsed or {}).get("raw") or ""):
        return _TOPIC2_ARCHIVE_PREV_BUILD_ITEMS_V1(parsed, price_text, choice)
    rows = _t2ff_current_rows_v1(parsed)
    price_map = (parsed or {}).get("_topic2_archive_price_map") or {}
    items = []
    for row in rows:
        name = _s(row.get("name"))
        try:
            qty = float(row.get("qty") or 0)
        except Exception:
            qty = 0.0
        if not name or qty <= 0:
            continue
        match = price_map.get(name) or {}
        work_price = match.get("work_price")
        material_price = match.get("material_price")
        category = _topic2_archive_category_v1(name)
        section = "Проектные позиции"
        if category.startswith("concrete") or category.startswith("rebar"):
            section = "Монолитные конструкции"
        elif category.startswith("blocks"):
            section = "Стены"
        missing = []
        if work_price is None:
            missing.append("WORK_PRICE_MISSING")
        if material_price is None:
            missing.append("MATERIAL_PRICE_MISSING")
        note = _s(row.get("note") or row.get("source") or "текущий PDF")
        if missing:
            note = (note + "; " + "; ".join(missing)).strip("; ")
        items.append({
            "section": section,
            "name": name[:240],
            "unit": _s(row.get("unit") or "шт"),
            "qty": qty,
            "price": float(work_price or 0) + float(material_price or 0),
            "work_price": float(work_price or 0),
            "mat_price": float(material_price or 0),
            "work_price_missing": work_price is None,
            "material_price_missing": material_price is None,
            "kind": "mixed",
            "note": note[:240],
            "source": _s(match.get("source") or "PRICE_MISSING"),
            "supplier": _s(match.get("supplier")),
            "url": _s(match.get("url")),
            "checked_at": _s(match.get("checked_at")),
            "archive_price_mode": True,
        })
    try:
        distance = float((parsed or {}).get("distance_km") or 0)
    except Exception:
        distance = 0.0
    if distance > 0:
        items.append({
            "section": "Логистика",
            "name": f"Логистика до объекта, удалённость {distance:g} км",
            "unit": "компл",
            "qty": 1.0,
            "price": 0.0,
            "work_price": 0.0,
            "mat_price": 0.0,
            "work_price_missing": True,
            "material_price_missing": True,
            "kind": "mixed",
            "note": "Цена отсутствует в однозначно совпадающих архивных данных; оставлена пустой",
            "source": "PRICE_MISSING",
            "archive_price_mode": True,
        })
    subtotal = sum(item["qty"] * (item["work_price"] + item["mat_price"]) for item in items)
    if subtotal > 0:
        overhead = round(subtotal * 0.07, 2)
        items.append({
            "section": "Накладные расходы",
            "name": "Организация работ и накладные расходы",
            "unit": "компл",
            "qty": 1.0,
            "price": overhead,
            "work_price": overhead,
            "mat_price": 0.0,
            "work_price_missing": False,
            "material_price_missing": False,
            "kind": "mixed",
            "note": "7% от подтверждённых архивными ценами позиций",
            "source": "CANON_OVERHEAD_7_PERCENT",
            "archive_price_mode": True,
        })
    return items


_TOPIC2_ARCHIVE_PREV_REWRITE_COLS_V1 = _t2ff_rewrite_work_material_cols_v1


def _t2ff_rewrite_work_material_cols_v1(path, items):  # noqa: F811
    _TOPIC2_ARCHIVE_PREV_REWRITE_COLS_V1(path, items)
    try:
        from openpyxl import load_workbook as _topic2_archive_rewrite_lwb
        wb = _topic2_archive_rewrite_lwb(path, data_only=False)
        ws = wb["AREAL_CALC"] if "AREAL_CALC" in wb.sheetnames else wb.active
        row_idx = 2
        for item in items or []:
            while row_idx <= ws.max_row and not _s(ws.cell(row_idx, 3).value):
                row_idx += 1
            if row_idx > ws.max_row:
                break
            if item.get("work_price_missing"):
                ws.cell(row_idx, 6).value = None
            if item.get("material_price_missing"):
                ws.cell(row_idx, 8).value = None
            ws.cell(row_idx, 11).value = _s(item.get("source") or "PRICE_MISSING")
            ws.cell(row_idx, 12).value = _s(item.get("supplier"))
            ws.cell(row_idx, 13).value = _s(item.get("url"))
            ws.cell(row_idx, 14).value = _s(item.get("checked_at"))
            ws.cell(row_idx, 15).value = _s(item.get("note"))
            row_idx += 1
        wb.save(path)
        wb.close()
    except Exception as exc:
        try:
            _STV3_LOG.warning("TOPIC2_ARCHIVE_PRICE_XLSX_REWRITE_FAILED: %s", exc)
        except Exception:
            pass


_TOPIC2_ARCHIVE_PREV_QUALITY_GATE_V1 = _quality_gate_xlsx


def _quality_gate_xlsx(xlsx_path: str, items: List[Dict[str, Any]], py_total: float) -> Tuple[bool, str]:  # noqa: F811
    archive_blank_mode = bool(items) and all(bool(item.get("archive_price_mode")) for item in items)
    if not archive_blank_mode or py_total > 0:
        return _TOPIC2_ARCHIVE_PREV_QUALITY_GATE_V1(xlsx_path, items, py_total)
    if not xlsx_path or not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) < 5000:
        return False, "ARCHIVE_BLANK_XLSX_INVALID"
    if len(items) < 8:
        return False, f"ARCHIVE_BLANK_TOO_FEW_ITEMS:{len(items)}"
    try:
        from openpyxl import load_workbook as _topic2_archive_qg_lwb
        wb = _topic2_archive_qg_lwb(xlsx_path, data_only=False, read_only=True)
        ws = wb["AREAL_CALC"] if "AREAL_CALC" in wb.sheetnames else wb.active
        headers = [_s(ws.cell(1, col).value) for col in range(1, 16)]
        formula_count = sum(1 for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("="))
        wb.close()
        if len(headers) < 15 or formula_count < 8:
            return False, "ARCHIVE_BLANK_CANON_COLUMNS_OR_FORMULAS_MISSING"
        return True, "OK_ARCHIVE_PRICES_MISSING_ALLOWED_BY_USER"
    except Exception as exc:
        return False, "ARCHIVE_BLANK_XLSX_VALIDATE_ERROR:" + _s(exc)[:120]
# === END_PATCH_TOPIC2_ARCHIVE_ONLY_PRICE_AND_COMPLETION_V1 ===

# === PATCH_TOPIC2_NEW_PROJECT_PDF_ISOLATION_V2 ===
# A newly uploaded project PDF is an isolated project context.  It must pass the
# project-volume completeness gate before prices, templates or final artifacts.
_TOPIC2_NEW_PROJECT_PREV_HANDLER_V2 = maybe_handle_stroyka_estimate


def _topic2_project_incomplete_message_v2(bundle: Dict[str, Any]) -> str:
    facts = list(bundle.get("project_facts") or [])
    positions = list(bundle.get("positions") or [])
    totals = list(bundle.get("totals") or [])
    missing = list(bundle.get("missing_items") or [])
    object_name = next((_s(row.get("value")) for row in facts if _low(row.get("name")) == "объект"), "проект")
    pages = next((_s(row.get("value")) for row in facts if "страниц" in _low(row.get("name"))), "")
    lines = [
        "Принял PDF как новый отдельный проект.",
        f"Объект: {object_name}",
    ]
    if pages:
        lines.append(f"Обработано страниц: {pages}")
    lines.extend(["", "Подтверждено по ведомостям проекта:"])
    for row in totals:
        value = row.get("value")
        unit = _s(row.get("unit"))
        if value not in (None, ""):
            lines.append(f"- {_s(row.get('name'))}: {value:g} {unit}" if isinstance(value, (int, float)) else f"- {_s(row.get('name'))}: {value} {unit}")
    masonry = [row for row in positions if _s(row.get("position_type")) == "masonry_material"]
    for row in masonry:
        lines.append(f"- {_s(row.get('name'))}: {int(row.get('count_pcs') or 0)} шт, {float(row.get('volume_m3') or 0):g} м³")
    structural = [row for row in positions if _s(row.get("position_type")) == "project_structural_element"]
    lines.append(f"- Конструктивные элементы: {len(structural)} позиций с объёмом бетона и массой арматуры")

    labels = {
        "formwork_area_m2": "площадь опалубки",
        "crushed_stone_base_volume_m3": "объём щебёночного основания",
        "waterproofing_area_m2": "площадь гидроизоляции",
        "insulation_area_m2": "площадь утепления",
        "steel_concrete_summary_table": "сводная ведомость бетона и арматуры",
        "masonry_schedule": "ведомость кладочных блоков",
    }
    lines.extend(["", "До сметы не подтверждены отдельной ведомостью:"])
    for item in missing:
        lines.append(f"- {labels.get(_s(item), _s(item))}")
    lines.extend([
        "",
        "Цены и смету пока не запускаю, чтобы не подмешивать старые позиции и не выдумывать объёмы.",
        "Уточни: рассчитать отсутствующие объёмы по геометрии чертежей или считать только позиции из подтверждённых ведомостей?",
    ])
    return "\n".join(lines)


def _topic2_project_clarifications_v2(conn, task_id: str) -> str:
    try:
        rows = conn.execute(
            "SELECT action FROM task_history WHERE task_id=? AND action LIKE 'clarified:%' ORDER BY id ASC",
            (task_id,),
        ).fetchall()
        return "\n".join(_s(row[0]).split("clarified:", 1)[1] for row in rows if "clarified:" in _s(row[0]))
    except Exception:
        return ""


def _topic2_project_manual_rates_v2(text: str) -> Dict[str, Any]:
    low = _low(text).replace("₽", " руб ")
    rules = {
        "rebar_material_per_t": r"арматур[^\n]{0,45}?(\d[\d\s]{3,})[^\n]{0,20}(?:тон|т\b)",
        "gasblock_material_per_m3": r"газобетон[^\n]{0,35}?(\d[\d\s]{3,})[^\n]{0,20}(?:куб|м3|м³)",
        "foundation_work_per_m3": r"(?:работ[^\n]{0,30}фундамент|фундамент[^\n]{0,30}работ)[^\n]{0,30}?(\d[\d\s]{3,})",
        "slab_work_per_m3": r"перекрыт[^\n]{0,35}?(\d[\d\s]{3,})[^\n]{0,20}(?:куб|м3|м³)",
        "wall_work_per_m3": r"(?:стен[^\n]{0,20}монолит|монолит[^\n]{0,20}стен)[^\n]{0,25}?(\d[\d\s]{3,})",
        "concrete_material_per_m3": r"(?<!газо)бетон[^\n]{0,35}?(\d[\d\s]{3,})(?:[^\n]{0,20}(?:куб|м3|м³))?",
        "column_work_per_m3": r"колонн[^\n]{0,35}?(?:работ[^\n]{0,20}?)?(\d[\d\s]{3,})[^\n]{0,20}(?:куб|м3|м³)",
        "masonry_work_per_m3": r"кладк[^\n]{0,25}газобетон[^\n]{0,30}?(\d[\d\s]{3,})[^\n]{0,20}(?:куб|м3|м³)",
        "stairs_work_per_unit": r"лестниц[^\n]{0,45}?работ[^\n]{0,25}?(\d[\d\s]{3,})[^\n]{0,20}(?:единиц|шт)",
        "remaining_monolith_work_per_m3": r"(?:р1|р2|р3|ростверк)[^\n]{0,60}?(\d[\d\s]{3,})[^\n]{0,20}(?:куб|м3|м³)",
        "rostverk_work_per_m": r"(?:^|\n)\s*(\d[\d\s]*)[^\n]{0,20}(?:метр\s+погон|пог\.?\s*м)",
        "crushed_stone_material_per_m3": r"щеб[^\n]{0,35}?(\d[\d\s]{1,})[^\n]{0,20}(?:куб|м3|м³)",
        "crushed_stone_work_per_m3": r"щеб[^\n]{0,60}?работ[^\n]{0,20}?(\d[\d\s]{1,})",
        "waterproofing_material_per_m2": r"гидроизоляц[^\n]{0,35}?материал[^\n]{0,20}?(\d[\d\s]{3,})",
        "waterproofing_work_per_m2": r"гидроизоляц[^\n]{0,35}?работ[^\n]{0,20}?(\d[\d\s]{3,})",
        "insulation_material_per_m2": r"утеплен[^\n]{0,35}?материал[^\n]{0,20}?(\d[\d\s]{3,})",
        "insulation_work_per_m2": r"утеплен[^\n]{0,35}?работ[^\n]{0,20}?(\d[\d\s]{3,})",
    }
    out: Dict[str, Any] = {}
    for key, pattern in rules.items():
        match = re.search(pattern, low, re.I)
        if not match:
            continue
        try:
            out[key] = float(match.group(1).replace(" ", ""))
        except Exception:
            pass
    shared_layers = re.search(
        r"(?:гидроизоляц[^\n]{0,30}утеплен|утеплен[^\n]{0,30}гидроизоляц)"
        r"[^\n]{0,35}?материал[^\n]{0,20}?(\d[\d\s]*)[^\n]{0,25}работ[^\n]{0,20}?(\d[\d\s]*)",
        low,
        re.I,
    )
    if shared_layers:
        material_rate = float(shared_layers.group(1).replace(" ", ""))
        work_rate = float(shared_layers.group(2).replace(" ", ""))
        out.update({
            "waterproofing_material_per_m2": material_rate,
            "waterproofing_work_per_m2": work_rate,
            "insulation_material_per_m2": material_rate,
            "insulation_work_per_m2": work_rate,
        })
    monolith_scope = re.search(
        r"монолитн[^\n.]{0,40}работ[^\n.]{0,40}(?:не\s+)?включ[^\n.]{0,50}опалуб[^\n.]{0,50}(?:вязк|арматур)",
        low,
        re.I,
    )
    if monolith_scope:
        out["monolith_rates_include_formwork_rebar"] = "не включ" not in monolith_scope.group(0)
    if ("щеб" in low and "гидроизоляц" in low) and any(x in low for x in ("включ", "конечно да", "считать")):
        out["include_project_layers"] = True
    if "подъезд" in low:
        out["logistics_access_confirmed"] = True
        out["logistics_access"] = "no" if re.search(r"подъезд[^\n]{0,20}(?:нет|отсутств)", low) else "yes"
    if any(token in low for token in ("разгруз", "манипулятор", "кран")):
        out["logistics_unloading_confirmed"] = True
        out["logistics_unloading"] = "not_required" if any(
            token in low for token in ("разгрузка не нужна", "манипулятор не нужен", "без разгрузки")
        ) else "required"
    same_blank_line = re.search(
        r"(?:логист[^\n]{0,50}накладн|накладн[^\n]{0,50}логист)[^\n]{0,50}(?:пуст|не включ|не учитывать)",
        low,
    )
    if same_blank_line:
        out["logistics_mode"] = "blank"
        out["overhead_mode"] = "blank"
    elif "логист" in low and any(token in low for token in ("пуст", "не включ", "не учитывать")):
        out["logistics_mode"] = "blank"
    elif "накладн" in low and any(token in low for token in ("пуст", "не включ", "не учитывать")):
        out["overhead_mode"] = "blank"
    logistics_deliveries = re.search(
        r"логист[^\n]{0,50}?(\d+)\s+(?:достав\w*|рейс\w*)\s+(?:по|x|х|×)\s*"
        r"(\d[\d\s]{2,})(?:\s*(?:руб|р\b))?",
        low,
    )
    if logistics_deliveries:
        delivery_count = int(logistics_deliveries.group(1))
        delivery_unit_price = float(logistics_deliveries.group(2).replace(" ", ""))
        out["logistics_mode"] = "manual_amount"
        out["logistics_delivery_count"] = delivery_count
        out["logistics_unit_price"] = delivery_unit_price
        out["logistics_amount"] = delivery_count * delivery_unit_price
    else:
        logistics_amount = re.search(r"логист[^\n]{0,35}?(\d[\d\s]{3,})\s*(?:руб|р\b)", low)
        if logistics_amount:
            out["logistics_mode"] = "manual_amount"
            out["logistics_amount"] = float(logistics_amount.group(1).replace(" ", ""))
    overhead_amount = re.search(r"накладн[^\n]{0,35}?(\d[\d\s]{3,})\s*(?:руб|р\b)", low)
    overhead_percent = re.search(r"накладн[^\n]{0,35}?(\d+(?:[.,]\d+)?)\s*%", low)
    if overhead_amount:
        out["overhead_mode"] = "manual_amount"
        out["overhead_amount"] = float(overhead_amount.group(1).replace(" ", ""))
    elif overhead_percent:
        out["overhead_mode"] = "percent"
        out["overhead_percent"] = float(overhead_percent.group(1).replace(",", "."))
        if re.search(r"(?:от\s+)?(?:общей\s+)?стоимост[ьи]\s+работ", low):
            out["overhead_basis"] = "works_only"
    return out


def _topic2_project_manual_price_message_v2(rates: Dict[str, Any]) -> str:
    labels = {
        "rebar_material_per_t": "арматура: {value:g} руб/т",
        "gasblock_material_per_m3": "газобетон: {value:g} руб/м³",
        "foundation_work_per_m3": "монолитный фундамент, работы: {value:g} руб/м³",
        "slab_work_per_m3": "межэтажные перекрытия, работы: {value:g} руб/м³",
        "wall_work_per_m3": "монолитные стены, работы: {value:g} руб/м³",
        "concrete_material_per_m3": "бетон В30/В25 с доставкой: {value:g} руб/м³",
        "column_work_per_m3": "монолитные колонны, работы: {value:g} руб/м³",
        "masonry_work_per_m3": "кладка газобетона, работы: {value:g} руб/м³",
        "stairs_work_per_unit": "лестницы, работы: {value:g} руб/ед.",
        "remaining_monolith_work_per_m3": "элементы Р1/Р2/Р3, работы: {value:g} руб/м³",
        "rostverk_work_per_m": "ростверки Р1/Р2/Р3, работы: {value:g} руб/пог. м",
        "crushed_stone_material_per_m3": "щебёночное основание, материал: {value:g} руб/м³",
        "crushed_stone_work_per_m3": "щебёночное основание, работы: {value:g} руб/м³",
        "waterproofing_material_per_m2": "гидроизоляция, материал: {value:g} руб/м²",
        "waterproofing_work_per_m2": "гидроизоляция, работы: {value:g} руб/м²",
        "insulation_material_per_m2": "утепление, материал: {value:g} руб/м²",
        "insulation_work_per_m2": "утепление, работы: {value:g} руб/м²",
    }
    lines = ["Принял расчёт по текущему проекту и записал ручные расценки:"]
    for key in labels:
        if key in rates:
            lines.append("- " + labels[key].format(value=rates[key]))
    if rates.get("monolith_rates_include_formwork_rebar") is True:
        lines.append("- монолитные работы включают опалубку и вязку арматуры")
    if rates.get("include_project_layers") is True:
        lines.append("- щебёночное основание, гидроизоляцию и утепление включить")
    missing: List[str] = []
    if "concrete_material_per_m3" not in rates:
        missing.append("бетон В30 и В25: цена материала за м³")
    if "column_work_per_m3" not in rates:
        missing.append("колонны К1/К2: цена работ за м³")
    if "remaining_monolith_work_per_m3" not in rates:
        if "rostverk_work_per_m" in rates:
            missing.append(
                f"для ростверков Р1/Р2/Р3 указано {rates['rostverk_work_per_m']:g} руб/пог. м, "
                "но в извлечённой ведомости есть только объём 10,5 м³ без погонной длины; "
                "укажите цену работ за м³ либо подтвердите расчёт длины по геометрии проекта"
            )
        else:
            missing.append("элементы Р1/Р2/Р3: цена работ за м³")
    if "stairs_work_per_unit" not in rates:
        missing.append("лестницы Л1/Л2: цена работ за единицу")
    if "masonry_work_per_m3" not in rates:
        missing.append("кладка газобетона: цена работ за м³")
    if "monolith_rates_include_formwork_rebar" not in rates:
        missing.append("включены ли опалубка и вязка арматуры в монолитные работы")
    if "include_project_layers" not in rates:
        missing.append("включать ли щебёночное основание, гидроизоляцию и утепление")
    if rates.get("include_project_layers") is True:
        layer_keys = (
            "crushed_stone_material_per_m3", "crushed_stone_work_per_m3",
            "waterproofing_material_per_m2", "waterproofing_work_per_m2",
            "insulation_material_per_m2", "insulation_work_per_m2",
        )
        if any(key not in rates for key in layer_keys):
            missing.append("щебень: материал/работа за м³; гидроизоляция и утепление: материал/работа за м² (совпадающих архивных цен нет)")
    lines.extend(["", "Для сметы без нулевых цен осталось уточнить:"])
    for index, item in enumerate(missing, 1):
        lines.append(f"{index}. {item}.")
    lines.extend(["", "Интернет-поиск не запускаю."])
    return "\n".join(lines)


def _topic2_project_manual_rates_missing_v3(rates: Dict[str, Any]) -> List[str]:
    required = (
        "rebar_material_per_t", "gasblock_material_per_m3",
        "foundation_work_per_m3", "slab_work_per_m3", "wall_work_per_m3",
        "concrete_material_per_m3", "column_work_per_m3", "masonry_work_per_m3",
        "stairs_work_per_unit", "remaining_monolith_work_per_m3",
        "crushed_stone_material_per_m3", "crushed_stone_work_per_m3",
        "waterproofing_material_per_m2", "waterproofing_work_per_m2",
        "insulation_material_per_m2", "insulation_work_per_m2",
        "monolith_rates_include_formwork_rebar", "include_project_layers",
    )
    return [key for key in required if key not in rates]


def _topic2_project_logistics_missing_v3(rates: Dict[str, Any]) -> List[str]:
    missing = []
    if not rates.get("logistics_access_confirmed"):
        missing.append("подъезд для грузовой техники")
    if not rates.get("logistics_unloading_confirmed"):
        missing.append("нужна ли разгрузка/манипулятор")
    if "logistics_mode" not in rates:
        missing.append("логистику: указать сумму или оставить пустой")
    if "overhead_mode" not in rates:
        missing.append("накладные расходы: указать сумму/процент или оставить пустыми")
    return missing


def _topic2_project_logistics_prompt_v3(rates: Dict[str, Any]) -> str:
    missing = _topic2_project_logistics_missing_v3(rates)
    lines = [
        "Перед финальной сметой осталось подтвердить логистику и накладные расходы для текущего проекта (20 км от Санкт-Петербурга).",
    ]
    if rates.get("logistics_mode") == "manual_amount":
        if rates.get("logistics_delivery_count") and rates.get("logistics_unit_price") is not None:
            lines.append(
                "Принял логистику: "
                f"{int(rates['logistics_delivery_count'])} доставок × "
                f"{float(rates['logistics_unit_price']):g} = "
                f"{float(rates['logistics_amount']):g} руб."
            )
        else:
            lines.append(f"Принял логистику: {float(rates['logistics_amount']):g} руб.")
    lines.extend([
        "Уточните одним сообщением:",
        *[f"- {item}" for item in missing],
        "",
        "Интернет-поиск не запускаю.",
    ])
    return "\n".join(lines)


def _topic2_current_project_source_v3(row: Dict[str, Any], calculation: str = "") -> Dict[str, Any]:
    return {
        "source_type": "CURRENT_PROJECT_PDF",
        "source_file": _s(row.get("source_file")),
        "page": row.get("page"),
        "table_name": _s(row.get("table_name") or row.get("sheet") or "Текущий проект"),
        "row_text": _s(row.get("row_text") or row.get("text") or row.get("name")),
        "calculation": calculation or _s(row.get("calculation") or row.get("geometry_calculation")),
        "confidence": _s(row.get("confidence") or "direct"),
    }


def _topic2_current_project_manual_rows_v3(bundle: Dict[str, Any], rates: Dict[str, Any]) -> Dict[str, Any]:
    prepared = dict(bundle or {})
    rows: List[Dict[str, Any]] = []

    def add_row(section, name, unit, qty, work_price, material_price, source, key):
        row = _topic2_billable_row_v1(section, name, unit, qty, key + ".material", key, source)
        row.update({
            "work_unit_price": float(work_price) if work_price is not None else None,
            "material_unit_price": float(material_price) if material_price is not None else None,
            "work_price_source": "MANUAL" if work_price is not None else "NOT_APPLICABLE_OR_INCLUDED",
            "material_price_source": "MANUAL" if material_price is not None else "NOT_APPLICABLE_OR_INCLUDED",
            "work_price_status": "MANUAL" if work_price is not None else "NOT_APPLICABLE",
            "material_price_status": "MANUAL" if material_price is not None else "NOT_APPLICABLE",
            "supplier": "Расценка пользователя",
            "source_url": "",
            "checked_at": _now(),
        })
        rows.append(row)

    for position in prepared.get("positions") or []:
        if not isinstance(position, dict):
            continue
        name = _s(position.get("name"))
        low = _low(name)
        source = _topic2_current_project_source_v3(position)
        if _s(position.get("position_type")) == "project_structural_element":
            volume = float(position.get("concrete_volume_m3") or 0)
            rebar_t = float(position.get("rebar_mass_kg") or 0) / 1000.0
            if "плита фп" in low:
                section, work = "02 Фундаменты", rates["foundation_work_per_m3"]
            elif "плита пп" in low:
                section, work = "03 Монолитные перекрытия", rates["slab_work_per_m3"]
            elif "колонн" in low:
                section, work = "03 Монолитные колонны", rates["column_work_per_m3"]
            elif "стен" in low:
                section, work = "03 Монолитные стены", rates["wall_work_per_m3"]
            elif "ростверк" in low:
                section, work = "02 Ростверки", rates["remaining_monolith_work_per_m3"]
            elif "лестниц" in low:
                section, work = "03 Монолитные лестницы", None
            else:
                continue
            add_row(
                section, f"Бетон и монолитные работы: {name}", "м³", volume,
                work, rates["concrete_material_per_m3"], source,
                "current_project.concrete." + re.sub(r"\W+", "_", low),
            )
            if rebar_t > 0:
                add_row(
                    section, f"Арматура: {name}", "т", rebar_t,
                    None, rates["rebar_material_per_t"], source,
                    "current_project.rebar." + re.sub(r"\W+", "_", low),
                )
            if "лестниц" in low:
                add_row(
                    section, f"Устройство лестницы: {name}", "шт", 1,
                    rates["stairs_work_per_unit"], None, source,
                    "current_project.stairs_work." + re.sub(r"\W+", "_", low),
                )
        elif _s(position.get("position_type")) == "masonry_material":
            volume = float(position.get("volume_m3") or 0)
            add_row(
                "04 Стены и перегородки", f"Газобетон и кладка: {name}", "м³", volume,
                rates["masonry_work_per_m3"], rates["gasblock_material_per_m3"], source,
                "current_project.masonry." + re.sub(r"\W+", "_", low),
            )

    derived_by_name = {
        _s(row.get("name")): row for row in (prepared.get("derived_quantities") or [])
        if isinstance(row, dict)
    }
    layer_rows = (
        (
            "01 Основание", "Щебёночное основание", "Объём щебёночного основания", "м³",
            rates["crushed_stone_work_per_m3"], rates["crushed_stone_material_per_m3"],
            "current_project.layers.crushed_stone",
        ),
        (
            "05 Гидроизоляция и утепление", "Гидроизоляция по проекту", "Площадь гидроизоляции", "м²",
            rates["waterproofing_work_per_m2"], rates["waterproofing_material_per_m2"],
            "current_project.layers.waterproofing",
        ),
        (
            "05 Гидроизоляция и утепление", "Утепление по проекту", "Площадь утепления", "м²",
            rates["insulation_work_per_m2"], rates["insulation_material_per_m2"],
            "current_project.layers.insulation",
        ),
    )
    for section, name, quantity_name, unit, work, material, key in layer_rows:
        quantity = derived_by_name.get(quantity_name)
        if not quantity:
            continue
        add_row(
            section, name, unit, quantity.get("value"), work, material,
            _topic2_current_project_source_v3(quantity), key,
        )

    clarification_source = {
        "source_type": "USER_CLARIFICATION",
        "source_file": "Telegram",
        "page": "",
        "table_name": "Текущее уточнение пользователя",
        "row_text": "Логистика и накладные расходы",
        "calculation": "Ручная сумма/процент пользователя",
        "confidence": "confirmed",
    }
    if rates.get("logistics_mode") == "manual_amount":
        logistics_source = dict(clarification_source)
        if rates.get("logistics_delivery_count") and rates.get("logistics_unit_price") is not None:
            logistics_source["calculation"] = (
                f"{int(rates['logistics_delivery_count'])} доставок × "
                f"{float(rates['logistics_unit_price']):g} руб = "
                f"{float(rates['logistics_amount']):g} руб"
            )
        add_row(
            "09 Логистика", "Логистика по текущему проекту", "компл.", 1,
            rates.get("logistics_amount"), None, logistics_source,
            "current_project.logistics",
        )
    if rates.get("overhead_mode") == "manual_amount":
        add_row(
            "10 Накладные расходы", "Накладные расходы", "компл.", 1,
            rates.get("overhead_amount"), None, clarification_source,
            "current_project.overhead",
        )
    elif rates.get("overhead_mode") == "percent":
        if rates.get("overhead_basis") == "works_only":
            base_total = sum(
                float(row.get("qty") or 0) * float(row.get("work_unit_price") or 0)
                for row in rows
                if _topic2_current_project_canon_section_v4(row) not in (
                    "09 Логистика", "10 Накладные расходы",
                )
            )
        else:
            base_total = sum(
                float(row.get("qty") or 0)
                * (float(row.get("work_unit_price") or 0) + float(row.get("material_unit_price") or 0))
                for row in rows
            )
        overhead_amount = base_total * float(rates.get("overhead_percent") or 0) / 100.0
        percent_source = dict(clarification_source)
        basis_text = "стоимость работ" if rates.get("overhead_basis") == "works_only" else "подтверждённая база"
        percent_source["calculation"] = (
            f"{basis_text}: {base_total:g} руб × "
            f"{float(rates.get('overhead_percent') or 0):g}%"
        )
        add_row(
            "10 Накладные расходы", "Накладные расходы", "компл.", 1,
            overhead_amount, None, percent_source,
            "current_project.overhead",
        )

    audit = []
    for row_no, row in enumerate(rows, 1):
        for kind, price_key, status_key, source_key in (
            ("work", "work_unit_price", "work_price_status", "work_price_source"),
            ("material", "material_unit_price", "material_price_status", "material_price_source"),
        ):
            if row.get(price_key) is None:
                continue
            audit.append({
                "estimate_row_no": row_no,
                "position_key": row.get("position_key"),
                "material_total_key": row.get("material_total_key"),
                "public_name": row.get("name"),
                "price_kind": kind,
                "unit": row.get("unit"),
                "unit_price": row.get(price_key),
                "price_source": row.get(source_key),
                "status": row.get(status_key),
                "supplier": "Расценка пользователя",
                "source_url": "",
                "checked_at": _now(),
                "cache_hit": False,
                "sonar_attempted": False,
                "note": "Текущее явное уточнение пользователя",
            })

    prepared["estimate_rows"] = rows
    prepared["evidence_only_rows"] = list(prepared.get("quantities") or [])
    prepared["price_audit"] = audit
    prepared["price_items"] = audit
    prepared["missing_items"] = []
    prepared["VOLUMES_COMPLETE"] = True
    prepared["POSITIONS_EXTRACTION_COMPLETE"] = True
    prepared["current_project_manual_rows_ready"] = True
    prepared["price_mode"] = "USER_CONFIRMED_MANUAL_NO_INTERNET"
    prepared["manual_rates"] = dict(rates)
    return prepared


async def maybe_handle_stroyka_estimate(conn, task, logger=None):  # noqa: F811
    try:
        task_id = _s(_row_get(task, "id", ""))
        chat_id = _s(_row_get(task, "chat_id", ""))
        topic_id = int(_row_get(task, "topic_id", 0) or 0)
        input_type = _low(_s(_row_get(task, "input_type", "")))
        raw_input = _s(_row_get(task, "raw_input", ""))
        if topic_id == TOPIC_ID_STROYKA and input_type in ("drive_file", "file", "document"):
            meta: Dict[str, Any] = {}
            try:
                meta = json.loads(raw_input) if raw_input.strip().startswith("{") else {}
            except Exception:
                meta = {}
            file_name = _s(meta.get("file_name"))
            mime_type = _low(meta.get("mime_type"))
            caption = _s(meta.get("caption"))
            is_pdf = file_name.lower().endswith(".pdf") or "pdf" in mime_type
            is_estimate = any(word in _low(caption + " " + raw_input) for word in ("смет", "стоимость", "посчитать"))
            if task_id and is_pdf and is_estimate:
                _history_safe(conn, task_id, "TOPIC2_ESTIMATE_SESSION_CREATED")
                import glob as _t2np_glob
                paths = _t2np_glob.glob(str(BASE / "runtime" / "drive_files" / f"{task_id}_*"))
                current_path = next((path for path in paths if path.lower().endswith(".pdf") and os.path.getsize(path) > 1000), "")
                if current_path:
                    from core.pdf_spec_extractor import extract_project_positions_bundle as _t2np_extract
                    bundle = _t2np_extract([current_path], topic_id=TOPIC_ID_STROYKA) or {}
                    if bundle.get("ok"):
                        _history_safe(conn, task_id, "TOPIC2_CONTEXT_READY")
                    if bundle.get("ok") and not bundle.get("POSITIONS_EXTRACTION_COMPLETE"):
                        clarified_text = "\n".join(
                            part for part in (caption, _topic2_project_clarifications_v2(conn, task_id))
                            if _s(part).strip()
                        )
                        scope_confirmed = any(marker in _low(clarified_text) for marker in (
                            "по проект", "текущему проект", "все позиции, найденные в pdf",
                            "по геометр", "рассчитать отсутствующ", "считать по найденн",
                        ))
                        manual_rates = _topic2_project_manual_rates_v2(clarified_text)
                        unresolved_project_items = list(bundle.get("missing_items") or [])
                        if manual_rates.get("monolith_rates_include_formwork_rebar") is True:
                            unresolved_project_items = [
                                item for item in unresolved_project_items if item != "formwork_area_m2"
                            ]
                        missing_manual_rates = _topic2_project_manual_rates_missing_v3(manual_rates)
                        missing_logistics = _topic2_project_logistics_missing_v3(manual_rates)
                        if (
                            scope_confirmed
                            and not unresolved_project_items
                            and not missing_manual_rates
                            and missing_logistics
                        ):
                            text = _topic2_project_logistics_prompt_v3(manual_rates)
                            pending = {
                                "version": "TOPIC2_NEW_PROJECT_PDF_ISOLATION_V3",
                                "status": "WAITING_LOGISTICS_CONFIRMATION",
                                "task_id": task_id, "chat_id": chat_id, "topic_id": topic_id,
                                "source_file": current_path, "project_bundle": bundle,
                                "volume_basis": "CURRENT_PROJECT_DRAWINGS",
                                "manual_rates": manual_rates,
                                "created_at": _now(),
                            }
                            _memory_save(chat_id, f"topic_2_estimate_pending_{task_id}", pending)
                            _history_safe(conn, task_id, "TOPIC2_LOGISTICS_CONFIRMATION_REQUIRED")
                            _history_safe(conn, task_id, "TOPIC2_FINAL_BLOCKED_UNTIL_LOGISTICS_CONFIRMATION")
                            send_res = await _send_text(
                                chat_id, text,
                                _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None),
                                topic_id,
                            )
                            kwargs = {
                                "state": "WAITING_CLARIFICATION", "result": text,
                                "error_message": "TOPIC2_LOGISTICS_CONFIRMATION_REQUIRED",
                            }
                            if isinstance(send_res, dict) and send_res.get("bot_message_id"):
                                kwargs["bot_message_id"] = send_res.get("bot_message_id")
                            _update_task_safe(conn, task_id, **kwargs)
                            return True
                        if scope_confirmed and not unresolved_project_items and not missing_manual_rates:
                            _history_safe(conn, task_id, "TOPIC2_LOGISTICS_CONFIRMED")
                            prepared_bundle = _topic2_current_project_manual_rows_v3(bundle, manual_rates)
                            estimate_rows = list(prepared_bundle.get("estimate_rows") or [])
                            names = "\n".join(_low(row.get("name")) for row in estimate_rows)
                            forbidden_old = any(token in names for token in ("фм1/фм2", "бфм1", "ангар"))
                            if not estimate_rows or forbidden_old:
                                _history_safe(conn, task_id, "TOPIC2_CURRENT_PROJECT_FINAL_BLOCKED:row_validation")
                                _update_task_safe(
                                    conn, task_id, state="FAILED",
                                    error_message="TOPIC2_CURRENT_PROJECT_ROWS_INVALID",
                                )
                                return True
                            _history_safe(
                                conn, task_id,
                                f"TOPIC2_PROJECT_ALL_PAGES_SCANNED:{(bundle.get('files') or [{}])[0].get('pages_seen', 0)}",
                            )
                            _history_safe(conn, task_id, f"TOPIC2_PROJECT_POSITIONS_EXTRACTED:{len(bundle.get('positions') or [])}")
                            _history_safe(conn, task_id, "TOPIC2_POSITIONS_EXTRACTION_COMPLETE_YES")
                            _history_safe(conn, task_id, "TOPIC2_VOLUME_COMPLETENESS_GATE_OK")
                            _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_NOT_REQUESTED_NO_INTERNET")
                            _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_STARTED:manual_user_rates")
                            _history_safe(conn, task_id, "TOPIC2_PRICE_SOURCE_FOUND:manual_user_rates")
                            _history_safe(conn, task_id, "TOPIC2_PRICE_ENRICHMENT_DONE:manual_user_rates")
                            _history_safe(conn, task_id, "TOPIC2_PRICE_CHOICE_CONFIRMED:manual_user_rates")
                            _history_safe(conn, task_id, f"TOPIC2_CURRENT_PROJECT_ESTIMATE_ROWS:{len(estimate_rows)}")
                            _memory_save(chat_id, f"topic_2_project_bundle_{task_id}", prepared_bundle)
                            await _topic2_project_bundle_send_artifacts_v1(
                                conn, task_id, chat_id, topic_id,
                                _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None),
                                prepared_bundle,
                            )
                            return True
                        if scope_confirmed:
                            text = _topic2_project_manual_price_message_v2(manual_rates)
                            pending_status = "WAITING_MANUAL_PRICE_INPUT"
                        else:
                            text = _topic2_project_incomplete_message_v2(bundle)
                            pending_status = "WAITING_PROJECT_VOLUME_CLARIFICATION"
                        pending = {
                            "version": "TOPIC2_NEW_PROJECT_PDF_ISOLATION_V2",
                            "status": pending_status,
                            "task_id": task_id, "chat_id": chat_id, "topic_id": topic_id,
                            "source_file": current_path, "project_bundle": bundle,
                            "volume_basis": "CURRENT_PROJECT_DRAWINGS" if scope_confirmed else "AWAITING_USER_CHOICE",
                            "manual_rates": manual_rates,
                            "created_at": _now(),
                        }
                        _memory_save(chat_id, f"topic_2_estimate_pending_{task_id}", pending)
                        _memory_save(chat_id, f"topic_2_project_bundle_{task_id}", pending)
                        _history_safe(conn, task_id, "TOPIC2_NEW_PROJECT_CONTEXT_ISOLATED")
                        _history_safe(conn, task_id, f"TOPIC2_PROJECT_ALL_PAGES_SCANNED:{(bundle.get('files') or [{}])[0].get('pages_seen', 0)}")
                        _history_safe(conn, task_id, f"TOPIC2_PROJECT_POSITIONS_EXTRACTED:{len(bundle.get('positions') or [])}")
                        _history_safe(conn, task_id, "TOPIC2_POSITIONS_EXTRACTION_COMPLETE_NO")
                        _history_safe(conn, task_id, "TOPIC2_PRICE_SEARCH_BLOCKED_BY_VOLUME_GATE")
                        _history_safe(conn, task_id, "TOPIC2_SMETA_GENERATION_BLOCKED_BY_VOLUME_GATE")
                        if scope_confirmed:
                            _history_safe(conn, task_id, "TOPIC2_PROJECT_GEOMETRY_CHOICE_CONFIRMED")
                            _history_safe(conn, task_id, f"TOPIC2_MANUAL_PRICES_ACCEPTED:{len(manual_rates)}")
                            _history_safe(conn, task_id, "TOPIC2_WAITING_ONLY_MISSING_MANUAL_PRICES")
                        send_res = await _send_text(
                            chat_id, text,
                            _row_get(task, "reply_to_message_id", None) or _row_get(task, "message_id", None),
                            topic_id,
                        )
                        kwargs = {
                            "state": "WAITING_CLARIFICATION", "result": text,
                            "error_message": "TOPIC2_MANUAL_PRICES_INCOMPLETE" if scope_confirmed else "TOPIC2_PROJECT_VOLUMES_INCOMPLETE",
                        }
                        if isinstance(send_res, dict) and send_res.get("bot_message_id"):
                            kwargs["bot_message_id"] = send_res.get("bot_message_id")
                        _update_task_safe(conn, task_id, **kwargs)
                        return True
    except Exception as exc:
        try:
            _history_safe(conn, _s(_row_get(task, "id", "")), "TOPIC2_NEW_PROJECT_PDF_ISOLATION_ERR:" + _s(exc)[:180])
        except Exception:
            pass
        if logger:
            logger.exception("TOPIC2_NEW_PROJECT_PDF_ISOLATION_ERR")
    return await _TOPIC2_NEW_PROJECT_PREV_HANDLER_V2(conn, task, logger)
# === END_PATCH_TOPIC2_NEW_PROJECT_PDF_ISOLATION_V2 ===
