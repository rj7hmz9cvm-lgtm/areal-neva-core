import os, logging, sqlite3, shutil
from typing import Optional
from datetime import datetime

logger = logging.getLogger(__name__)
MEMORY_DB = "/root/.areal-neva-core/data/memory.db"

def save_template(chat_id: str, topic_id: int, template_type: str, file_path: str) -> bool:
    try:
        conn = sqlite3.connect(MEMORY_DB)
        cur = conn.cursor()
        key = f"topic_{topic_id}_template_{template_type}"
        cur.execute("INSERT OR REPLACE INTO memory (chat_id, key, value, timestamp) VALUES (?,?,?,?)",
                    (chat_id, key, file_path, datetime.utcnow().isoformat()))
        conn.commit(); conn.close()
        return True
    except Exception as e:
        logger.error(f"save_template: {e}")
        return False

def get_template(chat_id: str, topic_id: int, template_type: str) -> Optional[str]:
    try:
        conn = sqlite3.connect(MEMORY_DB)
        cur = conn.cursor()
        cur.execute("SELECT value FROM memory WHERE chat_id=? AND key=?", (chat_id, f"topic_{topic_id}_template_{template_type}"))
        row = cur.fetchone()
        conn.close()
        return row[0] if row else None
    except:
        return None

def apply_template(template_path: str, output_path: str, data: list) -> bool:
    try:
        shutil.copy(template_path, output_path)
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        for i, row_data in enumerate(data, 2):
            for j, val in enumerate(row_data, 1):
                ws.cell(i, j, value=val)
        wb.save(output_path); wb.close()
        return True
    except Exception as e:
        logger.error(f"apply_template: {e}")
        return False

# === ALL_CONTOURS_TEMPLATE_MANAGER_V2 ===
def save_template(file_path, topic_id, intent):
    import shutil
    from pathlib import Path
    src=Path(file_path)
    if not src.exists():
        raise FileNotFoundError(str(file_path))
    safe="".join(c if c.isalnum() or c in ("_","-") else "_" for c in str(intent or "template"))
    out=Path("/root/.areal-neva-core/data/templates")
    out.mkdir(parents=True, exist_ok=True)
    dst=out/(str(int(topic_id or 0))+"_"+safe+(src.suffix or ".xlsx"))
    shutil.copy2(src,dst)
    return str(dst)
# === END_ALL_CONTOURS_TEMPLATE_MANAGER_V2 ===

# === FINAL_CODE_CONTOUR_TEMPLATE_APPLY_V1 ===
def apply_template(template_path, output_path, data_rows):
    from openpyxl import load_workbook
    wb=load_workbook(template_path)
    ws=wb.active
    start=2
    for r_idx,row in enumerate(data_rows or [], start):
        vals=row.values() if isinstance(row,dict) else row
        for c_idx,val in enumerate(list(vals),1):
            ws.cell(r_idx,c_idx,value=val)
    wb.save(output_path)
    return output_path
# === END_FINAL_CODE_CONTOUR_TEMPLATE_APPLY_V1 ===

# === TEMPLATE_SYSTEM_V41 ===

def template_learn_v41(file_path, chat_id=None, topic_id=0, template_type="project"):
    try:
        return save_template(str(chat_id or "default"), int(topic_id or 0), template_type, file_path)
    except TypeError:
        return save_template(file_path, topic_id, template_type)
    except Exception:
        return False


def template_priority_v41(chat_id=None, topic_id=0, template_type="project"):
    try:
        return get_template(str(chat_id or "default"), int(topic_id or 0), template_type)
    except TypeError:
        return get_template(topic_id, template_type)
    except Exception:
        return None


def project_template_engine_v41(template_path, output_path, data_rows=None):
    if not template_path:
        return {"success": False, "error": "TEMPLATE_NOT_FOUND"}
    ok = apply_template(template_path, output_path, data_rows or [])
    return {"success": bool(ok), "artifact_path": output_path if ok else None, "error": None if ok else "TEMPLATE_APPLY_FAILED"}

# === END_TEMPLATE_SYSTEM_V41 ===

# === CODE_CLOSE_V43_TEMPLATE_MANAGER ===

def _v43_template_dir():
    import os
    path = "/root/.areal-neva-core/data/templates"
    os.makedirs(path, exist_ok=True)
    return path

def template_learn_v43(file_path, topic_id=0, template_type="project"):
    import os, shutil, json
    base = _v43_template_dir()
    name = "topic_" + str(int(topic_id or 0)) + "_" + str(template_type or "project")
    ext = os.path.splitext(str(file_path))[1] or ".bin"
    dst = os.path.join(base, name + ext)
    shutil.copy2(file_path, dst)
    meta = {"topic_id": int(topic_id or 0), "template_type": template_type, "path": dst}
    with open(os.path.join(base, name + ".json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    return dst

def template_priority_v43(topic_id=0, template_type="project"):
    import os, json, glob
    base = _v43_template_dir()
    name = "topic_" + str(int(topic_id or 0)) + "_" + str(template_type or "project")
    meta = os.path.join(base, name + ".json")
    if os.path.exists(meta):
        try:
            data = json.load(open(meta, encoding="utf-8"))
            if os.path.exists(data.get("path","")):
                return data.get("path")
        except Exception:
            pass
    files = glob.glob(os.path.join(base, name + ".*"))
    files = [x for x in files if not x.endswith(".json")]
    return files[0] if files else None

def project_template_engine_v43(template_path, output_path, data_rows=None):
    import shutil, os
    if not template_path or not os.path.exists(template_path):
        return {"success": False, "error": "TEMPLATE_NOT_FOUND"}
    shutil.copy2(template_path, output_path)
    return {"success": True, "artifact_path": output_path, "template_used": template_path}

# === END_CODE_CLOSE_V43_TEMPLATE_MANAGER ===


# === PATCH_PROJECT_TEMPLATE_STORAGE_V1 ===
import json as _json_ptm
import re as _re_ptm
from pathlib import Path as _Path_ptm

_PTM_DIR = _Path_ptm("/root/.areal-neva-core/data/project_templates")

def save_project_template_model(model: dict, task_id: str = "", chat_id: str = "", topic_id: int = 0) -> str:
    _PTM_DIR.mkdir(parents=True, exist_ok=True)
    model = dict(model)
    model["task_id"] = task_id or model.get("task_id","")
    model["chat_id"] = str(chat_id or model.get("chat_id",""))
    model["topic_id"] = int(topic_id or model.get("topic_id",0) or 0)
    name = model.get("project_type","UNKNOWN")
    safe = _re_ptm.sub(r"[^A-Za-zА-Яа-я0-9_.-]+","_",f"{name}_{task_id[:8] if task_id else 'manual'}")
    path = _PTM_DIR / f"PROJECT_TEMPLATE_MODEL__{safe}.json"
    path.write_text(_json_ptm.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(path)

# === END PATCH_PROJECT_TEMPLATE_STORAGE_V1 ===
