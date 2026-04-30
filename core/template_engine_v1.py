# === TEMPLATE_ENGINE_V1 ===
import os, json, logging, glob
from typing import Optional, Dict, Any
logger = logging.getLogger(__name__)

BASE = "/root/.areal-neva-core"
TEMPLATE_DIR = f"{BASE}/data/templates"
TRIGGER_PHRASES = [
    "сделай так же", "по образцу", "как в прошлый раз",
    "аналогично", "такой же", "такую же", "такое же",
    "по шаблону", "используй шаблон", "как раньше"
]

def is_template_request(text: str) -> bool:
    low = text.lower()
    return any(t in low for t in TRIGGER_PHRASES)

def save_template(topic_id: int, file_path: str, template_type: str = "estimate") -> bool:
    """Сохранить файл как шаблон для топика"""
    try:
        os.makedirs(f"{TEMPLATE_DIR}/{template_type}", exist_ok=True)
        import shutil
        ext = os.path.splitext(file_path)[1]
        dest = f"{TEMPLATE_DIR}/{template_type}/ACTIVE__topic_{topic_id}{ext}"
        shutil.copy2(file_path, dest)
        meta = {
            "topic_id": topic_id,
            "type": template_type,
            "source": file_path,
            "saved_at": __import__("datetime").datetime.utcnow().isoformat()
        }
        with open(f"{TEMPLATE_DIR}/{template_type}/ACTIVE__topic_{topic_id}.json", "w") as f:
            json.dump(meta, f, ensure_ascii=False)
        logger.info("TEMPLATE_ENGINE_V1 saved topic=%s type=%s", topic_id, template_type)
        return True
    except Exception as e:
        logger.error("TEMPLATE_SAVE_ERR %s", e)
        return False

def get_template(topic_id: int, template_type: str = "estimate") -> Optional[str]:
    """Получить путь к шаблону топика"""
    try:
        patterns = [
            f"{TEMPLATE_DIR}/{template_type}/ACTIVE__topic_{topic_id}.*",
            f"{TEMPLATE_DIR}/estimate/ACTIVE__topic_{topic_id}.*",
        ]
        for pat in patterns:
            hits = [f for f in glob.glob(pat) if not f.endswith(".json")]
            if hits:
                return hits[0]
    except Exception as e:
        logger.warning("TEMPLATE_GET_ERR %s", e)
    return None

def apply_template_to_xlsx(template_path: str, rows: list, output_path: str) -> bool:
    """Применить структуру шаблона к новым данным"""
    try:
        from openpyxl import load_workbook
        import copy
        wb_tpl = load_workbook(template_path)
        ws_tpl = wb_tpl.active

        # Копируем структуру — заголовки из шаблона
        wb_new = load_workbook(template_path)
        ws_new = wb_new.active

        # Очищаем данные, оставляем заголовки (строка 1-2)
        header_rows = 2
        for row in ws_new.iter_rows(min_row=header_rows+1, max_row=ws_new.max_row):
            for cell in row:
                cell.value = None

        # Заполняем новыми данными с сохранением формул
        for i, item in enumerate(rows, start=header_rows+1):
            ws_new.cell(i, 1, value=i - header_rows)
            ws_new.cell(i, 2, value=str(item.get("name", "")))
            ws_new.cell(i, 3, value=str(item.get("unit", "шт")))
            ws_new.cell(i, 4, value=float(item.get("qty", 0) or 0))
            ws_new.cell(i, 5, value=float(item.get("price", 0) or 0))
            ws_new.cell(i, 6, f"=D{i}*E{i}")

        # Итог
        last = header_rows + len(rows)
        ws_new.cell(last+1, 6, f"=SUM(F{header_rows+1}:F{last})")

        wb_new.save(output_path)
        logger.info("TEMPLATE_APPLIED_V1 output=%s rows=%s", output_path, len(rows))
        return True
    except Exception as e:
        logger.error("TEMPLATE_APPLY_ERR %s", e)
        return False

def detect_template_type(file_name: str, intent: str = "") -> str:
    fn = file_name.lower()
    if any(e in fn for e in [".xlsx", ".xls", ".csv"]):
        return "estimate"
    if any(e in fn for e in [".docx", ".doc"]):
        return "technadzor"
    if "estimate" in intent or "смет" in intent:
        return "estimate"
    return "estimate"
# === END TEMPLATE_ENGINE_V1 ===
