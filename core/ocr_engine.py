# === FINAL_CLOSURE_BLOCKER_FIX_V1_OCR_TABLE_ENGINE ===
from __future__ import annotations

import csv
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

BASE = Path("/root/.areal-neva-core")
OUT = BASE / "outputs" / "ocr"
OUT.mkdir(parents=True, exist_ok=True)

UNIT_RE = re.compile(r"\b([мm]2|[мm]²|[мm]3|[мm]³|п\.?м|пог\.?м|шт|кг|тн|тонн|т|[мm]|мм|компл)\b", re.I)
NUM_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _safe_task_id(task_id: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_-]+", "_", _s(task_id))
    return s[:36] or datetime.now().strftime("%Y%m%d_%H%M%S")


def _num(v: Any) -> float:
    try:
        cleaned = _s(v).replace("\u00a0", " ").replace(" ", "").replace(",", ".")
        m = NUM_RE.search(cleaned)
        return float(m.group(0).replace(" ", "")) if m else 0.0
    except Exception:
        return 0.0


def _unit(v: Any) -> str:
    s = _s(v).lower()
    return (
        s.replace("m2", "м²")
        .replace("m3", "м³")
        .replace("m²", "м²")
        .replace("m³", "м³")
        .replace("м2", "м²")
        .replace("м3", "м³")
        .replace("пог.м", "п.м")
        .replace("погм", "п.м")
        .replace("пм", "п.м")
    )


def is_ocr_table_intent(text: str = "", file_name: str = "") -> bool:
    t = f"{text} {file_name}".lower().replace("ё", "е")
    return any(x in t for x in ["таблиц", "распознай", "ocr", "скан", "фото таблицы", "в excel", "в эксель"])


def _write_csv_diagnostic(task_id: str, message: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = OUT / f"OCR_TABLE__{_safe_task_id(task_id)}__{ts}.csv"
    rows: List[List[str]] = [["status", "message"], ["FAILED", message]]
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows(rows)
    return str(csv_path)


def _clean_line(text: str) -> str:
    return re.sub(r"\s+", " ", _s(text)).strip(" \t|;")


def _table_rows_from_cells(rows: Iterable[Iterable[Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for raw in rows or []:
        cells = [_clean_line(str(c)) for c in raw if _clean_line(str(c))]
        if not cells:
            continue
        joined = " | ".join(cells)
        item = _parse_table_line(joined)
        if item:
            item["source"] = "ocr_table_cells"
            out.append(item)
    return out


def _parse_table_line(line: str) -> Dict[str, Any]:
    clean = _clean_line(line)
    if len(clean) < 5:
        return {}
    unit_m = UNIT_RE.search(clean)
    if not unit_m:
        return {}
    numeric_source = f"{clean[:unit_m.start()]} {clean[unit_m.end():]}"
    nums = [_num(m.group(0)) for m in NUM_RE.finditer(numeric_source)]
    nums = [n for n in nums if n != 0]
    if not nums:
        return {}
    qty = nums[0]
    price = nums[1] if len(nums) >= 2 else 0.0
    total = nums[2] if len(nums) >= 3 else (round(qty * price, 2) if price else 0.0)

    name_part = clean[: unit_m.start()]
    if not re.search(r"[A-Za-zА-Яа-яЁё]", name_part):
        name_part = re.split(NUM_RE, clean, maxsplit=1)[0]
    name = re.sub(r"^\s*\d+[\).:-]?\s*", "", name_part).strip(" -|;:")
    if len(re.findall(r"[A-Za-zА-Яа-яЁё]", name)) < 3:
        return {}
    return {
        "name": name[:240],
        "unit": _unit(unit_m.group(1)),
        "qty": qty,
        "price": price,
        "total": total,
        "source": "ocr_text_line",
        "raw": clean[:500],
    }


def _extract_image_text(file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    from PIL import Image, ImageOps
    import pytesseract

    img = Image.open(file_path)
    img = ImageOps.grayscale(img)
    text = pytesseract.image_to_string(img, lang="rus+eng", config="--psm 6")
    lines = [_clean_line(x) for x in text.splitlines() if _clean_line(x)]
    return lines, [_parse_table_line(x) for x in lines if _parse_table_line(x)]


def _extract_pdf_text(file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    lines: List[str] = []
    rows: List[Dict[str, Any]] = []
    try:
        import pdfplumber

        with pdfplumber.open(file_path) as pdf:
            for page_no, page in enumerate(pdf.pages[:30], 1):
                try:
                    for table in page.extract_tables() or []:
                        parsed = _table_rows_from_cells(table or [])
                        for item in parsed:
                            item["page"] = page_no
                        rows.extend(parsed)
                except Exception:
                    pass
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
                for line in text.splitlines():
                    clean = _clean_line(line)
                    if clean:
                        lines.append(clean)
    except Exception:
        pass
    for line in lines:
        item = _parse_table_line(line)
        if item:
            rows.append(item)
    return lines, rows


def _extract_pdf_ocr(file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    from pdf2image import convert_from_path
    import pytesseract

    lines: List[str] = []
    rows: List[Dict[str, Any]] = []
    pages = convert_from_path(file_path, dpi=200, first_page=1, last_page=12)
    for page_no, page in enumerate(pages, 1):
        text = pytesseract.image_to_string(page, lang="rus+eng", config="--psm 6")
        for line in text.splitlines():
            clean = _clean_line(line)
            if not clean:
                continue
            lines.append(clean)
            item = _parse_table_line(clean)
            if item:
                item["page"] = page_no
                rows.append(item)
    return lines, rows


def _extract_spreadsheet_rows(file_path: str) -> Tuple[List[str], List[Dict[str, Any]], List[List[Any]]]:
    ext = Path(file_path).suffix.lower()
    raw_rows: List[List[Any]] = []
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(file_path)
        for sheet in book.sheets():
            for r in range(sheet.nrows):
                raw_rows.append(sheet.row_values(r))
    else:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    raw_rows.append(list(row or []))
        finally:
            wb.close()
    lines = [_clean_line(" | ".join(_s(c) for c in row if _s(c))) for row in raw_rows]
    lines = [x for x in lines if x]
    return lines, _table_rows_from_cells(raw_rows), raw_rows


def _extract_text_rows(file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
    lines = [_clean_line(x) for x in text.splitlines() if _clean_line(x)]
    return lines, [_parse_table_line(x) for x in lines if _parse_table_line(x)]


def _dedup_rows(rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen = set()
    for row in rows or []:
        if not row:
            continue
        key = (row.get("name"), row.get("unit"), row.get("qty"), row.get("price"), row.get("total"), row.get("raw"))
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _write_xlsx(task_id: str, lines: List[str], rows: List[Dict[str, Any]], raw_rows: List[List[Any]] | None = None) -> str:
    from openpyxl import Workbook

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = OUT / f"OCR_TABLE__{_safe_task_id(task_id)}__{ts}.xlsx"
    wb = Workbook()

    ws_text = wb.active
    ws_text.title = "OCR_TEXT"
    ws_text.append(["№", "Текст"])
    for i, line in enumerate(lines or [], 1):
        ws_text.append([i, line])
    ws_text.column_dimensions["A"].width = 8
    ws_text.column_dimensions["B"].width = 120

    ws_rows = wb.create_sheet("TABLE_ROWS")
    ws_rows.append(["№", "Наименование", "Ед.изм", "Кол-во", "Цена", "Сумма", "Источник", "Строка OCR"])
    for i, row in enumerate(rows or [], 2):
        ws_rows.cell(i, 1, value=i - 1)
        ws_rows.cell(i, 2, value=row.get("name"))
        ws_rows.cell(i, 3, value=row.get("unit"))
        ws_rows.cell(i, 4, value=float(row.get("qty") or 0))
        ws_rows.cell(i, 5, value=float(row.get("price") or 0))
        if row.get("price"):
            ws_rows.cell(i, 6, value=f"=D{i}*E{i}")
        else:
            ws_rows.cell(i, 6, value=float(row.get("total") or 0))
        ws_rows.cell(i, 7, value=row.get("source"))
        ws_rows.cell(i, 8, value=row.get("raw", ""))
    total_row = (len(rows or []) + 2)
    ws_rows.cell(total_row, 5, value="ИТОГО:")
    ws_rows.cell(total_row, 6, value=f"=SUM(F2:F{max(1, total_row - 1)})")
    for col, width in {"A": 8, "B": 70, "C": 12, "D": 14, "E": 14, "F": 16, "G": 22, "H": 90}.items():
        ws_rows.column_dimensions[col].width = width

    if raw_rows:
        ws_raw = wb.create_sheet("RAW_TABLE")
        for row in raw_rows:
            ws_raw.append(list(row or []))

    wb.save(out_xlsx)
    wb.close()
    return str(out_xlsx)


def _process_ocr_file(file_path: str, task_id: str, topic_id: int = 0) -> Dict[str, Any]:
    if not file_path or not os.path.exists(str(file_path)):
        diag = _write_csv_diagnostic(task_id, "FILE_NOT_FOUND")
        return {"success": False, "handled": True, "state": "FAILED", "artifact_path": diag, "error": "FILE_NOT_FOUND"}

    from core.engine_base import detect_real_file_type, upload_artifact_to_drive

    real_type = detect_real_file_type(file_path)
    ext = Path(file_path).suffix.lower()
    lines: List[str] = []
    rows: List[Dict[str, Any]] = []
    raw_rows: List[List[Any]] = []

    try:
        if real_type in ("jpg", "png", "image") or ext in (".jpg", ".jpeg", ".png", ".webp", ".heic"):
            lines, rows = _extract_image_text(file_path)
        elif real_type == "pdf":
            lines, rows = _extract_pdf_text(file_path)
            if not lines and not rows:
                lines, rows = _extract_pdf_ocr(file_path)
        elif real_type in ("xlsx", "xls") or ext in (".xlsx", ".xls"):
            lines, rows, raw_rows = _extract_spreadsheet_rows(file_path)
        elif real_type in ("csv", "txt") or ext in (".csv", ".txt"):
            lines, rows = _extract_text_rows(file_path)
        elif real_type == "invalid_pdf":
            diag = _write_csv_diagnostic(task_id, "INVALID_PDF_SIGNATURE")
            return {"success": False, "handled": True, "state": "FAILED", "artifact_path": diag, "error": "INVALID_PDF_SIGNATURE"}
        else:
            diag = _write_csv_diagnostic(task_id, f"UNSUPPORTED_OCR_FILE_TYPE:{real_type}")
            return {
                "success": False,
                "handled": True,
                "state": "FAILED",
                "artifact_path": diag,
                "error": f"UNSUPPORTED_OCR_FILE_TYPE:{real_type}",
            }

        rows = _dedup_rows(rows)
        if not lines and not rows and not raw_rows:
            diag = _write_csv_diagnostic(task_id, "OCR_NO_TEXT_EXTRACTED")
            return {"success": False, "handled": True, "state": "FAILED", "artifact_path": diag, "error": "OCR_NO_TEXT_EXTRACTED"}

        xlsx = _write_xlsx(task_id, lines, rows, raw_rows)
        drive_link = upload_artifact_to_drive(xlsx, task_id, int(topic_id or 0))
        table_note = f"Табличных строк: {len(rows)}" if rows else "Табличные строки не найдены, текст сохранён без выдумывания структуры"
        link_note = f"\nXLSX: {drive_link}" if drive_link else "\nXLSX подготовлен, Drive link не получен"
        return {
            "success": True,
            "handled": True,
            "state": "AWAITING_CONFIRMATION",
            "kind": "ocr_table",
            "artifact_path": xlsx,
            "excel_path": xlsx,
            "drive_link": drive_link,
            "result_text": f"OCR выполнен\nСтрок текста: {len(lines)}\n{table_note}{link_note}",
            "history": "FINAL_CLOSURE_BLOCKER_FIX_V1:OCR_REAL_ENGINE_DONE",
        }
    except Exception as e:
        diag = _write_csv_diagnostic(task_id, f"OCR_ENGINE_ERROR:{e}")
        return {
            "success": False,
            "handled": True,
            "state": "FAILED",
            "artifact_path": diag,
            "error": f"OCR_ENGINE_ERROR:{e}",
            "history": "FINAL_CLOSURE_BLOCKER_FIX_V1:OCR_REAL_ENGINE_FAILED",
        }


def process_ocr_table(text: str = "", task_id: str = "", file_path: str = "", file_name: str = "") -> Dict[str, Any]:
    if not is_ocr_table_intent(text, file_name):
        return {"ok": False, "handled": False, "reason": "NOT_OCR_TABLE"}
    if not file_path:
        return {
            "ok": True,
            "handled": True,
            "kind": "ocr_table",
            "state": "WAITING_CLARIFICATION",
            "message": "Пришли файл или фото таблицы для OCR. Без файла структуру не выдумываю.",
            "history": "FINAL_CLOSURE_BLOCKER_FIX_V1:OCR_WAITING_FILE",
        }
    res = _process_ocr_file(file_path=file_path, task_id=task_id, topic_id=0)
    res["ok"] = bool(res.get("success"))
    res["message"] = res.get("result_text") or res.get("error") or "OCR завершён"
    return res


async def process_image_to_excel(file_path: str, task_id: str, topic_id: int) -> Dict[str, Any]:
    return _process_ocr_file(file_path=file_path, task_id=task_id, topic_id=topic_id)


# === END_FINAL_CLOSURE_BLOCKER_FIX_V1_OCR_TABLE_ENGINE ===
