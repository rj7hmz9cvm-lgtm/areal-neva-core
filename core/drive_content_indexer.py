# === DRIVE_FILE_CONTENT_MEMORY_INDEX_V1 ===
from __future__ import annotations

import csv
import io
import json
import os
import re
import sqlite3
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv

BASE = "/root/.areal-neva-core"
MEM_DB = f"{BASE}/data/memory.db"
load_dotenv(f"{BASE}/.env", override=True)

MAX_TEXT = 50000
MAX_ROWS = 300


def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    try:
        return json.dumps(v, ensure_ascii=False)
    except Exception:
        return str(v).strip()


def _clean(text: str, limit: int = MAX_TEXT) -> str:
    text = (text or "").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()[:limit]


def _kind(file_name: str, mime_type: str = "") -> str:
    ext = os.path.splitext((file_name or "").lower())[1]
    mime = (mime_type or "").lower()
    if ext in (".xlsx", ".xlsm", ".csv") or "spreadsheet" in mime or mime == "text/csv":
        return "table"
    if ext in (".pdf", ".docx", ".doc", ".txt") or mime in ("application/pdf", "text/plain"):
        return "document"
    return "skip"


def _drive_service():
    from core.topic_drive_oauth import _oauth_service
    return _oauth_service()


def download_drive_file(file_id: str, file_name: str) -> Optional[str]:
    if not file_id:
        return None
    service = _drive_service()
    suffix = os.path.splitext(file_name or "")[1] or ".bin"
    fd, out = tempfile.mkstemp(prefix="drive_content_", suffix=suffix, dir="/tmp")
    os.close(fd)

    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    from googleapiclient.http import MediaIoBaseDownload
    with io.FileIO(out, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return out


def extract_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        parts = []
        for page in reader.pages[:80]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                pass
        return _clean("\n".join(parts), MAX_TEXT)
    except Exception as e:
        return f"PDF_PARSE_ERROR: {e}"


def extract_docx(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return _clean("\n".join(p.text for p in doc.paragraphs if p.text), MAX_TEXT)
    except Exception as e:
        return f"DOCX_PARSE_ERROR: {e}"


def extract_txt(path: str) -> str:
    try:
        return _clean(Path(path).read_text(encoding="utf-8", errors="ignore"), MAX_TEXT)
    except Exception as e:
        return f"TXT_PARSE_ERROR: {e}"


def extract_table(path: str, file_name: str) -> str:
    rows: List[str] = []
    ext = os.path.splitext((file_name or "").lower())[1]

    try:
        if ext == ".csv":
            with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
                reader = csv.reader(f)
                for idx, row in enumerate(reader):
                    rows.append(" | ".join(_s(x) for x in row))
                    if idx >= MAX_ROWS:
                        break
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            for ws in wb.worksheets[:5]:
                rows.append(f"=== SHEET: {ws.title} ===")
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    vals = [_s(x) for x in row if _s(x)]
                    if vals:
                        rows.append(" | ".join(vals))
                    if idx >= MAX_ROWS:
                        break
    except Exception as e:
        rows.append(f"TABLE_PARSE_ERROR: {e}")

    return _clean("\n".join(rows), MAX_TEXT)


def extract_content(local_path: str, file_name: str, mime_type: str = "") -> Dict[str, Any]:
    kind = _kind(file_name, mime_type)
    ext = os.path.splitext((file_name or "").lower())[1]

    if kind == "table":
        text = extract_table(local_path, file_name)
    elif kind == "document":
        if ext == ".pdf":
            text = extract_pdf(local_path)
        elif ext == ".docx":
            text = extract_docx(local_path)
        else:
            text = extract_txt(local_path)
    else:
        text = ""

    return {
        "ok": bool(text and not text.endswith("_PARSE_ERROR")),
        "kind": kind,
        "file_name": file_name,
        "mime_type": mime_type,
        "content": _clean(text, MAX_TEXT),
        "chars": len(text or ""),
    }


def save_file_content_memory(chat_id: str, topic_id: int, task_id: str, file_id: str, file_name: str, mime_type: str, content: str) -> Dict[str, Any]:
    if not content.strip():
        return {"ok": False, "reason": "EMPTY_CONTENT"}

    key = f"topic_{int(topic_id or 0)}_file_content_{task_id}"
    value = json.dumps({
        "task_id": task_id,
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "file_id": file_id,
        "file_name": file_name,
        "mime_type": mime_type,
        "content": _clean(content, MAX_TEXT),
    }, ensure_ascii=False)

    with sqlite3.connect(MEM_DB) as conn:
        conn.execute("CREATE TABLE IF NOT EXISTS memory (id TEXT PRIMARY KEY, chat_id TEXT, key TEXT, value TEXT, timestamp TEXT)")
        existing = conn.execute("SELECT 1 FROM memory WHERE chat_id=? AND key=? LIMIT 1", (str(chat_id), key)).fetchone()
        if existing:
            return {"ok": True, "key": key, "dedup": True}
        import hashlib
        mid = hashlib.sha1(f"{chat_id}:{key}".encode()).hexdigest()
        conn.execute(
            "INSERT OR IGNORE INTO memory (id, chat_id, key, value, timestamp) VALUES (?,?,?,?,datetime('now'))",
            (mid, str(chat_id), key, value),
        )
        conn.commit()

    return {"ok": True, "key": key, "dedup": False}


def index_drive_file_content(chat_id: str, topic_id: int, task_id: str, file_id: str, file_name: str, mime_type: str = "") -> Dict[str, Any]:
    local_path = None
    try:
        if _kind(file_name, mime_type) == "skip":
            return {"ok": False, "reason": "UNSUPPORTED_TYPE", "file_name": file_name}

        local_path = download_drive_file(file_id, file_name)
        if not local_path or not os.path.exists(local_path):
            return {"ok": False, "reason": "DOWNLOAD_FAILED", "file_name": file_name}

        extracted = extract_content(local_path, file_name, mime_type)
        if not extracted.get("content"):
            return {"ok": False, "reason": "EXTRACT_EMPTY", "file_name": file_name}

        saved = save_file_content_memory(
            chat_id=str(chat_id),
            topic_id=int(topic_id or 0),
            task_id=str(task_id),
            file_id=str(file_id),
            file_name=str(file_name),
            mime_type=str(mime_type or ""),
            content=str(extracted.get("content") or ""),
        )
        return {
            "ok": bool(saved.get("ok")),
            "reason": "INDEXED" if saved.get("ok") else saved.get("reason"),
            "key": saved.get("key"),
            "dedup": saved.get("dedup", False),
            "kind": extracted.get("kind"),
            "chars": extracted.get("chars"),
            "file_name": file_name,
        }
    except Exception as e:
        return {"ok": False, "reason": f"ERROR:{e}", "file_name": file_name}
    finally:
        if local_path:
            try:
                os.remove(local_path)
            except Exception:
                pass
# === END DRIVE_FILE_CONTENT_MEMORY_INDEX_V1 ===
