from __future__ import annotations
import csv
import logging
from pathlib import Path

logger = logging.getLogger("core.document_context")

TEXT_EXT = {".txt", ".md", ".csv"}
TABLE_EXT = {".xlsx", ".xls", ".csv"}
DOC_EXT = {".pdf", ".docx"}
IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp"}

def _safe_read_text(path: Path, limit: int = 12000) -> str:
    for enc in ("utf-8", "cp1251", "latin-1"):
        try:
            return path.read_text(encoding=enc, errors="ignore")[:limit]
        except Exception:
            continue
    return ""

def _read_csv_preview(path: Path, row_limit: int = 20, col_limit: int = 12) -> str:
    try:
        rows_out = []
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= row_limit:
                    break
                rows_out.append(" | ".join([str(x).strip() for x in row[:col_limit]]))
        return "\n".join(rows_out)
    except Exception as exc:
        logger.warning("csv read fail %s %s", path, exc)
        return ""

def _read_xlsx_preview(path: Path, row_limit: int = 25, col_limit: int = 12) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        parts = []
        for ws in wb.worksheets[:3]:
            parts.append(f"[SHEET] {ws.title}")
            count = 0
            for row in ws.iter_rows(values_only=True):
                vals = ["" if cell is None else str(cell).strip() for cell in row[:col_limit]]
                if any(vals):
                    parts.append(" | ".join(vals))
                    count += 1
                if count >= row_limit:
                    break
        return "\n".join(parts)[:15000]
    except Exception as exc:
        logger.warning("xlsx read fail %s %s", path, exc)
        return ""

def _read_docx_preview(path: Path, limit: int = 15000) -> str:
    try:
        from docx import Document
        doc = Document(str(path))
        parts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        return "\n".join(parts)[:limit]
    except Exception as exc:
        logger.warning("docx read fail %s %s", path, exc)
        return ""

def _read_pdf_preview(path: Path, limit: int = 18000) -> str:
    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages[:12]:
                txt = (page.extract_text() or "").strip()
                if txt:
                    parts.append(txt)
        text = "\n".join(parts).strip()
        if text:
            return text[:limit]
    except Exception as exc:
        logger.warning("pdfplumber fail %s %s", path, exc)
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages[:12]:
            txt = (page.extract_text() or "").strip()
            if txt:
                parts.append(txt)
        return "\n".join(parts)[:limit]
    except Exception as exc:
        logger.warning("pypdf fail %s %s", path, exc)
        return ""

def _ocr_image_preview(path: Path, limit: int = 10000) -> str:
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(str(path))
        text = pytesseract.image_to_string(img, lang="eng+rus")
        return (text or "").strip()[:limit]
    except Exception as exc:
        logger.warning("ocr fail %s %s", path, exc)
        return ""

def extract_document_context(path_str: str) -> dict:
    path = Path(path_str).expanduser()
    if not path.exists() or not path.is_file():
        return {"ok": False, "path": str(path), "error": "file_not_found", "text": ""}
    ext = path.suffix.lower()
    text = ""
    if ext in TEXT_EXT:
        text = _read_csv_preview(path) if ext == ".csv" else _safe_read_text(path)
    elif ext in TABLE_EXT:
        text = _read_csv_preview(path) if ext == ".csv" else _read_xlsx_preview(path)
    elif ext == ".docx":
        text = _read_docx_preview(path)
    elif ext == ".pdf":
        text = _read_pdf_preview(path)
    elif ext in IMAGE_EXT:
        text = _ocr_image_preview(path)
    elif ext in {".dwg", ".dxf"}:
        text = f"DWG/DXF detected: {path.name}"
    else:
        text = _safe_read_text(path)
    return {"ok": True, "path": str(path), "name": path.name, "ext": ext, "text": (text or "").strip()}

def format_document_context(path_str: str) -> str:
    data = extract_document_context(path_str)
    if not data.get("ok"):
        return ""
    text = (data.get("text") or "").strip()
    if not text:
        return f"[DOC] {data.get('name')} ({data.get('ext')})"
    return f"[DOC] {data.get('name')} ({data.get('ext')})\n{text[:12000]}"
