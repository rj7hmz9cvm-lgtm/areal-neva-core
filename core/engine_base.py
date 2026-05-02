import os, logging, hashlib, sqlite3, re
from typing import Dict, Any, Optional
from datetime import datetime, timezone

logger = logging.getLogger(__name__)
BASE = "/root/.areal-neva-core"
DB_PATH = f"{BASE}/data/core.db"

STAGES = ["INGESTED", "DOWNLOADED", "PARSED", "CLEANED", "NORMALIZED", "VALIDATED", "CALCULATED", "ARTIFACT_CREATED", "UPLOADED", "COMPLETED", "FAILED"]
UNIT_NORMALIZATION = {"м2": "м²", "кв.м": "м²", "м3": "м³", "куб.м": "м³", "шт": "шт", "кг": "кг", "т": "т", "тн": "т", "п.м": "п.м"}
FALSE_NUMBERS = ["B25", "B30", "B15", "A500", "A240", "A400", "12мм", "20мм", "10мм"]
BUILDING_DICT = {"бетон B25": "Бетон", "бетон B30": "Бетон", "доска 50х150": "Доска обрезная", "арматура A500": "Арматура"}


def _run_upload_sync(fn, *args, **kwargs):
    import asyncio
    import inspect
    import threading

    box = {"value": None, "error": None}

    def _runner():
        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            value = fn(*args, **kwargs)
            if inspect.isawaitable(value):
                value = loop.run_until_complete(value)
            box["value"] = value
        except Exception as e:
            box["error"] = e
        finally:
            try:
                loop.close()
            except Exception:
                pass
            try:
                asyncio.set_event_loop(None)
            except Exception:
                pass

    t = threading.Thread(target=_runner, daemon=True)
    t.start()
    t.join()

    if box["error"] is not None:
        raise box["error"]

    return box["value"]

def get_db(): return sqlite3.connect(DB_PATH)

def update_drive_file_stage(task_id: str, drive_file_id: str, stage: str) -> bool:
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM drive_files WHERE task_id=? AND drive_file_id=?", (task_id, drive_file_id))
        if cur.fetchone():
            cur.execute("UPDATE drive_files SET stage=? WHERE task_id=? AND drive_file_id=?", (stage, task_id, drive_file_id))
        else:
            cur.execute("INSERT INTO drive_files (task_id, drive_file_id, stage, created_at) VALUES (?,?,?,?)", (task_id, drive_file_id, stage, datetime.now(timezone.utc).isoformat()))
        conn.commit(); conn.close()
        return True
    except Exception as e:
        logger.error(f"update_drive_file_stage: {e}")
        return False


def detect_real_file_type(file_path: str) -> str:
    try:
        with open(file_path, "rb") as f:
            header = f.read(8)
    except Exception:
        header = b""

    ext = os.path.splitext(file_path)[1].lower()

    if header.startswith(b"%PDF"):
        return "pdf"
    if header.startswith(b"PK\x03\x04"):
        if ext in (".xlsx", ".xls"):
            return "xlsx"
        if ext in (".docx", ".doc"):
            return "docx"
        if ext == ".zip":
            return "zip"
        return "zip_or_office"
    if header.startswith(b"\xFF\xD8\xFF"):
        return "jpg"
    if header.startswith(b"\x89PNG"):
        return "png"
    if header.startswith(b"Rar!"):
        return "rar"
    if header.startswith(b"7z\xBC\xAF"):
        return "7z"
    if header.startswith(b"AC10") or ext in (".dwg", ".dxf"):
        return "dwg"

    ext_map = {
        ".csv": "csv",
        ".txt": "txt",
        ".heic": "image",
        ".webp": "image",
        ".jpg": "jpg",
        ".jpeg": "jpg",
        ".png": "png",
        ".pdf": "invalid_pdf",
    }
    return ext_map.get(ext, "unknown")


def calculate_file_hash(file_path: str) -> str:
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        for b in iter(lambda: f.read(4096), b""): sha.update(b)
    return sha.hexdigest()


# === PATCH_DRIVE_DIRECT_OAUTH_V1 ===
def _telegram_fallback_send(local_path: str, task_id: str, topic_id: int) -> str:
    """TELEGRAM_FALLBACK_V1 — отправить файл в Telegram если Drive недоступен"""
    try:
        import requests, os
        BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
        CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "-1003725299009")
        if not BOT_TOKEN or not os.path.exists(local_path):
            return ""
        caption = f"[DRIVE_UNAVAIL] Файл задачи {task_id[:8]} — Drive недоступен, отправляю напрямую"
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
        with open(local_path, "rb") as f:
            resp = requests.post(url, data={
                "chat_id": CHAT_ID,
                "message_thread_id": str(topic_id) if topic_id else "",
                "caption": caption,
            }, files={"document": f}, timeout=60)
        if resp.ok:
            result = resp.json()
            file_id = result.get("result", {}).get("document", {}).get("file_id", "")
            logger.info("TELEGRAM_FALLBACK_V1 sent file_id=%s task=%s", file_id, task_id)
            return f"telegram://file/{file_id}"
        else:
            logger.warning("TELEGRAM_FALLBACK_V1 failed status=%s", resp.status_code)
            return ""
    except Exception as e:
        logger.warning("TELEGRAM_FALLBACK_V1 err=%s", e)
        return ""

# === DRIVE_TOPIC_FOLDER_ENFORCER_V1 ===
def _drive_creds_v1():
    import os
    from dotenv import load_dotenv
    load_dotenv("/root/.areal-neva-core/.env", override=False)
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    creds = Credentials(
        None,
        refresh_token=os.environ["GDRIVE_REFRESH_TOKEN"],
        token_uri="https://oauth2.googleapis.com/token",
        client_id=os.environ["GDRIVE_CLIENT_ID"],
        client_secret=os.environ["GDRIVE_CLIENT_SECRET"],
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    creds.refresh(Request())
    return creds

def _drive_svc_v1():
    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=_drive_creds_v1(), cache_discovery=False)

def _drive_get_or_create_folder(svc, name: str, parent_id: str) -> str:
    safe = str(name or "").replace("'", "\'")
    q = f"mimeType=\'application/vnd.google-apps.folder\' and trashed=false and name=\'{safe}\' and \'{parent_id}\' in parents"
    r = svc.files().list(q=q, fields="files(id)", pageSize=1).execute()
    files = r.get("files") or []
    if files:
        return files[0]["id"]
    f = svc.files().create(
        body={"name": str(name), "mimeType": "application/vnd.google-apps.folder", "parents": [parent_id]},
        fields="id",
    ).execute()
    return f.get("id") or ""

def get_drive_topic_folder_id(topic_id: int, chat_id: str = "") -> str:
    import os
    from dotenv import load_dotenv
    load_dotenv("/root/.areal-neva-core/.env", override=False)
    svc = _drive_svc_v1()
    root = os.environ.get("DRIVE_INGEST_FOLDER_ID", "13No7_E7Mwj1n1awNQ-lzbohWGOiEM2PB")
    chat = str(chat_id or os.environ.get("TELEGRAM_CHAT_ID", "-1003725299009"))
    chat_folder = _drive_get_or_create_folder(svc, f"chat_{chat}", root)
    return _drive_get_or_create_folder(svc, f"topic_{int(topic_id or 0)}", chat_folder)

def upload_artifact_to_drive(file_path: str, task_id: str, topic_id: int):
    import logging, mimetypes, os
    _logger = logging.getLogger(__name__)
    if not file_path or not os.path.exists(str(file_path)):
        _logger.error("DRIVE_TOPIC_FOLDER_ENFORCER_V1_NOT_FOUND task=%s path=%s", task_id, file_path)
        return None
    try:
        from googleapiclient.http import MediaFileUpload
        svc = _drive_svc_v1()
        folder_id = get_drive_topic_folder_id(int(topic_id or 0))
        name = os.path.basename(str(file_path))
        mime = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        f = svc.files().create(
            body={"name": name, "parents": [folder_id]},
            media_body=MediaFileUpload(str(file_path), mimetype=mime, resumable=True),
            fields="id,webViewLink",
        ).execute()
        fid = f.get("id")
        if not fid:
            return None
        try:
            svc.permissions().create(fileId=fid, body={"role": "reader", "type": "anyone"}, fields="id").execute()
        except Exception as pe:
            _logger.warning("DRIVE_TOPIC_FOLDER_ENFORCER_V1_PERM_ERR task=%s err=%s", task_id, pe)
        link = f.get("webViewLink") or f"https://drive.google.com/file/d/{fid}/view"
        _logger.info("DRIVE_TOPIC_FOLDER_ENFORCER_V1_OK task=%s topic=%s link=%s", task_id, topic_id, link)
        return link
    except Exception as e:
        _logger.error("DRIVE_TOPIC_FOLDER_ENFORCER_V1_FAILED task=%s err=%s", task_id, e)
        return None
# === END_DRIVE_TOPIC_FOLDER_ENFORCER_V1 ===

def quality_gate(file_path: str, task_id: str, expected_type: str = "excel") -> Dict[str, Any]:
    err, warn = [], []
    if not os.path.exists(file_path): err.append("File not found")
    else:
        sz = os.path.getsize(file_path)
        if sz == 0: err.append("Empty file")
        elif sz > 50*1024*1024: warn.append("File >50MB")
    if expected_type == "excel" and file_path.endswith(('.xlsx','.xls')):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            has_formulas = any(cell.data_type == 'f' for sheet in wb for row in sheet.iter_rows() for cell in row)
            if not has_formulas: warn.append("No formulas found")
            wb.close()
        except: err.append("Excel validation failed")
    return {"passed": len(err)==0, "errors": err, "warnings": warn}

def normalize_unit(unit: str) -> str:
    return UNIT_NORMALIZATION.get(unit.lower().strip(), unit)

def is_false_number(val: str) -> bool:
    return any(fn in str(val) for fn in FALSE_NUMBERS)

def normalize_item_name(name: str) -> str:
    for k, v in BUILDING_DICT.items():
        if k in name.lower(): return v
    return name

def is_duplicate_task(conn, chat_id: str, topic_id: int, prompt: str, file_hash: str) -> bool:
    cur = conn.execute("SELECT id FROM tasks WHERE chat_id=? AND topic_id=? AND raw_input=? AND result LIKE ?", (chat_id, topic_id, prompt, f"%{file_hash}%"))
    return cur.fetchone() is not None

def should_retry(task_id: str) -> bool:
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM task_history WHERE task_id=? AND action='retry'", (task_id,))
        retries = cur.fetchone()[0]
        conn.close()
        return retries < 1
    except:
        return False

def mark_retry(task_id: str) -> None:
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("INSERT INTO task_history (task_id, action, created_at) VALUES (?,?,?)", (task_id, 'retry', datetime.now(timezone.utc).isoformat()))
        conn.commit(); conn.close()
    except: pass

def get_next_version(file_name: str, task_id: str) -> str:
    base, ext = os.path.splitext(file_name)
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM tasks WHERE result LIKE ?", (f"%{base}%",))
        count = cur.fetchone()[0]
        conn.close()
        return f"{base}_v{count+1}{ext}"
    except:
        return f"{base}_v2{ext}"
import fcntl

def acquire_task_lock(task_id: str) -> bool:
    lock_file = f"/tmp/task_{task_id}.lock"
    try:
        fd = open(lock_file, 'w')
        fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
        return True
    except:
        return False
import re

def sanitize_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', '_', name)[:100]
def check_file_size(file_path: str, max_mb: int = 50) -> bool:
    return os.path.getsize(file_path) <= max_mb * 1024 * 1024
def can_open_file(file_path: str) -> bool:
    try:
        if file_path.endswith(('.xlsx','.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(file_path); wb.close()
        elif file_path.endswith('.docx'):
            from docx import Document
            Document(file_path)
        elif file_path.endswith('.pdf'):
            from pypdf import PdfReader
            PdfReader(file_path)
        return True
    except:
        return False
