# === OCR_TABLE_TO_EXCEL_FULL_CLOSE_V1 ===
from __future__ import annotations

import json
import os
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List

def _safe(v: Any) -> str:
    return re.sub(r"[^A-Za-zА-Яа-я0-9_.-]+", "_", str(v or "ocr_table")).strip("._") or "ocr_table"

def _parse_rows(text: str) -> List[List[str]]:
    rows = []
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            data = data.get("rows") or data.get("items") or []
        if isinstance(data, list):
            for item in data:
                if isinstance(item, dict):
                    rows.append([
                        str(item.get("name") or item.get("наименование") or ""),
                        str(item.get("unit") or item.get("ед") or ""),
                        str(item.get("qty") or item.get("количество") or ""),
                        str(item.get("price") or item.get("цена") or ""),
                    ])
                elif isinstance(item, list):
                    rows.append([str(x) for x in item])
        if rows:
            return rows
    except Exception:
        pass

    for line in (text or "").splitlines():
        parts = [p.strip() for p in re.split(r"\s{2,}|\t|;", line) if p.strip()]
        if len(parts) >= 2:
            rows.append(parts[:6])
    return rows

def _write_xlsx(rows: List[List[str]], task_id: str) -> str:
    out = Path(tempfile.gettempdir()) / f"ocr_table_{_safe(task_id)}.xlsx"
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR_TABLE"
    headers = ["Наименование", "Ед", "Кол-во", "Цена", "Сумма"]
    ws.append(headers)
    for r in rows:
        name = r[0] if len(r) > 0 else ""
        unit = r[1] if len(r) > 1 else ""
        qty = r[2] if len(r) > 2 else ""
        price = r[3] if len(r) > 3 else ""
        ws.append([name, unit, qty, price, None])
        row = ws.max_row
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=4, value="ИТОГО")
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
    wb.save(out)
    wb.close()
    return str(out)

def _write_pdf_stub(rows: List[List[str]], task_id: str) -> str:
    out = Path(tempfile.gettempdir()) / f"ocr_table_{_safe(task_id)}.pdf"
    text = "OCR TABLE RESULT\\nRows: " + str(len(rows))
    stream = f"BT /F1 12 Tf 50 780 Td ({text}) Tj ET".encode()
    pdf = b"%PDF-1.4\n1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>endobj\n4 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n5 0 obj<< /Length " + str(len(stream)).encode() + b" >>stream\n" + stream + b"\nendstream endobj\ntrailer<< /Root 1 0 R >>\n%%EOF"
    out.write_bytes(pdf)
    return str(out)

def _zip(paths: List[str], task_id: str) -> str:
    out = Path(tempfile.gettempdir()) / f"ocr_table_package_{_safe(task_id)}.zip"
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in paths:
            if p and os.path.exists(p):
                z.write(p, arcname=os.path.basename(p))
    return str(out)

async def image_table_to_excel(local_path: str, task_id: str, user_text: str = "", topic_id: int = 0) -> Dict[str, Any]:
    if not local_path or not os.path.exists(local_path):
        return {"success": False, "error": "IMAGE_NOT_FOUND"}

    vision_text = ""
    try:
        from core.gemini_vision import analyze_image_file
        prompt = (
            "Распознай таблицу/смету/ВОР на изображении. "
            "Верни строго JSON: {\"rows\":[{\"name\":\"\",\"unit\":\"\",\"qty\":\"\",\"price\":\"\"}]}. "
            "Не считай руками, только извлеки строки."
        )
        vision_text = await analyze_image_file(local_path, prompt=prompt, timeout=90) or ""
    except Exception as e:
        return {"success": False, "error": f"VISION_UNAVAILABLE:{e}"}

    rows = _parse_rows(vision_text)
    if not rows:
        return {"success": False, "error": "NO_TABLE_ROWS_RECOGNIZED", "raw": vision_text[:2000]}

    xlsx = _write_xlsx(rows, task_id)
    pdf = _write_pdf_stub(rows, task_id)
    package = _zip([xlsx, pdf], task_id)

    return {
        "success": True,
        "engine": "OCR_TABLE_TO_EXCEL_FULL_CLOSE_V1",
        "summary": f"Фото таблицы распознано\\nСтрок: {len(rows)}\\nАртефакты: XLSX + PDF",
        "artifact_path": package,
        "artifact_name": f"ocr_table_package_{_safe(task_id)}.zip",
        "extra_artifacts": [xlsx, pdf],
        "rows": rows,
    }
# === END_OCR_TABLE_TO_EXCEL_FULL_CLOSE_V1 ===
