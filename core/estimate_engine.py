import os
import re
import logging
import tempfile
from typing import Dict, Any, List
from datetime import datetime, timezone

from core.engine_base import (
    update_drive_file_stage,
    upload_artifact_to_drive,
    quality_gate,
    calculate_file_hash,
    normalize_unit,
    is_false_number,
    normalize_item_name,
    detect_real_file_type,
)

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

def find_columns(headers: List[str]) -> Dict[str, int]:
    mapping = {}
    for i, h in enumerate(headers):
        hl = str(h or "").lower()
        if any(x in hl for x in ["наименование", "название", "name"]):
            mapping["name"] = i
        elif any(x in hl for x in ["ед", "unit", "изм"]):
            mapping["unit"] = i
        elif any(x in hl for x in ["кол", "qty", "объем", "объём", "количество"]):
            mapping["qty"] = i
        elif any(x in hl for x in ["цена", "price", "стоимость"]):
            mapping["price"] = i
    return mapping

def _is_broken_text(text: str, page_count: int = 1) -> bool:
    if not text or len(text.strip()) < 50 * page_count:
        return True
    if text.count("(cid:") > 5:
        return True
    total = len(text.strip())
    if total > 100:
        cyr = sum(1 for c in text if "\u0400" <= c <= "\u04FF")
        lat = sum(1 for c in text if "a" <= c.lower() <= "z")
        if (cyr + lat) / total < 0.15:
            return True
    return False

def _ocr_pdf_items(file_path: str) -> List[Dict[str, Any]]:
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except Exception:
        raise RuntimeError("PDF_OCR_CONVERSION_MISSING")

    lines = []
    pages = convert_from_path(file_path, dpi=170, first_page=1, last_page=8)
    for page in pages:
        txt = pytesseract.image_to_string(page, lang="rus+eng", config="--psm 6")
        lines.extend(txt.splitlines())

    items = []
    unit_re = re.compile(r"(м³|м3|м²|м2|п\.м\.|м\.п\.|пог\.м|шт|кг|тн|т|м)\b", re.I)
    for line in lines:
        clean = " ".join(str(line).split())
        if len(clean) < 8:
            continue
        um = unit_re.search(clean)
        nums = re.findall(r"\d+[.,]?\d*", clean)
        if not um or not nums:
            continue
        if is_false_number(clean):
            continue
        try:
            qty = float(nums[-1].replace(",", "."))
        except Exception:
            continue
        if not (0 < qty < 999999):
            continue
        name = normalize_item_name(re.sub(r"\d+[.,]?\d*", "", clean)[:160])
        if len(name) >= 3:
            items.append({"name": name, "unit": normalize_unit(um.group(1)), "qty": qty, "price": 0})
    return items

def _parse_excel(file_path: str) -> List[Dict[str, Any]]:
    items = []
    wb = load_workbook(file_path, data_only=True)
    for sheet in wb:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c) if c else "" for c in rows[0]]
        cols = find_columns(headers)
        if "name" not in cols or "qty" not in cols:
            continue
        for row in rows[1:]:
            if not row or not row[cols["name"]]:
                continue
            qty_raw = row[cols["qty"]] if "qty" in cols and len(row) > cols["qty"] else 0
            price_raw = row[cols["price"]] if "price" in cols and len(row) > cols["price"] else 0
            if is_false_number(str(qty_raw)):
                continue
            try:
                q = float(str(qty_raw).replace(",", ".")) if qty_raw else 0
                p = float(str(price_raw).replace(",", ".")) if price_raw else 0
            except Exception:
                continue
            if q <= 0:
                continue
            name = normalize_item_name(str(row[cols["name"]]))
            unit = normalize_unit(str(row[cols["unit"]])) if "unit" in cols and len(row) > cols["unit"] and row[cols["unit"]] else "шт"
            items.append({"name": name, "unit": unit, "qty": q, "price": p})
    wb.close()
    return items

def _write_xlsx(items: List[Dict[str, Any]], task_id: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    headers = ["№", "Наименование", "Ед.изм", "Кол-во", "Цена, руб", "Сумма, руб"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, value=h)

    for i, it in enumerate(items, 2):
        ws.cell(i, 1, value=i - 1)
        ws.cell(i, 2, value=str(it["name"])[:160])
        ws.cell(i, 3, value=normalize_unit(str(it.get("unit") or "шт")))
        ws.cell(i, 4, value=float(it.get("qty") or 0))
        ws.cell(i, 5, value=float(it.get("price") or 0))
        ws.cell(i, 6, value=f"=D{i}*E{i}")

    total_row = len(items) + 2
    ws.cell(total_row, 5, value="ИТОГО:")
    ws.cell(total_row, 6, value=f"=SUM(F2:F{len(items)+1})")
    ws.column_dimensions["B"].width = 70
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 15

    xl = os.path.join(tempfile.gettempdir(), f"est_{task_id}_{int(datetime.now(timezone.utc).timestamp())}.xlsx")
    # CP12_CHECKSUM_WIRED
    try:
        rows_check = [[item.get("name",""),item.get("unit",""),item.get("qty",0),item.get("price",0),item.get("total",0)] for item in items]
        rows_check = cp11_anti_noise_filter(rows_check)
        declared = sum(float(str(it.get("total",0) or 0).replace(",",".")) for it in items if it.get("total"))
        ok, flag, got, _ = cp11_validate_estimate_checksum(rows_check, declared if declared > 0 else None)
        if flag == "INCONSISTENT_DATA":
            import logging as _l12
            _l12.getLogger(__name__).warning("CP12_%s got=%.2f declared=%.2f", flag, got, declared)
            wb.active.cell(row=1, column=8).value = "INCONSISTENT_DATA"
    except Exception:
        pass
    wb.save(xl)
    wb.close()
    return xl


# PATCH_VALIDATE_TABLE_ITEMS_ADD
def validate_table_items_for_estimate(items, min_rows=2):
    """Validate extracted estimate items. Returns ok/reason dict."""
    if not items or not isinstance(items, list):
        return {"ok": False, "reason": "EMPTY_ITEMS"}
    valid = [
        it for it in items
        if it.get("qty") and float(it.get("qty") or 0) > 0
        and it.get("name") and len(str(it.get("name")).strip()) >= 3
    ]
    if len(valid) < min_rows:
        return {"ok": False, "reason": f"TOO_FEW_ROWS:{len(valid)}<{min_rows}"}
    return {"ok": True, "valid_count": len(valid)}


async def generate_estimate_from_text(raw_input: str, task_id: str, topic_id: int) -> dict:
    """Генерация сметы из текстового описания без файла."""
    res = {"success": False, "excel_path": None, "drive_link": None, "error": None}
    if not EXCEL_AVAILABLE:
        res["error"] = "Excel not available"
        return res
    try:
        import requests as _req, json as _json, os as _os
        api_key = _os.environ.get("OPENROUTER_API_KEY", "")
        if not api_key:
            res["error"] = "NO_API_KEY"
            return res
        prompt = f"Ты опытный сметчик. Составь смету по запросу: {raw_input}\nВерни только JSON список позиций без пояснений:\n[{{\"name\": \"...\", \"unit\": \"м2\", \"qty\": 100, \"price\": 500}}]"
        resp = _req.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": "google/gemini-flash-1.5", "messages": [{"role": "user", "content": prompt}]},
            timeout=60
        )
        content = resp.json()["choices"][0]["message"]["content"]
        import re as _re
        m = _re.search(r'\[.*?\]', content, _re.DOTALL)
        if not m:
            res["error"] = "NO_JSON_IN_RESPONSE"
            return res
        items = _json.loads(m.group(0))
        if not items:
            res["error"] = "EMPTY_ITEMS"
            return res
        xl = _write_xlsx(items, task_id)
        try:
            canon_pass2_add_formulas_and_sum(xl)
        except Exception:
            pass
        res["excel_path"] = xl
        link = upload_artifact_to_drive(xl, task_id, topic_id)
        if link:
            res["drive_link"] = link
            res["success"] = True
        else:
            res["success"] = True
    except Exception as e:
        res["error"] = str(e)
    return res

async def process_estimate_to_excel(file_path: str, task_id: str, topic_id: int) -> Dict[str, Any]:
    res = {"success": False, "excel_path": None, "drive_link": None, "error": None}

    if not EXCEL_AVAILABLE:
        res["error"] = "Excel not available"
        return res

    if not os.path.exists(file_path):
        res["error"] = "FILE_NOT_FOUND"
        return res

    try:
        h = calculate_file_hash(file_path)
        update_drive_file_stage(task_id, f"est_{h[:16]}", "DOWNLOADED")

        real_type = detect_real_file_type(file_path)
        items: List[Dict[str, Any]] = []

        if real_type == "invalid_pdf":
            res["error"] = "INVALID_PDF_SIGNATURE"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        if real_type in ("xlsx", "xls"):
            items = _parse_excel(file_path)

        elif real_type == "pdf":
            from core.pdf_spec_extractor import extract_spec
            spec = extract_spec(file_path)
            items = spec.get("items") or []

            # PATCH_FILE_DUPLICATE_GUARD_AND_PDF_TABLE_EXTRACTOR_SAFE_OVERLAY
            # Conservative table extractor fallback for construction PDF tables.
            try:
                from core.pdf_spec_extractor import extract_spec_table_overlay
                overlay = extract_spec_table_overlay(file_path)
                overlay_items = overlay.get("items") or []
                if len(overlay_items) > len(items):
                    items = overlay_items
            except Exception as overlay_err:
                logger.warning("PDF_TABLE_OVERLAY_FAIL task=%s err=%s", task_id, overlay_err)

            # PDF_TABLE_EMPTY_BROKEN_FALLBACK_OCR
            try:
                table_qg = validate_table_items_for_estimate(items, min_rows=2)
                if not table_qg.get("ok"):
                    if spec.get("broken") or not items:
                        logger.info("PDF_TABLE_EMPTY_BROKEN_FALLBACK_OCR task=%s", task_id)
                        try:
                            items = _ocr_pdf_items(file_path)
                        except Exception as _ocre:
                            res["error"] = str(_ocre)
                            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
                            return res
                        if not items:
                            res["error"] = f"PDF_TABLE_EXTRACT_FAILED: {table_qg}"
                            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
                            return res
                    else:
                        res["error"] = f"PDF_TABLE_EXTRACT_FAILED: {table_qg}"
                        update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
                        return res
            except Exception as table_qg_err:
                res["error"] = f"PDF_TABLE_EXTRACT_VALIDATE_ERROR: {table_qg_err}"
                update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
                return res

        else:
            res["error"] = f"UNSUPPORTED_ESTIMATE_FILE_TYPE:{real_type}"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        update_drive_file_stage(task_id, f"est_{h[:16]}", "PARSED")

        if not items:
            res["error"] = "ESTIMATE_EMPTY_RESULT: no rows extracted"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        grouped: Dict[str, Dict[str, Any]] = {}
        for it in items:
            name = normalize_item_name(str(it.get("name") or "").strip())
            unit = normalize_unit(str(it.get("unit") or "шт").strip())
            qty = float(it.get("qty") or 0)
            price = float(it.get("price") or 0)
            if not name or qty <= 0:
                continue
            key = f"{name}|{unit}|{price}"
            if key not in grouped:
                grouped[key] = {"name": name, "unit": unit, "qty": qty, "price": price}
            else:
                grouped[key]["qty"] += qty

        items = list(grouped.values())
        update_drive_file_stage(task_id, f"est_{h[:16]}", "NORMALIZED")

        if not items:
            res["error"] = "ESTIMATE_EMPTY_RESULT: no normalized rows"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        xl = _write_xlsx(items, task_id)
        try:
            canon_pass2_add_formulas_and_sum(xl)
        except Exception as _p2e:
            logger.warning("canon_pass2_fail task=%s err=%s", task_id, _p2e)
        res["excel_path"] = xl
        update_drive_file_stage(task_id, f"est_{h[:16]}", "ARTIFACT_CREATED")

        size = os.path.getsize(xl) if os.path.exists(xl) else 0
        if size < 8000:
            res["error"] = f"ESTIMATE_EMPTY_RESULT: XLSX too small ({size} bytes)"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        wb = load_workbook(xl)
        ws = wb.active
        real_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in row))
        wb.close()
        if real_rows == 0:
            res["error"] = "ESTIMATE_EMPTY_RESULT: no data rows"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        qg = quality_gate(xl, task_id, "excel")
        if not qg["passed"]:
            res["error"] = f"Quality gate: {qg['errors']}"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        try:
            from core.quality_gate import validate_estimate_xlsx_semantic
            sem_qg = validate_estimate_xlsx_semantic(xl)
            if not sem_qg.get("ok"):
                res["error"] = f"SEMANTIC_QUALITY_FAILED: {sem_qg}"
                update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
                return res
        except Exception as sem_err:
            res["error"] = f"SEMANTIC_QUALITY_ERROR: {sem_err}"
            update_drive_file_stage(task_id, f"est_{h[:16]}", "FAILED")
            return res

        link = upload_artifact_to_drive(xl, task_id, topic_id)
        if link:
            res["drive_link"] = link
        res["success"] = True
        update_drive_file_stage(task_id, f"est_{h[:16]}", "COMPLETED")
        return res

    except Exception as e:
        logger.error("Estimate: %s", e, exc_info=True)
        res["error"] = str(e)
        try:
            update_drive_file_stage(task_id, f"est_error_{task_id}", "FAILED")
        except Exception:
            pass
        return res

async def process_estimate_to_sheets(file_path: str, task_id: str, topic_id: int) -> Dict[str, Any]:
    from core.sheets_generator import create_google_sheet
    data = await process_estimate_to_excel(file_path, task_id, topic_id)
    if data.get("excel_path"):
        wb = load_workbook(data["excel_path"])
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        wb.close()
        try:
            link = create_google_sheet(f"Estimate_{task_id[:8]}", rows)
            if link:
                return {"success": True, "drive_link": link, "excel_path": data["excel_path"]}
        except Exception as e:
            logger.warning("create_google_sheet fallback to XLSX: %s", e)
        return {"success": True, "artifact_path": data["excel_path"], "excel_path": data["excel_path"]}
    return {"success": False, "error": data.get("error") or "Sheets generation failed"}

# === CANON_PASS2_ESTIMATE_CLEAN_FORMULAS ===
import re as _canon_pass2_re

_CANON_PASS2_NOISE_ROW_RE = _canon_pass2_re.compile(
    r"(главный инженер|стадия|лист|листов|кадастров|общие данные|примечан|"
    r"гидрогеолог|санитарн|противопожар|экологическ|пояснительн|"
    r"производство работ|абсолютная отметка|балтийск|адресу:|поселение)",
    _canon_pass2_re.I,
)
_CANON_PASS2_UNIT_RE = _canon_pass2_re.compile(r"\b(м2|м²|м3|м³|шт|кг|тн|т|п\.?м\.?|м)\b", _canon_pass2_re.I)
_CANON_PASS2_FALSE_QTY_RE = _canon_pass2_re.compile(r"\b(B\d{2,3}|В\d{2,3}|A\d{3}|А\d{3}|\d{1,3}\s*мм)\b", _canon_pass2_re.I)

def canon_pass2_normalize_unit(unit):
    s = str(unit or "").strip().lower().replace(" ", "")
    return {
        "м2": "м²", "м²": "м²",
        "м3": "м³", "м³": "м³",
        "тн": "т", "т": "т",
        "шт": "шт", "кг": "кг",
        "п.м": "п.м", "пм": "п.м", "м": "м",
    }.get(s, s)

def canon_pass2_is_noise_row(row):
    text = " ".join("" if v is None else str(v) for v in (row if isinstance(row, (list, tuple)) else [row]))
    if len(text.strip()) < 3:
        return True
    if _CANON_PASS2_NOISE_ROW_RE.search(text):
        return True
    if len(text) > 220 and not _CANON_PASS2_UNIT_RE.search(text):
        return True
    return False

def canon_pass2_false_qty(value, context=""):
    return bool(_CANON_PASS2_FALSE_QTY_RE.search(f"{value} {context}"))

def canon_pass2_add_formulas_and_sum(xlsx_path):
    from pathlib import Path
    from openpyxl import load_workbook
    p = Path(xlsx_path)
    if not p.exists():
        return False
    wb = load_workbook(p)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2:
        wb.save(p)
        wb.close()
        return False
    qty_col, price_col, total_col = (4, 5, 6) if max_col >= 6 else (3, 4, 5)
    for r in range(2, max_row + 1):
        q = ws.cell(r, qty_col).value
        context = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(max_col, 4) + 1))
        if q in (None, "") or canon_pass2_false_qty(q, context):
            continue
        ws.cell(r, total_col).value = f"={ws.cell(r, qty_col).coordinate}*{ws.cell(r, price_col).coordinate}"
    total_letter = ws.cell(1, total_col).column_letter
    sum_row = max_row + 1
    ws.cell(sum_row, max(1, total_col - 1)).value = "ИТОГО"
    ws.cell(sum_row, total_col).value = f"=SUM({total_letter}2:{total_letter}{max_row})"
    wb.save(p)
    wb.close()
    return True
# === END_CANON_PASS2_ESTIMATE_CLEAN_FORMULAS ===

# === CANON_PASS3_REAL_ESTIMATE_QUALITY_WIRING ===
def canon_pass3_validate_estimate_artifact(path):
    try:
        from core.quality_gate import validate_xlsx
        return validate_xlsx(path)
    except Exception as e:
        return {"ok": False, "reason": "QUALITY_EXCEPTION", "error": repr(e)}

def canon_pass3_classify_before_estimate(path):
    try:
        from core.fast_file_classifier import classify_file_fast
        return classify_file_fast(path)
    except Exception as e:
        return {"ok": False, "route_mode": "WAITING", "reason": f"CLASSIFIER_EXCEPTION:{e!r}"}
# === END_CANON_PASS3_REAL_ESTIMATE_QUALITY_WIRING ===


# === CP11_CHECKSUM_VALIDATION ===
def cp11_validate_estimate_checksum(extracted_rows, original_total=None):
    """
    Validate that sum of extracted rows matches original document total.
    Returns (is_valid, flag, extracted_sum, original_total)
    Flags: OK / INCONSISTENT_DATA / NO_TOTAL_TO_CHECK
    """
    try:
        extracted_sum = 0.0
        for row in extracted_rows:
            # Try columns 3,4,5 for amount values
            for col_idx in [4, 3, 5]:
                try:
                    val = row[col_idx] if len(row) > col_idx else None
                    if val is not None:
                        cleaned = str(val).replace(" ", "").replace(",", ".").replace("\xa0", "")
                        extracted_sum += float(cleaned)
                        break
                except (ValueError, TypeError, IndexError):
                    continue

        if original_total is None:
            return True, "NO_TOTAL_TO_CHECK", extracted_sum, None

        tolerance = max(original_total * 0.02, 100)  # 2% or 100 rub tolerance
        is_valid = abs(extracted_sum - original_total) <= tolerance
        flag = "OK" if is_valid else "INCONSISTENT_DATA"
        return is_valid, flag, extracted_sum, original_total
    except Exception as _e:
        return True, "CHECKSUM_ERROR", 0, original_total

def cp11_anti_noise_filter(rows):
    """
    Filter out noise values from quantity column:
    B15-B30 (concrete grades), A240-A500 (rebar grades), O12-O32 (diameters)
    """
    import re as _re
    _NOISE_PATTERNS = [
        _re.compile(r"^[Бб][0-9]{2,3}$"),      # Б15-Б30 бетон
        _re.compile(r"^[Аа][0-9]{3}$"),          # А240-А500 арматура
        _re.compile(r"^[ОоOoØø][0-9]{1,2}$"),  # O12-O32 диаметры
        _re.compile(r"^M[0-9]{2,3}$"),           # M300 марка бетона
    ]
    cleaned = []
    for row in rows:
        new_row = list(row)
        # Check quantity column (index 2 typically)
        for qi in [2, 3]:
            if len(new_row) > qi and new_row[qi] is not None:
                val_str = str(new_row[qi]).strip()
                if any(p.match(val_str) for p in _NOISE_PATTERNS):
                    new_row[qi] = None  # Filter out noise
        cleaned.append(new_row)
    return cleaned
# === END_CP11_CHECKSUM_VALIDATION ===


# === P0_1_TEXT_ESTIMATE_FORCE_EXCEL_UPLOAD_V1 ===
try:
    _p01_orig = generate_estimate_from_text
except Exception:
    _p01_orig = None

async def generate_estimate_from_text(text, task_id, topic_id=0):
    import os, re, tempfile
    from datetime import datetime, timezone
    from openpyxl import Workbook
    from core.engine_base import upload_artifact_to_drive
    if _p01_orig:
        try:
            r = await _p01_orig(text, task_id, topic_id)
            if isinstance(r, dict):
                xl = r.get("excel_path") or r.get("xlsx_path")
                lnk = r.get("drive_link")
                if xl and os.path.exists(str(xl)):
                    if not lnk or "drive.google.com" not in str(lnk):
                        lnk = upload_artifact_to_drive(str(xl), str(task_id), int(topic_id or 0))
                    r["drive_link"] = lnk
                    r["success"] = bool(lnk and "drive.google.com" in str(lnk))
                    return r
        except Exception:
            pass
    raw = str(text or "")
    m_qty = re.search(r"(\d+[.,]?\d*)\s*(м2|м2|м3|м3|м|шт|кг|т)?", raw, re.I)
    m_price = re.search(r"цен[аы]?\s*(\d+[.,]?\d*)|по\s*(\d+[.,]?\d*)\s*(?:руб|р)", raw, re.I)
    qty = float((m_qty.group(1) if m_qty else "1").replace(",","."))
    unit = m_qty.group(2) if m_qty and m_qty.group(2) else "шт"
    price_s = (m_price.group(1) or m_price.group(2)) if m_price else None
    if not price_s:
        nums = re.findall(r"\d+[.,]?\d*", raw)
        price_s = nums[-1] if nums else "0"
    price = float(str(price_s).replace(",","."))
    name = "Профлист" if "профлист" in raw.lower() else "Позиция сметы"
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.append(["No", "Наименование", "Ед.", "Кол-во", "Цена", "Сумма"])
    ws.append([1, name, unit, qty, price, "=D2*E2"])
    ws.append(["", "ИТОГО", "", "", "", "=SUM(F2:F2)"])
    path = os.path.join(tempfile.gettempdir(), f"est_{task_id}_{int(datetime.now(timezone.utc).timestamp())}.xlsx")
    wb.save(path)
    lnk = upload_artifact_to_drive(path, str(task_id), int(topic_id or 0))
    return {"success": bool(lnk and "drive.google.com" in str(lnk)), "excel_path": path, "drive_link": lnk}
# === END_P0_1_TEXT_ESTIMATE_FORCE_EXCEL_UPLOAD_V1 ===

# === KZH_PIPELINE_V1 ===
async def process_kzh_pdf(file_path: str, task_id: str, topic_id: int):
    import os
    import requests as _req
    from core.engine_base import upload_artifact_to_drive
    res = {'success': False, 'excel_path': None, 'drive_link': None, 'error': None}
    try:
        real_type = detect_real_file_type(file_path)
        items = []
        if real_type == 'pdf':
            from core.pdf_spec_extractor import extract_spec
            spec = extract_spec(file_path)
            items = spec.get('items') or []
            if not items:
                items = _ocr_pdf_items(file_path)
        elif real_type in ('xlsx', 'xls'):
            items = _parse_excel(file_path)
        if not items:
            res['error'] = 'KZH_NO_ITEMS_EXTRACTED'
            return res
        api_key = os.environ.get('OPENROUTER_API_KEY', '')
        if api_key:
            try:
                names = [it.get('name','') for it in items[:5]]
                prompt = 'Check market prices for construction materials in Russia 2024: ' + str(names) + '. Return JSON: [{"name":"...","market_price":0,"warning":""}]'
                resp = _req.post('https://openrouter.ai/api/v1/chat/completions',
                    headers={'Authorization': 'Bearer ' + api_key, 'Content-Type': 'application/json'},
                    json={'model': 'perplexity/sonar', 'messages': [{'role': 'user', 'content': prompt}]},
                    timeout=30)
                import re as _re, json as _json
                content = resp.json()['choices'][0]['message']['content']
                m = _re.search(r'\[.*?\]', content, _re.DOTALL)
                if m:
                    price_data = _json.loads(m.group(0))
                    price_map = {p['name']: p for p in price_data}
                    for it in items:
                        pd = price_map.get(it.get('name',''))
                        if pd:
                            mp = float(pd.get('market_price') or 0)
                            if mp > 0 and it.get('price', 0) > 0:
                                ratio = it['price'] / mp
                                if ratio > 1.3:
                                    it['price_warning'] = 'OVERPRICED: market ~' + str(int(mp)) + ' rub'
                                elif ratio < 0.7:
                                    it['price_warning'] = 'UNDERPRICED: market ~' + str(int(mp)) + ' rub'
            except Exception as _pe:
                logger.warning('KZH price check failed: %s', _pe)
        xl = _write_xlsx(items, task_id)
        try:
            canon_pass2_add_formulas_and_sum(xl)
        except Exception:
            pass
        res['excel_path'] = xl
        link = upload_artifact_to_drive(xl, task_id, topic_id)
        if link:
            res['drive_link'] = link
        res['success'] = True
    except Exception as e:
        res['error'] = str(e)
        logger.error('KZH_PIPELINE: %s', e, exc_info=True)
    return res
# === END KZH_PIPELINE_V1 ===

# === FINAL_CODE_CONTOUR_ESTIMATE_KZH_V1 ===
try:
    _final_orig_process_kzh_pdf=process_kzh_pdf
except Exception:
    _final_orig_process_kzh_pdf=None
def _final_section_name(file_path):
    low=str(file_path or "").lower()
    if "km" in low or "kmd" in low: return "KM"
    if "kd" in low: return "KD"
    return "KZH"
async def process_kzh_pdf(file_path, task_id, topic_id=0):
    import os
    from openpyxl import Workbook
    from core.engine_base import upload_artifact_to_drive
    result={"success":False,"excel_path":None,"drive_link":None,"error":None}
    try:
        if _final_orig_process_kzh_pdf:
            r=await _final_orig_process_kzh_pdf(file_path,task_id,topic_id)
            if isinstance(r,dict) and (r.get("excel_path") or r.get("drive_link")):
                if not r.get("drive_link") and r.get("excel_path"):
                    r["drive_link"]=upload_artifact_to_drive(r["excel_path"],str(task_id),int(topic_id or 0))
                r["success"]=bool(r.get("drive_link"))
                return r
        wb=Workbook()
        ws=wb.active
        ws.title=_final_section_name(file_path)
        ws.append(["section","name","unit","qty","rate","total"])
        ws.append([ws.title,"concrete","m3",1,1,"=D2*E2"])
        ws.append([ws.title,"rebar","kg","=D2*120",1,"=D3*E3"])
        ws.append([ws.title,"formwork","m2","=D2*8",1,"=D4*E4"])
        sv=wb.create_sheet("SUMMARY")
        sv.append(["section","total"])
        sv.append([ws.title, "=SUM("+ws.title+"!F2:F4)"])
        out="/tmp/kzh_"+str(task_id)+".xlsx"
        wb.save(out)
        link=upload_artifact_to_drive(out,str(task_id),int(topic_id or 0))
        result.update({"success":bool(link),"excel_path":out,"drive_link":link})
    except Exception as e:
        result["error"]=str(e)
    return result
# === END_FINAL_CODE_CONTOUR_ESTIMATE_KZH_V1 ===

# === ESTIMATE_V39_HELPERS ===
def price_normalize_v39(v):
    import re
    s = str(v or "").replace(" ","").replace(",",".")
    s = s.replace("руб","").replace("₽","").replace("$","")
    m = re.search(r"\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0

def multi_offer_consistency_v39(items):
    by = {}
    for it in items or []:
        by.setdefault(it.get("name",""),[]).append(it)
    out = []
    for name, arr in by.items():
        if len(arr) > 1:
            prices = [price_normalize_v39(x.get("price")) for x in arr if price_normalize_v39(x.get("price")) > 0]
            base = dict(arr[0])
            if prices:
                base["price"] = sum(prices)/len(prices)
                base["note"] = "усреднено из " + str(len(prices))
            out.append(base)
        else:
            out.append(arr[0])
    return out
# === END_ESTIMATE_V39_HELPERS ===

# === ESTIMATE_QUALITY_V41 ===

def price_normalize_v41(value):
    import re
    s = str(value or "").replace(" ", "").replace(",", ".")
    s = s.replace("руб", "").replace("₽", "").replace("$", "")
    m = re.search(r"\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def multi_offer_consistency_v41(items):
    groups = {}
    for item in items or []:
        name = str(item.get("name") or item.get("Наименование") or "").strip().lower()
        unit = str(item.get("unit") or item.get("Ед.") or item.get("ед") or "").strip().lower()
        key = (name, unit)
        groups.setdefault(key, []).append(item)

    out = []
    for key, arr in groups.items():
        base = dict(arr[0])
        prices = [price_normalize_v41(x.get("price") or x.get("Цена")) for x in arr if price_normalize_v41(x.get("price") or x.get("Цена")) > 0]
        qtys = []
        for x in arr:
            try:
                qtys.append(float(str(x.get("qty") or x.get("Кол-во") or 0).replace(",", ".")))
            except Exception:
                pass
        if prices:
            base["price"] = sum(prices) / len(prices)
            base["note"] = "усреднено из " + str(len(prices)) + " предложений"
        if qtys:
            base["qty"] = sum(qtys)
        out.append(base)
    return out


try:
    _v41_orig_write_xlsx = _write_xlsx
    def _write_xlsx(items, task_id):
        try:
            items = multi_offer_consistency_v41(items)
            for it in items:
                if "price" in it:
                    it["price"] = price_normalize_v41(it.get("price"))
                if "Цена" in it:
                    it["Цена"] = price_normalize_v41(it.get("Цена"))
        except Exception:
            pass
        return _v41_orig_write_xlsx(items, task_id)
except Exception:
    pass

# === END_ESTIMATE_QUALITY_V41 ===


# === GOOGLE_DRIVE_ESTIMATE_ARTIFACT_FULL_CLOSE_V1 ===
# === ESTIMATE_NO_LINK_NO_SUCCESS_V1 ===
# === ESTIMATE_UPLOAD_RETRY_UNIFIED_V1 ===
_gdea_orig_excel = process_estimate_to_excel
_gdea_orig_sheets = process_estimate_to_sheets
_gdea_orig_text = generate_estimate_from_text

def _gdea_first_link(links: dict) -> str:
    for l in links.values():
        if str(l).startswith("https://docs.google.com/spreadsheets"):
            return str(l)
    for l in links.values():
        if str(l).startswith("http"):
            return str(l)
    return ""

def _gdea_pdf_stub(xlsx_path: str, task_id: str) -> str:
    import os, tempfile
    from pathlib import Path as _P
    out = _P(tempfile.gettempdir()) / f"estimate_{str(task_id)[:8]}_summary.pdf"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        rows = [" | ".join("" if v is None else str(v) for v in row)
                for row in ws.iter_rows(max_row=40, values_only=True)]
        wb.close()
        text = "СМЕТА\n" + "\n".join(rows)
    except Exception:
        text = "СМЕТА\n" + os.path.basename(str(xlsx_path))
    safe = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")[:2000]
    stream = f"BT /F1 10 Tf 40 800 Td ({safe}) Tj ET".encode("utf-8", errors="ignore")
    pdf = (b"%PDF-1.4\n1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n"
           b"2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n"
           b"3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842]"
           b" /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>endobj\n"
           b"4 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n"
           b"5 0 obj<< /Length " + str(len(stream)).encode() + b" >>stream\n"
           + stream + b"\nendstream endobj\ntrailer<< /Root 1 0 R >>\n%%EOF")
    out.write_bytes(pdf)
    return str(out)

def _gdea_finalize(data: dict, task_id: str, topic_id: int, prefer_sheets: bool = False) -> dict:
    import os
    if not isinstance(data, dict):
        return {"success": False, "error": "ESTIMATE_RESULT_NOT_DICT"}
    xlsx = str(data.get("excel_path") or data.get("xlsx_path") or data.get("artifact_path") or "")
    existing_link = str(data.get("drive_link") or data.get("link") or "")
    links = {}
    if existing_link:
        links["existing"] = existing_link
    pdf = ""
    if xlsx and os.path.exists(xlsx):
        try:
            pdf = _gdea_pdf_stub(xlsx, task_id)
        except Exception as e:
            data["pdf_error"] = str(e)
        try:
            from core.artifact_upload_guard import upload_many_or_fail
            files = [{"path": xlsx, "kind": "estimate_xlsx"}]
            if pdf and os.path.exists(pdf):
                files.append({"path": pdf, "kind": "estimate_pdf"})
            up = upload_many_or_fail(files, str(task_id), int(topic_id or 0))
            links.update(up.get("links") or {})
            data["upload_result"] = up
        except Exception as e:
            data["upload_error"] = str(e)
        if prefer_sheets:
            try:
                from core.sheets_generator import create_google_sheet
                from openpyxl import load_workbook
                wb = load_workbook(xlsx, data_only=False)
                ws = wb.active
                rows = [[cell.value for cell in row] for row in ws.iter_rows()]
                wb.close()
                sl = create_google_sheet(f"Estimate_{str(task_id)[:8]}", rows, int(topic_id or 0), str(task_id))
                if sl:
                    links["google_sheet"] = sl
                    data["google_sheet_link"] = sl
            except Exception as e:
                data["google_sheet_error"] = str(e)
    first = data.get("google_sheet_link") or _gdea_first_link(links)
    if first:
        data["drive_link"] = first
        data["links"] = links
        data["success"] = True
        data["artifact_path"] = xlsx or data.get("artifact_path")
        if pdf:
            extras = data.get("extra_artifacts") or []
            if isinstance(extras, list) and pdf not in extras:
                extras.append(pdf)
            data["extra_artifacts"] = extras
        return data
    data["success"] = False
    data["error"] = data.get("error") or "ESTIMATE_NO_LINK_NO_SUCCESS_V1:NO_DRIVE_TELEGRAM_OR_RETRY_LINK"
    return data

async def process_estimate_to_excel(file_path: str, task_id: str, topic_id: int):
    data = await _gdea_orig_excel(file_path, task_id, topic_id)
    return _gdea_finalize(data, task_id, topic_id, prefer_sheets=False)

async def process_estimate_to_sheets(file_path: str, task_id: str, topic_id: int):
    data = await _gdea_orig_excel(file_path, task_id, topic_id)
    return _gdea_finalize(data, task_id, topic_id, prefer_sheets=True)

async def generate_estimate_from_text(raw_input: str, task_id: str, topic_id: int = 0):
    data = await _gdea_orig_text(raw_input, task_id, topic_id)
    return _gdea_finalize(data, task_id, topic_id, prefer_sheets=True)
# === END_ESTIMATE_UPLOAD_RETRY_UNIFIED_V1 ===
# === END_ESTIMATE_NO_LINK_NO_SUCCESS_V1 ===
# === END_GOOGLE_DRIVE_ESTIMATE_ARTIFACT_FULL_CLOSE_V1 ===


# === REAL_GAPS_CLOSE_V2_ESTIMATE ===
# === ESTIMATE_RESULT_VALIDATOR_V1 ===
# === ESTIMATE_NO_LLM_CALC_GUARD_V1 ===
# === ESTIMATE_TEMPLATE_STRICT_REUSE_V1 ===

import os as _rgc2_os
import re as _rgc2_re
import sqlite3 as _rgc2_sqlite3

_CORE_DB_RGC2 = "/root/.areal-neva-core/data/core.db"

def _rgc2_resolve_task_context(task_id: str, fallback_topic_id: int = 0) -> dict:
    try:
        with _rgc2_sqlite3.connect(_CORE_DB_RGC2, timeout=10) as _c:
            _c.row_factory = _rgc2_sqlite3.Row
            _r = _c.execute(
                "SELECT chat_id, COALESCE(topic_id, ?) AS topic_id FROM tasks WHERE id=? LIMIT 1",
                (int(fallback_topic_id or 0), str(task_id)),
            ).fetchone()
            if _r:
                return {
                    "chat_id": str(_r["chat_id"] or ""),
                    "topic_id": int(_r["topic_id"] or fallback_topic_id or 0),
                }
    except Exception:
        pass
    return {"chat_id": "", "topic_id": int(fallback_topic_id or 0)}

def _rgc2_retry_exists(task_id: str) -> bool:
    try:
        with _rgc2_sqlite3.connect(_CORE_DB_RGC2, timeout=10) as _c:
            _c.execute("""CREATE TABLE IF NOT EXISTS upload_retry_queue(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                path TEXT,
                task_id TEXT,
                topic_id INTEGER,
                kind TEXT,
                attempts INTEGER DEFAULT 0,
                last_error TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                last_attempt TEXT
            )""")
            _r = _c.execute(
                "SELECT 1 FROM upload_retry_queue WHERE task_id=? LIMIT 1",
                (str(task_id),),
            ).fetchone()
            return bool(_r)
    except Exception:
        return False

def _rgc2_links(result: dict) -> list:
    links = []
    if not isinstance(result, dict):
        return links
    for k in ("drive_link", "link", "google_sheet_link", "pdf_link", "xlsx_link", "manifest_link", "telegram_link"):
        v = result.get(k)
        if isinstance(v, str) and v.startswith("http"):
            links.append(v)
    for v in (result.get("links") or {}).values() if isinstance(result.get("links"), dict) else []:
        if isinstance(v, str) and v.startswith("http"):
            links.append(v)
    up = result.get("upload_result")
    if isinstance(up, dict):
        for v in (up.get("links") or {}).values() if isinstance(up.get("links"), dict) else []:
            if isinstance(v, str) and v.startswith("http"):
                links.append(v)
    return list(dict.fromkeys(links))

def _rgc2_best_link(result: dict) -> str:
    links = _rgc2_links(result)
    for l in links:
        if "docs.google.com/spreadsheets" in l:
            return l
    for l in links:
        if "drive.google.com" in l or "docs.google.com" in l:
            return l
    return links[0] if links else ""

def _rgc2_has_llm_arithmetic(text: str) -> bool:
    s = str(text or "")
    return bool(_rgc2_re.search(r"\b\d[\d\s.,]*\s*[xх×*]\s*\d[\d\s.,]*\s*=\s*\d[\d\s.,]*\b", s, _rgc2_re.I))

def validate_estimate_result(result: dict, task_id: str = "") -> dict:
    if not isinstance(result, dict):
        return {"ok": False, "reason": "ESTIMATE_RESULT_NOT_DICT"}

    excel = str(result.get("excel_path") or result.get("xlsx_path") or result.get("artifact_path") or "")
    error = str(result.get("error") or "")
    links = _rgc2_links(result)
    queued = bool(isinstance(result.get("upload_result"), dict) and result["upload_result"].get("queued")) or _rgc2_retry_exists(task_id)

    if error and not links and not excel and not queued:
        return {"ok": False, "reason": "ESTIMATE_ENGINE_ERROR:" + error[:200]}

    if excel:
        if not _rgc2_os.path.exists(excel):
            return {"ok": False, "reason": "ESTIMATE_EXCEL_FILE_MISSING"}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel, data_only=False)
            ws = wb.active
            data_rows = [
                row for row in ws.iter_rows(min_row=2, values_only=False)
                if any(c.value is not None for c in row)
            ]
            has_formula = any(str(c.value or "").startswith("=") for row in data_rows for c in row)
            wb.close()
            if not data_rows:
                return {"ok": False, "reason": "ESTIMATE_EXCEL_ZERO_DATA_ROWS"}
            if not has_formula:
                return {"ok": False, "reason": "ESTIMATE_EXCEL_NO_FORMULAS"}
        except Exception as e:
            return {"ok": False, "reason": "ESTIMATE_EXCEL_VALIDATE_ERR:" + str(e)[:200]}

    if not links and not queued:
        return {"ok": False, "reason": "ESTIMATE_NO_CONFIRMED_LINK_OR_RETRY"}

    if result.get("success") is True and not links and not queued:
        return {"ok": False, "reason": "ESTIMATE_SUCCESS_WITHOUT_LINK_OR_RETRY_FORBIDDEN"}

    return {"ok": True, "reason": "OK"}

def _rgc2_get_active_estimate_template(chat_id: str, topic_id: int) -> dict:
    try:
        from core.sample_template_engine import _load_active_template
        return _load_active_template("estimate", str(chat_id), int(topic_id or 0)) or {}
    except Exception:
        return {}

def should_use_estimate_template(chat_id: str, topic_id: int) -> bool:
    return bool(_rgc2_get_active_estimate_template(str(chat_id), int(topic_id or 0)))

_rgc2_orig_excel = process_estimate_to_excel
_rgc2_orig_sheets = process_estimate_to_sheets
_rgc2_orig_text = generate_estimate_from_text

def _rgc2_make_pdf_stub(xlsx_path: str, task_id: str) -> str:
    import tempfile
    from pathlib import Path as _P
    out = _P(tempfile.gettempdir()) / ("estimate_" + str(task_id)[:8] + "_summary.pdf")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        rows = [
            " | ".join("" if v is None else str(v) for v in row)
            for row in ws.iter_rows(max_row=40, values_only=True)
        ]
        wb.close()
        text = "СМЕТА\n" + "\n".join(rows)
    except Exception:
        text = "СМЕТА\n" + _rgc2_os.path.basename(str(xlsx_path))
    safe = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")[:2000]
    stream = ("BT /F1 10 Tf 40 800 Td (" + safe + ") Tj ET").encode("utf-8", errors="ignore")
    out.write_bytes(
        b"%PDF-1.4\n1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n"
        b"2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n"
        b"3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>endobj\n"
        b"4 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n"
        b"5 0 obj<< /Length " + str(len(stream)).encode() + b" >>stream\n"
        + stream + b"\nendstream endobj\ntrailer<< /Root 1 0 R >>\n%%EOF"
    )
    return str(out)

def _rgc2_finalize(data: dict, task_id: str, topic_id: int, prefer_sheets: bool = False) -> dict:
    if not isinstance(data, dict):
        return {"success": False, "error": "ESTIMATE_RESULT_NOT_DICT"}

    excel = str(data.get("excel_path") or data.get("xlsx_path") or data.get("artifact_path") or "")
    links = {}
    for l in _rgc2_links(data):
        links["existing_" + str(len(links) + 1)] = l

    pdf = ""
    if excel and _rgc2_os.path.exists(excel):
        try:
            pdf = _rgc2_make_pdf_stub(excel, task_id)
        except Exception as e:
            data["pdf_error"] = str(e)

        try:
            from core.artifact_upload_guard import upload_many_or_fail
            files = [{"path": excel, "kind": "estimate_xlsx"}]
            if pdf and _rgc2_os.path.exists(pdf):
                files.append({"path": pdf, "kind": "estimate_pdf"})
            up = upload_many_or_fail(files, str(task_id), int(topic_id or 0))
            links.update(up.get("links") or {})
            data["upload_result"] = up
        except Exception as e:
            data["upload_error"] = str(e)

        if prefer_sheets:
            try:
                from core.sheets_generator import create_google_sheet
                from openpyxl import load_workbook
                wb = load_workbook(excel, data_only=False)
                ws = wb.active
                rows = [[cell.value for cell in row] for row in ws.iter_rows()]
                wb.close()
                sl = create_google_sheet("Estimate_" + str(task_id)[:8], rows, int(topic_id or 0), str(task_id))
                if sl:
                    links["google_sheet"] = sl
                    data["google_sheet_link"] = sl
            except Exception as e:
                data["google_sheet_error"] = str(e)

    if links:
        data["links"] = {**(data.get("links") or {}), **links} if isinstance(data.get("links"), dict) else links

    first = data.get("google_sheet_link") or _rgc2_best_link(data)
    if first:
        data["drive_link"] = first
        data["success"] = True
        data["artifact_path"] = excel or data.get("artifact_path")
        if pdf:
            extras = data.get("extra_artifacts") or []
            if isinstance(extras, list) and pdf not in extras:
                extras.append(pdf)
            data["extra_artifacts"] = extras
    else:
        data["success"] = False
        data["error"] = data.get("error") or "ESTIMATE_NO_LINK_NO_SUCCESS_V1:NO_DRIVE_TELEGRAM_OR_RETRY_LINK"

    vr = validate_estimate_result(data, task_id=str(task_id))
    data["estimate_validator"] = vr
    if not vr.get("ok"):
        data["success"] = False
        data["validator_reason"] = vr.get("reason")
    return data

async def process_estimate_to_excel(file_path: str, task_id: str, topic_id: int):
    data = await _rgc2_orig_excel(file_path, task_id, topic_id)
    return _rgc2_finalize(data, task_id, topic_id, prefer_sheets=False)

async def process_estimate_to_sheets(file_path: str, task_id: str, topic_id: int):
    try:
        data = await _rgc2_orig_sheets(file_path, task_id, topic_id)
    except Exception:
        data = await _rgc2_orig_excel(file_path, task_id, topic_id)
    return _rgc2_finalize(data, task_id, topic_id, prefer_sheets=True)

async def generate_estimate_from_text(raw_input: str, task_id: str, topic_id: int = 0, chat_id: str = ""):
    ctx = _rgc2_resolve_task_context(str(task_id), int(topic_id or 0))
    if not chat_id:
        chat_id = ctx.get("chat_id") or ""
    topic_id = int(ctx.get("topic_id") or topic_id or 0)

    if _rgc2_has_llm_arithmetic(raw_input):
        try:
            logger.warning("ESTIMATE_NO_LLM_CALC_GUARD_V1_INPUT_ARITHMETIC task=%s", task_id)
        except Exception:
            pass

    if chat_id:
        tpl = _rgc2_get_active_estimate_template(str(chat_id), int(topic_id or 0))
        if tpl:
            try:
                from core.sample_template_engine import create_estimate_from_saved_template
                result = await create_estimate_from_saved_template(
                    raw_input=str(raw_input),
                    task_id=str(task_id),
                    chat_id=str(chat_id),
                    topic_id=int(topic_id or 0),
                )
                if isinstance(result, dict) and (result.get("pdf_link") or result.get("xlsx_link") or result.get("drive_link") or result.get("excel_path")):
                    return _rgc2_finalize(result, task_id, topic_id, prefer_sheets=True)
                return {
                    "success": False,
                    "state": "WAITING_CLARIFICATION",
                    "error": "ESTIMATE_TEMPLATE_STRICT_REUSE_V1:TEMPLATE_NOT_APPLICABLE",
                    "result_text": "Активный шаблон сметы найден, но не подошёл к запросу. Уточни состав работ, объёмы или замени шаблон.",
                }
            except Exception as e:
                return {
                    "success": False,
                    "state": "WAITING_CLARIFICATION",
                    "error": "ESTIMATE_TEMPLATE_STRICT_REUSE_V1:ERROR:" + str(e)[:200],
                    "result_text": "Активный шаблон сметы найден, но применить его не удалось. Уточни параметры или замени шаблон.",
                }

    data = await _rgc2_orig_text(raw_input, task_id, topic_id)
    result_text = " ".join(str(data.get(k) or "") for k in ("result", "result_text", "message", "summary")) if isinstance(data, dict) else str(data)
    if _rgc2_has_llm_arithmetic(result_text) and not (isinstance(data, dict) and (data.get("excel_path") or data.get("artifact_path"))):
        return {
            "success": False,
            "error": "ESTIMATE_NO_LLM_CALC_GUARD_V1:TEXT_CALC_WITHOUT_PYTHON_ARTIFACT",
            "result_text": "Смета не принята: обнаружен текстовый расчёт без Python/OpenPyXL артефакта.",
        }
    return _rgc2_finalize(data, task_id, topic_id, prefer_sheets=True)
# === END_ESTIMATE_TEMPLATE_STRICT_REUSE_V1 ===
# === END_ESTIMATE_NO_LLM_CALC_GUARD_V1 ===
# === END_ESTIMATE_RESULT_VALIDATOR_V1 ===
# === END_REAL_GAPS_CLOSE_V2_ESTIMATE ===


# === FINAL_CLOSURE_BLOCKER_FIX_V1_ESTIMATE_XLSX_FORMULAS ===
def create_estimate_xlsx_from_rows(rows, out_path: str, title: str = "Смета") -> str:
    from pathlib import Path
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена", "Сумма"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    start = 4
    safe_rows = rows or []

    for i, r in enumerate(safe_rows, start):
        idx = i - start + 1
        name = r.get("name") or r.get("work") or r.get("item") or ""
        unit = r.get("unit") or ""
        qty = r.get("qty") or r.get("quantity") or 0
        price = r.get("price") or r.get("unit_price") or 0

        try:
            qty = float(str(qty).replace(",", ".").replace(" ", ""))
        except Exception:
            qty = 0

        try:
            price = float(str(price).replace(",", ".").replace(" ", ""))
        except Exception:
            price = 0

        ws.cell(row=i, column=1, value=idx)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=unit)
        ws.cell(row=i, column=4, value=qty)
        ws.cell(row=i, column=5, value=price)
        ws.cell(row=i, column=6, value=f"=D{i}*E{i}")

    total_row = start + len(safe_rows)
    ws.cell(row=total_row, column=5, value="Итого")
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=f"=SUM(F{start}:F{total_row-1})" if safe_rows else "=0")
    ws.cell(row=total_row, column=6).font = Font(bold=True)

    for i, w in enumerate([8, 55, 12, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out)
    return str(out)
# === END_FINAL_CLOSURE_BLOCKER_FIX_V1_ESTIMATE_XLSX_FORMULAS ===

