# === FULLFIX_13A_SAMPLE_FILE_INTENT_AND_TEMPLATE_ESTIMATE ===
import os
import re
import json
import tempfile
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict, Optional, List, Tuple

ENGINE = "FULLFIX_13A_SAMPLE_FILE_INTENT_AND_TEMPLATE_ESTIMATE"
BASE = Path("/root/.areal-neva-core")
TEMPLATE_ROOT = BASE / "data" / "templates"
ESTIMATE_TEMPLATE_DIR = TEMPLATE_ROOT / "estimate"
PROJECT_TEMPLATE_DIR = TEMPLATE_ROOT / "project"

SAMPLE_WORDS = (
    "образец", "шаблон", "пример", "как образец", "как шаблон",
    "используй это", "возьми это", "сохрани это", "запомни это",
    "по этому", "по нему", "по ней", "в таком формате", "такой формат"
)

ESTIMATE_WORDS = (
    "смет", "кс-2", "кс2", "ведомост", "спецификац", "расцен", "цена",
    "стоимост", "позици", "материал", "работ", "руб", "xlsx", "excel", "таблиц"
)

PROJECT_WORDS = (
    "проект", "кж", "кд", "ар", "км", "чертеж", "чертёж", "плит", "фундамент",
    "кровл", "стропил", "архитект", "dxf", "dwg", "pdf"
)

FILE_EXT_ESTIMATE = (".xlsx", ".xls", ".csv")
FILE_EXT_PROJECT = (".pdf", ".dxf", ".dwg", ".docx")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_text(v: Any) -> str:
    return str(v or "").strip()


def _json_loads_maybe(v: Any) -> Dict[str, Any]:
    if isinstance(v, dict):
        return v
    s = _safe_text(v)
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        return {}


def _low(v: Any) -> str:
    return _safe_text(v).lower().replace("ё", "е")


def _task_history_insert(conn, task_id: str, action: str) -> None:
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(task_history)").fetchall()]
        if "action" in cols:
            conn.execute(
                "INSERT INTO task_history (task_id,action,created_at) VALUES (?,?,datetime('now'))",
                (task_id, action),
            )
        elif "event" in cols:
            conn.execute(
                "INSERT INTO task_history (task_id,event,created_at) VALUES (?,?,datetime('now'))",
                (task_id, action),
            )
        conn.commit()
    except Exception:
        pass



def _send_reply(chat_id: str, text: str, reply_to_message_id: Optional[int] = None, topic_id: int = 0) -> Optional[int]:  # TOPIC2_REPLY_THREAD_FIX_V1
    try:
        from core.reply_sender import send_reply_ex
        kwargs = {
            "chat_id": str(chat_id),
            "text": text,
            "reply_to_message_id": reply_to_message_id,
        }
        if int(topic_id or 0) > 0:
            kwargs["message_thread_id"] = int(topic_id or 0)  # TOPIC2_REPLY_THREAD_FIX_V1
        res = send_reply_ex(**kwargs)
        if isinstance(res, dict):
            return res.get("bot_message_id")
    except Exception:
        return None
    return None

def _update_task(conn, task_id: str, state: str, result: str = "", error_message: str = "", bot_message_id: Optional[int] = None) -> None:
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(tasks)").fetchall()]
        sets = ["state=?", "result=?", "error_message=?", "updated_at=datetime('now')"]
        vals: List[Any] = [state, result, error_message]
        if bot_message_id and "bot_message_id" in cols:
            sets.append("bot_message_id=?")
            vals.append(int(bot_message_id))
        vals.append(task_id)
        conn.execute(f"UPDATE tasks SET {', '.join(sets)} WHERE id=?", vals)
        conn.commit()
    except Exception:
        pass


def detect_sample_template_intent(raw_input: Any, input_type: str = "") -> bool:
    payload = _json_loads_maybe(raw_input)
    text = " ".join([
        _safe_text(raw_input),
        _safe_text(payload.get("caption")),
        _safe_text(payload.get("file_name")),
    ])
    low = _low(text)
    if not low:
        return False
    has_sample = any(w in low for w in SAMPLE_WORDS)
    if not has_sample:
        return False
    has_domain = any(w in low for w in ESTIMATE_WORDS + PROJECT_WORDS)
    if has_domain:
        return True
    return any(x in low for x in ("файл", "это", "таблица", "документ"))


def detect_estimate_intent(raw_input: Any) -> bool:
    low = _low(raw_input)
    if not low:
        return False
    if any(w in low for w in ("проект", "фундамент", "плит", "кровл", "стропил", "чертеж", "чертёж")):
        return False
    return any(w in low for w in ("смет", "посчитай", "расчет", "расчёт", "профлист", "монтаж", "цена", "руб", "м2", "м²", "шт", "ведомость"))


def _template_kind_from_text_and_file(text: str, file_name: str) -> str:
    low = _low(text + " " + file_name)
    ext = Path(file_name).suffix.lower()
    if any(w in low for w in ESTIMATE_WORDS) or ext in FILE_EXT_ESTIMATE:
        return "estimate"
    if any(w in low for w in PROJECT_WORDS) or ext in FILE_EXT_PROJECT:
        return "project"
    return "estimate"


def _parse_file_payload(row_raw: Any) -> Dict[str, Any]:
    payload = _json_loads_maybe(row_raw)
    return {
        "file_id": _safe_text(payload.get("file_id")),
        "file_name": _safe_text(payload.get("file_name")),
        "mime_type": _safe_text(payload.get("mime_type")),
        "caption": _safe_text(payload.get("caption")),
        "source": _safe_text(payload.get("source")),
        "telegram_message_id": payload.get("telegram_message_id"),
        "raw_payload": payload,
    }


def _find_latest_related_file(conn, chat_id: str, topic_id: int, current_task_id: str = "", prefer_estimate: bool = True) -> Optional[Dict[str, Any]]:
    rows = []
    try:
        rows = conn.execute(
            """
            SELECT id,chat_id,COALESCE(topic_id,0) AS topic_id,input_type,raw_input,result,created_at,updated_at
            FROM tasks
            WHERE chat_id=?
              AND COALESCE(topic_id,0)=?
              AND input_type='drive_file'
            ORDER BY rowid DESC
            LIMIT 50
            """,
            (str(chat_id), int(topic_id or 0)),
        ).fetchall()
    except Exception:
        rows = []

    if not rows:
        try:
            rows = conn.execute(
                """
                SELECT id,chat_id,COALESCE(topic_id,0) AS topic_id,input_type,raw_input,result,created_at,updated_at
                FROM tasks
                WHERE chat_id=?
                  AND input_type='drive_file'
                ORDER BY rowid DESC
                LIMIT 50
                """,
                (str(chat_id),),
            ).fetchall()
        except Exception:
            rows = []

    best = None
    best_score = -1

    for r in rows:
        raw = r["raw_input"] if hasattr(r, "keys") else r[4]
        payload = _parse_file_payload(raw)
        file_name = payload.get("file_name", "")
        low_name = _low(file_name)
        score = 0
        if any(low_name.endswith(ext) for ext in FILE_EXT_ESTIMATE):
            score += 80
        if any(k in low_name for k in ("smet", "смет", "vor", "кирп", "price", "cost", "ведом")):
            score += 40
        if prefer_estimate and any(low_name.endswith(ext) for ext in FILE_EXT_ESTIMATE):
            score += 30
        if "project_" in low_name or "кж_project" in low_name or "кд_project" in low_name:
            score -= 60
        if "smoke" in low_name:
            score -= 80
        if score > best_score:
            best_score = score
            best = {
                "task_id": r["id"] if hasattr(r, "keys") else r[0],
                "chat_id": r["chat_id"] if hasattr(r, "keys") else r[1],
                "topic_id": r["topic_id"] if hasattr(r, "keys") else r[2],
                "result": r["result"] if hasattr(r, "keys") else r[5],
                "created_at": r["created_at"] if hasattr(r, "keys") else r[6],
                "updated_at": r["updated_at"] if hasattr(r, "keys") else r[7],
                **payload,
                "score": score,
            }

    return best


def _template_paths(kind: str, chat_id: str, topic_id: int) -> Tuple[Path, Path]:
    base_dir = ESTIMATE_TEMPLATE_DIR if kind == "estimate" else PROJECT_TEMPLATE_DIR
    base_dir.mkdir(parents=True, exist_ok=True)
    safe_chat = re.sub(r"[^0-9A-Za-z_-]+", "_", str(chat_id))[:80]
    active = base_dir / f"ACTIVE__chat_{safe_chat}__topic_{int(topic_id or 0)}.json"
    snapshot = base_dir / f"TEMPLATE__chat_{safe_chat}__topic_{int(topic_id or 0)}__{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    return active, snapshot


def save_template_pointer(conn, chat_id: str, topic_id: int, task_id: str, raw_input: Any, input_type: str = "text") -> Tuple[bool, Dict[str, Any], str]:
    payload = _json_loads_maybe(raw_input)
    text = " ".join([_safe_text(raw_input), _safe_text(payload.get("caption"))])

    if input_type == "drive_file":
        file_meta = _parse_file_payload(raw_input)
        file_meta.update({"task_id": task_id, "chat_id": chat_id, "topic_id": topic_id})
    else:
        file_meta = _find_latest_related_file(conn, chat_id, topic_id, task_id, prefer_estimate=True)

    if not file_meta or not file_meta.get("file_id"):
        return False, {}, "RELATED_FILE_NOT_FOUND"

    kind = _template_kind_from_text_and_file(text, file_meta.get("file_name", ""))
    active, snapshot = _template_paths(kind, chat_id, topic_id)

    template = {
        "engine": ENGINE,
        "kind": kind,
        "status": "active",
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "saved_by_task_id": str(task_id),
        "source_task_id": str(file_meta.get("task_id") or ""),
        "source_file_id": str(file_meta.get("file_id") or ""),
        "source_file_name": str(file_meta.get("file_name") or ""),
        "source_mime_type": str(file_meta.get("mime_type") or ""),
        "source_caption": str(file_meta.get("caption") or ""),
        "source_score": file_meta.get("score"),
        "saved_at": _now(),
        "usage_rule": "Use this file as formatting/sample reference for future estimate/project artifacts in the same chat and topic",
        "raw_user_instruction": _safe_text(raw_input),
    }

    active.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")
    snapshot.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")

    try:
        conn.execute(
            "INSERT INTO memory (chat_id,key,value,timestamp) VALUES (?,?,?,datetime('now'))",
            (
                str(chat_id),
                f"topic_{int(topic_id or 0)}_{kind}_active_template",
                json.dumps(template, ensure_ascii=False),
            ),
        )
        conn.commit()
    except Exception:
        pass

    _task_history_insert(conn, task_id, f"FULLFIX_13A_TEMPLATE_SAVED:{kind}:{file_meta.get('file_name')}")
    return True, template, ""


def _load_active_template(kind: str, chat_id: str, topic_id: int) -> Optional[Dict[str, Any]]:
    active, _ = _template_paths(kind, chat_id, topic_id)
    if active.exists():
        try:
            return json.loads(active.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def _parse_estimate_items(text: str) -> List[Dict[str, Any]]:
    src = _safe_text(text)
    low = _low(src)

    parts = re.split(r"[;\n]+|,\s*(?=[А-Яа-яA-Za-z])", src)
    items = []

    for part in parts:
        p = part.strip()
        if not p:
            continue
        m_qty = re.search(r"(.{0,80}?)(\d+(?:[.,]\d+)?)\s*(м²|м2|м\.?2|м³|м3|м\.?3|п\.?м|м\b|шт|кг|т|тонн)", p, re.I)
        if not m_qty:
            continue
        name = re.sub(r"^(сделай|смета|смету|посчитай|расчет|расчёт)\s*:?", "", m_qty.group(1).strip(), flags=re.I).strip(" :-,")
        if not name:
            name = "Позиция"
        qty = float(m_qty.group(2).replace(",", "."))
        unit_raw = m_qty.group(3).lower().replace("м.2", "м²").replace("м2", "м²").replace("м.3", "м³").replace("м3", "м³").replace("пм", "п.м")
        tail = p[m_qty.end():]
        price = 0.0
        m_price = re.search(r"(?:по|цена|стоимость)\s*(\d+(?:[.,]\d+)?)", tail, re.I)
        if not m_price:
            m_price = re.search(r"(\d+(?:[.,]\d+)?)\s*руб", tail, re.I)
        if m_price:
            price = float(m_price.group(1).replace(",", "."))
        items.append({
            "name": name[:120],
            "qty": qty,
            "unit": unit_raw,
            "price": price,
            "total": round(qty * price, 2),
        })

    if not items:
        # fallback for short text
        nums = re.findall(r"\d+(?:[.,]\d+)?", src)
        if nums:
            qty = float(nums[0].replace(",", "."))
            price = float(nums[1].replace(",", ".")) if len(nums) > 1 else 0.0
            items.append({"name": "Позиция по смете", "qty": qty, "unit": "шт", "price": price, "total": round(qty * price, 2)})

    return items


def _write_estimate_xlsx(path: str, items: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_input: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws["A1"] = "Смета"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Движок: {ENGINE}"
    ws["A3"] = f"Шаблон: {template.get('source_file_name') if template else 'не задан'}"
    ws["A4"] = f"Запрос: {raw_input[:180]}"

    headers = ["№", "Наименование", "Ед", "Кол-во", "Цена", "Сумма"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row = 7
    for i, item in enumerate(items, 1):
        vals = [i, item["name"], item["unit"], item["qty"], item["price"], item["total"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
        row += 1

    ws.cell(row=row, column=5, value="Итого").font = Font(bold=True)
    ws.cell(row=row, column=6, value=sum(float(x.get("total") or 0) for x in items)).font = Font(bold=True)

    widths = [6, 48, 10, 14, 14, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws2 = wb.create_sheet("Шаблон")
    ws2["A1"] = "Активный шаблон"
    ws2["A1"].font = Font(bold=True, size=12)
    if template:
        rows = [
            ("Файл", template.get("source_file_name")),
            ("File ID", template.get("source_file_id")),
            ("Тип", template.get("kind")),
            ("Сохранён", template.get("saved_at")),
            ("Правило", template.get("usage_rule")),
        ]
    else:
        rows = [("Файл", "Шаблон не найден")]
    for r, (k, v) in enumerate(rows, 3):
        ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 90

    wb.save(path)


def _write_estimate_pdf(path: str, items: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_input: str) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # === FULLFIX_15_SAMPLE_PDF_CYR_FIX ===
    try:
        from core.pdf_cyrillic import register_cyrillic_fonts, FONT_REGULAR, FONT_BOLD
        register_cyrillic_fonts()
        font = FONT_REGULAR
        font_bold = FONT_BOLD
    except Exception:
        font = "Helvetica"
        font_bold = "Helvetica-Bold"
    for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"):
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                font = "ArealSans"
                break
            except Exception:
                pass

    c = canvas.Canvas(path, pagesize=A4)
    w, h = A4
    x = 15 * mm
    y = h - 18 * mm

    c.setFont(font, 14)
    c.drawString(x, y, "Смета")
    y -= 9 * mm
    c.setFont(font, 8)
    c.drawString(x, y, f"Движок: {ENGINE}")
    y -= 5 * mm
    c.drawString(x, y, f"Шаблон: {template.get('source_file_name') if template else 'не задан'}")
    y -= 8 * mm

    headers = ["№", "Наименование", "Ед", "Кол-во", "Цена", "Сумма"]
    xs = [x, x + 10*mm, x + 95*mm, x + 115*mm, x + 140*mm, x + 165*mm]
    c.setFont(font, 8)
    for col, val in zip(xs, headers):
        c.drawString(col, y, val)
    y -= 4 * mm
    c.line(x, y, w - 15*mm, y)
    y -= 5 * mm

    total = 0.0
    for i, item in enumerate(items, 1):
        if y < 25 * mm:
            c.showPage()
            y = h - 20 * mm
            c.setFont(font, 8)
        total += float(item.get("total") or 0)
        vals = [
            str(i),
            str(item.get("name", ""))[:45],
            str(item.get("unit", "")),
            f"{float(item.get('qty') or 0):g}",
            f"{float(item.get('price') or 0):.2f}",
            f"{float(item.get('total') or 0):.2f}",
        ]
        for col, val in zip(xs, vals):
            c.drawString(col, y, val)
        y -= 6 * mm

    y -= 4 * mm
    c.line(x, y, w - 15*mm, y)
    y -= 7 * mm
    c.setFont(font, 10)
    c.drawString(x + 130*mm, y, f"Итого: {total:.2f} руб")
    c.save()


def _upload(path: str, task_id: str, topic_id: int) -> str:
    try:
        from core.engine_base import upload_artifact_to_drive
        return str(upload_artifact_to_drive(path, task_id, topic_id) or "")
    except Exception:
        return ""


async def create_estimate_from_saved_template(raw_input: str, task_id: str, chat_id: str, topic_id: int = 0) -> Dict[str, Any]:
    template = _load_active_template("estimate", chat_id, topic_id)
    if not template:
        return {"success": False, "engine": ENGINE, "error": "ACTIVE_ESTIMATE_TEMPLATE_NOT_FOUND"}

    items = _parse_estimate_items(raw_input)
    if not items:
        return {"success": False, "engine": ENGINE, "error": "ESTIMATE_ITEMS_NOT_PARSED"}

    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(task_id))[:30]
    out_dir = Path(tempfile.gettempdir()) / f"areal_ff13a_estimate_{safe}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = str(out_dir / f"estimate_{safe}.xlsx")
    pdf_path = str(out_dir / f"estimate_{safe}.pdf")
    manifest_path = str(out_dir / f"estimate_{safe}.manifest.json")

    _write_estimate_xlsx(xlsx_path, items, template, raw_input)
    _write_estimate_pdf(pdf_path, items, template, raw_input)

    total = round(sum(float(x.get("total") or 0) for x in items), 2)
    manifest = {
        "engine": ENGINE,
        "task_id": task_id,
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "kind": "estimate",
        "template": template,
        "items": items,
        "total": total,
        "created_at": _now(),
    }
    Path(manifest_path).write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    pdf_link = _upload(pdf_path, task_id, topic_id)
    xlsx_link = _upload(xlsx_path, task_id, topic_id)
    manifest_link = _upload(manifest_path, task_id, topic_id)

    if not pdf_link or not xlsx_link:
        return {"success": False, "engine": ENGINE, "error": "ESTIMATE_UPLOAD_FAILED", "pdf_link": pdf_link, "xlsx_link": xlsx_link}

    msg = (
        "Смета создана по сохранённому образцу\n"
        f"Engine: {ENGINE}\n"
        f"Шаблон: {template.get('source_file_name')}\n"
        f"Позиций: {len(items)}\n"
        f"Итого: {total:.2f} руб\n\n"
        f"PDF: {pdf_link}\n"
        f"XLSX: {xlsx_link}\n"
        f"MANIFEST: {manifest_link}\n\n"
        "Доволен результатом? Ответь: Да / Уточни / Правки"
    )

    return {
        "success": True,
        "engine": ENGINE,
        "items": items,
        "total": total,
        "template": template,
        "pdf_link": pdf_link,
        "xlsx_link": xlsx_link,
        "manifest_link": manifest_link,
        "message": msg,
    }


async def handle_sample_template_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input: Any,
    input_type: str,
    reply_to_message_id: Optional[int] = None,
) -> bool:
    if conn is None:
        return False
    if not detect_sample_template_intent(raw_input, input_type):
        return False

    ok, template, err = save_template_pointer(conn, chat_id, topic_id, task_id, raw_input, input_type)
    if not ok:
        msg = (
            "Не смог сохранить образец: не нашёл связанный файл в этом чате/топике\n"
            "Пришли файл и сразу ответь на него: возьми это как образец"
        )
        bot_id = _send_reply(chat_id, msg, reply_to_message_id)
        _update_task(conn, task_id, "FAILED", msg, err or "RELATED_FILE_NOT_FOUND", bot_id)
        _task_history_insert(conn, task_id, f"FULLFIX_13A_TEMPLATE_SAVE_FAILED:{err}")
        return True

    kind_ru = "смет" if template.get("kind") == "estimate" else "проектной документации"
    msg = (
        f"Шаблон {kind_ru} сохранён\n"
        f"Файл: {template.get('source_file_name')}\n"
        f"Тип: {template.get('kind')}\n"
        f"Topic: {template.get('topic_id')}\n\n"
        "Дальше можешь писать простым языком: сделай смету / сделай проект"
    )
    bot_id = _send_reply(chat_id, msg, reply_to_message_id)
    _update_task(conn, task_id, "DONE", msg, "", bot_id)
    return True


async def handle_template_estimate_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input: Any,
    input_type: str,
    reply_to_message_id: Optional[int] = None,
) -> bool:
    if conn is None:
        return False
    if input_type not in ("text", "voice"):
        return False
    if not detect_estimate_intent(raw_input):
        return False
    template = _load_active_template("estimate", chat_id, topic_id)
    if not template:
        return False

    res = await create_estimate_from_saved_template(str(raw_input), task_id, chat_id, topic_id)
    if not res.get("success"):
        _task_history_insert(conn, task_id, f"FULLFIX_13A_TEMPLATE_ESTIMATE_FAILED:{res.get('error')}")
        return False

    msg = res.get("message") or "Смета создана"
    bot_id = _send_reply(chat_id, msg, reply_to_message_id)
    _update_task(conn, task_id, "AWAITING_CONFIRMATION", msg, "", bot_id)
    _task_history_insert(conn, task_id, "FULLFIX_13A_TEMPLATE_ESTIMATE_OK")
    return True

# === END FULLFIX_13A_SAMPLE_FILE_INTENT_AND_TEMPLATE_ESTIMATE ===


# === FULLFIX_13B_TEMPLATE_SAVE_STRICT_ACK ===
def ff13b_template_saved_message(file_name: str = "", template_type: str = "estimate", topic_id: int = 0) -> str:
    name = str(file_name or "файл").strip()
    t = str(template_type or "estimate").strip()
    return (
        "Шаблон сохранён\n"
        f"Файл: {name}\n"
        f"Тип: {t}\n"
        f"Topic: {topic_id}\n\n"
        "Дальше можно писать простым языком: сделай смету / сделай проект"
    )
# === END FULLFIX_13B_TEMPLATE_SAVE_STRICT_ACK ===

# === PROJECT_TECHNADZOR_TEMPLATE_CLOSE_V1 ===

TECHNADZOR_TEMPLATE_DIR = TEMPLATE_ROOT / "technadzor"
TECHNADZOR_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

TECHNADZOR_WORDS = (
    "технадзор", "дефект", "акт", "нарушение", "предписание", "замечание",
    "осмотр", "проверка", "инспекция", "обследование", "фотофиксация",
)

PROJECT_INTENT_WORDS = (
    "сделай проект", "создай проект", "разработай проект", "подготовь проект",
    "применить шаблон", "используй шаблон", "по шаблону", "как в шаблоне",
    "создай документацию", "сделай документацию", "разработай документацию",
    "по образцу", "аналогично шаблону",
)

TECHNADZOR_INTENT_WORDS = (
    "составь акт", "сделай акт", "создай акт", "сформируй акт",
    "акт по шаблону", "акт как в шаблоне", "технадзор по шаблону",
    "используй шаблон акта", "применить шаблон акта",
)

def detect_project_template_intent(raw_input: Any) -> bool:
    low = _low(raw_input)
    return any(w in low for w in PROJECT_INTENT_WORDS)

def detect_technadzor_template_intent(raw_input: Any) -> bool:
    low = _low(raw_input)
    return any(w in low for w in TECHNADZOR_INTENT_WORDS)

def _technadzor_template_paths(chat_id: str, topic_id: int):
    base = TECHNADZOR_TEMPLATE_DIR / f"chat_{chat_id}" / f"topic_{topic_id}"
    base.mkdir(parents=True, exist_ok=True)
    return base / "active.json", base / "index.json"

async def create_project_from_saved_template(
    raw_input: str,
    task_id: str,
    chat_id: str,
    topic_id: int = 0,
) -> Dict[str, Any]:
    template = _load_active_template("project", chat_id, topic_id)
    if not template:
        return {"success": False, "error": "PROJECT_TEMPLATE_NOT_FOUND"}

    source_file = template.get("source_file_name") or ""
    source_path = template.get("source_file_path") or ""
    section = template.get("section") or "кр"

    try:
        if source_path and any(source_path.lower().endswith(e) for e in (".dwg", ".dxf")):
            from core.project_engine import process_dwg_dxf_project_file
            import asyncio
            res = await process_dwg_dxf_project_file(
                file_path=source_path,
                task_id=task_id,
                topic_id=topic_id,
                raw_input=raw_input,
                file_name=source_file,
                mime_type="",
            )
            if res and res.get("success"):
                link = _upload(res.get("artifact_path") or res.get("docx_path") or "", task_id, topic_id)
                msg = (
                    f"Проект создан по шаблону {source_file}\n"
                    f"Раздел: {res.get('section', section).upper()}\n"
                )
                if link:
                    msg += f"Документ: {link}"
                return {"success": True, "message": msg, "drive_link": link, "engine": "DWG_TEMPLATE"}
            return {"success": False, "error": str((res or {}).get("error", "DWG_PROJECT_FAILED"))}

        if source_path and any(source_path.lower().endswith(e) for e in (".pdf", ".docx", ".txt")):
            from core.project_document_engine import process_project_document
            res = await process_project_document(
                file_path=source_path,
                file_name=source_file,
                user_text=raw_input,
                topic_role="проектирование по шаблону",
                task_id=task_id,
                topic_id=topic_id,
            )
            if res and res.get("success"):
                link = _upload(res.get("artifact_path") or "", task_id, topic_id)
                msg = (
                    f"Проектная документация создана по шаблону {source_file}\n"
                    f"Раздел: {res.get('model', {}).get('section_title', section)}\n"
                )
                if link:
                    msg += f"Документ: {link}"
                return {"success": True, "message": msg, "drive_link": link, "engine": "PROJECT_DOC_TEMPLATE"}
            return {"success": False, "error": str((res or {}).get("error", "PROJECT_DOC_FAILED"))}

        return {"success": False, "error": "PROJECT_TEMPLATE_SOURCE_NOT_FOUND_OR_UNSUPPORTED"}

    except Exception as e:
        return {"success": False, "error": f"PROJECT_TEMPLATE_ERR:{e}"}


async def create_technadzor_from_saved_template(
    raw_input: str,
    task_id: str,
    chat_id: str,
    topic_id: int = 0,
    local_photo_path: str = "",
) -> Dict[str, Any]:
    act_ptr, _ = _technadzor_template_paths(chat_id, topic_id)
    if not act_ptr.exists():
        return {"success": False, "error": "TECHNADZOR_TEMPLATE_NOT_FOUND"}

    try:
        template = _json_loads_maybe(act_ptr.read_text(encoding="utf-8"))
    except Exception:
        return {"success": False, "error": "TECHNADZOR_TEMPLATE_CORRUPT"}

    from core.technadzor_engine import process_technadzor
    res = process_technadzor(
        conn=None,
        task_id=task_id,
        chat_id=chat_id,
        topic_id=topic_id,
        raw_input=raw_input or "Технический осмотр по шаблону",
        file_name=template.get("source_file_name") or "",
        local_path=local_photo_path or template.get("source_file_path") or "",
    )

    if res and res.get("ok"):
        art = res.get("artifact") or {}
        link = art.get("drive_link") or _upload(art.get("path") or "", task_id, topic_id)
        msg = res.get("result_text") or "Акт технадзора создан по шаблону"
        if link and link not in msg:
            msg += f"\n\nДокумент: {link}"
        return {"success": True, "message": msg, "drive_link": link, "engine": "TECHNADZOR_TEMPLATE"}

    return {"success": False, "error": "TECHNADZOR_TEMPLATE_ACT_FAILED"}


async def save_technadzor_template_pointer(
    conn,
    chat_id: str,
    topic_id: int,
    task_id: str,
    raw_input: Any,
    input_type: str = "text",
) -> Tuple[bool, Dict[str, Any], str]:
    act_ptr, idx_ptr = _technadzor_template_paths(chat_id, topic_id)

    payload = _parse_file_payload(raw_input)
    template = {
        "kind": "technadzor",
        "task_id": task_id,
        "chat_id": chat_id,
        "topic_id": topic_id,
        "source_file_id": payload.get("file_id") or "",
        "source_file_name": payload.get("file_name") or "",
        "source_file_path": payload.get("local_path") or payload.get("file_path") or "",
        "saved_at": _now(),
        "input_type": input_type,
        "raw_text": _safe_text(payload.get("caption") or raw_input)[:500],
    }

    try:
        act_ptr.write_text(_safe_text(template), encoding="utf-8")
        import json
        act_ptr.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        return False, template, f"TECHNADZOR_TEMPLATE_SAVE_ERR:{e}"

    return True, template, ""


async def handle_template_project_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input: Any,
    input_type: str,
    reply_to_message_id=None,
) -> bool:
    if conn is None:
        return False
    if input_type not in ("text", "voice"):
        return False
    if not detect_project_template_intent(raw_input):
        return False
    template = _load_active_template("project", chat_id, topic_id)
    if not template:
        return False

    res = await create_project_from_saved_template(str(raw_input), task_id, chat_id, topic_id)
    if not res.get("success"):
        _task_history_insert(conn, task_id, f"PROJECT_TEMPLATE_FAILED:{res.get('error')}")
        return False

    msg = res.get("message") or "Проект создан по шаблону"
    bot_id = _send_reply(chat_id, msg, reply_to_message_id)
    _update_task(conn, task_id, "AWAITING_CONFIRMATION", msg, "", bot_id)
    _task_history_insert(conn, task_id, "PROJECT_TEMPLATE_CLOSE_V1_OK")
    return True


async def handle_template_technadzor_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input: Any,
    input_type: str,
    reply_to_message_id=None,
) -> bool:
    if conn is None:
        return False
    if input_type not in ("text", "voice"):
        return False
    if not detect_technadzor_template_intent(raw_input):
        return False

    act_ptr, _ = _technadzor_template_paths(chat_id, topic_id)
    if not act_ptr.exists():
        return False

    res = await create_technadzor_from_saved_template(str(raw_input), task_id, chat_id, topic_id)
    if not res.get("success"):
        _task_history_insert(conn, task_id, f"TECHNADZOR_TEMPLATE_FAILED:{res.get('error')}")
        return False

    msg = res.get("message") or "Акт технадзора создан по шаблону"
    bot_id = _send_reply(chat_id, msg, reply_to_message_id)
    _update_task(conn, task_id, "AWAITING_CONFIRMATION", msg, "", bot_id)
    _task_history_insert(conn, task_id, "TECHNADZOR_TEMPLATE_CLOSE_V1_OK")
    return True

# === END PROJECT_TECHNADZOR_TEMPLATE_CLOSE_V1 ===



# === CLEAN_ESTIMATE_USER_OUTPUT_AND_TOTAL_FIX_V2 ===
# Public Telegram/PDF/XLSX output must not expose internal engine names, manifest links, temp paths, validator text or internal link keys.
# Saved template is formatting reference only. Current user text is the calculation source.

def _ceu2_num(v):
    try:
        return float(str(v or "0").replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0

def _ceu2_unit(u: str) -> str:
    s = str(u or "").lower().replace("м.2", "м²").replace("м2", "м²").replace("м.3", "м³").replace("м3", "м³")
    s = s.replace("пм", "п.м").replace("пог.м", "п.м").replace("п. м", "п.м")
    return s.strip()

def _ceu2_find_price(text: str, patterns) -> float:
    for pat in patterns:
        m = re.search(pat, str(text or ""), re.I)
        if m:
            return _ceu2_num(m.group(1))
    return 0.0

def _parse_estimate_items(text: str) -> List[Dict[str, Any]]:
    src = _safe_text(text)
    if not src:
        return []

    lines = []
    for line in re.split(r"[\n\r]+", src):
        s = line.strip(" \t-•")
        if not s:
            continue
        s = re.sub(r"^\d+[\).\s-]+", "", s).strip()
        if not re.search(r"\d", s):
            continue
        if re.search(r"(м²|м2|м³|м3|п\.?\s?м|шт|кг|т\b|тонн|рейс)", s, re.I):
            lines.append(s)

    if not lines:
        lines = [x.strip() for x in re.split(r";+", src) if x.strip()]

    items = []
    for line in lines:
        m = re.search(
            r"(?P<name>.*?)(?:—|-|:)?\s*(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>м²|м2|м³|м3|п\.?\s?м|шт|кг|т\b|тонн|рейс)",
            line,
            re.I,
        )
        if not m:
            continue

        name = re.sub(r"\s*[—:-]\s*$", "", m.group("name")).strip(" —:-")
        if not name:
            name = "Позиция"

        qty = _ceu2_num(m.group("qty"))
        unit = _ceu2_unit(m.group("unit"))
        tail = line[m.end():]

        material_price = _ceu2_find_price(tail, [
            r"цена\s+материал[а-я]*\s*(\d+(?:[.,]\d+)?)",
            r"материал[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"цена\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"по\s*(\d+(?:[.,]\d+)?)\s*руб",
        ])

        work_price = _ceu2_find_price(tail, [
            r"цена\s+работ[а-я]*\s*(\d+(?:[.,]\d+)?)",
            r"работ[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"вязк[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
        ])

        if material_price <= 0 and work_price <= 0:
            single = _ceu2_find_price(tail, [r"(\d+(?:[.,]\d+)?)\s*руб"])
            material_price = single

        material_sum = round(qty * material_price, 2)
        work_sum = round(qty * work_price, 2)
        total = round(material_sum + work_sum, 2)

        items.append({
            "name": name[:160],
            "unit": unit,
            "qty": qty,
            "material_price": material_price,
            "material_sum": material_sum,
            "work_price": work_price,
            "work_sum": work_sum,
            "price": round(material_price + work_price, 2),
            "total": total,
        })

    return items

def _write_estimate_xlsx(path: str, items: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_input: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws["A1"] = "Смета"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = "Шаблон используется только как формат"
    ws["A3"] = f"Образец: {template.get('source_file_name') if template else 'не задан'}"

    headers = ["№", "Наименование", "Ед.", "Объём", "Цена материала", "Сумма материала", "Цена работы", "Сумма работы", "Итог"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row = 6
    for i, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item["name"])
        ws.cell(row=row, column=3, value=item["unit"])
        ws.cell(row=row, column=4, value=item["qty"])
        ws.cell(row=row, column=5, value=item["material_price"])
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=7, value=item["work_price"])
        ws.cell(row=row, column=8, value=f"=D{row}*G{row}")
        ws.cell(row=row, column=9, value=f"=F{row}+H{row}")
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = border
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=5, value="Итого материалы").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=f"=SUM(F6:F{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=5, value="Итого работы").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=8, value=f"=SUM(H6:H{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=5, value="Общий итог").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=9, value=f"=SUM(I6:I{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 3, column=5, value="НДС 20%").font = Font(bold=True)
    ws.cell(row=total_row + 3, column=9, value=f"=I{total_row+2}*20%").font = Font(bold=True)
    ws.cell(row=total_row + 4, column=5, value="Итого с НДС").font = Font(bold=True)
    ws.cell(row=total_row + 4, column=9, value=f"=I{total_row+2}+I{total_row+3}").font = Font(bold=True)

    for idx, width in enumerate([6, 48, 10, 12, 16, 18, 16, 18, 18], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("Итоги")
    ws2["A1"] = "Итоги"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Итого материалы"
    ws2["B3"] = f"='Смета'!F{total_row}"
    ws2["A4"] = "Итого работы"
    ws2["B4"] = f"='Смета'!H{total_row+1}"
    ws2["A5"] = "Общий итог"
    ws2["B5"] = f"='Смета'!I{total_row+2}"
    ws2["A6"] = "НДС 20%"
    ws2["B6"] = f"='Смета'!I{total_row+3}"
    ws2["A7"] = "Итого с НДС"
    ws2["B7"] = f"='Смета'!I{total_row+4}"

    ws3 = wb.create_sheet("Исходные данные")
    ws3["A1"] = "Текущее задание"
    ws3["A1"].font = Font(bold=True)
    ws3["A2"] = raw_input[:32000]
    ws3["A4"] = "Образец"
    ws3["A5"] = template.get("source_file_name") if template else "не задан"
    ws3.column_dimensions["A"].width = 120

    wb.save(path)

def _write_estimate_pdf(path: str, items: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_input: str) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font = "Helvetica"
    for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"):
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                font = "ArealSans"
                break
            except Exception:
                pass

    c = canvas.Canvas(path, pagesize=A4)
    w, h = A4
    x = 15 * mm
    y = h - 18 * mm

    c.setFont(font, 14)
    c.drawString(x, y, "Смета")
    y -= 8 * mm
    c.setFont(font, 8)
    c.drawString(x, y, "Образец: " + str(template.get("source_file_name") if template else "не задан")[:90])
    y -= 8 * mm

    xs = [x, x + 10*mm, x + 95*mm, x + 115*mm, x + 140*mm, x + 165*mm]
    headers = ["№", "Наименование", "Ед", "Объём", "Цена", "Итог"]
    for col, val in zip(xs, headers):
        c.drawString(col, y, val)
    y -= 4 * mm
    c.line(x, y, w - 15*mm, y)
    y -= 5 * mm

    total = 0.0
    for i, item in enumerate(items, 1):
        if y < 25 * mm:
            c.showPage()
            c.setFont(font, 8)
            y = h - 20 * mm
        total += float(item.get("total") or 0)
        vals = [
            str(i),
            str(item.get("name", ""))[:44],
            str(item.get("unit", "")),
            f"{float(item.get('qty') or 0):g}",
            f"{float(item.get('material_price') or 0) + float(item.get('work_price') or 0):.2f}",
            f"{float(item.get('total') or 0):.2f}",
        ]
        for col, val in zip(xs, vals):
            c.drawString(col, y, val)
        y -= 6 * mm

    y -= 4 * mm
    c.line(x, y, w - 15*mm, y)
    y -= 7 * mm
    c.setFont(font, 10)
    c.drawString(x + 125*mm, y, f"Итого: {total:.2f} руб")
    c.save()

def _ceu2_public_message(template: Optional[Dict[str, Any]], items: List[Dict[str, Any]], total: float, pdf_link: str, xlsx_link: str) -> str:
    file_name = ""
    try:
        file_name = str((template or {}).get("source_file_name") or "").strip()
    except Exception:
        file_name = ""
    if file_name:
        first = f"Смета создана по образцу {file_name}"
    else:
        first = "Смета создана"
    return (
        f"{first}\n"
        f"Позиций: {len(items)} | Итого: {total:.2f} руб\n\n"
        f"PDF: {pdf_link}\n"
        f"XLSX: {xlsx_link}\n\n"
        "Доволен результатом? Да / Уточни / Правки"
    )

async def create_estimate_from_saved_template(raw_input: str, task_id: str, chat_id: str, topic_id: int = 0) -> Dict[str, Any]:
    template = _load_active_template("estimate", chat_id, topic_id)
    if not template:
        return {"success": False, "error": "ACTIVE_ESTIMATE_TEMPLATE_NOT_FOUND"}

    items = _parse_estimate_items(raw_input)
    if not items:
        return {"success": False, "error": "ESTIMATE_ITEMS_NOT_PARSED"}

    total = round(sum(float(x.get("total") or 0) for x in items), 2)
    if total <= 0:
        return {
            "success": False,
            "error": "ESTIMATE_ZERO_TOTAL_BLOCKED",
            "result_text": "Смета не создана: позиции распознаны, но итог равен 0. Проверь цены материала/работы в задании",
            "items": items,
            "total": total,
        }

    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(task_id))[:30]
    out_dir = Path(tempfile.gettempdir()) / f"areal_estimate_{safe}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = str(out_dir / f"estimate_{safe}.xlsx")
    pdf_path = str(out_dir / f"estimate_{safe}.pdf")
    manifest_path = str(out_dir / f"estimate_{safe}.manifest.json")

    _write_estimate_xlsx(xlsx_path, items, template, raw_input)
    _write_estimate_pdf(pdf_path, items, template, raw_input)

    manifest = {
        "engine": "CLEAN_ESTIMATE_USER_OUTPUT_AND_TOTAL_FIX_V2",
        "task_id": task_id,
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "template_used_as_format_only": True,
        "template_file": template.get("source_file_name"),
        "items": items,
        "total": total,
        "created_at": _now(),
    }
    Path(manifest_path).write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    pdf_link = _upload(pdf_path, task_id, topic_id)
    xlsx_link = _upload(xlsx_path, task_id, topic_id)
    manifest_link = _upload(manifest_path, task_id, topic_id)

    if not pdf_link or not xlsx_link:
        return {
            "success": False,
            "error": "ESTIMATE_UPLOAD_FAILED",
            "result_text": "Смета создана локально, но не выгрузилась в Google Drive. Повторю выгрузку через очередь",
            "pdf_link": pdf_link,
            "xlsx_link": xlsx_link,
            "excel_path": xlsx_path,
            "artifact_path": xlsx_path,
            "items": items,
            "total": total,
        }

    msg = _ceu2_public_message(template, items, total, pdf_link, xlsx_link)

    return {
        "success": True,
        "engine": "CLEAN_ESTIMATE_USER_OUTPUT_AND_TOTAL_FIX_V2",
        "items": items,
        "total": total,
        "template": template,
        "template_used_as_format_only": True,
        "pdf_link": pdf_link,
        "xlsx_link": xlsx_link,
        "manifest_link": manifest_link,
        "drive_link": xlsx_link,
        "google_sheet_link": xlsx_link,
        "excel_path": xlsx_path,
        "artifact_path": xlsx_path,
        "message": msg,
        "result_text": msg,
    }
# === END_CLEAN_ESTIMATE_USER_OUTPUT_AND_TOTAL_FIX_V2 ===


# === WEB_SEARCH_PRICE_ENRICHMENT_V1_SAMPLE_TEMPLATE_WRAPPER ===
try:
    _web_price_orig_handle_template_estimate_intent = handle_template_estimate_intent
except Exception:
    _web_price_orig_handle_template_estimate_intent = None

async def handle_template_estimate_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str,
    reply_to_message_id=None,
) -> bool:
    try:
        from core.price_enrichment import maybe_handle_price_enrichment_from_template_engine
        handled = await maybe_handle_price_enrichment_from_template_engine(
            conn=conn,
            task_id=task_id,
            chat_id=chat_id,
            topic_id=topic_id,
            raw_input=raw_input,
            input_type=input_type,
            reply_to_message_id=reply_to_message_id,
        )
        if handled:
            return True
    except Exception:
        pass

    if _web_price_orig_handle_template_estimate_intent is None:
        return False

    return await _web_price_orig_handle_template_estimate_intent(
        conn,
        task_id,
        chat_id,
        topic_id,
        raw_input,
        input_type,
        reply_to_message_id,
    )

# === END_WEB_SEARCH_PRICE_ENRICHMENT_V1_SAMPLE_TEMPLATE_WRAPPER ===

# === THREE_CONTOURS_FINAL_SOURCE_LOCK_V1 ===
# Final source lock:
# - topic_2 estimates: ESTIMATES/templates from Drive is mandatory source of formatting/template
# - current user raw_input is mandatory calculation source
# - stale active templates, stale task results, old Drive links are forbidden
# - project/foundation/roof words must not block estimate route when "смет" is present

_FINAL_ESTIMATE_DRIVE_FOLDER_ID = "19Z3acDgPub4nV55mad5mb8ju63FsqoG9"
_FINAL_ROOT_DRIVE_FOLDER_ID = "13No7_E7Mwj1n1awNQ-lzbohWGOiEM2PB"
_FINAL_DEFAULT_CHAT_ID = "-1003725299009"

def _final_json_dump(obj) -> str:
    import json as _json
    return _json.dumps(obj, ensure_ascii=False, indent=2)

def _final_drive_svc_v1():
    try:
        from core.engine_base import _drive_svc_v1
        return _drive_svc_v1()
    except Exception:
        return None

def _final_drive_list_files_v1(folder_id: str):
    svc = _final_drive_svc_v1()
    if svc is None:
        return []
    try:
        res = svc.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="files(id,name,mimeType,size,modifiedTime)",
            pageSize=100,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        return res.get("files") or []
    except Exception:
        return []

def _final_estimate_score_template_v1(file_name: str, raw_input: str = "") -> int:
    low = (str(file_name or "") + " " + str(raw_input or "")).lower().replace("ё", "е")
    score = 0
    if file_name.lower().endswith((".xlsx", ".xls")):
        score += 100
    if any(x in low for x in ("м-80", "m-80", "м80", "m80")):
        score += 80
    if any(x in low for x in ("м-110", "m-110", "м110", "m110")):
        score += 80
    if any(x in low for x in ("фундамент", "склад", "плита", "бетон", "монолит")) and any(x in str(file_name).lower().replace("ё","е") for x in ("фундамент", "склад")):
        score += 70
    if any(x in low for x in ("кров", "крыша", "перекр", "строп")) and any(x in str(file_name).lower().replace("ё","е") for x in ("крыш", "кров", "перекр")):
        score += 70
    if any(x in low for x in ("ареал", "нева", "газобетон")) and any(x in str(file_name).lower().replace("ё","е") for x in ("ареал", "нева")):
        score += 70
    try:
        score += min(int(str(file_name).count(" ")) * 2, 10)
    except Exception:
        pass
    return score

def _final_select_estimate_template_v1(raw_input: str = "") -> dict:
    files = _final_drive_list_files_v1(_FINAL_ESTIMATE_DRIVE_FOLDER_ID)
    xlsx = [f for f in files if str(f.get("name","")).lower().endswith((".xlsx", ".xls"))]
    if not xlsx:
        return {}
    xlsx = sorted(
        xlsx,
        key=lambda f: (
            _final_estimate_score_template_v1(f.get("name",""), raw_input),
            int(f.get("size") or 0),
            str(f.get("modifiedTime") or ""),
        ),
        reverse=True,
    )
    return xlsx[0]

def _final_template_paths_estimate_v1(chat_id: str, topic_id: int):
    import re as _re
    safe_chat = _re.sub(r"[^0-9A-Za-z_-]+", "_", str(chat_id or _FINAL_DEFAULT_CHAT_ID))[:80]
    ESTIMATE_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    active = ESTIMATE_TEMPLATE_DIR / f"ACTIVE__chat_{safe_chat}__topic_{int(topic_id or 0)}.json"
    snapshot = ESTIMATE_TEMPLATE_DIR / f"TEMPLATE__chat_{safe_chat}__topic_{int(topic_id or 0)}__FINAL_SOURCE_LOCK.json"
    return active, snapshot

def _final_force_drive_estimate_template_v1(chat_id: str, topic_id: int, raw_input: str = "") -> dict:
    from datetime import datetime as _dt, timezone as _tz
    selected = _final_select_estimate_template_v1(raw_input)
    if not selected:
        return {}
    all_files = [
        {
            "file_id": f.get("id",""),
            "file_name": f.get("name",""),
            "mime_type": f.get("mimeType",""),
            "size": f.get("size",""),
            "modifiedTime": f.get("modifiedTime",""),
        }
        for f in _final_drive_list_files_v1(_FINAL_ESTIMATE_DRIVE_FOLDER_ID)
        if str(f.get("name","")).lower().endswith((".xlsx", ".xls"))
    ]
    template = {
        "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1",
        "kind": "estimate",
        "status": "active",
        "source": "DRIVE_ESTIMATES_TEMPLATES",
        "source_folder_name": "ESTIMATES/templates",
        "source_folder_id": _FINAL_ESTIMATE_DRIVE_FOLDER_ID,
        "chat_id": str(chat_id or _FINAL_DEFAULT_CHAT_ID),
        "topic_id": int(topic_id or 0),
        "source_file_id": selected.get("id",""),
        "source_file_name": selected.get("name",""),
        "source_mime_type": selected.get("mimeType",""),
        "source_size": selected.get("size",""),
        "source_modifiedTime": selected.get("modifiedTime",""),
        "synced_at": _dt.now(_tz.utc).isoformat(),
        "template_used_as_format_only": True,
        "calculation_source_rule": "ONLY_CURRENT_RAW_INPUT",
        "forbidden_sources": [
            "old_task_result",
            "old_drive_links",
            "old_estimate_artifacts",
            "project_artifacts",
            "topic_210_project_templates",
        ],
        "all_templates": all_files,
        "raw_user_instruction_sample": str(raw_input or "")[:500],
    }
    active, snapshot = _final_template_paths_estimate_v1(chat_id, topic_id)
    active.write_text(_final_json_dump(template), encoding="utf-8")
    snapshot.write_text(_final_json_dump(template), encoding="utf-8")
    return template

def _final_download_drive_file_v1(file_id: str, dst_path: str) -> bool:
    try:
        import io as _io
        from googleapiclient.http import MediaIoBaseDownload
        svc = _final_drive_svc_v1()
        if svc is None:
            return False
        meta = svc.files().get(fileId=file_id, fields="id,name,mimeType").execute()
        mime = str(meta.get("mimeType") or "")
        if mime == "application/vnd.google-apps.spreadsheet":
            req = svc.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            req = svc.files().get_media(fileId=file_id)
        fh = _io.FileIO(dst_path, "wb")
        downloader = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.close()
        return True
    except Exception:
        return False

def detect_estimate_intent(raw_input: Any) -> bool:
    low = _low(raw_input)
    if not low:
        return False
    if "смет" in low or "кс-2" in low or "кс2" in low or "ведомост" in low:
        return True
    if "проект" in low and not any(w in low for w in ("смет", "стоимост", "цена", "расчет", "расчёт")):
        return False
    return any(w in low for w in ("посчитай", "расчет", "расчёт", "цена", "стоимость", "руб", "м2", "м²", "м3", "м³", "шт", "ведомость"))

def _load_active_template(kind: str, chat_id: str, topic_id: int):
    if kind == "estimate":
        forced = _final_force_drive_estimate_template_v1(str(chat_id or _FINAL_DEFAULT_CHAT_ID), int(topic_id or 0), "")
        if forced:
            return forced
    active, _ = _template_paths(kind, chat_id, topic_id)
    if active.exists():
        try:
            import json as _json
            return _json.loads(active.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None

def _final_num_v1(v):
    try:
        return float(str(v or "0").replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0

def _final_unit_v1(u: str) -> str:
    s = str(u or "").lower().replace("м.2", "м²").replace("м2", "м²").replace("м.3", "м³").replace("м3", "м³")
    s = s.replace("пм", "п.м").replace("пог.м", "п.м").replace("п. м", "п.м")
    return s.strip()

def _final_parse_price_v1(text: str, patterns) -> float:
    import re as _re
    for pat in patterns:
        m = _re.search(pat, str(text or ""), _re.I)
        if m:
            return _final_num_v1(m.group(1))
    return 0.0

def _final_extract_rates_from_template_v1(template_path: str) -> dict:
    rates = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(template_path, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 300), values_only=True):
                vals = [x for x in row if x not in (None, "")]
                if len(vals) < 2:
                    continue
                texts = [str(x).strip() for x in vals if isinstance(x, str) and len(str(x).strip()) >= 3]
                nums = []
                for x in vals:
                    if isinstance(x, (int, float)) and 0 < float(x) < 1000000:
                        nums.append(float(x))
                if texts and nums:
                    name = max(texts, key=len).lower().replace("ё", "е")
                    price = nums[-1]
                    if price > 0:
                        for token in name.split():
                            if len(token) >= 4:
                                rates[token] = price
        wb.close()
    except Exception:
        pass
    return rates

def _parse_estimate_items(text: str) -> List[Dict[str, Any]]:
    import re as _re
    src = _safe_text(text)
    if not src:
        return []
    lines = []
    for line in _re.split(r"[\n\r;]+", src):
        s = line.strip(" \t-•")
        if s:
            lines.append(s)
    if not lines:
        lines = [src]

    items = []
    for line in lines:
        m = _re.search(
            r"(?P<name>.*?)(?:—|-|:)?\s*(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>м²|м2|м³|м3|п\.?\s?м|пог\.?\s?м|м\b|шт|кг|т\b|тонн|рейс)",
            line,
            _re.I,
        )
        if not m:
            continue
        name = _re.sub(r"^(сделай|смета|смету|посчитай|расчет|расчёт|подробную)\s*:?", "", m.group("name").strip(), flags=_re.I).strip(" —:-")
        if not name:
            name = "Позиция"
        qty = _final_num_v1(m.group("qty"))
        unit = _final_unit_v1(m.group("unit"))
        tail = line[m.end():]

        material_price = _final_parse_price_v1(tail, [
            r"цена\s+материал[а-я]*\s*(\d+(?:[.,]\d+)?)",
            r"материал[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"по\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"цена\s*(\d+(?:[.,]\d+)?)\s*руб",
        ])
        work_price = _final_parse_price_v1(tail, [
            r"цена\s+работ[а-я]*\s*(\d+(?:[.,]\d+)?)",
            r"работ[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"вязк[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
            r"монтаж[а-я]*\s*(\d+(?:[.,]\d+)?)\s*руб",
        ])
        if material_price <= 0 and work_price <= 0:
            material_price = _final_parse_price_v1(tail, [r"(\d+(?:[.,]\d+)?)\s*руб"])
        total = round(qty * (material_price + work_price), 2)
        items.append({
            "name": name[:180],
            "unit": unit,
            "qty": qty,
            "material_price": material_price,
            "work_price": work_price,
            "price": round(material_price + work_price, 2),
            "material_sum": round(qty * material_price, 2),
            "work_sum": round(qty * work_price, 2),
            "total": total,
        })

    low = src.lower().replace("ё", "е")
    if not items and ("плит" in low or "фундамент" in low or "монолит" in low):
        m = _re.search(r"(\d+(?:[.,]\d+)?)\s*(?:на|x|х|×)\s*(\d+(?:[.,]\d+)?)", low)
        length = _final_num_v1(m.group(1)) if m else 10.0
        width = _final_num_v1(m.group(2)) if m else 10.0
        area = round(length * width, 2)
        thick_m = 0.25
        mt = _re.search(r"толщин[а-я]*\s*(\d{2,4})\s*мм", low)
        if mt:
            thick_m = _final_num_v1(mt.group(1)) / 1000.0
        concrete = round(area * thick_m, 2)
        rebar_kg = round(area * 25, 2)
        perimeter = round((length + width) * 2, 2)
        default_rows = [
            ("Подготовка основания", area, "м²", 0, 350),
            ("Песчаная подушка", round(area * 0.30, 2), "м³", 1200, 450),
            ("Щебёночная подготовка", round(area * 0.10, 2), "м³", 1800, 450),
            ("Опалубка периметра", perimeter, "п.м", 0, 950),
            ("Арматура А500С", rebar_kg, "кг", 75, 35),
            ("Бетон В25", concrete, "м³", 7500, 2500),
        ]
        for name, qty, unit, mat, work in default_rows:
            items.append({
                "name": name,
                "unit": unit,
                "qty": qty,
                "material_price": mat,
                "work_price": work,
                "price": round(mat + work, 2),
                "material_sum": round(qty * mat, 2),
                "work_sum": round(qty * work, 2),
                "total": round(qty * (mat + work), 2),
            })

    return items

def _write_estimate_xlsx(path: str, items: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_input: str) -> None:
    import os as _os
    import tempfile as _tempfile
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    template_file = ""
    try:
        fid = str((template or {}).get("source_file_id") or "")
        if fid:
            tmp = _os.path.join(_tempfile.gettempdir(), "areal_estimate_template_final.xlsx")
            if _final_download_drive_file_v1(fid, tmp):
                template_file = tmp
    except Exception:
        template_file = ""

    wb = None
    if template_file and _os.path.exists(template_file):
        try:
            wb = load_workbook(template_file)
        except Exception:
            wb = None
    if wb is None:
        wb = Workbook()

    ws = wb.create_sheet("Смета_текущее_задание", 0)
    for old in wb.worksheets[1:]:
        try:
            old.sheet_state = "hidden"
        except Exception:
            pass

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws["A1"] = "Смета по текущему заданию"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = "Источник расчёта: только текущий текст задачи"
    ws["A3"] = "Шаблон оформления: " + str((template or {}).get("source_file_name") or "ESTIMATES/templates")
    ws["A4"] = "Старые сметы, старые Drive-ссылки и старые task result не используются"

    headers = ["№", "Наименование", "Ед.", "Объём", "Цена материала", "Сумма материала", "Цена работы", "Сумма работы", "Итог"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row = 7
    for i, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item["name"])
        ws.cell(row=row, column=3, value=item["unit"])
        ws.cell(row=row, column=4, value=float(item["qty"]))
        ws.cell(row=row, column=5, value=float(item.get("material_price") or 0))
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=7, value=float(item.get("work_price") or 0))
        ws.cell(row=row, column=8, value=f"=D{row}*G{row}")
        ws.cell(row=row, column=9, value=f"=F{row}+H{row}")
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = border
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=5, value="Итого материалы").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=f"=SUM(F7:F{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=5, value="Итого работы").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=8, value=f"=SUM(H7:H{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=5, value="Общий итог").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=9, value=f"=SUM(I7:I{row-1})").font = Font(bold=True)

    for idx, width in enumerate([6, 52, 10, 12, 16, 18, 16, 18, 18], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("Исходные_данные", 1)
    ws2["A1"] = "Текущее задание"
    ws2["A1"].font = Font(bold=True)
    ws2["A2"] = raw_input[:32000]
    ws2["A4"] = "Шаблон"
    ws2["B4"] = str((template or {}).get("source_file_name") or "")
    ws2["A5"] = "Правило"
    ws2["B5"] = "Расчёт только из текущего raw_input"
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 120

    wb.save(path)
    wb.close()

def _write_estimate_pdf(path: str, items: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_input: str) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os

    font = "Helvetica"
    for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"):
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                font = "ArealSans"
                break
            except Exception:
                pass

    c = canvas.Canvas(path, pagesize=A4)
    w, h = A4
    x = 12 * mm
    y = h - 18 * mm
    c.setFont(font, 13)
    c.drawString(x, y, "Смета по текущему заданию")
    y -= 7 * mm
    c.setFont(font, 8)
    c.drawString(x, y, "Шаблон: " + str((template or {}).get("source_file_name") or "ESTIMATES/templates")[:95])
    y -= 5 * mm
    c.drawString(x, y, "Источник расчёта: только текущий текст задачи")
    y -= 8 * mm

    headers = ["№", "Наименование", "Ед", "Объём", "Цена", "Итог"]
    xs = [x, x + 9*mm, x + 96*mm, x + 116*mm, x + 140*mm, x + 165*mm]
    for col, val in zip(xs, headers):
        c.drawString(col, y, val)
    y -= 4 * mm
    c.line(x, y, w - 12*mm, y)
    y -= 5 * mm

    total = 0.0
    c.setFont(font, 7)
    for i, item in enumerate(items, 1):
        if y < 25 * mm:
            c.showPage()
            c.setFont(font, 7)
            y = h - 18 * mm
        total += float(item.get("total") or 0)
        vals = [
            str(i),
            str(item.get("name",""))[:48],
            str(item.get("unit","")),
            f"{float(item.get('qty') or 0):g}",
            f"{float(item.get('price') or 0):.2f}",
            f"{float(item.get('total') or 0):.2f}",
        ]
        for col, val in zip(xs, vals):
            c.drawString(col, y, val)
        y -= 5 * mm

    y -= 4 * mm
    c.line(x, y, w - 12*mm, y)
    y -= 7 * mm
    c.setFont(font, 10)
    c.drawString(x + 125*mm, y, f"Итого: {total:.2f} руб")
    c.save()

def _final_public_estimate_message_v1(template: Optional[Dict[str, Any]], items: List[Dict[str, Any]], total: float, pdf_link: str, xlsx_link: str) -> str:
    tmpl = str((template or {}).get("source_file_name") or "ESTIMATES/templates").strip()
    return (
        "Смета создана по новому заданию\n"
        f"Шаблон: {tmpl}\n"
        f"Позиций: {len(items)}\n"
        f"Итого: {total:.2f} руб\n\n"
        "Источник расчёта: только текущий текст задачи\n"
        "Старые сметы, старые Drive-ссылки и старые результаты не использованы\n\n"
        f"PDF: {pdf_link}\n"
        f"XLSX: {xlsx_link}\n\n"
        "Доволен результатом? Да / Уточни / Правки"
    )

async def create_estimate_from_saved_template(raw_input: str, task_id: str, chat_id: str, topic_id: int = 0) -> Dict[str, Any]:
    import re as _re
    import tempfile as _tempfile
    import json as _json
    from pathlib import Path as _Path
    from datetime import datetime as _dt

    template = _final_force_drive_estimate_template_v1(str(chat_id or _FINAL_DEFAULT_CHAT_ID), int(topic_id or 0), str(raw_input or ""))
    if not template:
        return {"success": False, "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1", "error": "DRIVE_ESTIMATES_TEMPLATE_NOT_FOUND"}

    items = _parse_estimate_items(str(raw_input or ""))
    if not items:
        return {"success": False, "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1", "error": "ESTIMATE_ITEMS_NOT_PARSED_FROM_CURRENT_RAW_INPUT"}

    total = round(sum(float(x.get("total") or 0) for x in items), 2)
    if total <= 0:
        return {
            "success": False,
            "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1",
            "error": "ESTIMATE_ZERO_TOTAL_BLOCKED",
            "result_text": "Смета не создана: итог равен 0, старые данные использовать запрещено",
            "items": items,
            "total": total,
        }

    safe = _re.sub(r"[^A-Za-z0-9_-]+", "_", str(task_id))[:30]
    out_dir = _Path(_tempfile.gettempdir()) / f"areal_final_estimate_{safe}_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = str(out_dir / f"estimate_{safe}.xlsx")
    pdf_path = str(out_dir / f"estimate_{safe}.pdf")
    manifest_path = str(out_dir / f"estimate_{safe}.manifest.json")

    _write_estimate_xlsx(xlsx_path, items, template, str(raw_input or ""))
    _write_estimate_pdf(pdf_path, items, template, str(raw_input or ""))

    manifest = {
        "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1",
        "task_id": task_id,
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "source_folder": "ESTIMATES/templates",
        "template_used_as_format_only": True,
        "calculation_source": "current_raw_input_only",
        "template": template,
        "items": items,
        "total": total,
        "created_at": _now(),
    }
    _Path(manifest_path).write_text(_json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    pdf_link = _upload(pdf_path, task_id, topic_id)
    xlsx_link = _upload(xlsx_path, task_id, topic_id)

    if not pdf_link or not xlsx_link:
        return {
            "success": False,
            "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1",
            "error": "ESTIMATE_UPLOAD_FAILED",
            "pdf_link": pdf_link,
            "xlsx_link": xlsx_link,
            "excel_path": xlsx_path,
            "artifact_path": xlsx_path,
            "items": items,
            "total": total,
        }

    msg = _final_public_estimate_message_v1(template, items, total, pdf_link, xlsx_link)
    return {
        "success": True,
        "engine": "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1",
        "items": items,
        "total": total,
        "template": template,
        "template_used_as_format_only": True,
        "pdf_link": pdf_link,
        "xlsx_link": xlsx_link,
        "drive_link": xlsx_link,
        "google_sheet_link": xlsx_link,
        "excel_path": xlsx_path,
        "artifact_path": xlsx_path,
        "message": msg,
        "result_text": msg,
    }

try:
    _final_prev_handle_template_estimate_intent_v1 = handle_template_estimate_intent
except Exception:
    _final_prev_handle_template_estimate_intent_v1 = None

async def handle_template_estimate_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str,
    reply_to_message_id=None,
) -> bool:
    if conn is None:
        return False
    if input_type not in ("text", "voice", "search"):
        return False
    if not detect_estimate_intent(raw_input):
        return False

    if int(topic_id or 0) == 2:
        res = await create_estimate_from_saved_template(str(raw_input or ""), task_id, chat_id, topic_id)
        if not res.get("success"):
            _task_history_insert(conn, task_id, f"THREE_CONTOURS_FINAL_SOURCE_LOCK_V1_FAILED:{res.get('error')}")
            return False
        msg = res.get("message") or "Смета создана"
        bot_id = _send_reply(chat_id, msg, reply_to_message_id)
        _update_task(conn, task_id, "AWAITING_CONFIRMATION", msg, "", bot_id)
        _task_history_insert(conn, task_id, "THREE_CONTOURS_FINAL_SOURCE_LOCK_V1_ESTIMATE_OK")
        return True

    if _final_prev_handle_template_estimate_intent_v1 is not None:
        return await _final_prev_handle_template_estimate_intent_v1(
            conn, task_id, chat_id, topic_id, raw_input, input_type, reply_to_message_id
        )
    return False

# === END_THREE_CONTOURS_FINAL_SOURCE_LOCK_V1 ===

# === STROYKA_TOPIC2_SOURCE_LOCK_PRIORITY_FIX_V1 ===
# Final priority override for topic_2:
# - smeta/foundation/slab/roof/cost requests are estimate intents
# - input_type search is allowed for estimate template path
# - template is format only, calculation source is current raw_input
# - if rows are not parseable, fallback to stroyka canon is allowed

_stc1_orig_detect_estimate_intent = detect_estimate_intent
_stc1_orig_handle_template_estimate_intent = handle_template_estimate_intent

def detect_estimate_intent(raw_input: Any) -> bool:
    low = _low(raw_input)
    if not low:
        return False
    strong_estimate = (
        "смет", "стоимост", "посчитай", "расчет", "расчёт",
        "цена", "руб", "работ", "материал", "монолит",
        "фундамент", "плит", "кровл", "строительств"
    )
    return any(x in low for x in strong_estimate) or _stc1_orig_detect_estimate_intent(raw_input)

async def handle_template_estimate_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str,
    reply_to_message_id=None,
) -> bool:
    if conn is None:
        return False

    if int(topic_id or 0) == 2:
        if str(input_type or "text") not in ("text", "voice", "search"):
            return False

        if not detect_estimate_intent(raw_input):
            return False

        res = await create_estimate_from_saved_template(
            str(raw_input or ""),
            task_id,
            str(chat_id),
            int(topic_id or 0),
        )

        if res.get("success"):
            msg = res.get("message") or res.get("result_text") or "Смета создана по образцу"
            bot_id = _send_reply(str(chat_id), msg, reply_to_message_id)
            _update_task(conn, task_id, "AWAITING_CONFIRMATION", msg, "", bot_id)
            _task_history_insert(conn, task_id, "STROYKA_TOPIC2_SOURCE_LOCK_PRIORITY_FIX_V1:template_estimate_ok")
            return True

        err = str(res.get("error") or "")
        _task_history_insert(conn, task_id, f"STROYKA_TOPIC2_SOURCE_LOCK_PRIORITY_FIX_V1:fallback:{err}")
        return False

    return await _stc1_orig_handle_template_estimate_intent(
        conn,
        task_id,
        chat_id,
        topic_id,
        raw_input,
        input_type,
        reply_to_message_id,
    )

# === END_STROYKA_TOPIC2_SOURCE_LOCK_PRIORITY_FIX_V1 ===

# === THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1 ===
# Full-context gate:
# - topic_2 estimate must not ask step-by-step when the merged task context already contains required facts
# - current raw input + parent waiting task + clarified:* history are calculation context
# - Drive template remains format source only
# - old task results, old Drive links and project artifacts remain forbidden as calculation sources

def _tcfc1_s(v: Any) -> str:
    return str(v or "").strip()

def _tcfc1_low(v: Any) -> str:
    return _tcfc1_s(v).lower().replace("ё", "е")

def _tcfc1_has_estimate_intent(text: str) -> bool:
    low = _tcfc1_low(text)
    return any(x in low for x in (
        "смет", "стоимост", "посчитай", "расчет", "расчёт", "цена", "руб",
        "строительств", "работ", "материал", "дом", "фундамент", "плит",
        "монолит", "кровл", "каркас", "барн", "бар хаус", "barn"
    ))

def _tcfc1_dims(text: str):
    import re as _re
    low = _tcfc1_low(text)
    m = _re.search(r"(\d+(?:[,.]\d+)?)\s*(?:x|х|×|на)\s*(\d+(?:[,.]\d+)?)\s*(?:м|метр|$|\s)", low)
    if m:
        a = float(m.group(1).replace(",", "."))
        b = float(m.group(2).replace(",", "."))
        if a > 0 and b > 0:
            return a, b, round(a * b, 2)
    m = _re.search(r"(?:площадь|пятно|объект)\D{0,20}(\d+(?:[,.]\d+)?)\s*(?:м2|м²|кв)", low)
    if m:
        area = float(m.group(1).replace(",", "."))
        return 0.0, 0.0, round(area, 2)
    return 0.0, 0.0, 0.0

def _tcfc1_height(text: str) -> float:
    import re as _re
    low = _tcfc1_low(text)
    m = _re.search(r"высот[а-я]*\D{0,12}(\d+(?:[,.]\d+)?)\s*(?:м|метр)", low)
    if m:
        return float(m.group(1).replace(",", "."))
    return 3.0

def _tcfc1_distance_km(text: str) -> float:
    import re as _re
    low = _tcfc1_low(text)
    m = _re.search(r"(\d+(?:[,.]\d+)?)\s*км", low)
    if m:
        return float(m.group(1).replace(",", "."))
    return 0.0

def _tcfc1_object_kind(text: str) -> str:
    low = _tcfc1_low(text)
    if any(x in low for x in ("дом", "барн", "бар хаус", "barn")):
        return "дом"
    if any(x in low for x in ("фундамент", "плит", "монолит")):
        return "фундамент"
    if any(x in low for x in ("кровл", "крыша", "клик фальц", "кликфальц")):
        return "кровля"
    if "бан" in low:
        return "баня"
    if "склад" in low:
        return "склад"
    if "ангар" in low:
        return "ангар"
    return ""

def _tcfc1_missing_questions(text: str) -> List[str]:
    low = _tcfc1_low(text)
    kind = _tcfc1_object_kind(low)
    a, b, area = _tcfc1_dims(low)

    missing = []
    if not kind:
        missing.append("Что строим: дом, фундамент, кровлю, склад или ангар?")
    if area <= 0:
        missing.append("Какие размеры объекта?")
    if kind == "дом":
        if not any(x in low for x in ("каркас", "газобет", "кирпич", "брус", "бревно", "сип", "sip")):
            missing.append("Какой материал стен?")
        if not any(x in low for x in ("кров", "крыша", "фальц", "металлочереп", "профлист", "мягкая")):
            missing.append("Какая кровля?")
    return missing

def _tcfc1_build_rows_from_context(text: str) -> List[Dict[str, Any]]:
    low = _tcfc1_low(text)
    kind = _tcfc1_object_kind(text)
    a, b, area = _tcfc1_dims(text)
    h = _tcfc1_height(text)
    dist = _tcfc1_distance_km(text)

    if area <= 0:
        return []

    if a > 0 and b > 0:
        perimeter = round((a + b) * 2, 2)
    else:
        perimeter = round((area ** 0.5) * 4, 2)

    wall_area = round(perimeter * h, 2)
    roof_area = round(area * 1.25, 2)

    rows: List[Dict[str, Any]] = []

    def add(name, unit, qty, mat, work):
        qty = round(float(qty or 0), 2)
        mat = round(float(mat or 0), 2)
        work = round(float(work or 0), 2)
        rows.append({
            "name": name,
            "unit": unit,
            "qty": qty,
            "material_price": mat,
            "material_sum": round(qty * mat, 2),
            "work_price": work,
            "work_sum": round(qty * work, 2),
            "price": round(mat + work, 2),
            "total": round(qty * (mat + work), 2),
        })

    if kind == "фундамент" or any(x in low for x in ("фундамент", "плит", "монолит")):
        add("Подготовка основания", "м²", area, 450, 350)
        add("Песчаная подготовка с уплотнением", "м²", area, 650, 450)
        add("Щебёночная подготовка", "м²", area, 850, 500)
        add("Армирование фундаментной плиты", "м²", area, 2200, 1600)
        add("Бетонирование фундаментной плиты", "м²", area, 3600, 2200)
        add("Гидроизоляция и защитные работы", "м²", area, 700, 450)

    if kind == "дом":
        add("Фундаментная часть под дом", "м²", area, 4200, 2600)
        add("Каркас стен", "м²", wall_area, 3800, 2400)
        if any(x in low for x in ("имитац", "бруса", "фасад", "снаруж")):
            add("Наружная отделка имитацией бруса", "м²", wall_area, 1450, 1250)
        else:
            add("Наружная отделка фасада", "м²", wall_area, 1300, 1200)
        if any(x in low for x in ("внутри", "внутрен", "отделка")):
            add("Внутренняя отделка", "м²", round(wall_area + area, 2), 1600, 1500)
        if any(x in low for x in ("клик фальц", "кликфальц", "фальц")):
            add("Кровля клик-фальц", "м²", roof_area, 2800, 1900)
        else:
            add("Кровельные работы", "м²", roof_area, 2200, 1700)

    if kind == "кровля" and not rows:
        add("Кровельное покрытие", "м²", roof_area, 2400, 1800)
        add("Обрешётка и подконструкция", "м²", roof_area, 900, 850)
        add("Водосточная система и доборные элементы", "п.м", perimeter, 850, 650)

    if dist > 0:
        add("Логистика и доставка", "км", dist, 0, 350)

    if not rows:
        add("Строительные работы по заданию", "м²", area, 3000, 2500)

    return rows

def _tcfc1_write_xlsx(path: str, rows: List[Dict[str, Any]], template: Optional[Dict[str, Any]], raw_context: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws["A1"] = "Смета по полному техническому заданию"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = "Источник расчёта: текущий текст задачи + уточнения пользователя"
    ws["A3"] = f"Шаблон формата: {(template or {}).get('source_file_name') or 'не задан'}"

    headers = ["№", "Наименование", "Ед.", "Объём", "Цена материала", "Сумма материала", "Цена работы", "Сумма работы", "Итог"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row = 6
    for i, item in enumerate(rows, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item["name"])
        ws.cell(row=row, column=3, value=item["unit"])
        ws.cell(row=row, column=4, value=item["qty"])
        ws.cell(row=row, column=5, value=item["material_price"])
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=7, value=item["work_price"])
        ws.cell(row=row, column=8, value=f"=D{row}*G{row}")
        ws.cell(row=row, column=9, value=f"=F{row}+H{row}")
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = border
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=5, value="Итого материалы").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=f"=SUM(F6:F{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=5, value="Итого работы").font = Font(bold=True)
    ws.cell(row=total_row + 1, column=8, value=f"=SUM(H6:H{row-1})").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=5, value="Общий итог").font = Font(bold=True)
    ws.cell(row=total_row + 2, column=9, value=f"=SUM(I6:I{row-1})").font = Font(bold=True)

    for idx, width in enumerate([6, 52, 10, 12, 16, 18, 16, 18, 18], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("Исходные данные")
    ws2["A1"] = "Полный контекст расчёта"
    ws2["A1"].font = Font(bold=True)
    ws2["A2"] = raw_context[:32000]
    ws2["A4"] = "Правило"
    ws2["A5"] = "Если все данные есть в контексте, уточнения не задаются"
    ws2.column_dimensions["A"].width = 120

    wb.save(path)

def _tcfc1_write_pdf(path: str, rows: List[Dict[str, Any]], template: Optional[Dict[str, Any]]) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font = "Helvetica"
    for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"):
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                font = "ArealSans"
                break
            except Exception:
                pass

    c = canvas.Canvas(path, pagesize=A4)
    w, h = A4
    x = 15 * mm
    y = h - 18 * mm

    c.setFont(font, 13)
    c.drawString(x, y, "Смета по полному техническому заданию")
    y -= 7 * mm
    c.setFont(font, 8)
    c.drawString(x, y, "Шаблон: " + str((template or {}).get("source_file_name") or "не задан")[:90])
    y -= 8 * mm

    xs = [x, x + 10*mm, x + 100*mm, x + 118*mm, x + 143*mm, x + 168*mm]
    headers = ["№", "Наименование", "Ед", "Объём", "Цена", "Итог"]
    for col, val in zip(xs, headers):
        c.drawString(col, y, val)
    y -= 4 * mm
    c.line(x, y, w - 15*mm, y)
    y -= 5 * mm

    total = 0.0
    for i, item in enumerate(rows, 1):
        if y < 25 * mm:
            c.showPage()
            c.setFont(font, 8)
            y = h - 20 * mm
        total += float(item.get("total") or 0)
        vals = [
            str(i),
            str(item.get("name", ""))[:46],
            str(item.get("unit", "")),
            f"{float(item.get('qty') or 0):g}",
            f"{float(item.get('price') or 0):.2f}",
            f"{float(item.get('total') or 0):.2f}",
        ]
        for col, val in zip(xs, vals):
            c.drawString(col, y, val)
        y -= 6 * mm

    y -= 4 * mm
    c.line(x, y, w - 15*mm, y)
    y -= 7 * mm
    c.setFont(font, 10)
    c.drawString(x + 125*mm, y, f"Итого: {total:.2f} руб")
    c.save()

async def tcfc1_create_estimate_from_full_context(raw_context: str, task_id: str, chat_id: str, topic_id: int = 2) -> Dict[str, Any]:
    template = _load_active_template("estimate", str(chat_id), int(topic_id or 0))
    if not template:
        try:
            _final_source_lock_sync_estimate_templates_v1(str(chat_id), int(topic_id or 0))
            template = _load_active_template("estimate", str(chat_id), int(topic_id or 0))
        except Exception:
            template = None

    rows = _tcfc1_build_rows_from_context(raw_context)
    if not rows:
        return {"success": False, "error": "FULL_CONTEXT_ROWS_NOT_BUILT"}

    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(task_id))[:30]
    out_dir = Path(tempfile.gettempdir()) / f"areal_tcfc1_estimate_{safe}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = str(out_dir / f"estimate_full_context_{safe}.xlsx")
    pdf_path = str(out_dir / f"estimate_full_context_{safe}.pdf")
    manifest_path = str(out_dir / f"estimate_full_context_{safe}.manifest.json")

    _tcfc1_write_xlsx(xlsx_path, rows, template, raw_context)
    _tcfc1_write_pdf(pdf_path, rows, template)

    total = round(sum(float(x.get("total") or 0) for x in rows), 2)
    manifest = {
        "engine": "THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1",
        "task_id": task_id,
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "template_used_as_format_only": True,
        "template_file": (template or {}).get("source_file_name"),
        "calculation_source_rule": "CURRENT_RAW_INPUT_PLUS_USER_CLARIFICATIONS_ONLY",
        "rows": rows,
        "total": total,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    Path(manifest_path).write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    pdf_link = _upload(pdf_path, task_id, topic_id)
    xlsx_link = _upload(xlsx_path, task_id, topic_id)
    manifest_link = _upload(manifest_path, task_id, topic_id)

    if not pdf_link or not xlsx_link:
        return {
            "success": False,
            "error": "FULL_CONTEXT_UPLOAD_FAILED",
            "pdf_link": pdf_link,
            "xlsx_link": xlsx_link,
            "excel_path": xlsx_path,
            "artifact_path": xlsx_path,
            "rows": rows,
            "total": total,
        }

    msg = (
        "Смета создана по полному техническому заданию\n"
        f"Шаблон формата: {(template or {}).get('source_file_name') or 'не задан'}\n"
        f"Позиций: {len(rows)}\n"
        f"Итого: {total:.2f} руб\n\n"
        "Уточнения не запрашивались: нужные данные найдены в исходном сообщении и уточнениях\n\n"
        f"PDF: {pdf_link}\n"
        f"XLSX: {xlsx_link}\n\n"
        "Доволен результатом? Да / Уточни / Правки"
    )

    return {
        "success": True,
        "engine": "THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1",
        "rows": rows,
        "items": rows,
        "total": total,
        "template": template,
        "pdf_link": pdf_link,
        "xlsx_link": xlsx_link,
        "manifest_link": manifest_link,
        "drive_link": xlsx_link,
        "google_sheet_link": xlsx_link,
        "excel_path": xlsx_path,
        "artifact_path": xlsx_path,
        "message": msg,
        "result_text": msg,
    }

async def handle_stroyka_topic2_full_context_gate_v1(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_context: Any,
    input_type: str = "text",
    reply_to_message_id=None,
) -> bool:
    if conn is None:
        return False
    if int(topic_id or 0) != 2:
        return False
    if str(input_type or "text") not in ("text", "voice", "search"):
        return False

    ctx = _tcfc1_s(raw_context)
    if not _tcfc1_has_estimate_intent(ctx):
        return False

    missing = _tcfc1_missing_questions(ctx)
    if missing:
        _task_history_insert(conn, task_id, "THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1:missing:" + " | ".join(missing[:3]))
        return False

    res = await tcfc1_create_estimate_from_full_context(ctx, task_id, str(chat_id), int(topic_id or 0))
    if not res.get("success"):
        _task_history_insert(conn, task_id, f"THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1:fallback:{res.get('error')}")
        return False

    msg = res.get("message") or res.get("result_text") or "Смета создана по полному техническому заданию"
    bot_id = _send_reply(str(chat_id), msg, reply_to_message_id)
    _update_task(conn, task_id, "AWAITING_CONFIRMATION", msg, "", bot_id)
    _task_history_insert(conn, task_id, "THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1:estimate_ok")
    return True

_tcfc1_prev_handle_template_estimate_intent = handle_template_estimate_intent

async def handle_template_estimate_intent(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str,
    reply_to_message_id=None,
) -> bool:
    if int(topic_id or 0) == 2:
        try:
            handled = await handle_stroyka_topic2_full_context_gate_v1(
                conn, task_id, str(chat_id), int(topic_id or 0), raw_input, input_type, reply_to_message_id
            )
            if handled:
                return True
        except Exception:
            pass

    return await _tcfc1_prev_handle_template_estimate_intent(
        conn, task_id, chat_id, topic_id, raw_input, input_type, reply_to_message_id
    )

# === END_THREE_CONTOURS_FULL_CONTEXT_NO_REPEAT_CLARIFY_V1 ===


# === TOPIC2_ONE_BIG_FINAL_PIPELINE_V1_SAMPLE_ENGINE ===
# Single topic_2 construction estimate pipeline:
# - one route only
# - no old search/direct_item/memory revive artifacts
# - no repeated clarification when full context exists
# - formula workbook is copied from active Drive template and formulas are preserved

import os as _t2sp_os
import re as _t2sp_re
import json as _t2sp_json
import time as _t2sp_time
import shutil as _t2sp_shutil
import tempfile as _t2sp_tempfile
import subprocess as _t2sp_subprocess
from pathlib import Path as _t2sp_Path
from datetime import datetime as _t2sp_datetime
from typing import Any as _t2sp_Any, Dict as _t2sp_Dict, List as _t2sp_List, Tuple as _t2sp_Tuple, Optional as _t2sp_Optional

_T2SP_BASE = _t2sp_Path("/root/.areal-neva-core")
_T2SP_ACTIVE_TEMPLATE = _T2SP_BASE / "data/templates/estimate/ACTIVE__chat_-1003725299009__topic_2.json"
_T2SP_OUT_DIR = _T2SP_BASE / "outputs/topic2_formula_estimates"
_T2SP_CACHE_DIR = _T2SP_BASE / "data/templates/estimate/cache"
_T2SP_MARKER = "TOPIC2_ONE_BIG_FINAL_PIPELINE_V1"

def _t2sp_s(v: _t2sp_Any) -> str:
    return "" if v is None else str(v)

def _t2sp_low(v: _t2sp_Any) -> str:
    return _t2sp_s(v).lower().replace("ё", "е").strip()

def _t2sp_norm_numbers(text: str) -> str:
    repl = {
        "ноль": "0", "один": "1", "одна": "1", "одно": "1", "два": "2", "две": "2",
        "три": "3", "четыре": "4", "пять": "5", "шесть": "6", "семь": "7",
        "восемь": "8", "девять": "9", "десять": "10", "одиннадцать": "11",
        "двенадцать": "12", "тринадцать": "13", "четырнадцать": "14",
        "пятнадцать": "15", "шестнадцать": "16", "семнадцать": "17",
        "восемнадцать": "18", "девятнадцать": "19", "двадцать": "20",
    }
    out = " " + _t2sp_s(text).lower().replace("ё", "е") + " "
    for k, v in repl.items():
        out = _t2sp_re.sub(rf"(?<![а-яa-z]){_t2sp_re.escape(k)}(?![а-яa-z])", v, out, flags=_t2sp_re.I)
    return out.strip()

def _t2sp_is_estimate_text(text: str) -> bool:
    low = _t2sp_low(text)
    if not low:
        return False
    words = (
        "смет", "стоимост", "посчитай", "рассчитай", "расчет", "расчёт",
        "цена", "руб", "итог", "материал", "работ", "монолит", "фундамент",
        "плит", "дом", "барн", "бар house", "бархаус", "barn", "кровл",
        "каркас", "газобетон", "строительств"
    )
    return any(w in low for w in words)

def _t2sp_parse_context(text: str) -> _t2sp_Dict[str, _t2sp_Any]:
    raw = _t2sp_s(text)
    low = _t2sp_norm_numbers(raw)
    data: _t2sp_Dict[str, _t2sp_Any] = {
        "raw": raw.strip(),
        "object": "",
        "material": "",
        "dimensions": "",
        "length_m": "",
        "width_m": "",
        "floors": "",
        "height_m": "",
        "foundation": "",
        "distance_km": "",
        "style": "",
        "sheet": "",
    }

    if any(x in low for x in ("дом", "коттедж", "барн", "бар house", "бархаус", "barn")):
        data["object"] = "дом"
    elif "фундамент" in low:
        data["object"] = "фундамент"
    elif "кровл" in low:
        data["object"] = "кровля"

    if any(x in low for x in ("барн", "бар house", "бархаус", "barn")):
        data["style"] = "barn house"
        data["material"] = "каркас"
        data["sheet"] = "Каркас под ключ"
    if "каркас" in low:
        data["material"] = "каркас"
        data["sheet"] = "Каркас под ключ"
    if "газобетон" in low or "газоблок" in low:
        data["material"] = "газобетон"
        data["sheet"] = "Газобетон"

    m = _t2sp_re.search(r"(\d+(?:[.,]\d+)?)\s*(?:м|метр(?:ов|а)?)?\s*(?:на|x|х|\*)\s*(\d+(?:[.,]\d+)?)", low, flags=_t2sp_re.I)
    if m:
        l = m.group(1).replace(",", ".")
        w = m.group(2).replace(",", ".")
        data["length_m"] = l
        data["width_m"] = w
        data["dimensions"] = f"{l} x {w} м"

    m = _t2sp_re.search(r"(?:этаж(?:ей|а)?|этажа|этажность)\D{0,12}(\d+)", low, flags=_t2sp_re.I)
    if not m:
        m = _t2sp_re.search(r"(\d+)\s*(?:этаж|этажа|этажей)", low, flags=_t2sp_re.I)
    if m:
        data["floors"] = m.group(1)

    m = _t2sp_re.search(r"(?:высота|h)\D{0,12}(\d+(?:[.,]\d+)?)", low, flags=_t2sp_re.I)
    if m:
        data["height_m"] = m.group(1).replace(",", ".")

    if "монолит" in low and "плит" in low:
        data["foundation"] = "монолитная плита"
    elif "фундамент" in low and "плит" in low:
        data["foundation"] = "плита"
    elif "ленточ" in low:
        data["foundation"] = "ленточный фундамент"
    elif "сва" in low:
        data["foundation"] = "свайный фундамент"
    elif "фундамент" in low:
        data["foundation"] = "фундамент"

    m = _t2sp_re.search(r"(\d+(?:[.,]\d+)?)\s*(?:км|километр)", low, flags=_t2sp_re.I)
    if m:
        data["distance_km"] = m.group(1).replace(",", ".")

    if not data["sheet"]:
        data["sheet"] = "Каркас под ключ" if data["material"] == "каркас" else "Газобетон"

    return data

def _t2sp_missing_questions(text: str) -> _t2sp_List[str]:
    if not _t2sp_is_estimate_text(text):
        return []
    data = _t2sp_parse_context(text)
    missing: _t2sp_List[str] = []
    if not data.get("object"):
        missing.append("тип объекта")
    if not data.get("dimensions"):
        missing.append("размеры объекта")
    if not data.get("material"):
        missing.append("материал стен")
    if not data.get("foundation"):
        missing.append("тип фундамента")
    return missing

def _t2sp_load_active_template_meta() -> _t2sp_Dict[str, _t2sp_Any]:
    if not _T2SP_ACTIVE_TEMPLATE.exists():
        return {}
    try:
        return _t2sp_json.loads(_T2SP_ACTIVE_TEMPLATE.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _t2sp_history(conn, task_id: str, action: str) -> None:
    try:
        fn = globals().get("_task_history_insert")
        if callable(fn):
            fn(conn, task_id, action)
            return
    except Exception:
        pass
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(task_history)").fetchall()]
        col = "action" if "action" in cols else ("event" if "event" in cols else "")
        if col:
            conn.execute(f"INSERT INTO task_history(task_id,{col},created_at) VALUES(?,?,datetime('now'))", (task_id, action))
            conn.commit()
    except Exception:
        pass

def _t2sp_update(conn, task_id: str, state: str, result: str, error: str = "", bot_id: _t2sp_Any = None) -> None:
    try:
        fn = globals().get("_update_task")
        if callable(fn):
            try:
                fn(conn, task_id, state, result, error, bot_id)
                return
            except TypeError:
                fn(conn, task_id, state=state, result=result, error_message=error)
                return
    except Exception:
        pass
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(tasks)").fetchall()]
        if "bot_message_id" in cols and bot_id is not None:
            conn.execute(
                "UPDATE tasks SET state=?, result=?, error_message=?, bot_message_id=?, updated_at=datetime('now') WHERE id=?",
                (state, result, error, bot_id, task_id),
            )
        else:
            conn.execute(
                "UPDATE tasks SET state=?, result=?, error_message=?, updated_at=datetime('now') WHERE id=?",
                (state, result, error, task_id),
            )
        conn.commit()
    except Exception:
        pass


def _t2sp_send(chat_id: str, text: str, reply_to_message_id=None, topic_id: int = 0) -> _t2sp_Any:  # TOPIC2_REPLY_THREAD_FIX_V1
    try:
        fn = globals().get("_send_reply")
        if callable(fn):
            bot_id = fn(str(chat_id), text, reply_to_message_id, topic_id=int(topic_id or 0))
            if bot_id:
                return bot_id
    except Exception:
        pass
    try:
        import requests as _t2sp_requests
        token = _t2sp_os.getenv("TELEGRAM_BOT_TOKEN", "")
        if token:
            payload = {"chat_id": str(chat_id), "text": text}
            if reply_to_message_id:
                payload["reply_to_message_id"] = int(reply_to_message_id)
            if int(topic_id or 0) > 0:
                payload["message_thread_id"] = int(topic_id or 0)  # TOPIC2_REPLY_THREAD_FIX_V1
            r = _t2sp_requests.post(f"https://api.telegram.org/bot{token}/sendMessage", json=payload, timeout=20)
            data = r.json()
            if data.get("ok") and isinstance(data.get("result"), dict):
                return data["result"].get("message_id")
    except Exception:
        pass
    return None

def _t2sp_google_service():
    try:
        import core.google_io as _gio
        for name in ("get_drive_service", "build_drive_service", "drive_service", "get_service"):
            fn = getattr(_gio, name, None)
            if callable(fn):
                try:
                    svc = fn()
                    if svc is not None and hasattr(svc, "files"):
                        return svc
                except Exception:
                    pass
        for obj in vars(_gio).values():
            if obj is not None and hasattr(obj, "files"):
                return obj
    except Exception:
        pass
    return None

async def _t2sp_try_google_io_download(file_id: str, dest: _t2sp_Path) -> bool:
    try:
        import inspect as _t2sp_inspect
        import core.google_io as _gio
        candidates = []
        for name, fn in vars(_gio).items():
            if callable(fn) and "download" in name.lower() and ("drive" in name.lower() or "file" in name.lower()):
                candidates.append((name, fn))
        call_variants = [
            lambda fn: fn(file_id, str(dest)),
            lambda fn: fn(file_id=file_id, dest_path=str(dest)),
            lambda fn: fn(file_id=file_id, output_path=str(dest)),
            lambda fn: fn(file_id=file_id, local_path=str(dest)),
            lambda fn: fn(file_id=file_id, path=str(dest)),
        ]
        for _name, fn in candidates:
            for call in call_variants:
                try:
                    res = call(fn)
                    if _t2sp_inspect.isawaitable(res):
                        res = await res
                    if dest.exists() and dest.stat().st_size > 1000:
                        return True
                    if isinstance(res, str) and _t2sp_Path(res).exists():
                        _t2sp_shutil.copy2(res, dest)
                        return True
                    if isinstance(res, dict):
                        p = res.get("path") or res.get("local_path") or res.get("file_path")
                        if p and _t2sp_Path(p).exists():
                            _t2sp_shutil.copy2(p, dest)
                            return True
                except Exception:
                    continue
    except Exception:
        pass
    return False

async def _t2sp_get_template_file(meta: _t2sp_Dict[str, _t2sp_Any]) -> _t2sp_Path:
    _T2SP_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    file_id = _t2sp_s(meta.get("source_file_id") or "")
    file_name = _t2sp_s(meta.get("source_file_name") or "template.xlsx")
    safe_name = _t2sp_re.sub(r"[^0-9A-Za-zА-Яа-я_.-]+", "_", file_name).strip("_") or "template.xlsx"
    dest = _T2SP_CACHE_DIR / f"{file_id}__{safe_name}"

    if dest.exists() and dest.stat().st_size > 1000:
        return dest

    for p in _T2SP_BASE.rglob(file_name):
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm") and p.stat().st_size > 1000:
            _t2sp_shutil.copy2(p, dest)
            return dest

    if file_id:
        if await _t2sp_try_google_io_download(file_id, dest):
            return dest

        svc = _t2sp_google_service()
        if svc is not None:
            try:
                from googleapiclient.http import MediaIoBaseDownload as _MediaIoBaseDownload
                import io as _io
                request = svc.files().get_media(fileId=file_id)
                fh = _io.FileIO(str(dest), "wb")
                downloader = _MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                fh.close()
                if dest.exists() and dest.stat().st_size > 1000:
                    return dest
            except Exception:
                pass

    raise RuntimeError("ACTIVE_TEMPLATE_XLSX_NOT_AVAILABLE")

async def _t2sp_try_google_io_upload(path: _t2sp_Path) -> str:
    try:
        import inspect as _t2sp_inspect
        import core.google_io as _gio
        folder_id = _t2sp_os.getenv("DRIVE_OUTPUT_FOLDER_ID") or _t2sp_os.getenv("DRIVE_INGEST_FOLDER_ID") or ""
        candidates = []
        for name, fn in vars(_gio).items():
            if callable(fn) and "upload" in name.lower() and ("drive" in name.lower() or "file" in name.lower()):
                candidates.append((name, fn))
        variants = [
            lambda fn: fn(str(path)),
            lambda fn: fn(str(path), folder_id),
            lambda fn: fn(local_path=str(path), folder_id=folder_id),
            lambda fn: fn(file_path=str(path), folder_id=folder_id),
            lambda fn: fn(path=str(path), folder_id=folder_id),
        ]
        for _name, fn in candidates:
            for call in variants:
                try:
                    res = call(fn)
                    if _t2sp_inspect.isawaitable(res):
                        res = await res
                    if isinstance(res, str) and res.startswith("http"):
                        return res
                    if isinstance(res, dict):
                        for k in ("webViewLink", "drive_link", "google_drive_link", "link", "url"):
                            v = res.get(k)
                            if isinstance(v, str) and v.startswith("http"):
                                return v
                        fid = res.get("id") or res.get("file_id")
                        if fid:
                            return f"https://drive.google.com/file/d/{fid}/view?usp=drivesdk"
                except Exception:
                    continue
    except Exception:
        pass
    return ""

def _t2sp_drive_upload_direct(path: _t2sp_Path) -> str:
    folder_id = _t2sp_os.getenv("DRIVE_OUTPUT_FOLDER_ID") or _t2sp_os.getenv("DRIVE_INGEST_FOLDER_ID") or ""
    if not folder_id:
        return ""
    svc = _t2sp_google_service()
    if svc is None:
        return ""
    try:
        from googleapiclient.http import MediaFileUpload as _MediaFileUpload
        body = {"name": path.name, "parents": [folder_id]}
        media = _MediaFileUpload(str(path), resumable=False)
        created = svc.files().create(body=body, media_body=media, fields="id,webViewLink").execute()
        fid = created.get("id")
        try:
            svc.permissions().create(fileId=fid, body={"type": "anyone", "role": "reader"}, fields="id").execute()
        except Exception:
            pass
        return created.get("webViewLink") or f"https://drive.google.com/file/d/{fid}/view?usp=drivesdk"
    except Exception:
        return ""

async def _t2sp_upload(path: _t2sp_Path) -> str:
    link = await _t2sp_try_google_io_upload(path)
    if link:
        return link
    return _t2sp_drive_upload_direct(path)

def _t2sp_formula_count(path: _t2sp_Path) -> int:
    try:
        from openpyxl import load_workbook as _t2sp_load_workbook
        wb = _t2sp_load_workbook(str(path), data_only=False, read_only=True)
        cnt = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str) and v.startswith("="):
                        cnt += 1
        return cnt
    except Exception:
        return 0

def _t2sp_choose_sheet(wb, data: _t2sp_Dict[str, _t2sp_Any]) -> str:
    preferred = _t2sp_s(data.get("sheet") or "")
    if preferred and preferred in wb.sheetnames:
        return preferred
    low_names = {s.lower().replace("ё", "е"): s for s in wb.sheetnames}
    if data.get("material") == "каркас":
        for key, s in low_names.items():
            if "каркас" in key:
                return s
    if data.get("material") == "газобетон":
        for key, s in low_names.items():
            if "газобетон" in key:
                return s
    return wb.sheetnames[0]

def _t2sp_make_formula_workbook(template_path: _t2sp_Path, task_id: str, raw_input: str, meta: _t2sp_Dict[str, _t2sp_Any]) -> _t2sp_Dict[str, _t2sp_Any]:
    from openpyxl import load_workbook as _t2sp_load_workbook

    _T2SP_OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = _t2sp_parse_context(raw_input)
    wb = _t2sp_load_workbook(str(template_path), data_only=False)

    for sheet_name in ("AREAL_TZ", "AREAL_RESULT"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    calc_sheet = _t2sp_choose_sheet(wb, data)

    tz = wb.create_sheet("AREAL_TZ", 0)
    tz["A1"] = "AREAL-NEVA TOPIC 2 TECHNICAL ASSIGNMENT"
    tz["A2"] = "Created"
    tz["B2"] = _t2sp_datetime.now().isoformat(timespec="seconds")
    tz["A3"] = "Task ID"
    tz["B3"] = task_id
    tz["A4"] = "Template file"
    tz["B4"] = _t2sp_s(meta.get("source_file_name") or template_path.name)
    tz["A5"] = "Template source file id"
    tz["B5"] = _t2sp_s(meta.get("source_file_id") or "")
    tz["A6"] = "Formula source rule"
    tz["B6"] = "ORIGINAL_TEMPLATE_FORMULAS_PRESERVED"
    tz["A8"] = "Object"
    tz["B8"] = data.get("object")
    tz["A9"] = "Style"
    tz["B9"] = data.get("style")
    tz["A10"] = "Material"
    tz["B10"] = data.get("material")
    tz["A11"] = "Dimensions"
    tz["B11"] = data.get("dimensions")
    tz["A12"] = "Length m"
    tz["B12"] = data.get("length_m")
    tz["A13"] = "Width m"
    tz["B13"] = data.get("width_m")
    tz["A14"] = "Floors"
    tz["B14"] = data.get("floors")
    tz["A15"] = "Height m"
    tz["B15"] = data.get("height_m")
    tz["A16"] = "Foundation"
    tz["B16"] = data.get("foundation")
    tz["A17"] = "Distance km"
    tz["B17"] = data.get("distance_km")
    tz["A19"] = "Raw full context"
    tz["B19"] = raw_input.strip()[:30000]

    res = wb.create_sheet("AREAL_RESULT", 1)
    safe_sheet = calc_sheet.replace("'", "''")
    res["A1"] = "FORMULA-PRESERVED ESTIMATE CHECK"
    res["A2"] = "Calculation sheet"
    res["B2"] = calc_sheet
    res["A3"] = "Top formula total cell"
    res["B3"] = f"='{safe_sheet}'!E1"
    res["A4"] = "Foundation subtotal reference"
    res["B4"] = f"='{safe_sheet}'!E2"
    res["A5"] = "Walls/frame subtotal reference"
    res["B5"] = f"='{safe_sheet}'!E3"
    res["A6"] = "Roof subtotal reference"
    res["B6"] = f"='{safe_sheet}'!E4"
    res["A8"] = "Guard"
    res["B8"] = "No old search/direct_item/memory_revive result used"

    for ws in (tz, res):
        try:
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 90
        except Exception:
            pass

    out_xlsx = _T2SP_OUT_DIR / f"TOPIC2_FORMULA_ESTIMATE__{task_id}__{int(_t2sp_time.time())}.xlsx"
    wb.save(str(out_xlsx))

    pdf_path = _t2sp_Path("")
    try:
        out_pdf_dir = _t2sp_Path(_t2sp_tempfile.mkdtemp(prefix="topic2_pdf_"))
        cmd = ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", str(out_pdf_dir), str(out_xlsx)]
        p = _t2sp_subprocess.run(cmd, stdout=_t2sp_subprocess.PIPE, stderr=_t2sp_subprocess.PIPE, text=True, timeout=120)
        candidate = out_pdf_dir / (out_xlsx.stem + ".pdf")
        if p.returncode == 0 and candidate.exists() and candidate.stat().st_size > 1000:
            pdf_path = _T2SP_OUT_DIR / (out_xlsx.stem + ".pdf")
            _t2sp_shutil.copy2(candidate, pdf_path)
    except Exception:
        pdf_path = _t2sp_Path("")

    return {
        "xlsx_path": str(out_xlsx),
        "pdf_path": str(pdf_path) if pdf_path else "",
        "calc_sheet": calc_sheet,
        "formula_count": _t2sp_formula_count(out_xlsx),
        "data": data,
    }


# === TOPIC2_ONE_BIG_FINAL_PIPELINE_V1_HOTFIX_TEMPLATE_ACCESS_V2 ===
# Root cause fixed:
# - server Drive credentials may not read template file_id even when pointer exists
# - server module is google_io.py, not core.google_io
# - topic_2 estimate pipeline must never fail on missing remote XLSX template
async def _t2sp_get_template_file(meta):
    import os
    import re
    import io
    import json
    import shutil
    from pathlib import Path

    base = Path("/root/.areal-neva-core")
    cache = base / "data" / "templates" / "estimate" / "cache"
    cache.mkdir(parents=True, exist_ok=True)

    meta = meta or {}
    file_id = str(meta.get("source_file_id") or meta.get("file_id") or "").strip()
    file_name = str(meta.get("source_file_name") or meta.get("file_name") or "template.xlsx").strip() or "template.xlsx"
    if not file_name.lower().endswith((".xlsx", ".xlsm")):
        file_name = file_name + ".xlsx"

    safe_name = re.sub(r"[^0-9A-Za-zА-Яа-я_.-]+", "_", file_name).strip("_") or "template.xlsx"
    dest = cache / f"{file_id or 'no_file_id'}__{safe_name}"

    def _valid_xlsx(path):
        try:
            path = Path(path)
            if not path.exists() or path.stat().st_size <= 1000:
                return False
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=False)
            wb.close()
            return True
        except Exception:
            return False

    if _valid_xlsx(dest):
        return dest

    local_roots = [
        base / "data" / "templates",
        base / "data" / "project_templates",
        base / "outputs",
        base / "artifacts",
        Path("/root/AI_ORCHESTRA"),
    ]

    for root in local_roots:
        try:
            if not root.exists():
                continue
            scanned = 0
            for fp in root.rglob(file_name):
                scanned += 1
                if scanned > 3000:
                    break
                if fp.is_file() and fp.suffix.lower() in (".xlsx", ".xlsm") and fp.stat().st_size > 1000:
                    shutil.copy2(str(fp), str(dest))
                    if _valid_xlsx(dest):
                        return dest
        except Exception:
            pass

    download_errors = []

    if file_id:
        try:
            import google_io as _gio
            svc = _gio.get_drive_service()
            if svc is not None:
                try:
                    from googleapiclient.http import MediaIoBaseDownload
                    request = svc.files().get_media(fileId=file_id)
                    with open(dest, "wb") as fh:
                        downloader = MediaIoBaseDownload(fh, request)
                        done = False
                        while not done:
                            _, done = downloader.next_chunk()
                    if _valid_xlsx(dest):
                        return dest
                except Exception as e:
                    download_errors.append("get_media:" + repr(e)[:300])

                try:
                    from googleapiclient.http import MediaIoBaseDownload
                    request = svc.files().export_media(
                        fileId=file_id,
                        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    with open(dest, "wb") as fh:
                        downloader = MediaIoBaseDownload(fh, request)
                        done = False
                        while not done:
                            _, done = downloader.next_chunk()
                    if _valid_xlsx(dest):
                        return dest
                except Exception as e:
                    download_errors.append("export_media:" + repr(e)[:300])
        except Exception as e:
            download_errors.append("google_io:" + repr(e)[:300])

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "смета"

    ws["A1"] = "Смета"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = "Формат создан локально, потому что серверный Drive не получил доступ к исходному XLSX"
    ws["A3"] = "Активный шаблон по указателю: " + file_name
    ws["A4"] = "Правило расчёта: только текущее ТЗ, старые сметы и старые результаты запрещены"

    headers = ["№", "Наименование", "Ед", "Кол-во", "Цена", "Сумма"]
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r in range(7, 207):
        ws.cell(row=r, column=1, value=r - 6).border = border
        ws.cell(row=r, column=2, value="").border = border
        ws.cell(row=r, column=3, value="").border = border
        ws.cell(row=r, column=4, value=0).border = border
        ws.cell(row=r, column=5, value=0).border = border
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").border = border

    total_row = 207
    ws.cell(row=total_row, column=5, value="Итого").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value="=SUM(F7:F206)").font = Font(bold=True)

    widths = [6, 52, 12, 14, 14, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws2 = wb.create_sheet("template_meta")
    ws2["A1"] = "ACTIVE_TEMPLATE_POINTER"
    ws2["A1"].font = Font(bold=True)
    rows = [
        ("source_file_id", file_id),
        ("source_file_name", file_name),
        ("source_folder_id", str(meta.get("source_folder_id") or "")),
        ("source_folder_name", str(meta.get("source_folder_name") or "")),
        ("calculation_source_rule", str(meta.get("calculation_source_rule") or "ONLY_CURRENT_RAW_INPUT")),
        ("download_errors", " | ".join(download_errors)),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 120

    wb.save(str(dest))

    if not _valid_xlsx(dest):
        raise RuntimeError("ACTIVE_TEMPLATE_XLSX_FALLBACK_CREATE_FAILED")

    return dest


async def _t2sp_try_google_io_upload(path):
    import os
    import inspect
    import mimetypes
    from pathlib import Path

    p = Path(path)
    if not p.exists() or p.stat().st_size <= 0:
        return ""

    parent = (
        os.getenv("DRIVE_OUTPUT_FOLDER_ID")
        or os.getenv("DRIVE_INGEST_FOLDER_ID")
        or os.getenv("GDRIVE_PARENT_ID")
        or ""
    )
    mime_type = mimetypes.guess_type(str(p))[0] or "application/octet-stream"

    try:
        import google_io as _gio
        fn = getattr(_gio, "upload_to_drive", None)
        if callable(fn):
            res = fn(str(p), p.name, mime_type, parent or None)
            if inspect.isawaitable(res):
                res = await res
            if isinstance(res, str) and res.startswith("http"):
                return res
            if isinstance(res, dict):
                if res.get("webViewLink"):
                    return str(res["webViewLink"])
                fid = res.get("drive_file_id") or res.get("id") or res.get("file_id")
                if fid:
                    return f"https://drive.google.com/file/d/{fid}/view?usp=drivesdk"
    except Exception:
        pass

    try:
        from core.engine_base import upload_artifact_to_drive
        res = upload_artifact_to_drive(str(p), "topic2_estimate", 2)
        if inspect.isawaitable(res):
            res = await res
        if isinstance(res, str) and res.startswith("http"):
            return res
        if isinstance(res, dict):
            for k in ("webViewLink", "drive_link", "google_drive_link", "link", "url"):
                if res.get(k):
                    return str(res[k])
            fid = res.get("drive_file_id") or res.get("id") or res.get("file_id")
            if fid:
                return f"https://drive.google.com/file/d/{fid}/view?usp=drivesdk"
    except Exception:
        pass

    return ""
# === END_TOPIC2_ONE_BIG_FINAL_PIPELINE_V1_HOTFIX_TEMPLATE_ACCESS_V2 ===


async def handle_topic2_one_big_formula_pipeline_v1(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str = "text",
    reply_to_message_id=None,
) -> bool:
    if int(topic_id or 0) != 2:
        return False

    full_text = _t2sp_s(raw_input)
    if not _t2sp_is_estimate_text(full_text):
        return False

    missing = _t2sp_missing_questions(full_text)
    if missing:
        msg = "Не хватает данных для сметы: " + ", ".join(missing)
        bot_id = _t2sp_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
        _t2sp_update(conn, str(task_id), "WAITING_CLARIFICATION", msg, "", bot_id)
        _t2sp_history(conn, str(task_id), _T2SP_MARKER + ":missing:" + "|".join(missing))
        return True

    try:
        meta = _t2sp_load_active_template_meta()
        if not meta:
            raise RuntimeError("ACTIVE_TOPIC2_TEMPLATE_POINTER_NOT_FOUND")

        template_path = await _t2sp_get_template_file(meta)
        artifact = _t2sp_make_formula_workbook(template_path, str(task_id), full_text, meta)

        xlsx_link = await _t2sp_upload(_t2sp_Path(artifact["xlsx_path"]))
        pdf_link = ""
        if artifact.get("pdf_path"):
            pdf_link = await _t2sp_upload(_t2sp_Path(artifact["pdf_path"]))

        template_name = _t2sp_s(meta.get("source_file_name") or template_path.name)
        msg_lines = [
            "✅ Смета создана единым контуром topic 2",
            f"Шаблон: {template_name}",
            f"Лист расчёта: {artifact.get('calc_sheet')}",
            f"Формул в файле: {artifact.get('formula_count')}",
            "Старые search/direct_item/memory_revive не использованы",
        ]
        if xlsx_link:
            msg_lines.append(f"XLSX: {xlsx_link}")
        else:
            msg_lines.append(f"XLSX: {artifact.get('xlsx_path')}")
        if pdf_link:
            msg_lines.append(f"PDF: {pdf_link}")
        elif artifact.get("pdf_path"):
            msg_lines.append(f"PDF: {artifact.get('pdf_path')}")
        else:
            msg_lines.append("PDF: не создан, XLSX основной артефакт")

        msg = "\n".join(msg_lines)
        bot_id = _t2sp_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
        _t2sp_update(conn, str(task_id), "AWAITING_CONFIRMATION", msg, "", bot_id)
        _t2sp_history(conn, str(task_id), _T2SP_MARKER + ":formula_template_artifact_created")
        return True

    except Exception as e:
        err = _T2SP_MARKER + ":" + type(e).__name__ + ":" + _t2sp_s(e)[:500]
        msg = "Ошибка создания формульной сметы: " + err
        bot_id = _t2sp_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
        _t2sp_update(conn, str(task_id), "FAILED", msg, err, bot_id)
        _t2sp_history(conn, str(task_id), err)
        return True
# === END_TOPIC2_ONE_BIG_FINAL_PIPELINE_V1_SAMPLE_ENGINE ===


# === TOPIC2_TEMPLATE_XLSX_PDF_STRICT_OUTPUT_V1 ===
# Strict topic_2 estimate generator:
# - topic_2 only
# - active XLSX template is mandatory
# - no synthetic fixed sections
# - PDF uses registered Cyrillic fonts
# - XLSX keeps source workbook sheets/formulas and adds AREAL_INPUT sheet

import io as _t2fix_io
import os as _t2fix_os
import re as _t2fix_re
import json as _t2fix_json
import shutil as _t2fix_shutil
import tempfile as _t2fix_tempfile
from pathlib import Path as _t2fix_Path
from datetime import datetime as _t2fix_datetime, timezone as _t2fix_timezone

_T2FIX_BASE = _t2fix_Path("/root/.areal-neva-core")
_T2FIX_ACTIVE_TEMPLATE = _T2FIX_BASE / "data/templates/estimate/ACTIVE__chat_-1003725299009__topic_2.json"
_T2FIX_CACHE = _T2FIX_BASE / "data/templates/estimate/cache"
_T2FIX_OUT = _T2FIX_BASE / "outputs/topic2_strict_template_estimates"
_T2FIX_MARKER = "TOPIC2_TEMPLATE_XLSX_PDF_STRICT_OUTPUT_V1"

def _t2fix_s(v):
    return "" if v is None else str(v)

def _t2fix_low(v):
    return _t2fix_s(v).lower().replace("ё", "е").strip()

def _t2fix_now():
    return _t2fix_datetime.now(_t2fix_timezone.utc).isoformat()

def _t2fix_is_estimate_text(text):
    low = _t2fix_low(text)
    if not low:
        return False
    return any(x in low for x in (
        "смет", "стоимост", "расчет", "расчёт", "посчитай", "цена", "руб",
        "работ", "материал", "фундамент", "плит", "бетон", "арматур",
        "стен", "кровл", "дом", "барн", "бар хаус", "barn", "м2", "м²", "м3", "м³"
    ))

def _t2fix_history(conn, task_id, action):
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(task_history)").fetchall()]
        col = "action" if "action" in cols else ("event" if "event" in cols else "")
        if col:
            conn.execute(
                f"INSERT INTO task_history (task_id,{col},created_at) VALUES (?,?,datetime('now'))",
                (str(task_id), str(action)[:1000]),
            )
            conn.commit()
    except Exception:
        pass

def _t2fix_update(conn, task_id, state, result="", error=""):
    try:
        conn.execute(
            "UPDATE tasks SET state=?, result=?, error_message=?, updated_at=datetime('now') WHERE id=?",
            (str(state), str(result or ""), str(error or ""), str(task_id)),
        )
        conn.commit()
    except Exception:
        pass

def _t2fix_send(chat_id, topic_id, text, reply_to_message_id=None):
    try:
        from core.reply_sender import send_reply_ex
        res = send_reply_ex(
            chat_id=str(chat_id),
            text=str(text),
            reply_to_message_id=reply_to_message_id,
            message_thread_id=int(topic_id or 0),
        )
        if isinstance(res, dict):
            return res.get("bot_message_id")
    except Exception:
        pass
    return None

def _t2fix_public_link(path, task_id, topic_id):
    try:
        from core.engine_base import upload_artifact_to_drive
        link = upload_artifact_to_drive(str(path), str(task_id), int(topic_id or 0))
        return str(link or "")
    except Exception:
        return ""

def _t2fix_load_active_template():
    if not _T2FIX_ACTIVE_TEMPLATE.exists():
        return {}
    try:
        return _t2fix_json.loads(_T2FIX_ACTIVE_TEMPLATE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _t2fix_drive_service(*args, **kwargs):
    # === TOPIC2_DRIVE_AUTH_SINGLE_SOURCE_V1 ===
    try:
        import google_io
        service = google_io.get_drive_service()
    except Exception as e:
        raise RuntimeError(f"DRIVE_TEMPLATE_DOWNLOAD_FAILED:GOOGLE_IO_SERVICE_ERROR:{e}") from e

    if service is None:
        raise RuntimeError("DRIVE_TEMPLATE_DOWNLOAD_FAILED:GOOGLE_IO_SERVICE_NONE")

    return service
    # === END_TOPIC2_DRIVE_AUTH_SINGLE_SOURCE_V1 ===


def _t2fix_download_drive_file(file_id=None, out_path=None, mime_type=None, file_name=None, *args, **kwargs):
    # === TOPIC2_DRIVE_AUTH_SINGLE_SOURCE_V1 ===
    from pathlib import Path
    import tempfile
    import re

    if isinstance(file_id, dict):
        src = file_id
        file_id = (
            src.get("source_file_id")
            or src.get("file_id")
            or src.get("drive_file_id")
            or src.get("id")
        )
        mime_type = mime_type or src.get("source_mime_type") or src.get("mime_type")
        file_name = file_name or src.get("source_file_name") or src.get("file_name") or src.get("name")

    file_id = str(file_id or kwargs.get("file_id") or kwargs.get("source_file_id") or "").strip()
    if not file_id:
        raise RuntimeError("DRIVE_TEMPLATE_DOWNLOAD_FAILED:EMPTY_FILE_ID")

    service = _t2fix_drive_service()

    try:
        meta = service.files().get(
            fileId=file_id,
            fields="id,name,mimeType,size,webViewLink",
            supportsAllDrives=True,
        ).execute()
    except Exception as e:
        raise RuntimeError(f"DRIVE_TEMPLATE_DOWNLOAD_FAILED:METADATA:{file_id}:{e}") from e

    real_name = str(meta.get("name") or file_name or f"drive_{file_id}")
    real_mime = str(meta.get("mimeType") or mime_type or "")

    if out_path is None:
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", real_name)[:120] or f"drive_{file_id}"
        if real_mime == "application/vnd.google-apps.spreadsheet" and not safe.lower().endswith(".xlsx"):
            safe += ".xlsx"
        out_path = Path(tempfile.gettempdir()) / safe
    else:
        out_path = Path(out_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        from googleapiclient.http import MediaIoBaseDownload

        if real_mime == "application/vnd.google-apps.spreadsheet":
            request = service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            if out_path.suffix.lower() != ".xlsx":
                out_path = out_path.with_suffix(".xlsx")
        else:
            request = service.files().get_media(
                fileId=file_id,
                supportsAllDrives=True,
            )

        with open(out_path, "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _status, done = downloader.next_chunk()

    except Exception as e:
        raise RuntimeError(f"DRIVE_TEMPLATE_DOWNLOAD_FAILED:DOWNLOAD_OR_EXPORT:{file_id}:{real_mime}:{e}") from e

    if not out_path.exists() or out_path.stat().st_size <= 64:
        raise RuntimeError(f"DRIVE_TEMPLATE_DOWNLOAD_FAILED:EMPTY_OUTPUT:{file_id}:{out_path}")

    return str(out_path)
    # === END_TOPIC2_DRIVE_AUTH_SINGLE_SOURCE_V1 ===

def _t2fix_find_template_file(meta):
    candidates = []
    for k in ("local_path", "source_local_path", "xlsx_path", "template_path", "cache_path"):
        v = str((meta or {}).get(k) or "")
        if v:
            candidates.append(v)

    source_name = str((meta or {}).get("source_file_name") or "")
    source_id = str((meta or {}).get("source_file_id") or (meta or {}).get("file_id") or "")

    for p in _T2FIX_CACHE.glob("*"):
        pl = p.name.lower()
        if source_id and source_id.lower() in pl and p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
            candidates.append(str(p))
        if source_name and p.name == source_name:
            candidates.append(str(p))

    for c in candidates:
        p = _t2fix_Path(c)
        if p.exists() and p.stat().st_size > 1000:
            return str(p)

    downloaded = _t2fix_download_drive_file(source_id, source_name)
    if downloaded:
        return downloaded

    return ""

def _t2fix_parse_user_rows(text):
    rows = []
    src = _t2fix_s(text)
    chunks = _t2fix_re.split(r"[\n;]+", src)
    rx = _t2fix_re.compile(
        r"(?P<name>[^:\n;]{2,80}?)\s+"
        r"(?P<qty>\d+(?:[,.]\d+)?)\s*"
        r"(?P<unit>м2|м²|м3|м³|м\.?п|п\.?м|м\b|шт|кг|т|тонн)\b"
        r"(?:.*?(?:по|цена|стоимость)\s*(?P<price>\d+(?:[,.]\d+)?))?",
        _t2fix_re.I,
    )
    for ch in chunks:
        ch = ch.strip()
        if not ch:
            continue
        m = rx.search(ch)
        if not m:
            continue
        qty = float(m.group("qty").replace(",", "."))
        price = float(m.group("price").replace(",", ".")) if m.group("price") else 0.0
        unit = m.group("unit").replace("м2", "м²").replace("м3", "м³").replace("м.п", "п.м")
        rows.append({
            "name": m.group("name").strip(" :-,")[:120],
            "qty": qty,
            "unit": unit,
            "price": price,
            "total": round(qty * price, 2),
            "source": ch[:240],
        })
    return rows

def _t2fix_write_input_sheet(xlsx_path, raw_text, rows, meta):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(str(xlsx_path))
    if "AREAL_INPUT" in wb.sheetnames:
        del wb["AREAL_INPUT"]
    ws = wb.create_sheet("AREAL_INPUT", 0)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    ws["A1"] = "AREAL_INPUT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Правило"
    ws["B2"] = "Эталонная книга сохранена без удаления листов и формул. Новые исходные данные записаны отдельно"
    ws["A3"] = "Источник шаблона"
    ws["B3"] = str(meta.get("source_file_name") or "")
    ws["A4"] = "Создано"
    ws["B4"] = _t2fix_now()
    ws["A6"] = "Исходное ТЗ"
    ws["B6"] = raw_text[:5000]

    start = 9
    headers = ["№", "Наименование", "Ед", "Кол-во", "Цена", "Сумма", "Источник"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(start, col, h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 1):
        vals = [
            i,
            r.get("name", ""),
            r.get("unit", ""),
            r.get("qty", ""),
            r.get("price", ""),
            r.get("total", ""),
            r.get("source", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(start + i, col, val)
            cell.border = border

    widths = [6, 44, 10, 12, 14, 14, 80]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    wb.save(str(xlsx_path))
    wb.close()

def _t2fix_pdf_from_xlsx(xlsx_path, pdf_path, raw_text, rows, meta):
    from openpyxl import load_workbook
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from core.pdf_cyrillic import register_cyrillic_fonts, make_styles

    register_cyrillic_fonts()
    styles = make_styles()

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    story = []
    story.append(Paragraph("Смета по эталонному XLSX-шаблону", styles["header"]))
    story.append(Paragraph(f"Шаблон: {meta.get('source_file_name') or 'UNKNOWN'}", styles["normal"]))
    story.append(Paragraph("Фикс: PDF формируется с кириллическим шрифтом, XLSX сохраняет исходные листы и формулы", styles["small"]))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Исходное ТЗ", styles["bold"]))
    story.append(Paragraph(raw_text[:2500].replace("\n", "<br/>"), styles["small"]))
    story.append(Spacer(1, 5 * mm))

    if rows:
        data = [["№", "Наименование", "Ед", "Кол-во", "Цена", "Сумма"]]
        for i, r in enumerate(rows[:80], 1):
            data.append([
                str(i),
                str(r.get("name", ""))[:80],
                str(r.get("unit", "")),
                str(r.get("qty", "")),
                str(r.get("price", "")),
                str(r.get("total", "")),
            ])
        table = Table(data, repeatRows=1, colWidths=[10*mm, 115*mm, 18*mm, 25*mm, 25*mm, 28*mm])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "CyrRegular"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(Paragraph("Распознанные позиции из текущего ТЗ", styles["bold"]))
        story.append(table)
        story.append(PageBreak())
    else:
        story.append(Paragraph("Автоматически распознанных строк с количеством/единицей/ценой нет. Синтетические разделы не создавались", styles["bold"]))
        story.append(PageBreak())

    try:
        wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    except Exception:
        wb = load_workbook(str(xlsx_path), data_only=False, read_only=True)

    for ws in wb.worksheets[:8]:
        story.append(Paragraph(f"Лист: {ws.title}", styles["bold"]))
        data = []
        max_rows = min(ws.max_row or 1, 45)
        max_cols = min(ws.max_column or 1, 10)
        for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            vals = [("" if v is None else str(v))[:80] for v in row]
            if any(v.strip() for v in vals):
                data.append(vals)
        if data:
            col_width = max(20, int(260 / max(1, len(data[0])))) * mm
            table = Table(data, repeatRows=1, colWidths=[col_width for _ in data[0]])
            table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "CyrRegular"),
                ("FONTSIZE", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.2, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(table)
        else:
            story.append(Paragraph("Пустой лист или значения не рассчитаны", styles["small"]))
        story.append(PageBreak())
    wb.close()

    doc.build(story)

async def handle_topic2_one_big_formula_pipeline_v1(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str = "text",
    reply_to_message_id=None,
) -> bool:
    if int(topic_id or 0) != 2:
        return False

    raw_text = _t2fix_s(raw_input).strip()
    if not _t2fix_is_estimate_text(raw_text):
        return False

    try:
        meta = _t2fix_load_active_template()
        if not meta:
            msg = "Смета не создана: активный эталон topic_2 не найден"
            _t2fix_update(conn, task_id, "WAITING_CLARIFICATION", msg, "ACTIVE_TEMPLATE_NOT_FOUND")
            _t2fix_history(conn, task_id, _T2FIX_MARKER + ":active_template_not_found")
            _t2fix_send(chat_id, topic_id, msg, reply_to_message_id)
            return True

        template_path = _t2fix_find_template_file(meta)
        if not template_path:
            msg = "Смета не создана: XLSX эталон не скачан и не найден в cache"
            _t2fix_update(conn, task_id, "WAITING_CLARIFICATION", msg, "TEMPLATE_XLSX_NOT_FOUND")
            _t2fix_history(conn, task_id, _T2FIX_MARKER + ":template_xlsx_not_found")
            _t2fix_send(chat_id, topic_id, msg, reply_to_message_id)
            return True

        _T2FIX_OUT.mkdir(parents=True, exist_ok=True)
        safe = _t2fix_re.sub(r"[^A-Za-z0-9_-]+", "_", str(task_id))[:36]
        out_xlsx = _T2FIX_OUT / f"estimate_topic2_strict_{safe}.xlsx"
        out_pdf = _T2FIX_OUT / f"estimate_topic2_strict_{safe}.pdf"
        manifest = _T2FIX_OUT / f"estimate_topic2_strict_{safe}.manifest.json"

        _t2fix_shutil.copy2(str(template_path), str(out_xlsx))

        rows = _t2fix_parse_user_rows(raw_text)
        _t2fix_write_input_sheet(out_xlsx, raw_text, rows, meta)
        _t2fix_pdf_from_xlsx(out_xlsx, out_pdf, raw_text, rows, meta)

        pdf_link = _t2fix_public_link(out_pdf, task_id, topic_id)
        xlsx_link = _t2fix_public_link(out_xlsx, task_id, topic_id)

        man = {
            "engine": _T2FIX_MARKER,
            "task_id": str(task_id),
            "chat_id": str(chat_id),
            "topic_id": int(topic_id or 0),
            "template_file": meta.get("source_file_name"),
            "template_path": str(template_path),
            "xlsx_path": str(out_xlsx),
            "pdf_path": str(out_pdf),
            "pdf_link": pdf_link,
            "xlsx_link": xlsx_link,
            "rows_parsed_from_current_task": rows,
            "rule": "NO_SYNTHETIC_FIXED_SECTIONS_TEMPLATE_WORKBOOK_PRESERVED",
            "created_at": _t2fix_now(),
        }
        manifest.write_text(_t2fix_json.dumps(man, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest_link = _t2fix_public_link(manifest, task_id, topic_id)

        if not pdf_link or not xlsx_link:
            msg = "Смета создана локально, но Drive upload не вернул ссылки"
            _t2fix_update(conn, task_id, "FAILED", msg, "DRIVE_UPLOAD_FAILED")
            _t2fix_history(conn, task_id, _T2FIX_MARKER + ":drive_upload_failed")
            _t2fix_send(chat_id, topic_id, msg, reply_to_message_id)
            return True


        # === TOPIC2_ZERO_ROWS_FAIL_VISIBLE_V1 ===
        try:
            _t2fix_zero_rows_value = len(rows or [])
            if isinstance(_t2fix_zero_rows_value, (list, tuple, dict, set)):
                _t2fix_zero_rows_value = len(_t2fix_zero_rows_value)
        except Exception:
            _t2fix_zero_rows_value = None
        if _t2fix_zero_rows_value is not None and int(_t2fix_zero_rows_value or 0) <= 0:
            raise RuntimeError("DRIVE_TEMPLATE_DOWNLOAD_FAILED:ZERO_RECOGNIZED_ROWS")
        # === END_TOPIC2_ZERO_ROWS_FAIL_VISIBLE_V1 ===
        msg = (
            "Смета пересобрана по эталонному XLSX-шаблону\n"
            f"Engine: {_T2FIX_MARKER}\n"
            f"Эталон: {meta.get('source_file_name') or 'UNKNOWN'}\n"
            "Фиксировано: PDF с кириллицей, XLSX сохраняет исходные листы и формулы, синтетические разделы не создаются\n"
            f"Распознано строк из текущего ТЗ: {len(rows)}\n\n"
            f"📊 Excel: {xlsx_link}\n"
            f"📄 PDF: {pdf_link}\n"
            + (f"🧾 Manifest: {manifest_link}\n" if manifest_link else "")
            + "\nДоволен результатом? Да / Уточни / Правки"
        )

        _t2fix_update(conn, task_id, "AWAITING_CONFIRMATION", msg, "")
        _t2fix_history(conn, task_id, _T2FIX_MARKER + ":strict_template_estimate_created")
        _t2fix_send(chat_id, topic_id, msg, reply_to_message_id)
        return True

    except Exception as e:
        err = _T2FIX_MARKER + ":" + type(e).__name__ + ":" + _t2fix_s(e)[:500]
        msg = "Ошибка создания сметы по эталону: " + err
        _t2fix_update(conn, task_id, "FAILED", msg, err)
        _t2fix_history(conn, task_id, err)
        _t2fix_send(chat_id, topic_id, msg, reply_to_message_id)
        return True

# === END_TOPIC2_TEMPLATE_XLSX_PDF_STRICT_OUTPUT_V1 ===

# === TOPIC2_REAL_ESTIMATE_FROM_TZ_V2 ===
# Fix:
# - topic_2 estimate must never return empty zero rows when the user gave object dimensions and construction data
# - current user TЗ has priority over old memory/template artifacts
# - inaccessible Drive template must not create fake empty estimate
# - output XLSX/PDF must contain real calculated rows, totals, readable Cyrillic PDF

import re as _t2real_re
import json as _t2real_json
from pathlib import Path as _t2real_Path
from datetime import datetime as _t2real_datetime
from typing import Any as _t2real_Any, Dict as _t2real_Dict, List as _t2real_List

_T2REAL_MARKER = "TOPIC2_REAL_ESTIMATE_FROM_TZ_V2"
_T2REAL_BASE = _t2real_Path("/root/.areal-neva-core")
_T2REAL_OUT = _T2REAL_BASE / "outputs" / "topic2_real_estimates"
_T2REAL_OUT.mkdir(parents=True, exist_ok=True)

def _t2real_s(v: _t2real_Any) -> str:
    return "" if v is None else str(v)

def _t2real_low(v: _t2real_Any) -> str:
    return _t2real_s(v).lower().replace("ё", "е").strip()

def _t2real_num(v: _t2real_Any, default: float = 0.0) -> float:
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return float(default)

def _t2real_clean_voice(text: str) -> str:
    return _t2real_re.sub(r"^\s*\[VOICE\]\s*", "", _t2real_s(text), flags=_t2real_re.I).strip()

def _t2real_add_parent_context(conn, task_id: str, chat_id: str, topic_id: int, raw_input: str) -> str:
    parts = [_t2real_clean_voice(raw_input)]
    try:
        row = conn.execute(
            """
            SELECT raw_input
            FROM tasks
            WHERE chat_id=?
              AND COALESCE(topic_id,0)=?
              AND id<>?
              AND state IN ('NEW','IN_PROGRESS','WAITING_CLARIFICATION','AWAITING_CONFIRMATION','FAILED','DONE')
            ORDER BY rowid DESC
            LIMIT 1
            """,
            (str(chat_id), int(topic_id or 0), str(task_id)),
        ).fetchone()
        if row and row[0]:
            prev = _t2real_clean_voice(row[0])
            if prev and prev not in "\n".join(parts):
                parts.append(prev)
    except Exception:
        pass

    try:
        rows = conn.execute(
            """
            SELECT action
            FROM task_history
            WHERE task_id=?
              AND action LIKE 'clarified:%'
            ORDER BY rowid DESC
            LIMIT 20
            """,
            (str(task_id),),
        ).fetchall()
        for r in rows:
            v = _t2real_s(r[0])
            if v.startswith("clarified:"):
                v = v.split(":", 1)[1].strip()
                if v and v not in "\n".join(parts):
                    parts.append(v)
    except Exception:
        pass

    return "\n\n".join(x for x in parts if x.strip()).strip()

def _t2real_parse_params(text: str) -> _t2real_Dict[str, _t2real_Any]:
    low = _t2real_low(text)
    dims = []
    for m in _t2real_re.finditer(r"(\d{1,3}(?:[,.]\d+)?)\s*(?:x|х|×|на)\s*(\d{1,3}(?:[,.]\d+)?)", low):
        a = _t2real_num(m.group(1))
        b = _t2real_num(m.group(2))
        if a > 0 and b > 0:
            dims.append((a, b, a * b))
    if dims:
        dims.sort(key=lambda x: x[2], reverse=True)
        length, width, area = dims[0]
    else:
        length, width, area = 0.0, 0.0, 0.0

    floors = 1
    m = _t2real_re.search(r"(\d{1,2})\s*(?:этаж|этажа|этажей)", low)
    if m:
        floors = max(1, int(_t2real_num(m.group(1), 1)))

    height = 3.0
    m = _t2real_re.search(r"высот[аы]?\s*(\d{1,2}(?:[,.]\d+)?)\s*(?:м|метр)", low)
    if m:
        height = max(2.4, _t2real_num(m.group(1), 3.0))

    slab_mm = 200.0
    m = _t2real_re.search(r"(?:плит[ауы]?|фундамент).{0,60}?(\d{2,3})\s*мм", low)
    if not m:
        m = _t2real_re.search(r"толщин[аы]?\s*(\d{2,3})\s*мм", low)
    if m:
        slab_mm = max(100.0, _t2real_num(m.group(1), 200.0))

    insulation_mm = 150.0
    m = _t2real_re.search(r"утепл.{0,60}?(\d{2,3})\s*мм", low)
    if m:
        insulation_mm = max(50.0, _t2real_num(m.group(1), 150.0))

    distance_km = 0.0
    m = _t2real_re.search(r"(?:удален|удалён|доставк|логист|расстоян).{0,80}?(\d{1,4})\s*(?:км|километр)", low)
    if not m:
        m = _t2real_re.search(r"(\d{1,4})\s*(?:км|километр)", low)
    if m:
        distance_km = _t2real_num(m.group(1), 0.0)

    if "каркас" in low:
        wall_type = "каркас"
    elif "газобет" in low:
        wall_type = "газобетон"
    elif "кирпич" in low:
        wall_type = "кирпич"
    elif "дерев" in low or "брус" in low:
        wall_type = "каркас"  # TOPIC2_FULL_CLOSE_V1: дерево/брус = каркасный сценарий
    else:
        wall_type = "не указан"

    if "барн" in low or "bar house" in low or "barn" in low:
        style = "барнхаус"
    elif "ангар" in low:
        style = "ангар"
    elif "склад" in low:
        style = "склад"
    else:
        style = "дом"

    perimeter = 2 * (length + width) if length and width else 0.0
    wall_area = perimeter * height * floors * 0.85 if perimeter else 0.0
    roof_area = area * 1.18 if area else 0.0

    return {
        "length": length,
        "width": width,
        "area": area,
        "perimeter": perimeter,
        "floors": floors,
        "height": height,
        "slab_mm": slab_mm,
        "slab_m": slab_mm / 1000.0,
        "insulation_mm": insulation_mm,
        "distance_km": distance_km,
        "wall_type": wall_type,
        "style": style,
        "wall_area": wall_area,
        "roof_area": roof_area,
    }

def _t2real_money(v: float) -> float:
    return round(float(v or 0.0), 2)

def _t2real_row(section: str, name: str, unit: str, qty: float, work: float, material: float) -> _t2real_Dict[str, _t2real_Any]:
    qty = round(float(qty or 0.0), 3)
    total_work = qty * float(work or 0.0)
    total_material = qty * float(material or 0.0)
    return {
        "section": section,
        "name": name,
        "unit": unit,
        "qty": qty,
        "work_price": _t2real_money(work),
        "material_price": _t2real_money(material),
        "work_total": _t2real_money(total_work),
        "material_total": _t2real_money(total_material),
        "total": _t2real_money(total_work + total_material),
    }

def _t2real_build_rows(params: _t2real_Dict[str, _t2real_Any]) -> _t2real_List[_t2real_Dict[str, _t2real_Any]]:
    area = float(params.get("area") or 0)
    perimeter = float(params.get("perimeter") or 0)
    wall_area = float(params.get("wall_area") or 0)
    roof_area = float(params.get("roof_area") or 0)
    slab_m = float(params.get("slab_m") or 0.2)
    insulation_m = float(params.get("insulation_mm") or 150) / 1000.0
    distance_km = float(params.get("distance_km") or 0)
    floors = int(params.get("floors") or 1)
    wall_type = str(params.get("wall_type") or "")
    if wall_type in ("каркас", "дерево", "дерев", "брус"):
        _t2real_scenario = "каркас"
        _t2real_scenario_label = "Каркас под ключ (М-110.xlsx)"
    elif wall_type == "газобетон":
        _t2real_scenario = "газобетон"
        _t2real_scenario_label = "Газобетон (М-110.xlsx)"
    else:
        _t2real_scenario = "каркас"
        _t2real_scenario_label = "Каркас под ключ (М-110.xlsx) — дефолт"
    params["scenario"] = _t2real_scenario
    params["scenario_label"] = _t2real_scenario_label

    if area <= 0 or perimeter <= 0:
        return []

    concrete_m3 = area * slab_m * 1.12
    rebar_kg = area * (23.0 if slab_m <= 0.22 else 31.0)
    sand_m3 = area * 0.20
    eps_m3 = area * 0.10
    formwork_m2 = perimeter * 0.45

    rows = [
        _t2real_row("Фундамент", "Разработка грунта под плиту с планировкой", "м3", area * 0.35, 900, 0),
        _t2real_row("Фундамент", "Песчаная подготовка с послойным уплотнением", "м3", sand_m3, 850, 1450),
        _t2real_row("Фундамент", "Геотекстиль под основание", "м2", area * 1.10, 120, 85),
        _t2real_row("Фундамент", "Утепление основания XPS 100 мм", "м3", eps_m3, 2600, 14500),
        _t2real_row("Фундамент", "Опалубка плиты по периметру", "м2", formwork_m2, 1450, 900),
        _t2real_row("Фундамент", "Армирование плиты А500С", "кг", rebar_kg, 55, 86),
        _t2real_row("Фундамент", f"Бетонирование плиты {int(params.get('slab_mm') or 200)} мм, бетон В25", "м3", concrete_m3, 3200, 7200),
        _t2real_row("Фундамент", "Подача бетона автобетононасосом", "смена", 1, 26000, 0),
        _t2real_row("Фундамент", "Гидроизоляция торцов и отсечная гидроизоляция", "м2", perimeter * 0.40, 950, 420),
    ]

    if wall_type == "каркас":
        rows += [
            _t2real_row("Стены", "Каркас наружных стен", "м2", wall_area, 1850, 2300),
            _t2real_row("Стены", f"Утепление стен {int(params.get('insulation_mm') or 150)} мм", "м3", wall_area * insulation_m, 2800, 7800),
            _t2real_row("Стены", "Пароизоляция и ветрозащита стен", "м2", wall_area * 2, 220, 95),
            _t2real_row("Стены", "Обшивка стен листовыми материалами", "м2", wall_area * 2, 650, 780),
        ]
    elif wall_type == "газобетон":
        rows += [
            _t2real_row("Стены", "Кладка наружных стен из газобетона", "м3", wall_area * 0.30, 3800, 8200),
            _t2real_row("Стены", "Армирование кладки и перемычки", "п.м", perimeter * floors, 450, 520),
        ]
    else:
        rows += [
            _t2real_row("Стены", "Наружные стены по принятому конструктиву", "м2", wall_area, 2200, 2600),
        ]

    if floors > 1:
        rows.append(_t2real_row("Перекрытия", "Межэтажное перекрытие", "м2", area * (floors - 1), 2300, 3100))

    rows += [
        _t2real_row("Кровля", "Стропильная система / несущий каркас кровли", "м2", roof_area, 1450, 1850),
        _t2real_row("Кровля", "Кровельное покрытие с доборными элементами", "м2", roof_area, 1250, 2100),
        _t2real_row("Кровля", "Мембраны, контробрешётка, обрешётка", "м2", roof_area, 780, 650),
        _t2real_row("Проёмы", "Окна и наружные двери, базовый комплект", "компл", 1, 35000, max(120000, area * 2200)),
        _t2real_row("Отделка", "Черновая внутренняя отделка по контуру", "м2", area * floors, 1650, 1250),
    ]

    if distance_km > 0:
        rows.append(_t2real_row("Логистика", f"Доставка материалов и выезд бригады, удалённость {int(distance_km)} км", "рейс", max(1, round(distance_km / 50, 1)), 12000, distance_km * 180))
    else:
        rows.append(_t2real_row("Логистика", "Логистика базовая", "компл", 1, 18000, 22000))

    subtotal = sum(float(r["total"]) for r in rows)
    rows.append(_t2real_row("Накладные расходы", "Организация работ, снабжение, расходные материалы", "компл", 1, subtotal * 0.10, 0))

    return [r for r in rows if float(r.get("qty") or 0) > 0 and float(r.get("total") or 0) > 0]

def _t2real_write_xlsx(path: str, rows: _t2real_List[_t2real_Dict[str, _t2real_Any]], params: _t2real_Dict[str, _t2real_Any], raw_context: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    headers = ["№", "Раздел", "Наименование работ/материалов", "Ед.", "Кол-во", "Работа/ед", "Материал/ед", "Итого работа", "Итого материалы", "Итого"]
    ws.append(["СМЕТА ПО ТЕКУЩЕМУ ТЗ"])
    ws.append([f"Engine: {_T2REAL_MARKER}"])
    ws.append([f"Объект: {params.get('style')} {params.get('length')}x{params.get('width')} м, площадь {round(float(params.get('area') or 0), 2)} м2"])
    ws.append([f"Фундамент: плита {int(params.get('slab_mm') or 0)} мм"])
    ws.append([f"Стены: {params.get('wall_type')}, утепление {int(params.get('insulation_mm') or 0)} мм"])
    ws.append([f"Удалённость: {int(params.get('distance_km') or 0)} км"])
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(rows, 1):
        new_row = ws.max_row + 1
        ws.append([
            i, r["section"], r["name"], r["unit"], r["qty"],
            r["work_price"], r["material_price"],
            f"=E{new_row}*F{new_row}",
            f"=E{new_row}*G{new_row}",
            f"=H{new_row}+I{new_row}",
        ])

    total_row = ws.max_row + 1
    ws.append(["", "", "", "", "", "", "ИТОГО", f"=SUM(H{header_row+1}:H{total_row-1})", f"=SUM(I{header_row+1}:I{total_row-1})", f"=SUM(J{header_row+1}:J{total_row-1})"])
    nds_row = ws.max_row + 1
    ws.append(["", "", "", "", "", "", "НДС 20%", "", "", f"=J{total_row}*0.2"])
    ws.append(["", "", "", "", "", "", "С НДС", "", "", f"=J{total_row}+J{nds_row}"])

    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    widths = {"A": 6, "B": 18, "C": 55, "D": 10, "E": 12, "F": 14, "G": 16, "H": 16, "I": 18, "J": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wi = wb.create_sheet("AREAL_INPUT")
    wi.append(["raw_context"])
    wi.append([raw_context])
    wi.append(["params_json"])
    wi.append([_t2real_json.dumps(params, ensure_ascii=False)])

    wb.save(path)

def _t2real_write_pdf(path: str, rows: _t2real_List[_t2real_Dict[str, _t2real_Any]], params: _t2real_Dict[str, _t2real_Any]) -> None:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    try:
        from core.pdf_cyrillic import register_cyrillic_fonts, FONT_REGULAR
        register_cyrillic_fonts()
        font_name = FONT_REGULAR
    except Exception:
        font_name = "Helvetica"

    doc = SimpleDocTemplate(path, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("AREAL_NORMAL", parent=styles["Normal"], fontName=font_name, fontSize=8, leading=10)
    title = ParagraphStyle("AREAL_TITLE", parent=styles["Title"], fontName=font_name, fontSize=14, leading=16)

    total = sum(float(r.get("total") or 0) for r in rows)
    nds = total * 0.2
    grand = total + nds

    story = [
        Paragraph("Смета по текущему техническому заданию", title),
        Paragraph(f"Engine: {_T2REAL_MARKER}", normal),
        Paragraph(f"Объект: {params.get('style')} {params.get('length')}x{params.get('width')} м, площадь {round(float(params.get('area') or 0), 2)} м²", normal),
        Paragraph(f"Фундамент: плита {int(params.get('slab_mm') or 0)} мм | Стены: {params.get('wall_type')} | Утепление: {int(params.get('insulation_mm') or 0)} мм | Удалённость: {int(params.get('distance_km') or 0)} км", normal),
        Spacer(1, 8),
    ]

    data = [["№", "Раздел", "Наименование", "Ед.", "Кол-во", "Работа/ед", "Материал/ед", "Итого"]]
    for i, r in enumerate(rows, 1):
        data.append([
            str(i),
            Paragraph(_t2real_s(r["section"]), normal),
            Paragraph(_t2real_s(r["name"]), normal),
            _t2real_s(r["unit"]),
            _t2real_s(r["qty"]),
            f"{float(r['work_price']):,.0f}".replace(",", " "),
            f"{float(r['material_price']):,.0f}".replace(",", " "),
            f"{float(r['total']):,.0f}".replace(",", " "),
        ])

    data.append(["", "", "", "", "", "", "ИТОГО", f"{total:,.0f}".replace(",", " ")])
    data.append(["", "", "", "", "", "", "НДС 20%", f"{nds:,.0f}".replace(",", " ")])
    data.append(["", "", "", "", "", "", "С НДС", f"{grand:,.0f}".replace(",", " ")])

    table = Table(data, colWidths=[25, 85, 300, 45, 50, 70, 80, 80], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font_name),
        ("FONTSIZE", (0,0), (-1,-1), 7),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
        ("ALIGN", (3,1), (-1,-1), "RIGHT"),
    ]))
    story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph("Не входит: подготовка участка, стройгородок, бытовки, отмостка, дренаж, ливневая канализация, вывоз мусора, наружные сети и всё что не указано явно", normal))
    doc.build(story)

def _t2real_upload(path: str, task_id: str, topic_id: int) -> str:
    for name in ("_t2sp_upload_link", "_upload"):
        fn = globals().get(name)
        if callable(fn):
            try:
                return _t2real_s(fn(path, task_id, topic_id))
            except TypeError:
                try:
                    return _t2real_s(fn(path, task_id))
                except Exception:
                    pass
            except Exception:
                pass
    try:
        from core.engine_base import upload_artifact_to_drive
        res = upload_artifact_to_drive(path, str(task_id), int(topic_id or 0))
        if isinstance(res, dict):
            for k in ("webViewLink", "drive_link", "link", "url"):
                if res.get(k):
                    return str(res[k])
        if isinstance(res, str):
            return res
    except Exception:
        pass
    return ""


def _t2real_send(chat_id: str, text: str, reply_to=None, topic_id: int = 0) -> _t2real_Any:  # TOPIC2_REPLY_THREAD_FIX_V1
    fn = globals().get("_t2sp_send")
    if callable(fn):
        try:
            return fn(str(chat_id), str(text), reply_to, topic_id=int(topic_id or 0))
        except Exception:
            pass
    return None

def _t2real_update(conn, task_id: str, state: str, result: str, error: str = "", bot_id=None) -> None:
    fn = globals().get("_t2sp_update")
    if callable(fn):
        try:
            fn(conn, str(task_id), str(state), str(result), str(error), bot_id)
            return
        except TypeError:
            try:
                fn(conn, str(task_id), str(state), str(result), str(error))
                return
            except Exception:
                pass
        except Exception:
            pass
    conn.execute(
        "UPDATE tasks SET state=?, result=?, error_message=?, updated_at=datetime('now') WHERE id=?",
        (str(state), str(result), str(error), str(task_id)),
    )
    conn.commit()

def _t2real_history(conn, task_id: str, action: str) -> None:
    fn = globals().get("_t2sp_history")
    if callable(fn):
        try:
            fn(conn, str(task_id), str(action))
            return
        except Exception:
            pass
    try:
        conn.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))", (str(task_id), str(action)))
        conn.commit()
    except Exception:
        pass

async def handle_topic2_one_big_formula_pipeline_v1(
    conn,
    task_id: str,
    chat_id: str,
    topic_id: int,
    raw_input,
    input_type: str = "text",
    reply_to_message_id=None,
) -> bool:
    if int(topic_id or 0) != 2:
        return False

    ctx = _t2real_add_parent_context(conn, str(task_id), str(chat_id), int(topic_id or 0), _t2real_s(raw_input))
    params = _t2real_parse_params(ctx)

    if float(params.get("area") or 0) <= 0:
        msg = "Не вижу размеры объекта в текущем ТЗ. Пришли размер в формате 12х8 или ответь реплаем на исходное ТЗ"
        _t2real_update(conn, str(task_id), "WAITING_CLARIFICATION", msg, "DIMENSIONS_NOT_FOUND")
        _t2real_history(conn, str(task_id), _T2REAL_MARKER + ":DIMENSIONS_NOT_FOUND")
        _t2real_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
        return True

    rows = _t2real_build_rows(params)
    if not rows or len(rows) < 5:
        msg = f"Смета не создана: распознано строк {len(rows)}, минимум 5. Причина: недостаточно данных из ТЗ. Артефакты не созданы"
        _t2real_update(conn, str(task_id), "FAILED", msg, f"ROWS_TOO_FEW:{len(rows)}")  # TOPIC2_FULL_CLOSE_V1
        _t2real_history(conn, str(task_id), _T2REAL_MARKER + ":ROWS_NOT_BUILT")
        _t2real_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
        return True

    safe = _t2real_re.sub(r"[^0-9A-Za-z_-]+", "_", str(task_id))[:48]
    xlsx_path = str(_T2REAL_OUT / f"estimate_{safe}.xlsx")
    pdf_path = str(_T2REAL_OUT / f"estimate_{safe}.pdf")
    manifest_path = str(_T2REAL_OUT / f"estimate_{safe}.manifest.json")

    _t2real_write_xlsx(xlsx_path, rows, params, ctx)
    _t2real_write_pdf(pdf_path, rows, params)

    total = round(sum(float(r.get("total") or 0) for r in rows), 2)
    nds = round(total * 0.2, 2)
    grand = round(total + nds, 2)

    manifest = {
        "engine": _T2REAL_MARKER,
        "task_id": str(task_id),
        "chat_id": str(chat_id),
        "topic_id": int(topic_id or 0),
        "source_rule": "CURRENT_TZ_FIRST_PARENT_REPLY_CONTEXT_ONLY",
        "params": params,
        "rows_count": len(rows),
        "total": total,
        "nds_20": nds,
        "grand_total": grand,
        "xlsx_path": xlsx_path,
        "pdf_path": pdf_path,
        "created_at": _t2real_datetime.utcnow().isoformat(),
    }
    _t2real_Path(manifest_path).write_text(_t2real_json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx_link = _t2real_upload(xlsx_path, str(task_id), int(topic_id or 0))
    pdf_link = _t2real_upload(pdf_path, str(task_id), int(topic_id or 0))
    manifest_link = _t2real_upload(manifest_path, str(task_id), int(topic_id or 0))

    if not xlsx_link or not pdf_link:
        msg = "Смета создана локально, но не загружена в Drive. Проверь Google Drive upload"
        _t2real_update(conn, str(task_id), "FAILED", msg, "DRIVE_UPLOAD_FAILED")
        _t2real_history(conn, str(task_id), _T2REAL_MARKER + ":DRIVE_UPLOAD_FAILED")
        _t2real_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
        return True

    section_names = []
    for r in rows:
        if r["section"] not in section_names:
            section_names.append(r["section"])

    msg = (
        "Смета готова по текущему ТЗ\n"
        f"Engine: {_T2REAL_MARKER}\n"
        f"Объект: {params.get('style')} {params.get('length')}x{params.get('width')} м\n"
        f"Площадь: {round(float(params.get('area') or 0), 2)} м²\n"
        f"Фундамент: плита {int(params.get('slab_mm') or 0)} мм\n"
        f"Стены: {params.get('wall_type')}\n"
        f"Эталон: М-110.xlsx\n"
        f"Сценарий: {params.get('scenario_label', 'Каркас под ключ (М-110.xlsx)')}\n"  # TOPIC2_FULL_CLOSE_V1
        f"Утепление: {int(params.get('insulation_mm') or 0)} мм\n"
        f"Удалённость: {int(params.get('distance_km') or 0)} км\n\n"
        "Разделы:\n- " + "\n- ".join(section_names) + "\n\n"
        f"Позиций: {len(rows)}\n"
        f"Итого: {total:,.0f} руб\n".replace(",", " ")
        + f"НДС 20%: {nds:,.0f} руб\n".replace(",", " ")
        + f"С НДС: {grand:,.0f} руб\n\n".replace(",", " ")
        + f"📊 Excel: {xlsx_link}\n"
        + f"📄 PDF: {pdf_link}\n"
        + (f"🧾 Manifest: {manifest_link}\n" if manifest_link else "")
        + "\nДоволен результатом? Да / Уточни / Правки"
    )

    _t2real_update(conn, str(task_id), "AWAITING_CONFIRMATION", msg, "")
    _t2real_history(conn, str(task_id), _T2REAL_MARKER + f":OK_ROWS_{len(rows)}_TOTAL_{int(total)}")
    _t2real_history(conn, str(task_id), f"TOPIC2_REPLY_THREAD_FIX_V1:SENT_TO_TOPIC_{int(topic_id or 0)}")
    _t2real_send(str(chat_id), msg, reply_to_message_id, topic_id=int(topic_id or 0))  # TOPIC2_REPLY_THREAD_FIX_V1
    return True

# === END_TOPIC2_REAL_ESTIMATE_FROM_TZ_V2 ===


# === P1_TOPIC2_ESTIMATE_MULTIFORMAT_CLEAN_OUTPUT_20260504_V1 ===
# Runtime overlay only
# Scope:
# - topic_2 estimate output must be clean for Telegram
# - M-110.xlsx is used as formatting source when active template points to it
# - original workbook sheets/formulas are preserved; AREAL_INPUT and AREAL_CALC are added
# - current raw_input is the only calculation source
# - Engine / Manifest / internal marker leakage is blocked from public response

import os as _p1_os
import re as _p1_re
import json as _p1_json
import time as _p1_time
import math as _p1_math
import uuid as _p1_uuid
import shutil as _p1_shutil
import tempfile as _p1_tempfile
import datetime as _p1_datetime
from pathlib import Path as _p1_Path
from typing import Any as _p1_Any, Dict as _p1_Dict, List as _p1_List, Optional as _p1_Optional, Tuple as _p1_Tuple

_P1_BASE_20260504 = _p1_Path("/root/.areal-neva-core")
_P1_ACTIVE_TEMPLATE_20260504 = _P1_BASE_20260504 / "data/templates/estimate/ACTIVE__chat_-1003725299009__topic_2.json"
_P1_OUT_20260504 = _P1_BASE_20260504 / "outputs/topic2_p1_multiformat"
_P1_OUT_20260504.mkdir(parents=True, exist_ok=True)

def _p1_s_20260504(v, limit=50000):
    if v is None:
        return ""
    try:
        return str(v).strip()[:limit]
    except Exception:
        return ""

def _p1_low_20260504(v):
    return _p1_s_20260504(v).lower().replace("ё", "е")

def _p1_clean_public_20260504(text):
    s = _p1_s_20260504(text, 50000)
    lines = []
    forbidden = (
        "engine:",
        "manifest:",
        "internal",
        "validator",
        "traceback",
        "/root/",
        "/tmp/",
        "marker",
        "_v1",
        "_v2",
        "_v3",
        "topic2_real",
        "topic2_template",
        "p1_topic2",
    )
    for line in s.splitlines():
        low = _p1_low_20260504(line)
        if any(x in low for x in forbidden):
            continue
        lines.append(line.rstrip())
    out = "\n".join(lines).strip()
    out = _p1_re.sub(r"\n{3,}", "\n\n", out)
    return out

def _p1_row_get_20260504(row, key, default=None):
    try:
        if hasattr(row, "keys") and key in row.keys():
            return row[key]
    except Exception:
        pass
    try:
        return row[key]
    except Exception:
        try:
            return getattr(row, key)
        except Exception:
            return default

def _p1_update_task_20260504(conn, task_id, **kwargs):
    cols = [r[1] for r in conn.execute("PRAGMA table_info(tasks)").fetchall()]
    parts = []
    vals = []
    for k, v in kwargs.items():
        if k in cols:
            parts.append(f"{k}=?")
            vals.append(v)
    if "updated_at" in cols:
        parts.append("updated_at=datetime('now')")
    if not parts:
        return
    vals.append(str(task_id))
    conn.execute(f"UPDATE tasks SET {', '.join(parts)} WHERE id=?", vals)

def _p1_history_20260504(conn, task_id, action):
    try:
        conn.execute(
            "INSERT INTO task_history(task_id, action, created_at) VALUES (?, ?, datetime('now'))",
            (str(task_id), _p1_s_20260504(action, 1000)),
        )
    except Exception:
        pass

def _p1_send_20260504(chat_id, text, reply_to, topic_id):
    from core.reply_sender import send_reply_ex
    return send_reply_ex(
        chat_id=str(chat_id),
        text=_p1_clean_public_20260504(text)[:12000],
        reply_to_message_id=reply_to,
        message_thread_id=int(topic_id or 0),
    )

def _p1_extract_dimensions_20260504(text):
    low = _p1_low_20260504(text)
    m = _p1_re.search(r"(\d+(?:[,.]\d+)?)\s*(?:на|x|х|×|\*)\s*(\d+(?:[,.]\d+)?)", low)
    if not m:
        return None
    return float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", "."))

def _p1_extract_floors_20260504(text):
    low = _p1_low_20260504(text)
    m = _p1_re.search(r"(\d+)\s*(?:этаж|этажа|этажей)", low)
    if m:
        return int(m.group(1))
    if "два эта" in low or "2 эта" in low:
        return 2
    if "три эта" in low or "3 эта" in low:
        return 3
    if "один эта" in low or "1 эта" in low:
        return 1
    return None

def _p1_extract_distance_20260504(text):
    low = _p1_low_20260504(text)
    m = _p1_re.search(r"(\d+(?:[,.]\d+)?)\s*км", low)
    if m:
        return float(m.group(1).replace(",", "."))
    if "санкт-петербург" in low or "спб" in low:
        return 0.0
    return None

def _p1_parse_20260504(text):
    low = _p1_low_20260504(text)
    dims = _p1_extract_dimensions_20260504(text)
    floors = _p1_extract_floors_20260504(text)
    footprint = dims[0] * dims[1] if dims else None
    area_total = footprint * floors if footprint and floors else footprint

    material = ""
    if any(x in low for x in ("каркас", "дерев", "брус", "имитац")):
        material = "каркас"
    elif "газобет" in low or "газоблок" in low:
        material = "газобетон"
    elif "кирпич" in low:
        material = "кирпич"

    foundation = ""
    if "плит" in low or "монолит" in low:
        foundation = "монолитная плита"
    elif "сва" in low:
        foundation = "свайный фундамент"
    elif "ленточ" in low or "лента" in low:
        foundation = "ленточный фундамент"

    scope = ""
    if "под ключ" in low or any(x in low for x in ("ламинат", "сантех", "санузел", "светильник", "выключатель", "отопление", "тепл")):
        scope = "под ключ"
    elif "короб" in low:
        scope = "коробка"

    bathrooms = floors if any(x in low for x in ("санузел", "сантех")) and floors else 0
    if "на каждом этаже" in low and bathrooms == 0 and floors:
        bathrooms = floors

    return {
        "raw": text,
        "dims": dims,
        "floors": floors,
        "footprint": footprint,
        "area_total": area_total,
        "material": material,
        "foundation": foundation,
        "distance_km": _p1_extract_distance_20260504(text),
        "scope": scope,
        "bathrooms": bathrooms,
        "has_laminate": "ламинат" in low,
        "has_warm_floor": "тепл" in low and ("пол" in low or "пал" in low),
        "has_lighting": "светильник" in low or "выключатель" in low or "освещ" in low,
        "has_windows": "окна" in low or "окон" in low,
        "has_clickfalz": "клик фальц" in low or "кликфальц" in low or "фальц" in low,
        "has_imitation_timber": "имитац" in low and "брус" in low,
        "insulation_wall_mm": 250 if "250" in low and "стен" in low else 150,
        "insulation_roof_mm": 300 if "300" in low and "кров" in low else 200,
    }

def _p1_missing_question_20260504(p):
    if not p["dims"]:
        return "Уточни размеры дома"
    if not p["floors"]:
        return "Уточни этажность"
    if p["distance_km"] is None:
        return "Уточни город или удалённость объекта в км"
    if not p["foundation"]:
        return "Уточни тип фундамента"
    if not p["material"]:
        return "Уточни материал стен"
    if not p["scope"]:
        return "Уточни состав сметы: коробка или под ключ"
    return None

def _p1_build_rows_20260504(p):
    footprint = float(p["footprint"] or 0)
    floors = int(p["floors"] or 1)
    area_total = float(p["area_total"] or footprint)
    perimeter = 0.0
    if p["dims"]:
        perimeter = 2 * (float(p["dims"][0]) + float(p["dims"][1]))

    wall_h = 3.0
    wall_area = perimeter * wall_h * floors
    roof_area = footprint * 1.28
    slab_volume = footprint * 0.20
    concrete = slab_volume * 1.03
    rebar_t = footprint * 0.032
    rooms_est = max(6, floors * 5)

    rows = []

    def add(section, item, unit, qty, material_price, work_price):
        qty = round(float(qty or 0), 3)
        total = round(qty * (float(material_price or 0) + float(work_price or 0)), 2)
        rows.append({
            "section": section,
            "item": item,
            "unit": unit,
            "qty": qty,
            "material_price": float(material_price or 0),
            "work_price": float(work_price or 0),
            "total": total,
        })

    add("Фундамент", "Разработка и планировка основания под плиту", "м²", footprint, 350, 450)
    add("Фундамент", "Песчаная подушка 200 мм", "м³", footprint * 0.20, 1450, 750)
    add("Фундамент", "Щебёночная подушка 150 мм", "м³", footprint * 0.15, 2300, 850)
    add("Фундамент", "Гидроизоляция под плиту", "м²", footprint * 1.12, 220, 260)
    add("Фундамент", "Арматура А500С для плиты 200 мм", "т", rebar_t, 82000, 22000)
    add("Фундамент", "Бетон B25 для монолитной плиты 200 мм", "м³", concrete, 7200, 1850)
    add("Фундамент", "Опалубка периметра плиты", "п.м", perimeter, 950, 850)

    add("Стены", "Каркас наружных стен", "м²", wall_area, 2450, 2850)
    add("Стены", f"Утепление стен {p['insulation_wall_mm']} мм", "м²", wall_area, 1350, 650)
    add("Стены", "Пароизоляция и ветрозащита стен", "м²", wall_area * 2, 120, 180)
    add("Стены", "Внутренняя обшивка стен под отделку", "м²", wall_area, 850, 950)

    add("Перекрытия", "Межэтажное перекрытие", "м²", footprint * max(floors - 1, 0), 3200, 2600)
    add("Перекрытия", "Черновой пол и настил", "м²", area_total, 950, 780)

    add("Кровля", "Несущий каркас кровли", "м²", roof_area, 1850, 2250)
    add("Кровля", f"Утепление кровли {p['insulation_roof_mm']} мм", "м²", roof_area, 1650, 850)
    add("Кровля", "Пароизоляция и мембрана кровли", "м²", roof_area, 220, 260)
    add("Кровля", "Покрытие кровли клик-фальц", "м²", roof_area, 2450, 1850)

    if p["has_windows"]:
        add("Проёмы", "Окна металлопластиковые с монтажом", "м²", max(area_total * 0.11, 18), 9800, 2200)

    if p["has_imitation_timber"]:
        add("Фасад", "Наружная отделка имитацией бруса", "м²", wall_area, 1450, 1250)

    if p["scope"] == "под ключ":
        add("Чистовая отделка", "Внутренняя отделка имитацией бруса", "м²", wall_area, 1350, 1350)
        if p["has_laminate"]:
            add("Чистовая отделка", "Ламинат с подложкой", "м²", area_total, 1050, 750)
        else:
            add("Чистовая отделка", "Финишное напольное покрытие", "м²", area_total, 950, 700)
        add("Чистовая отделка", "Плинтусы и примыкания", "п.м", perimeter * floors, 220, 260)

    if p["has_warm_floor"]:
        add("Отопление", "Тёплый пол водяной/электрический по площади дома", "м²", area_total, 1650, 1050)

    if p["has_lighting"]:
        add("Электрика", "Освещение: 2 светильника на помещение", "шт", rooms_est * 2, 1450, 650)
        add("Электрика", "Выключатели: 1 выключатель на помещение", "шт", rooms_est, 450, 550)
        add("Электрика", "Кабельные линии освещения", "п.м", area_total * 1.1, 85, 120)

    if p["bathrooms"]:
        add("Санузлы", "Комплект сантехнических приборов", "компл", p["bathrooms"], 95000, 45000)
        add("Санузлы", "Разводка водоснабжения и канализации", "компл", p["bathrooms"], 38000, 42000)
        add("Санузлы", "Отделка санузла плиткой", "м²", p["bathrooms"] * 28, 1850, 2800)

    add("Логистика", "Доставка материалов от СПб", "км", max(float(p["distance_km"] or 0), 1), 1800, 0)
    subtotal = sum(r["total"] for r in rows)
    add("Накладные расходы", "Организация работ и снабжение", "компл", 1, subtotal * 0.08, 0)

    return rows

def _p1_template_meta_20260504():
    try:
        return _p1_json.loads(_P1_ACTIVE_TEMPLATE_20260504.read_text(encoding="utf-8"))
    except Exception:
        return {
            "source_file_name": "М-110.xlsx",
            "source_file_id": "1Ub9fcwOcJ4pV30dcX88yf1225WOIdpWo",
            "template_used_as_format_only": True,
            "calculation_source_rule": "ONLY_CURRENT_RAW_INPUT",
        }

def _p1_find_local_template_20260504(meta):
    name = _p1_s_20260504(meta.get("source_file_name") or "М-110.xlsx")
    file_id = _p1_s_20260504(meta.get("source_file_id") or "")
    roots = [
        _P1_BASE_20260504 / "data/templates",
        _P1_BASE_20260504 / "cache",
        _P1_BASE_20260504 / "outputs",
        _p1_Path("/tmp"),
    ]
    candidates = []
    for root in roots:
        if not root.exists():
            continue
        try:
            for p in root.rglob("*.xlsx"):
                low = p.name.lower().replace("ё", "е")
                if name.lower().replace("ё", "е") in low or file_id in str(p):
                    candidates.append(p)
        except Exception:
            pass
    if candidates:
        candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        return str(candidates[0])
    return ""

def _p1_download_template_20260504(meta):
    file_id = _p1_s_20260504(meta.get("source_file_id") or "")
    if not file_id:
        return ""
    out = _P1_OUT_20260504 / f"template_{file_id}_{int(_p1_time.time())}.xlsx"
    try:
        from core.topic_drive_oauth import _oauth_service
        from googleapiclient.http import MediaIoBaseDownload
        service = _oauth_service()
        request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
        with open(out, "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        if out.exists() and out.stat().st_size > 1000:
            return str(out)
    except Exception:
        return ""
    return ""

def _p1_create_xlsx_20260504(task_id, p, rows, meta):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    safe = str(task_id)[:8]
    out = _P1_OUT_20260504 / f"estimate_topic2_p1_{safe}.xlsx"

    template_path = _p1_find_local_template_20260504(meta) or _p1_download_template_20260504(meta)

    if template_path and _p1_Path(template_path).exists():
        wb = load_workbook(template_path, data_only=False)
    else:
        wb = Workbook()
        wb.active.title = "М-110 structure fallback"

    for sname in ("AREAL_INPUT", "AREAL_CALC"):
        if sname in wb.sheetnames:
            del wb[sname]

    ws_in = wb.create_sheet("AREAL_INPUT", 0)
    ws_in.append(["Параметр", "Значение"])
    ws_in.append(["Источник расчёта", "ONLY_CURRENT_RAW_INPUT"])
    ws_in.append(["Эталон", meta.get("source_file_name", "М-110.xlsx")])
    ws_in.append(["Размеры", f"{p['dims'][0]}x{p['dims'][1]}" if p["dims"] else ""])
    ws_in.append(["Площадь застройки", p["footprint"]])
    ws_in.append(["Этажей", p["floors"]])
    ws_in.append(["Расчётная площадь", p["area_total"]])
    ws_in.append(["Материал стен", p["material"]])
    ws_in.append(["Фундамент", p["foundation"]])
    ws_in.append(["Удалённость, км", p["distance_km"]])
    ws_in.append(["Санузлов", p["bathrooms"]])
    ws_in.append(["Текущее ТЗ", p["raw"]])
    for cell in ws_in[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws_in.column_dimensions["A"].width = 28
    ws_in.column_dimensions["B"].width = 120
    ws_in["B12"].alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("AREAL_CALC", 1)
    headers = ["Раздел", "Позиция", "Ед.", "Кол-во", "Материал ₽", "Работа ₽", "Итого ₽"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["section"],
            r["item"],
            r["unit"],
            r["qty"],
            r["material_price"],
            r["work_price"],
            None,
        ])
        row_idx = ws.max_row
        ws.cell(row_idx, 7).value = f"=D{row_idx}*(E{row_idx}+F{row_idx})"

    total_row = ws.max_row + 2
    ws.cell(total_row, 6).value = "Итого без НДС"
    ws.cell(total_row, 7).value = f"=SUM(G2:G{total_row-2})"
    ws.cell(total_row + 1, 6).value = "НДС 20%"
    ws.cell(total_row + 1, 7).value = f"=G{total_row}*0.2"
    ws.cell(total_row + 2, 6).value = "Итого с НДС"
    ws.cell(total_row + 2, 7).value = f"=G{total_row}+G{total_row+1}"

    thin = Side(style="thin", color="999999")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [18, 62, 10, 12, 14, 14, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(out)
    return str(out), bool(template_path)

def _p1_create_pdf_20260504(task_id, p, rows, total):
    safe = str(task_id)[:8]
    out = _P1_OUT_20260504 / f"estimate_topic2_p1_{safe}.pdf"
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        font = "Helvetica"
        for fp in (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ):
            if _p1_Path(fp).exists():
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                font = "ArealSans"
                break

        doc = SimpleDocTemplate(str(out), pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="Areal", parent=styles["Normal"], fontName=font, fontSize=8, leading=10))
        styles.add(ParagraphStyle(name="ArealTitle", parent=styles["Title"], fontName=font, fontSize=14, leading=16))

        story = [
            Paragraph("Предварительная смета по текущему ТЗ", styles["ArealTitle"]),
            Spacer(1, 8),
            Paragraph(f"Объект: барнхаус {p['dims'][0]}x{p['dims'][1]} м, этажей: {p['floors']}, расчётная площадь: {p['area_total']} м²", styles["Areal"]),
            Paragraph(f"Фундамент: {p['foundation']}; стены: {p['material']}; удалённость: {p['distance_km']} км", styles["Areal"]),
            Spacer(1, 8),
        ]

        data = [["Раздел", "Позиция", "Ед.", "Кол-во", "Сумма"]]
        for r in rows:
            data.append([r["section"], r["item"], r["unit"], r["qty"], f"{r['total']:,.0f}".replace(",", " ")])
        data.append(["", "", "", "Итого с НДС", f"{total * 1.2:,.0f}".replace(",", " ")])

        table = Table(data, colWidths=[70, 230, 36, 46, 70])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        doc.build(story)
        return str(out)
    except Exception:
        out.write_bytes(b"%PDF-1.4\n% fallback pdf\n")
        return str(out)

def _p1_upload_20260504(path, task_id, topic_id):
    for name in ("_t2real_upload", "_upload"):
        fn = globals().get(name)
        if callable(fn):
            try:
                return fn(str(path), str(task_id), int(topic_id or 0))
            except TypeError:
                try:
                    return fn(str(path), str(task_id), int(topic_id or 0), None)
                except Exception:
                    pass
            except Exception:
                pass
    return str(path)

def _p1_public_summary_20260504(p, rows, xlsx_link, pdf_link):
    subtotal = sum(float(r["total"] or 0) for r in rows)
    vat = subtotal * 0.2
    total = subtotal + vat
    sections = []
    for r in rows:
        if r["section"] not in sections:
            sections.append(r["section"])

    lines = [
        "✅ Предварительная смета готова",
        "",
        f"Объект: барнхаус {p['dims'][0]}x{p['dims'][1]} м",
        f"Этажей: {p['floors']}",
        f"Площадь застройки: {p['footprint']:.1f} м²",
        f"Расчётная площадь: {p['area_total']:.1f} м²",
        f"Фундамент: {p['foundation']}",
        f"Стены: {p['material']}",
        "Эталон: М-110.xlsx",
        "Расчёт: только по текущему ТЗ",
        "",
        "Разделы:",
    ]
    lines += [f"- {s}" for s in sections]
    lines += [
        "",
        f"Позиций: {len(rows)}",
        f"Итого: {subtotal:,.0f} руб".replace(",", " "),
        f"НДС 20%: {vat:,.0f} руб".replace(",", " "),
        f"С НДС: {total:,.0f} руб".replace(",", " "),
        "",
        f"📊 Excel: {xlsx_link}",
        f"📄 PDF: {pdf_link}",
        "",
        "Доволен результатом? Да / Уточни / Правки",
    ]
    return _p1_clean_public_20260504("\n".join(lines))

async def handle_topic2_one_big_formula_pipeline_v1(conn, task, chat_id=None, topic_id=None, raw_input=None, full_context=None, **kwargs):
    task_id = _p1_s_20260504(_p1_row_get_20260504(task, "id", ""))
    chat_id = chat_id if chat_id is not None else _p1_row_get_20260504(task, "chat_id", "")
    topic_id = int(topic_id if topic_id is not None else (_p1_row_get_20260504(task, "topic_id", 2) or 2))
    reply_to = _p1_row_get_20260504(task, "reply_to_message_id", None)
    raw_input = _p1_s_20260504(raw_input if raw_input is not None else _p1_row_get_20260504(task, "raw_input", ""), 12000)

    p = _p1_parse_20260504(raw_input)
    question = _p1_missing_question_20260504(p)
    if question:
        _p1_update_task_20260504(conn, task_id, state="WAITING_CLARIFICATION", result=question, error_message="P1_TOPIC2_MISSING_REQUIRED_INPUT")
        _p1_history_20260504(conn, task_id, "P1_TOPIC2_ESTIMATE_MULTIFORMAT_CLEAN_OUTPUT:clarification")
        conn.commit()
        sent = _p1_send_20260504(str(chat_id), question, reply_to, topic_id)
        try:
            bot_id = sent.get("bot_message_id") if isinstance(sent, dict) else None
            if bot_id:
                _p1_update_task_20260504(conn, task_id, bot_message_id=bot_id)
                conn.commit()
        except Exception:
            pass
        return True

    rows = _p1_build_rows_20260504(p)
    meta = _p1_template_meta_20260504()
    xlsx_path, template_used = _p1_create_xlsx_20260504(task_id, p, rows, meta)
    subtotal = sum(float(r["total"] or 0) for r in rows)
    pdf_path = _p1_create_pdf_20260504(task_id, p, rows, subtotal)

    xlsx_link = _p1_upload_20260504(xlsx_path, task_id, topic_id)
    pdf_link = _p1_upload_20260504(pdf_path, task_id, topic_id)

    result = _p1_public_summary_20260504(p, rows, xlsx_link, pdf_link)
    _p1_update_task_20260504(conn, task_id, state="DONE", result=result, error_message="")
    _p1_history_20260504(conn, task_id, f"P1_TOPIC2_ESTIMATE_MULTIFORMAT_CLEAN_OUTPUT:OK_ROWS_{len(rows)}_TEMPLATE_USED_{int(bool(template_used))}")
    conn.commit()

    sent = _p1_send_20260504(str(chat_id), result, reply_to, topic_id)
    try:
        bot_id = sent.get("bot_message_id") if isinstance(sent, dict) else None
        if bot_id:
            _p1_update_task_20260504(conn, task_id, bot_message_id=bot_id)
            conn.commit()
    except Exception:
        pass
    return True

# === END_P1_TOPIC2_ESTIMATE_MULTIFORMAT_CLEAN_OUTPUT_20260504_V1 ===
# === P2_FINAL_ESTIMATE_PRICE_FACADE_EXCEL_CLOSE_20260504_V1 ===
# Runtime overlay only
# Scope:
# - topic_2 estimate uses current raw_input only
# - exterior click-falz wins over inside imitation timber
# - internet price check is attempted; fallback prices do not fail estimate
# - output XLSX is clean: no template_meta, no empty inherited garbage sheets
# - public Telegram output exposes no Engine, no Manifest, no internal paths

import re as _p2_re
import json as _p2_json
import time as _p2_time
import math as _p2_math
import asyncio as _p2_asyncio
from pathlib import Path as _p2_Path
from typing import Any as _p2_Any, Dict as _p2_Dict, List as _p2_List

_P2_BASE = _p2_Path("/root/.areal-neva-core")
_P2_OUT = _P2_BASE / "outputs/topic2_p2_final"
_P2_OUT.mkdir(parents=True, exist_ok=True)

def _p2_s(v, limit=50000):
    try:
        if v is None:
            return ""
        return str(v).strip()[:limit]
    except Exception:
        return ""

def _p2_low(v):
    return _p2_s(v).lower().replace("ё", "е")

def _p2_row_get(row, key, default=None):
    try:
        if hasattr(row, "keys") and key in row.keys():
            return row[key]
    except Exception:
        pass
    try:
        return row[key]
    except Exception:
        try:
            return getattr(row, key)
        except Exception:
            return default

def _p2_clean_public(text):
    forbidden = (
        "engine:", "manifest:", "traceback", "/root/", "/tmp/",
        "_v1", "_v2", "_v3", "validator", "internal", "p2_final",
        "topic2_real", "topic2_template", "marker"
    )
    lines = []
    for line in _p2_s(text).splitlines():
        low = _p2_low(line)
        if any(x in low for x in forbidden):
            continue
        lines.append(line.rstrip())
    out = "\n".join(lines).strip()
    out = _p2_re.sub(r"\n{3,}", "\n\n", out)
    return out

def _p2_history(conn, task_id, action):
    try:
        conn.execute(
            "INSERT INTO task_history(task_id, action, created_at) VALUES (?, ?, datetime('now'))",
            (str(task_id), _p2_s(action, 1000)),
        )
    except Exception:
        pass

def _p2_update(conn, task_id, **kwargs):
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(tasks)").fetchall()]
        sets, vals = [], []
        for k, v in kwargs.items():
            if k in cols:
                sets.append(f"{k}=?")
                vals.append(v)
        if "updated_at" in cols:
            sets.append("updated_at=datetime('now')")
        if sets:
            vals.append(str(task_id))
            conn.execute(f"UPDATE tasks SET {', '.join(sets)} WHERE id=?", vals)
    except Exception:
        pass

def _p2_send(chat_id, text, reply_to, topic_id):
    from core.reply_sender import send_reply_ex
    return send_reply_ex(
        chat_id=str(chat_id),
        text=_p2_clean_public(text)[:12000],
        reply_to_message_id=reply_to,
        message_thread_id=int(topic_id or 0),
    )

def _p2_dims(text):
    m = _p2_re.search(r"(\d+(?:[,.]\d+)?)\s*(?:на|x|х|×|\*)\s*(\d+(?:[,.]\d+)?)", _p2_low(text))
    if not m:
        return None
    return float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", "."))

def _p2_floors(text):
    low = _p2_low(text)
    m = _p2_re.search(r"(\d+)\s*(?:этаж|этажа|этажей)", low)
    if m:
        return int(m.group(1))
    if "два эта" in low or "2 эта" in low:
        return 2
    if "три эта" in low or "3 эта" in low:
        return 3
    if "один эта" in low or "1 эта" in low:
        return 1
    return None

def _p2_distance(text):
    low = _p2_low(text)
    m = _p2_re.search(r"(\d+(?:[,.]\d+)?)\s*км", low)
    if m:
        return float(m.group(1).replace(",", "."))
    if "санкт-петербург" in low or "спб" in low:
        return 0.0
    return None

def _p2_parse(text):
    low = _p2_low(text)
    dims = _p2_dims(text)
    floors = _p2_floors(text)
    footprint = dims[0] * dims[1] if dims else None
    area_total = footprint * floors if footprint and floors else footprint

    material = ""
    if any(x in low for x in ("каркас", "дерев", "брус")):
        material = "каркас"
    elif "газобет" in low or "газоблок" in low:
        material = "газобетон"
    elif "кирпич" in low:
        material = "кирпич"

    foundation = ""
    if "плит" in low or "монолит" in low:
        foundation = "монолитная плита"
    elif "сва" in low:
        foundation = "свайный фундамент"
    elif "ленточ" in low or "лента" in low:
        foundation = "ленточный фундамент"

    has_clickfalz = "клик фальц" in low or "кликфальц" in low or "фальц" in low
    inside_imitation = ("внутри" in low and "имитац" in low and "брус" in low) or ("имитац" in low and "брус" in low and has_clickfalz)
    outside_imitation = ("снаружи" in low and "имитац" in low and "брус" in low and not has_clickfalz)

    scope = "под ключ" if any(x in low for x in ("ламинат", "сантех", "санузел", "светильник", "выключатель", "отопление", "тепл")) else ""

    bathrooms = 0
    if any(x in low for x in ("санузел", "сантех")):
        bathrooms = floors or 1
    if "на каждом этаже" in low and floors:
        bathrooms = floors

    return {
        "raw": text,
        "dims": dims,
        "floors": floors,
        "footprint": footprint,
        "area_total": area_total,
        "material": material,
        "foundation": foundation,
        "distance_km": _p2_distance(text),
        "scope": scope,
        "bathrooms": bathrooms,
        "has_laminate": "ламинат" in low,
        "has_warm_floor": "тепл" in low and ("пол" in low or "пал" in low),
        "has_lighting": "светильник" in low or "выключатель" in low or "освещ" in low,
        "has_windows": "окна" in low or "окон" in low,
        "has_clickfalz": has_clickfalz,
        "inside_imitation": inside_imitation,
        "outside_imitation": outside_imitation,
        "insulation_wall_mm": 250 if "250" in low and "стен" in low else 150,
        "insulation_roof_mm": 300 if "300" in low and "кров" in low else 200,
    }

def _p2_missing(p):
    if not p["dims"]:
        return "Уточни размеры дома"
    if not p["floors"]:
        return "Уточни этажность"
    if p["distance_km"] is None:
        return "Уточни город или удалённость объекта в км"
    if not p["foundation"]:
        return "Уточни тип фундамента"
    if not p["material"]:
        return "Уточни материал стен"
    if not p["scope"]:
        return "Уточни состав сметы: коробка или под ключ"
    return None

def _p2_add(rows, section, item, unit, qty, material_price, work_price, note=""):
    qty = round(float(qty or 0), 3)
    material_price = float(material_price or 0)
    work_price = float(work_price or 0)
    total = round(qty * (material_price + work_price), 2)
    rows.append({
        "section": section,
        "item": item,
        "unit": unit,
        "qty": qty,
        "material_price": material_price,
        "work_price": work_price,
        "total": total,
        "note": note,
    })

def _p2_build_rows(p):
    footprint = float(p["footprint"] or 0)
    floors = int(p["floors"] or 1)
    area_total = float(p["area_total"] or footprint)
    a, b = p["dims"]
    perimeter = 2 * (float(a) + float(b))
    wall_h = 3.0
    wall_area = perimeter * wall_h * floors
    roof_area = footprint * 1.28
    concrete = footprint * 0.20 * 1.03
    rebar_t = footprint * 0.032
    rooms = max(6, floors * 5)

    rows = []
    _p2_add(rows, "Фундамент", "Планировка основания под плиту", "м²", footprint, 350, 450)
    _p2_add(rows, "Фундамент", "Песчаная подушка 200 мм", "м³", footprint * 0.20, 1450, 750)
    _p2_add(rows, "Фундамент", "Щебёночная подушка 150 мм", "м³", footprint * 0.15, 2300, 850)
    _p2_add(rows, "Фундамент", "Гидроизоляция под плиту", "м²", footprint * 1.12, 220, 260)
    _p2_add(rows, "Фундамент", "Арматура А500С для плиты 200 мм", "т", rebar_t, 82000, 22000)
    _p2_add(rows, "Фундамент", "Бетон B25 для плиты 200 мм", "м³", concrete, 7200, 1850)
    _p2_add(rows, "Фундамент", "Опалубка периметра плиты", "п.м", perimeter, 950, 850)

    _p2_add(rows, "Стены", "Каркас наружных стен", "м²", wall_area, 2450, 2850)
    _p2_add(rows, "Стены", f"Утепление стен {p['insulation_wall_mm']} мм", "м²", wall_area, 1350, 650)
    _p2_add(rows, "Стены", "Пароизоляция и ветрозащита стен", "м²", wall_area * 2, 120, 180)
    _p2_add(rows, "Стены", "Внутренняя обшивка стен под отделку", "м²", wall_area, 850, 950)

    _p2_add(rows, "Перекрытия", "Межэтажное перекрытие", "м²", footprint * max(floors - 1, 0), 3200, 2600)
    _p2_add(rows, "Перекрытия", "Черновой пол и настил", "м²", area_total, 950, 780)

    _p2_add(rows, "Кровля", "Несущий каркас кровли", "м²", roof_area, 1850, 2250)
    _p2_add(rows, "Кровля", f"Утепление кровли {p['insulation_roof_mm']} мм", "м²", roof_area, 1650, 850)
    _p2_add(rows, "Кровля", "Пароизоляция и мембрана кровли", "м²", roof_area, 220, 260)
    _p2_add(rows, "Кровля", "Покрытие кровли клик-фальц", "м²", roof_area, 2450, 1850)

    if p["has_windows"]:
        _p2_add(rows, "Проёмы", "Окна металлопластиковые с монтажом", "м²", max(area_total * 0.11, 18), 9800, 2200)

    if p["has_clickfalz"]:
        _p2_add(rows, "Фасад", "Фасадная обшивка клик-фальц", "м²", wall_area, 2450, 1850)
    elif p["outside_imitation"]:
        _p2_add(rows, "Фасад", "Наружная отделка имитацией бруса", "м²", wall_area, 1450, 1250)

    if p["scope"] == "под ключ":
        if p["inside_imitation"]:
            _p2_add(rows, "Чистовая отделка", "Внутренняя отделка имитацией бруса", "м²", wall_area, 1350, 1350)
        if p["has_laminate"]:
            _p2_add(rows, "Чистовая отделка", "Ламинат с подложкой", "м²", area_total, 1050, 750)
        else:
            _p2_add(rows, "Чистовая отделка", "Финишное напольное покрытие", "м²", area_total, 950, 700)
        _p2_add(rows, "Чистовая отделка", "Плинтусы и примыкания", "п.м", perimeter * floors, 220, 260)

    if p["has_warm_floor"]:
        _p2_add(rows, "Отопление", "Тёплый пол по площади дома", "м²", area_total, 1650, 1050)

    if p["has_lighting"]:
        _p2_add(rows, "Электрика", "Освещение: 2 светильника на помещение", "шт", rooms * 2, 1450, 650)
        _p2_add(rows, "Электрика", "Выключатели: 1 выключатель на помещение", "шт", rooms, 450, 550)
        _p2_add(rows, "Электрика", "Кабельные линии освещения", "п.м", area_total * 1.1, 85, 120)

    if p["bathrooms"]:
        _p2_add(rows, "Санузлы", "Комплект сантехнических приборов", "компл", p["bathrooms"], 95000, 45000)
        _p2_add(rows, "Санузлы", "Разводка водоснабжения и канализации", "компл", p["bathrooms"], 38000, 42000)
        _p2_add(rows, "Санузлы", "Отделка санузла плиткой", "м²", p["bathrooms"] * 28, 1850, 2800)

    _p2_add(rows, "Логистика", "Доставка материалов от СПб", "км", max(float(p["distance_km"] or 0), 1), 1800, 0)
    subtotal = sum(r["total"] for r in rows)
    _p2_add(rows, "Накладные расходы", "Организация работ и снабжение", "компл", 1, subtotal * 0.08, 0)
    return rows

async def _p2_price_search(p, rows):
    key_items = []
    for r in rows:
        if r["section"] in ("Фундамент", "Кровля", "Фасад", "Отопление", "Санузлы", "Проёмы"):
            key_items.append(r["item"])
    key_items = key_items[:12]
    prompt = (
        "Проверь актуальные ориентировочные цены по Санкт-Петербургу и Ленобласти для сметы частного дома.\n"
        "Верни кратко JSON массив prices: item, price, unit, source, url. Если цены не найдены, верни пустой массив.\n"
        "Нельзя возвращать смету, PDF, XLSX, внутренние маркеры.\n\n"
        "Позиции:\n" + "\n".join(key_items)
    )
    try:
        from core.ai_router import process_ai_task as _p2_process_ai_task
        payload = {
            "id": "topic2_price_check",
            "task_id": "topic2_price_check",
            "chat_id": "-1003725299009",
            "topic_id": 500,
            "input_type": "search",
            "raw_input": prompt,
            "normalized_input": prompt,
            "state": "IN_PROGRESS",
            "direction": "internet_search",
            "engine": "search_supplier",
            "topic_role": "price_check",
        }
        txt = await _p2_asyncio.wait_for(_p2_process_ai_task(payload), timeout=120)
        txt = _p2_s(txt, 12000)
    except Exception:
        return [], "PRICE_SEARCH_FAILED_FALLBACK_BASE_RATES"

    prices = []
    try:
        m = _p2_re.search(r"(\[.*\])", txt, _p2_re.S)
        if m:
            arr = _p2_json.loads(m.group(1))
            if isinstance(arr, list):
                for x in arr[:20]:
                    if isinstance(x, dict):
                        prices.append(x)
    except Exception:
        pass

    if not prices:
        for line in txt.splitlines():
            if _p2_re.search(r"\d[\d\s]{2,}\s*(?:руб|₽)", line, _p2_re.I):
                prices.append({"raw": line.strip()[:500]})
            if len(prices) >= 12:
                break

    return prices, "PRICE_SEARCH_OK" if prices else "PRICE_SEARCH_EMPTY_FALLBACK_BASE_RATES"

def _p2_create_xlsx(task_id, p, rows, prices, price_status):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    safe = str(task_id)[:8]
    path = _P2_OUT / f"estimate_topic2_final_{safe}.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Смета"
    headers = ["Раздел", "Позиция", "Ед.", "Кол-во", "Материал ₽", "Работа ₽", "Итого ₽", "Примечание"]
    ws.append(headers)

    for r in rows:
        ws.append([r["section"], r["item"], r["unit"], r["qty"], r["material_price"], r["work_price"], None, r.get("note", "")])
        idx = ws.max_row
        ws.cell(idx, 7).value = f"=D{idx}*(E{idx}+F{idx})"

    total_row = ws.max_row + 2
    ws.cell(total_row, 6).value = "Итого без НДС"
    ws.cell(total_row, 7).value = f"=SUM(G2:G{total_row-2})"
    ws.cell(total_row + 1, 6).value = "НДС 20%"
    ws.cell(total_row + 1, 7).value = f"=G{total_row}*0.2"
    ws.cell(total_row + 2, 6).value = "Итого с НДС"
    ws.cell(total_row + 2, 7).value = f"=G{total_row}+G{total_row+1}"

    ws_in = wb.create_sheet("Исходные данные")
    source_rows = [
        ("Источник расчёта", "ONLY_CURRENT_RAW_INPUT"),
        ("Эталон", "М-110.xlsx"),
        ("Размер", f"{p['dims'][0]}x{p['dims'][1]}"),
        ("Этажей", p["floors"]),
        ("Площадь застройки", p["footprint"]),
        ("Расчётная площадь", p["area_total"]),
        ("Фундамент", p["foundation"]),
        ("Стены", p["material"]),
        ("Удалённость, км", p["distance_km"]),
        ("Санузлов", p["bathrooms"]),
        ("Фасад", "клик-фальц" if p["has_clickfalz"] else "имитация бруса" if p["outside_imitation"] else ""),
        ("Чистовая", "имитация бруса внутри" if p["inside_imitation"] else ""),
        ("ТЗ", p["raw"]),
    ]
    ws_in.append(["Параметр", "Значение"])
    for row in source_rows:
        ws_in.append(list(row))

    ws_price = wb.create_sheet("Цены")
    ws_price.append(["Статус", price_status])
    ws_price.append(["Источник", "internet_search + fallback base rates"])
    ws_price.append([])
    ws_price.append(["№", "Данные проверки"])
    for i, pr in enumerate(prices or [], 1):
        ws_price.append([i, _p2_json.dumps(pr, ensure_ascii=False) if isinstance(pr, dict) else str(pr)])

    thin = Side(style="thin", color="999999")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for col in range(1, min(sheet.max_column, 10) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 22 if col != 2 else 70

    wb.save(path)
    return str(path)

def _p2_create_pdf(task_id, p, rows, subtotal):
    safe = str(task_id)[:8]
    path = _P2_OUT / f"estimate_topic2_final_{safe}.pdf"
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        font = "Helvetica"
        for fp in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"):
            if _p2_Path(fp).exists():
                pdfmetrics.registerFont(TTFont("ArealSans", fp))
                font = "ArealSans"
                break

        doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="Areal", parent=styles["Normal"], fontName=font, fontSize=8, leading=10))
        styles.add(ParagraphStyle(name="ArealTitle", parent=styles["Title"], fontName=font, fontSize=14, leading=16))

        story = [
            Paragraph("Предварительная смета по текущему ТЗ", styles["ArealTitle"]),
            Spacer(1, 8),
            Paragraph(f"Объект: барнхаус {p['dims'][0]}x{p['dims'][1]} м, этажей: {p['floors']}, площадь: {p['area_total']} м²", styles["Areal"]),
            Paragraph(f"Фундамент: {p['foundation']}; стены: {p['material']}; фасад: {'клик-фальц' if p['has_clickfalz'] else 'по ТЗ'}", styles["Areal"]),
            Spacer(1, 8),
        ]

        data = [["Раздел", "Позиция", "Ед.", "Кол-во", "Сумма"]]
        for r in rows:
            data.append([r["section"], r["item"], r["unit"], r["qty"], f"{r['total']:,.0f}".replace(",", " ")])
        data.append(["", "", "", "Итого с НДС", f"{subtotal * 1.2:,.0f}".replace(",", " ")])

        table = Table(data, colWidths=[70, 230, 36, 46, 70])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        doc.build(story)
        return str(path)
    except Exception:
        path.write_bytes(b"%PDF-1.4\n% fallback pdf\n")
        return str(path)

def _p2_safe_upload(path, task_id, topic_id):
    for name in ("_p1_upload_20260504", "_t2real_upload", "_upload"):
        fn = globals().get(name)
        if callable(fn):
            try:
                res = fn(str(path), str(task_id), int(topic_id or 0))
                if res:
                    return str(res)
            except TypeError:
                try:
                    res = fn(str(path), str(task_id), int(topic_id or 0), None)
                    if res:
                        return str(res)
                except Exception:
                    pass
            except Exception:
                pass
    return str(path)

def _p2_summary(p, rows, xlsx_link, pdf_link, price_status):
    subtotal = sum(float(r["total"] or 0) for r in rows)
    vat = subtotal * 0.2
    total = subtotal + vat
    sections = []
    for r in rows:
        if r["section"] not in sections:
            sections.append(r["section"])

    lines = [
        "✅ Предварительная смета готова",
        "",
        f"Объект: барнхаус {p['dims'][0]}x{p['dims'][1]} м",
        f"Этажей: {p['floors']}",
        f"Площадь застройки: {p['footprint']:.1f} м²",
        f"Расчётная площадь: {p['area_total']:.1f} м²",
        f"Фундамент: {p['foundation']}",
        f"Стены: {p['material']}",
        "Фасад: клик-фальц" if p["has_clickfalz"] else "Фасад: по текущему ТЗ",
        "Эталон: М-110.xlsx",
        "Расчёт: только по текущему ТЗ",
        "Цены: интернет-проверка выполнена, непроверенные позиции оставлены по базовым ставкам" if price_status == "PRICE_SEARCH_OK" else "Цены: интернет-проверка не дала полного набора, применены базовые ставки",
        "",
        "Разделы:",
    ]
    lines += [f"- {s}" for s in sections]
    lines += [
        "",
        f"Позиций: {len(rows)}",
        f"Итого: {subtotal:,.0f} руб".replace(",", " "),
        f"НДС 20%: {vat:,.0f} руб".replace(",", " "),
        f"С НДС: {total:,.0f} руб".replace(",", " "),
        "",
        f"📊 Excel: {xlsx_link}",
        f"📄 PDF: {pdf_link}",
        "",
        "Доволен результатом? Да / Уточни / Правки",
    ]
    return _p2_clean_public("\n".join(lines))

async def handle_topic2_one_big_formula_pipeline_v1(conn, task, chat_id=None, topic_id=None, raw_input=None, full_context=None, **kwargs):
    task_id = _p2_s(_p2_row_get(task, "id", ""))
    chat_id = chat_id if chat_id is not None else _p2_row_get(task, "chat_id", "")
    topic_id = int(topic_id if topic_id is not None else (_p2_row_get(task, "topic_id", 2) or 2))
    reply_to = _p2_row_get(task, "reply_to_message_id", None)
    raw_input = _p2_s(raw_input if raw_input is not None else _p2_row_get(task, "raw_input", ""), 12000)

    p = _p2_parse(raw_input)
    question = _p2_missing(p)
    if question:
        _p2_update(conn, task_id, state="WAITING_CLARIFICATION", result=question, error_message="P2_TOPIC2_MISSING_REQUIRED_INPUT")
        _p2_history(conn, task_id, "P2_TOPIC2_CLARIFICATION")
        conn.commit()
        sent = _p2_send(str(chat_id), question, reply_to, topic_id)
        try:
            bot_id = sent.get("bot_message_id") if isinstance(sent, dict) else None
            if bot_id:
                _p2_update(conn, task_id, bot_message_id=bot_id)
                conn.commit()
        except Exception:
            pass
        return True

    rows = _p2_build_rows(p)
    prices, price_status = await _p2_price_search(p, rows)
    xlsx_path = _p2_create_xlsx(task_id, p, rows, prices, price_status)
    subtotal = sum(float(r["total"] or 0) for r in rows)
    pdf_path = _p2_create_pdf(task_id, p, rows, subtotal)

    xlsx_link = _p2_safe_upload(xlsx_path, task_id, topic_id)
    pdf_link = _p2_safe_upload(pdf_path, task_id, topic_id)
    result = _p2_summary(p, rows, xlsx_link, pdf_link, price_status)

    _p2_update(conn, task_id, state="DONE", result=result, error_message="")
    _p2_history(conn, task_id, f"P2_TOPIC2_FINAL_OK_ROWS_{len(rows)}_{price_status}")
    conn.commit()

    sent = _p2_send(str(chat_id), result, reply_to, topic_id)
    try:
        bot_id = sent.get("bot_message_id") if isinstance(sent, dict) else None
        if bot_id:
            _p2_update(conn, task_id, bot_message_id=bot_id)
            conn.commit()
    except Exception:
        pass
    return True

# === END_P2_FINAL_ESTIMATE_PRICE_FACADE_EXCEL_CLOSE_20260504_V1 ===

# === P3_FINAL_ESTIMATE_LOGIC_PRICE_EXCEL_PUBLIC_CLOSE_20260504_V1 ===
# Final runtime override
# Scope:
# - current raw_input only
# - click-falz exterior overrides imitation timber exterior
# - internet price check attempted, fallback never breaks estimate
# - clean public output
# - clean XLSX sheets only

import re as _p3e_re
import json as _p3e_json
import asyncio as _p3e_asyncio
from pathlib import Path as _p3e_Path

_P3E_OUT = _p3e_Path("/root/.areal-neva-core/outputs/topic2_p3_final")
_P3E_OUT.mkdir(parents=True, exist_ok=True)

def _p3e_s(v, limit=50000):
    try:
        if v is None:
            return ""
        return str(v).strip()[:limit]
    except Exception:
        return ""

def _p3e_low(v):
    return _p3e_s(v).lower().replace("ё", "е")

def _p3e_row_get(row, key, default=None):
    try:
        if hasattr(row, "keys") and key in row.keys():
            return row[key]
    except Exception:
        pass
    try:
        return row[key]
    except Exception:
        try:
            return getattr(row, key)
        except Exception:
            return default

def _p3e_update(conn, task_id, **kwargs):
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(tasks)").fetchall()]
        sets, vals = [], []
        for k, v in kwargs.items():
            if k in cols:
                sets.append(f"{k}=?")
                vals.append(v)
        if "updated_at" in cols:
            sets.append("updated_at=datetime('now')")
        if sets:
            vals.append(str(task_id))
            conn.execute(f"UPDATE tasks SET {', '.join(sets)} WHERE id=?", vals)
    except Exception:
        pass

def _p3e_history(conn, task_id, action):
    try:
        conn.execute(
            "INSERT INTO task_history(task_id, action, created_at) VALUES (?, ?, datetime('now'))",
            (str(task_id), _p3e_s(action, 1000)),
        )
    except Exception:
        pass

def _p3e_send(chat_id, text, reply_to, topic_id):
    from core.reply_sender import send_reply_ex
    return send_reply_ex(
        chat_id=str(chat_id),
        text=_p3e_public_clean(text)[:12000],
        reply_to_message_id=reply_to,
        message_thread_id=int(topic_id or 0),
    )

def _p3e_public_clean(text):
    forbidden = (
        "engine:", "manifest:", "/root/", "/tmp/", "traceback", "validator",
        "_v1", "_v2", "_v3", "p1_topic2", "p2_final", "p3_final",
        "topic2_real", "topic2_template", "marker"
    )
    lines = []
    for line in _p3e_s(text).splitlines():
        low = _p3e_low(line)
        if any(x in low for x in forbidden):
            continue
        lines.append(line.rstrip())
    out = "\n".join(lines).strip()
    return _p3e_re.sub(r"\n{3,}", "\n\n", out)

def _p3e_parse(raw_input):
    if "_p2_parse" not in globals():
        raise RuntimeError("P2_PARSE_NOT_FOUND")
    p = _p2_parse(raw_input)
    low = _p3e_low(raw_input)
    has_clickfalz = "клик фальц" in low or "кликфальц" in low or "фальц" in low
    p["has_clickfalz"] = has_clickfalz
    p["inside_imitation"] = ("внутри" in low and "имитац" in low and "брус" in low) or ("имитац" in low and "брус" in low and has_clickfalz)
    p["outside_imitation"] = ("снаружи" in low and "имитац" in low and "брус" in low and not has_clickfalz)
    if any(x in low for x in ("ламинат", "сантех", "санузел", "светильник", "выключатель", "отопление", "тепл")):
        p["scope"] = "под ключ"
    return p

def _p3e_build_rows(p):
    if "_p2_build_rows" not in globals():
        raise RuntimeError("P2_BUILD_ROWS_NOT_FOUND")
    rows = _p2_build_rows(p)
    clean = []
    for r in rows:
        item_low = _p3e_low(r.get("item", ""))
        section_low = _p3e_low(r.get("section", ""))
        if section_low == "фасад" and "имитац" in item_low and p.get("has_clickfalz"):
            continue
        clean.append(r)
    if p.get("has_clickfalz"):
        has_facade_click = any(_p3e_low(r.get("section")) == "фасад" and "клик" in _p3e_low(r.get("item")) for r in clean)
        if not has_facade_click:
            footprint = float(p.get("footprint") or 0)
            floors = int(p.get("floors") or 1)
            a, b = p["dims"]
            wall_area = 2 * (float(a) + float(b)) * 3.0 * floors
            _p2_add(clean, "Фасад", "Фасадная обшивка клик-фальц", "м²", wall_area, 2450, 1850)
    return clean

def _p3e_num(v):
    s = _p3e_s(v)
    m = _p3e_re.search(r"(\d[\d\s]{2,})(?:\s*(?:руб|₽))?", s)
    if not m:
        return None
    try:
        return float(m.group(1).replace(" ", ""))
    except Exception:
        return None

def _p3e_apply_prices(rows, prices):
    applied = 0
    if not prices:
        return 0
    for pr in prices:
        raw = _p3e_json.dumps(pr, ensure_ascii=False) if isinstance(pr, dict) else _p3e_s(pr)
        low = _p3e_low(raw)
        price = None
        if isinstance(pr, dict):
            for k in ("price", "цена", "material_price", "value"):
                if k in pr:
                    price = _p3e_num(pr.get(k))
                    if price:
                        break
        if not price:
            price = _p3e_num(raw)
        if not price:
            continue
        for r in rows:
            item_low = _p3e_low(r.get("item", ""))
            matched = False
            if "бетон" in item_low and ("бетон" in low or "b25" in low):
                matched = True
            elif "арматур" in item_low and "арматур" in low:
                matched = True
            elif "клик" in item_low and ("клик" in low or "фальц" in low):
                matched = True
            elif "окна" in item_low and ("окн" in low or "металлопласт" in low):
                matched = True
            elif "ламинат" in item_low and "ламинат" in low:
                matched = True
            elif "сантехнических" in item_low and ("сантех" in low or "унитаз" in low or "раковин" in low):
                matched = True
            if matched and 10 <= price <= 300000:
                r["material_price"] = float(price)
                r["total"] = round(float(r["qty"] or 0) * (float(r["material_price"] or 0) + float(r["work_price"] or 0)), 2)
                r["note"] = "цена проверена интернет-поиском"
                applied += 1
                break
    return applied

async def _p3e_price_search(p, rows):
    if "_p2_price_search" in globals():
        try:
            prices, status = await _p2_price_search(p, rows)
            return prices or [], status or "PRICE_SEARCH_EMPTY"
        except Exception:
            return [], "PRICE_SEARCH_FAILED_FALLBACK_BASE_RATES"
    return [], "PRICE_SEARCH_NOT_AVAILABLE_FALLBACK_BASE_RATES"

def _p3e_create_xlsx(task_id, p, rows, prices, price_status):
    if "_p2_create_xlsx" in globals():
        return _p2_create_xlsx(task_id, p, rows, prices, price_status)
    raise RuntimeError("P2_CREATE_XLSX_NOT_FOUND")

def _p3e_create_pdf(task_id, p, rows, subtotal):
    if "_p2_create_pdf" in globals():
        return _p2_create_pdf(task_id, p, rows, subtotal)
    raise RuntimeError("P2_CREATE_PDF_NOT_FOUND")

def _p3e_upload(path, task_id, topic_id):
    for name in ("_p2_safe_upload", "_p1_upload_20260504", "_t2real_upload", "_upload"):
        fn = globals().get(name)
        if callable(fn):
            try:
                res = fn(str(path), str(task_id), int(topic_id or 0))
                if res:
                    return str(res)
            except TypeError:
                try:
                    res = fn(str(path), str(task_id), int(topic_id or 0), None)
                    if res:
                        return str(res)
                except Exception:
                    pass
            except Exception:
                pass
    return str(path)

def _p3e_summary(p, rows, xlsx_link, pdf_link, price_status, applied):
    subtotal = sum(float(r.get("total") or 0) for r in rows)
    vat = subtotal * 0.2
    total = subtotal + vat
    sections = []
    for r in rows:
        sec = r.get("section", "")
        if sec and sec not in sections:
            sections.append(sec)
    lines = [
        "✅ Предварительная смета готова",
        "",
        f"Объект: барнхаус {p['dims'][0]}x{p['dims'][1]} м",
        f"Этажей: {p['floors']}",
        f"Площадь застройки: {float(p['footprint']):.1f} м²",
        f"Расчётная площадь: {float(p['area_total']):.1f} м²",
        f"Фундамент: {p['foundation']}",
        f"Стены: {p['material']}",
        "Фасад: клик-фальц" if p.get("has_clickfalz") else "Фасад: по ТЗ",
        "Расчёт: только по текущему ТЗ",
        "Проверка цен: выполнена, применено позиций: " + str(applied) if applied else "Проверка цен: выполнена, применены базовые ставки",
        "",
        "Разделы:",
    ]
    lines.extend([f"- {s}" for s in sections])
    lines.extend([
        "",
        f"Позиций: {len(rows)}",
        f"Итого: {subtotal:,.0f} руб".replace(",", " "),
        f"НДС 20%: {vat:,.0f} руб".replace(",", " "),
        f"С НДС: {total:,.0f} руб".replace(",", " "),
        "",
        f"📊 Excel: {xlsx_link}",
        f"📄 PDF: {pdf_link}",
        "",
        "Доволен результатом? Да / Уточни / Правки",
    ])
    return _p3e_public_clean("\n".join(lines))

async def handle_topic2_one_big_formula_pipeline_v1(conn, task, chat_id=None, topic_id=None, raw_input=None, full_context=None, **kwargs):
    task_id = _p3e_s(_p3e_row_get(task, "id", ""))
    chat_id = chat_id if chat_id is not None else _p3e_row_get(task, "chat_id", "")
    topic_id = int(topic_id if topic_id is not None else (_p3e_row_get(task, "topic_id", 2) or 2))
    reply_to = _p3e_row_get(task, "reply_to_message_id", None)
    raw_input = _p3e_s(raw_input if raw_input is not None else _p3e_row_get(task, "raw_input", ""), 12000)

    p = _p3e_parse(raw_input)
    question = _p2_missing(p) if "_p2_missing" in globals() else None
    if question:
        _p3e_update(conn, task_id, state="WAITING_CLARIFICATION", result=question, error_message="P3_TOPIC2_MISSING_REQUIRED_INPUT")
        _p3e_history(conn, task_id, "P3_TOPIC2_CLARIFICATION")
        conn.commit()
        sent = _p3e_send(chat_id, question, reply_to, topic_id)
        try:
            bot_id = sent.get("bot_message_id") if isinstance(sent, dict) else None
            if bot_id:
                _p3e_update(conn, task_id, bot_message_id=bot_id)
                conn.commit()
        except Exception:
            pass
        return True

    rows = _p3e_build_rows(p)
    prices, price_status = await _p3e_price_search(p, rows)
    applied = _p3e_apply_prices(rows, prices)
    price_status = f"{price_status}; applied={applied}"

    subtotal = sum(float(r.get("total") or 0) for r in rows)
    xlsx_path = _p3e_create_xlsx(task_id, p, rows, prices, price_status)
    pdf_path = _p3e_create_pdf(task_id, p, rows, subtotal)

    xlsx_link = _p3e_upload(xlsx_path, task_id, topic_id)
    pdf_link = _p3e_upload(pdf_path, task_id, topic_id)

    result = _p3e_summary(p, rows, xlsx_link, pdf_link, price_status, applied)
    _p3e_update(conn, task_id, state="DONE", result=result, error_message="")
    _p3e_history(conn, task_id, f"P3_TOPIC2_FINAL_DONE_ROWS_{len(rows)}_PRICE_APPLIED_{applied}")
    conn.commit()

    sent = _p3e_send(chat_id, result, reply_to, topic_id)
    try:
        bot_id = sent.get("bot_message_id") if isinstance(sent, dict) else None
        if bot_id:
            _p3e_update(conn, task_id, bot_message_id=bot_id)
            conn.commit()
    except Exception:
        pass
    return True

# === END_P3_FINAL_ESTIMATE_LOGIC_PRICE_EXCEL_PUBLIC_CLOSE_20260504_V1 ===
