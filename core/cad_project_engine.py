# === FULLFIX_07_CAD_PROJECT_DOCUMENTATION_CLOSURE ===
import os
import re
import json
import math
import glob
import tempfile
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Dict, List, Tuple

BASE = "/root/.areal-neva-core"
TEMPLATE_DIR = f"{BASE}/data/project_templates"

ENGINE = "FULLFIX_07_CAD_PROJECT_DOCUMENTATION_CLOSURE"

DEFAULT_FOUNDATION_SHEETS = [
    {"mark": "КЖ", "number": "0", "title": "Титульный лист"},
    {"mark": "КЖ", "number": "1", "title": "Общие данные"},
    {"mark": "КЖ", "number": "2", "title": "Ведомость листов"},
    {"mark": "КЖ", "number": "3", "title": "План фундаментной плиты"},
    {"mark": "КЖ", "number": "4", "title": "Разрез 1-1"},
    {"mark": "КЖ", "number": "5", "title": "Схема нижнего армирования"},
    {"mark": "КЖ", "number": "6", "title": "Схема верхнего армирования"},
    {"mark": "КЖ", "number": "7", "title": "Узлы и детали"},
    {"mark": "КЖ", "number": "8", "title": "Спецификация материалов"},
    {"mark": "КЖ", "number": "9", "title": "Ведомость расхода стали"},
]

DEFAULT_ROOF_SHEETS = [
    {"mark": "КД", "number": "0", "title": "Титульный лист"},
    {"mark": "КД", "number": "1", "title": "Общие данные"},
    {"mark": "КД", "number": "2", "title": "Ведомость листов"},
    {"mark": "КД", "number": "3", "title": "План кровли"},
    {"mark": "КД", "number": "4", "title": "План стропильной системы"},
    {"mark": "КД", "number": "5", "title": "Разрезы"},
    {"mark": "КД", "number": "6", "title": "Узлы кровли"},
    {"mark": "КД", "number": "7", "title": "Спецификация древесины"},
    {"mark": "КД", "number": "8", "title": "Спецификация крепежа"},
]

NORMATIVE_NOTES = [
    "СП 63.13330.2018 Бетонные и железобетонные конструкции",
    "СП 20.13330.2016 Нагрузки и воздействия",
    "ГОСТ 21.501-2018 Правила выполнения рабочей документации архитектурных и конструктивных решений",
    "ГОСТ 21.101-2020 Основные требования к проектной и рабочей документации",
    "ГОСТ 34028-2016 Прокат арматурный для железобетонных конструкций",
]

REBAR_WEIGHT_KG_M = {
    6: 0.222,
    8: 0.395,
    10: 0.617,
    12: 0.888,
    14: 1.21,
    16: 1.58,
    18: 2.00,
    20: 2.47,
    22: 2.98,
    25: 3.85,
}

def _clean(v: Any, limit: int = 10000) -> str:
    return str(v or "").replace("\x00", " ").strip()[:limit]

def _safe_name(v: Any) -> str:
    s = re.sub(r"[^A-Za-zА-Яа-я0-9_.-]+", "_", _clean(v, 80))
    return s.strip("_") or "project"

def _font_name() -> str:
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        ]
        for path in candidates:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont("ArealSans", path))
                return "ArealSans"
    except Exception:
        pass
    return "Helvetica"

def _load_templates() -> List[Dict[str, Any]]:
    out = []
    for p in sorted(glob.glob(f"{TEMPLATE_DIR}/PROJECT_TEMPLATE_MODEL__*.json"), key=os.path.getmtime, reverse=True):
        try:
            data = json.loads(Path(p).read_text(encoding="utf-8"))
            if isinstance(data, dict):
                data["template_file"] = p
                out.append(data)
        except Exception:
            pass
    return out

def _choose_template(section: str, topic_id: int = 0) -> Dict[str, Any]:
    templates = _load_templates()
    if not templates:
        return {}
    section = _clean(section).upper()
    for tpl in templates:
        if topic_id and int(tpl.get("topic_id", 0) or 0) == int(topic_id) and _clean(tpl.get("project_type")).upper() == section:
            return tpl
    for tpl in templates:
        if _clean(tpl.get("project_type")).upper() == section:
            return tpl
    return templates[0]

def _num(text: str, default: float) -> float:
    try:
        return float(str(text).replace(",", "."))
    except Exception:
        return default

def _parse_mm(text: str, patterns: List[str], default: int) -> int:
    low = text.lower()
    for pat in patterns:
        m = re.search(pat, low, re.I)
        if m:
            return int(float(m.group(1).replace(",", ".")))
    return default

def parse_project_request(raw_input: str, template_hint: str = "") -> Dict[str, Any]:
    text = _clean(raw_input + " " + template_hint, 6000)
    low = text.lower()

    section = "КЖ"
    project_kind = "foundation_slab"
    if any(x in low for x in ["кров", "строп", "кд"]):
        section = "КД"
        project_kind = "roof"
    if any(x in low for x in ["ар ", "архитект", "планиров"]):
        section = "АР"
        project_kind = "architectural"

    length_m = 10.0
    width_m = 10.0
    m = re.search(r"(\d+(?:[,.]\d+)?)\s*[xх×]\s*(\d+(?:[,.]\d+)?)\s*(?:м|m)?", low)
    if m:
        length_m = _num(m.group(1), 10.0)
        width_m = _num(m.group(2), 10.0)

    slab_mm = _parse_mm(low, [
        r"толщин[аы]?\D{0,30}(\d{2,4})\s*мм",
        r"плит[аы]?\D{0,30}(\d{2,4})\s*мм",
        r"бетон\D{0,30}(\d{2,4})\s*мм",
    ], 250)

    sand_mm = _parse_mm(low, [
        r"песчан\D{0,30}(\d{2,4})\s*мм",
        r"песок\D{0,30}(\d{2,4})\s*мм",
    ], 300)

    gravel_mm = _parse_mm(low, [
        r"щеб[её]н\D{0,30}(\d{2,4})\s*мм",
        r"щебень\D{0,30}(\d{2,4})\s*мм",
        r"основан\D{0,30}(\d{2,4})\s*мм",
    ], 150)

    rebar_step_mm = _parse_mm(low, [
        r"шаг\D{0,30}(\d{2,4})\s*мм",
        r"арматур\D{0,40}(\d{2,4})\s*мм",
    ], 200)

    rebar_diam_mm = 12
    md = re.search(r"(?:ø|ф|d|диаметр)\s*(\d{1,2})", low, re.I)
    if md:
        rebar_diam_mm = int(md.group(1))
    else:
        md = re.search(r"арматур[аы]?\D{0,30}(\d{1,2})(?!\d)", low, re.I)
        if md:
            rebar_diam_mm = int(md.group(1))

    concrete_class = "B25"
    mc = re.search(r"\b[вb]\s?(\d{2,3}(?:[,.]\d)?)\b", text, re.I)
    if mc:
        concrete_class = "B" + mc.group(1).replace(",", ".")

    rebar_class = "A500"
    mr = re.search(r"\b[аa]\s?500[сc]?\b", text, re.I)
    if mr:
        rebar_class = "A500C" if "c" in mr.group(0).lower() or "с" in mr.group(0).lower() else "A500"

    return {
        "project_name": "Проект фундаментной плиты" if project_kind == "foundation_slab" else "Проект по образцу",
        "project_kind": project_kind,
        "section": section,
        "length_m": length_m,
        "width_m": width_m,
        "slab_mm": slab_mm,
        "sand_mm": sand_mm,
        "gravel_mm": gravel_mm,
        "rebar_diam_mm": rebar_diam_mm,
        "rebar_step_mm": rebar_step_mm,
        "rebar_class": rebar_class,
        "concrete_class": concrete_class,
        "cover_mm": 40,
        "input": raw_input,
    }

def _normalize_sheet_register(template: Dict[str, Any], data: Dict[str, Any]) -> List[Dict[str, str]]:
    section = data.get("section") or "КЖ"
    raw = template.get("sheet_register") or []
    sheets: List[Dict[str, str]] = []
    seen = set()

    for i, sh in enumerate(raw, 1):
        if isinstance(sh, dict):
            title = _clean(sh.get("title") or sh.get("name") or sh.get("sheet") or "", 120)
            number = _clean(sh.get("number") or sh.get("num") or str(i), 20)
            mark = _clean(sh.get("mark") or section, 20)
        else:
            title = _clean(sh, 120)
            number = str(i)
            mark = section
        if not title:
            continue
        key = title.lower()
        if key in seen:
            continue
        seen.add(key)
        sheets.append({"mark": mark, "number": number, "title": title})

    sections = template.get("sections") or []
    if len(sheets) < 6 and sections:
        keys = ["общие", "ведомость", "план", "разрез", "армир", "спецификац", "узел", "фасад", "схема"]
        for sec in sections:
            title = _clean(sec, 120)
            if not title:
                continue
            low = title.lower()
            if not any(k in low for k in keys):
                continue
            if low in seen:
                continue
            seen.add(low)
            sheets.append({"mark": section, "number": str(len(sheets) + 1), "title": title})
            if len(sheets) >= 12:
                break

    base = DEFAULT_ROOF_SHEETS if data.get("project_kind") == "roof" else DEFAULT_FOUNDATION_SHEETS
    for sh in base:
        low = sh["title"].lower()
        if low not in seen:
            sheets.append({"mark": section, "number": str(len(sheets)), "title": sh["title"]})
            seen.add(low)

    fixed = []
    for idx, sh in enumerate(sheets[:20], 1):
        title = _clean(sh.get("title"), 120)
        mark = _clean(sh.get("mark"), 20) or section
        num = _clean(sh.get("number"), 20) or str(idx)
        fixed.append({"mark": mark, "number": num, "title": title})
    return fixed

def _calc_foundation(data: Dict[str, Any]) -> Dict[str, Any]:
    L = float(data["length_m"])
    W = float(data["width_m"])
    area = L * W
    slab_m = data["slab_mm"] / 1000.0
    sand_m = data["sand_mm"] / 1000.0
    gravel_m = data["gravel_mm"] / 1000.0
    step_m = data["rebar_step_mm"] / 1000.0
    d = int(data["rebar_diam_mm"])
    bars_x = int(math.floor(W / step_m)) + 1
    bars_y = int(math.floor(L / step_m)) + 1
    rebar_m_one_layer = bars_x * L + bars_y * W
    rebar_m_total = rebar_m_one_layer * 2
    weight = REBAR_WEIGHT_KG_M.get(d, (d * d) / 162.0)
    rebar_kg = rebar_m_total * weight
    return {
        "area_m2": round(area, 3),
        "concrete_m3": round(area * slab_m, 3),
        "sand_m3": round(area * sand_m, 3),
        "gravel_m3": round(area * gravel_m, 3),
        "bars_x": bars_x,
        "bars_y": bars_y,
        "rebar_m_total": round(rebar_m_total, 1),
        "rebar_kg": round(rebar_kg, 1),
        "rebar_t": round(rebar_kg / 1000.0, 3),
    }

def _frame(c, w, h, title: str, sheet_no: int, total: int, font: str, data: Dict[str, Any]) -> None:
    from reportlab.lib.units import mm
    c.setLineWidth(0.7)
    c.rect(10*mm, 10*mm, w - 20*mm, h - 20*mm)
    c.line(10*mm, 35*mm, w - 10*mm, 35*mm)
    c.line(w - 135*mm, 10*mm, w - 135*mm, 35*mm)
    c.line(w - 80*mm, 10*mm, w - 80*mm, 35*mm)
    c.line(w - 35*mm, 10*mm, w - 35*mm, 35*mm)
    c.setFont(font, 9)
    c.drawString(14*mm, 23*mm, _clean(title, 90))
    c.drawString(w - 132*mm, 23*mm, f"Раздел {data.get('section','КЖ')}")
    c.drawString(w - 77*mm, 23*mm, f"Лист {sheet_no}")
    c.drawString(w - 32*mm, 23*mm, f"Листов {total}")
    c.setFont(font, 7)
    c.drawString(14*mm, 15*mm, f"{ENGINE} · {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.setFont(font, 14)
    c.drawString(18*mm, h - 18*mm, _clean(title, 120))
    c.setFont(font, 8)
    c.drawString(18*mm, h - 25*mm, f"{data.get('project_name','Проект')} · {data.get('length_m')}x{data.get('width_m')} м")

def _draw_lines(c, lines: List[str], x_mm: float, y_mm: float, font: str, size: int = 10, step: float = 7.5) -> None:
    from reportlab.lib.units import mm
    y = y_mm
    c.setFont(font, size)
    for line in lines:
        c.drawString(x_mm*mm, y*mm, _clean(line, 150))
        y -= step

def _draw_plan(c, w, h, font: str, data: Dict[str, Any], calc: Dict[str, Any], rebar: bool = False) -> None:
    from reportlab.lib.units import mm
    L = float(data["length_m"])
    W = float(data["width_m"])
    x0 = 55*mm
    y0 = 55*mm
    scale = min((w - 115*mm) / (L * 1000), (h - 130*mm) / (W * 1000))
    rw = L * 1000 * scale
    rh = W * 1000 * scale
    c.setLineWidth(1.2)
    c.rect(x0, y0, rw, rh)
    c.setFont(font, 9)
    c.drawString(x0, y0 - 8*mm, f"{L:g} м")
    c.saveState()
    c.translate(x0 - 8*mm, y0)
    c.rotate(90)
    c.drawString(0, 0, f"{W:g} м")
    c.restoreState()
    c.setDash(5, 3)
    c.line(x0, y0 + rh / 2, x0 + rw, y0 + rh / 2)
    c.line(x0 + rw / 2, y0, x0 + rw / 2, y0 + rh)
    c.setDash()

    if rebar:
        step_px = max(1.8*mm, data["rebar_step_mm"] * scale)
        c.setLineWidth(0.25)
        x = x0 + step_px
        while x < x0 + rw:
            c.line(x, y0, x, y0 + rh)
            x += step_px
        y = y0 + step_px
        while y < y0 + rh:
            c.line(x0, y, x0 + rw, y)
            y += step_px
        txt = f"Армирование: {data['rebar_class']} Ø{data['rebar_diam_mm']} шаг {data['rebar_step_mm']} мм, верхняя и нижняя сетка"
    else:
        txt = f"Плита {L:g}x{W:g} м, площадь {calc['area_m2']} м²"

    c.setFont(font, 10)
    c.drawString(25*mm, (h/mm - 42)*mm, txt)

def _draw_section(c, w, h, font: str, data: Dict[str, Any]) -> None:
    from reportlab.lib.units import mm
    bx = 55*mm
    by = 70*mm
    total = data["slab_mm"] + data["gravel_mm"] + data["sand_mm"]
    k = 105*mm / total
    layers = [
        ("Фундаментная плита", data["slab_mm"], f"Бетон {data['concrete_class']}, защитный слой {data['cover_mm']} мм"),
        ("Щебёночное основание", data["gravel_mm"], "Уплотнение послойно"),
        ("Песчаная подушка", data["sand_mm"], "Уплотнение послойно"),
    ]
    y = by
    c.setLineWidth(0.8)
    for name, th, note in reversed(layers):
        hh = th * k
        c.rect(bx, y, 230*mm, hh)
        c.setFont(font, 10)
        c.drawString(bx + 5*mm, y + hh/2, f"{name}: {th} мм · {note}")
        y += hh
    c.setFont(font, 10)
    c.drawString(25*mm, 235*mm, "Разрез 1-1")
    c.drawString(25*mm, 225*mm, f"Армирование: {data['rebar_class']} Ø{data['rebar_diam_mm']} шаг {data['rebar_step_mm']} мм")

def _draw_nodes(c, w, h, font: str, data: Dict[str, Any]) -> None:
    from reportlab.lib.units import mm
    c.setFont(font, 11)
    c.drawString(25*mm, 245*mm, "Типовые узлы")
    bx, by = 45*mm, 95*mm
    for i, title in enumerate(["Узел края плиты", "Узел защитного слоя", "Узел основания"], 0):
        x = bx + i * 115*mm
        c.setLineWidth(0.8)
        c.rect(x, by, 90*mm, 75*mm)
        c.line(x, by + 22*mm, x + 90*mm, by + 22*mm)
        c.line(x + 20*mm, by, x + 20*mm, by + 75*mm)
        c.setFont(font, 8)
        c.drawString(x + 4*mm, by + 63*mm, title)
        c.drawString(x + 4*mm, by + 15*mm, f"Защитный слой {data['cover_mm']} мм")
        c.drawString(x + 4*mm, by + 7*mm, f"Ø{data['rebar_diam_mm']} {data['rebar_class']}")

def _spec_rows(data: Dict[str, Any], calc: Dict[str, Any]) -> List[Tuple[str, str, Any, str]]:
    return [
        (f"Бетон {data['concrete_class']} для фундаментной плиты", "м³", calc["concrete_m3"], "по объёму плиты"),
        ("Песчаная подушка", "м³", calc["sand_m3"], "послойное уплотнение"),
        ("Щебёночное основание", "м³", calc["gravel_m3"], "послойное уплотнение"),
        (f"Арматура {data['rebar_class']} Ø{data['rebar_diam_mm']}", "п.м", calc["rebar_m_total"], "верхняя и нижняя сетка"),
        (f"Арматура {data['rebar_class']} Ø{data['rebar_diam_mm']}", "т", calc["rebar_t"], "расчётный вес"),
    ]

def _draw_spec(c, w, h, font: str, rows: List[Tuple[str, str, Any, str]]) -> None:
    from reportlab.lib.units import mm
    x = 25*mm
    y = 235*mm
    c.setFont(font, 10)
    headers = ["№", "Наименование", "Ед.", "Кол-во", "Примечание"]
    widths = [12*mm, 150*mm, 22*mm, 32*mm, 110*mm]
    c.setLineWidth(0.5)
    cx = x
    for head, ww in zip(headers, widths):
        c.rect(cx, y, ww, 8*mm)
        c.drawString(cx + 2*mm, y + 2.3*mm, head)
        cx += ww
    y -= 8*mm
    for i, row in enumerate(rows, 1):
        vals = [str(i), row[0], row[1], str(row[2]), row[3]]
        cx = x
        for val, ww in zip(vals, widths):
            c.rect(cx, y, ww, 8*mm)
            c.drawString(cx + 2*mm, y + 2.3*mm, _clean(val, 55))
            cx += ww
        y -= 8*mm

def write_project_pdf(path: str, data: Dict[str, Any], template: Dict[str, Any], sheets: List[Dict[str, str]], calc: Dict[str, Any]) -> str:
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    font = _font_name()
    page_size = landscape(A3)
    w, h = page_size
    c = canvas.Canvas(path, pagesize=page_size)
    rows = _spec_rows(data, calc)
    total = len(sheets)

    for idx, sh in enumerate(sheets, 1):
        title = sh["title"]
        low = title.lower()
        _frame(c, w, h, f"{sh['mark']}-{sh['number']} {title}", idx, total, font, data)

        if "титул" in low:
            c.setFont(font, 18)
            c.drawCentredString(w/2, h - 85*mm, data["project_name"])
            c.setFont(font, 14)
            c.drawCentredString(w/2, h - 100*mm, f"Раздел {data['section']}")
            c.setFont(font, 11)
            c.drawCentredString(w/2, h - 116*mm, f"Параметры: {data['length_m']:g}x{data['width_m']:g} м, плита {data['slab_mm']} мм")
            c.drawCentredString(w/2, h - 130*mm, "Сформировано по сохранённому шаблону пользователя")
        elif "общ" in low or "данн" in low:
            lines = [
                f"Наименование: {data['project_name']}",
                f"Раздел: {data['section']}",
                f"Размер плиты: {data['length_m']:g} x {data['width_m']:g} м",
                f"Толщина плиты: {data['slab_mm']} мм",
                f"Основание: щебень {data['gravel_mm']} мм, песок {data['sand_mm']} мм",
                f"Бетон: {data['concrete_class']}",
                f"Арматура: {data['rebar_class']} Ø{data['rebar_diam_mm']} шаг {data['rebar_step_mm']} мм",
                f"Площадь: {calc['area_m2']} м², бетон: {calc['concrete_m3']} м³",
                "",
                "Нормативная база:",
            ] + NORMATIVE_NOTES
            _draw_lines(c, lines, 25, 245, font, 10)
        elif "ведомость лист" in low or "состав" in low:
            lines = [f"{i}. {x['mark']}-{x['number']} {x['title']}" for i, x in enumerate(sheets, 1)]
            _draw_lines(c, lines, 25, 245, font, 10)
        elif "план" in low and "арм" not in low:
            _draw_plan(c, w, h, font, data, calc, rebar=False)
        elif "разрез" in low or "сечен" in low:
            _draw_section(c, w, h, font, data)
        elif "ниж" in low and "арм" in low:
            _draw_plan(c, w, h, font, data, calc, rebar=True)
            c.setFont(font, 10)
            c.drawString(25*mm, 220*mm, "Нижняя сетка армирования")
        elif "верх" in low and "арм" in low:
            _draw_plan(c, w, h, font, data, calc, rebar=True)
            c.setFont(font, 10)
            c.drawString(25*mm, 220*mm, "Верхняя сетка армирования")
        elif "арм" in low or "сетк" in low:
            _draw_plan(c, w, h, font, data, calc, rebar=True)
        elif "узел" in low or "детал" in low:
            _draw_nodes(c, w, h, font, data)
        elif "специф" in low or "материал" in low or "стали" in low:
            _draw_spec(c, w, h, font, rows)
        else:
            _draw_lines(c, [
                f"Лист: {title}",
                f"Раздел: {data['section']}",
                f"Размер: {data['length_m']:g}x{data['width_m']:g} м",
                f"Бетон: {data['concrete_class']}",
                f"Арматура: {data['rebar_class']} Ø{data['rebar_diam_mm']} шаг {data['rebar_step_mm']} мм",
            ], 25, 245, font, 10)

        c.showPage()

    c.save()
    return path

def write_project_dxf(path: str, data: Dict[str, Any], calc: Dict[str, Any]) -> str:
    import ezdxf
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 4
    msp = doc.modelspace()

    for name, color in [("KJ_OUTLINE", 7), ("KJ_REBAR", 1), ("KJ_AXES", 3), ("KJ_TEXT", 2), ("KJ_SECTION", 5)]:
        doc.layers.new(name=name, dxfattribs={"color": color})

    L = float(data["length_m"]) * 1000
    W = float(data["width_m"]) * 1000
    step = float(data["rebar_step_mm"])

    pts = [(0,0), (L,0), (L,W), (0,W), (0,0)]
    msp.add_lwpolyline(pts, dxfattribs={"layer": "KJ_OUTLINE", "closed": False})
    msp.add_line((L/2, 0), (L/2, W), dxfattribs={"layer": "KJ_AXES"})
    msp.add_line((0, W/2), (L, W/2), dxfattribs={"layer": "KJ_AXES"})

    x = step
    while x < L:
        msp.add_line((x, 0), (x, W), dxfattribs={"layer": "KJ_REBAR"})
        x += step
    y = step
    while y < W:
        msp.add_line((0, y), (L, y), dxfattribs={"layer": "KJ_REBAR"})
        y += step

    msp.add_text(
        f"{data['project_name']} {data['length_m']:g}x{data['width_m']:g}m",
        dxfattribs={"layer": "KJ_TEXT", "height": 250}
    ).set_placement((0, -900))

    msp.add_text(
        f"{data['concrete_class']} · {data['rebar_class']} D{data['rebar_diam_mm']} step {data['rebar_step_mm']}mm",
        dxfattribs={"layer": "KJ_TEXT", "height": 220}
    ).set_placement((0, -1300))

    sx = L + 1500
    y0 = 0
    layers = [
        ("Sand", data["sand_mm"]),
        ("Gravel", data["gravel_mm"]),
        ("Slab", data["slab_mm"]),
    ]
    for name, th in layers:
        msp.add_lwpolyline([(sx,y0),(sx+4000,y0),(sx+4000,y0+th),(sx,y0+th),(sx,y0)], dxfattribs={"layer": "KJ_SECTION"})
        msp.add_text(f"{name} {th}mm", dxfattribs={"layer": "KJ_TEXT", "height": 160}).set_placement((sx+4200, y0 + th/2))
        y0 += th

    doc.saveas(path)
    return path

def write_project_xlsx(path: str, data: Dict[str, Any], sheets: List[Dict[str, str]], calc: Dict[str, Any], template: Dict[str, Any]) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    headers = ["№", "Наименование", "Ед.изм", "Кол-во", "Примечание"]
    rows = _spec_rows(data, calc)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{data['project_name']} · {data['section']}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEFF")

    for r, row in enumerate(rows, 4):
        ws.cell(r, 1, r - 3)
        ws.cell(r, 2, row[0])
        ws.cell(r, 3, row[1])
        ws.cell(r, 4, row[2])
        ws.cell(r, 5, row[3])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 35

    ws2 = wb.create_sheet("Ведомость листов")
    ws2.append(["№", "Марка", "Номер", "Наименование"])
    for i, sh in enumerate(sheets, 1):
        ws2.append([i, sh["mark"], sh["number"], sh["title"]])
    for col in ["A", "B", "C", "D"]:
        ws2.column_dimensions[col].width = 25

    ws3 = wb.create_sheet("Расчёт")
    for k, v in [
        ("Длина, м", data["length_m"]),
        ("Ширина, м", data["width_m"]),
        ("Площадь, м2", calc["area_m2"]),
        ("Бетон, м3", calc["concrete_m3"]),
        ("Песок, м3", calc["sand_m3"]),
        ("Щебень, м3", calc["gravel_m3"]),
        ("Арматура, п.м", calc["rebar_m_total"]),
        ("Арматура, т", calc["rebar_t"]),
    ]:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 18

    wb.save(path)
    return path

def write_project_manifest(path: str, data: Dict[str, Any], template: Dict[str, Any], sheets: List[Dict[str, str]], calc: Dict[str, Any], files: Dict[str, str], links: Dict[str, str], task_id: str, topic_id: int) -> str:
    manifest = {
        "schema": "AREAL_PROJECT_DOCUMENTATION_PACKAGE_V1",
        "engine": ENGINE,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "task_id": task_id,
        "topic_id": topic_id,
        "input": data.get("input"),
        "section": data.get("section"),
        "project_kind": data.get("project_kind"),
        "template_file": template.get("template_file"),
        "template_project_type": template.get("project_type"),
        "sheet_count": len(sheets),
        "sheet_register": sheets,
        "parameters": data,
        "calculation": calc,
        "files": files,
        "links": links,
        "normative_notes": NORMATIVE_NOTES,
        "status": "ARTIFACTS_CREATED_AND_UPLOADED",
    }
    Path(path).write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return path

def _validate_pdf(path: str, min_pages: int) -> Tuple[bool, str]:
    if not os.path.exists(path) or os.path.getsize(path) < 5000:
        return False, "PDF_FILE_TOO_SMALL"
    try:
        from pypdf import PdfReader
        pages = len(PdfReader(path).pages)
        if pages < min_pages:
            return False, f"PDF_PAGE_COUNT_TOO_LOW:{pages}"
    except Exception as e:
        return False, f"PDF_VALIDATE_ERROR:{str(e)[:100]}"
    return True, ""

def _upload(path: str, task_id: str, topic_id: int) -> str:
    from core.engine_base import upload_artifact_to_drive
    link = upload_artifact_to_drive(path, task_id, topic_id)
    return str(link or "")

def create_full_project_package(raw_input: str, task_id: str, topic_id: int = 0, template_hint: str = "") -> Dict[str, Any]:
    data = parse_project_request(raw_input, template_hint)
    template = _choose_template(data["section"], topic_id)
    sheets = _normalize_sheet_register(template, data)

    if len(sheets) < 8:
        return {
            "success": False,
            "error": f"SHEET_REGISTER_TOO_SHORT:{len(sheets)}",
            "engine": ENGINE,
            "section": data["section"],
            "sheet_count": len(sheets),
        }

    calc = _calc_foundation(data)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_task = _safe_name(task_id)[:20]
    out_dir = Path(tempfile.gettempdir()) / f"areal_project_full_{safe_task}_{stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)

    base_name = f"{data['section']}_project_{safe_task}"
    pdf_path = str(out_dir / f"{base_name}.pdf")
    dxf_path = str(out_dir / f"{base_name}.dxf")
    xlsx_path = str(out_dir / f"{base_name}.xlsx")
    manifest_path = str(out_dir / f"{base_name}.manifest.json")

    res = {
        "success": False,
        "engine": ENGINE,
        "section": data["section"],
        "sheet_count": len(sheets),
        "template_file": template.get("template_file"),
        "pdf_path": pdf_path,
        "dxf_path": dxf_path,
        "xlsx_path": xlsx_path,
        "manifest_path": manifest_path,
        "pdf_link": "",
        "dxf_link": "",
        "xlsx_link": "",
        "manifest_link": "",
        "error": None,
        "data": data,
        "calculation": calc,
    }

    try:
        write_project_pdf(pdf_path, data, template, sheets, calc)
        write_project_dxf(dxf_path, data, calc)
        write_project_xlsx(xlsx_path, data, sheets, calc, template)

        ok, err = _validate_pdf(pdf_path, min_pages=8)
        if not ok:
            res["error"] = err
            return res
        if not os.path.exists(dxf_path) or os.path.getsize(dxf_path) < 1500:
            res["error"] = "DXF_FILE_TOO_SMALL"
            return res
        if not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) < 3000:
            res["error"] = "XLSX_FILE_TOO_SMALL"
            return res

        pdf_link = _upload(pdf_path, task_id, topic_id)
        dxf_link = _upload(dxf_path, task_id, topic_id)
        xlsx_link = _upload(xlsx_path, task_id, topic_id)

        links = {"pdf": pdf_link, "dxf": dxf_link, "xlsx": xlsx_link}
        files = {"pdf": pdf_path, "dxf": dxf_path, "xlsx": xlsx_path}
        write_project_manifest(manifest_path, data, template, sheets, calc, files, links, task_id, topic_id)
        manifest_link = _upload(manifest_path, task_id, topic_id)

        if not pdf_link:
            res["error"] = "PDF_UPLOAD_FAILED"
            return res
        if not dxf_link:
            res["error"] = "DXF_UPLOAD_FAILED"
            return res
        if not xlsx_link:
            res["error"] = "XLSX_UPLOAD_FAILED"
            return res

        res.update({
            "success": True,
            "pdf_link": pdf_link,
            "dxf_link": dxf_link,
            "xlsx_link": xlsx_link,
            "manifest_link": manifest_link,
        })
        return res
    except Exception as e:
        res["error"] = str(e)[:500]
        return res

def is_project_design_request(text: str) -> bool:
    low = _clean(text, 2000).lower()
    triggers = [
        "создай проект",
        "сделай проект",
        "разработай проект",
        "готовый проект",
        "проект фундамент",
        "проект фундаментной плиты",
        "проект кровли",
        "проект по образцу",
        "проект по шаблону",
        "полный проект",
        "проектная документация",
        "рабочая документация",
        "выдай проект",
        "нужен проект",
    ]
    return any(x in low for x in triggers)

def format_project_result_message(res: Dict[str, Any]) -> str:
    if not res.get("success"):
        return "Проект не создан: " + _clean(res.get("error") or "ошибка генерации", 300)
    data = res.get("data") or {}
    calc = res.get("calculation") or {}
    lines = [
        "Проектная документация создана",
        f"Движок: {res.get('engine')}",
        f"Раздел: {res.get('section')}",
        f"Листов PDF: {res.get('sheet_count')}",
        f"Размер: {data.get('length_m')} x {data.get('width_m')} м",
        f"Плита: {data.get('slab_mm')} мм",
        f"Бетон: {data.get('concrete_class')}",
        f"Арматура: {data.get('rebar_class')} Ø{data.get('rebar_diam_mm')} шаг {data.get('rebar_step_mm')} мм",
        f"Бетон: {calc.get('concrete_m3')} м³",
        f"Арматура: {calc.get('rebar_t')} т",
        "",
        f"PDF: {res.get('pdf_link')}",
        f"DXF: {res.get('dxf_link')}",
        f"XLSX: {res.get('xlsx_link')}",
    ]
    if res.get("manifest_link"):
        lines.append(f"MANIFEST: {res.get('manifest_link')}")
    lines.append("")
    lines.append("Доволен результатом? Ответь: Да / Уточни / Правки")
    return "\n".join(lines)

async def create_project_pdf_dxf_artifact(raw_input: str, task_id: str, topic_id: int = 0, template_hint: str = "") -> Dict[str, Any]:
    return create_full_project_package(raw_input, task_id, topic_id, template_hint)

# === END FULLFIX_07_CAD_PROJECT_DOCUMENTATION_CLOSURE ===
