# === FULLFIX_16_ESTIMATE_UNIFIED_P0_SAFE ===
import os, re, logging, sqlite3, traceback
logger = logging.getLogger(__name__)
ENGINE = "FULLFIX_16_ESTIMATE_UNIFIED_P0_SAFE"
RUNTIME_DIR = "/root/.areal-neva-core/runtime"
CORE_DB = "/root/.areal-neva-core/data/core.db"
MEMORY_DB = "/root/.areal-neva-core/data/memory.db"
os.makedirs(RUNTIME_DIR, exist_ok=True)

_STRIP_RE = re.compile(r"(?im)^\s*MANIFEST\s*:\s*https?://\S+\s*$")

def _strip_manifest(text):
    t = str(text or "")
    t = _STRIP_RE.sub("", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()

def parse_estimate_rows(text):
    try:
        from core.sample_template_engine import parse_estimate_items
        rows = parse_estimate_items(text)
        if rows:
            return rows
    except Exception:
        pass
    rows = []
    seen = set()
    pat = re.compile(
        r"([а-яёА-ЯЁa-zA-Z][а-яёА-ЯЁa-zA-Z0-9 \-/\.\"]{1,60}?)"
        r"\s+(\d+(?:[.,]\d+)?)\s*"
        r"(м²|м2|м³|м3|п\.м|м\.п|шт|кг|тн|т|компл\.?|л|м)\s*"
        r"(?:(?:цена|по|x|х|@)?\s*(\d+(?:[.,]\d+)?)(?:\s*руб\.?)?)?",
        re.I | re.U
    )
    skip = {"итого", "всего", "смета", "смету", "сделай", "составь"}
    for m in pat.finditer(str(text or "")):
        name = m.group(1).strip().rstrip(",:. ")
        name = re.sub(r"^(сделай|составь|посчитай|смету|смета|по|на)\s+", "", name, flags=re.I|re.U)
        name = name.strip(" ,:;.-")
        if not name or name.lower() in skip or len(name) < 2:
            continue
        qty = float(m.group(2).replace(",", "."))
        unit = m.group(3).replace("м2","м²").replace("м3","м³")
        price = float(m.group(4).replace(",", ".")) if m.group(4) else 0.0
        key = (name.lower(), qty, unit, price)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"name": name, "qty": qty, "unit": unit, "price": price, "total": round(qty*price, 2)})
    return rows

def generate_xlsx(rows, task_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    path = os.path.join(RUNTIME_DIR, "estimate_" + str(task_id)[:8] + ".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.merge_cells("A1:F1")
    ws["A1"] = "СМЕТА"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    hdrs = ["№", "Наименование", "Ед.", "Кол-во", "Цена, руб.", "Сумма, руб."]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for col, w in zip("ABCDEF", [6, 45, 10, 12, 16, 16]):
        ws.column_dimensions[col].width = w
    for i, row in enumerate(rows, 1):
        r = i + 2
        for c, v in enumerate([i, row["name"], row["unit"], row["qty"], row["price"], "=D"+str(r)+"*E"+str(r)], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    tr = len(rows) + 3
    ws.cell(row=tr, column=2, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=tr, column=6, value="=SUM(F3:F"+str(tr-1)+")").font = Font(bold=True)
    for c in range(1, 7):
        ws.cell(row=tr, column=c).border = thin
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:F" + str(max(tr, 3))
    wb.save(path)
    return path

def generate_pdf(rows, task_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
    from reportlab.lib import colors
    from core.pdf_cyrillic import register_cyrillic_fonts, make_styles, make_paragraph, clean_pdf_text, FONT_BOLD
    path = os.path.join(RUNTIME_DIR, "estimate_" + str(task_id)[:8] + ".pdf")
    register_cyrillic_fonts()
    styles = make_styles()
    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=20, bottomMargin=20, leftMargin=20, rightMargin=20)
    hdr = [make_paragraph(h, "bold", styles) for h in ["№", "Наименование", "Ед.", "Кол-во", "Цена, руб.", "Сумма, руб."]]
    data = [hdr]
    total = 0.0
    for i, row in enumerate(rows, 1):
        t = round(float(row["qty"]) * float(row["price"]), 2)
        total += t
        data.append([
            make_paragraph(str(i), "normal", styles),
            make_paragraph(clean_pdf_text(row["name"]), "normal", styles),
            make_paragraph(row["unit"], "normal", styles),
            make_paragraph(str(row["qty"]), "normal", styles),
            make_paragraph("%.2f" % row["price"], "normal", styles),
            make_paragraph("%.2f" % t, "normal", styles),
        ])
    data.append([make_paragraph("", "normal", styles), make_paragraph("ИТОГО", "bold", styles),
                 make_paragraph("", "normal", styles), make_paragraph("", "normal", styles),
                 make_paragraph("", "normal", styles), make_paragraph("%.2f" % total, "bold", styles)])
    story = [make_paragraph("СМЕТА", "header", styles), Spacer(1, 8)]
    tbl = Table(data, colWidths=[22, 190, 32, 52, 70, 70])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#444444")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),FONT_BOLD),
        ("GRID",(0,0),(-1,-1),0.4,colors.black),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#FFFFCC")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(tbl)
    doc.build(story)
    return path

# === MAIN ENTRY — opens its own DB connection, no cross-thread SQLite ===
def process_estimate_task_sync(task_id, chat_id, topic_id, raw_input):
    from core.artifact_upload_guard import upload_many_or_fail
    from core.reply_sender import send_reply_ex
    try:
        rows = parse_estimate_rows(raw_input)
        if not rows:
            msg = "Смета не создана: не нашёл строки «позиция количество единица цена»"
            with sqlite3.connect(CORE_DB, timeout=30) as c:
                c.execute("UPDATE tasks SET state='FAILED',result=?,error_message=?,updated_at=datetime('now') WHERE id=?",
                    (msg, "NO_ESTIMATE_ROWS", task_id))
                c.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                    (task_id, "state:FAILED:no_rows"))
                c.commit()
            send_reply_ex(chat_id=str(chat_id), text=msg, reply_to_message_id=None)
            return False

        xlsx_path = generate_xlsx(rows, task_id)
        pdf_path = generate_pdf(rows, task_id)
        up = upload_many_or_fail(
            [{"path": pdf_path, "kind": "estimate_pdf"}, {"path": xlsx_path, "kind": "estimate_xlsx"}],
            task_id, topic_id
        )
        pdf_r = up.get("results", {}).get(pdf_path, {})
        xlsx_r = up.get("results", {}).get(xlsx_path, {})
        pdf_link = pdf_r.get("link") if pdf_r.get("success") else ""
        xlsx_link = xlsx_r.get("link") if xlsx_r.get("success") else ""
        total = round(sum(float(r["qty"]) * float(r["price"]) for r in rows), 2)

        if not (pdf_link or xlsx_link):
            msg = "Смета рассчитана, Drive upload не выполнен. Позиций: " + str(len(rows)) + ". Итого: %.2f руб" % total
            with sqlite3.connect(CORE_DB, timeout=30) as c:
                c.execute("UPDATE tasks SET state='FAILED',result=?,error_message=?,updated_at=datetime('now') WHERE id=?",
                    (msg, "UPLOAD_FAILED", task_id))
                c.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                    (task_id, "state:FAILED:upload_failed"))
                c.commit()
            send_reply_ex(chat_id=str(chat_id), text=msg, reply_to_message_id=None)
            return False

        lines = ["Смета готова.", "Позиций: " + str(len(rows)) + ". Итого: %.2f руб" % total]
        if pdf_link:
            lines.append("PDF: " + pdf_link)
        if xlsx_link:
            lines.append("XLSX: " + xlsx_link)
        lines.append("")
        lines.append("Доволен результатом? Ответь: Да / Уточни / Правки")
        result_text = _strip_manifest("\n".join(lines))

        with sqlite3.connect(CORE_DB, timeout=30) as c:
            c.execute("UPDATE tasks SET state='AWAITING_CONFIRMATION',result=?,updated_at=datetime('now') WHERE id=?",
                (result_text, task_id))
            c.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                (task_id, "state:AWAITING_CONFIRMATION:estimate_unified"))
            c.commit()

        br = send_reply_ex(chat_id=str(chat_id), text=result_text, reply_to_message_id=None)
        bmid = None
        if isinstance(br, dict):
            bmid = br.get("bot_message_id") or br.get("message_id")
        elif hasattr(br, "message_id"):
            bmid = br.message_id
        if bmid:
            with sqlite3.connect(CORE_DB, timeout=30) as c:
                c.execute("UPDATE tasks SET bot_message_id=?,updated_at=datetime('now') WHERE id=?", (str(bmid), task_id))
                c.commit()

        try:
            with sqlite3.connect(MEMORY_DB, timeout=10) as mc:
                import json
                mc.execute("INSERT OR REPLACE INTO memory(chat_id,key,value,timestamp) VALUES(?,?,?,datetime('now'))",
                    (str(chat_id), "topic_"+str(topic_id)+"_last_estimate",
                     json.dumps({"task_id": task_id, "rows": len(rows), "total": total}, ensure_ascii=False)))
                mc.commit()
        except Exception as me:
            logger.warning("ESTIMATE_MEMORY_SAVE err=%s", me)

        return True

    except Exception as e:
        err = traceback.format_exc()
        logger.error("ESTIMATE_UNIFIED_ERROR task=%s err=%s trace=%s", task_id, e, err)
        msg = "Смета не создана: внутренняя ошибка"
        try:
            with sqlite3.connect(CORE_DB, timeout=30) as c:
                c.execute("UPDATE tasks SET state='FAILED',result=?,error_message=?,updated_at=datetime('now') WHERE id=?",
                    (msg, str(e)[:500], task_id))
                c.execute("INSERT INTO task_history(task_id,action,created_at) VALUES(?,?,datetime('now'))",
                    (task_id, "state:FAILED:exception"))
                c.commit()
        except Exception:
            pass
        try:
            from core.reply_sender import send_reply_ex
            send_reply_ex(chat_id=str(chat_id), text=msg, reply_to_message_id=None)
        except Exception:
            pass
        return False

async def process_estimate_task(conn, task_id, chat_id, topic_id, raw_input):
    # conn intentionally NOT passed to sync function — opens its own connection
    return process_estimate_task_sync(task_id, chat_id, topic_id, raw_input)
# === END FULLFIX_16_ESTIMATE_UNIFIED_P0_SAFE ===
